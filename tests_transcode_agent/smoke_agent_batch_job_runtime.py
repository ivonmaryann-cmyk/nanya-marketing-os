from __future__ import annotations

from pathlib import Path
import shutil
import sys
from tempfile import TemporaryDirectory

import openpyxl
from werkzeug.datastructures import FileStorage


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app import db
from fangzheng_web_app import transcode_agent_rules as agent_rules
from fangzheng_web_app import transcode_rules
from fangzheng_web_app.db import create_job, get_job
from fangzheng_web_app.transcode_agent_rules import FEATURE_KEY
from fangzheng_web_app.transcode_agent_service import (
    _comparison_code_display,
    _comparison_code_differences,
    _normalize_comparison_code,
    calculate_transcode_agent_quote,
    run_transcode_agent_job,
)


DRAFT_PATH = ROOT / "docs/develop0707/客户特殊规则结构化草稿_按原表行_20260708.xlsx"
SAMPLE_PATH = ROOT / "tests_transcode_agent/fixtures/transcode_agent_499_regression.xlsx"
EXPECTED_RULE_COUNT = 196
EXPECTED_CONFIRMED_PARSE_CODES = {
    7: "DA000581141104920RWA1C",
    28: "2A01000HH43004900YWF1T",
    29: "2A009001143004900YWF1T",
    94: "DA00050HH41004900RWA1C",
    95: "DA00100HH41004900RWA1C",
    96: "DA00053HH41304930RWA1C",
    97: "DA00060HH43304930RWA1C",
    98: "DA00065HH43004900RWA1C",
    99: "DA00065TT43304930RRA1C",
    100: "DA00050HH43004900RWA1C",
    101: "DA00065HH43004900RRA1C",
    102: "CG00152HH18001600RZD3C",
    103: "CG00203HH18001600RZD3C",
    214: "2B001001143004900YWA1C",
    411: "2B001021120402441YWA1C",
    412: "2B00710HH20402441YWA1C",
    413: "2B00710HH20242426YWA1C",
    414: "2B001521120242426YWA1C",
}


def main() -> None:
    assert DRAFT_PATH.exists(), DRAFT_PATH
    assert SAMPLE_PATH.exists(), SAMPLE_PATH

    with TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        original_db_path = db.DATABASE_PATH
        original_agent_rules_dir = agent_rules.TRANSCODE_AGENT_RULES_DIR
        original_agent_versions_dir = agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR
        original_base_versions_dir = transcode_rules.TRANSCODE_RULES_VERSIONS_DIR
        try:
            db.DATABASE_PATH = temp_dir / "storage/app.db"
            db.DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
            db.init_db()

            transcode_rules.TRANSCODE_RULES_VERSIONS_DIR = temp_dir / "rules/transcode/versions"
            transcode_rules.TRANSCODE_RULES_VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
            base_version = transcode_rules.ensure_default_transcode_rule_version()

            agent_rules.TRANSCODE_AGENT_RULES_DIR = temp_dir / "rules/transcode_agent"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = agent_rules.TRANSCODE_AGENT_RULES_DIR / "versions"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
            with DRAFT_PATH.open("rb") as source:
                upload = FileStorage(
                    stream=source,
                    filename=DRAFT_PATH.name,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                agent_version = agent_rules.save_new_transcode_agent_rule_version(
                    upload,
                    updated_by="smoke",
                    remark="batch job runtime smoke",
                )
            assert len(agent_rules.load_transcode_agent_rules(agent_version)) == EXPECTED_RULE_COUNT

            employee_id = "smoke"
            job_dir = temp_dir / "jobs" / employee_id
            job_dir.mkdir(parents=True, exist_ok=True)
            input_path = job_dir / SAMPLE_PATH.name
            shutil.copy2(SAMPLE_PATH, input_path)

            job_id = create_job(
                employee_id,
                SAMPLE_PATH.name,
                str(input_path),
                f"base:{base_version};agent:{agent_version}",
                feature=FEATURE_KEY,
            )
            run_transcode_agent_job(job_id, employee_id)

            job = get_job(job_id)
            assert job is not None
            assert job["status"] == "completed", dict(job)
            assert job["total_rows"] == 499, dict(job)
            assert job["skip_count"] == 145, dict(job)
            assert job["success_count"] == 354, dict(job)
            assert job["fail_count"] == 0, dict(job)
            assert job["confirm_count"] == 0, dict(job)
            assert str(job["rule_version"]) == f"base:{base_version};agent:{agent_version}", dict(job)

            result_path = Path(job["stored_result_path"])
            assert result_path.exists(), result_path
            workbook = openpyxl.load_workbook(result_path, data_only=True)
            for sheet_name in ["字段证据链", "规则命中汇总", "问题分类明细", "问题聚合修复清单", "证据评分影子对比"]:
                assert sheet_name in workbook.sheetnames, workbook.sheetnames
            assert "待确认清单" not in workbook.sheetnames, workbook.sheetnames
            main_ws = workbook[workbook.sheetnames[0]]
            main_headers = [cell.value for cell in main_ws[1]]
            result_index = max(
                index for index, header in enumerate(main_headers, start=1)
                if header == "Agent转码结果"
            )
            output_status_index = main_headers.index("结果对比") + 1
            assert "待人工确认码值" not in main_headers, main_headers
            difference_index = main_headers.index("22位码值差异") + 1
            system_analysis_index = main_headers.index("系统分析原因") + 1
            assert difference_index == result_index + 1, main_headers
            assert output_status_index == difference_index + 1, main_headers
            assert system_analysis_index == len(main_headers), main_headers
            assert "需要人工确认" not in main_headers, main_headers
            transcode_status_index = main_headers.index("转码状态") + 1
            confirmation_index = main_headers.index("人工确认提示") + 1
            product_name_index = next(
                index for index, header in enumerate(main_headers, start=1)
                if str(header or "").replace(" ", "") == "品名"
            )
            confirmation_count = 0
            comparison_counts = {True: 0, False: 0, "跳过": 0}
            for excel_row, row in enumerate(main_ws.iter_rows(min_row=2, values_only=True), start=2):
                result = str(row[result_index - 1] or "")
                expected = bool(result) and not result.startswith(("待确认", "未识别", "跳过"))
                if result.startswith("跳过"):
                    expected_comparison = "跳过"
                elif not expected:
                    expected_comparison = "未出码"
                else:
                    expected_comparison = (
                        _normalize_comparison_code(result)
                        == _normalize_comparison_code(row[product_name_index - 1])
                    )
                assert row[output_status_index - 1] == expected_comparison, row
                system_reason = str(row[system_analysis_index - 1] or "")
                if expected_comparison is False:
                    assert system_reason.startswith("码值不一致："), row
                    assert "定位：" in system_reason and "建议：" in system_reason, row
                else:
                    assert not system_reason, row
                expected_display = str(
                    _comparison_code_display(result, row[product_name_index - 1], expected_comparison)
                )
                assert row[difference_index - 1] == expected_display, row
                result_fill = main_ws.cell(row=excel_row, column=result_index).fill.fgColor.rgb or ""
                difference_fill = main_ws.cell(row=excel_row, column=difference_index).fill.fgColor.rgb or ""
                assert difference_fill == result_fill, (excel_row, result_fill, difference_fill)
                comparison_counts[expected_comparison] = comparison_counts.get(expected_comparison, 0) + 1
                if excel_row in EXPECTED_CONFIRMED_PARSE_CODES:
                    assert result.split("*", 1)[0][:22] == EXPECTED_CONFIRMED_PARSE_CODES[excel_row], row
                confirmation = str(row[confirmation_index - 1] or "")
                if confirmation:
                    confirmation_count += 1
                    assert confirmation.startswith("待确认："), row
                    assert expected is True, row
                    assert row[transcode_status_index - 1] == "待人工确认", row
                    fill_color = main_ws.cell(row=excel_row, column=result_index).fill.fgColor.rgb or ""
                    assert fill_color.endswith("FFCDD2"), (excel_row, fill_color, row)
                else:
                    expected_status = "跳过" if result.startswith("跳过") else "可直接采用"
                    assert row[transcode_status_index - 1] == expected_status, row
                if expected_comparison is False:
                    assert row[transcode_status_index - 1] == "可直接采用", row
                    assert not confirmation, row
                    differences = _comparison_code_differences(result, row[product_name_index - 1])
                    assert differences, row
                    fill_color = main_ws.cell(row=excel_row, column=result_index).fill.fgColor.rgb or ""
                    assert fill_color.endswith("FFCDD2"), (excel_row, fill_color, row)
            assert confirmation_count == 0, confirmation_count
            assert comparison_counts == {True: 316, False: 38, "跳过": 145}, comparison_counts

            issue_detail_ws = workbook["问题分类明细"]
            issue_detail_headers = [cell.value for cell in issue_detail_ws[1]]
            for header in ["问题分类", "错误字段", "已命中规则ID", "候选未命中规则ID", "问题指纹"]:
                assert header in issue_detail_headers, issue_detail_headers
            issue_detail_rows = list(issue_detail_ws.iter_rows(min_row=2, values_only=True))
            assert len(issue_detail_rows) >= comparison_counts[False], len(issue_detail_rows)
            assert all(row[4] is False for row in issue_detail_rows), issue_detail_rows[:3]
            assert all(row[5] and row[6] and row[14] for row in issue_detail_rows), issue_detail_rows[:3]
            issue_aggregate_ws = workbook["问题聚合修复清单"]
            issue_aggregate_rows = list(issue_aggregate_ws.iter_rows(min_row=2, values_only=True))
            assert issue_aggregate_rows, "no aggregated issue rows"
            affected_counts = [int(row[3]) for row in issue_aggregate_rows]
            assert affected_counts == sorted(affected_counts, reverse=True), affected_counts
            assert sum(affected_counts) == len(issue_detail_rows), (affected_counts, len(issue_detail_rows))

            rich_workbook = openpyxl.load_workbook(result_path, rich_text=True)
            rich_ws = rich_workbook[rich_workbook.sheetnames[0]]
            false_row = next(
                row
                for row in range(2, rich_ws.max_row + 1)
                if rich_ws.cell(row=row, column=output_status_index).value is False
            )
            rich_value = rich_ws.cell(row=false_row, column=difference_index).value
            assert isinstance(rich_value, openpyxl.cell.rich_text.CellRichText), rich_value
            error_blocks = [
                block for block in rich_value
                if isinstance(block, openpyxl.cell.rich_text.TextBlock)
            ]
            assert error_blocks, rich_value
            assert all((block.font.color.rgb or "").upper() == "FF9C0006" for block in error_blocks), error_blocks
            rich_workbook.close()
            for sheet in workbook.worksheets[1:]:
                populated_rows = [
                    row for row in sheet.iter_rows(values_only=True)
                    if any(value is not None and str(value).strip() for value in row)
                ]
                assert len(populated_rows) > 1, (sheet.title, populated_rows)

            evidence_ws = workbook["字段证据链"]
            summary_ws = workbook["规则命中汇总"]
            evidence_headers = [cell.value for cell in evidence_ws[1]]
            assert evidence_headers[:5] == ["行号", "客户代码", "客户", "规格", "状态"], evidence_headers
            evidence_rows = list(evidence_ws.iter_rows(min_row=2, values_only=True))
            agent_evidence_rows = [row for row in evidence_rows if row[9] == "Agent规则覆盖"]
            assert agent_evidence_rows, "no Agent rule evidence rows"
            assert not any(row[5] == "结构码" for row in agent_evidence_rows), agent_evidence_rows[:5]

            summary_values = {
                str(row[0] or ""): row[1]
                for row in summary_ws.iter_rows(min_row=1, max_col=2, values_only=True)
                if row and row[0]
            }
            assert summary_values.get("当前Agent机器规则数") == EXPECTED_RULE_COUNT, summary_values
            assert summary_values.get("总行数") == 499, summary_values
            assert summary_values.get("跳过PP/RC/%") == 145, summary_values

            invalid_quote = calculate_transcode_agent_quote(
                "NYUNKNOWN 0.8mm 1/1 37*49 HTE",
                customer="解析失败测试",
            )
            assert invalid_quote["status"] == "失败", invalid_quote
            assert not invalid_quote["result"], invalid_quote
            assert not invalid_quote["candidate_code"], invalid_quote

            print(
                "batch job runtime smoke passed "
                f"base_version={base_version} agent_version={agent_version} "
                f"rows={job['total_rows']} success={job['success_count']} "
                f"fail={job['fail_count']} skip={job['skip_count']} "
                f"agent_evidence_rows={len(agent_evidence_rows)}"
            )
        finally:
            db.DATABASE_PATH = original_db_path
            agent_rules.TRANSCODE_AGENT_RULES_DIR = original_agent_rules_dir
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = original_agent_versions_dir
            transcode_rules.TRANSCODE_RULES_VERSIONS_DIR = original_base_versions_dir


if __name__ == "__main__":
    main()
