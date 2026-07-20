from __future__ import annotations

import os
import shutil
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl
from werkzeug.datastructures import FileStorage


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app import db
from fangzheng_web_app import transcode_agent_rules as agent_rules
from fangzheng_web_app import transcode_rules
from fangzheng_web_app import transcode_semantic_rules as semantic_rules
from fangzheng_web_app import transcode_agent_service as agent_service
from fangzheng_web_app.db import create_job, get_job
from fangzheng_web_app.transcode_evidence_model import EvidenceModelRuntime
from fangzheng_web_app.transcode_agent_rules import FEATURE_KEY
from fangzheng_web_app.transcode_agent_service import FORMAL_RESULT_HEADER, run_transcode_agent_job


DRAFT_PATH = ROOT / "docs/develop0707/客户特殊规则结构化草稿_按原表行_20260708.xlsx"
SEMANTIC_RULE_WORKBOOK = ROOT / "docs/develop0707/营销转码Agent最终模型语义规则表_20260710.xlsx"
UPLOAD_SAMPLE_PATH = ROOT / "tests_transcode_agent/fixtures/transcode_agent_10_manual_upload.xlsx"


def main() -> None:
    assert DRAFT_PATH.exists(), DRAFT_PATH
    assert SEMANTIC_RULE_WORKBOOK.exists(), SEMANTIC_RULE_WORKBOOK
    assert UPLOAD_SAMPLE_PATH.exists(), UPLOAD_SAMPLE_PATH
    with TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        original_db_path = db.DATABASE_PATH
        original_agent_rules_dir = agent_rules.TRANSCODE_AGENT_RULES_DIR
        original_agent_versions_dir = agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR
        original_base_versions_dir = transcode_rules.TRANSCODE_RULES_VERSIONS_DIR
        original_semantic_rules_dir = semantic_rules.TRANSCODE_SEMANTIC_RULES_DIR
        original_semantic_versions_dir = semantic_rules.TRANSCODE_SEMANTIC_RULES_VERSIONS_DIR
        original_mode = os.environ.get("TRANSCODE_SEMANTIC_RULE_RUNTIME_MODE")
        original_gate_mode = os.environ.get("TRANSCODE_EVIDENCE_GATE_MODE")
        original_override_mode = os.environ.get("TRANSCODE_SEMANTIC_OVERRIDE_MODE")
        try:
            db.DATABASE_PATH = temp_dir / "storage/app.db"
            db.DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
            db.init_db()

            transcode_rules.TRANSCODE_RULES_VERSIONS_DIR = temp_dir / "rules/transcode/versions"
            transcode_rules.TRANSCODE_RULES_VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
            base_version = transcode_rules.ensure_default_transcode_rule_version()

            agent_rules.TRANSCODE_AGENT_RULES_DIR = temp_dir / "rules/transcode_agent"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = agent_rules.TRANSCODE_AGENT_RULES_DIR / "versions"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
            with DRAFT_PATH.open("rb") as source:
                upload = FileStorage(
                    stream=source,
                    filename=DRAFT_PATH.name,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                agent_version = agent_rules.save_new_transcode_agent_rule_version(
                    upload,
                    updated_by="smoke",
                    remark="semantic shadow batch smoke",
                )

            semantic_rules.TRANSCODE_SEMANTIC_RULES_DIR = temp_dir / "rules/transcode_semantic"
            semantic_rules.TRANSCODE_SEMANTIC_RULES_VERSIONS_DIR = (
                semantic_rules.TRANSCODE_SEMANTIC_RULES_DIR / "versions"
            )
            semantic_rules.TRANSCODE_SEMANTIC_RULES_VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
            semantic_version = semantic_rules.publish_transcode_semantic_rule_version(
                SEMANTIC_RULE_WORKBOOK,
                updated_by="smoke",
                approval_basis="P2-2 shadow smoke",
                remark="semantic shadow batch smoke",
            )

            os.environ["TRANSCODE_SEMANTIC_RULE_RUNTIME_MODE"] = "off"
            os.environ["TRANSCODE_EVIDENCE_GATE_MODE"] = "shadow"
            os.environ["TRANSCODE_SEMANTIC_OVERRIDE_MODE"] = "off"
            off_job = _run_job(temp_dir, "off", base_version, agent_version, semantic_version)
            os.environ["TRANSCODE_SEMANTIC_RULE_RUNTIME_MODE"] = "shadow"
            shadow_job = _run_job(temp_dir, "shadow", base_version, agent_version, semantic_version)
            os.environ["TRANSCODE_SEMANTIC_OVERRIDE_MODE"] = "enforce"
            os.environ["TRANSCODE_EVIDENCE_GATE_MODE"] = "enforce"
            enforce_job = _run_job(temp_dir, "enforce", base_version, agent_version, semantic_version)
            os.environ["TRANSCODE_SEMANTIC_OVERRIDE_MODE"] = "off"
            os.environ["TRANSCODE_EVIDENCE_GATE_MODE"] = "shadow"
            original_model_loader = agent_service.load_evidence_model_runtime
            fake_client = FakeEvidenceClient()
            agent_service.load_evidence_model_runtime = lambda: EvidenceModelRuntime(
                mode="shadow",
                client=fake_client,
                model="fake-evidence-model",
            )
            try:
                model_job = _run_job(temp_dir, "model", base_version, agent_version, semantic_version)
            finally:
                agent_service.load_evidence_model_runtime = original_model_loader

            assert off_job["total_rows"] == shadow_job["total_rows"] == 10
            assert off_job["success_count"] == shadow_job["success_count"]
            assert off_job["fail_count"] == shadow_job["fail_count"]
            assert off_job["skip_count"] == shadow_job["skip_count"] == 0

            off_wb = openpyxl.load_workbook(off_job["stored_result_path"], data_only=True)
            shadow_wb = openpyxl.load_workbook(shadow_job["stored_result_path"], data_only=True)
            model_wb = openpyxl.load_workbook(model_job["stored_result_path"], data_only=True)
            enforce_wb = openpyxl.load_workbook(enforce_job["stored_result_path"], data_only=True)
            assert _formal_results(off_wb) == _formal_results(shadow_wb)
            assert _formal_results(shadow_wb) == _formal_results(model_wb)
            enforce_summary = {
                str(row[0] or ""): row[1]
                for row in enforce_wb["规则命中汇总"].iter_rows(min_row=1, max_col=2, values_only=True)
                if row and row[0]
            }
            assert enforce_summary.get("正式证据门禁模式") == "enforce", enforce_summary
            assert int(enforce_summary.get("正式证据门禁拦截") or 0) >= 0, enforce_summary
            enforce_score_ws = enforce_wb["证据评分影子对比"]
            enforce_headers = [cell.value for cell in enforce_score_ws[1]]
            blocked_index = enforce_headers.index("正式门禁拦截")
            enforce_rows = list(enforce_score_ws.iter_rows(min_row=2, values_only=True))
            assert enforce_rows
            assert all(row[blocked_index] in {"是", "否"} for row in enforce_rows)
            assert "模型语义影子证据" in shadow_wb.sheetnames
            assert "证据评分影子对比" in shadow_wb.sheetnames

            shadow_rows = list(shadow_wb["模型语义影子证据"].iter_rows(min_row=2, values_only=True))
            assert shadow_rows, "shadow evidence is empty"
            assert any(row[12] == "缺少输入" and "订单备注" in str(row[13] or "") for row in shadow_rows)
            assert any(row[12] == "命中" for row in shadow_rows)

            summary = {
                str(row[0] or ""): row[1]
                for row in shadow_wb["规则命中汇总"].iter_rows(min_row=1, max_col=2, values_only=True)
                if row and row[0]
            }
            assert summary.get("模型语义影子版本") == semantic_version, summary
            assert summary.get("模型语义正式规则数") == 51, summary
            assert int(summary.get("模型语义影子评估数") or 0) > 0, summary
            assert int(summary.get("模型语义缺少输入") or 0) > 0, summary
            assert summary.get("模型语义运行时影响") == "影子观察；不覆盖编码和评分", summary
            assert summary.get("证据影子评分行数") == 10, summary
            assert summary.get("证据影子模型调用") == 0, summary
            score_rows = list(shadow_wb["证据评分影子对比"].iter_rows(min_row=2, values_only=True))
            assert len(score_rows) == 80, len(score_rows)
            assert all(row[26] == "否" for row in score_rows), score_rows[:3]
            assert any(row[17] in {"ambiguous", "missing_evidence"} for row in score_rows)

            model_summary = {
                str(row[0] or ""): row[1]
                for row in model_wb["规则命中汇总"].iter_rows(min_row=1, max_col=2, values_only=True)
                if row and row[0]
            }
            assert 0 < model_summary.get("证据影子模型调用") <= 10, model_summary
            assert model_summary.get("证据影子模型成功") == model_summary.get("证据影子模型调用"), model_summary
            assert model_summary.get("证据影子模型失败") == 0, model_summary
            model_score_rows = list(model_wb["证据评分影子对比"].iter_rows(min_row=2, values_only=True))
            assert any(row[26] == "是" for row in model_score_rows)
            assert any(row[36] == "模型证据审查仅更新影子对比，不覆盖当前分数和90分门禁" for row in model_score_rows)

            print(
                "semantic shadow batch smoke passed "
                f"version={semantic_version} rows=10 formal_results_unchanged "
                f"evaluations={summary['模型语义影子评估数']} "
                f"missing={summary['模型语义缺少输入']}"
            )
        finally:
            db.DATABASE_PATH = original_db_path
            agent_rules.TRANSCODE_AGENT_RULES_DIR = original_agent_rules_dir
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = original_agent_versions_dir
            transcode_rules.TRANSCODE_RULES_VERSIONS_DIR = original_base_versions_dir
            semantic_rules.TRANSCODE_SEMANTIC_RULES_DIR = original_semantic_rules_dir
            semantic_rules.TRANSCODE_SEMANTIC_RULES_VERSIONS_DIR = original_semantic_versions_dir
            if original_mode is None:
                os.environ.pop("TRANSCODE_SEMANTIC_RULE_RUNTIME_MODE", None)
            else:
                os.environ["TRANSCODE_SEMANTIC_RULE_RUNTIME_MODE"] = original_mode
            if original_gate_mode is None:
                os.environ.pop("TRANSCODE_EVIDENCE_GATE_MODE", None)
            else:
                os.environ["TRANSCODE_EVIDENCE_GATE_MODE"] = original_gate_mode
            if original_override_mode is None:
                os.environ.pop("TRANSCODE_SEMANTIC_OVERRIDE_MODE", None)
            else:
                os.environ["TRANSCODE_SEMANTIC_OVERRIDE_MODE"] = original_override_mode


def _run_job(
    temp_dir: Path,
    label: str,
    base_version: str,
    agent_version: str,
    semantic_version: str,
):
    employee_id = f"smoke-{label}"
    job_dir = temp_dir / "jobs" / employee_id
    job_dir.mkdir(parents=True, exist_ok=True)
    input_path = job_dir / UPLOAD_SAMPLE_PATH.name
    shutil.copy2(UPLOAD_SAMPLE_PATH, input_path)
    job_id = create_job(
        employee_id,
        UPLOAD_SAMPLE_PATH.name,
        str(input_path),
        f"base:{base_version};agent:{agent_version};semantic:{semantic_version}",
        feature=FEATURE_KEY,
    )
    run_transcode_agent_job(job_id, employee_id)
    job = get_job(job_id)
    assert job is not None and job["status"] == "completed", dict(job or {})
    assert Path(job["stored_result_path"]).exists(), dict(job)
    return job


def _formal_results(workbook) -> list[str]:
    ws = workbook["转码需求表"] if "转码需求表" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    result_index = headers.index(FORMAL_RESULT_HEADER) + 1
    return [str(ws.cell(row=row, column=result_index).value or "") for row in range(2, ws.max_row + 1)]


class FakeEvidenceClient:
    def __init__(self):
        self.calls = []

    def review_evidence(self, **kwargs):
        self.calls.append(kwargs)
        return {
            "schema_version": "1.0",
            "field_reviews": [
                {
                    "field": item["field"],
                    "verdict": "ambiguous",
                    "source_field": "",
                    "evidence_text": "",
                    "reason": "批量集成测试保持程序歧义结论",
                }
                for item in kwargs["field_evidence"]
            ],
            "hard_blockers": [],
            "model_confidence": "high",
        }


if __name__ == "__main__":
    main()
