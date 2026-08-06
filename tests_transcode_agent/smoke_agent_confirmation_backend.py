from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app import db
from fangzheng_web_app import create_app
from fangzheng_web_app.transcode_agent_service import (
    FORMAL_RESULT_HEADER,
    PENDING_RESULT_HEADER,
    SYSTEM_ANALYSIS_HEADER,
    _build_confirmation_items,
    _business_rule_plain_text,
    _business_rule_source_label,
    _load_runtime,
    _refresh_analysis_after_semantic_overrides,
    _trace_record_state,
    activate_transcode_agent_pending_rule,
    analyze_spec,
    confirm_transcode_agent_item,
    delete_transcode_agent_pending_rule,
    list_transcode_agent_confirmations,
    list_transcode_agent_pending_rules,
    reevaluate_transcode_agent_confirmations,
    refresh_transcode_agent_audit_sheet,
    update_transcode_agent_pending_rule,
    verify_all_transcode_agent_rows,
    verify_transcode_agent_row,
)
from fangzheng_web_app.transcode_semantic_overrides import apply_confirmed_semantic_overrides
from fangzheng_web_app.transcode_customer_rule_admin import (
    CustomerRuleMaintenanceError,
    ensure_customer_rule_maintenance_tables,
    make_customer_key,
)


def _analysis() -> dict:
    steps = {
        "step1_glue_code": "2B",
        "step2_thick_code": "00800",
        "step3_copper_code": "11",
        "step4_size_code": "37004900",
        "step5_glue_cat_code": "Y",
        "step6_copper_type_code": "W",
        "step7_grade_code": "A1",
        "step8_tc_code": "T",
        "step9_struct_code": "*",
        "errors": [],
    }
    evidence = []
    field_codes = [
        ("glue", "胶系", "2B"),
        ("thickness", "厚度", "00800"),
        ("copper", "铜厚", "11"),
        ("size", "尺寸", "37004900"),
        ("glue_category", "胶水类别", "Y"),
        ("copper_type", "铜箔类型", "W"),
        ("grade", "基板级别", "A1"),
        ("total_core", "总/芯厚", "T"),
    ]
    for key, label, code in field_codes:
        evidence.append(
            {
                "field_key": key,
                "field": label,
                "code": code,
                "score": 60 if key == "grade" else 100,
                "gate": True,
                "hit_type": "规则冲突" if key == "grade" else "标准解析",
                "source": "测试",
                "evidence": "grade_code: A1 vs AC" if key == "grade" else code,
            }
        )
    return {
        "row": 2,
        "customer_code": "100001",
        "customer": "确认中心测试客户",
        "spec": "NY2150 0.8mm 1/1 37*49 HTE 含铜 汽车板",
        "context_text": "订单备注：汽车板",
        "status": "待确认",
        "formal_code": "",
        "candidate_code": "2B008001137004900YWA1T*",
        "overall_score": 60,
        "reason": "Agent规则冲突：grade_code: A1 vs AC",
        "summary": "",
        "field_evidence": evidence,
        "applied_rules": [],
        "conflicts": ["grade_code: A1 vs AC"],
        "engine_steps": steps,
    }


def main() -> None:
    assert (
        _business_rule_plain_text("业务正式规则；TGM-MASTER-0051；2HNN NY2150H 2H")
        == "2HNN NY2150H 2H"
    )
    assert _business_rule_source_label("transcode_rules.xlsx/铜箔规格") == "基础映射"
    assert _business_rule_source_label("TAR-CYB-20260721-002") == "已确认规则"
    assert (
        _trace_record_state(
            {"excel_row": 2, "transcode_status": "已出码需核对"},
            {"confirmed"},
            set(),
        )
        == "automatic"
    )
    assert (
        _trace_record_state(
            {"excel_row": 2, "transcode_status": "人工已核对"},
            {"confirmed"},
            set(),
        )
        == "confirmed"
    )
    engine, tables, rules, mappings, _, _ = _load_runtime()
    runtime_analysis = analyze_spec(
        engine,
        tables,
        rules,
        "NY2150 0.8mm 1/1 37*49 HTE 含铜",
        agent_mapping_tables=mappings,
        customer="确认中心测试客户",
        excel_row=2,
    )
    applied, conflicts = apply_confirmed_semantic_overrides(
        engine,
        tables,
        runtime_analysis,
        [
            {
                "status": "命中",
                "rule_id": "TCR-RUNTIME-TEST",
                "business_field": "基板级别",
                "source_text": "订单备注汽车板时AC",
                "source_column": "确认中心",
                "model": "确认中心人工规则",
                "target_fields": ["grade_intent"],
                "normalized_values": ["AC"],
                "priority": 200,
                "condition_results": [{"field": "订单备注", "matched": True}],
            }
        ],
        allow_order_remark_priority=True,
    )
    assert not conflicts
    assert applied and applied[0]["source"] == "已确认人工长期规则", applied
    _refresh_analysis_after_semantic_overrides(runtime_analysis, applied, [])
    assert runtime_analysis["status"] == "成功", runtime_analysis
    assert runtime_analysis["overall_score"] == 100, runtime_analysis["field_evidence"]
    assert runtime_analysis["engine_steps"]["step7_grade_code"] == "AC"

    protected_analysis = {
        "engine_steps": {"step7_grade_code": "AC", "errors": []},
        "applied_rules": [
            {
                "field": "grade_code",
                "new": "AC",
                "rule_type": "客户人工长期规则",
            }
        ],
    }
    protected_applied, protected_conflicts = apply_confirmed_semantic_overrides(
        engine,
        tables,
        protected_analysis,
        [
            {
                "status": "命中",
                "rule_id": "MODEL-SHOULD-NOT-OVERRIDE",
                "business_field": "基板级别",
                "source_text": "模型结果A1",
                "source_column": "CCL特殊规则",
                "model": "deepseek",
                "target_fields": ["grade_intent"],
                "normalized_values": ["A1"],
                "priority": 200,
                "condition_results": [{"field": "订单备注", "matched": True}],
            }
        ],
        allow_order_remark_priority=True,
    )
    assert not protected_applied
    assert protected_conflicts, protected_conflicts
    assert protected_analysis["engine_steps"]["step7_grade_code"] == "AC"

    model_normalized_analysis = {
        "engine_steps": {"step7_grade_code": "A1", "errors": []},
        "applied_rules": [],
        "spec": "NY2150 0.8mm 1/1 37*49 HTE 含铜",
        "conflicts": [],
    }
    model_normalized_applied, model_normalized_conflicts = apply_confirmed_semantic_overrides(
        engine,
        tables,
        model_normalized_analysis,
        [
            {
                "status": "命中",
                "rule_id": "TCR-MODEL-NORMALIZED-TEST",
                "business_field": "基板级别",
                "source_text": "订单备注指向汽车板时基板级别为AC",
                "source_column": "确认中心",
                "model": "确认中心人工规则",
                "target_fields": ["grade_intent"],
                "normalized_values": ["AC"],
                "priority": 200,
                "condition_results": [{"field": "订单备注", "matched": True}],
                "model_normalized": True,
            }
        ],
        allow_order_remark_priority=True,
    )
    assert model_normalized_applied, model_normalized_applied
    assert not model_normalized_conflicts, model_normalized_conflicts
    _refresh_analysis_after_semantic_overrides(
        model_normalized_analysis,
        model_normalized_applied,
        model_normalized_conflicts,
    )
    assert model_normalized_analysis["status"] == "待确认", model_normalized_analysis
    assert model_normalized_analysis["overall_score"] == 98, model_normalized_analysis
    assert model_normalized_analysis["engine_steps"]["step7_grade_code"] == "AC"
    assert any(
        item.get("model_normalized") and item.get("new") == "AC"
        for item in model_normalized_analysis["applied_rules"]
    )

    with TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        original_database_path = db.DATABASE_PATH
        db.DATABASE_PATH = temp_dir / "app.db"
        try:
            db.init_db()
            ensure_customer_rule_maintenance_tables()
            output_path = temp_dir / "result.xlsx"
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "转码需求表"
            worksheet.append(
                [
                    "客户简称",
                    "客户规格",
                    FORMAL_RESULT_HEADER,
                    PENDING_RESULT_HEADER,
                    "22位码值差异",
                    "结果对比",
                    "转码状态",
                    "人工确认提示",
                    SYSTEM_ANALYSIS_HEADER,
                ]
            )
            worksheet.append(
                [
                    "确认中心测试客户",
                    "NY2150 0.8mm 1/1 37*49 HTE 含铜 汽车板",
                    "",
                    "2B008001137004900YWA1T*",
                    "",
                    "待人工确认",
                    "待人工确认",
                    "待确认：基板级别冲突",
                    "基板级别存在规则冲突",
                ]
            )
            worksheet.append(
                [
                    "无需确认测试客户",
                    "NY2150 1.0mm 1/1 37*49 HTE 含铜",
                    "2B010001137004900YWA1T*",
                    "",
                    "2B010001137004900YWA1T",
                    True,
                    "可直接采用",
                    "",
                    "已验证基础默认规则",
                ]
            )
            worksheet.append(
                [
                    "对比异常测试客户",
                    "NY2150 1.0mm 1/1 37*49 HTE 含铜",
                    "2B010001137004900YWA1T*",
                    "",
                    "2B010001137004900YWA1（AC）T",
                    False,
                    "待人工确认",
                    "待确认：基板级别不一致",
                    "100分正式码与测试文件正确码不一致",
                ]
            )
            evidence_sheet = workbook.create_sheet("字段证据链")
            evidence_sheet.append(
                ["行号", "客户代码", "客户", "规格", "状态", "字段", "原始值", "编码片段", "置信度", "命中方式", "规则来源", "证据", "规则ID", "规则类型", "来源行号"]
            )
            evidence_sheet.append([3, "", "无需确认测试客户", "NY2150 1.0mm 1/1 37*49 HTE 含铜", "成功", "基板级别", "基础默认", "A1", 100, "已验证基础默认", "基础规则", "基础默认可执行", "", "base", ""])
            workbook.save(output_path)

            job_id = db.create_job(
                "tester",
                "confirmation.xlsx",
                str(output_path),
                "test",
                feature="transcode_agent",
            )
            db.update_job_status(
                job_id,
                status="awaiting_confirmation",
                stored_result_path=str(output_path),
                success_count=1,
                fail_count=1,
                confirm_count=1,
            )
            analysis = _analysis()
            items = _build_confirmation_items([analysis])
            assert len(items) == 1, items
            item_ids = db.replace_transcode_agent_confirmation_items(job_id, "tester", items)

            listing = list_transcode_agent_confirmations(job_id, "tester")
            assert listing["confirm_count"] == 1
            assert listing["items"][0]["field_key"] == "grade"
            assert listing["record_total"] == 3, listing
            listing_states = [record["record_state"] for record in listing["records"]]
            assert listing_states == [
                "pending",
                "automatic",
                "automatic",
            ], [(record["excel_row"], record["comparison"], record["transcode_status"], record["record_state"]) for record in listing["records"]]
            assert any(str(record["comparison"]).lower() == "false" for record in listing["records"]), listing["records"]

            result = confirm_transcode_agent_item(
                item_ids[0],
                "tester",
                confirmed_code="AC",
                basis="订单备注明确为汽车板",
                save_long_term=True,
                long_term_rule={
                    "condition_field": "订单备注",
                    "condition_operator": "contains_any",
                    "condition_value": "汽车板",
                    "second_confirmed": True,
                },
            )
            assert result["status"] == "成功", result
            assert result["formal_code"] == "2B008001137004900YWACT*", result
            assert result["remaining"] == 0, result
            assert str(result["long_term_rule_id"]).startswith("TCR-"), result
            assert result["pending_rule_id"], result
            confirmed_listing = list_transcode_agent_confirmations(job_id, "tester")
            assert confirmed_listing["success_count"] == 2
            assert confirmed_listing["verify_count"] == 2, confirmed_listing
            assert confirmed_listing["items"][0]["score"] == 100
            assert [record["record_state"] for record in confirmed_listing["records"]] == [
                "confirmed",
                "automatic",
                "automatic",
            ]

            job = db.get_job(job_id)
            assert job["status"] == "awaiting_verification", dict(job)
            assert job["success_count"] == 2, dict(job)
            assert job["fail_count"] == 0, dict(job)
            assert job["confirm_count"] == 0, dict(job)
            assert job["verify_count"] == 2, dict(job)

            batch_verify_result = verify_all_transcode_agent_rows(
                job_id,
                "tester",
                basis="测试批量核对",
            )
            assert batch_verify_result["verified_count"] == 2, batch_verify_result
            verified_listing = list_transcode_agent_confirmations(job_id, "tester")
            assert verified_listing["verify_count"] == 0, verified_listing
            assert {record["record_state"] for record in verified_listing["records"]} == {
                "confirmed"
            }, verified_listing["records"]
            assert db.get_job(job_id)["status"] == "completed"

            refresh_transcode_agent_audit_sheet(job_id)
            result_book = openpyxl.load_workbook(output_path, data_only=True)
            result_sheet = result_book["转码需求表"]
            assert result_sheet.cell(2, 3).value == "2B008001137004900YWACT*"
            assert result_sheet.cell(2, 4).value is None
            assert result_sheet.cell(2, 5).value == "人工已核对"
            assert result_sheet.cell(2, 6).value == "人工已核对"
            assert result_sheet.cell(2, 7).value == "人工已核对"
            assert str(result_sheet.cell(2, 9).value).startswith("当前任务人工确认：")
            for column in (3, 5, 6, 7, 8, 9):
                fill = result_sheet.cell(2, column).fill.fgColor.rgb or ""
                assert fill.endswith("C8E6C9"), (column, fill)
            assert "人工确认审计" in result_book.sheetnames
            assert "人工核对审计" in result_book.sheetnames
            audit_sheet = result_book["人工确认审计"]
            audit_headers = [cell.value for cell in audit_sheet[1]]
            assert audit_sheet.cell(2, audit_headers.index("是否待生效规则") + 1).value == "是"
            assert (
                audit_sheet.cell(2, audit_headers.index("待生效规则ID") + 1).value
                == result["long_term_rule_id"]
            )
            assert (
                audit_sheet.cell(2, audit_headers.index("待生效记录ID") + 1).value
                == str(result["pending_rule_id"])
            )
            with db.get_connection() as connection:
                saved_override = connection.execute(
                    "SELECT rule_json FROM transcode_customer_rule_overrides WHERE rule_id = ?",
                    (result["long_term_rule_id"],),
                ).fetchone()
            assert saved_override is None
            pending_listing = list_transcode_agent_pending_rules(
                "admin-tester",
                include_all=True,
            )
            assert len(pending_listing) == 1, pending_listing
            assert pending_listing[0]["id"] == result["pending_rule_id"]
            assert pending_listing[0]["condition_summary"] == "订单备注 包含任一 汽车板"
            db.create_user("admin-tester", role="admin")
            try:
                activate_transcode_agent_pending_rule(
                    result["pending_rule_id"],
                    "tester",
                )
            except CustomerRuleMaintenanceError:
                pass
            else:
                raise AssertionError("普通用户不应能确认待生效规则")
            update_transcode_agent_pending_rule(
                result["pending_rule_id"],
                "tester",
                {
                    "customer_code": "100001",
                    "customer_name": "确认中心测试客户",
                    "business_field": "基板级别",
                    "target_field": "grade_intent",
                    "target_value": "AC",
                    "condition_field": "订单备注",
                    "condition_operator": "contains_any",
                    "condition_value": "汽车板",
                    "source_text": "订单备注出现汽车板时基板级别为AC（测试编辑）",
                    "priority": "200",
                    "enabled": "1",
                    "semantic_enabled": "1",
                },
            )
            updated_pending = list_transcode_agent_pending_rules(
                "admin-tester",
                include_all=True,
            )[0]
            assert "测试编辑" in updated_pending["source_text"]
            activate_transcode_agent_pending_rule(
                result["pending_rule_id"],
                "admin-tester",
            )
            try:
                activate_transcode_agent_pending_rule(
                    result["pending_rule_id"],
                    "admin-tester",
                )
            except (LookupError, CustomerRuleMaintenanceError):
                pass
            else:
                raise AssertionError("已生效的待生效规则不应重复确认")
            delete_target = db.create_transcode_agent_pending_rule(
                rule_id="TCR-PENDING-DELETE-TEST",
                rule_json=(
                    '{"rule_id":"TCR-PENDING-DELETE-TEST",'
                    '"customer_code":"100001","customer_name":"确认中心测试客户",'
                    '"business_field":"基板级别","target_fields":["grade_intent"],'
                    '"normalized_values":["AC"],"conditions":['
                    '{"field":"订单备注","operator":"contains_any","value":"删除测试"}],'
                    '"source_text":"删除测试"}'
                ),
                employee_id="tester",
                customer_code="100001",
                customer_name="确认中心测试客户",
                business_field="基板级别",
                target_value="AC",
                condition_summary="订单备注 包含任一 删除测试",
                source_task_id=job_id,
                source_excel_row=2,
            )
            delete_transcode_agent_pending_rule(delete_target, "tester")
            assert (
                db.get_transcode_agent_pending_rule(delete_target)["status"]
                == "deleted"
            )
            with db.get_connection() as connection:
                saved_override = connection.execute(
                    "SELECT rule_json FROM transcode_customer_rule_overrides WHERE rule_id = ?",
                    (result["long_term_rule_id"],),
                ).fetchone()
            assert saved_override is not None
            assert list_transcode_agent_pending_rules("tester") == []

            app = create_app()
            app.testing = True
            db.change_user_password("tester", "tester-ready")
            db.change_user_password("admin-tester", "admin-ready")
            client = app.test_client()
            with client.session_transaction() as session:
                session["employee_id"] = "tester"
            page_response = client.get(
                f"/features/transcode-agent/confirmations/{job_id}"
            )
            assert page_response.status_code == 200, (
                page_response.status_code,
                page_response.location,
                page_response.data[:500],
            )
            assert "转码记录与人工确认".encode("utf-8") in page_response.data
            assert "已出码需核对".encode("utf-8") in page_response.data
            admin_client = app.test_client()
            with admin_client.session_transaction() as session:
                session["employee_id"] = "admin-tester"
            pending_page_response = admin_client.get(
                "/admin/transcode-rule-center",
                query_string={"section": "submitted"},
            )
            assert pending_page_response.status_code == 200, (
                pending_page_response.status_code,
                pending_page_response.location,
                pending_page_response.data[:800],
            )
            assert "已提交待生效规则".encode("utf-8") in pending_page_response.data
            api_response = client.get(
                f"/api/transcode-agent/jobs/{job_id}/confirmations"
            )
            assert api_response.status_code == 200
            assert api_response.get_json()["confirm_count"] == 0
            assert api_response.get_json()["verify_count"] == 0
            assert api_response.get_json()["record_total"] == 3
            rule_page_response = client.get(
                "/admin/transcode-agent-customer-rules",
                query_string={
                    "customer_key": make_customer_key("100001", "确认中心测试客户"),
                    "business_field": "基板级别",
                    "rule_id": result["long_term_rule_id"],
                },
            )
            assert rule_page_response.status_code == 200
            assert result["long_term_rule_id"].encode("utf-8") in rule_page_response.data
            assert "确认中心待生效规则".encode("utf-8") in rule_page_response.data
            assert "包含任一".encode("utf-8") in rule_page_response.data

            reeval_path = temp_dir / "reevaluate.xlsx"
            reeval_book = openpyxl.Workbook()
            reeval_sheet = reeval_book.active
            reeval_sheet.title = "转码需求表"
            reeval_sheet.append(
                [
                    "客户简称",
                    "品名",
                    "客户规格",
                    "订单备注",
                    FORMAL_RESULT_HEADER,
                    PENDING_RESULT_HEADER,
                    "22位码值差异",
                    "结果对比",
                    "转码状态",
                    "人工确认提示",
                    SYSTEM_ANALYSIS_HEADER,
                ]
            )
            reeval_spec = '覆铜板 36±3MIL 2/2OZ 74*49" TG≥150 HTE 不连铜 NY2150H ANTI-CAF'
            reeval_sheet.append(
                [
                    "广东依顿",
                    "2H010502274004900RWACTA",
                    reeval_spec,
                    "下汽车板",
                    "",
                    "2H010502274004900RWACT*XXXXXX",
                    "待人工确认",
                    "待人工确认",
                    "待人工确认",
                    "待确认：基板级别98分",
                    "历史评分快照",
                ]
            )
            reeval_book.save(reeval_path)
            reeval_job_id = db.create_job(
                "tester",
                "reevaluate.xlsx",
                str(reeval_path),
                "test",
                feature="transcode_agent",
            )
            db.update_job_status(
                reeval_job_id,
                status="awaiting_confirmation",
                stored_result_path=str(reeval_path),
                success_count=0,
                fail_count=1,
                confirm_count=1,
            )
            old_analysis = _analysis()
            old_analysis.update(
                {
                    "customer": "广东依顿",
                    "customer_code": "",
                    "spec": reeval_spec,
                    "context_text": "下汽车板",
                    "candidate_code": "2H010502274004900RWACT*XXXXXX",
                    "overall_score": 98,
                }
            )
            old_analysis["field_evidence"][-2]["score"] = 98
            old_analysis["field_evidence"][-2]["code"] = "AC"
            old_analysis["field_evidence"][-2]["evidence"] = "如果订单备注是汽车板则=AC"
            old_analysis["engine_steps"]["step7_grade_code"] = "AC"
            old_items = _build_confirmation_items([old_analysis])
            old_items[0].update(
                {
                    "customer_name": "广东依顿",
                    "spec": reeval_spec,
                    "context_text": "下汽车板",
                    "current_code": "AC",
                    "score": 98,
                    "pending_code": "2H010502274004900RWACT*XXXXXX",
                }
            )
            db.replace_transcode_agent_confirmation_items(
                reeval_job_id,
                "tester",
                old_items,
            )
            reeval_result = reevaluate_transcode_agent_confirmations(reeval_job_id, "tester")
            assert reeval_result["resolved_rows"] == 1, reeval_result
            assert reeval_result["remaining"] == 0, reeval_result
            reeval_listing = list_transcode_agent_confirmations(reeval_job_id, "tester")
            assert reeval_listing["status"] == "awaiting_verification", reeval_listing
            assert reeval_listing["verify_count"] == 1, reeval_listing
            assert reeval_listing["items"][0]["status"] == "auto_resolved", reeval_listing
            assert reeval_listing["items"][0]["score"] == 100, reeval_listing
            assert reeval_listing["records"][0]["record_state"] == "automatic", reeval_listing
            reeval_output = openpyxl.load_workbook(reeval_path, data_only=True)["转码需求表"]
            assert reeval_output.cell(2, 5).value == "2H010502274004900RWACT*XXXXXX"
            assert reeval_output.cell(2, 6).value is None
            assert reeval_output.cell(2, 8).value is True
            assert reeval_output.cell(2, 9).value == "已出码需核对"
            verify_transcode_agent_row(
                reeval_job_id,
                2,
                "tester",
                code=reeval_output.cell(2, 5).value,
                basis="重评后核对无误",
            )
            assert db.get_job(reeval_job_id)["status"] == "completed"
        finally:
            db.DATABASE_PATH = original_database_path

    print("confirmation backend smoke passed")


if __name__ == "__main__":
    main()
