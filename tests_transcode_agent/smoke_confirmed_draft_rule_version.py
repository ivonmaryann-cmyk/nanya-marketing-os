from pathlib import Path
import sys
from tempfile import TemporaryDirectory

import openpyxl
from werkzeug.datastructures import FileStorage

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app import db
from fangzheng_web_app import transcode_agent_rules as agent_rules


DRAFT_PATH = ROOT / "docs/develop0707/客户特殊规则结构化草稿_按原表行_20260708.xlsx"


def main() -> None:
    with TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        original_rules_dir = agent_rules.TRANSCODE_AGENT_RULES_DIR
        original_versions_dir = agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR
        original_db_path = db.DATABASE_PATH
        try:
            agent_rules.TRANSCODE_AGENT_RULES_DIR = temp_dir / "rules/transcode_agent"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = agent_rules.TRANSCODE_AGENT_RULES_DIR / "versions"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
            db.DATABASE_PATH = temp_dir / "storage/app.db"
            db.DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
            db.init_db()

            with DRAFT_PATH.open("rb") as source:
                upload = FileStorage(
                    stream=source,
                    filename=DRAFT_PATH.name,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                version = agent_rules.save_new_transcode_agent_rule_version(
                    upload,
                    updated_by="smoke",
                    remark="confirmed draft smoke",
                )

            version_dir = agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR / version
            machine_rule_path = version_dir / agent_rules.RULE_FILENAME
            original_path = version_dir / agent_rules.ORIGINAL_FILENAME
            structured_draft_path = version_dir / agent_rules.STRUCTURED_DRAFT_FILENAME
            assert version_dir.exists(), version_dir
            assert machine_rule_path.exists(), machine_rule_path
            assert original_path.exists(), original_path
            assert structured_draft_path.exists(), structured_draft_path
            assert agent_rules.get_active_transcode_agent_rule_version() == version

            loaded_rules = agent_rules.load_transcode_agent_rules(version)
            assert len(loaded_rules) == 200, len(loaded_rules)
            assert not any(rule.get("覆盖字段") == "struct_code" for rule in loaded_rules)
            assert not any("订单" in (rule.get("条件文本", "") + rule.get("覆盖值", "")) for rule in loaded_rules)

            workbook = openpyxl.load_workbook(machine_rule_path, data_only=True)
            assert "机器规则" in workbook.sheetnames
            assert "转换说明" in workbook.sheetnames
            assert workbook["机器规则"].max_row == 201

            print(f"smoke passed version={version} rules={len(loaded_rules)}")
        finally:
            agent_rules.TRANSCODE_AGENT_RULES_DIR = original_rules_dir
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = original_versions_dir
            db.DATABASE_PATH = original_db_path


if __name__ == "__main__":
    main()
