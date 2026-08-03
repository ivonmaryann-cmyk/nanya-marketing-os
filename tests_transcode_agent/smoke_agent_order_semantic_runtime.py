from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app.transcode_agent_service import _append_order_semantic_model_sheet
from fangzheng_web_app.transcode_order_semantic_model import (
    OrderSemanticRuntime,
    build_order_semantic_cache_key,
    normalize_order_shadow,
    should_normalize_order,
)


class FakeOrderSemanticClient:
    def __init__(self) -> None:
        self.calls: list[dict] = []

    def normalize(self, **kwargs):
        self.calls.append(kwargs)
        return {
            "schema_version": "1.0",
            "task_type": "order_normalization",
            "material_scope": "CCL",
            "semantic_items": [
                {
                    "semantic_type": "keyword_present",
                    "target_field": "grade_intent",
                    "normalized_value": "automotive",
                    "stated_target_value": "",
                    "conditions": [],
                    "source_field": "订单备注",
                    "evidence_text": "汽车板",
                    "confidence": "high",
                    "deterministic_preferred": False,
                }
            ],
            "ambiguities": [],
            "missing_inputs": [],
            "model_confidence": "high",
        }


def main() -> None:
    evaluations = [
        {
            "rule_id": "CSR-TEST-001",
            "status": "命中",
            "business_field": "基板级别",
            "target_fields": ["grade_intent"],
            "normalized_values": ["automotive"],
            "source_text": "订单备注有汽车板时下汽车板料号",
        }
    ]
    source_fields = {"订单规格": "NY2150 0.8 1/1 37*49", "订单备注": "汽车板"}
    evaluations[0]["condition_results"] = [
        {"field": "订单备注", "matched": True}
    ]
    assert should_normalize_order(evaluations, source_fields["订单备注"])
    assert not should_normalize_order(evaluations, "")
    assert not should_normalize_order([], source_fields["订单备注"])
    assert build_order_semantic_cache_key("100001", "测试客户", source_fields, evaluations) == (
        build_order_semantic_cache_key("100001", "测试客户", source_fields, evaluations)
    )

    client = FakeOrderSemanticClient()
    runtime = OrderSemanticRuntime(
        mode="shadow",
        client=client,
        model="deepseek-v4-pro",
        max_calls=10,
    )
    result = normalize_order_shadow(
        runtime,
        customer_code="100001",
        customer_name="测试客户",
        source_fields=source_fields,
        semantic_evaluations=evaluations,
    )
    assert result["semantic_items"][0]["normalized_value"] == "automotive"
    assert len(client.calls) == 1
    assert client.calls[0]["task_type"] == "order_normalization"

    workbook = openpyxl.Workbook()
    _append_order_semantic_model_sheet(
        workbook,
        [
            {
                "row": 2,
                "customer_code": "100001",
                "customer": "测试客户",
                "spec": source_fields["订单规格"],
                "order_semantic_model": {
                    "mode": "shadow",
                    "status": "成功",
                    "model": "deepseek-v4-pro",
                    "cached": False,
                    "source_fields": source_fields,
                    "rule_ids": ["CSR-TEST-001"],
                    "result": result,
                },
            }
        ],
    )
    sheet = workbook["模型实时语义标准化"]
    assert sheet.max_row == 2
    assert sheet.cell(2, 6).value == "成功"
    assert "automotive" in sheet.cell(2, 12).value
    print("order semantic runtime smoke passed calls=1 export_rows=1 formal_code_unchanged")


if __name__ == "__main__":
    main()
