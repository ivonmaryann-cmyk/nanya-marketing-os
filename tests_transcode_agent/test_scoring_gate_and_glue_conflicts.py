from __future__ import annotations

from contextlib import contextmanager
import importlib.util
from pathlib import Path
import sqlite3

from openpyxl import Workbook, load_workbook

from fangzheng_web_app import transcode_rule_center
from fangzheng_web_app.transcode_agent_glue_resolver import (
    clear_agent_glue_index_cache,
    is_retired_agent_glue_mapping,
    resolve_agent_glue,
)
from fangzheng_web_app.transcode_agent_service import (
    _apply_agent_field_mappings,
    _apply_agent_rules,
    _build_code_from_steps,
    _enforce_retired_glue_runtime_guard,
    _remove_retired_glue_mappings,
    _score_field,
    _validate_confirmation_code,
)
from fangzheng_web_app.transcode_confirmation_policy import decide_confirmation


ROOT = Path(__file__).resolve().parents[1]


def _load_script_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _read_mapping_workbook(path: Path) -> dict[str, list[dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: dict[str, list[dict]] = {}
    for worksheet in workbook.worksheets:
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        tables[worksheet.title] = [
            {
                headers[index]: value
                for index, value in enumerate(values)
                if index < len(headers) and headers[index] and value not in (None, "")
            }
            for values in worksheet.iter_rows(min_row=2, values_only=True)
            if any(value not in (None, "") for value in values)
        ]
    workbook.close()
    return tables


def _write_rows(path: Path, title: str, headers: list[str], rows: list[list[str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def _is_enabled(row: dict) -> bool:
    return str(row.get("启用", "是") or "是").strip().lower() not in {
        "否",
        "false",
        "0",
        "停用",
        "禁用",
    }


def _active_agent_mapping_path() -> Path:
    with sqlite3.connect(ROOT / "storage/app.db") as connection:
        version = connection.execute(
            "SELECT value FROM settings WHERE key = ?",
            ("active_transcode_agent_rule_version",),
        ).fetchone()[0]
    return (
        ROOT
        / "storage/rules/transcode_agent/versions"
        / version
        / "transcode_agent_mapping_tables.xlsx"
    )


def _rule(*, source: str, rule_type: str, rule_id: str = "RULE-1") -> dict:
    return {
        "rule_id": rule_id,
        "source": source,
        "rule_type": rule_type,
        "text": "测试规则",
    }


def test_conflict_is_scored_before_formal_agent_override() -> None:
    score, hit_type, _, note = _score_field(
        "glue",
        "AA",
        {},
        [],
        {"glue_code": _rule(source="最新版胶系名称优先命中", rule_type="Agent胶系主数据映射")},
        ["最新版胶系主表同名多码待补口径：NYTEST候选AA/BB"],
    )

    assert score == 60
    assert hit_type == "规则冲突"
    assert "同名多码" in note


def test_historical_correct_code_sample_cannot_score_100() -> None:
    score, hit_type, _, note = _score_field(
        "grade",
        "AC",
        {},
        [],
        {
            "grade_code": _rule(
                source="正确码回归客户字段映射",
                rule_type="辅助客户字段映射",
                rule_id="AFM-HISTORY-1",
            )
        },
        [],
    )

    assert score == 99
    assert hit_type == "历史样本建议"
    assert "不能单独正式出码" in note

    decision = decide_confirmation(
        errors=[],
        conflicts=[],
        candidate_code="2B008001137004900YWA1T*",
        field_evidence=[
            {
                "field_key": "grade",
                "field": "基板级别",
                "score": score,
                "gate": True,
                "evidence": note,
                "rule_id": "AFM-HISTORY-1",
            }
        ],
    )
    assert decision["status"] == "待确认"
    assert decision["formal_code"] == ""


def test_business_formal_rules_remain_100_and_unknown_sources_do_not() -> None:
    base_score, base_type, _, _ = _score_field(
        "glue", "2B", {}, [], {}, []
    )
    formal_score, formal_type, _, _ = _score_field(
        "grade",
        "AC",
        {},
        [],
        {
            "grade_code": _rule(
                source="客户特殊规则正式表",
                rule_type="确认草稿机器规则",
            )
        },
        [],
    )
    unknown_score, unknown_type, _, _ = _score_field(
        "grade",
        "AC",
        {},
        [],
        {"grade_code": _rule(source="临时推断", rule_type="未知规则")},
        [],
    )

    assert base_score == 99
    assert base_type == "解析来源待确认"
    assert formal_score == 100
    assert formal_type == "业务正式规则"
    assert unknown_score == 99
    assert unknown_type == "规则来源待确认"


def test_confirmed_base_mapping_and_defaults_have_explicit_100_sources() -> None:
    steps = {
        "glue_model": "NY2150",
        "step1_glue_code": "2B",
        "thickness_raw": "0.8mm",
        "thickness_mm": 0.8,
        "copper_spec_raw": "1/1",
        "size_w": 41.0,
        "size_h": 49.0,
        "size_note": "标准尺寸",
        "glue_category": "普通",
        "order_type": "总厚",
    }

    glue = _score_field("glue", "2B", steps, [], {}, [])
    grade = _score_field("grade", "A1", steps, [], {}, [])
    copper_type = _score_field("copper_type", "W", steps, [], {}, [])

    assert glue[:2] == (100, "正式基础映射")
    assert grade[:2] == (99, "解析来源待确认")
    assert copper_type[:3] == (
        100,
        "已确认默认规则",
        "BASE-DEFAULT-COPPER-TYPE-W",
    )

    automotive_steps = dict(steps, agent_input_text="NY3150HF 汽车专用")
    automotive = _score_field("grade", "AC", automotive_steps, [], {}, [])
    assert automotive[:3] == (
        100,
        "业务确定性等级规则",
        "BASE-GRADE-AUTOMOTIVE-AC",
    )


def test_model_normalization_and_priority_candidate_never_score_100() -> None:
    model_score = _score_field(
        "grade",
        "AC",
        {},
        [],
        {
            "grade_code": _rule(
                source="已批准模型语义映射",
                rule_type="已批准模型语义规则/受控映射",
            )
        },
        [],
    )
    priority_score = _score_field(
        "glue",
        "AA",
        {},
        [],
        {
            "glue_code": _rule(
                source="最新版胶系名称优先命中",
                rule_type="Agent胶系主数据映射",
            )
        },
        [],
    )

    assert model_score[0] == 98
    assert model_score[1] == "模型语义标准化"
    assert priority_score[0] == 99
    assert priority_score[1] == "胶系候选口径"


def _machine_rule(rule_id: str, output: str, priority: int) -> dict:
    return {
        "规则ID": rule_id,
        "启用": "是",
        "待确认": "否",
        "强制执行": "是",
        "物料类别": "CCL",
        "客户简称": "测试客户",
        "条件胶系": "NY2150",
        "覆盖字段": "grade_code",
        "覆盖值": output,
        "优先级": str(priority),
        "规则文本": f"NY2150基板级别={output}",
        "命中来源": "客户特殊规则正式表",
    }


def test_same_machine_condition_multi_output_conflicts_before_priority_override() -> None:
    steps = {"glue_model": "NY2150", "step7_grade_code": "A1", "errors": []}
    applied, conflicts = _apply_agent_rules(
        [
            _machine_rule("RULE-HIGH", "AC", 999),
            _machine_rule("RULE-LOW", "A1", 1),
        ],
        "",
        "测试客户",
        "NY2150 0.8mm 1/1 41x49",
        "",
        steps,
        [],
    )

    assert steps["step7_grade_code"] == "AC"
    assert applied[0]["rule_id"] == "RULE-HIGH"
    assert any("同一条件存在多个输出A1/AC" in conflict for conflict in conflicts)
    score = _score_field(
        "grade",
        "AC",
        steps,
        [],
        {"grade_code": applied[0]},
        conflicts,
    )
    assert score[0] == 60
    assert score[1] == "规则冲突"


def test_historical_mapping_conflict_downgrades_without_overriding_formal_rule() -> None:
    steps = {
        "glue_model": "NY2150",
        "agent_glue_name": "NY2150",
        "step7_grade_code": "AC",
        "errors": [],
    }
    formal = {
        "rule_id": "FORMAL-AC",
        "field": "grade_code",
        "old": "A1",
        "new": "AC",
        "text": "正式规则输出AC",
        "source": "客户特殊规则正式表",
        "rule_type": "确认草稿机器规则",
    }
    mappings = {
        "客户字段映射": [
            {
                "映射ID": "HISTORY-A1",
                "启用": "是",
                "客户简称": "测试客户",
                "条件胶系": "NY2150",
                "覆盖字段": "grade_code",
                "覆盖值": "A1",
                "规则文本": "历史正确码为A1",
            }
        ]
    }

    evidence = _apply_agent_field_mappings(
        object(),
        mappings,
        "",
        "测试客户",
        "NY2150 0.8mm 1/1 41x49",
        "",
        steps,
        [],
        [formal],
    )

    assert steps["step7_grade_code"] == "AC"
    assert evidence and evidence[0]["historical_suggested"] == "A1"
    score = _score_field(
        "grade",
        "AC",
        steps,
        [],
        {"grade_code": evidence[0]},
        [],
    )
    assert score[0] == 99
    assert score[1] == "历史样本建议"


def test_selection_rule_priority_does_not_hide_multi_output_conflict() -> None:
    mappings = {
        "Agent胶系选择规则": [
            {
                "映射ID": "SELECT-HIGH",
                "启用": "是",
                "胶系名称": "NYTEST",
                "输出胶系代码": "AA",
                "条件客户简称": "测试客户",
                "优先级": "999",
            },
            {
                "映射ID": "SELECT-LOW",
                "启用": "是",
                "胶系名称": "NYTEST",
                "输出胶系代码": "BB",
                "条件客户简称": "测试客户",
                "优先级": "1",
            },
        ]
    }

    result = resolve_agent_glue(
        mappings,
        spec="NYTEST 0.8mm 1/1 41x49",
        customer_name="测试客户",
    )

    assert result and result["code"] == "AA"
    assert result["uncertain"] is True
    assert {item["code"] for item in result["candidates"]} == {"AA", "BB"}
    assert "不使用优先级消除冲突" in result["conflict"]


def test_runtime_and_manual_confirmation_cannot_restore_ny_a1_to_2z() -> None:
    steps = {
        "glue_model": "NY-A1 白纹改善",
        "step1_glue_code": "2Z",
        "step2_thick_code": "00800",
        "step3_copper_code": "11",
        "step4_size_code": "41004900",
        "step5_glue_cat_code": "Y",
        "step6_copper_type_code": "W",
        "step7_grade_code": "A1",
        "step8_tc_code": "T",
        "step9_struct_code": "*",
        "errors": [],
    }
    errors: list[str] = []
    conflicts: list[str] = []

    _enforce_retired_glue_runtime_guard(steps, errors, conflicts)

    assert steps["step1_glue_code"] == ""
    assert errors
    assert any("NY-A1→2Z已废弃" in item for item in conflicts)
    assert _build_code_from_steps(steps, errors) == ""

    analysis = {"engine_steps": {"glue_model": "NY-A1", "step1_glue_code": "RC"}}
    try:
        _validate_confirmation_code("glue", "2Z", analysis)
    except ValueError as exc:
        assert "已废弃" in str(exc)
    else:
        raise AssertionError("NY-A1→2Z must be rejected")


def test_same_name_multi_code_glue_is_kept_but_marked_conflict() -> None:
    mappings = {
        "Agent胶系主表": [
            {
                "映射ID": "MASTER-1",
                "启用": "是",
                "胶系编号": "AANN",
                "胶系名称": "NYTEST",
                "输出胶系代码": "AA",
                "来源行号": "10",
            },
            {
                "映射ID": "MASTER-2",
                "启用": "是",
                "胶系编号": "BBNN",
                "胶系名称": "NYTEST",
                "输出胶系代码": "BB",
                "来源行号": "11",
            },
        ]
    }

    result = resolve_agent_glue(mappings, spec="NYTEST 0.8mm 1/1 41x49")

    assert result is not None
    assert result["status"] == "matched"
    assert result["uncertain"] is True
    assert {item["code"] for item in result["candidates"]} == {"AA", "BB"}
    assert "同名多码" in result["conflict"]

    decision = decide_confirmation(
        errors=[],
        conflicts=[result["conflict"]],
        candidate_code="AA008001137004900YWA1T*",
        field_evidence=[
            {
                "field_key": "glue",
                "field": "胶系",
                "score": 100,
                "gate": True,
                "evidence": "最新版正式胶系表",
            }
        ],
    )
    assert decision["status"] == "待确认"
    assert decision["formal_code"] == ""
    assert decision["overall_score"] == 60


def test_ny_a1_to_2z_is_retired_without_retiring_other_legacy_aliases() -> None:
    assert is_retired_agent_glue_mapping(
        {"映射ID": "TGM-SELECT-0002", "胶系名称": "NY-A1", "输出胶系代码": "2Z"}
    )
    assert not is_retired_agent_glue_mapping(
        {"兼容名称": "2ZZN", "输出胶系代码": "RC"}
    )

    mappings = {
        "Agent胶系主表": [
            {
                "映射ID": "TGM-MASTER-0058",
                "启用": "是",
                "胶系名称": "NY-A1",
                "输出胶系代码": "2Z",
                "来源行号": "58",
            },
            {
                "映射ID": "TGM-MASTER-NEW",
                "启用": "是",
                "胶系名称": "NY-A1",
                "输出胶系代码": "RC",
                "来源行号": "100",
            },
        ],
        "Agent胶系兼容别名": [
            {
                "映射ID": "TGM-ALIAS-LEGACY",
                "启用": "是",
                "兼容名称": "2ZZN",
                "标准胶系名称": "NY-A1",
                "输出胶系代码": "RC",
            }
        ],
        "Agent胶系选择规则": [
            {
                "映射ID": "TGM-SELECT-0002",
                "启用": "是",
                "胶系名称": "NY-A1",
                "输出胶系代码": "2Z",
                "条件客户简称": "测试客户",
                "优先级": "999",
            }
        ],
    }

    latest = resolve_agent_glue(
        mappings,
        spec="NY-A1 0.8mm 1/1 41x49",
        customer_name="测试客户",
    )
    legacy = resolve_agent_glue(mappings, spec="2ZZN 0.8mm 1/1 41x49")

    assert latest and latest["code"] == "RC"
    assert latest.get("uncertain") is not True
    assert legacy and legacy["code"] == "RC"


def test_retired_ny_a1_mapping_is_removed_from_runtime_and_business_lookup() -> None:
    runtime_tables = {
        "glue_exact_map": {"NY-A1": "2Z", "NY2150": "2B"},
        "glue_model_map": {"NY-A1": "2Z", "NY2170": "2C"},
    }
    _remove_retired_glue_mappings(runtime_tables)
    assert runtime_tables["glue_exact_map"] == {"NY2150": "2B"}
    assert runtime_tables["glue_model_map"] == {"NY2170": "2C"}

    business_tables = transcode_rule_center.build_rule_center_lookup_tables(
        {
            "glue_exact_map": {"NY-A1": "2Z", "NY2150": "2B"},
            "glue_model_map": {"NY-A1": "2Z"},
        },
        {"Agent胶系主表": []},
        official_grade_codes=set(),
        standard_sizes={},
        high_speed_mil={},
        copper_micron={},
        copper_types=[],
        copper_valid=set(),
        size_ranges=[],
    )
    assert business_tables["rule_center_glue_code"] == {"NY2150": "2B"}


def test_retired_rule_is_filtered_from_merged_agent_assets(monkeypatch) -> None:
    class _Result:
        def fetchall(self) -> list:
            return []

    class _Connection:
        def execute(self, *_args, **_kwargs) -> _Result:
            return _Result()

    @contextmanager
    def _db_cursor():
        yield _Connection()

    monkeypatch.setattr(transcode_rule_center, "ensure_rule_center_tables", lambda: None)
    monkeypatch.setattr(transcode_rule_center, "db_cursor", _db_cursor)
    merged = transcode_rule_center.merge_agent_mapping_overrides(
        {
            "Agent胶系主表": [],
            "Agent胶系兼容别名": [],
            "Agent胶系选择规则": [
                {
                    "映射ID": "TGM-SELECT-0002",
                    "启用": "是",
                    "胶系名称": "NY-A1",
                    "输出胶系代码": "2Z",
                },
                {
                    "映射ID": "TGM-SELECT-KEEP",
                    "启用": "是",
                    "胶系名称": "NY2150",
                    "输出胶系代码": "2B",
                },
            ],
        }
    )

    assert [row["映射ID"] for row in merged["Agent胶系选择规则"]] == [
        "TGM-SELECT-KEEP"
    ]


def test_generator_cannot_reintroduce_ny_a1_2z_and_keeps_2zzn_as_rc(
    tmp_path: Path,
) -> None:
    builder = _load_script_module(
        "build_agent_glue_assets_for_test",
        ROOT
        / "model_skills/customer-special-rule-maintenance/scripts/build_agent_glue_assets.py",
    )
    source = tmp_path / "latest.xlsx"
    legacy = tmp_path / "legacy.xlsx"
    output = tmp_path / "mapping.xlsx"
    _write_rows(
        source,
        "Sheet1",
        ["胶系编号", "胶系名称", "胶系分类", "旧胶系编号"],
        [
            ["2ZNA", "NY-A1", "NY-A1(A1)", "RC"],
            ["2ZZN", "NY-A1", "NY-A1(A1)", "2Z"],
            ["6CNN", "NY6300S", "NY6300(6300)", "6C"],
            ["R6C3", "NY6300S", "NY6300(6300)", "B1"],
        ],
    )
    _write_rows(
        legacy,
        "胶系代码",
        ["NYEOS胶系编号", "胶系名称", "特性描述", "旧胶系编号"],
        [
            ["2ZNN", "NY-A1", "NY-A1", "2Z"],
            ["2BNN", "NY2150", "NY2150", "2B"],
        ],
    )
    _write_rows(output, "占位", ["占位"], [])

    builder.build_assets(source, legacy, output)
    mappings = _read_mapping_workbook(output)
    generated_rows = [
        row
        for sheet in ("Agent胶系主表", "Agent胶系兼容别名", "Agent胶系选择规则")
        for row in mappings[sheet]
    ]
    assert not any(
        _is_enabled(row) and is_retired_agent_glue_mapping(row)
        for row in generated_rows
    )

    retired_history = [
        row
        for row in mappings["Agent胶系主表"]
        if is_retired_agent_glue_mapping(row)
    ]
    assert len(retired_history) == 1
    assert retired_history[0]["启用"] == "否"
    assert "禁用历史记录" in retired_history[0]["备注"]

    legacy_2zzn = [
        row
        for row in mappings["Agent胶系兼容别名"]
        if str(row.get("兼容名称") or "").upper() == "2ZZN"
    ]
    assert len(legacy_2zzn) == 1
    assert legacy_2zzn[0]["输出胶系代码"] == "RC"
    assert legacy_2zzn[0]["启用"] == "是"

    no_condition_selections = mappings["Agent胶系选择规则"]
    assert no_condition_selections
    assert all(row["启用"] == "否" for row in no_condition_selections)

    clear_agent_glue_index_cache()
    latest = resolve_agent_glue(mappings, spec="NY-A1 0.8mm 1/1 41x49")
    historical = resolve_agent_glue(mappings, spec="2ZZN 0.8mm 1/1 41x49")
    conflict = resolve_agent_glue(mappings, spec="NY6300S 0.8mm 1/1 41x49")
    assert latest and latest["code"] == "RC"
    assert historical and historical["code"] == "RC"
    assert conflict and conflict.get("uncertain") is True
    assert {item["code"] for item in conflict["candidates"]} == {"6C", "B1"}


def test_active_asset_has_no_enabled_ny_a1_2z_rows_in_active_and_default_assets() -> None:
    paths = (
        _active_agent_mapping_path(),
        ROOT
        / "fangzheng_web_app/default_rules/transcode_agent/transcode_agent_mapping_tables.xlsx",
    )
    for path in paths:
        mappings = _read_mapping_workbook(path)
        assert not any(
            _is_enabled(row) and is_retired_agent_glue_mapping(row)
            for rows in mappings.values()
            for row in rows
        ), path
        clear_agent_glue_index_cache()
        current = resolve_agent_glue(
            mappings,
            spec="NY-A1 0.8mm 1/1 41x49",
            customer_name="测试客户",
        )
        historical = resolve_agent_glue(mappings, spec="2ZZN 0.8mm 1/1 41x49")
        assert current and current["code"] == "RC", path
        assert historical and historical["code"] == "RC", path
        assert current["code"] != "2Z", path
        assert historical["code"] != "2Z", path


def test_active_base_glue_table_maps_ny_a1_to_rc() -> None:
    path = (
        ROOT
        / "storage/rules/transcode/versions/transcode_bootstrap_20260703_191711"
        / "transcode_rules.xlsx"
    )
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["胶系代码"]
    headers = [str(cell.value or "").strip() for cell in worksheet[1]]
    rows = [
        dict(zip(headers, values))
        for values in worksheet.iter_rows(min_row=2, values_only=True)
        if str(values[1] or "").strip().upper() == "NY-A1"
    ]
    workbook.close()

    assert rows
    assert {str(row["旧胶系编号"] or "").strip().upper() for row in rows} == {"RC"}


def test_all_master_multi_code_names_without_selection_condition_are_pending() -> None:
    audit_module = _load_script_module(
        "audit_rule_assets_full_for_test",
        ROOT / "tests_transcode_agent/audit_rule_assets_full.py",
    )
    mappings = _read_mapping_workbook(_active_agent_mapping_path())
    conflicts = audit_module._glue_conflicts(mappings["Agent胶系主表"])
    assert conflicts

    for conflict_group in conflicts:
        glue_name = conflict_group["rows"][0]["胶系名称"]
        clear_agent_glue_index_cache()
        result = resolve_agent_glue(
            mappings,
            spec=f"{glue_name} 0.8mm 1/1 41x49",
        )
        assert result and result.get("uncertain") is True, glue_name
        assert {item["code"] for item in result["candidates"]} == set(
            conflict_group["codes"]
        ), glue_name

        score, hit_type, _, note = _score_field(
            "glue",
            result["code"],
            {},
            [],
            {
                "glue_code": _rule(
                    source="最新版胶系名称优先命中",
                    rule_type="Agent胶系主数据映射",
                    rule_id=result.get("rule_id", ""),
                )
            },
            [result["conflict"]],
        )
        assert score < 100, glue_name
        assert hit_type == "规则冲突", glue_name
        assert "同名多码" in note, glue_name

        decision = decide_confirmation(
            errors=[],
            conflicts=[result["conflict"]],
            candidate_code=f"{result['code']}008001137004900YWA1T*",
            field_evidence=[
                {
                    "field_key": "glue",
                    "field": "胶系",
                    "score": score,
                    "gate": True,
                    "evidence": note,
                }
            ],
        )
        assert decision["status"] == "待确认", glue_name
        assert decision["formal_code"] == "", glue_name
        assert decision["overall_score"] < 100, glue_name
