from __future__ import annotations

from pathlib import Path
import sys

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app import transcode_agent_service as service
from fangzheng_web_app import transcode_agent_engine as engine
from fangzheng_web_app.transcode_semantic_overrides import apply_confirmed_semantic_overrides


def main() -> None:
    assert engine.extract_thickness_mm("0.184mm (不含铜 0.114mm 2116*1)")[0] == 0.114
    assert engine.extract_thickness_mm("1.515/15")[0] == 1.5
    assert engine.get_thickness_mode("芯板 0.26mm 含铜") == "total"
    assert engine.get_thickness_mode("0.57mm EXCLUDING COPPER") == "core"
    assert engine.extract_copper_spec("16mil 1/H 43*49") == "H/1"
    assert engine.extract_copper_spec("61um/61um") == "S/S"
    assert engine.extract_copper_spec("2.5/2.5OZ") == "E/E"
    assert engine.extract_copper_spec("(HTE) 012012 510*614") == "T/T"
    assert engine.extract_copper_spec("0.508mm不含铜/18um/NY6300SN/DK=3.53/C级公差") == "H/H"
    assert engine.extract_copper_spec("NY2140 1.00MM 35/00 43X49.3") == "1/0"
    assert engine.get_copper_type_code("IGAV UV") == "I"
    assert engine.get_grade_code("LED", "", "NY2150", {}, {}) == "AM"
    assert engine.get_grade_code("车载板", "", "NY3150HF", {}, {}) == "AC"
    assert ("940", "1245") in service._extract_all_raw_size_pairs("NY2150 35/35 940x1245 0,80mm")
    assert service._mapping_row_enabled({"启用": "Y"})
    correct_code = "2B008001137004900YWA1T"
    assert service._export_result_comparison(
        {"status": "成功", "formal_code": f"{correct_code}*"},
        f"{correct_code}Axxxxxxx",
    ) is True
    assert service._export_result_comparison(
        {"status": "待确认", "formal_code": f"{correct_code}*"},
        "AH008001137004900RWA1TAxxxxxxx",
    ) is False
    assert service._export_result_comparison({"status": "失败", "formal_code": ""}, correct_code) == "未出码"
    assert service._export_result_comparison(
        {"status": "成功", "formal_code": f"{correct_code}*"},
        "客户规格描述，不含正确制造码",
    ) == "无法对比"
    assert service._export_result_comparison({"status": "跳过", "formal_code": ""}, correct_code) == "跳过"

    analysis = {
        "engine_steps": {"step7_grade_code": "A1"},
        "applied_rules": [{"field": "grade_code", "source": "Agent客户映射"}],
    }
    evaluations = [
        {
            "status": "命中",
            "target_fields": ["grade_intent"],
            "normalized_values": ["AC"],
            "rule_id": "TSR-TEST",
        }
    ]
    applied, conflicts = apply_confirmed_semantic_overrides(
        engine,
        {"grade_code_map": {"AC": "汽车板"}},
        analysis,
        evaluations,
    )
    assert applied == [] and conflicts == []
    assert analysis["engine_steps"]["step7_grade_code"] == "A1"

    mapping_path = ROOT / "fangzheng_web_app/default_rules/transcode_agent/transcode_agent_mapping_tables.xlsx"
    workbook = openpyxl.load_workbook(mapping_path, read_only=True, data_only=True)
    assert _populated_row_count(workbook["客户字段映射"]) >= 194
    assert _populated_row_count(workbook["客户厚度映射"]) >= 9
    assert _populated_row_count(workbook["客户尺寸映射"]) >= 62
    field_mappings = _mapping_rows(workbook["客户字段映射"])
    mapping_ids = {str(row.get("映射ID") or "") for row in field_mappings}
    assert "TAM-FIELD-MANUAL-20260714-0008" in mapping_ids
    assert "TAM-FIELD-MODEL-20260714-0022" in mapping_ids

    semantic_steps = {
        "glue_model": "NY2150",
        "step3_copper_code": "H5",
        "step6_copper_type_code": "W",
        "errors": [],
    }
    semantic_applied = service._apply_agent_field_mappings(
        engine,
        {"客户字段映射": field_mappings},
        "103962",
        "深众齐翌",
        "NY2150 0.2mm 5um+18um/5um+18um 20.28*24.22",
        "",
        semantic_steps,
        semantic_steps["errors"],
        [],
    )
    assert semantic_steps["step3_copper_code"] == "XX"
    assert semantic_steps["step6_copper_type_code"] == "X"
    model_applied = [item for item in semantic_applied if item["rule_id"].startswith("TAM-FIELD-MODEL-")]
    assert len(model_applied) == 2
    assert all(item["source"] == "已批准模型语义映射" for item in model_applied)

    size_steps = {"glue_model": "NY2170", "step4_size_code": "82304900", "errors": []}
    service._apply_agent_field_mappings(
        engine,
        {"客户字段映射": field_mappings},
        "104397",
        "江中信华",
        "1.5（含铜厚）15/15 NY2170 TG170 2083*1245",
        "",
        size_steps,
        size_steps["errors"],
        [],
    )
    assert size_steps["step4_size_code"] == "82304930"
    print("1070 confirmed deterministic fixes smoke: PASS")


def _populated_row_count(worksheet) -> int:
    return sum(1 for row in worksheet.iter_rows(values_only=True) if row and row[0])


def _mapping_rows(worksheet) -> list[dict]:
    values = list(worksheet.iter_rows(values_only=True))
    headers = [str(value or "").strip() for value in values[0]]
    return [dict(zip(headers, row)) for row in values[1:] if row and row[0]]


if __name__ == "__main__":
    main()
