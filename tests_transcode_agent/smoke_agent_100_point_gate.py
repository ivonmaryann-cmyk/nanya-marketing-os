from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app.transcode_agent_service import (
    FIELD_GATE_THRESHOLD,
    _refresh_analysis_after_semantic_overrides,
    _save_agent_result,
    _load_runtime,
    analyze_spec,
)


def main() -> None:
    assert FIELD_GATE_THRESHOLD == 100
    engine, tables, rules, mappings, _, _ = _load_runtime()
    analysis = analyze_spec(
        engine,
        tables,
        rules,
        "NY2150 0.8mm 1/1 37*49 HTE 含铜",
        agent_mapping_tables=mappings,
        customer="测试客户",
        excel_row=2,
    )
    assert analysis["status"] == "成功", analysis
    assert analysis["overall_score"] == 100, analysis["field_evidence"]
    assert analysis["formal_code"] == analysis["candidate_code"]

    _refresh_analysis_after_semantic_overrides(analysis, [], ["基板级别存在多个语义结果"])
    assert analysis["status"] == "待确认", analysis
    assert analysis["formal_code"] == "", analysis
    assert analysis["candidate_code"], analysis

    with TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        source = temp_dir / "source.xlsx"
        output = temp_dir / "output.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "转码需求表"
        worksheet.append(["客户简称", "客户规格", "Agent转码结果"])
        worksheet.append(["测试客户", analysis["spec"], ""])
        workbook.save(source)

        request_frame = pd.DataFrame(
            [
                ["客户简称", "客户规格", "Agent转码结果"],
                ["测试客户", analysis["spec"], ""],
            ]
        )
        _save_agent_result(
            str(source),
            str(output),
            request_frame,
            2,
            [analysis],
            [],
            {},
            1,
        )
        result_book = openpyxl.load_workbook(output, data_only=True)
        result_sheet = result_book["转码需求表"]
        headers = [cell.value for cell in result_sheet[1]]
        formal_col = headers.index("Agent转码结果") + 1
        pending_col = headers.index("待人工确认码值") + 1
        comparison_col = headers.index("结果对比") + 1
        status_col = headers.index("转码状态") + 1
        assert result_sheet.cell(2, formal_col).value is None
        assert result_sheet.cell(2, pending_col).value == analysis["candidate_code"]
        assert result_sheet.cell(2, comparison_col).value == "待人工确认"
        assert result_sheet.cell(2, status_col).value == "待人工确认"

    print("100-point formal gate smoke passed formal/pending output separation verified")


if __name__ == "__main__":
    main()
