from __future__ import annotations

import os
import re
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app.transcode_agent_rules import (
    MAPPING_TABLE_HEADERS,
    load_transcode_agent_mapping_tables,
)
from fangzheng_web_app.transcode_agent_service import _load_runtime, analyze_spec


DEFAULT_BASELINE = Path(
    "/Users/ny/Desktop/南亚/产品包/转码采集规则源/CCL转码测试数据_业务确认基线_20260721.xlsx"
)
DEFAULT_MAPPING = (
    ROOT / "fangzheng_web_app/default_rules/transcode_agent/transcode_agent_mapping_tables.xlsx"
)
LEGACY_AGENT_VERSION = "transcode_agent_rules_20260721_111515"


def main() -> None:
    baseline = Path(os.environ.get("TRANSCODE_GLUE_REGRESSION_FILE") or DEFAULT_BASELINE)
    if not baseline.exists():
        print(f"agent glue regression skipped: missing {baseline}")
        return

    candidate_mappings = _load_mapping_workbook(DEFAULT_MAPPING)
    legacy_mappings = load_transcode_agent_mapping_tables(LEGACY_AGENT_VERSION)
    engine, tables, rules, _, _, _ = _load_runtime()
    worksheet = openpyxl.load_workbook(baseline, read_only=True, data_only=True).active
    regressions = []
    improvements = []
    changed = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
        cells = list(values) + [None] * 4
        customer_code, customer_name, correct_code, spec = cells[:4]
        if not str(spec or "").strip():
            continue
        outputs = {}
        candidate_master_backed = False
        for label, mappings in (("legacy", legacy_mappings), ("candidate", candidate_mappings)):
            analysis = analyze_spec(
                engine,
                tables,
                rules,
                str(spec),
                agent_mapping_tables=mappings,
                customer=str(customer_name or ""),
                customer_code=str(customer_code or ""),
                excel_row=row_number,
            )
            outputs[label] = _code22(analysis.get("candidate_code"))
            if label == "candidate":
                candidate_master_backed = any(
                    str(item.get("rule_type") or "") == "Agent胶系主数据映射"
                    and str(item.get("new") or "").upper() == outputs[label][:2]
                    for item in analysis.get("applied_rules") or []
                )
        correct = _code22(correct_code)
        if outputs["legacy"] != outputs["candidate"]:
            changed.append(row_number)
        if outputs["legacy"] == correct and outputs["candidate"] != correct:
            regressions.append((
                row_number,
                customer_name,
                correct,
                outputs,
                spec,
                candidate_master_backed,
            ))
        if outputs["legacy"] != correct and outputs["candidate"] == correct:
            improvements.append((row_number, customer_name, correct, outputs, spec))

    expected_latest_master_migrations = []
    unexpected_regressions = []
    for regression in regressions:
        _row_number, _customer_name, correct, outputs, _spec, master_backed = regression
        if master_backed and _only_glue_fields_changed(correct, outputs["candidate"]):
            expected_latest_master_migrations.append(regression)
        else:
            unexpected_regressions.append(regression)

    assert not unexpected_regressions, unexpected_regressions[:10]
    print(
        "agent glue regression passed "
        f"rows={worksheet.max_row - 1} changed={len(changed)} "
        f"unexpected_regressions={len(unexpected_regressions)} "
        f"latest_master_migrations={len(expected_latest_master_migrations)} "
        f"improvements={len(improvements)}"
    )


def _load_mapping_workbook(path: Path) -> dict[str, list[dict]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tables = {sheet: [] for sheet in MAPPING_TABLE_HEADERS}
    for sheet_name in tables:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {
                headers[index]: str(value or "").strip()
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
            if row.get("映射ID"):
                tables[sheet_name].append(row)
    return tables


def _code22(value) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()[:22]


def _only_glue_fields_changed(expected: str, candidate: str) -> bool:
    """Allow a confirmed latest-master migration only at glue/category positions."""
    if len(expected) != 22 or len(candidate) != 22 or expected == candidate:
        return False
    return expected[2:17] == candidate[2:17] and expected[18:22] == candidate[18:22]


if __name__ == "__main__":
    main()
