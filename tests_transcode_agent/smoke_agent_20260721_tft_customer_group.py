from pathlib import Path
import sys
from tempfile import TemporaryDirectory

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app.transcode_agent_service import (
    _approved_policy_difference_reason,
    _export_result_comparison,
    _load_runtime,
    _save_agent_result,
    analyze_spec,
)


ENGINE, TABLES, RULES, MAPPINGS, _, _ = _load_runtime()


def quote(customer_code: str, customer: str, spec: str, order_remark: str = "") -> dict:
    return analyze_spec(
        ENGINE,
        TABLES,
        RULES,
        spec,
        agent_mapping_tables=MAPPINGS,
        customer=customer,
        customer_code=customer_code,
        context_text=order_remark,
        excel_row=2,
    )


def assert_codes(result: dict, glue: str, category: str, grade: str | None = None) -> None:
    steps = result["engine_steps"]
    assert steps["step1_glue_code"] == glue, result
    assert steps["step5_glue_cat_code"] == category, result
    if grade is not None:
        assert steps["step7_grade_code"] == grade, result


def main() -> None:
    plain = quote("", "测试客户", "NY3150HF 0.5mm 1/1 41*49 HTE 不含铜 无卤")
    assert_codes(plain, "AH", "R", "A1")
    spaced_plain = quote("", "测试客户", "NY3150 HF 0.5mm 1/1 41*49 HTE 不含铜 无卤")
    assert_codes(spaced_plain, "AH", "R", "A1")
    hyphen_plain = quote("", "测试客户", "NY3150-HF 0.5mm 1/1 41*49 HTE 不含铜 无卤")
    assert_codes(hyphen_plain, "AH", "R", "A1")
    old_plain_code = "3B005001141004900YWATC"
    assert _export_result_comparison(plain, old_plain_code) is True
    assert "历史正确码待更新" in _approved_policy_difference_reason(plain, old_plain_code)
    assert _export_result_comparison(plain, "3B006001141004900YWATC") is False
    _assert_policy_confirmation_export(plain, old_plain_code)

    spec_tft = quote("", "测试客户", "NY3150HF 0.5mm 1/1 41*49 HTE TFT 不含铜 无卤")
    assert_codes(spec_tft, "3B", "Y", "AT")
    spaced_tft = quote("", "测试客户", "NY3150 HF 0.5mm 1/1 41*49 HTE TFT 不含铜 无卤")
    assert_codes(spaced_tft, "3B", "Y", "AT")
    assert _export_result_comparison(spec_tft, "AH005001141004900RWA1C") is True

    remark_tft = quote(
        "",
        "测试客户",
        "NY3150HF 0.5mm 1/1 41*49 HTE 不含铜 无卤",
        order_remark="客户要求 TFT 产品",
    )
    assert_codes(remark_tft, "3B", "Y", "AT")
    assert _export_result_comparison(remark_tft, "AH005001141004900RWA1C") is True

    old_customer_plain = quote(
        "104359",
        "淮安特创",
        "NY3150HF 0.5mm 1/1 41*49 HTE 不含铜 无卤",
    )
    assert_codes(old_customer_plain, "AH", "R", "A1")
    old_customer_tft = quote(
        "104359",
        "淮安特创",
        "NY3150HF 0.5mm 1/1 41*49 HTE TFT 不含铜 无卤",
    )
    assert_codes(old_customer_tft, "3B", "Y", "AT")

    zhuhai = quote("103890", "珠海景旺", "NY2150 0.8mm 1/1 41*49 HTE")
    assert_codes(zhuhai, "2T", "R")
    assert _export_result_comparison(zhuhai, "AS008001141004900YWA1T") is True
    assert _export_result_comparison(zhuhai, "AS009001141004900YWA1T") is False
    jiangxi = quote("123018", "江西景旺", "NY2150 0.8mm 1/1 41*49 HTE")
    assert_codes(jiangxi, "2B", "Y")

    zhuhai_h = quote("103890", "珠海景旺", "NY2150H 0.8mm 1/1 41*49 HTE")
    assert zhuhai_h["engine_steps"]["step1_glue_code"] != "2T", zhuhai_h
    jiangxi_h = quote("123018", "江西景旺", "NY2150H 0.8mm 1/1 41*49 HTE")
    assert jiangxi_h["engine_steps"]["step1_glue_code"] != "2B", jiangxi_h

    longer_model = quote("", "测试客户", "NY3150HFP 0.5mm 1/1 41*49 HTE 不含铜 无卤")
    assert longer_model["engine_steps"]["step1_glue_code"] != "AH", longer_model

    shennan_spec = "覆铜板 FR4.0 0.089 HV2H/HV2H 0.124 36X48 2A NY-P1"
    for customer_code, customer in (
        ("104370", "南通深南"),
        ("103673", "深南电路"),
        ("104299", "无锡深南"),
        ("", "深南集团"),
    ):
        result = quote(customer_code, customer, shennan_spec)
        assert result["engine_steps"]["step4_size_code"] == "37004900", result
        assert result["engine_steps"]["customer_rule_group"] == "深南集团", result

    print("20260721 TFT and customer group smoke: PASS")


def _assert_policy_confirmation_export(analysis: dict, old_code: str) -> None:
    with TemporaryDirectory() as temp_dir:
        source_path = Path(temp_dir) / "source.xlsx"
        output_path = Path(temp_dir) / "result.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["客户简称", "品名", "客户规格", "Agent转码结果"])
        worksheet.append([
            analysis["customer"],
            old_code,
            analysis["spec"],
            analysis["formal_code"],
        ])
        workbook.save(source_path)
        frame = pd.DataFrame([
            ["客户简称", "品名", "客户规格", "Agent转码结果"],
            [analysis["customer"], old_code, analysis["spec"], analysis["formal_code"]],
        ])
        _save_agent_result(
            str(source_path),
            str(output_path),
            frame,
            3,
            [analysis],
            RULES,
            MAPPINGS,
            0,
        )
        result_book = openpyxl.load_workbook(output_path, data_only=True)
        result_sheet = result_book["Sheet1"]
        headers = {str(cell.value or ""): cell.column for cell in result_sheet[1]}
        assert result_sheet.cell(2, headers["状态"]).value == "已出码需核对"
        prompt = str(result_sheet.cell(2, headers["说明"]).value or "")
        assert prompt == "系统100分出码，待人工核对", prompt
        fill = result_sheet.cell(2, headers["Agent转码结果"]).fill.fgColor.rgb or ""
        assert fill.endswith("FFCDD2"), fill


if __name__ == "__main__":
    main()
