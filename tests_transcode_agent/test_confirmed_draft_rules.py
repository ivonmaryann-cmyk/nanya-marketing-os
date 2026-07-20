from pathlib import Path

from fangzheng_web_app.transcode_agent_rules import (
    build_machine_rule_workbook,
    load_transcode_agent_rules,
    parse_customer_special_master,
)


ROOT = Path(__file__).resolve().parents[1]
DRAFT_PATH = ROOT / "docs/develop0707/客户特殊规则结构化草稿_按原表行_20260708.xlsx"


def test_confirmed_draft_generates_agent_rules(tmp_path):
    rules, summary = parse_customer_special_master(DRAFT_PATH)

    assert summary["source_type"] == "confirmed_structured_draft"
    assert summary["rule_count"] == 200
    assert summary["status_counts"]
    assert rules
    assert {rule["覆盖字段"] for rule in rules} <= {
        "glue_code",
        "thickness_code",
        "copper_code",
        "size_code",
        "glue_category_code",
        "copper_type_code",
        "grade_code",
        "tc_code",
    }
    assert all(rule["强制执行"] == "是" for rule in rules)
    assert all(rule["待确认"] == "否" for rule in rules)
    assert all(rule["物料类别"] == "CCL" for rule in rules)
    assert not any(rule["覆盖字段"] == "struct_code" for rule in rules)
    assert not any("订单" in (rule.get("条件文本", "") + rule.get("覆盖值", "")) for rule in rules)

    output_path = tmp_path / "transcode_agent_rules.xlsx"
    build_machine_rule_workbook(output_path, rules, summary)
    loaded = load_transcode_agent_rules_from_path(output_path)
    assert len(loaded) == len(rules)


def test_key_customer_rules_are_table_driven():
    rules, _summary = parse_customer_special_master(DRAFT_PATH)

    assert _find_rule(rules, "珠海景旺", "glue_code", "RV", "NY3150HC")
    assert _find_rule(rules, "方正F7", "tc_code", "C", "")
    assert _find_rule(rules, "上海山崎", "glue_code", "2A", "NY1140")
    assert _find_rule(rules, "江苏广谦", "tc_code", "T", ">=0.8")
    assert _find_rule(rules, "江苏广谦", "tc_code", "C", "<0.8")
    assert _find_rule(rules, "广德三生", "tc_code", "T", ">=0.8")
    assert _find_rule(rules, "广德三生", "tc_code", "C", "<0.8")
    assert _find_rule(rules, "川英创力", "glue_code", "2A", "TG140")
    assert _find_rule(rules, "川英创力", "glue_code", "2B", "TG150")
    assert _find_rule(rules, "川英创力", "glue_code", "2C", "TG170")
    assert _find_rule(rules, "川英创力", "copper_code", "FF", "W/W")
    assert _find_rule(rules, "赣州景旺", "glue_code", "RV", "NY3150HC")
    assert _find_rule(rules, "川华兴宇", "grade_code", "A1", ">=0.9")
    assert _find_rule(rules, "川华兴宇", "grade_code", "AC", "<0.9")
    assert _find_rule(rules, "黄石广合", "grade_code", "AT", "43MIL")
    assert _find_rule(rules, "惠州泰和", "grade_code", "AT", "<=1.2")
    assert _find_rule(rules, "广华升鑫、深华升鑫", "glue_code", "AL", "NY-A2")
    assert _find_rule(rules, "广华升鑫、深华升鑫", "grade_code", "AC", "NY-A2")
    assert _find_rule(rules, "万安裕维", "glue_code", "2A", "__NO_TG__")
    assert _find_rule(rules, "万安裕维", "glue_code", "2A", "TG130")
    assert _find_rule(rules, "万安裕维", "glue_code", "3B", "TG150和无卤素")
    assert _find_rule(rules, "万安裕维", "glue_code", "3H", "TG150&无卤素&CTI600")
    assert _find_rule(rules, "珠海乐健", "grade_code", "AC", "")
    assert _find_rule(rules, "珠海乐健", "grade_code", "A1", "1.1")
    assert _find_rule(rules, "赣州乐健", "grade_code", "AC", "")
    assert _find_rule(rules, "赣州乐健", "grade_code", "A1", "H/H,1/1,2/2")


def load_transcode_agent_rules_from_path(path: Path) -> list[dict]:
    original = load_transcode_agent_rules
    # Keep the test independent from the active local rule version.
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook["机器规则"]
    headers = [str(worksheet.cell(1, col).value or "").strip() for col in range(1, worksheet.max_column + 1)]
    rows = []
    for row_idx in range(2, worksheet.max_row + 1):
        item = {
            headers[col - 1]: str(worksheet.cell(row_idx, col).value or "").strip()
            for col in range(1, worksheet.max_column + 1)
            if headers[col - 1]
        }
        if item.get("规则ID"):
            rows.append(item)
    assert original is load_transcode_agent_rules
    return rows


def _find_rule(rules: list[dict], customer_name: str, field: str, value: str, condition_part: str) -> bool:
    for rule in rules:
        if rule.get("客户简称") != customer_name:
            continue
        if rule.get("覆盖字段") != field or rule.get("覆盖值") != value:
            continue
        condition_text = " ".join(
            [
                rule.get("条件文本", ""),
                rule.get("条件胶系", ""),
                rule.get("条件厚度", ""),
                rule.get("条件铜厚", ""),
                rule.get("条件关键词", ""),
            ]
        )
        if not condition_part or condition_part in condition_text:
            return True
    return False
