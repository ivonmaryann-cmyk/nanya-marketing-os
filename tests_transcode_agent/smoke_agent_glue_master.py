from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app.transcode_agent_rules import MAPPING_TABLE_HEADERS
from fangzheng_web_app.transcode_agent_service import _load_runtime, analyze_spec


ASSET = ROOT / "fangzheng_web_app/default_rules/transcode_agent/transcode_agent_mapping_tables.xlsx"


def main() -> None:
    mappings = _load_mapping_workbook(ASSET)
    assert len(mappings["Agent胶系主表"]) == 282
    assert len(mappings["Agent胶系兼容别名"]) == 182
    assert len(mappings["Agent胶系选择规则"]) == 7

    engine, tables, rules, _, _, _ = _load_runtime()
    cases = [
        ("江西旭昇", "FR4-NY-2140L-板厚1.1mm-铜厚15/15um-尺寸82*49.3英寸", "2K"),
        ("沪士二厂", 'NY6300SL 0050" 1/1 688*535 1078*2 RTF3 SNY', "6S"),
        ("珠海杰赛", '0.508mm不含铜/18um/NY6300SN/DK=3.53/C级公差 41"x49"', "RB"),
        ("昆山竞陆", "铜基板 FR-4 无卤 0.71mm H/H 41*49 TG>=260 SHNY 3170LK", "3D"),
        ("测试客户", "3CNA NY3170HF 汽车板 0.8mm 1/1 41*49", "3C"),
        ("测试客户", "R3E3 NY3170HF 汽车板 0.8mm 1/1 41*49", "BD"),
        ("测试客户", "1AON NY1140 O料 UL专用 0.8mm 1/1 41*49", "1A"),
        ("南通深南", "NY6300 0.406 H/H 43*49 HTE", "6W"),
        ("南通深南", "NY6300(C) 0.406 H/H 43*49 HTE", "R1"),
        ("江西景旺", 'CCL NY-A1 1.5mm H/H 含铜 82inX49.3in HTE', "RC"),
        ("历史测试客户", 'CCL 2ZZN NY-A1 1.5mm H/H 含铜 82inX49.3in HTE', "RC"),
        ("惠州智恩", '南亚新材NY-A2 TG170 0.25mm 3/3oz 不含铜 82"*49"', "AL"),
    ]
    for customer, spec, expected in cases:
        analysis = analyze_spec(
            engine,
            tables,
            rules,
            spec,
            agent_mapping_tables=mappings,
            customer=customer,
        )
        actual = str((analysis.get("engine_steps") or {}).get("step1_glue_code") or "")
        assert actual == expected, (customer, spec, expected, actual, analysis.get("reason"))
        if "NY-A1" in spec:
            steps = analysis.get("engine_steps") or {}
            assert not steps.get("agent_glue_uncertain")
            assert not steps.get("agent_glue_candidates")
    print(
        "agent glue master smoke passed "
        f"master=282 aliases=182 selections=7 cases={len(cases)}"
    )


def _load_mapping_workbook(path: Path) -> dict[str, list[dict]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tables = {sheet: [] for sheet in MAPPING_TABLE_HEADERS}
    for sheet_name in tables:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = {
                headers[index]: str(value or "").strip()
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
            if row.get("映射ID"):
                tables[sheet_name].append(row)
    return tables


if __name__ == "__main__":
    main()
