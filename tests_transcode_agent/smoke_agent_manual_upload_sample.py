from pathlib import Path
import shutil
import sys
from tempfile import TemporaryDirectory

import openpyxl
from werkzeug.datastructures import FileStorage


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app import db
from fangzheng_web_app import transcode_agent_rules as agent_rules
from fangzheng_web_app import transcode_rules
from fangzheng_web_app.db import create_job, get_job
from fangzheng_web_app.transcode_agent_rules import FEATURE_KEY
from fangzheng_web_app.transcode_agent_service import run_transcode_agent_job


DRAFT_PATH = ROOT / "docs/develop0707/客户特殊规则结构化草稿_按原表行_20260708.xlsx"
UPLOAD_SAMPLE_PATH = ROOT / "tests_transcode_agent/fixtures/transcode_agent_10_manual_upload.xlsx"


def main() -> None:
    assert UPLOAD_SAMPLE_PATH.exists(), UPLOAD_SAMPLE_PATH
    with TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        original_db_path = db.DATABASE_PATH
        original_agent_rules_dir = agent_rules.TRANSCODE_AGENT_RULES_DIR
        original_agent_versions_dir = agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR
        original_base_versions_dir = transcode_rules.TRANSCODE_RULES_VERSIONS_DIR
        try:
            db.DATABASE_PATH = temp_dir / "storage/app.db"
            db.DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
            db.init_db()

            transcode_rules.TRANSCODE_RULES_VERSIONS_DIR = temp_dir / "rules/transcode/versions"
            transcode_rules.TRANSCODE_RULES_VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
            base_version = transcode_rules.ensure_default_transcode_rule_version()

            agent_rules.TRANSCODE_AGENT_RULES_DIR = temp_dir / "rules/transcode_agent"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = agent_rules.TRANSCODE_AGENT_RULES_DIR / "versions"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
            with DRAFT_PATH.open("rb") as source:
                upload = FileStorage(
                    stream=source,
                    filename=DRAFT_PATH.name,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                agent_version = agent_rules.save_new_transcode_agent_rule_version(
                    upload,
                    updated_by="smoke",
                    remark="manual upload sample smoke",
                )

            employee_id = "smoke"
            job_dir = temp_dir / "jobs" / employee_id
            job_dir.mkdir(parents=True, exist_ok=True)
            input_path = job_dir / UPLOAD_SAMPLE_PATH.name
            shutil.copy2(UPLOAD_SAMPLE_PATH, input_path)

            job_id = create_job(
                employee_id,
                UPLOAD_SAMPLE_PATH.name,
                str(input_path),
                f"base:{base_version};agent:{agent_version}",
                feature=FEATURE_KEY,
            )
            run_transcode_agent_job(job_id, employee_id)
            job = get_job(job_id)
            assert job["status"] == "completed", dict(job)
            assert job["total_rows"] == 10, dict(job)
            assert job["skip_count"] == 0, dict(job)
            assert Path(job["stored_result_path"]).exists(), dict(job)

            workbook = openpyxl.load_workbook(job["stored_result_path"], data_only=True)
            main_ws = workbook[workbook.sheetnames[0]]
            main_headers = [cell.value for cell in main_ws[1]]
            for header in ["22位码值差异", "结果对比", "转码状态", "人工确认提示", "系统分析原因"]:
                assert header in main_headers, main_headers
            assert "待人工确认码值" not in main_headers, main_headers
            assert main_headers[-1] == "系统分析原因", main_headers
            assert "需要人工确认" not in main_headers, main_headers
            output_status_index = main_headers.index("结果对比") + 1
            difference_index = main_headers.index("22位码值差异") + 1
            transcode_status_index = main_headers.index("转码状态") + 1
            system_analysis_index = main_headers.index("系统分析原因") + 1
            assert all(
                main_ws.cell(row=row, column=output_status_index).value == "无法对比"
                for row in range(2, main_ws.max_row + 1)
            )
            assert all(
                main_ws.cell(row=row, column=difference_index).value == "无法对比"
                for row in range(2, main_ws.max_row + 1)
            )
            assert all(
                main_ws.cell(row=row, column=transcode_status_index).value in {"可直接采用", "待人工确认", "未出码", "跳过"}
                for row in range(2, main_ws.max_row + 1)
            )
            assert all(
                main_ws.cell(row=row, column=system_analysis_index).value is None
                for row in range(2, main_ws.max_row + 1)
            )
            assert "字段证据链" in workbook.sheetnames
            assert "规则命中汇总" in workbook.sheetnames
            if "技术待支持清单" in workbook.sheetnames:
                assert workbook["技术待支持清单"].max_row > 1
            if "模型语义影子证据" in workbook.sheetnames:
                assert workbook["模型语义影子证据"].max_row > 1
            assert "证据评分影子对比" in workbook.sheetnames
            score_headers = [cell.value for cell in workbook["证据评分影子对比"][1]]
            for header in ["正式证据门禁模式", "程序证据分", "正式有效分", "正式门禁拦截", "正式门禁原因"]:
                assert header in score_headers, score_headers
            evidence_rows = list(workbook["字段证据链"].iter_rows(min_row=2, values_only=True))
            _assert_evidence(evidence_rows, "珠海景旺", "胶系", "RV", "Agent规则覆盖")
            _assert_evidence(evidence_rows, "方正F7", "铜箔类型", "P", "Agent规则覆盖")
            _assert_evidence(evidence_rows, "南通深南", "尺寸", "37004900", "Agent尺寸映射")
            _assert_evidence(evidence_rows, "惠州特创", "尺寸", "37304930", "Agent尺寸映射")
            _assert_evidence(evidence_rows, "安徽万奔", "尺寸", "41304930", "Agent尺寸映射")
            _assert_evidence(evidence_rows, "江福昌发", "尺寸", "74304130", "Agent尺寸映射")
            _assert_evidence(evidence_rows, "江苏瀚宇", "厚度", "00079", "Agent厚度映射")
            _assert_evidence(evidence_rows, "江苏瀚宇", "总/芯厚", "C", "Agent总芯厚映射")
            _assert_evidence(evidence_rows, "江苏瀚宇", "总/芯厚", "T", "Agent总芯厚映射")
            _assert_evidence(evidence_rows, "无新美亚", "总/芯厚", "C", "Agent物料编码口径")
            _assert_evidence(evidence_rows, "无新美亚", "总/芯厚", "T", "Agent物料编码口径")
            summary_rows = list(workbook["规则命中汇总"].iter_rows(values_only=True))
            _assert_summary_count(summary_rows, "Agent尺寸映射", 4)
            _assert_summary_count(summary_rows, "Agent厚度映射", 1)
            _assert_summary_count(summary_rows, "Agent总芯厚映射", 2)
            _assert_summary_count(summary_rows, "Agent物料编码口径", 2)
            pending_rows = (
                list(workbook["技术待支持清单"].iter_rows(min_row=2, values_only=True))
                if "技术待支持清单" in workbook.sheetnames
                else []
            )
            _assert_no_pending_item(pending_rows, "外部尺寸表引用", "无新美亚")

            print(
                "manual upload sample smoke passed "
                f"rows={job['total_rows']} success={job['success_count']} "
                f"fail={job['fail_count']} skip={job['skip_count']}"
            )
        finally:
            db.DATABASE_PATH = original_db_path
            agent_rules.TRANSCODE_AGENT_RULES_DIR = original_agent_rules_dir
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = original_agent_versions_dir
            transcode_rules.TRANSCODE_RULES_VERSIONS_DIR = original_base_versions_dir


def _assert_evidence(rows, customer: str, field: str, code: str, hit_type: str) -> None:
    for row in rows:
        if row[2] == customer and row[5] == field and row[7] == code and row[9] == hit_type:
            return
    raise AssertionError((customer, field, code, hit_type))


def _assert_no_evidence(rows, customer: str, field: str, hit_type: str) -> None:
    for row in rows:
        if row[2] == customer and row[5] == field and row[9] == hit_type:
            raise AssertionError((customer, field, hit_type, row))


def _assert_summary_count(rows, label: str, minimum_count: int) -> None:
    for row in rows:
        if row and row[0] == label and int(row[1] or 0) >= minimum_count:
            return
    raise AssertionError((label, minimum_count))


def _assert_pending_item(rows, technical_type: str, customer: str, keyword: str) -> None:
    for row in rows:
        if row and row[1] == technical_type and row[3] == customer and keyword in str(row[6] or ""):
            return
    raise AssertionError((technical_type, customer, keyword))


def _assert_no_pending_item(rows, technical_type: str, customer: str) -> None:
    for row in rows:
        if row and row[1] == technical_type and row[3] == customer:
            raise AssertionError((technical_type, customer, row))


if __name__ == "__main__":
    main()
