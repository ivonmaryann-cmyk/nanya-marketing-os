from __future__ import annotations

import zipfile

import openpyxl
import pytest

from fangzheng_web_app import transcode_agent_service as service
from fangzheng_web_app.transcode_agent_service import (
    _atomic_save_workbook,
    _load_transcode_agent_trace_records,
)


def _write_workbook(path, value: str) -> None:
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = value
    workbook.save(path)


def test_atomic_save_keeps_existing_result_readable_until_replace(tmp_path):
    target = tmp_path / "result.xlsx"
    _write_workbook(target, "old")

    class DelayedWorkbook:
        def save(self, temp_path):
            assert zipfile.is_zipfile(target)
            assert openpyxl.load_workbook(target, read_only=True).active["A1"].value == "old"
            _write_workbook(temp_path, "new")
            assert openpyxl.load_workbook(target, read_only=True).active["A1"].value == "old"

    _atomic_save_workbook(DelayedWorkbook(), target)

    assert zipfile.is_zipfile(target)
    assert openpyxl.load_workbook(target, read_only=True).active["A1"].value == "new"
    assert not list(tmp_path.glob("*.tmp.xlsx"))


def test_atomic_save_failure_preserves_existing_result(tmp_path):
    target = tmp_path / "result.xlsx"
    _write_workbook(target, "old")

    class FailingWorkbook:
        def save(self, temp_path):
            temp_path.write_bytes(b"partial")
            raise RuntimeError("save failed")

    with pytest.raises(RuntimeError, match="save failed"):
        _atomic_save_workbook(FailingWorkbook(), target)

    assert zipfile.is_zipfile(target)
    assert openpyxl.load_workbook(target, read_only=True).active["A1"].value == "old"
    assert not list(tmp_path.glob("*.tmp.xlsx"))


def test_corrupt_legacy_result_does_not_crash_confirmation_page(tmp_path):
    target = tmp_path / "corrupt.xlsx"
    target.write_bytes(b"partial xlsx")

    assert _load_transcode_agent_trace_records(str(target)) == []


def test_batch_reevaluation_updates_all_rows_with_one_save(tmp_path, monkeypatch):
    target = tmp_path / "result.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "转码需求表"
    sheet.append(
        [
            "品名",
            service.FORMAL_RESULT_HEADER,
            service.PENDING_RESULT_HEADER,
            service.CODE_DIFFERENCE_HEADER,
            service.OUTPUT_STATUS_HEADER,
            service.TRANSCODE_STATUS_HEADER,
            service.CONFIRMATION_HEADER,
            service.SYSTEM_ANALYSIS_HEADER,
        ]
    )
    sheet.append(["", "", "old-2", "", "", "待人工确认", "待确认", ""])
    sheet.append(["", "", "old-3", "", "", "待人工确认", "待确认", ""])
    workbook.save(target)

    save_calls = 0
    original_save = service._atomic_save_workbook

    def counted_save(updated_workbook, output_path):
        nonlocal save_calls
        save_calls += 1
        original_save(updated_workbook, output_path)

    monkeypatch.setattr(service, "_atomic_save_workbook", counted_save)
    service._update_automatic_reevaluation_workbook_rows(
        str(target),
        [
            (2, {"formal_code": "CODE-2", "candidate_code": "CODE-2"}),
            (3, {"formal_code": "CODE-3", "candidate_code": "CODE-3"}),
        ],
    )

    assert save_calls == 1
    updated = openpyxl.load_workbook(target)
    result = updated["转码需求表"]
    assert result.cell(2, 2).value == "CODE-2"
    assert result.cell(3, 2).value == "CODE-3"
    assert result.cell(2, 3).value is None
    assert result.cell(3, 3).value is None
    assert result.cell(2, 6).value == "可直接采用"
    assert result.cell(3, 6).value == "可直接采用"
