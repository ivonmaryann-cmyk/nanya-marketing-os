from __future__ import annotations

from pathlib import Path

import openpyxl

from fangzheng_web_app import db
from fangzheng_web_app import create_app
from fangzheng_web_app import routes
from fangzheng_web_app import transcode_agent_service as service
from fangzheng_web_app.transcode_agent_rules import FEATURE_KEY


def test_single_input_creates_one_row_batch_job(monkeypatch, tmp_path):
    monkeypatch.setattr(db, "DATABASE_PATH", tmp_path / "app.db")
    monkeypatch.setattr(service, "JOBS_DIR", tmp_path / "jobs")
    db.init_db()

    launched: list[tuple[int, str, str]] = []
    monkeypatch.setattr(
        service,
        "launch_job_process",
        lambda job_id, feature, employee_id: launched.append(
            (job_id, feature, employee_id)
        ),
    )

    job_id = service.queue_transcode_agent_single_job(
        "tester",
        customer_code="103901",
        customer="广东依顿",
        spec='NY2150 0.8mm 1/1 37*49" HTE 含铜',
        order_remark="下汽车板",
    )

    job = db.get_job(job_id)
    assert job is not None
    assert job["feature"] == FEATURE_KEY
    assert launched == [(job_id, FEATURE_KEY, "tester")]

    input_path = Path(job["stored_input_path"])
    workbook = openpyxl.load_workbook(input_path, data_only=True)
    worksheet = workbook["转码需求表"]
    assert [cell.value for cell in worksheet[1]] == [
        "客户代码",
        "客户简称",
        "客户规格",
        "订单备注",
    ]
    assert [cell.value for cell in worksheet[2]] == [
        "103901",
        "广东依顿",
        'NY2150 0.8mm 1/1 37*49" HTE 含铜',
        "下汽车板",
    ]
    workbook.close()

    service.run_transcode_agent_job(job_id, "tester")
    completed = db.get_job(job_id)
    assert completed is not None
    assert completed["total_rows"] == 1
    assert completed["status"] in {"completed", "awaiting_confirmation"}
    result_path = Path(completed["stored_result_path"])
    assert result_path.exists()

    result_book = openpyxl.load_workbook(result_path, data_only=True)
    assert "字段证据链" in result_book.sheetnames
    assert "证据评分影子对比" in result_book.sheetnames
    result_book.close()


def test_single_job_api_returns_shared_task_url(monkeypatch, tmp_path):
    monkeypatch.setattr(db, "DATABASE_PATH", tmp_path / "app.db")
    db.init_db()
    captured: dict[str, object] = {}

    def fake_queue(employee_id: str, **kwargs) -> int:
        captured.update(employee_id=employee_id, **kwargs)
        return 321

    monkeypatch.setattr(routes, "queue_transcode_agent_single_job", fake_queue)
    app = create_app()
    app.config.update(TESTING=True)

    with app.test_client() as client:
        with client.session_transaction() as session:
            session["employee_id"] = "tester"
        response = client.post(
            "/api/transcode-agent/single-jobs",
            json={
                "customer_code": "103901",
                "customer": "广东依顿",
                "spec": "NY2150 0.8mm 1/1 37*49 HTE",
                "order_remark": "下汽车板",
            },
        )

    assert response.status_code == 202
    payload = response.get_json()
    assert payload["job_id"] == 321
    assert payload["task_url"].endswith(
        "/features/transcode-agent?job_id=321&auto_confirm=1"
    )
    assert captured == {
        "employee_id": "tester",
        "customer_code": "103901",
        "customer": "广东依顿",
        "spec": "NY2150 0.8mm 1/1 37*49 HTE",
        "order_remark": "下汽车板",
    }
