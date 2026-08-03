from pathlib import Path
import sys
from tempfile import TemporaryDirectory

import openpyxl
from werkzeug.datastructures import FileStorage


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app import db
from fangzheng_web_app import transcode_agent_rules as agent_rules


DRAFT_PATH = ROOT / "docs/develop0707/客户特殊规则结构化草稿_按原表行_20260708.xlsx"


def main() -> None:
    with TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        original_rules_dir = agent_rules.TRANSCODE_AGENT_RULES_DIR
        original_versions_dir = agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR
        original_db_path = db.DATABASE_PATH
        try:
            agent_rules.TRANSCODE_AGENT_RULES_DIR = temp_dir / "rules/transcode_agent"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = agent_rules.TRANSCODE_AGENT_RULES_DIR / "versions"
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR.mkdir(parents=True, exist_ok=True)
            db.DATABASE_PATH = temp_dir / "storage/app.db"
            db.DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
            db.init_db()

            with DRAFT_PATH.open("rb") as source:
                upload = FileStorage(
                    stream=source,
                    filename=DRAFT_PATH.name,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                version = agent_rules.save_new_transcode_agent_rule_version(
                    upload,
                    updated_by="smoke",
                    remark="mapping table smoke",
                )

            machine_rules = agent_rules.load_transcode_agent_rules(version)
            assert len(machine_rules) == 196, len(machine_rules)

            mapping_path = agent_rules.get_transcode_agent_mapping_table_file_path(version)
            assert mapping_path.exists(), mapping_path
            workbook = openpyxl.load_workbook(mapping_path, data_only=True)
            expected_sheets = [
                "Agent胶系主表",
                "Agent胶系兼容别名",
                "Agent胶系选择规则",
                "Agent基础条件规则",
                "客户规则组",
                "客户字段映射",
                "客户单边尺寸映射",
                "客户尺寸映射",
                "客户尺寸算法",
                "客户厚度映射",
                "客户物料编码口径",
                "外部尺寸表引用",
                "待接入规则",
                "转换说明",
            ]
            assert workbook.sheetnames == expected_sheets, workbook.sheetnames
            assert workbook["Agent胶系主表"].max_row == 1
            assert workbook["Agent胶系兼容别名"].max_row == 1
            assert workbook["Agent胶系选择规则"].max_row == 1
            assert workbook["Agent基础条件规则"].max_row >= 1
            assert workbook["客户规则组"].max_row >= 1
            assert workbook["客户字段映射"].max_row >= 1
            assert workbook["客户单边尺寸映射"].max_row == 25
            assert workbook["客户尺寸映射"].max_row == 49
            assert workbook["客户尺寸算法"].max_row == 2
            assert workbook["客户厚度映射"].max_row == 3
            assert workbook["客户物料编码口径"].max_row == 3
            assert workbook["外部尺寸表引用"].max_row == 2
            assert workbook["待接入规则"].max_row == 1

            assert _find_row(workbook["客户单边尺寸映射"], "江福昌发", "1888", "74.3")
            assert _find_row(workbook["客户尺寸映射"], "南通深南", "36", "48", "37", "49")
            assert _find_row(workbook["客户尺寸算法"], "安徽万奔", "尺寸加大", "0.3", "0.3")
            assert _find_row(workbook["客户厚度映射"], "江苏瀚宇", "003", "0.079", "3")
            assert _find_row(workbook["客户物料编码口径"], "无新美亚", "631", "C")
            assert _find_row(workbook["外部尺寸表引用"], "无新美亚", "是", "新美亚规格尺寸对照表.xlsx")

            unresolved_tables = {sheet: [] for sheet in agent_rules.MAPPING_TABLE_HEADERS}
            agent_rules._append_pending_mapping_row(
                unresolved_tables,
                {"启用": "否", "客户代码": "TEST", "客户简称": "测试客户", "来源行号": "999"},
                "需按客户未定义等级机制处理",
                "尚无可执行映射",
            )
            assert len(unresolved_tables["待接入规则"]) == 1

            print(
                "mapping table smoke passed "
                f"version={version} machine_rules={len(machine_rules)} "
                "side=24 size=48 size_algo=1 thickness=2 material=2 external=1 pending=0"
            )
        finally:
            agent_rules.TRANSCODE_AGENT_RULES_DIR = original_rules_dir
            agent_rules.TRANSCODE_AGENT_RULES_VERSIONS_DIR = original_versions_dir
            db.DATABASE_PATH = original_db_path


def _find_row(worksheet, *needles: str) -> bool:
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        text = " ".join(str(value or "") for value in row)
        if all(needle in text for needle in needles):
            return True
    return False


if __name__ == "__main__":
    main()
