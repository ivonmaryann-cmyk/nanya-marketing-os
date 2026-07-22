from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from fangzheng_web_app.inventory_detail_service import (
    A_HEADERS,
    B_HEADERS,
    PLAN_A_HEADERS,
    PLAN_A_MODE,
    build_inventory_workbook,
    load_inventory_rows,
    load_plan_a_rows,
)


HEADERS = [
    "品号",
    "品名",
    "数量",
    "规格",
    "类别",
    "厚度",
    "铜箔",
    "尺寸",
    "颜色",
    "水印",
    "等级",
    "库位信息",
    "芯/总厚",
    "单重",
    "折合大板数量",
]

PLAN_HEADERS = [
    "品名",
    "品号",
    "规格",
    "排版结构",
    "上海_江西",
    "胶系",
    "胶系名称",
    "厚度",
    "铜箔",
    "单重",
    "库龄",
    "数量",
    "折合大板",
]


def write_input(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def write_plan_input(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(PLAN_HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


class InventoryDetailTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_filter_grade_merge_and_quantity_source(self) -> None:
        shanghai = self.root / "shanghai.xlsx"
        jiangxi = self.root / "jiangxi.xlsx"
        write_input(
            shanghai,
            [
                ["P1", "品一", 2, "  NY2150   0.10 1/1 41*49 ", "NY2150", 0.1, "1/1", "41*49", "", "", "A1", "无订单A级", "总厚", 1.2, 999],
                ["P2", "品二", 3, "NY2150 0.10 1/1 41*49", "NY2150", 0.1, "1/1", "41*49", "", "", "A2", "无订单A级", "总厚", 1.3, 999],
                ["P2-2", "品二-不同水印", 4, "NY2150 0.10 1/1 41*49", "NY2150", 0.1, "1/1", "41*49", "", "有水印", "A2", "无订单A级", "总厚", 1.3, 999],
                ["P2-3", "品二-不同类别", 14, "NY2150 0.10 1/1 41*49", "NY2170", 0.1, "1/1", "41*49", "", "", "A2", "无订单A级", "总厚", 1.3, 999],
                ["P2-4", "品二-不同厚度", 15, "NY2150 0.10 1/1 41*49", "NY2150", 0.11, "1/1", "41*49", "", "", "A2", "无订单A级", "总厚", 1.3, 999],
                ["P2-5", "品二-不同铜箔", 16, "NY2150 0.10 1/1 41*49", "NY2150", 0.1, "1/2", "41*49", "", "", "A2", "无订单A级", "总厚", 1.3, 999],
                ["P2-6", "品二-不同尺寸", 17, "NY2150 0.10 1/1 41*49", "NY2150", 0.1, "1/1", "43*49", "", "", "A2", "无订单A级", "总厚", 1.3, 999],
                ["P2-7", "品二-不同规格", 18, "NY2150 0.10 1/1 43*49", "NY2150", 0.1, "1/1", "41*49", "", "", "A2", "无订单A级", "总厚", 1.3, 999],
                ["P3", "品三", 7, "NY2150 0.20 1/2 41*49", "NY2150", 0.2, "1/2", "41*49", "", "", "B3", "无订单B级", "总厚", 1.4, 999],
                ["P4", "删除", 11, "NY2150 0.30 1/1 41*49", "NY2150", 0.3, "1/1", "41*49", "", "", "50%", "无订单", "总厚", 1.5, 999],
                ["P5", "删除", 12, "NY2150 0.40 1/1 41*49", "NY2150", 0.4, "1/1", "41*49", "", "", "", "无订单", "总厚", 1.5, 999],
                ["P6", "删除", 13, "NY2150 0.50 1/1 41*49", "NY2150", 0.5, "1/1", "41*49", "", "", 0.5, "无订单", "总厚", 1.5, 999],
            ],
        )
        write_input(
            jiangxi,
            [
                ["J1", "江一", 5, "NY2150 0.10 1/1 41*49", "NY2150", 0.1, "1/1", "41*49", "", "", "A1", "无订单管制品", "总厚", 1.0, 999],
                ["J2", "删除", 6, "NY2150 报废 0.10 1/1 41*49", "NY2150", 0.1, "1/1", "41*49", "", "", "A1", "无订单A级", "总厚", 1.0, 999],
                ["J3", "江三", "非数字", "NY2150 0.20 1/1 41*49", "NY2150", 0.2, "1/1", "41*49", "", "", "B3", "无订单退货", "总厚", 1.0, 999],
                ["J4", "删除", 8, "NY2150 0.30 1/1 41*49", "NY2150", 0.3, "1/1", "41*49", "", "", "A1", "有订单", "总厚", 1.0, 999],
            ],
        )

        rows, stats = load_inventory_rows(shanghai, jiangxi)

        self.assertEqual({"A": 8, "B": 2}, {grade: len(items) for grade, items in rows.items()})
        shanghai_a_rows = [row for row in rows["A"] if row.plant == "上海"]
        self.assertEqual(7, len(shanghai_a_rows))
        shanghai_a = next(row for row in shanghai_a_rows if row.product_no == "P1")
        shanghai_watermark = next(row for row in shanghai_a_rows if row.watermark == "有水印")
        jiangxi_a = next(row for row in rows["A"] if row.plant == "江西")
        self.assertEqual(5, shanghai_a.quantity)
        self.assertEqual(1998, shanghai_a.folded_large_quantity)
        self.assertEqual("P1", shanghai_a.product_no)
        self.assertEqual(4, shanghai_watermark.quantity)
        self.assertEqual(999, shanghai_watermark.folded_large_quantity)
        self.assertEqual(5, jiangxi_a.quantity)
        self.assertEqual(999, jiangxi_a.folded_large_quantity)
        self.assertEqual(3, stats["shanghai"]["filtered"])
        self.assertEqual(2, stats["jiangxi"]["filtered"])
        jiangxi_b = next(row for row in rows["B"] if row.plant == "江西")
        self.assertEqual(0, jiangxi_b.quantity)
        self.assertTrue(jiangxi_b.parse_status.startswith("待确认"))

    def test_workbook_structure_and_text_copper(self) -> None:
        shanghai = self.root / "shanghai.xlsx"
        jiangxi = self.root / "jiangxi.xlsx"
        write_input(
            shanghai,
            [["P1", "品一", 2, "NY2150 0.10 1/1 41*49", "NY2150", 0.1, "1/1", "41*49", "", "", "A1", "无订单A级", "总厚", 1.2, 888]],
        )
        write_input(
            jiangxi,
            [["J1", "品二", 3, "NY2150 0.20 1/2 41*49", "NY2150", 0.2, "1/2", "41*49", "", "", "B3", "无订单B级", "总厚", 1.0, 777]],
        )
        rows, _ = load_inventory_rows(shanghai, jiangxi)

        for grade, headers in (("A", A_HEADERS), ("B", B_HEADERS)):
            output = self.root / f"{grade}.xlsx"
            build_inventory_workbook(rows[grade], grade, output)
            workbook = load_workbook(output, data_only=False)
            try:
                self.assertIn("导航仪表盘", workbook.sheetnames)
                self.assertIn("全部明细", workbook.sheetnames)
                self.assertIn("胶系厚度汇总", workbook.sheetnames)
                self.assertIn("光板明细", workbook.sheetnames)
                self.assertIn("异常待确认", workbook.sheetnames)
                all_sheet = workbook["全部明细"]
                self.assertEqual(headers, [cell.value for cell in all_sheet[1]])
                self.assertEqual(24 if grade == "A" else 21, len(headers))
                self.assertIn("折合大板", headers)
                self.assertNotIn("排版结构", headers)
                self.assertEqual(1, len(all_sheet.tables))
                self.assertEqual(1, len(workbook["导航仪表盘"]._charts))
                copper_col = headers.index("铜箔") + 1
                copper_cell = all_sheet.cell(2, copper_col)
                self.assertEqual("s", copper_cell.data_type)
                self.assertIn(copper_cell.value, {"1/1", "1/2"})
                summary_sheet = workbook["胶系厚度汇总"]
                self.assertEqual("折合大板合计", summary_sheet["F3"].value)
                self.assertEqual(888 if grade == "A" else 777, summary_sheet["F4"].value)
            finally:
                workbook.close()

    def test_plan_a_filter_composite_merge_and_age_note(self) -> None:
        source = self.root / "plan-a.xlsx"
        write_plan_input(
            source,
            [
                ["首个品名", "P1", " NY2150 0.10 1/1 41*49 A级 有水印 ", "1*10", "江西", "1A", "来源胶系", 0.1, "1/1", 1.2, 3, 2, 4],
                ["第二品名", "P2", "NY2150 0.10 1/1 41*49 A级 有水印", "2*10", "江西", "1A", "来源胶系", 0.1, "1/1", 1.3, 9, 3, 5],
                ["上海同规格", "S1", "NY2150 0.10 1/1 41*49 A级 有水印", "1*10", "上海", "1A", "来源胶系", 0.1, "1/1", 1.1, 5, 7, 8],
                ["不同类别", "P3", "NY2150 0.10 1/1 41*49 A级 有水印", "1*10", "江西", "1B", "来源胶系", 0.1, "1/1", 1.2, 4, 11, 12],
                ["排除", "B1", "NY2150 0.10 1/1 41*49 b3  ", "1*10", "江西", "1A", "来源胶系", 0.1, "1/1", 1.2, 1, 99, 99],
            ],
        )

        rows, stats = load_plan_a_rows(source)

        self.assertEqual(4, stats["plan_a"]["kept"])
        self.assertEqual(1, stats["plan_a"]["filtered"])
        self.assertEqual(3, len(rows))
        merged = next(row for row in rows if row.plant == "江西" and row.product_no == "P1")
        self.assertEqual(5, merged.quantity)
        self.assertEqual(9, merged.folded_large_quantity)
        self.assertEqual("首个品名", merged.product_name)
        self.assertEqual("1*10", merged.layout)
        self.assertEqual("来源胶系", merged.glue)
        self.assertIn("库龄：3–9个月", merged.special_note)
        self.assertEqual(7, next(row for row in rows if row.plant == "上海").quantity)

    def test_plan_a_workbook_has_reference_25_columns(self) -> None:
        source = self.root / "plan-a.xlsx"
        output = self.root / "plan-a-output.xlsx"
        write_plan_input(
            source,
            [["品名", "P1", "NY2150 0.10 1/1 41*49 A级 无水印", "1*10", "江西", "1A", "NY2150", 0.1, "1/1", 1.2, 6, 2, 4]],
        )
        rows, _ = load_plan_a_rows(source)
        build_inventory_workbook(rows, "A", output, workbook_mode=PLAN_A_MODE)

        workbook = load_workbook(output, data_only=False)
        try:
            all_sheet = workbook["全部明细"]
            self.assertEqual(25, len(PLAN_A_HEADERS))
            self.assertEqual(PLAN_A_HEADERS, [cell.value for cell in all_sheet[1]])
            self.assertEqual("排版结构", all_sheet["D1"].value)
            self.assertEqual(2, all_sheet["J2"].value)
            self.assertEqual(4, all_sheet["K2"].value)
            self.assertIn("计划A级", workbook["导航仪表盘"]["A1"].value)
            self.assertEqual(1, len(workbook["导航仪表盘"]._charts))
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
