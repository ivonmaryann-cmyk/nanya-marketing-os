from __future__ import annotations

import json
import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from .db import get_setting, set_setting
from .paths import DEFAULT_RULES_DIR, TRANSCODE_AGENT_RULES_DIR, TRANSCODE_AGENT_RULES_VERSIONS_DIR
from .transcode_agent_standard import OFFICIAL_GRADE_CODES


FEATURE_KEY = "transcode_agent"
RULE_FILENAME = "transcode_agent_rules.xlsx"
ORIGINAL_FILENAME = "customer_special_master.xlsx"
STRUCTURED_DRAFT_FILENAME = "customer_special_structured_draft.xlsx"
MAPPING_TABLE_FILENAME = "transcode_agent_mapping_tables.xlsx"
DEFAULT_AGENT_RULE_VERSION = "transcode_agent_bundled_v1"
DEFAULT_AGENT_RULE_DIR = DEFAULT_RULES_DIR / "transcode_agent"

MACHINE_RULE_HEADERS = [
    "规则ID",
    "启用",
    "客户代码",
    "客户简称",
    "物料类别",
    "来源字段",
    "原始字段",
    "规则文本",
    "条件文本",
    "条件胶系",
    "条件关键词",
    "条件铜厚",
    "条件厚度",
    "条件尺寸",
    "覆盖字段",
    "覆盖值",
    "命中来源",
    "优先级",
    "强制执行",
    "待确认",
    "来源行号",
    "规则解释",
    "跳过原因",
]

FIELD_TO_OVERRIDE = {
    "胶系": "glue_code",
    "基板厚度": "thickness_code",
    "铜箔规格": "copper_code",
    "基板尺寸": "size_code",
    "胶水类别": "glue_category_code",
    "铜箔类型+印字/非印字": "copper_type_code",
    "基板级别": "grade_code",
    "总/芯厚": "tc_code",
    "组合结构": "struct_code",
    "铜箔厂商": "copper_vendor",
    "玻布厂商": "cloth_vendor",
    "配方代码": "formula_code",
    "PP长度": "pp_length",
    "PP级别": "pp_grade",
    "PP窄幅宽": "pp_width",
    "GT长短秒": "pp_gt",
    "树脂含量": "pp_rc",
    "小片尺寸": "pp_piece_size",
}

EXECUTABLE_FIELDS = {
    "glue_code",
    "thickness_code",
    "copper_code",
    "size_code",
    "glue_category_code",
    "copper_type_code",
    "grade_code",
    "tc_code",
}

# 这里只放开规则值校验；是否触发仍必须由客户规则或明确规格条件决定。
GRADE_CODES = set(OFFICIAL_GRADE_CODES)
GLUE_MODEL_TO_CODE = {
    "NY2140": "2A",
    "2140": "2A",
    "NY2150": "2B",
    "2150": "2B",
    "NY2170": "2C",
    "2170": "2C",
}
TECHNICAL_DETERMINISTIC_STRUCTURED_SOURCE_ROWS = {"43", "185", "206", "207", "213"}
TECHNICAL_RAW_CCL_DETERMINISTIC_SOURCE_ROWS = {"206", "207", "213"}
TECHNICAL_FULLY_SUPPORTED_SOURCE_ROWS = {"23", "206", "207", "213"}

CONFIRMED_DRAFT_COLUMNS = {
    "结构化处理状态",
    "通用特殊规则_结构化",
    "CCL特殊规则_结构化",
}

CONFIRMED_DRAFT_STATUSES = {"已结构化草稿", "可执行草稿"}
CONFIRMED_RULE_STATUS = "可执行草稿"

IGNORED_CONFIRMED_FIELDS = {"组合结构", "铜箔厂商", "玻布厂商", "配方代码"}
IGNORED_CONFIRMED_OVERRIDES = {
    "struct_code",
    "copper_vendor",
    "cloth_vendor",
    "formula_code",
    "reference_copper_vendor",
    "reference_glass_vendor",
    "non_transcode_note",
}
ORDER_SEMANTIC_TERMS = ("订单", "第5码", "第5位", "客户订单", "订单内容", "订单字段")

COPPER_TYPE_VALUE_MAP = {
    "HVLP5": "J",
    "HVLP4": "Z",
    "HVLP3": "K",
    "HVLP2": "P",
    "HVLP1": "O",
    "HVLP": "O",
    "RTF4": "G",
    "RTF3": "A",
    "RTF2": "B",
    "RTF1": "R",
    "RTF": "R",
    "VLP": "L",
    "有水印": "Q",
}

MAPPING_TABLE_HEADERS = {
    "Agent胶系主表": [
        "映射ID", "启用", "胶系编号", "胶系名称", "胶系分类", "输出胶系代码",
        "来源文件", "来源行号", "备注",
    ],
    "Agent胶系兼容别名": [
        "映射ID", "启用", "兼容名称", "标准胶系编号", "标准胶系名称",
        "输出胶系代码", "来源批次", "规则文本", "备注",
    ],
    "Agent胶系选择规则": [
        "映射ID", "启用", "胶系名称", "条件客户代码", "条件客户简称", "条件关键词",
        "输出胶系代码", "优先级", "来源批次", "规则文本", "备注",
    ],
    "Agent基础条件规则": [
        "映射ID", "启用", "物料类别", "条件胶系", "条件关键词", "关键词模式",
        "覆盖胶系代码", "覆盖胶水类别", "覆盖基板级别", "来源批次", "规则文本", "备注",
    ],
    "客户规则组": [
        "映射ID", "启用", "规则组ID", "规则组名称", "客户代码", "客户简称",
        "主规则客户代码", "主规则客户简称", "来源批次", "备注",
    ],
    "客户字段映射": [
        "映射ID",
        "启用",
        "客户代码",
        "客户简称",
        "来源批次",
        "来源行号",
        "样本数",
        "条件胶系",
        "条件关键词",
        "覆盖字段",
        "覆盖值",
        "规则文本",
        "备注",
    ],
    "客户单边尺寸映射": [
        "映射ID",
        "启用",
        "客户代码",
        "客户简称",
        "来源行号",
        "客户单边尺寸",
        "厂内单边尺寸",
        "适用字段",
        "来源字段",
        "规则文本",
        "备注",
    ],
    "客户尺寸映射": [
        "映射ID",
        "启用",
        "客户代码",
        "客户简称",
        "来源行号",
        "客户尺寸W",
        "客户尺寸H",
        "厂内尺寸W",
        "厂内尺寸H",
        "目标size_code",
        "来源字段",
        "规则文本",
        "备注",
    ],
    "客户尺寸算法": [
        "映射ID",
        "启用",
        "客户代码",
        "客户简称",
        "来源行号",
        "算法类型",
        "加大W",
        "加大H",
        "适用条件",
        "来源字段",
        "规则文本",
        "备注",
    ],
    "客户厚度映射": [
        "映射ID",
        "启用",
        "客户代码",
        "客户简称",
        "来源行号",
        "客户厚度写法",
        "厚度mm",
        "厚度mil",
        "总芯厚口径",
        "来源字段",
        "规则文本",
        "备注",
    ],
    "客户物料编码口径": [
        "映射ID",
        "启用",
        "客户代码",
        "客户简称",
        "来源行号",
        "物料编码模式",
        "命中值",
        "总芯厚口径",
        "来源字段",
        "规则文本",
        "备注",
    ],
    "外部尺寸表引用": [
        "映射ID",
        "启用",
        "客户代码",
        "客户简称",
        "来源行号",
        "引用文件",
        "引用Sheet",
        "来源字段",
        "规则文本",
        "备注",
    ],
    "待接入规则": [
        "映射ID",
        "启用",
        "客户代码",
        "客户简称",
        "来源行号",
        "技术类型",
        "原始规则",
        "规则来源说明",
        "建议处理",
        "备注",
    ],
}


def _history_key() -> str:
    return "transcode_agent_rule_history"


def _active_key() -> str:
    return "active_transcode_agent_rule_version"


def _read_history() -> list[dict]:
    raw = get_setting(_history_key(), "[]") or "[]"
    try:
        data = json.loads(raw)
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        return []


def _write_history(history: list[dict]) -> None:
    set_setting(_history_key(), json.dumps(history[:50], ensure_ascii=False))


def get_transcode_agent_rule_history() -> list[dict]:
    return _read_history()


def get_active_transcode_agent_rule_version() -> str:
    return get_setting(_active_key(), "") or ""


def ensure_default_transcode_agent_rule_version() -> str:
    active = get_active_transcode_agent_rule_version()
    if active and get_transcode_agent_rule_file_path(active).exists():
        return active
    required = [
        ORIGINAL_FILENAME,
        STRUCTURED_DRAFT_FILENAME,
        RULE_FILENAME,
        MAPPING_TABLE_FILENAME,
    ]
    missing = [name for name in required if not (DEFAULT_AGENT_RULE_DIR / name).exists()]
    if missing:
        raise FileNotFoundError(f"内置营销转码Agent规则资产缺失：{', '.join(missing)}")
    version_dir = TRANSCODE_AGENT_RULES_VERSIONS_DIR / DEFAULT_AGENT_RULE_VERSION
    version_dir.mkdir(parents=True, exist_ok=True)
    for filename in required:
        shutil.copy2(DEFAULT_AGENT_RULE_DIR / filename, version_dir / filename)
    set_setting(_active_key(), DEFAULT_AGENT_RULE_VERSION)
    _write_history([
        {
            "version": DEFAULT_AGENT_RULE_VERSION,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_by": "system",
            "remark": "首次启动加载项目内置营销转码Agent规则",
            "source_file": ORIGINAL_FILENAME,
            "rule_count": len(load_transcode_agent_rules(DEFAULT_AGENT_RULE_VERSION)),
        }
    ])
    return DEFAULT_AGENT_RULE_VERSION


def get_transcode_agent_rule_dir(version: str | None = None) -> Path:
    rule_version = version or get_active_transcode_agent_rule_version()
    return TRANSCODE_AGENT_RULES_VERSIONS_DIR / rule_version if rule_version else TRANSCODE_AGENT_RULES_DIR


def get_transcode_agent_rule_file_path(version: str | None = None) -> Path:
    return get_transcode_agent_rule_dir(version) / RULE_FILENAME


def get_transcode_agent_original_file_path(version: str | None = None) -> Path:
    return get_transcode_agent_rule_dir(version) / ORIGINAL_FILENAME


def get_transcode_agent_mapping_table_file_path(version: str | None = None) -> Path:
    return get_transcode_agent_rule_dir(version) / MAPPING_TABLE_FILENAME


def publish_transcode_agent_glue_asset_version(
    mapping_file: Path | None = None,
    *,
    updated_by: str,
    remark: str = "",
) -> str:
    source_mapping = Path(mapping_file or (DEFAULT_AGENT_RULE_DIR / MAPPING_TABLE_FILENAME))
    if not source_mapping.exists():
        raise FileNotFoundError(f"未找到Agent映射资产：{source_mapping}")
    workbook = openpyxl.load_workbook(source_mapping, read_only=True, data_only=True)
    required_glue_sheets = {"Agent胶系主表", "Agent胶系兼容别名", "Agent胶系选择规则"}
    missing = sorted(required_glue_sheets - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"Agent胶系资产缺少Sheet：{', '.join(missing)}")
    master_sheet = workbook["Agent胶系主表"]
    first_data_row = next(
        master_sheet.iter_rows(min_row=2, max_row=2, values_only=True),
        None,
    )
    if not first_data_row or not any(str(value or "").strip() for value in first_data_row):
        raise ValueError("Agent胶系主表没有可发布数据")

    current_version = get_active_transcode_agent_rule_version() or ensure_default_transcode_agent_rule_version()
    current_dir = get_transcode_agent_rule_dir(current_version)
    version = datetime.now().strftime("transcode_agent_rules_%Y%m%d_%H%M%S")
    version_dir = TRANSCODE_AGENT_RULES_VERSIONS_DIR / version
    shutil.copytree(current_dir, version_dir)
    shutil.copy2(source_mapping, version_dir / MAPPING_TABLE_FILENAME)
    set_setting(_active_key(), version)

    history = _read_history()
    history.insert(
        0,
        {
            "version": version,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_by": updated_by,
            "remark": remark or "发布营销转码Agent最新版胶系主表及历史兼容映射",
            "source_file": source_mapping.name,
            "rule_count": len(load_transcode_agent_rules(version)),
            "base_version": current_version,
            "updated_sheets": "Agent胶系主表, Agent胶系兼容别名, Agent胶系选择规则",
        },
    )
    _write_history(history)
    return version


def publish_cleaned_active_mapping_version(*, updated_by: str, remark: str = "") -> str:
    current_version = get_active_transcode_agent_rule_version() or ensure_default_transcode_agent_rule_version()
    current_dir = get_transcode_agent_rule_dir(current_version)
    structured_draft = current_dir / STRUCTURED_DRAFT_FILENAME
    if not structured_draft.exists():
        raise FileNotFoundError(f"当前Agent规则版本缺少确认草稿：{structured_draft}")

    rebuilt_tables, _ = parse_confirmed_customer_special_mapping_tables(structured_draft)
    pending_rows = rebuilt_tables.get("待接入规则", [])

    version = datetime.now().strftime("transcode_agent_rules_%Y%m%d_%H%M%S")
    version_dir = TRANSCODE_AGENT_RULES_VERSIONS_DIR / version
    shutil.copytree(current_dir, version_dir)
    mapping_path = version_dir / MAPPING_TABLE_FILENAME
    workbook = openpyxl.load_workbook(mapping_path)
    worksheet = workbook["待接入规则"]
    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    headers = MAPPING_TABLE_HEADERS["待接入规则"]
    for row in pending_rows:
        worksheet.append([row.get(header, "") for header in headers])
    workbook.save(mapping_path)
    set_setting(_active_key(), version)

    history = _read_history()
    history.insert(
        0,
        {
            "version": version,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_by": updated_by,
            "remark": remark or "清理已分流的重复待接入规则，保留当前全部已发布映射",
            "source_file": structured_draft.name,
            "rule_count": len(load_transcode_agent_rules(current_version)),
            "base_version": current_version,
            "updated_sheets": "待接入规则,转换说明",
        },
    )
    _write_history(history)
    return version


def save_new_transcode_agent_rule_version(
    rule_file: FileStorage,
    *,
    updated_by: str,
    remark: str = "",
) -> str:
    if not rule_file or not rule_file.filename:
        raise ValueError("请上传客户特殊清单结构化 Excel 文件")
    if not rule_file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("客户特殊清单仅支持 .xlsx / .xlsm 文件")

    version = datetime.now().strftime("transcode_agent_rules_%Y%m%d_%H%M%S")
    version_dir = TRANSCODE_AGENT_RULES_VERSIONS_DIR / version
    version_dir.mkdir(parents=True, exist_ok=True)
    original_path = version_dir / ORIGINAL_FILENAME
    rule_path = version_dir / RULE_FILENAME
    mapping_path = version_dir / MAPPING_TABLE_FILENAME
    rule_file.save(original_path)

    rules, summary = parse_customer_special_master(original_path)
    if summary.get("source_type") == "confirmed_structured_draft":
        shutil.copy2(original_path, version_dir / STRUCTURED_DRAFT_FILENAME)
        mapping_tables, mapping_summary = parse_confirmed_customer_special_mapping_tables(original_path)
        # 正确码回归样本形成的客户字段映射不属于草稿生成物。
        # 上传新草稿时继承当前已发布映射，避免确定性回归规则丢失。
        published_tables = load_transcode_agent_mapping_tables()
        for maintained_sheet in (
            "Agent胶系主表",
            "Agent胶系兼容别名",
            "Agent胶系选择规则",
            "Agent基础条件规则",
            "客户规则组",
            "客户字段映射",
        ):
            mapping_tables[maintained_sheet] = published_tables.get(maintained_sheet, [])
        build_mapping_table_workbook(mapping_path, mapping_tables, mapping_summary)
    build_machine_rule_workbook(rule_path, rules, summary)
    set_setting(_active_key(), version)

    history = _read_history()
    history.insert(
        0,
        {
            "version": version,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_by": updated_by,
            "remark": remark or "上传客户特殊清单并转换为营销转码Agent规则",
            "source_file": secure_filename(rule_file.filename) or ORIGINAL_FILENAME,
            "rule_count": len(rules),
        },
    )
    _write_history(history)
    return version


def parse_customer_special_master(path: Path) -> tuple[list[dict], dict]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    if _is_confirmed_structured_draft(workbook):
        return parse_confirmed_customer_special_draft(path, workbook=workbook)

    worksheet = workbook["客户特殊清单"] if "客户特殊清单" in workbook.sheetnames else workbook.worksheets[0]
    header_map = {
        _clean(worksheet.cell(1, col).value): col
        for col in range(1, worksheet.max_column + 1)
        if _clean(worksheet.cell(1, col).value)
    }
    required = ["客户代码", "客户简称", "CCL特殊规则", "PP特殊规则", "PP小片特殊规则"]
    missing = [name for name in required if name not in header_map]
    if missing:
        raise ValueError(f"客户特殊清单缺少列：{', '.join(missing)}")

    rows: list[dict] = []
    for row_idx in range(2, worksheet.max_row + 1):
        customer_code = _clean(worksheet.cell(row_idx, header_map["客户代码"]).value)
        customer_name = _clean(worksheet.cell(row_idx, header_map["客户简称"]).value)
        if not customer_code and not customer_name:
            continue
        if _is_template_row(customer_code, customer_name):
            continue
        for source_col, material_type in [
            ("CCL特殊规则", "CCL"),
            ("PP特殊规则", "PP"),
            ("PP小片特殊规则", "PP小片"),
        ]:
            source_text = _clean_multiline(worksheet.cell(row_idx, header_map[source_col]).value)
            rows.extend(_parse_structured_cell(customer_code, customer_name, material_type, source_col, source_text, row_idx))
        common_text = _clean_multiline(worksheet.cell(row_idx, header_map.get("通用特殊规则", 0)).value) if "通用特殊规则" in header_map else ""
        if common_text:
            rows.extend(_parse_free_text_rule(customer_code, customer_name, "通用", "通用特殊规则", common_text, row_idx))
        non_transcode_text = _clean_multiline(worksheet.cell(row_idx, header_map.get("非影响转码备注", 0)).value) if "非影响转码备注" in header_map else ""
        if non_transcode_text:
            rows.append(_make_rule(customer_code, customer_name, "通用", "非影响转码备注", "非影响转码备注", non_transcode_text, row_idx, "non_transcode_note", "", pending="否"))

    for idx, rule in enumerate(rows, 1):
        rule["规则ID"] = f"TAR-{idx:05d}"

    summary = {
        "source_type": "customer_special_master",
        "source_path": str(path),
        "rule_count": len(rows),
        "customer_count": len({(_clean(rule["客户代码"]), _clean(rule["客户简称"])) for rule in rows}),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return rows, summary


def parse_confirmed_customer_special_draft(path: Path, *, workbook=None) -> tuple[list[dict], dict]:
    workbook = workbook or openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook["结构化草稿"]
    header_map = _worksheet_header_map(worksheet)
    missing = [name for name in CONFIRMED_DRAFT_COLUMNS if name not in header_map]
    if missing:
        raise ValueError(f"确认草稿缺少列：{', '.join(missing)}")

    rows: list[dict] = []
    skipped: list[dict] = []
    status_counts: dict[str, int] = {}
    source_columns = [
        ("CCL特殊规则_结构化", "CCL特殊规则", 1000),
        ("通用特殊规则_结构化", "通用特殊规则", 900),
    ]
    for row_idx in range(2, worksheet.max_row + 1):
        row_status = _cell_by_header(worksheet, row_idx, header_map, "结构化处理状态")
        status_counts[row_status or "空"] = status_counts.get(row_status or "空", 0) + 1
        customer_code = _cell_by_header(worksheet, row_idx, header_map, "客户代码")
        customer_name = _cell_by_header(worksheet, row_idx, header_map, "客户简称")
        source_row = _cell_by_header(worksheet, row_idx, header_map, "来源行号") or str(row_idx)
        row_source_note = _cell_by_header(worksheet, row_idx, header_map, "规则来源说明", multiline=True)
        parse_structured_rules = (
            row_status in CONFIRMED_DRAFT_STATUSES
            or source_row in TECHNICAL_DETERMINISTIC_STRUCTURED_SOURCE_ROWS
        )
        if parse_structured_rules:
            for column_name, source_label, base_priority in source_columns:
                source_text = _cell_by_header(worksheet, row_idx, header_map, column_name, multiline=True)
                for rule_text in _iter_confirmed_rule_lines(source_text):
                    rule, reason = _parse_confirmed_rule_line(
                        rule_text,
                        customer_code,
                        customer_name,
                        source_label,
                        source_row,
                        base_priority,
                    )
                    if rule:
                        rows.append(rule)
                    elif reason:
                        skipped.append(
                            {
                                "来源行号": source_row,
                                "客户代码": customer_code,
                                "客户简称": customer_name,
                                "来源字段": source_label,
                                "规则文本": rule_text,
                                "跳过原因": reason,
                            }
                        )
                rows.extend(
                    _parse_supplemental_review_rules(
                        source_text,
                        customer_code,
                        customer_name,
                        source_label,
                        source_row,
                        base_priority,
                    )
                )
        if row_status in CONFIRMED_DRAFT_STATUSES or source_row in TECHNICAL_RAW_CCL_DETERMINISTIC_SOURCE_ROWS:
            raw_ccl_text = _cell_by_header(worksheet, row_idx, header_map, "CCL特殊规则", multiline=True)
            rows.extend(
                _parse_raw_ccl_deterministic_rules(
                    raw_ccl_text,
                    customer_code,
                    customer_name,
                    source_row,
                    1060,
                )
            )
            rows.extend(_parse_lejian_ccl_grade_rules(raw_ccl_text, customer_code, customer_name, source_row, 1060))
        if parse_structured_rules:
            rows.extend(
                _parse_supplemental_review_rules(
                    row_source_note,
                    customer_code,
                    customer_name,
                    "规则来源说明",
                    source_row,
                    1100,
                )
            )

    rows = _dedupe_confirmed_rules(rows)
    rows, conflict_count = _drop_conflicting_confirmed_rules(rows, skipped)
    for idx, rule in enumerate(rows, 1):
        rule["规则ID"] = f"TAR-{idx:05d}"

    summary = {
        "source_type": "confirmed_structured_draft",
        "source_path": str(path),
        "rule_count": len(rows),
        "customer_count": len({(_clean(rule["客户代码"]), _clean(rule["客户简称"])) for rule in rows}),
        "skipped_count": len(skipped),
        "conflict_count": conflict_count,
        "status_counts": json.dumps(status_counts, ensure_ascii=False),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    if skipped:
        summary["skipped_preview"] = "；".join(
            f"{item['来源行号']}:{item['客户简称']}:{item['跳过原因']}" for item in skipped[:20]
        )
    return rows, summary


def build_machine_rule_workbook(path: Path, rules: list[dict], summary: dict) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "机器规则"
    ws.append(MACHINE_RULE_HEADERS)
    for rule in rules:
        ws.append([rule.get(header, "") for header in MACHINE_RULE_HEADERS])

    ws_summary = workbook.create_sheet("转换说明")
    ws_summary.append(["项目", "值"])
    for key, value in summary.items():
        ws_summary.append([key, value])
    ws_summary.append(["说明", "该文件由客户级特殊清单转换生成，供营销转码Agent读取；旧转码功能不读取该规则。"])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in enumerate([14, 8, 14, 18, 10, 18, 18, 48, 34, 22, 28, 20, 18, 18, 18, 18, 16, 8, 10, 10, 10, 36, 36], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 90
    workbook.save(path)


def parse_confirmed_customer_special_mapping_tables(path: Path, *, workbook=None) -> tuple[dict[str, list[dict]], dict]:
    workbook = workbook or openpyxl.load_workbook(path, data_only=True)
    if not _is_confirmed_structured_draft(workbook):
        return {sheet: [] for sheet in MAPPING_TABLE_HEADERS}, {
            "source_type": "no_mapping_tables",
            "source_path": str(path),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    worksheet = workbook["结构化草稿"]
    header_map = _worksheet_header_map(worksheet)
    tables: dict[str, list[dict]] = {sheet: [] for sheet in MAPPING_TABLE_HEADERS}
    status_counts: dict[str, int] = {}
    technical_rows = 0

    for row_idx in range(2, worksheet.max_row + 1):
        row_status = _cell_by_header(worksheet, row_idx, header_map, "结构化处理状态")
        status_counts[row_status or "空"] = status_counts.get(row_status or "空", 0) + 1
        if "技术待支持" not in row_status:
            continue
        technical_rows += 1
        customer_code = _cell_by_header(worksheet, row_idx, header_map, "客户代码")
        customer_name = _cell_by_header(worksheet, row_idx, header_map, "客户简称")
        source_row = _cell_by_header(worksheet, row_idx, header_map, "来源行号") or str(row_idx)
        common_text = _cell_by_header(worksheet, row_idx, header_map, "通用特殊规则", multiline=True)
        ccl_text = _cell_by_header(worksheet, row_idx, header_map, "CCL特殊规则", multiline=True)
        feedback = _cell_by_header(worksheet, row_idx, header_map, "业务反馈", multiline=True)
        source_note = _cell_by_header(worksheet, row_idx, header_map, "规则来源说明", multiline=True)
        combined_text = "\n".join(item for item in [common_text, ccl_text, feedback, source_note] if item)
        base = {
            "启用": "否",
            "客户代码": customer_code,
            "客户简称": customer_name,
            "来源行号": source_row,
        }

        size_text = _extract_ccl_field_section(ccl_text, "基板尺寸")
        thickness_text = _extract_ccl_field_section(ccl_text, "基板厚度")
        tc_text = _extract_ccl_field_section(ccl_text, "总/芯厚")

        _append_size_mapping_rows(tables, base, size_text, ccl_text)
        _append_size_algorithm_rows(tables, base, size_text, combined_text)
        _append_thickness_mapping_rows(tables, base, thickness_text, tc_text, ccl_text)
        _append_material_code_rows(tables, base, tc_text, ccl_text)
        _append_external_reference_rows(tables, base, combined_text)
        if source_row not in TECHNICAL_FULLY_SUPPORTED_SOURCE_ROWS:
            _append_pending_mapping_row(tables, base, combined_text, source_note)

    for sheet_name, rows in tables.items():
        for idx, row in enumerate(rows, 1):
            row["映射ID"] = f"TAM-{_mapping_sheet_code(sheet_name)}-{idx:05d}"

    summary = {
        "source_type": "confirmed_structured_draft_mapping_tables",
        "source_path": str(path),
        "technical_row_count": technical_rows,
        "status_counts": json.dumps(status_counts, ensure_ascii=False),
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    for sheet_name, rows in tables.items():
        summary[f"{sheet_name}_count"] = len(rows)
    return tables, summary


def build_mapping_table_workbook(path: Path, tables: dict[str, list[dict]], summary: dict) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="7030A0")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name, headers in MAPPING_TABLE_HEADERS.items():
        ws = workbook.create_sheet(sheet_name)
        ws.append(headers)
        for row in tables.get(sheet_name, []):
            ws.append([row.get(header, "") for header in headers])
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for data_row in ws.iter_rows(min_row=2):
            for cell in data_row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, width in enumerate(_mapping_sheet_widths(sheet_name, len(headers)), 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws_summary = workbook.create_sheet("转换说明")
    ws_summary.append(["项目", "值"])
    for key, value in summary.items():
        ws_summary.append([key, value])
    ws_summary.append(["说明", "该文件由确认草稿中的技术待支持规则生成；已启用的辅助映射供营销转码Agent运行时读取，未启用项保留为后续接入线索。"])
    ws_summary.freeze_panes = "A2"
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 90
    workbook.save(path)


def _extract_ccl_field_section(text: str, field_name: str) -> str:
    source = _clean_multiline(text)
    if not source:
        return ""
    field_names = [
        "胶系",
        "基板厚度",
        "铜箔规格",
        "基板尺寸",
        "胶水类别",
        "铜箔类型+印字/非印字",
        "基板级别",
        "总/芯厚",
        "组合结构",
        "铜箔厂商",
        "玻布厂商",
        "配方代码",
    ]
    alternatives = "|".join(re.escape(name) for name in field_names if name != field_name)
    pattern = rf"{re.escape(field_name)}\s*[:：]\s*(.*?)(?=\s*(?:{alternatives})\s*[:：]|$)"
    match = re.search(pattern, source, flags=re.IGNORECASE | re.DOTALL)
    return _clean(match.group(1)) if match else ""


def _append_size_mapping_rows(tables: dict[str, list[dict]], base: dict, size_text: str, original_text: str) -> None:
    source = _clean(size_text).replace("×", "*")
    if not source:
        return
    full_size_pattern = r"(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)\s*=\s*(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)"
    for customer_w, customer_h, factory_w, factory_h in re.findall(full_size_pattern, source):
        tables["客户尺寸映射"].append(
            {
                **base,
                "启用": "是",
                "客户尺寸W": customer_w,
                "客户尺寸H": customer_h,
                "厂内尺寸W": factory_w,
                "厂内尺寸H": factory_h,
                "目标size_code": "",
                "来源字段": "CCL特殊规则/基板尺寸",
                "规则文本": f"{customer_w}*{customer_h}={factory_w}*{factory_h}",
                "备注": "运行时按客户完整尺寸映射覆盖size_code",
            }
        )

    source_without_full_sizes = re.sub(full_size_pattern, " ", source)
    for customer_side, factory_side in re.findall(
        r"(?<![\d.*])(\d{3,4})\s*=\s*(\d+(?:\.\d+)?)(?!\s*(?:mil|mm|\*))",
        source_without_full_sizes,
        flags=re.IGNORECASE,
    ):
        tables["客户单边尺寸映射"].append(
            {
                **base,
                "启用": "是",
                "客户单边尺寸": customer_side,
                "厂内单边尺寸": factory_side,
                "适用字段": "宽或高",
                "来源字段": "CCL特殊规则/基板尺寸",
                "规则文本": f"{customer_side}={factory_side}",
                "备注": "运行时按客户单边尺寸替换后覆盖size_code",
            }
        )


def _append_size_algorithm_rows(tables: dict[str, list[dict]], base: dict, size_text: str, combined_text: str) -> None:
    source = _clean(f"{size_text} {combined_text}")
    if "加大" not in source and "增加" not in source:
        return
    match = re.search(r"(?:加大|增加)\s*(\d+(?:\.\d+)?)", source)
    if not match:
        return
    delta = match.group(1)
    tables["客户尺寸算法"].append(
        {
            **base,
            "启用": "是",
            "算法类型": "尺寸加大",
            "加大W": delta,
            "加大H": delta,
            "适用条件": "默认全部尺寸",
            "来源字段": "CCL特殊规则/基板尺寸",
            "规则文本": size_text or combined_text,
            "备注": "运行时按尺寸加大算法覆盖size_code",
        }
    )


def _append_thickness_mapping_rows(tables: dict[str, list[dict]], base: dict, thickness_text: str, tc_text: str, original_text: str) -> None:
    source = _clean(f"{thickness_text} {tc_text}")
    if not source:
        return
    for alias, mil_value, mm_value in re.findall(
        r"(\d{2,4})\s*=\s*(\d+(?:\.\d+)?)\s*mil\s*=\s*(\d+(?:\.\d+)?)",
        source,
        flags=re.IGNORECASE,
    ):
        tables["客户厚度映射"].append(
            {
                **base,
                "启用": "是",
                "客户厚度写法": alias,
                "厚度mm": mm_value,
                "厚度mil": mil_value,
                "总芯厚口径": "",
                "来源字段": "CCL特殊规则/基板厚度",
                "规则文本": f"{alias}={mil_value}mil={mm_value}",
                "备注": "运行时按客户厚度别名覆盖thickness_code",
            }
        )
    for mil_value, tc_label in re.findall(r"(\d+(?:\.\d+)?)\s*mil\s*(?:含)?以上\s*(总厚|芯厚)", source, flags=re.IGNORECASE):
        tables["客户厚度映射"].append(
            {
                **base,
                "启用": "是",
                "客户厚度写法": f"{mil_value}mil含以上",
                "厚度mm": "",
                "厚度mil": mil_value,
                "总芯厚口径": "T" if tc_label == "总厚" else "C",
                "来源字段": "CCL特殊规则/总/芯厚",
                "规则文本": f"{mil_value}mil含以上{tc_label}",
                "备注": "运行时按客户厚度阈值覆盖总/芯厚口径",
            }
        )


def _append_material_code_rows(tables: dict[str, list[dict]], base: dict, tc_text: str, original_text: str) -> None:
    source = _clean(tc_text)
    if not source:
        return
    for code, tc_label in re.findall(r"(\d{3,})[.。]?\s*是\s*(芯厚|总厚)", source):
        tables["客户物料编码口径"].append(
            {
                **base,
                "启用": "是",
                "物料编码模式": "包含",
                "命中值": code,
                "总芯厚口径": "C" if tc_label == "芯厚" else "T",
                "来源字段": "CCL特殊规则/总/芯厚",
                "规则文本": f"{code}={tc_label}",
                "备注": "客户限定物料编码口径映射，运行时按数字边界匹配",
            }
        )


def _append_external_reference_rows(tables: dict[str, list[dict]], base: dict, combined_text: str) -> None:
    source = _clean(combined_text)
    if "新美亚" not in source or ("尺寸转换表" not in source and "尺寸对照表" not in source):
        return
    tables["外部尺寸表引用"].append(
        {
            **base,
            "启用": "是",
            "引用文件": "docs/develop0707/新美亚规格尺寸对照表.xlsx",
            "引用Sheet": "",
            "来源字段": "业务反馈/规则来源说明",
            "规则文本": source,
            "备注": "运行时读取新美亚尺寸对照表；631按表换算，632按英寸*25.4四舍五入",
        }
    )


def _append_pending_mapping_row(tables: dict[str, list[dict]], base: dict, combined_text: str, source_note: str) -> None:
    source = _clean_multiline(combined_text)
    if not source:
        return
    if _mapping_rule_is_already_routed(tables, base, source):
        return
    tables["待接入规则"].append(
        {
            **base,
            "技术类型": _detect_pending_mapping_type(source),
            "原始规则": source,
            "规则来源说明": source_note,
            "建议处理": _suggest_pending_mapping_action(source),
            "备注": "无法分流到已有映射、订单语义或当前编码范围的规则，需人工复核",
        }
    )


def _mapping_rule_is_already_routed(tables: dict[str, list[dict]], base: dict, source: str) -> bool:
    source_row = _clean(base.get("来源行号"))
    customer_code = _clean(base.get("客户代码"))
    customer_name = _clean(base.get("客户简称"))
    for sheet_name in (
        "客户单边尺寸映射",
        "客户尺寸映射",
        "客户尺寸算法",
        "客户厚度映射",
        "客户物料编码口径",
        "外部尺寸表引用",
    ):
        for row in tables.get(sheet_name, []):
            if _clean(row.get("来源行号")) != source_row:
                continue
            if _clean(row.get("客户代码")) != customer_code:
                continue
            if _clean(row.get("客户简称")) == customer_name:
                return True

    normalized = _clean(source).upper()
    if "订单" in normalized or "备注" in normalized or "第5码" in normalized:
        return True
    if "J0J0F0" in normalized or "24-30位" in normalized or "24~30位" in normalized:
        return True
    return False


def _detect_pending_mapping_type(text: str) -> str:
    source = _clean(text).upper()
    if "尺寸转换表" in source or "尺寸对照表" in source:
        return "外部尺寸表"
    if "尺寸" in source and ("加大" in source or "增加" in source):
        return "尺寸算法"
    if "尺寸" in source and "=" in source:
        return "尺寸映射"
    if "MIL" in source or "厚度" in source:
        return "厚度映射"
    if "物料编码" in source:
        return "物料编码口径"
    if "订单" in source or "备注" in source or "第5码" in source:
        return "订单语义待处理"
    if "J0J0F0" in source:
        return "后缀/扩展字段"
    if "TG" in source:
        return "胶系语义映射"
    return "待分类"


def _suggest_pending_mapping_action(text: str) -> str:
    mapping_type = _detect_pending_mapping_type(text)
    return {
        "外部尺寸表": "接入外部尺寸表读取和客户尺寸映射",
        "尺寸算法": "接入客户尺寸算法表并覆盖size_code",
        "尺寸映射": "接入客户尺寸/单边尺寸映射并覆盖size_code",
        "厚度映射": "接入客户厚度映射并覆盖thickness_code/tc_code",
        "物料编码口径": "确认输入字段后接入物料编码条件匹配",
        "订单语义待处理": "进入订单语义模块，不混入规格解析",
        "后缀/扩展字段": "确认编码后缀字段模型后再接入",
        "胶系语义映射": "接入客户胶系语义/默认胶系映射",
    }.get(mapping_type, "人工复核后再决定辅助映射类型")


def _mapping_sheet_code(sheet_name: str) -> str:
    return {
        "客户单边尺寸映射": "SIDE",
        "客户尺寸映射": "SIZE",
        "客户尺寸算法": "SIZA",
        "客户厚度映射": "THIK",
        "客户物料编码口径": "MATL",
        "外部尺寸表引用": "EXTS",
        "待接入规则": "PEND",
    }.get(sheet_name, "MAP")


def _mapping_sheet_widths(sheet_name: str, header_count: int) -> list[int]:
    defaults = [16, 8, 14, 18, 10, 18, 18, 18, 18, 26, 48, 36, 36]
    if sheet_name == "待接入规则":
        return [16, 8, 14, 18, 10, 18, 70, 70, 42, 36]
    if sheet_name == "外部尺寸表引用":
        return [16, 8, 14, 18, 10, 42, 18, 24, 70, 36]
    return defaults[:header_count]


def _is_confirmed_structured_draft(workbook) -> bool:
    if "结构化草稿" not in workbook.sheetnames:
        return False
    header_map = _worksheet_header_map(workbook["结构化草稿"])
    return CONFIRMED_DRAFT_COLUMNS.issubset(set(header_map))


def _worksheet_header_map(worksheet) -> dict[str, int]:
    return {
        _clean(worksheet.cell(1, col).value): col
        for col in range(1, worksheet.max_column + 1)
        if _clean(worksheet.cell(1, col).value)
    }


def _cell_by_header(worksheet, row_idx: int, header_map: dict[str, int], header: str, *, multiline: bool = False) -> str:
    col = header_map.get(header)
    if not col:
        return ""
    value = worksheet.cell(row_idx, col).value
    return _clean_multiline(value) if multiline else _clean(value)


def _iter_confirmed_rule_lines(text: str) -> list[str]:
    lines: list[str] = []
    for raw_line in _clean_multiline(text).splitlines():
        line = raw_line.strip()
        if not line or "字段=" not in line or "覆盖字段=" not in line:
            continue
        lines.append(line)
    return lines


def _parse_raw_ccl_deterministic_rules(
    text: str,
    customer_code: str,
    customer_name: str,
    source_row: str,
    base_priority: int,
) -> list[dict]:
    source = _clean_multiline(text)
    if not source:
        return []
    rules: list[dict] = []
    tc_text = _extract_ccl_field_section(source, "总/芯厚")
    rules.extend(_parse_raw_tc_threshold_rules(tc_text, customer_code, customer_name, source_row, base_priority))

    if source_row == "213":
        glue_text = _extract_ccl_field_section(source, "胶系")
        rules.extend(_parse_raw_tg_glue_rules(glue_text, customer_code, customer_name, source_row, base_priority))
    return rules


def _parse_raw_tc_threshold_rules(
    text: str,
    customer_code: str,
    customer_name: str,
    source_row: str,
    base_priority: int,
) -> list[dict]:
    source = _clean(text)
    if not source:
        return []
    rules: list[dict] = []
    threshold_pattern = (
        r"(\d+(?:\.\d+)?)\s*(?:含)?以上\s*(总厚|芯厚)"
        r".*?"
        r"(\d+(?:\.\d+)?)\s*以下\s*(总厚|芯厚)"
    )
    for upper_number, upper_label, lower_number, lower_label in re.findall(threshold_pattern, source):
        if upper_number != lower_number:
            continue
        for operator, label in ((f">={upper_number}", upper_label), (f"<{lower_number}", lower_label)):
            value = "T" if label == "总厚" else "C"
            rule_text = f"总/芯厚:{source}"
            rule = _make_rule(
                customer_code,
                customer_name,
                "CCL",
                "CCL特殊规则",
                "总/芯厚",
                rule_text,
                source_row,
                "tc_code",
                value,
                condition_text=f"厚度{operator}",
                pending="否",
            )
            rule["条件胶系"] = ""
            rule["条件关键词"] = ""
            rule["条件铜厚"] = ""
            rule["条件厚度"] = operator
            rule["条件尺寸"] = ""
            rule["命中来源"] = "客户特殊规则确认草稿/原始CCL确定性解析"
            rule["优先级"] = str(base_priority + 40)
            rule["规则解释"] = "由原始CCL总/芯厚阈值重建，避免结构化列把>=与<条件拆丢"
            rules.append(rule)
    return rules


def _parse_raw_tg_glue_rules(
    text: str,
    customer_code: str,
    customer_name: str,
    source_row: str,
    base_priority: int,
) -> list[dict]:
    source = _clean(text)
    if not source or "TG" not in source.upper():
        return []
    rules: list[dict] = []
    for clause in _split_clauses(source):
        clause_upper = clause.upper()
        if "订单" in clause or "=" not in clause:
            continue
        condition_text, value_raw = _split_condition_value(clause)
        has_default_tg_condition = (
            "没标注TG" in condition_text or "未标注TG" in condition_text or "没有标注TG" in condition_text
        )
        if not has_default_tg_condition and "TG130" not in clause_upper:
            continue
        value = _normalize_override_value("glue_code", value_raw)
        if not value:
            continue
        condition_keywords: list[str] = []
        if has_default_tg_condition:
            condition_keywords.append("__NO_TG__")
        if "无卤素" in condition_text or "CTI" in clause_upper:
            condition_keywords.append(
                condition_text.replace("和", "&").replace("且", "&").replace("＋", "&").replace("+", "&")
            )
        else:
            condition_keywords.extend(re.findall(r"TG\d+", clause_upper))
        condition_keywords = list(dict.fromkeys(condition_keywords))
        for keyword in condition_keywords:
            rule = _make_rule(
                customer_code,
                customer_name,
                "CCL",
                "CCL特殊规则",
                "胶系",
                f"胶系:{clause}",
                source_row,
                "glue_code",
                value,
                condition_text=f"关键词:{keyword}",
                pending="否",
            )
            rule["条件胶系"] = ""
            rule["条件关键词"] = keyword
            rule["条件铜厚"] = ""
            rule["条件厚度"] = ""
            rule["条件尺寸"] = ""
            rule["命中来源"] = "客户特殊规则确认草稿/原始CCL确定性解析"
            specificity = 30 if keyword == "__NO_TG__" else 40 + keyword.count("&") * 10 + keyword.count("和") * 10
            rule["优先级"] = str(base_priority + specificity)
            rule["规则解释"] = "由原始CCL胶系TG语义规则生成"
            rules.append(rule)
    return rules


def _parse_lejian_ccl_grade_rules(
    text: str,
    customer_code: str,
    customer_name: str,
    source_row: str,
    base_priority: int,
) -> list[dict]:
    source = _clean(text)
    if "乐健" not in _clean(customer_name) or "非汽车板" not in source or "基板" not in source:
        return []

    rules: list[dict] = []

    def add_rule(
        rule_text: str,
        value: str,
        *,
        condition_text: str = "",
        condition_glue: str = "",
        condition_thickness: str = "",
        condition_copper: str = "",
        priority_bonus: int = 0,
    ) -> None:
        rule = _make_rule(
            customer_code,
            customer_name,
            "CCL",
            "CCL特殊规则",
            "基板级别",
            rule_text,
            source_row,
            "grade_code",
            value,
            condition_text=condition_text,
            pending="否",
        )
        rule["条件胶系"] = condition_glue
        rule["条件关键词"] = ""
        rule["条件铜厚"] = condition_copper
        rule["条件厚度"] = condition_thickness
        rule["条件尺寸"] = ""
        rule["命中来源"] = "客户特殊规则确认草稿/原始CCL确定性解析"
        rule["优先级"] = str(base_priority + priority_bonus)
        rule["规则解释"] = "业务反馈确认PP不考虑，仅按CCL基板汽车板/非汽车板条件执行"
        rules.append(rule)

    add_rule(
        "乐健CCL基板默认下汽车板",
        "AC",
        priority_bonus=20,
    )
    add_rule(
        "NY2150 板厚1.1mm 全部下非汽车板",
        "A1",
        condition_text="胶系:NY2150；厚度:1.1",
        condition_glue="NY2150",
        condition_thickness="1.1",
        priority_bonus=90,
    )
    add_rule(
        "NY2150 板厚1.5/1.45mm 且铜厚H/H、1/1、2/2全部下非汽车板",
        "A1",
        condition_text="胶系:NY2150；厚度:1.5/1.45；铜厚:H/H,1/1,2/2",
        condition_glue="NY2150",
        condition_thickness="1.5/1.45",
        condition_copper="H/H,1/1,2/2",
        priority_bonus=100,
    )
    return rules


def _parse_supplemental_review_rules(
    text: str,
    customer_code: str,
    customer_name: str,
    source_label: str,
    source_row: str,
    base_priority: int,
) -> list[dict]:
    rules: list[dict] = []
    for raw_line in _clean_multiline(text).splitlines():
        line = _clean(raw_line)
        if source_label != "规则来源说明" and "状态=可执行草稿" not in line:
            continue
        if not any(term in line for term in ("厚度", "总厚")) or not any(token in line for token in ("为", "等于", "=")):
            continue
        total_thickness_actions = re.findall(
            r"总厚\s*(\d+(?:\.\d+)?)\s*(?:MM|mm)?(?:（含）|含)?以下(?:就)?(?:等于|=)\s*([A-Z0-9]{2})",
            line,
        )
        for thickness, code in total_thickness_actions:
            code = code.upper()
            if code not in GRADE_CODES:
                continue
            condition_text = f"厚度:<={thickness}"
            rule = _make_rule(
                customer_code,
                customer_name,
                "CCL",
                source_label,
                "业务反馈复核",
                line,
                source_row,
                "grade_code",
                code,
                condition_text=condition_text,
                pending="否",
            )
            rule["条件胶系"] = ""
            rule["条件厚度"] = f"<={thickness}"
            rule["条件关键词"] = ""
            rule["条件铜厚"] = ""
            rule["条件尺寸"] = ""
            rule["命中来源"] = "客户特殊规则确认草稿/业务反馈复核"
            rule["优先级"] = str(base_priority + 60)
            rule["规则解释"] = "由业务反馈复核行补充生成"
            rules.append(rule)
        glue_tokens = _extract_condition_glue(line).split("/")
        glue_tokens = [token for token in glue_tokens if token]
        if not glue_tokens:
            continue
        thickness_actions = re.findall(
            r"厚度\s*(>=|<=|>|<|≥|≤)\s*(\d+(?:\.\d+)?)\s*(?:MM|mm)?\s*为\s*([A-Z0-9]{2})",
            line,
        )
        for glue in glue_tokens:
            for operator, thickness, code in thickness_actions:
                code = code.upper()
                if code not in GRADE_CODES:
                    continue
                condition_text = f"胶系:{glue}；厚度:{operator}{thickness}"
                rule = _make_rule(
                    customer_code,
                    customer_name,
                    "CCL",
                    source_label,
                    "业务反馈复核",
                    line,
                    source_row,
                    "grade_code",
                    code,
                    condition_text=condition_text,
                    pending="否",
                )
                rule["条件胶系"] = glue
                rule["条件厚度"] = f"{operator}{thickness}"
                rule["条件关键词"] = ""
                rule["条件铜厚"] = ""
                rule["条件尺寸"] = ""
                rule["命中来源"] = "客户特殊规则确认草稿/业务反馈复核"
                rule["优先级"] = str(base_priority + 60)
                rule["规则解释"] = "由业务反馈复核行补充生成"
                rules.append(rule)
    return rules


def _parse_confirmed_rule_line(
    rule_text: str,
    customer_code: str,
    customer_name: str,
    source_label: str,
    source_row: str,
    base_priority: int,
) -> tuple[dict | None, str]:
    parts = _parse_structured_kv_line(rule_text)
    field_name = parts.get("字段", "")
    override_field = parts.get("覆盖字段", "")
    override_value_raw = parts.get("覆盖值", "")
    rule_status = parts.get("状态", "")
    condition_text = parts.get("条件", "")
    explanation = parts.get("依据", "") or parts.get("说明", "") or parts.get("确认点", "")

    if (
        _clean(customer_name) == "方正F7"
        and override_field == "tc_code"
        and _clean(override_value_raw).upper() == "C"
        and not condition_text
    ):
        condition_text = "厚度:<0.8"
        explanation = "方正F7默认输入芯厚；>=0.8mm由Agent总芯厚算法转总厚"
    if (
        _clean(customer_name) == "江苏博敏"
        and override_field == "grade_code"
        and re.fullmatch(r"AH\s+AD", _clean(override_value_raw).upper())
    ):
        override_value_raw = "AD"
        explanation = "原文AH为客户胶系代码，AD为基板级别代码"

    if rule_status != CONFIRMED_RULE_STATUS:
        return None, f"规则状态非{CONFIRMED_RULE_STATUS}"
    if field_name in IGNORED_CONFIRMED_FIELDS or override_field in IGNORED_CONFIRMED_OVERRIDES:
        return None, "本阶段排除字段"
    if override_field not in EXECUTABLE_FIELDS:
        return None, "覆盖字段不在本阶段执行白名单"
    customer_spec_absence_rule = (
        override_field == "grade_code"
        and "客户规格没有Q" in _clean(condition_text).upper()
    )
    if (
        _contains_order_semantic(condition_text) or _contains_order_semantic(override_value_raw)
    ) and not customer_spec_absence_rule:
        return None, "订单语义规则后续专门处理"
    if _contains_unsupported_negative_condition(condition_text) and not customer_spec_absence_rule:
        return None, "负向/排除条件本阶段暂不执行"

    normalized_value = _normalize_override_value(override_field, override_value_raw)
    if not normalized_value:
        return None, "覆盖值无法标准化"

    priority = base_priority
    condition_source = f"{condition_text}；{override_value_raw}"
    condition_glue = _extract_condition_glue(condition_source)
    condition_keywords = _extract_condition_keywords(condition_text)
    condition_copper = _extract_condition_copper(condition_source)
    condition_thickness = _extract_condition_thickness(condition_source)
    condition_size = _extract_condition_size(condition_source) if override_field == "size_code" else ""
    if override_field == "glue_code" and re.search(r"TG\d+", condition_keywords.upper()):
        condition_glue = ""
    if customer_code or customer_name:
        priority += 20
    if condition_glue:
        priority += 20
    if condition_copper or condition_thickness or condition_size:
        priority += 10
    if condition_keywords:
        priority += 10 + _condition_keyword_specificity_bonus(condition_keywords)

    rule = _make_rule(
        customer_code,
        customer_name,
        "CCL",
        source_label,
        field_name,
        rule_text,
        source_row,
        override_field,
        normalized_value,
        condition_text=condition_text,
        pending="否",
    )
    rule["条件胶系"] = condition_glue
    rule["条件关键词"] = condition_keywords
    rule["条件铜厚"] = condition_copper
    rule["条件厚度"] = condition_thickness
    rule["条件尺寸"] = condition_size
    rule["命中来源"] = "客户特殊规则确认草稿"
    rule["优先级"] = str(priority)
    rule["规则解释"] = explanation
    rule["跳过原因"] = ""
    return rule, ""


def _condition_keyword_specificity_bonus(value: str) -> int:
    bonus = 0
    for token in re.split(r"[/,，、;；]+", value):
        keyword = _clean(token)
        if not keyword:
            continue
        parts = [part for part in re.split(r"&|\+|和|且", keyword) if part]
        if len(parts) > 1:
            bonus += (len(parts) - 1) * 10
    return bonus


def _parse_structured_kv_line(line: str) -> dict[str, str]:
    text = re.sub(r"^【[^】]+】", "", _clean(line)).strip()
    result: dict[str, str] = {}
    for part in re.split(r"\s*\|\s*", text):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        key = _clean(key)
        value = _clean(value)
        if key:
            result[key] = value
    return result


def _contains_order_semantic(text: str) -> bool:
    source = _clean(text)
    return any(term in source for term in ORDER_SEMANTIC_TERMS)


def _contains_unsupported_negative_condition(text: str) -> bool:
    source = _clean(text)
    return bool(re.search(r"除.+外|非[A-Z0-9]|不是|其余", source, flags=re.IGNORECASE))


def _drop_conflicting_confirmed_rules(rows: list[dict], skipped: list[dict]) -> tuple[list[dict], int]:
    grouped: dict[tuple[str, str, str, str, str, str, str], list[dict]] = {}
    for rule in rows:
        key = (
            _clean(rule.get("客户代码")),
            _clean(rule.get("客户简称")),
            _clean(rule.get("来源字段")),
            _clean(rule.get("覆盖字段")),
            _clean(rule.get("条件文本")),
            _clean(rule.get("条件胶系")),
            _clean(rule.get("条件关键词")),
        )
        grouped.setdefault(key, []).append(rule)

    conflict_ids: set[int] = set()
    for items in grouped.values():
        values = {_clean(item.get("覆盖值")) for item in items}
        if len(values) <= 1:
            continue
        for item in items:
            conflict_ids.add(id(item))
            skipped.append(
                {
                    "来源行号": item.get("来源行号", ""),
                    "客户代码": item.get("客户代码", ""),
                    "客户简称": item.get("客户简称", ""),
                    "来源字段": item.get("来源字段", ""),
                    "规则文本": item.get("规则文本", ""),
                    "跳过原因": "同客户同条件同字段存在不同覆盖值",
                }
            )
    if not conflict_ids:
        return rows, 0
    return [rule for rule in rows if id(rule) not in conflict_ids], len(conflict_ids)


def _dedupe_confirmed_rules(rows: list[dict]) -> list[dict]:
    seen: set[tuple[str, ...]] = set()
    deduped: list[dict] = []
    for rule in rows:
        key = (
            _clean(rule.get("客户代码")),
            _clean(rule.get("客户简称")),
            _clean(rule.get("来源字段")),
            _clean(rule.get("覆盖字段")),
            _clean(rule.get("覆盖值")),
            _clean(rule.get("条件文本")),
            _clean(rule.get("条件胶系")),
            _clean(rule.get("条件厚度")),
            _clean(rule.get("条件铜厚")),
            _clean(rule.get("条件尺寸")),
            _clean(rule.get("条件关键词")),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(rule)
    return deduped


def load_transcode_agent_rules(
    version: str | None = None,
    *,
    include_maintenance: bool = True,
) -> list[dict]:
    path = get_transcode_agent_rule_file_path(version)
    if not path.exists():
        return []
    workbook = openpyxl.load_workbook(path, data_only=True)
    if "机器规则" not in workbook.sheetnames:
        return []
    worksheet = workbook["机器规则"]
    headers = [_clean(worksheet.cell(1, col).value) for col in range(1, worksheet.max_column + 1)]
    rules = []
    for row_idx in range(2, worksheet.max_row + 1):
        item = {headers[col - 1]: _clean(worksheet.cell(row_idx, col).value) for col in range(1, worksheet.max_column + 1) if headers[col - 1]}
        if item.get("规则ID"):
            rules.append(item)
    if not include_maintenance:
        return rules
    from .transcode_customer_rule_admin import merge_agent_rule_overrides
    from .transcode_rule_center import merge_base_rule_overrides

    return merge_base_rule_overrides(merge_agent_rule_overrides(rules))


def load_transcode_agent_mapping_tables(version: str | None = None) -> dict[str, list[dict]]:
    path = get_transcode_agent_mapping_table_file_path(version)
    tables: dict[str, list[dict]] = {sheet: [] for sheet in MAPPING_TABLE_HEADERS}
    if not path.exists():
        return tables
    workbook = openpyxl.load_workbook(path, data_only=True)
    for sheet_name, rows in tables.items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        headers = [_clean(worksheet.cell(1, col).value) for col in range(1, worksheet.max_column + 1)]
        for row_idx in range(2, worksheet.max_row + 1):
            item = {
                headers[col - 1]: _clean(worksheet.cell(row_idx, col).value)
                for col in range(1, worksheet.max_column + 1)
                if headers[col - 1]
            }
            if item.get("映射ID"):
                rows.append(item)
    from .transcode_rule_center import merge_agent_mapping_overrides

    return merge_agent_mapping_overrides(tables)


def get_transcode_agent_rule_count() -> int:
    return len(load_transcode_agent_rules())


def export_transcode_agent_rules(export_type: str) -> Path:
    rule_path = get_transcode_agent_rule_file_path()
    original_path = get_transcode_agent_original_file_path()
    if export_type == "machine":
        if not rule_path.exists():
            raise FileNotFoundError("还没有上传并转换过营销转码Agent规则")
        return rule_path
    if export_type == "original":
        if not original_path.exists():
            raise FileNotFoundError("还没有上传客户特殊清单原文件")
        return original_path
    if export_type == "full":
        if not rule_path.exists():
            raise FileNotFoundError("还没有上传并转换过营销转码Agent规则")
        export_path = TRANSCODE_AGENT_RULES_DIR / "transcode_agent_rule_package.xlsx"
        shutil.copy2(rule_path, export_path)
        return export_path
    raise ValueError("未知导出类型")


def _parse_structured_cell(customer_code: str, customer_name: str, material_type: str, source_col: str, text: str, source_row: int) -> list[dict]:
    if not text:
        return []
    rules = []
    for line in text.splitlines():
        line = _clean(line)
        if not line or "：" not in line:
            continue
        field_name, raw_value = line.split("：", 1)
        field_name = _clean(field_name)
        raw_value = raw_value.strip("；; \t")
        if not raw_value:
            continue
        override_field = FIELD_TO_OVERRIDE.get(field_name, "")
        for clause in _split_clauses(raw_value):
            if not clause:
                continue
            condition_text, override_value = _split_condition_value(clause)
            normalized_value = _normalize_override_value(override_field, override_value)
            pending = "否" if override_field in EXECUTABLE_FIELDS and normalized_value else "是"
            if _looks_uncertain(clause):
                pending = "是"
            rules.append(
                _make_rule(
                    customer_code,
                    customer_name,
                    material_type,
                    source_col,
                    field_name,
                    clause,
                    source_row,
                    override_field,
                    normalized_value,
                    condition_text=condition_text,
                    pending=pending,
                )
            )
    return rules


def _parse_free_text_rule(customer_code: str, customer_name: str, material_type: str, source_col: str, text: str, source_row: int) -> list[dict]:
    rules = []
    for clause in _split_clauses(text):
        condition_text, override_value = _split_condition_value(clause)
        override_field = _guess_free_text_override_field(clause)
        normalized_value = _normalize_override_value(override_field, override_value)
        rules.append(
            _make_rule(
                customer_code,
                customer_name,
                material_type,
                source_col,
                "通用",
                clause,
                source_row,
                override_field,
                normalized_value,
                condition_text=condition_text,
                pending="否" if override_field in EXECUTABLE_FIELDS and normalized_value else "是",
            )
        )
    return rules


def _make_rule(
    customer_code: str,
    customer_name: str,
    material_type: str,
    source_col: str,
    original_field: str,
    rule_text: str,
    source_row: int,
    override_field: str,
    override_value: str,
    *,
    condition_text: str = "",
    pending: str = "否",
) -> dict:
    condition_glue = _extract_condition_glue(condition_text or rule_text)
    condition_keywords = _extract_condition_keywords(condition_text or rule_text)
    condition_copper = _extract_condition_copper(condition_text or rule_text)
    condition_thickness = _extract_condition_thickness(condition_text or rule_text)
    condition_size = _extract_condition_size(condition_text or rule_text)
    priority = 100
    if condition_keywords:
        priority += 30
    if condition_copper:
        priority += 20
    if condition_thickness or condition_size:
        priority += 10
    if condition_glue:
        priority += 10
    return {
        "规则ID": "",
        "启用": "是",
        "客户代码": customer_code,
        "客户简称": customer_name,
        "物料类别": material_type,
        "来源字段": source_col,
        "原始字段": original_field,
        "规则文本": rule_text,
        "条件文本": condition_text,
        "条件胶系": condition_glue,
        "条件关键词": condition_keywords,
        "条件铜厚": condition_copper,
        "条件厚度": condition_thickness,
        "条件尺寸": condition_size,
        "覆盖字段": override_field,
        "覆盖值": override_value,
        "命中来源": "客户特殊清单结构化母表",
        "优先级": str(priority),
        "强制执行": "是" if override_field in EXECUTABLE_FIELDS and override_value and pending != "是" else "否",
        "待确认": pending,
        "来源行号": str(source_row),
        "规则解释": "",
        "跳过原因": "",
    }


def _split_clauses(value: str) -> list[str]:
    normalized = _clean(value).replace("；", ";")
    pieces = re.split(r";+|(?=当[^;；]*?=)|(?=如果[^;；]*?=)|[,，](?=[^,，;；]{0,30}=)", normalized)
    return [piece.strip(" ,，;；") for piece in pieces if piece.strip(" ,，;；")]


def _split_condition_value(clause: str) -> tuple[str, str]:
    text = _clean(clause)
    if "=" not in text:
        return text, ""
    left, right = text.rsplit("=", 1)
    return left.strip(" ,，"), right.strip(" ,，;；")


def _normalize_override_value(override_field: str, value: str) -> str:
    raw = _clean(value).upper()
    raw = raw.replace("（", "(").replace("）", ")")
    if not raw:
        return ""
    if "=" in raw:
        tail = raw.rsplit("=", 1)[1].strip()
        if tail:
            raw = tail
    if override_field == "copper_code":
        compact = raw.replace(" ", "")
        if "1.5/1.5" in compact or "F/F" in compact:
            return "FF"
        if "R/R" in compact:
            return "FF"
        if "J/J" in compact:
            return "JJ"
        if "H/H" in compact:
            return "JJ"
        if "K/K" in compact:
            return "KK"
        embedded_code = re.search(r"(?:^|[^A-Z0-9])(FF|JJ|KK|HH|TT|11|22|33|44|55|66|AA|BB|GG)(?:$|[^A-Z0-9])", raw)
        if embedded_code:
            return embedded_code.group(1)
        if re.fullmatch(r"[A-Z0-9]{2}", compact):
            return compact
    if override_field == "copper_type_code":
        code_match = re.match(r"^([A-Z0-9])(?:\(|（)", raw)
        if code_match:
            return code_match.group(1)
        for keyword, code in COPPER_TYPE_VALUE_MAP.items():
            if keyword in raw:
                return code
        if re.fullmatch(r"[A-Z0-9]", raw):
            return raw
    if override_field == "tc_code":
        if "芯" in raw or raw == "C":
            return "C"
        if "总" in raw or raw == "T":
            return "T"
    if override_field == "struct_code":
        if re.fullmatch(r"[A-Z0-9*]", raw):
            return raw
    if override_field == "grade_code":
        code_match = re.search(r"(?:^|[^A-Z0-9])(A1|A2|AC|AD|AH|AL|AP|AM|AT|F1)(?:$|[^A-Z0-9])", raw)
        if code_match:
            return code_match.group(1)
        if "汽车" in raw:
            return "AC"
        if re.fullmatch(r"[A-Z0-9]{2,4}", raw):
            return raw
    if override_field == "glue_category_code":
        if "普通" in raw or raw == "Y":
            return "Y"
        if "特殊" in raw or raw == "R":
            return "R"
    if override_field == "size_code":
        digits = re.sub(r"\D", "", raw)
        if 1 <= len(digits) <= 8:
            return digits.zfill(8)
    if override_field == "glue_code":
        compact = raw.replace(" ", "")
        if compact in GLUE_MODEL_TO_CODE:
            return GLUE_MODEL_TO_CODE[compact]
        model_match = re.fullmatch(r"NY[-]?(\d{4})", compact)
        if model_match and model_match.group(1) in GLUE_MODEL_TO_CODE:
            return GLUE_MODEL_TO_CODE[model_match.group(1)]
        if re.fullmatch(r"[A-Z0-9]{2,4}", compact):
            return compact
    return raw if override_field not in EXECUTABLE_FIELDS else ""


def _extract_condition_glue(text: str) -> str:
    candidates = re.findall(r"NY[-]?[A-Z0-9]+(?:HF|HC|M2|M|H|L)?", _clean(text).upper())
    return "/".join(dict.fromkeys(candidates))


def _extract_condition_copper(text: str) -> str:
    compact = _clean(text).upper().replace(" ", "")
    values = []
    for token in ("R/R", "F/F", "J/J", "K/K", "H/H", "W/W", "1/1", "1.5/1.5", "2/2", "0.5/0.5"):
        if token in compact:
            values.append(token)
    return "/".join(values)


def _extract_condition_thickness(text: str) -> str:
    source = _clean(text)
    matches = re.findall(r"(?:厚度|板厚)\s*[:：=]?\s*([<>]=?|≥|≤)?\s*(\d+(?:\.\d+)?)\s*(mil|MM|mm)?", source, flags=re.IGNORECASE)
    values = []
    for operator, number, unit in matches:
        normalized_number = _normalize_condition_thickness_number(number, unit)
        if normalized_number:
            values.append(f"{operator or ''}{normalized_number}{unit or ''}")
    return "/".join(dict.fromkeys(values))


def _normalize_condition_thickness_number(number: str, unit: str) -> str:
    if unit:
        return number
    if "." in number:
        return number
    if len(number) == 1 and number in {"8", "9"}:
        return f"0.{number}"
    if len(number) == 2 and number.isdigit() and int(number) >= 10:
        return f"0.{number}"
    return number


def _extract_condition_size(text: str) -> str:
    source = _clean(text).replace("×", "*").lower()
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*\*\s*(\d+(?:\.\d+)?)", source)
    return "/".join(dict.fromkeys(f"{width}*{height}" for width, height in matches))


def _extract_condition_keywords(text: str) -> str:
    source = _clean(text)
    explicit = re.findall(r"关键词\s*[:：]\s*([^；;|]+)", source, flags=re.IGNORECASE)
    if explicit:
        keywords = []
        for item in explicit:
            keyword = item.strip(" ，,")
            if keyword in {"当胶系", "胶系", "基板级别", "客户为"}:
                continue
            if keyword:
                keywords.append(keyword)
        return "/".join(keywords)
    match = re.search(r"(?:备注|订单|客户规格|物料描述)[^有]{0,8}有(.+?)(?:字样|时|$)", source)
    if match:
        return match.group(1).replace("或", "/").replace("、", "/").strip(" ，,")
    if "汽车板" in source:
        return "汽车板"
    if "MINILED" in source.upper():
        return "MINILED"
    return ""


def _guess_free_text_override_field(text: str) -> str:
    upper = _clean(text).upper()
    if "汽车板" in upper and ("AC" in upper or "料号" in upper):
        return "grade_code"
    if "HVLP" in upper or "RTF" in upper or "VLP" in upper:
        return "copper_type_code"
    if "R/R" in upper or "F/F" in upper or "J/J" in upper or "K/K" in upper:
        return "copper_code"
    return ""


def _looks_uncertain(text: str) -> bool:
    return bool(re.search(r"待确认|需确认|参考|详见|注意|特殊留意|可能|或", _clean(text)))


def _clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip()


def _clean_multiline(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\r\n", "\n").replace("\r", "\n")
    if text.strip().lower() == "nan":
        return ""
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def _is_template_row(customer_code: str, customer_name: str) -> bool:
    return not customer_code and not customer_name
