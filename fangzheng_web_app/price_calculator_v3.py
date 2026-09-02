# -*- coding: utf-8 -*-
"""
方正价格自动计算工具 v3.0
==========================
更新内容（v3.0）：
  1. 内置方正价格表和基板对照表（data_price.pkl / data_account.pkl）
     用户只需上传只含「价格计算」Sheet 的精简 Excel 即可
     如果上传的 Excel 中包含「方正价格」Sheet，则优先使用 Excel 中的数据
  2. 非标准宽度铜箔价格修正：宽度不等于49时，价格 × (实际宽度/48) × 1.07
  3. 双面铜箔类型判断：RTF2/RTF 这类双代码，取铜厚较厚一面对应的代码
     铜厚大小：H(0.5) < 1 < 2

使用方法：
  1. 将只含「价格计算」Sheet 的 Excel 文件放到 input/ 文件夹
  2. 运行：python price_calculator_v3.py
  3. 结果保存在 output/ 文件夹
"""
import pandas as pd
import numpy as np
import re
import os
import shutil
import math
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# ★ 配置区
# ============================================================
# 标准尺寸列表（原始长*宽，均为整数）-> 对应方正价格列名
STANDARD_SIZES = {
    (37, 49): '36"*48"',
    (41, 49): '40"*48"',
    (43, 49): '42"*48"',
}
# 减1后尺寸 -> 价格列名（用于非标准尺寸大板推导）
SIZE_COL_MAP = {
    (36, 48): '36"*48"',
    (40, 48): '40"*48"',
    (42, 48): '42"*48"',
}
# PP 计算公式固定参数
PP_FIXED_WIDTH = 48
PP_FIXED_DIV = 144
PP_SMALL_PIECE_TAIL_THRESHOLD = 16.5
PP_SMALL_PIECE_DIVISOR_LE_THRESHOLD = 3
PP_SMALL_PIECE_DIVISOR_GT_THRESHOLD = 2
ROLL_FIXED_WIDTH = 48
ROLL_MM_DIVISOR = 0.0254
FINAL_PRICE_DECIMALS = 2
PP_ROLL_PRICE_COLUMNS = ('36"*48"', '40"*48"', '42"*48"')
PP_ROLL_LENGTH_PATTERN = r'\d+(?:\.\d+)?\s*(?:M|\u7c73)\s*/\s*(?:ROLL|\u5377)'
# 允许近似取价的最大板厚差异。例：0.140mm 只允许 0.130~0.150mm。
MAX_THICKNESS_DELTA = 0.01
# mil 转 mm 系数
MIL_TO_MM = 0.0254
# 标准宽度（不需要修正系数）
STANDARD_WIDTH = 49
# 非标准宽度修正系数
NON_STD_CORRECTION = 1.07
# 铜厚大小映射（H=0.5oz）
CU_THICK_ORDER = {'H': 0.5, '1': 1.0, '2': 2.0}
DEFAULT_FOIL_TYPE = 'HTE'
TAIL_BOARD_H_VALUES = {43}
TAIL_BOARD_WIDTH_MAP = {
    37: ('36"*48"', 1),
    41: ('40"*48"', 1),
    43: ('42"*48"', 1),
    74: ('36"*48"', 2),
    82: ('40"*48"', 2),
    86: ('42"*48"', 2),
}
CALCULATION_SHEET_NAME = '价格计算'
RULE_SHEET_NAMES = {'方正价格', '基板对账', '基板对账表', '基板对照', '基板对照表', '计算说明', '规则说明'}
DESCRIPTION_HEADER_KEYWORDS = ('物料描述', '物料', '品名', '规格', '描述', '产品描述')
DESCRIPTION_HEADER_EXCLUDE_KEYWORDS = ('编码', '代码', '料号', '编号', 'PO', '数量', '单位', '交货期', '订单', '供应商')
DESCRIPTION_CONTENT_RE = re.compile(
    r'(?:\bPP\b|RC\s*\d+\s*%|\d+\.?\d*\s*mm|\d+\.?\d*"?\s*[*xX×]\s*\d+\.?\d*"?|M/Roll|HVLP|RTF|HTE|\d+\s*/\s*\d+|H\s*/\s*H)',
    re.IGNORECASE,
)

# 文件路径配置
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, 'input')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
DATA_PRICE_PKL = os.path.join(BASE_DIR, 'data_price.pkl')
DATA_ACCOUNT_PKL = os.path.join(BASE_DIR, 'data_account.pkl')

# ============================================================
# 内置数据加载（全局缓存，只加载一次）
# ============================================================
_BUILTIN_PRICE = None
_BUILTIN_ACCOUNT = None

def load_builtin_data():
    """加载内置数据（pkl文件），全局缓存"""
    global _BUILTIN_PRICE, _BUILTIN_ACCOUNT
    if _BUILTIN_PRICE is None:
        if not os.path.exists(DATA_PRICE_PKL):
            raise FileNotFoundError(
                f"内置价格数据文件不存在：{DATA_PRICE_PKL}\n"
                "请确保 data_price.pkl 和 data_account.pkl 与程序在同一文件夹。"
            )
        _BUILTIN_PRICE = pd.read_pickle(DATA_PRICE_PKL)
        _BUILTIN_ACCOUNT = pd.read_pickle(DATA_ACCOUNT_PKL)
        log(f"内置数据已加载：方正价格 {len(_BUILTIN_PRICE)} 行，基板对账 {len(_BUILTIN_ACCOUNT)} 行")
    return _BUILTIN_PRICE, _BUILTIN_ACCOUNT

# ============================================================
# 工具函数
# ============================================================
def log(msg, level='INFO'):
    timestamp = datetime.now().strftime('%H:%M:%S')
    print(f"[{timestamp}] [{level}] {msg}")

def ensure_dirs():
    for d in [INPUT_DIR, OUTPUT_DIR]:
        os.makedirs(d, exist_ok=True)

def round_price(value):
    quant = Decimal("1").scaleb(-FINAL_PRICE_DECIMALS)
    return float(Decimal(str(value)).quantize(quant, rounding=ROUND_HALF_UP))

def format_price(value):
    return f"{round_price(value):.{FINAL_PRICE_DECIMALS}f}"

def pp_piece_divisor(h):
    """小片 PP 按纬向尺寸判断开数：纬向<=16.5 开3，否则开2。"""
    return (
        PP_SMALL_PIECE_DIVISOR_LE_THRESHOLD
        if h <= PP_SMALL_PIECE_TAIL_THRESHOLD
        else PP_SMALL_PIECE_DIVISOR_GT_THRESHOLD
    )

def normalize_str(s):
    if pd.isna(s):
        return ''
    s = str(s).strip()
    s = s.replace('\xa0', ' ')   # 不间断空格 → 普通空格
    s = s.replace('\u3000', ' ')  # 全角空格 → 普通空格
    s = s.replace('（', '(').replace('）', ')')
    # 将多个连续空格合并为单个
    import re as _re
    s = _re.sub(r' +', ' ', s)
    return s

# ============================================================
# 解析函数
# ============================================================
def parse_thickness(text):
    """提取厚度（mm），支持 mil 和 mm"""
    mil_match = re.search(r'(\d+\.?\d*)\s*mil', text, re.IGNORECASE)
    if mil_match:
        return round(float(mil_match.group(1)) * MIL_TO_MM, 4)
    mm_match = re.search(r'(\d+\.?\d*)\s*mm', text, re.IGNORECASE)
    if mm_match:
        return float(mm_match.group(1))
    return None

def format_thickness(t):
    """将厚度浮点数格式化为方正价格表中的字符串格式"""
    s = f"{t:.4f}".rstrip('0').rstrip('.')
    return s

# 铜厚降级映射：查询时先用原始铜厚，找不到再用降级后的标准铜厚
CU_THICK_FALLBACK = {
    'H/1': '1/1',
    'H/2': '2/2',
    '1/2': '2/2',
}

def parse_copper_thickness(text):
    """提取铜厚，格式：H/H, 1/1, 2/2, H/1, 1/2 等
    返回原始铜厚字符串，查询时再按需降级
    """
    match = re.search(r'(\d+\.?\d*|H|T)\s*/\s*(\d+\.?\d*|H|T)', text, re.IGNORECASE)
    if match:
        return f"{match.group(1).upper()}/{match.group(2).upper()}"
    return None

def cu_thick_value(cu_str):
    """将铜厚字符串转换为数值，用于比较大小"""
    cu_str = str(cu_str).strip().upper()
    return CU_THICK_ORDER.get(cu_str, float(cu_str) if cu_str.replace('.','').isdigit() else 0)

# 铜箔类型别名映射（特殊型号 -> 标准代码）
FOIL_ALIAS = {
    'HS2-M2-VS': 'HVLP2',
    'HS2-M2-VSP': 'HVLP2',
}

# 方正订单中部分 PP 胶系会保留产品型号中的 0；报价表使用的是简写型号。
PP_GLUE_ALIASES = {
    'NY3170HFP': ('NY317HFP',),
}

def parse_foil_type(text, cu_thick=None):
    """
    提取铜箔类型，支持双代码（RTF2/RTF）
    策略：先尝试整体匹配（如 RTF2/RTF 作为一个完整代码），
    找不到时再按铜厚拆分取较厚一面。
    铜厚大小：H(0.5) < 1 < 2
    """
    # 先尝试匹配双代码格式，如 RTF2/RTF 或 (RTF2/RTF)
    dual_match = re.search(
        r'\(?(HVLP\d*|RTF\d*|HTE|HS[\w\-]+|RG\d+)\s*/\s*(HVLP\d*|RTF\d*|HTE|HS[\w\-]+|RG\d+)\)?',
        text, re.IGNORECASE
    )
    if dual_match:
        foil_front = dual_match.group(1).upper()
        foil_back = dual_match.group(2).upper()
        # 先返回整体代码（如 RTF2/RTF），由查询函数决定是否需要拆分
        combined = f"{foil_front}/{foil_back}"
        log(f"  检测到双代码铜箔：{combined}")
        return combined
    
    # 单代码匹配（括号内或独立出现）
    match = re.search(
        r'\((HVLP\d*|RTF\d*|HTE|HS[\w\-]+|RG\d+|HVLP[\w]+|RTF[\w]+)\)',
        text, re.IGNORECASE
    )
    if match:
        raw = match.group(1).upper()
        return FOIL_ALIAS.get(raw, raw)
    
    # 无括号单代码（支持裸RTF、HVLP等不带数字的型号）
    match2 = re.search(
        r'\b(HVLP\d*|RTF\d*|HTE|HS2-M2-VSP?|RG\d+)\b',
        text, re.IGNORECASE
    )
    if match2:
        raw = match2.group(1).upper()
        return FOIL_ALIAS.get(raw, raw)
    
    return None

def parse_laminate(text):
    """提取叠构，转换为方正价格表格式：1078x2"""
    # 优先匹配括号格式：(1078*2) 或 (2*1078)
    match = re.search(r'\((\d+)\*(\d+)\)', text)
    if match:
        a, b = int(match.group(1)), int(match.group(2))
        if a > b:
            return f"{a}x{b}"
        else:
            return f"{b}x{a}"
    # 兼容无括号格式：1x1078 或 2x1078
    match2 = re.search(r'\b(\d+)[xX\*](\d+)\b', text)
    if match2:
        a, b = int(match2.group(1)), int(match2.group(2))
        # 排除尺寸（两个数都较大）和厚度
        if min(a, b) <= 10 and max(a, b) >= 100:
            if a > b:
                return f"{a}x{b}"
            else:
                return f"{b}x{a}"
    return None

def parse_size(text):
    """
    提取尺寸（长*宽），返回 (float, float) 或 None
    支持：37"*49", 28.6"*43, 37*49 等格式
    """
    if re.search(r'\d+M/Roll', text, re.IGNORECASE):
        return None
    
    patterns = [
        r'(\d+\.?\d*)"?\s*[*×xX]\s*(\d+\.?\d*)"',
        r'(\d+\.?\d*)"\s*(\d+\.?\d*)"',
        r'(\d+\.?\d*)\'?"\s*[*×xX]\s*(\d+\.?\d*)',
        r'(\d+\.?\d*)\s*[*×xX]\s*(\d+\.?\d*)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            w = float(match.group(1))
            h = float(match.group(2))
            if w > 100 or h > 100:
                continue
            return (w, h)
    return None

def parse_rc_percent(text):
    """提取 PP 的 RC% 数值"""
    match = re.search(r'RC\s*(\d+)\s*%', text, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None

def size_to_code(w, h):
    """将尺寸转换为8位代码：28.6*49 -> 28604900"""
    def to_4digit(v):
        s = str(v).replace('.', '')
        return s.ljust(4, '0')[:4]
    return to_4digit(w) + to_4digit(h)

def is_standard_size(w, h):
    return (round(w), round(h)) in STANDARD_SIZES

def get_standard_size_col(w, h):
    return STANDARD_SIZES.get((round(w), round(h)))

def select_calculation_sheet_name(sheet_names):
    """Pick the sheet that contains uploaded material rows."""
    names = list(sheet_names)
    if CALCULATION_SHEET_NAME in names:
        return CALCULATION_SHEET_NAME
    for name in names:
        if str(name).strip() not in RULE_SHEET_NAMES:
            return name
    if names:
        return names[0]
    raise ValueError("Excel 文件中没有可读取的 Sheet")

def _looks_like_description(value):
    if value is None or pd.isna(value):
        return False
    text = str(value).strip()
    if not text or text.lower() in {'nan', 'none'}:
        return False
    return DESCRIPTION_CONTENT_RE.search(text) is not None

def detect_description_column_dataframe(df_calc):
    """Return zero-based material description column index."""
    for idx, col in enumerate(df_calc.columns):
        name = str(col).strip()
        if (
            any(keyword in name for keyword in DESCRIPTION_HEADER_KEYWORDS)
            and not any(keyword in name for keyword in DESCRIPTION_HEADER_EXCLUDE_KEYWORDS)
        ):
            return idx

    best_idx = 3 if len(df_calc.columns) > 3 else 0
    best_score = -1
    for idx in range(len(df_calc.columns)):
        values = df_calc.iloc[:, idx].head(80)
        score = sum(1 for value in values if _looks_like_description(value))
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx

def detect_description_column_openpyxl(ws):
    """Return one-based material description column index for openpyxl sheets."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for idx, value in enumerate(header_row, start=1):
        name = str(value or '').strip()
        if (
            any(keyword in name for keyword in DESCRIPTION_HEADER_KEYWORDS)
            and not any(keyword in name for keyword in DESCRIPTION_HEADER_EXCLUDE_KEYWORDS)
        ):
            return idx

    max_col = ws.max_column or 1
    best_idx = 4 if max_col >= 4 else 1
    best_score = -1
    max_row = min(ws.max_row or 1, 80)
    for col_idx in range(1, max_col + 1):
        score = 0
        for row_idx in range(2, max_row + 1):
            if _looks_like_description(ws.cell(row=row_idx, column=col_idx).value):
                score += 1
        if score > best_score:
            best_score = score
            best_idx = col_idx
    return best_idx

# ============================================================
# 非标准宽度修正（v3新增）
# ============================================================
def apply_width_correction(price, w, h):
    """
    铜箔计算中，宽度不等于49时，价格需要修正：
    最终价格 × (实际宽度 / 48) × 1.07
    注意：这里的宽度是原始尺寸（未-1的），但修正用原始宽度
    """
    # 取宽度（h方向，即第二个数）
    actual_h = round(h)
    if actual_h != STANDARD_WIDTH:
        corrected = price * (actual_h / 48) * NON_STD_CORRECTION
        log(f"  非标准宽度 {actual_h}（≠49），价格修正：{price} × ({actual_h}/48) × {NON_STD_CORRECTION} = {round(corrected, 4)}")
        return round(corrected, 4), True
    return price, False

# ============================================================
# PP 胶系名称转换
# ============================================================
def _try_convert_pp_glue(raw_glue):
    """尝试转换PP胶系名称"""
    if re.search(r'\d+P\(C\)$', raw_glue):
        return re.sub(r'P\(C\)$', '(C)', raw_glue)
    if not raw_glue.endswith('P') and not raw_glue.endswith(')'):
        return raw_glue + 'P'
    return raw_glue

def convert_pp_glue(raw_glue, pp_rows):
    """PP 胶系名称转换（带表查询验证）"""
    if (pp_rows['型号'].astype(str).str.strip() == raw_glue).any():
        return raw_glue
    converted = _try_convert_pp_glue(raw_glue)
    if (pp_rows['型号'].astype(str).str.strip() == converted).any():
        return converted
    return raw_glue

# ============================================================
# 数据加载（支持精简Excel + 内置数据）
# ============================================================
def load_excel(filepath):
    """
    加载 Excel 文件
    - 优先使用「价格计算」Sheet；如果没有，则自动使用第一个业务 Sheet
    - 如果包含「方正价格」Sheet，优先使用；否则使用内置数据
    - 如果包含「基板对账」Sheet，优先使用；否则使用内置数据
    """
    log(f"正在加载文件：{os.path.basename(filepath)}")
    
    xl = pd.ExcelFile(filepath)
    sheet_names = xl.sheet_names
    log(f"  文件包含 Sheet：{sheet_names}")
    
    # 加载价格计算表：客户文件常见为 Sheet1，因此不强制要求固定 Sheet 名。
    calc_sheet_name = select_calculation_sheet_name(sheet_names)
    df_calc = pd.read_excel(filepath, sheet_name=calc_sheet_name, header=0)
    log(f"  计算Sheet：{calc_sheet_name}，{len(df_calc)} 行")
    
    # 加载方正价格表（优先Excel，否则内置）
    if '方正价格' in sheet_names:
        df_price = pd.read_excel(filepath, sheet_name='方正价格', header=17)
        df_price.columns = [str(c).strip() for c in df_price.columns]
        log(f"  方正价格（来自Excel）：{len(df_price)} 行")
    else:
        builtin_price, _ = load_builtin_data()
        df_price = builtin_price
        log(f"  方正价格（内置数据）：{len(df_price)} 行")
    
    # 加载基板对照表（优先Excel，否则内置）
    if '基板对照' in sheet_names:
        df_account = pd.read_excel(filepath, sheet_name='基板对照', header=0)
        log(f"  基板对照（来自Excel）：{len(df_account)} 行")
    elif '基板对照表' in sheet_names:
        df_account = pd.read_excel(filepath, sheet_name='基板对照表', header=0)
        log(f"  基板对照（来自Excel）：{len(df_account)} 行")
    elif '基板对账' in sheet_names:
        df_account = pd.read_excel(filepath, sheet_name='基板对账', header=0)
        df_account.columns = [str(c).strip() for c in df_account.columns]
        log(f"  基板对账（来自Excel）：{len(df_account)} 行")
    else:
        _, builtin_account = load_builtin_data()
        df_account = builtin_account
        log(f"  基板对账（内置数据）：{len(df_account)} 行")
    
    return df_calc, df_price, df_account

# ============================================================
# 方正价格表查询
# ============================================================
def query_ccl_price(df_price, glue, thickness_str, cu_thick, foil_type, laminate, size_col, cu_thick_raw=None):
    """
    查询 CCL 铜箔价格
    支持：
    1. 铜箔双代码（RTF2/RTF）整体匹配，失败则按铜厚拆分
    2. 铜厚降级：H/1→先查H/1，找不到再查降级后的标准铜厚
    3. 叠构包含匹配：如 3313 匹配价格表中的 2313/3313
    """
    ccl_rows = df_price[df_price['CCL'].astype(str).str.strip() == 'CCL']
    
    def _query(foil, cu):
        mask = (ccl_rows['型号'].astype(str).str.strip() == glue) & \
               (ccl_rows['不含铜板厚/（mm)'].astype(str).str.strip() == thickness_str) & \
               (ccl_rows['铜厚'].astype(str).str.strip() == cu) & \
               (ccl_rows['铜箔'].astype(str).str.strip() == foil) & \
               (ccl_rows['叠构'].astype(str).str.strip() == laminate)
        result = ccl_rows[mask]
        # 叠构包含匹配：如果精确匹配失败，尝试匹配包含该叠构的行（如 3313 匹配 2313/3313）
        if len(result) == 0:
            mask2 = (ccl_rows['型号'].astype(str).str.strip() == glue) & \
                    (ccl_rows['不含铜板厚/（mm)'].astype(str).str.strip() == thickness_str) & \
                    (ccl_rows['铜厚'].astype(str).str.strip() == cu) & \
                    (ccl_rows['铜箔'].astype(str).str.strip() == foil) & \
                    (ccl_rows['叠构'].astype(str).str.contains(laminate, na=False))
            result = ccl_rows[mask2]
        return result
    
    def _query_with_foil(foil):
        """先用原始铜厚查，找不到再用降级铜厚查"""
        # 第一步：用原始铜厚查
        result = _query(foil, cu_thick)
        cu_used = cu_thick
        # 第二步：铜厚降级
        if len(result) == 0 and cu_thick in CU_THICK_FALLBACK:
            fallback_cu = CU_THICK_FALLBACK[cu_thick]
            log(f"  铜厚 {cu_thick} 未匹配，降级为 {fallback_cu}")
            result = _query(foil, fallback_cu)
            cu_used = fallback_cu
        return result, cu_used
    
    # 第一步：整体匹配铜箔（如 RTF2/RTF 作为完整代码）
    result, cu_used = _query_with_foil(foil_type)
    foil_used = foil_type
    
    # 第二步：若整体匹配失败且是双代码，按铜厚拆分取较厚一面
    if len(result) == 0 and '/' in foil_type:
        parts = foil_type.split('/')
        if len(parts) == 2:
            foil_front, foil_back = parts[0], parts[1]
            if cu_thick_raw:
                cu_parts = cu_thick_raw.split('/')
                if len(cu_parts) == 2:
                    front_val = cu_thick_value(cu_parts[0])
                    back_val = cu_thick_value(cu_parts[1])
                    chosen = foil_back if back_val >= front_val else foil_front
                else:
                    chosen = foil_back
            else:
                chosen = foil_back
            log(f"  整体匹配失败，按铜厚拆分取 {chosen}")
            result, cu_used = _query_with_foil(chosen)
            foil_used = chosen
    
    if len(result) == 0:
        return None, None, f"未找到匹配：胶系={glue}, 厚度={thickness_str}, 铜厚={cu_thick}, 铜箔={foil_type}, 叠构={laminate}"
    
    if len(result) > 1:
        log(f"  [WARN] 找到 {len(result)} 行匹配，取第一行", 'WARN')
    
    row = result.iloc[0]
    price = row.get(size_col)
    
    if pd.isna(price):
        return None, result.index[0], f"价格列 {size_col} 为空"
    
    return float(price), result.index[0], None

def _pp_laminate_contains(series, laminate_str):
    """PP叠构按分隔 token 匹配，支持 3313 命中 2313/3313。"""
    pattern = rf'(^|[/,，、\s]){re.escape(laminate_str)}($|[/,，、\s])'
    return series.astype(str).str.strip().str.contains(pattern, na=False, regex=True)


def query_pp_price(df_price, glue, laminate_type, rc_percent):
    """查询 PP 价格（RMB/SF）"""
    pp_rows = df_price[df_price['CCL'].astype(str).str.strip() == 'PP']
    
    glue_candidates = [glue]
    glue_candidates.extend(PP_GLUE_ALIASES.get(glue.upper(), ()))
    converted = _try_convert_pp_glue(glue)
    if converted != glue:
        glue_candidates.append(converted)
    
    laminate_str = str(laminate_type)
    
    for glue_try in glue_candidates:
        model_mask = pp_rows['型号'].astype(str).str.strip() == glue_try
        exact_mask = model_mask & (pp_rows['不含铜板厚/（mm)'].astype(str).str.strip() == laminate_str)
        contains_mask = model_mask & _pp_laminate_contains(pp_rows['不含铜板厚/（mm)'], laminate_str)

        for candidates in (pp_rows[exact_mask], pp_rows[contains_mask]):
            if len(candidates) == 0:
                continue
            result = _match_cu_thick(candidates, rc_percent)
            if len(result) > 0:
                row = result.iloc[0]
                rmb_sf = row.get('RMB/SF')
                if not pd.isna(rmb_sf):
                    return float(rmb_sf), result.index[0], None
    
    glue_tried = '/'.join(set(glue_candidates))
    return None, None, f"PP未找到匹配：胶系={glue_tried}, 叠构={laminate_type}, RC%={rc_percent}"


def pp_roll_price_from_row(row):
    """读取 PP 报价行中的整卷价格。"""
    for col in PP_ROLL_PRICE_COLUMNS:
        value = row.get(col)
        if not pd.isna(value):
            return float(value)
    return None


def query_pp_roll_price(df_price, glue, laminate_type, rc_percent):
    """查询 PP 报价单整卷价格。"""
    _, row_idx, err = query_pp_price(df_price, glue, laminate_type, rc_percent)
    if err:
        return None
    row = df_price.loc[row_idx]
    return pp_roll_price_from_row(row)


def is_pp_roll_desc(desc):
    """Return True for PP roll specs such as 49.5"*300M/Roll or 49.5" ... 300M/卷."""
    return bool(re.search(PP_ROLL_LENGTH_PATTERN, normalize_str(desc), re.IGNORECASE))


def extract_pp_roll_width(desc):
    """Extract roll width in inch from supported PP roll specs."""
    desc = normalize_str(desc)
    joined_pattern = rf'(\d+\.?\d*)"?\s*[*脳xX×]\s*{PP_ROLL_LENGTH_PATTERN}'
    match = re.search(joined_pattern, desc, re.IGNORECASE)
    if match:
        return float(match.group(1))

    separated_pattern = rf'(\d+\.?\d*)\s*(?:"|IN|INCH|英寸)\s*.{{0,80}}?{PP_ROLL_LENGTH_PATTERN}'
    match = re.search(separated_pattern, desc, re.IGNORECASE)
    if match:
        return float(match.group(1))
    return None


def _parse_pp_price_key(desc):
    """从 PP 规格中提取报价表匹配键。非 PP 返回空。"""
    desc = normalize_str(desc)
    is_pp_prefix = bool(re.match(r'^PP\s+', desc, re.IGNORECASE))
    is_roll = is_pp_roll_desc(desc)
    is_pp_implicit = (
        not is_pp_prefix
        and bool(re.search(r'RC\s*\d+\s*%', desc, re.IGNORECASE))
        and not bool(re.search(r'\d+\.?\d*\s*mm', desc, re.IGNORECASE))
        and not is_roll
    )
    if not (is_pp_prefix or is_pp_implicit or is_roll):
        return None, None, None

    match = re.match(r'PP\s+([\w\-\(\)\.]+)', desc)
    if match:
        raw_glue = match.group(1).strip()
        lam_match = re.search(r'PP\s+[\w\-\(\)\.]+\s+(\d+)\s+RC', desc, re.IGNORECASE)
        if not lam_match and is_roll:
            lam_match = re.search(r'\b(1078|1080|1035|2116|2313|3313|106|1067)\b', desc)
        laminate_type = lam_match.group(1) if lam_match else None
    else:
        match2 = re.match(r'([\w\-\(\)\.]+)\s+(\d+)\s+RC', desc, re.IGNORECASE)
        if not match2:
            return None, None, None
        raw_glue = match2.group(1).strip()
        laminate_type = match2.group(2)

    rc_percent = parse_rc_percent(desc)
    if not raw_glue or not laminate_type or rc_percent is None:
        return None, None, None
    return raw_glue, laminate_type, rc_percent


def calculate_pp_roll_price(desc, df_price):
    """返回 PP 报价单中的整卷价格；非 PP 或未命中时返回 None。"""
    raw_glue, laminate_type, rc_percent = _parse_pp_price_key(desc)
    if raw_glue is None:
        return None
    return query_pp_roll_price(df_price, raw_glue, laminate_type, rc_percent)


def output_price_for_desc(desc, price, pp_roll_price):
    """方正结果只输出一列：PP卷料用整卷价，其它用原计算价。"""
    if is_pp_roll_desc(desc) and pp_roll_price not in (None, ''):
        return pp_roll_price
    return price

def _match_cu_thick(candidates, rc_value):
    """在候选行中匹配铜厚值"""
    numeric_cu = pd.to_numeric(candidates['铜厚'], errors='coerce')
    exact = candidates[numeric_cu == rc_value]
    if len(exact) > 0:
        return exact
    
    for idx, row in candidates.iterrows():
        cu_str = str(row['铜厚']).strip()
        if _in_cu_range(cu_str, rc_value):
            return candidates.loc[[idx]]
    
    return candidates.iloc[0:0]

def _in_cu_range(cu_str, rc_value):
    """判断 RC% 值是否在铜厚范围字符串内
    支持格式：≤45, >=50, ≥50, >49, ＞57（全角）, 46-49, <=45 等
    """
    # 将全角符号转为半角，统一处理
    cu_str = cu_str.replace('＞', '>').replace('＜', '<')
    # ≥ 或 >= 格式（大于等于）
    m = re.match(r'[≥≧]+\s*(\d+)|>=\s*(\d+)', cu_str)
    if m:
        threshold = int(m.group(1) or m.group(2))
        return rc_value >= threshold
    # ≤ 或 <= 格式（小于等于）
    m = re.match(r'[≤≦]+\s*(\d+)|<=\s*(\d+)', cu_str)
    if m:
        threshold = int(m.group(1) or m.group(2))
        return rc_value <= threshold
    # > 格式（严格大于）
    m = re.match(r'>\s*(\d+)', cu_str)
    if m:
        return rc_value > int(m.group(1))
    # < 格式（严格小于）
    m = re.match(r'<\s*(\d+)', cu_str)
    if m:
        return rc_value < int(m.group(1))
    # a-b 范围格式
    m = re.match(r'(\d+)\s*-\s*(\d+)', cu_str)
    if m:
        return int(m.group(1)) <= rc_value <= int(m.group(2))
    return False

# ============================================================
# 基板对照表查询（非标准尺寸）
# ============================================================
def query_nonstandard_size(df_account, w, h):
    """在基板对照表中查询非标准尺寸对应的大板信息"""
    code = size_to_code(w, h)
    log(f"  非标准尺寸代码：{code}（{w}x{h}）")
    
    mask = df_account['品名'].astype(str).str.contains(code, na=False)
    result = df_account[mask]
    
    if len(result) == 0:
        return None, None, None, f"基板对照表未找到尺寸代码：{code}（{w}x{h}）"
    
    row = result.iloc[0]
    qty = int(row['小片数量'])
    big_spec = str(row['大板规格'])
    
    size_match = re.search(r'(\d+\.?\d*)\s*x\s*(\d+\.?\d*)', big_spec)
    if not size_match:
        return None, None, None, f"无法从大板规格中提取尺寸：{big_spec}"
    
    big_w = float(size_match.group(1))
    big_h = float(size_match.group(2))
    
    log(f"  大板规格：{big_w}x{big_h}，小片数量：{qty}")
    return big_w, big_h, qty, None

def get_price_col_from_big(big_w, big_h):
    """从大板尺寸推导价格列与拼板倍率。"""
    std_h = round(big_h)
    
    # 第一步：检查大板本身是否接近标准尺寸（如 37.3x49.3 ≈ 37x49）
    for key, col in SIZE_COL_MAP.items():
        target_w, target_h = key[0] + 1, key[1] + 1
        if abs(big_w - target_w) <= 1.5 and abs(big_h - target_h) <= 1.5:
            log(f"  大板 {big_w}x{big_h} 接近标准尺寸 {target_w}x{target_h}，直接使用价格列 {col}")
            return target_w, target_h, col, 1
    
    # 第二步：尝试按 2 等分
    std_w = round(big_w / 2)
    price_key = (std_w - 1, std_h - 1)
    size_col = SIZE_COL_MAP.get(price_key)
    if size_col:
        return std_w, std_h, size_col, 2
    
    # 第三步：尝试按 3 等分
    std_w3 = round(big_w / 3)
    price_key3 = (std_w3 - 1, std_h - 1)
    size_col3 = SIZE_COL_MAP.get(price_key3)
    if size_col3:
        return std_w3, std_h, size_col3, 3
    
    # 第四步：宽松匹配（小数点误差）
    for divisor in [2, 3]:
        approx_w = big_w / divisor
        for key, col in SIZE_COL_MAP.items():
            target_w, target_h = key[0] + 1, key[1] + 1
            if abs(approx_w - target_w) <= 1.5 and abs(std_h - target_h) <= 1.5:
                return target_w, target_h, col, divisor
    
    return round(big_w / 2), std_h, None, None

# ============================================================
# 核心计算逻辑
# ============================================================
def calculate_price(desc, df_price, df_account):
    """根据物料描述计算价格，返回：(价格, 计算说明, 错误信息)"""
    desc = normalize_str(desc)
    
    if not desc or desc == 'nan':
        return None, '', '物料描述为空'
    
    is_pp_prefix = bool(re.match(r'^PP\s+', desc, re.IGNORECASE))
    is_roll = is_pp_roll_desc(desc)
    
    if is_roll:
        return _calc_roll(desc, df_price)
    
    # 判断是否为非 PP 开头的 PP 行：型号以P结尾且包含 RC%，且无mm厚度信息
    # 如：NY2170P 1080 RC68% 21.6"x24.6" 有卤 CAF
    is_pp_implicit = (
        not is_pp_prefix
        and bool(re.search(r'RC\s*\d+\s*%', desc, re.IGNORECASE))
        and not bool(re.search(r'\d+\.?\d*\s*mm', desc, re.IGNORECASE))
        and not is_roll
    )
    
    if is_pp_prefix or is_pp_implicit:
        return _calc_pp(desc, df_price)
    else:
        return _calc_ccl(desc, df_price, df_account)

def _calc_ccl(desc, df_price, df_account):
    """CCL 铜箔价格计算"""
    
    # 提取胶系
    glue_match = re.match(r'([\w\-\(\)\.]+)\s+\d', desc)
    if not glue_match:
        parts = desc.split()
        glue = parts[0] if parts else None
    else:
        glue = glue_match.group(1).strip()
    
    if not glue:
        return None, '', f'无法提取胶系：{desc}'
    
    # 提取厚度
    thickness = parse_thickness(desc)
    if thickness is None:
        return None, '', f'无法提取厚度：{desc}'
    thickness_str = format_thickness(thickness)
    thickness_raw = thickness  # 保存原始厚度值，用于后续近似匹配
    
    # 提取铜厚（先提取，供铜箔类型判断使用）
    cu_thick = parse_copper_thickness(desc)
    if not cu_thick:
        return None, '', f'无法提取铜厚：{desc}'
    
    # 提取铜箔类型（传入铜厚，支持双代码判断）
    foil_type = parse_foil_type(desc, cu_thick)
    foil_default_note = ""
    if not foil_type:
        foil_type = DEFAULT_FOIL_TYPE
        foil_default_note = "（描述未写铜箔，默认HTE）"
    
    # 提取叠构
    laminate = parse_laminate(desc)
    if not laminate:
        return None, '', f'无法提取叠构（如1078*2）：{desc}'
    
    # 提取尺寸
    size = parse_size(desc)
    if not size:
        return None, '', f'无法提取尺寸：{desc}'
    w, h = size
    
    log(f"  解析：胶系={glue}, 厚度={thickness_str}mm, 铜厚={cu_thick}, 铜箔={foil_type}, 叠构={laminate}, 尺寸={w}x{h}")
    
    def _try_query_with_nearest_thickness(size_col, thickness_str_in, thickness_raw_in):
        """先用精确厚度查，失败则取最近厚度重试"""
        price, row_idx, err = query_ccl_price(
            df_price, glue, thickness_str_in, cu_thick, foil_type, laminate, size_col, cu_thick_raw=cu_thick
        )
        if not err:
            return price, row_idx, err, thickness_str_in
        # 精确匹配失败，尝试取最近厚度
        ccl_rows = df_price[df_price['CCL'].astype(str).str.strip() == 'CCL']
        cands = ccl_rows[
            (ccl_rows['型号'].astype(str).str.strip() == glue) &
            (ccl_rows['铜厚'].astype(str).str.strip().isin([cu_thick, CU_THICK_FALLBACK.get(cu_thick, cu_thick)])) &
            (ccl_rows['铜箔'].astype(str).str.strip().isin([foil_type, foil_type.split('/')[0] if '/' in foil_type else foil_type,
                                                                foil_type.split('/')[-1] if '/' in foil_type else foil_type])) &
            (ccl_rows['叠构'].astype(str).str.contains(laminate, na=False))
        ]
        if len(cands) == 0:
            return None, None, err, thickness_str_in
        # 找最近厚度
        thick_vals = pd.to_numeric(cands['不含铜板厚/（mm)'], errors='coerce')
        thick_vals = thick_vals.dropna()
        if len(thick_vals) == 0:
            return None, None, err, thickness_str_in
        nearest_val = thick_vals.iloc[(thick_vals - thickness_raw_in).abs().argsort()].iloc[0]
        nearest_str = format_thickness(nearest_val)
        delta = abs(float(nearest_val) - float(thickness_raw_in))
        if delta > MAX_THICKNESS_DELTA + 1e-9:
            return (
                None,
                None,
                f"未找到匹配：厚度={thickness_str_in}，最近厚度={nearest_str}，差异={delta:.3f}mm，超过允许范围±{MAX_THICKNESS_DELTA:.2f}mm",
                thickness_str_in,
            )
        log(f"  厚度 {thickness_str_in} 未匹配，取最近厚度 {nearest_str}")
        price2, row_idx2, err2 = query_ccl_price(
            df_price, glue, nearest_str, cu_thick, foil_type, laminate, size_col, cu_thick_raw=cu_thick
        )
        return price2, row_idx2, err2, nearest_str

    if is_standard_size(w, h):
        # ===== 第一步：标准尺寸 =====
        size_col = get_standard_size_col(w, h)
        log(f"  标准尺寸 {w}x{h} -> 价格列：{size_col}")
        
        price, row_idx, err, thickness_used = _try_query_with_nearest_thickness(size_col, thickness_str, thickness_raw)
        if err:
            return None, '', err
        
        # 非标准宽度修正（标准尺寸不需要，因为标准尺寸宽度都是49）
        approx_note = f"（近似厚度{thickness_str}→{thickness_used}）" if thickness_used != thickness_str else ""
        note = (f"[标准尺寸{approx_note}] 胶系={glue} | 厚度={thickness_used}mm | 铜厚={cu_thick} | "
                f"铜箔={foil_type}{foil_default_note} | 叠构={laminate} | "
                f"尺寸{w}x{h}→{round(w)-1}x{round(h)-1} | 列={size_col} | 价格={price}")
        return round_price(price), note, None
    
    else:
        # ===== 第二步：非标准尺寸 =====
        log(f"  非标准尺寸 {w}x{h}，查询基板对照表...")
        
        big_w, big_h, qty, err = query_nonstandard_size(df_account, w, h)
        if err:
            actual_h = round(h)
            actual_w = round(w)
            fallback = TAIL_BOARD_WIDTH_MAP.get(actual_w)
            if actual_h in TAIL_BOARD_H_VALUES and fallback:
                size_col, multiplier = fallback
                price_unit, row_idx, price_err, thickness_used = _try_query_with_nearest_thickness(size_col, thickness_str, thickness_raw)
                if price_err:
                    return None, '', price_err
                final_price = round_price(multiplier * price_unit * actual_h / 48 * NON_STD_CORRECTION)
                approx_note = f"（近似厚度{thickness_str}→{thickness_used}）" if thickness_used != thickness_str else ""
                note = (f"[尾板尺寸回退{approx_note}] 胶系={glue} | 厚度={thickness_used}mm | 铜厚={cu_thick} | "
                        f"铜箔={foil_type}{foil_default_note} | 叠构={laminate} | 小片{w}x{h} | "
                        f"基板对账未命中，按列={size_col} 标准价={price_unit} | 倍率={multiplier} | "
                        f"{multiplier}×{price_unit}×{actual_h}/48×{NON_STD_CORRECTION}={format_price(final_price)}")
                return final_price, note, None
            return None, '', err
        
        std_w, std_h, size_col, multiplier = get_price_col_from_big(big_w, big_h)
        
        if not size_col:
            return None, '', f"大板 {big_w}x{big_h} 无法确定价格列"
        
        log(f"  大板 {big_w}x{big_h} -> 标准 {std_w}x{std_h} -> 价格列 {size_col} -> 倍率 {multiplier}")
        
        price_unit, row_idx, err, thickness_used = _try_query_with_nearest_thickness(size_col, thickness_str, thickness_raw)
        if err:
            return None, '', err
        
        # 非标准尺寸按大板对应的拼板倍率折算：
        # 1倍：大板本身就是标准板；2/3倍：大板由多个标准板拼成
        base_price = multiplier * price_unit / qty
        tail_board_factor = None
        packed_h = round(h * max(1, int((big_h / h) + 0.5)))
        if packed_h in TAIL_BOARD_H_VALUES:
            tail_board_factor = (packed_h, 48, NON_STD_CORRECTION)
            final_price = round_price(base_price * packed_h / 48 * NON_STD_CORRECTION)
        else:
            final_price = round_price(base_price)
        
        approx_note = f"（近似厚度{thickness_str}→{thickness_used}）" if thickness_used != thickness_str else ""
        note = (f"[非标准尺寸{approx_note}] 胶系={glue} | 厚度={thickness_used}mm | 铜厚={cu_thick} | "
                f"铜箔={foil_type}{foil_default_note} | 叠构={laminate} | 小片{w}x{h} | "
                f"大板{big_w}x{big_h}(共{qty}片) | 标准{std_w}x{std_h}→{std_w-1}x{std_h-1} | "
                f"列={size_col} | 单价={price_unit} | 倍率={multiplier} | "
                f"{multiplier}×{price_unit}/{qty}={round_price(base_price)}")
        if tail_board_factor:
            actual_h, standard_width, correction = tail_board_factor
            note += f" | 尾板修正={actual_h}/{standard_width}×{correction} | {multiplier}×{price_unit}/{qty}×{actual_h}/{standard_width}×{correction}={final_price}"
        
        return final_price, note, None

def _calc_roll(desc, df_price):
    """
    卷料价格计算（如 200M/Roll、300M/Roll）
    逻辑：与 PP 相同，叠构从 M/Roll 前的数字提取
    """
    # 提取胶系，兼容 PP 开头和隐式 PP（如 NY2150P 1080 RC71%）
    match = re.match(r'PP\s+([\w\-\(\)\.]+)', desc)
    if match:
        raw_glue = match.group(1).strip()
    else:
        match2 = re.match(r'([\w\-\(\)\.]+)\s+(\d+)\s+RC', desc, re.IGNORECASE)
        if not match2:
            return None, '', f'卷料无法提取胶系：{desc[:60]}'
        raw_glue = match2.group(1).strip()
    
    # 提取叠构类型（如 1080、1078）
    lam_match = re.search(r'PP\s+[\w\-\(\)\.]+\s+(\d+)\s+RC', desc, re.IGNORECASE)
    if not lam_match and not match:
        lam_match = re.match(r'[\w\-\(\)\.]+\s+(\d+)\s+RC', desc, re.IGNORECASE)
    if not lam_match:
        # 尝试从叠构字段提取
        lam_match2 = re.search(r'\b(1078|1080|1035|2116|3313|106|1067)\b', desc)
        if not lam_match2:
            return None, '', f'卷料无法提取叠构类型：{desc[:60]}'
        laminate_type = lam_match2.group(1)
    else:
        laminate_type = lam_match.group(1)
    
    # 提取 RC%
    rc_percent = parse_rc_percent(desc)
    if rc_percent is None:
        return None, '', f'卷料无法提取 RC%：{desc[:60]}'
    
    # 提取卷料宽度（如 49.5"*200M/Roll 或 49.5" ... 300M/卷 中的 49.5）
    w = extract_pp_roll_width(desc)
    if w is None:
        return None, '', f'卷料无法提取宽度（格式如 49.5"*200M/Roll）：{desc[:60]}'
    
    log(f"  卷料解析：原始胶系={raw_glue}, 叠构={laminate_type}, RC%={rc_percent}, 宽度={w}")
    
    rmb_sf, row_idx, err = query_pp_price(df_price, raw_glue, laminate_type, rc_percent)
    if err:
        return None, '', err
    
    price = ROLL_FIXED_WIDTH / ROLL_MM_DIVISOR / PP_FIXED_DIV * rmb_sf
    
    note = (f"[卷料/PP] 原始胶系={raw_glue} | 叠构={laminate_type} | RC%={rc_percent} | "
            f"宽度={w} | RMB/SF={rmb_sf} | "
            f"公式={ROLL_FIXED_WIDTH}/{ROLL_MM_DIVISOR}/{PP_FIXED_DIV}×{rmb_sf} = {format_price(price)}")
    return round_price(price), note, None


def _calc_pp(desc, df_price):
    """PP 物料价格计算（支持 PP 开头和非 PP 开头两种格式）"""
    
    # 尝试匹配 PP 开头格式：PP NY2170P 1080 RC68%
    match = re.match(r'PP\s+([\w\-\(\)\.]+)', desc)
    if match:
        raw_glue = match.group(1).strip()
        lam_match = re.search(r'PP\s+[\w\-\(\)\.]+\s+(\d+)\s+RC', desc, re.IGNORECASE)
    else:
        # 非 PP 开头格式：NY2170P 1080 RC68% 21.6"x24.6" 有卤 CAF
        match2 = re.match(r'([\w\-\(\)\.]+)\s+(\d+)\s+RC', desc, re.IGNORECASE)
        if not match2:
            return None, '', f'PP 无法提取胶系：{desc}'
        raw_glue = match2.group(1).strip()
        lam_match = match2  # group(2) 就是叠构类型
    
    if not lam_match:
        return None, '', f'PP 无法提取叠构类型：{desc}'
    
    # PP 开头格式用 group(1)，非 PP 开头格式用 group(2)
    laminate_type = lam_match.group(1) if match else lam_match.group(2)
    
    rc_percent = parse_rc_percent(desc)
    if rc_percent is None:
        return None, '', f'PP 无法提取 RC%：{desc}'
    
    size = parse_size(desc)
    if not size:
        return None, '', f'PP 无法提取尺寸：{desc}'
    w, h = size
    
    log(f"  PP解析：原始胶系={raw_glue}, 叠构={laminate_type}, RC%={rc_percent}, 尺寸={w}x{h}")
    
    rmb_sf, row_idx, err = query_pp_price(df_price, raw_glue, laminate_type, rc_percent)
    if err:
        return None, '', err
    
    divisor = pp_piece_divisor(h)
    if divisor == 0:
        return None, '', f'PP 计算除数为0：纬向h={h}，阈值={PP_SMALL_PIECE_TAIL_THRESHOLD}'
    
    price = w * PP_FIXED_WIDTH / PP_FIXED_DIV / divisor * rmb_sf
    
    pp_rows = df_price[df_price['CCL'].astype(str).str.strip() == 'PP']
    glue_used = convert_pp_glue(raw_glue, pp_rows)
    
    note = (f"[PP] 原始胶系={raw_glue}→匹配={glue_used} | 叠构={laminate_type} | RC%={rc_percent} | "
            f"尺寸={w}x{h} | RMB/SF={rmb_sf} | "
            f"开数=纬向h={h} {'<=' if h <= PP_SMALL_PIECE_TAIL_THRESHOLD else '>'} {PP_SMALL_PIECE_TAIL_THRESHOLD}，按1开{divisor} | "
            f"公式={w}×{PP_FIXED_WIDTH}/{PP_FIXED_DIV}/{divisor}×{rmb_sf} "
            f"= {format_price(price)}")
    return round_price(price), note, None

# ============================================================
# 主处理流程
# ============================================================
def process_file(filepath):
    """处理单个 Excel 文件"""
    log(f"\n{'='*60}")
    log(f"开始处理：{os.path.basename(filepath)}")
    log(f"{'='*60}")
    
    df_calc, df_price, df_account = load_excel(filepath)
    desc_col = detect_description_column_dataframe(df_calc)
    log(f"  物料描述列：第 {desc_col + 1} 列")
    
    results = []
    success_count = 0
    fail_count = 0
    skip_count = 0
    
    for idx, row in df_calc.iterrows():
        raw_desc = row.iloc[desc_col] if desc_col < len(row) else ''
        desc = str(raw_desc).strip() if pd.notna(raw_desc) else ''
        pp_roll_price = calculate_pp_roll_price(desc, df_price)
        pp_roll_price_value = round_price(pp_roll_price) if pp_roll_price is not None else ''
        
        if not desc or desc == 'nan':
            results.append({'行号': idx+2, '物料描述': '', '价格': '', '输出价格': '', '说明': '', '状态': '跳过'})
            skip_count += 1
            continue
        
        log(f"\n行 {idx+2}：{desc[:70]}")
        
        price, note, err = calculate_price(desc, df_price, df_account)
        
        if err:
            log(f"  失败：{err}", 'ERROR')
            results.append({'行号': idx+2, '物料描述': desc, '价格': '', '输出价格': '', '说明': err, '状态': '失败'})
            fail_count += 1
        else:
            output_price = output_price_for_desc(desc, price, pp_roll_price_value)
            log(f"  价格：{output_price}")
            results.append({'行号': idx+2, '物料描述': desc, '价格': price, '输出价格': output_price, '说明': note, '状态': '成功'})
            success_count += 1
            if len(df_calc.columns) > 8:
                df_calc.iloc[idx, 8] = output_price
    
    log(f"\n{'='*60}")
    log(f"处理完成：成功 {success_count} 行，失败 {fail_count} 行，跳过 {skip_count} 行")
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = os.path.splitext(os.path.basename(filepath))[0]
    output_path = os.path.join(OUTPUT_DIR, f"{base_name}_计算结果_{timestamp}.xlsx")
    
    save_result(filepath, df_calc, results, output_path)
    log(f"结果已保存：{output_path}")
    
    return output_path, results, success_count, fail_count

def save_result(source_path, df_calc, results, output_path, sheet_name=None):
    """保存结果到 Excel，保留原始格式，并添加计算说明 Sheet"""
    shutil.copy2(source_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    calc_sheet_name = sheet_name or select_calculation_sheet_name(wb.sheetnames)
    ws_calc = wb[calc_sheet_name]
    
    price_col = _last_value_column(ws_calc) + 1
    header_row = _detect_header_row(ws_calc)
    header_cell = ws_calc.cell(row=header_row, column=price_col, value='方正计算价格')
    header_cell.font = Font(bold=True)
    
    for r in results:
        if r['状态'] == '跳过':
            continue
        row_num = r['行号']
        output_price = r.get('输出价格', r.get('价格', ''))
        if r['状态'] == '成功' and output_price != '':
            ws_calc.cell(row=row_num, column=price_col, value=output_price)
            ws_calc.cell(row=row_num, column=price_col).fill = PatternFill(
                start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'
            )
        elif r['状态'] == '失败':
            ws_calc.cell(row=row_num, column=price_col, value='未找到')
            ws_calc.cell(row=row_num, column=price_col).fill = PatternFill(
                start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'
            )
    
    # 添加计算说明 Sheet
    if '计算说明' in wb.sheetnames:
        del wb['计算说明']
    ws_note = wb.create_sheet('计算说明')
    
    headers = ['行号', '物料描述', '计算价格', '计算说明', '状态']
    for col, h in enumerate(headers, 1):
        cell = ws_note.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    data_row = 2
    for r in results:
        if r['状态'] == '跳过':
            continue
        ws_note.cell(row=data_row, column=1, value=r['行号'])
        ws_note.cell(row=data_row, column=2, value=str(r['物料描述'])[:120])
        ws_note.cell(row=data_row, column=3, value=r.get('输出价格', r.get('价格', '')))
        ws_note.cell(row=data_row, column=4, value=str(r['说明'])[:200])
        ws_note.cell(row=data_row, column=5, value=r['状态'])
        
        if r['状态'] == '成功':
            ws_note.cell(row=data_row, column=5).fill = PatternFill(
                start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'
            )
        elif r['状态'] == '失败':
            ws_note.cell(row=data_row, column=5).fill = PatternFill(
                start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'
            )
        data_row += 1
    
    ws_note.column_dimensions['A'].width = 8
    ws_note.column_dimensions['B'].width = 65
    ws_note.column_dimensions['C'].width = 12
    ws_note.column_dimensions['D'].width = 90
    ws_note.column_dimensions['E'].width = 10
    
    wb.save(output_path)


def _last_value_column(ws) -> int:
    last_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ''):
                last_col = max(last_col, cell.column)
    return last_col


def _detect_header_row(ws) -> int:
    header_names = {'物料描述', '规格', '客户规格', '材料描述', '物料规格'}
    for row in range(1, min(ws.max_row, 30) + 1):
        values = {str(cell.value).strip() for cell in ws[row] if cell.value not in (None, '')}
        if values & header_names:
            return row
    return 1

# ============================================================
# 对外接口（供 GUI 调用）
# ============================================================
def run_calculation(filepath, progress_callback=None):
    """
    供 GUI 调用的接口
    progress_callback(current, total, message) 用于更新进度
    返回：(output_path, results, success_count, fail_count)
    """
    ensure_dirs()
    return process_file(filepath)

# ============================================================
# 入口
# ============================================================
def main():
    ensure_dirs()
    
    print("=" * 60)
    print("  方正价格自动计算工具 v3.0")
    print("=" * 60)
    print(f"  输入目录：{INPUT_DIR}")
    print(f"  输出目录：{OUTPUT_DIR}")
    print()
    
    excel_files = [
        f for f in os.listdir(INPUT_DIR)
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')
    ]
    
    if not excel_files:
        print("未在 input/ 目录中找到 Excel 文件")
        print(f"  请将 Excel 文件放入：{INPUT_DIR}")
        input("\n按 Enter 键退出...")
        return
    
    print(f"找到 {len(excel_files)} 个文件：")
    for f in excel_files:
        print(f"  - {f}")
    print()
    
    total_success = 0
    total_fail = 0
    
    for filename in excel_files:
        filepath = os.path.join(INPUT_DIR, filename)
        try:
            output_path, results, s, f = process_file(filepath)
            total_success += s
            total_fail += f
            print(f"\n成功 {filename} -> {os.path.basename(output_path)}")
            print(f"  成功：{s} 行，失败：{f} 行")
        except Exception as e:
            print(f"\n失败 {filename} 处理失败：{e}")
            import traceback
            traceback.print_exc()
    
    print("\n" + "=" * 60)
    print(f"  全部完成！成功 {total_success} 行，失败 {total_fail} 行")
    print(f"  结果文件保存在：{OUTPUT_DIR}")
    print("=" * 60)
    input("\n按 Enter 键退出...")

def save_result_v3(source_path, results, output_path, sheet_name=None):
    """
    GUI 专用保存函数：支持精简版 Excel（只含价格计算 Sheet）
    内置数据版本不需要方正价格和基板对账 Sheet
    """
    save_result(source_path, None, results, output_path, sheet_name=sheet_name)


if __name__ == '__main__':
    main()
