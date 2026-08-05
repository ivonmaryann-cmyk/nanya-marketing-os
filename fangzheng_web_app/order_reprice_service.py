from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from .db import append_job_log, create_job, get_job, update_job_status
from .job_control import launch_job_process
from .paths import JOBS_DIR


FEATURE = "order_reprice"
RULE_VERSION = "订单改价内置规则 v1"
EXCEL_EXTENSIONS = (".xlsx", ".xlsm", ".xls")
JINGWANG_PRICE_CHECK_MODE = "jingwang_price_check"
JINGWANG_FACTORY_MERGE_MODE = "jingwang_factory_merge"
JINGWANG_CUSTOMER_PRICE_CHECK_MODE = "jingwang_customer_price_check"
JINGWANG_MERGED_COLUMNS = [
    "来源",
    "合并表行号",
    "单别单号",
    "项次",
    "客户单号",
    "客户产品编号",
    "数量",
    "单位",
    "单价",
    "总金额",
]
JINGWANG_FACTORY_MERGED_COLUMNS = JINGWANG_MERGED_COLUMNS + ["规格", "客户规格", "品名"]
JINGWANG_ALLOWED_FACTORY_GROUPS = {"乐健集团", "景旺集团"}
JINGWANG_FACTORY_FILTER_COLUMNS = [
    "来源",
    "原始行号",
    "单别单号",
    "项次",
    "客户单号",
    "客户产品编号",
    "单价",
    "集团",
    "过滤原因",
]
JINGWANG_MERGED_CUSTOMER_COLUMNS = [
    "客户匹配状态",
    "客户匹配数量",
    "对应客户数据行号",
    "客户物料编号",
    "客户采购单号",
    "客户价格",
    "客户金额",
    "PP米数",
    "厂内核对价格",
    "价格核对结果",
    "价格差额",
    "客户匹配说明",
]

MODE_LABELS = {
    "block1": "第一块：客户明细与430厂内明细匹配",
    "block2": "第二块：430价格核对",
    "block3": "第三块：客户改价结果与411厂内价格核对",
    JINGWANG_FACTORY_MERGE_MODE: "景旺厂内数据合并",
    JINGWANG_CUSTOMER_PRICE_CHECK_MODE: "景旺客户数据核价",
    JINGWANG_PRICE_CHECK_MODE: "景旺订单核价",
}


@dataclass
class MatchResult:
    status: str
    factory_indexes: list[int]
    factory_items: list[str]
    note: str = ""


@dataclass
class QuoteRow:
    source_file: str
    sheet_name: str
    excel_row: int
    product: str
    material_type: str
    values: dict[str, Any]
    size_prices: dict[str, float]


@dataclass
class ParsedSpec:
    material_type: str
    product: str | None
    thickness: float | None = None
    copper: str | None = None
    foil: str | None = None
    copper_state: str | None = None
    stack: str | None = None
    size_key: str | None = None
    glass: str | None = None
    rc: float | None = None
    length_m: float | None = None


def queue_order_reprice_job(
    employee_id: str,
    mode: str,
    customer_file=None,
    factory_file=None,
    quote_files: list | None = None,
    *,
    customer_key: str = "shenghong",
    factory_701_file=None,
    factory_411_file=None,
    merged_file=None,
) -> int:
    if mode not in MODE_LABELS:
        raise ValueError("未知订单改价处理块")
    quote_files = quote_files or []

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    job_dir = employee_dir / f"{timestamp}_order_reprice_{mode}"
    job_dir.mkdir(parents=True, exist_ok=True)

    if mode == JINGWANG_FACTORY_MERGE_MODE:
        if not factory_701_file or not factory_411_file:
            raise ValueError("景旺厂内数据合并需要上传厂内701和厂内411两份 Excel")
        factory_701_path = _save_upload(factory_701_file, job_dir, "factory_701")
        factory_411_path = _save_upload(factory_411_file, job_dir, "factory_411")
        quote_paths = []
        manifest = {
            "mode": mode,
            "label": MODE_LABELS[mode],
            "customer_key": "jingwang",
            "customer_label": "景旺",
            "factory_701_path": str(factory_701_path),
            "factory_411_path": str(factory_411_path),
            "quote_paths": [],
        }
        source_filename = f"{MODE_LABELS[mode]}：{factory_701_file.filename} + {factory_411_file.filename}"
    elif mode == JINGWANG_CUSTOMER_PRICE_CHECK_MODE:
        if not customer_file or not merged_file:
            raise ValueError("景旺客户数据核价需要上传701+411合并表和客户数据两份 Excel")
        customer_path = _save_upload(customer_file, job_dir, "customer")
        merged_path = _save_upload(merged_file, job_dir, "merged")
        quote_paths = []
        manifest = {
            "mode": mode,
            "label": MODE_LABELS[mode],
            "customer_key": "jingwang",
            "customer_label": "景旺",
            "customer_path": str(customer_path),
            "merged_path": str(merged_path),
            "quote_paths": [],
        }
        source_filename = f"{MODE_LABELS[mode]}：{merged_file.filename} + {customer_file.filename}"
    elif mode == JINGWANG_PRICE_CHECK_MODE:
        if not factory_701_file or not factory_411_file:
            raise ValueError("景旺订单核价需要上传厂内701、厂内411和客户数据三份 Excel")
        customer_path = _save_upload(customer_file, job_dir, "customer")
        factory_701_path = _save_upload(factory_701_file, job_dir, "factory_701")
        factory_411_path = _save_upload(factory_411_file, job_dir, "factory_411")
        quote_paths = []
        manifest = {
            "mode": mode,
            "label": MODE_LABELS[mode],
            "customer_key": "jingwang",
            "customer_label": "景旺",
            "customer_path": str(customer_path),
            "factory_701_path": str(factory_701_path),
            "factory_411_path": str(factory_411_path),
            "quote_paths": [],
        }
        source_filename = (
            f"{MODE_LABELS[mode]}：{factory_701_file.filename} + "
            f"{factory_411_file.filename} + {customer_file.filename}"
        )
    else:
        if mode != "block2" and not factory_file:
            raise ValueError("请上传厂内明细 Excel 文件")
        customer_path = _save_upload(customer_file, job_dir, "customer")
        factory_path = _save_upload(factory_file, job_dir, "factory") if factory_file else None
        quote_paths = [_save_upload(file_obj, job_dir, f"quote_{idx}") for idx, file_obj in enumerate(quote_files, 1)]
        manifest = {
            "mode": mode,
            "label": MODE_LABELS[mode],
            "customer_key": customer_key or "shenghong",
            "customer_label": "胜宏",
            "customer_path": str(customer_path),
            "factory_path": str(factory_path) if factory_path else "",
            "quote_paths": [str(path) for path in quote_paths],
        }
        source_filename = f"{MODE_LABELS[mode]}：{customer_file.filename}"
        if factory_file:
            source_filename += f" + {factory_file.filename}"
        if quote_paths:
            source_filename += f" + {len(quote_paths)}份报价单"

    manifest_path = job_dir / "order_reprice_manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    job_id = create_job(employee_id, source_filename, str(manifest_path), RULE_VERSION, feature=FEATURE)
    launch_job_process(job_id, FEATURE, employee_id)
    return job_id


def run_order_reprice_job(job_id: int, employee_id: str) -> None:
    job = get_job(job_id)
    if not job or job["employee_id"] != employee_id:
        return

    update_job_status(job_id, status="running")
    try:
        manifest = json.loads(Path(job["stored_input_path"]).read_text(encoding="utf-8"))
        mode = manifest["mode"]
        append_job_log(job_id, f"开始处理订单改价任务：{MODE_LABELS.get(mode, mode)}")

        customer_path = Path(manifest["customer_path"]) if manifest.get("customer_path") else None
        factory_path = Path(manifest["factory_path"]) if manifest.get("factory_path") else None
        quote_paths = [Path(path) for path in manifest.get("quote_paths", [])]

        if mode == JINGWANG_FACTORY_MERGE_MODE:
            result = process_jingwang_factory_merge(
                Path(manifest["factory_701_path"]),
                Path(manifest["factory_411_path"]),
                job_id=job_id,
            )
        elif mode == JINGWANG_CUSTOMER_PRICE_CHECK_MODE:
            result = process_jingwang_customer_price_check(
                customer_path,
                Path(manifest["merged_path"]),
                job_id=job_id,
            )
        elif mode == JINGWANG_PRICE_CHECK_MODE:
            result = process_jingwang_price_check(
                customer_path,
                Path(manifest["factory_701_path"]),
                Path(manifest["factory_411_path"]),
                job_id=job_id,
            )
        elif mode == "block1":
            result = process_block1(customer_path, factory_path, job_id=job_id)
        elif mode == "block2":
            result = process_block2(customer_path, factory_path, quote_paths, job_id=job_id)
        elif mode == "block3":
            result = process_block3(customer_path, factory_path, job_id=job_id)
        else:
            raise ValueError(f"未知订单改价处理块：{mode}")

        output_path = Path(job["stored_input_path"]).with_name(
            f"订单改价结果_{mode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        save_result_workbook(result["sheets"], output_path)
        update_job_status(
            job_id,
            status="completed",
            stored_result_path=str(output_path),
            success_count=result["success_count"],
            fail_count=result["fail_count"],
            skip_count=result.get("skip_count", 0),
            current_row=result["total_rows"],
            total_rows=result["total_rows"],
            completed=True,
        )
        append_job_log(
            job_id,
            f"订单改价任务完成：成功 {result['success_count']}，失败 {result['fail_count']}，结果已生成",
        )
    except Exception as exc:
        append_job_log(job_id, f"订单改价任务失败：{exc}")
        update_job_status(job_id, status="failed", error_message=str(exc), completed=True)
        raise


def process_block1(customer_path: Path, factory_path: Path, *, job_id: int | None = None) -> dict:
    customer_df = read_business_sheet(customer_path, {"采购订单号", "项次", "料件编号"})
    factory_df = read_business_sheet(factory_path, {"客户订单", "项次", "客户产品编号"})
    matches = match_customer_to_factory(
        customer_df,
        factory_df,
        customer_cols=("采购订单号", "项次", "料件编号"),
        factory_cols=("客户订单", "项次", "客户产品编号"),
    )
    return build_match_result(customer_df, factory_df, matches, job_id=job_id)


def process_block2(
    customer_path: Path,
    factory_path: Path | None,
    quote_paths: list[Path],
    *,
    job_id: int | None = None,
) -> dict:
    if not quote_paths:
        raise ValueError("第二块需要上传至少一份胜宏报价单")

    customer_df = read_block2_customer_sheet(customer_path)
    has_tax_price = "含税单价" in customer_df.columns
    matches: dict[int, MatchResult] = {}
    if factory_path is not None:
        factory_df = read_business_sheet(factory_path, {"客户订单", "项次", "客户产品编号"})
        matches = match_customer_to_factory(
            customer_df,
            factory_df,
            customer_cols=("采购订单号", "项次", "料件编号"),
            factory_cols=("客户订单", "项次", "客户产品编号"),
        )
    quote_rows = load_quote_rows(quote_paths)
    append_job_log(job_id, f"已读取胜宏报价单：{len(quote_rows)} 条报价记录", total_rows=len(customer_df)) if job_id else None

    output = customer_df.copy()
    notes: list[dict[str, Any]] = []
    success = 0
    fail = 0
    skipped = 0
    actual_qty_values = []
    factory_items = []
    quote_prices = []
    compare_results = []
    quote_notes = []
    price_diffs = []

    for pos, (idx, row) in enumerate(customer_df.iterrows(), 1):
        spec = _text(row.get("规格"))
        actual_qty = calculate_actual_quantity(row.get("采购单位"), row.get("采购量"), spec)
        actual_qty_values.append(actual_qty)

        match = matches.get(idx, MatchResult("未匹配", [], [], ""))
        factory_items.append(",".join(match.factory_items))

        quote = find_quote_price(spec, quote_rows)
        quote_prices.append(quote["price"])
        quote_notes.append(quote["note"])

        if quote.get("status") == "skipped":
            compare = "不输出"
            price_diffs.append(None)
            skipped += 1
        elif quote["price"] is None:
            compare = "未命中报价"
            price_diffs.append(None)
            fail += 1
            notes.append(
                {
                    "行号": pos + 1,
                    "采购订单号": row.get("采购订单号") or row.get("订购单号"),
                    "项次": row.get("项次"),
                    "料件编号": row.get("料件编号"),
                    "规格": spec,
                    "原因": quote["note"],
                }
            )
        elif not has_tax_price or _number(row.get("含税单价")) is None:
            compare = "无需比对"
            success += 1
        else:
            diff = _price_diff(quote["price"], row.get("含税单价"))
            price_diffs.append(diff)
            if _prices_equal(quote["price"], row.get("含税单价")):
                compare = "一致"
                success += 1
            else:
                compare = "不一致"
                fail += 1
            compare_results.append(compare)
            if job_id and (pos == 1 or pos % 50 == 0 or pos == len(customer_df)):
                append_job_log(
                    job_id,
                    f"第二块价格核对进度：{pos}/{len(customer_df)}",
                    success_count=success,
                    fail_count=fail,
                    current_row=pos,
                    total_rows=len(customer_df),
                )
            continue
        if compare in {"无需比对"}:
            price_diffs.append(None)
        compare_results.append(compare)

        if job_id and (pos == 1 or pos % 50 == 0 or pos == len(customer_df)):
            append_job_log(
                job_id,
                f"第二块价格核对进度：{pos}/{len(customer_df)}",
                success_count=success,
                fail_count=fail,
                current_row=pos,
                total_rows=len(customer_df),
            )

    output["实际订单数量"] = actual_qty_values
    output["厂内项次"] = factory_items
    output["报价单价格"] = quote_prices
    if has_tax_price:
        output["价格差额"] = price_diffs
    output["价格比对结果"] = compare_results
    output["报价命中说明"] = quote_notes

    summary = summary_sheet(
        [
            ("总记录数", len(output)),
            ("价格一致数量", compare_results.count("一致")),
            ("价格不一致数量", compare_results.count("不一致")),
            ("无需比对数量", compare_results.count("无需比对")),
            ("不输出数量", compare_results.count("不输出")),
            ("未命中报价数量", compare_results.count("未命中报价")),
            ("未匹配厂内数量", sum(1 for item in matches.values() if item.status == "未匹配") if matches else 0),
        ]
    )
    return {
        "sheets": [
            ("客户价格核对结果", output),
            ("匹配汇总", summary),
            ("报价未命中说明", pd.DataFrame(notes)),
        ],
        "success_count": success,
        "fail_count": fail,
        "skip_count": skipped,
        "total_rows": len(output),
    }


def process_block3(customer_path: Path, factory_path: Path, *, job_id: int | None = None) -> dict:
    customer_df = read_business_sheet(customer_path, {"采购订单号", "项次", "料件编号", "规格", "含税单价"})
    factory_df = read_business_sheet(factory_path, {"客户单号", "项次", "客户产品编号", "单价"})
    matches = match_customer_to_factory(
        customer_df,
        factory_df,
        customer_cols=("采购订单号", "项次", "料件编号"),
        factory_cols=("客户单号", "项次", "客户产品编号"),
    )

    output = customer_df.copy()
    statuses = []
    factory_items = []
    factory_prices = []
    factory_price_notes = []
    compare_results = []
    success = 0
    fail = 0

    for pos, (idx, row) in enumerate(customer_df.iterrows(), 1):
        match = matches.get(idx, MatchResult("未匹配", [], [], ""))
        statuses.append(match.status)
        factory_items.append(",".join(match.factory_items))
        price_parts = [
            _block3_factory_price(factory_df.loc[fidx], row.get("规格"))
            for fidx in _sort_factory_indexes_by_item(factory_df, match.factory_indexes)
        ]
        matched_price, price_note = _pick_block3_price(price_parts, len(match.factory_indexes) > 1)
        factory_price_notes.append(price_note)
        factory_prices.append(matched_price)
        if matched_price is None:
            result = "未匹配"
            fail += 1
        elif _prices_equal(matched_price, row.get("含税单价")):
            result = "正确"
            success += 1
        else:
            result = "错误"
            fail += 1
        compare_results.append(result)

        if job_id and (pos == 1 or pos % 50 == 0 or pos == len(customer_df)):
            append_job_log(
                job_id,
                f"第三块价格核对进度：{pos}/{len(customer_df)}",
                success_count=success,
                fail_count=fail,
                current_row=pos,
                total_rows=len(customer_df),
            )

    output["匹配状态"] = statuses
    output["厂内项次"] = factory_items
    output["厂内匹配价格"] = factory_prices
    output["厂内价格换算说明"] = factory_price_notes
    output["价格核对结果"] = compare_results
    summary = summary_sheet(
        [
            ("总记录数", len(output)),
            ("价格正确数量", compare_results.count("正确")),
            ("价格错误数量", compare_results.count("错误")),
            ("未匹配数量", compare_results.count("未匹配")),
        ]
    )
    return {
        "sheets": [("客户改价核对结果", output), ("核对汇总", summary)],
        "success_count": success,
        "fail_count": fail,
        "skip_count": 0,
        "total_rows": len(output),
    }


def process_jingwang_factory_merge(
    factory_701_path: Path,
    factory_411_path: Path,
    *,
    job_id: int | None = None,
) -> dict:
    factory_701_df = read_business_sheet(
        factory_701_path,
        {"订单号/出货单号", "项次", "客户单号", "客户产品编号", "数量", "单位", "单价", "本币金额", "所属集团"},
    )
    factory_411_df = read_business_sheet(
        factory_411_path,
        {"单别单号", "项次", "客户单号", "客户产品编号", "订单数量", "单位", "单价", "总金额", "所属集团公司"},
    )
    if job_id:
        append_job_log(
            job_id,
            f"已读取景旺厂内文件：701 {len(factory_701_df)} 行，411 {len(factory_411_df)} 行",
            total_rows=len(factory_701_df) + len(factory_411_df),
        )
    filtered_701_df, filtered_411_df, filter_df, filter_stats = filter_jingwang_factory_rows(factory_701_df, factory_411_df)
    merged_df, covered_df = build_jingwang_factory_merge(filtered_701_df, filtered_411_df)
    summary = summary_sheet(
        [
            ("701原始行数", len(factory_701_df)),
            ("411原始行数", len(factory_411_df)),
            ("701过滤删除数", filter_stats["filtered_701"]),
            ("411过滤删除数", filter_stats["filtered_411"]),
            ("过滤后701行数", len(filtered_701_df)),
            ("过滤后411行数", len(filtered_411_df)),
            ("单价为0删除数", filter_stats["zero_price"]),
            ("集团不符合删除数", filter_stats["invalid_group"]),
            ("411被701覆盖数量", len(covered_df)),
            ("合并表行数", len(merged_df)),
        ]
    )
    if job_id:
        append_job_log(
            job_id,
            (
                f"景旺厂内数据合并完成：过滤删除 {len(filter_df)} 行，"
                f"合并表 {len(merged_df)} 行，411被701覆盖 {len(covered_df)} 行"
            ),
            success_count=len(merged_df),
            fail_count=0,
            current_row=len(factory_701_df) + len(factory_411_df),
            total_rows=len(factory_701_df) + len(factory_411_df),
        )
    return {
        "sheets": [
            ("701+411合并表", merged_df),
            ("411被701覆盖明细", covered_df),
            ("厂内过滤明细", filter_df),
            ("合并汇总", summary),
        ],
        "success_count": len(merged_df),
        "fail_count": 0,
        "skip_count": 0,
        "total_rows": len(merged_df),
    }


def filter_jingwang_factory_rows(
    factory_701_df: pd.DataFrame,
    factory_411_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, int]]:
    filtered_701_df, filter_701_rows, stats_701 = _filter_jingwang_factory_df(
        factory_701_df,
        source="701",
        group_col="所属集团",
        order_col="订单号/出货单号",
    )
    filtered_411_df, filter_411_rows, stats_411 = _filter_jingwang_factory_df(
        factory_411_df,
        source="411",
        group_col="所属集团公司",
        order_col="单别单号",
    )
    filter_df = pd.DataFrame(filter_701_rows + filter_411_rows, columns=JINGWANG_FACTORY_FILTER_COLUMNS)
    stats = {
        "filtered_701": len(filter_701_rows),
        "filtered_411": len(filter_411_rows),
        "zero_price": stats_701["zero_price"] + stats_411["zero_price"],
        "invalid_group": stats_701["invalid_group"] + stats_411["invalid_group"],
    }
    return filtered_701_df, filtered_411_df, filter_df, stats


def _filter_jingwang_factory_df(
    df: pd.DataFrame,
    *,
    source: str,
    group_col: str,
    order_col: str,
) -> tuple[pd.DataFrame, list[dict[str, Any]], dict[str, int]]:
    keep_indexes: list[int] = []
    filtered_rows: list[dict[str, Any]] = []
    zero_price_count = 0
    invalid_group_count = 0

    for original_pos, (idx, row) in enumerate(df.iterrows(), 1):
        reasons: list[str] = []
        unit_price = _number(row.get("单价"))
        group = _text(row.get(group_col))
        if unit_price == 0:
            reasons.append("单价为0")
            zero_price_count += 1
        if group not in JINGWANG_ALLOWED_FACTORY_GROUPS:
            reasons.append("集团不符合")
            invalid_group_count += 1

        if reasons:
            filtered_rows.append(
                {
                    "来源": source,
                    "原始行号": original_pos,
                    "单别单号": row.get(order_col),
                    "项次": row.get("项次"),
                    "客户单号": row.get("客户单号"),
                    "客户产品编号": row.get("客户产品编号"),
                    "单价": row.get("单价"),
                    "集团": row.get(group_col),
                    "过滤原因": "；".join(reasons),
                }
            )
            continue
        keep_indexes.append(idx)

    return df.loc[keep_indexes].reset_index(drop=True), filtered_rows, {
        "zero_price": zero_price_count,
        "invalid_group": invalid_group_count,
    }


def build_jingwang_factory_merge(factory_701_df: pd.DataFrame, factory_411_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    merged_rows: list[dict[str, Any]] = []
    key_to_701: dict[tuple[str, str], dict[str, Any]] = {}

    for _, row in factory_701_df.iterrows():
        record = {
            "来源": "701",
            "合并表行号": len(merged_rows) + 1,
            "单别单号": row.get("订单号/出货单号"),
            "项次": row.get("项次"),
            "客户单号": row.get("客户单号"),
            "客户产品编号": row.get("客户产品编号"),
            "数量": row.get("数量"),
            "单位": row.get("单位"),
            "单价": row.get("单价"),
            "总金额": row.get("本币金额"),
            "规格": row.get("规格"),
            "客户规格": row.get("客户规格"),
            "品名": row.get("品名"),
        }
        merged_rows.append(record)
        key = (_norm_identifier(row.get("订单号/出货单号")), _norm_item(row.get("项次")))
        if all(key):
            key_to_701.setdefault(key, record)

    covered_411_rows: list[dict[str, Any]] = []
    for _, row in factory_411_df.iterrows():
        key = (_norm_identifier(row.get("单别单号")), _norm_item(row.get("项次")))
        covered_by = key_to_701.get(key) if all(key) else None
        record = {
            "来源": "411",
            "合并表行号": None,
            "单别单号": row.get("单别单号"),
            "项次": row.get("项次"),
            "客户单号": row.get("客户单号"),
            "客户产品编号": row.get("客户产品编号"),
            "数量": row.get("订单数量"),
            "单位": row.get("单位"),
            "单价": row.get("单价"),
            "总金额": row.get("总金额"),
            "规格": row.get("规格"),
            "客户规格": row.get("客户规格"),
            "品名": row.get("品名"),
        }
        if covered_by:
            covered = dict(record)
            covered["覆盖701合并表行号"] = covered_by["合并表行号"]
            covered["覆盖原因"] = "701订单号/出货单号+项次命中411单别单号+项次，保留701"
            covered_411_rows.append(covered)
            continue
        record["合并表行号"] = len(merged_rows) + 1
        merged_rows.append(record)

    merged_df = pd.DataFrame(merged_rows, columns=JINGWANG_FACTORY_MERGED_COLUMNS)
    covered_columns = JINGWANG_FACTORY_MERGED_COLUMNS + ["覆盖701合并表行号", "覆盖原因"]
    covered_df = pd.DataFrame(covered_411_rows, columns=covered_columns)
    return merged_df, covered_df


def process_jingwang_customer_price_check(
    customer_path: Path,
    merged_path: Path,
    *,
    job_id: int | None = None,
) -> dict:
    customer_df = read_business_sheet(customer_path, {"采购单号", "物料编号", "价格"})
    merged_df = read_business_sheet(merged_path, {"来源", "合并表行号", "客户单号", "客户产品编号", "单价"})
    result = build_jingwang_customer_price_check(customer_df, merged_df, job_id=job_id)
    return result


def build_jingwang_customer_price_check(
    customer_df: pd.DataFrame,
    merged_df: pd.DataFrame,
    *,
    job_id: int | None = None,
) -> dict:
    customer_po_col = _jingwang_customer_po_column(customer_df)
    merged_rows = merged_df.to_dict("records")

    customer_index: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for customer_pos, (_, row) in enumerate(customer_df.iterrows(), 1):
        key = (_norm_identifier(row.get("物料编号")), _norm_identifier(row.get(customer_po_col)))
        if not all(key):
            continue
        customer_index.setdefault(key, []).append(
            {
                "客户数据行号": customer_pos,
                "物料编号": row.get("物料编号"),
                "采购单号": row.get(customer_po_col),
                "物料名称": row.get("物料名称"),
                "价格": row.get("价格"),
                "价格(两位)": _money_2(row.get("价格")),
                "金额": row.get("金额"),
            }
        )

    merged_customer_matched = 0
    merged_customer_unmatched = 0
    merged_customer_price_equal = 0
    merged_customer_price_mismatch = 0
    for record in merged_rows:
        key = (_norm_identifier(record.get("客户产品编号")), _norm_identifier(record.get("客户单号")))
        customer_matches = customer_index.get(key, []) if all(key) else []
        if not all(key):
            factory_price, pp_length, price_note = _jingwang_factory_check_price(record)
            record.update(
                {
                    "客户匹配状态": "失败",
                    "客户匹配数量": 0,
                    "对应客户数据行号": None,
                    "客户物料编号": None,
                    "客户采购单号": None,
                    "客户价格": None,
                    "客户金额": None,
                    "PP米数": pp_length,
                    "厂内核对价格": _decimal_to_float(factory_price),
                    "价格核对结果": "未匹配",
                    "价格差额": None,
                    "客户匹配说明": f"客户产品编号或客户单号为空；{price_note}",
                }
            )
            merged_customer_unmatched += 1
            continue
        if not customer_matches:
            factory_price, pp_length, price_note = _jingwang_factory_check_price(record)
            record.update(
                {
                    "客户匹配状态": "失败",
                    "客户匹配数量": 0,
                    "对应客户数据行号": None,
                    "客户物料编号": None,
                    "客户采购单号": None,
                    "客户价格": None,
                    "客户金额": None,
                    "PP米数": pp_length,
                    "厂内核对价格": _decimal_to_float(factory_price),
                    "价格核对结果": "未匹配",
                    "价格差额": None,
                    "客户匹配说明": f"合并表客户产品编号+客户单号未命中客户数据；{price_note}",
                }
            )
            merged_customer_unmatched += 1
            continue

        customer_match = customer_matches[0]
        customer_price = customer_match["价格(两位)"]
        factory_price, pp_length, price_note = _jingwang_factory_check_price(record, customer_match)
        diff = customer_price - factory_price if customer_price is not None and factory_price is not None else None
        is_equal = diff == Decimal("0.00") if diff is not None else False
        record.update(
            {
                "客户匹配状态": "成功",
                "客户匹配数量": len(customer_matches),
                "对应客户数据行号": customer_match["客户数据行号"],
                "客户物料编号": customer_match["物料编号"],
                "客户采购单号": customer_match["采购单号"],
                "客户价格": customer_match["价格"],
                "客户金额": customer_match["金额"],
                "PP米数": pp_length,
                "厂内核对价格": _decimal_to_float(factory_price),
                "价格核对结果": "价格一致" if is_equal else "价格不一致",
                "价格差额": _decimal_to_float(diff),
                "客户匹配说明": f"{price_note}；价格两位小数一致"
                if is_equal
                else (
                    f"{price_note}；同一键匹配到多条客户数据，当前展示第一条"
                    if len(customer_matches) > 1
                    else f"{price_note}；价格两位小数不一致"
                ),
            }
        )
        merged_customer_matched += 1
        if is_equal:
            merged_customer_price_equal += 1
        else:
            merged_customer_price_mismatch += 1

    checked_merged_df = pd.DataFrame(merged_rows, columns=JINGWANG_FACTORY_MERGED_COLUMNS + JINGWANG_MERGED_CUSTOMER_COLUMNS)

    merged_index: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for record in merged_rows:
        key = (_norm_identifier(record.get("客户产品编号")), _norm_identifier(record.get("客户单号")))
        if all(key):
            merged_index.setdefault(key, []).append(record)

    output = customer_df.copy()
    match_statuses: list[str] = []
    compare_results: list[str] = []
    merged_row_numbers: list[str | None] = []
    factory_sources: list[Any] = []
    factory_prices: list[Any] = []
    factory_check_prices: list[Any] = []
    price_diffs: list[Any] = []
    notes: list[str] = []
    matched = 0
    unmatched = 0
    price_equal = 0
    price_mismatch = 0

    for pos, (_, row) in enumerate(customer_df.iterrows(), 1):
        key = (_norm_identifier(row.get("物料编号")), _norm_identifier(row.get(customer_po_col)))
        candidates = merged_index.get(key, []) if all(key) else []
        customer_price = _money_2(row.get("价格"))

        if not all(key):
            match_statuses.append("失败")
            compare_results.append("未匹配")
            merged_row_numbers.append(None)
            factory_sources.append(None)
            factory_prices.append(None)
            factory_check_prices.append(None)
            price_diffs.append(None)
            notes.append("物料编号或采购单号为空")
            unmatched += 1
        elif not candidates:
            match_statuses.append("失败")
            compare_results.append("未匹配")
            merged_row_numbers.append(None)
            factory_sources.append(None)
            factory_prices.append(None)
            factory_check_prices.append(None)
            price_diffs.append(None)
            notes.append("客户物料编号+采购单号未命中701+411合并表")
            unmatched += 1
        else:
            matched += 1
            selection = _jingwang_select_factory_candidate(candidates, row, customer_price)
            selected = selection["selected"]
            match_statuses.append("成功")
            compare_results.append(selection["compare_result"])
            merged_row_numbers.append(selection["row_numbers"])
            factory_sources.append(selected.get("来源") if selected else None)
            factory_prices.append(selected.get("单价") if selected else None)
            factory_check_prices.append(selection["factory_check_price"])
            price_diffs.append(selection["price_diff"])
            notes.append(selection["note"])
            if selection["compare_result"] == "价格一致":
                price_equal += 1
            else:
                price_mismatch += 1

        if job_id and (pos == 1 or pos % 50 == 0 or pos == len(customer_df)):
            append_job_log(
                job_id,
                f"景旺订单核价进度：{pos}/{len(customer_df)}",
                success_count=price_equal,
                fail_count=pos - price_equal,
                current_row=pos,
                total_rows=len(customer_df),
            )

    output["行号"] = merged_row_numbers
    output["厂内来源"] = factory_sources
    output["厂内单价"] = factory_prices
    output["匹配状态"] = match_statuses
    output["对比结果"] = compare_results
    output["厂内核对价格"] = factory_check_prices
    output["价格差额"] = price_diffs
    output["核价说明"] = notes

    summary = summary_sheet(
        [
            ("合并表行数", len(merged_df)),
            ("合并表匹配客户数量", merged_customer_matched),
            ("合并表未匹配客户数量", merged_customer_unmatched),
            ("合并表价格一致数量", merged_customer_price_equal),
            ("合并表价格不一致数量", merged_customer_price_mismatch),
            ("客户总记录数", len(output)),
            ("客户已匹配数量", matched),
            ("客户未匹配数量", unmatched),
            ("价格一致数量", price_equal),
            ("价格不一致数量", price_mismatch),
            ("异常数量", len(output) - price_equal),
        ]
    )

    return {
        "sheets": [
            ("客户核价结果", output),
            ("701+411合并表", checked_merged_df),
            ("核价汇总", summary),
        ],
        "success_count": price_equal,
        "fail_count": len(output) - price_equal,
        "skip_count": 0,
        "total_rows": len(output),
    }


def process_jingwang_price_check(
    customer_path: Path,
    factory_701_path: Path,
    factory_411_path: Path,
    *,
    job_id: int | None = None,
) -> dict:
    factory_701_df = read_business_sheet(
        factory_701_path,
        {"订单号/出货单号", "项次", "客户单号", "客户产品编号", "数量", "单位", "单价", "本币金额", "所属集团"},
    )
    factory_411_df = read_business_sheet(
        factory_411_path,
        {"单别单号", "项次", "客户单号", "客户产品编号", "订单数量", "单位", "单价", "总金额", "所属集团公司"},
    )
    customer_df = read_business_sheet(customer_path, {"采购单号", "物料编号", "价格"})
    if job_id:
        append_job_log(
            job_id,
            (
                f"已读取景旺文件：701 {len(factory_701_df)} 行，"
                f"411 {len(factory_411_df)} 行，客户数据 {len(customer_df)} 行"
            ),
            total_rows=len(customer_df),
        )

    filtered_701_df, filtered_411_df, filter_df, filter_stats = filter_jingwang_factory_rows(factory_701_df, factory_411_df)
    merged_df, covered_df = build_jingwang_factory_merge(filtered_701_df, filtered_411_df)
    result = build_jingwang_customer_price_check(customer_df, merged_df, job_id=job_id)
    summary = result["sheets"][-1][1]
    factory_summary = summary_sheet(
        [
            ("701原始行数", len(factory_701_df)),
            ("411原始行数", len(factory_411_df)),
            ("701过滤删除数", filter_stats["filtered_701"]),
            ("411过滤删除数", filter_stats["filtered_411"]),
            ("过滤后701行数", len(filtered_701_df)),
            ("过滤后411行数", len(filtered_411_df)),
            ("单价为0删除数", filter_stats["zero_price"]),
            ("集团不符合删除数", filter_stats["invalid_group"]),
            ("411被701覆盖数量", len(covered_df)),
            ("合并表行数", len(merged_df)),
        ]
    )
    summary = pd.concat([summary, factory_summary], ignore_index=True)
    result["sheets"] = [
        ("客户核价结果", result["sheets"][0][1]),
        ("701+411合并表", result["sheets"][1][1]),
        ("411被701覆盖明细", covered_df),
        ("厂内过滤明细", filter_df),
        ("核价汇总", summary),
    ]
    return result


def build_match_result(
    customer_df: pd.DataFrame,
    factory_df: pd.DataFrame,
    matches: dict[int, MatchResult],
    *,
    job_id: int | None = None,
) -> dict:
    customer_result = customer_df.copy()
    customer_statuses = []
    customer_factory_items = []
    factory_to_customer: dict[int, list[str]] = {}
    success = 0
    fail = 0

    for pos, (idx, _) in enumerate(customer_df.iterrows(), 1):
        match = matches.get(idx, MatchResult("未匹配", [], [], ""))
        customer_statuses.append(match.status)
        customer_factory_items.append(",".join(match.factory_items))
        if match.status in {"已匹配", "拆分匹配"}:
            success += 1
        else:
            fail += 1
        for fidx in match.factory_indexes:
            factory_to_customer.setdefault(fidx, []).append(_text(customer_df.loc[idx].get("项次")))
        if job_id and (pos == 1 or pos % 50 == 0 or pos == len(customer_df)):
            append_job_log(
                job_id,
                f"匹配进度：{pos}/{len(customer_df)}",
                success_count=success,
                fail_count=fail,
                current_row=pos,
                total_rows=len(customer_df),
            )

    customer_result["匹配状态"] = customer_statuses
    customer_result["厂内项次"] = customer_factory_items

    factory_result = factory_df.copy()
    factory_statuses = []
    factory_customer_items = []
    for idx in factory_df.index:
        linked = factory_to_customer.get(idx, [])
        factory_statuses.append("已匹配" if linked else "未匹配")
        factory_customer_items.append(",".join(linked))
    factory_result["匹配状态"] = factory_statuses
    factory_result["客户项次"] = factory_customer_items

    summary = summary_sheet(
        [
            ("客户总记录数", len(customer_result)),
            ("客户已匹配数量", success),
            ("客户未匹配数量", fail),
            ("厂内总记录数", len(factory_result)),
            ("厂内已匹配数量", factory_statuses.count("已匹配")),
            ("厂内未匹配数量", factory_statuses.count("未匹配")),
        ]
    )
    return {
        "sheets": [
            ("客户明细匹配结果", customer_result),
            ("厂内明细匹配结果", factory_result),
            ("匹配汇总", summary),
        ],
        "success_count": success,
        "fail_count": fail,
        "skip_count": 0,
        "total_rows": len(customer_result),
    }


def match_customer_to_factory(
    customer_df: pd.DataFrame,
    factory_df: pd.DataFrame,
    *,
    customer_cols: tuple[str, str, str],
    factory_cols: tuple[str, str, str],
) -> dict[int, MatchResult]:
    c_po, c_item, c_mat = customer_cols
    f_po, f_item, f_mat = factory_cols

    factory_exact: dict[tuple[str, str, str], list[int]] = {}
    factory_by_order_mat: dict[tuple[str, str], list[int]] = {}
    for fidx, row in factory_df.iterrows():
        key = (_norm(row.get(f_po)), _norm_item(row.get(f_item)), _norm(row.get(f_mat)))
        if not all(key):
            continue
        factory_exact.setdefault(key, []).append(fidx)
        factory_by_order_mat.setdefault((key[0], key[2]), []).append(fidx)

    customer_items_by_order_mat: dict[tuple[str, str], set[str]] = {}
    for _, row in customer_df.iterrows():
        po = _norm(row.get(c_po))
        item = _norm_item(row.get(c_item))
        mat = _norm(row.get(c_mat))
        if po and item and mat:
            customer_items_by_order_mat.setdefault((po, mat), set()).add(item)

    results: dict[int, MatchResult] = {}
    for cidx, row in customer_df.iterrows():
        po = _norm(row.get(c_po))
        item = _norm_item(row.get(c_item))
        mat = _norm(row.get(c_mat))
        if not po or not item or not mat:
            results[cidx] = MatchResult("未匹配", [], [], "关键字段为空")
            continue

        exact = factory_exact.get((po, item, mat), [])
        if exact:
            results[cidx] = MatchResult("已匹配", exact, [_norm_item(factory_df.loc[idx].get(f_item)) for idx in exact])
            continue

        candidates = []
        ambiguous_items = []
        for fidx in factory_by_order_mat.get((po, mat), []):
            factory_item = _norm_item(factory_df.loc[fidx].get(f_item))
            if not factory_item.startswith(item) or factory_item == item:
                continue
            possible_customer_items = [
                candidate
                for candidate in customer_items_by_order_mat.get((po, mat), set())
                if factory_item.startswith(candidate) and factory_item != candidate
            ]
            if len(possible_customer_items) > 1:
                ambiguous_items.append(factory_item)
            else:
                candidates.append(fidx)
        if ambiguous_items:
            results[cidx] = MatchResult("拆分项次歧义", [], [], f"厂内项次可能对应多个客户项次：{','.join(ambiguous_items)}")
        elif candidates:
            results[cidx] = MatchResult(
                "拆分匹配",
                candidates,
                [_norm_item(factory_df.loc[idx].get(f_item)) for idx in candidates],
            )
        else:
            results[cidx] = MatchResult("未匹配", [], [], "未找到对应厂内记录")
    return results


def load_quote_rows(paths: list[Path]) -> list[QuoteRow]:
    rows: list[QuoteRow] = []
    for path in paths:
        workbook = pd.ExcelFile(path)
        for sheet_name in workbook.sheet_names:
            raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=object)
            for header_idx in range(len(raw)):
                headers = [_clean_header(value) for value in raw.iloc[header_idx].tolist()]
                if _is_ccl_header(headers):
                    rows.extend(_parse_ccl_rows(raw, headers, header_idx, path.name, sheet_name))
                elif _is_pp_header(headers):
                    rows.extend(_parse_pp_rows(raw, headers, header_idx, path.name, sheet_name))
    return rows


def _parse_ccl_rows(raw, headers: list[str], header_idx: int, source_file: str, sheet_name: str) -> list[QuoteRow]:
    product_col = _find_col(headers, {"产品型号", "产品名称"})
    thickness_col = _find_col(headers, {"厚度mm", "厚度（mm）", "厚度"})
    copper_col = _find_col(headers, {"铜箔", "铜箔厚度"})
    foil_col = _find_col(headers, {"铜箔特性", "铜箔类型"})
    state_col = _find_col(headers, {"板材类型", "是否含铜"})
    stack_col = _find_col(headers, {"叠构", "组合结构"})
    sf_col = _find_col(headers, {"SF单价", "SF价格"})
    size_cols = {headers[idx].upper().replace("X", "*"): idx for idx, value in enumerate(headers) if re.fullmatch(r"\d+(?:\.\d+)?[*X]\d+(?:\.\d+)?", value.upper())}
    if product_col is None or thickness_col is None or sf_col is None:
        return []

    rows = []
    for row_idx in range(header_idx + 1, len(raw)):
        values = raw.iloc[row_idx].tolist()
        if _looks_like_header(values) or _row_is_empty(values):
            continue
        product = _norm_product(values[product_col] if product_col < len(values) else "")
        if not product:
            continue
        thickness_value = values[thickness_col] if thickness_col is not None and thickness_col < len(values) else None
        stack_value = values[stack_col] if stack_col is not None and stack_col < len(values) else None
        row_values = {
            "厚度": _number(thickness_value),
            "厚度候选": _numbers(thickness_value),
            "铜箔": _norm_copper(values[copper_col] if copper_col is not None and copper_col < len(values) else None),
            "铜箔特性": _norm_foil(values[foil_col] if foil_col is not None and foil_col < len(values) else None),
            "板材类型": _norm_state(values[state_col] if state_col is not None and state_col < len(values) else None),
            "板材类型候选": _norm_state_candidates(values[state_col] if state_col is not None and state_col < len(values) else None),
            "叠构": _norm_stack(stack_value),
            "叠构候选": _norm_stack_candidates(stack_value),
            "SF单价": _number(values[sf_col] if sf_col is not None and sf_col < len(values) else None),
        }
        price_map = {}
        for size_key, col_idx in size_cols.items():
            price = _number(values[col_idx] if col_idx < len(values) else None)
            if price is not None:
                price_map[_norm_size_key(size_key)] = price
        if row_values["SF单价"] is None and not price_map:
            continue
        rows.append(QuoteRow(source_file, sheet_name, row_idx + 1, product, "CCL", row_values, price_map))
    return rows


def _parse_pp_rows(raw, headers: list[str], header_idx: int, source_file: str, sheet_name: str) -> list[QuoteRow]:
    product_col = _find_col(headers, {"产品型号", "产品名称"})
    glass_col = _find_col(headers, {"布种"})
    rc_col = _find_col(headers, {"RC值", "树脂含量"})
    length_col = _find_col(headers, {"长度M", "长度(M)", "长度", "每卷长度", "RL"})
    price_col = _find_col(headers, {"RL价格", "每卷价格"})
    if price_col is None:
        price_col = _find_col(headers, {"单价M", "单价/M"})
    if product_col is None or glass_col is None or rc_col is None or price_col is None:
        return []

    rows = []
    for row_idx in range(header_idx + 1, len(raw)):
        values = raw.iloc[row_idx].tolist()
        if _looks_like_header(values) or _row_is_empty(values):
            continue
        product = _norm_product(values[product_col] if product_col < len(values) else "")
        glass = _norm(values[glass_col] if glass_col < len(values) else "")
        price = _number(values[price_col] if price_col < len(values) else None)
        if not product or not glass or price is None:
            continue
        rows.append(
            QuoteRow(
                source_file,
                sheet_name,
                row_idx + 1,
                product,
                "PP",
                {
                    "布种": _norm_glass(glass),
                    "RC": values[rc_col] if rc_col < len(values) else None,
                    "长度": _number(values[length_col] if length_col is not None and length_col < len(values) else None),
                    "价格": price,
                },
                {},
            )
        )
    return rows


def find_quote_price(spec: str, quote_rows: list[QuoteRow]) -> dict[str, Any]:
    parsed = parse_spec(spec)
    if parsed.product is None:
        return {"price": None, "note": "无法从规格提取产品型号"}
    if parsed.material_type == "PP" and parsed.length_m is None:
        return {"price": None, "note": "PP小片不输出结果", "status": "skipped"}
    product_aliases = _quote_product_aliases(parsed.product, parsed.material_type)
    candidates = [
        row
        for row in quote_rows
        if row.material_type == parsed.material_type and _norm_product(row.product) in product_aliases
    ]
    candidates.sort(key=lambda row: product_aliases.index(_norm_product(row.product)))
    if not candidates:
        note = f"报价单未找到产品型号 {parsed.product}"
        if len(product_aliases) > 1:
            note += f"（已尝试 {', '.join(product_aliases[1:])}）"
        return {"price": None, "note": note}
    if parsed.material_type == "PP":
        return _find_pp_quote(parsed, candidates)
    return _find_ccl_quote(parsed, candidates)


def _find_pp_quote(parsed: ParsedSpec, candidates: list[QuoteRow]) -> dict[str, Any]:
    for row in candidates:
        if parsed.glass and not _token_matches(row.values.get("布种"), parsed.glass):
            continue
        if parsed.rc is not None and not _rc_matches(row.values.get("RC"), parsed.rc):
            continue
        if parsed.length_m is not None and row.values.get("长度") is None:
            continue
        if parsed.length_m is not None and not _float_equal(parsed.length_m, row.values["长度"]):
            continue
        return {
            "price": row.values.get("价格"),
            "note": f"命中 {row.source_file}/{row.sheet_name} 第{row.excel_row}行",
        }
    return {"price": None, "note": "未命中PP报价：产品、布种、RC或长度不匹配"}


def _find_ccl_quote(parsed: ParsedSpec, candidates: list[QuoteRow]) -> dict[str, Any]:
    for row in candidates:
        if parsed.thickness is not None and not _number_choice_matches(row.values.get("厚度候选") or row.values.get("厚度"), parsed.thickness, 0.003):
            continue
        if parsed.copper and row.values.get("铜箔") and parsed.copper != row.values["铜箔"]:
            continue
        if parsed.foil and row.values.get("铜箔特性") and parsed.foil != row.values["铜箔特性"]:
            continue
        if parsed.copper_state and not _state_matches(row.values, parsed.copper_state):
            continue
        if parsed.stack and not _stack_matches(row.values, parsed.stack):
            continue
        price = _ccl_row_price(parsed, row)
        if price is not None:
            return {
                "price": price["price"],
                "note": f"命中 {row.source_file}/{row.sheet_name} 第{row.excel_row}行{price['note']}",
            }
    return {"price": None, "note": "未命中CCL报价：型号、厚度、铜箔、叠构或尺寸不匹配"}


def _ccl_row_price(parsed: ParsedSpec, row: QuoteRow) -> dict[str, Any] | None:
    price = row.size_prices.get(parsed.size_key or "") if parsed.size_key else None
    note = ""
    if price is None and parsed.size_key:
        split_match = _split_ccl_piece_price(parsed.size_key, row.size_prices)
        if split_match is not None:
            price = split_match["price"]
            note = f"，基板小片按{split_match['parent_size']} 1开{split_match['split_count']}折算"
    if price is None and row.values.get("SF单价") is not None:
        price = row.values["SF单价"]
    if price is None:
        return None
    return {"price": price, "note": note}


def _split_ccl_piece_price(size_key: str, size_prices: dict[str, float]) -> dict[str, Any] | None:
    child = _size_tuple(size_key)
    if child is None:
        return None
    options = []
    for parent_key, parent_price in size_prices.items():
        parent = _size_tuple(parent_key)
        if parent is None or parent_price is None:
            continue
        for parent_w, parent_h in (parent, (parent[1], parent[0])):
            split_count = _split_count(parent_w, parent_h, child[0], child[1])
            if split_count is not None and split_count > 1:
                options.append(
                    {
                        "price": parent_price / split_count,
                        "parent_size": parent_key,
                        "split_count": split_count,
                    }
                )
    if not options:
        return None
    return sorted(options, key=lambda item: item["split_count"])[0]


def _size_tuple(size_key: str) -> tuple[float, float] | None:
    parts = [_number(part) for part in re.split(r"[*Xx]", _text(size_key))]
    if len(parts) < 2 or parts[0] is None or parts[1] is None:
        return None
    return parts[0], parts[1]


def _split_count(parent_w: float, parent_h: float, child_w: float, child_h: float) -> int | None:
    if child_w <= 0 or child_h <= 0 or child_w - parent_w > 0.6 or child_h - parent_h > 0.6:
        return None
    cols = _near_int(parent_w / child_w)
    rows = _near_int(parent_h / child_h)
    if cols is None or rows is None:
        return None
    return cols * rows


def _near_int(value: float, tolerance: float = 0.08) -> int | None:
    rounded = round(value)
    if rounded < 1:
        return None
    return rounded if abs(value - rounded) <= tolerance else None


def parse_spec(spec: str) -> ParsedSpec:
    text = _text(spec).upper().replace("×", "*").replace("Ｘ", "*")
    product_match = re.search(r"(NY[-A-Z0-9()]+)\s*:", text)
    product = product_match.group(1) if product_match else None
    if not product:
        product_match = re.search(r"\b(NY[-A-Z0-9()]+P?)\b", text)
        product = product_match.group(1) if product_match else None
    material_type = "PP" if product and _product_is_pp(product) else "CCL"
    if material_type == "PP":
        glass_match = re.search(r":\s*([0-9]{2,4})\s*RC", text)
        rc_match = re.search(r"RC\s*=?\s*(\d+(?:\.\d+)?)\s*%?", text)
        length_match = re.search(r"(\d+(?:\.\d+)?)\s*M\s*/?\s*R", text)
        return ParsedSpec(
            material_type="PP",
            product=product,
            glass=_norm_glass(glass_match.group(1)) if glass_match else None,
            rc=float(rc_match.group(1)) if rc_match else None,
            length_m=float(length_match.group(1)) if length_match else None,
        )

    thickness_match = re.search(r"(\d+(?:\.\d+)?)\s*MM", text)
    copper_match = re.search(r"MM\s*([A-Z0-9.]+/[A-Z0-9.]+)", text)
    foil_match = re.search(r"\(([^)）]+)\)", text)
    size_match = re.search(r"(\d+(?:\.\d+)?)\s*[*X]\s*(\d+(?:\.\d+)?)", text)
    stack_match = re.search(r"\(([^()）]*\d{2,4}[^()）]*)\)\s*$", text)
    return ParsedSpec(
        material_type="CCL",
        product=product,
        thickness=float(thickness_match.group(1)) if thickness_match else None,
        copper=_norm_copper(copper_match.group(1)) if copper_match else None,
        foil=_norm_foil(foil_match.group(1)) if foil_match else None,
        copper_state="含铜" if "含铜" in text and "不含铜" not in text else "不含铜" if "不含铜" in text else None,
        stack=_norm_stack(stack_match.group(1)) if stack_match else None,
        size_key=_norm_size_key(f"{size_match.group(1)}*{size_match.group(2)}") if size_match else None,
    )


def calculate_actual_quantity(unit: Any, qty: Any, spec: str) -> float | None:
    quantity = _number(qty)
    if quantity is None:
        return None
    if _text(unit) == "卷":
        parsed = parse_spec(spec)
        if parsed.length_m is not None:
            return parsed.length_m * quantity
    return quantity


def _block3_factory_price(factory_row: pd.Series, customer_spec: Any) -> tuple[float | None, str, str]:
    factory_item = _norm_item(factory_row.get("项次"))
    unit_price = _number(factory_row.get("单价"))
    if unit_price is None:
        return None, "厂内单价为空", factory_item

    specs = [
        customer_spec,
        factory_row.get("规格"),
        factory_row.get("客户规格"),
        factory_row.get("品名"),
    ]
    if not any(_is_pp_spec(spec) for spec in specs):
        base_price = _block3_non_pp_price(factory_row)
        if base_price is None:
            return None, "非PP: 单价为空", factory_item
        return base_price, f"非PP: 单价{base_price:g}", factory_item

    length = _pp_length_from_specs(specs)
    if length is None:
        return unit_price, f"PP长度缺失，未换算: 单价{unit_price:g}", factory_item

    converted = unit_price * length
    return converted, f"PP: 单价{unit_price:g}×{length:g}={converted:g}", factory_item


def _pick_block3_price(price_parts: list[tuple[float | None, str, str]], is_split: bool) -> tuple[float | None, str]:
    notes = []
    for price, note, factory_item in price_parts:
        if price is not None:
            if is_split:
                prefix = f"拆分取用项次{factory_item}，未累加"
                return price, f"{prefix}; {note}" if note else prefix
            return price, note
        if note:
            notes.append(note)
    return None, "; ".join(notes)


def _sort_factory_indexes_by_item(factory_df: pd.DataFrame, factory_indexes: list[int]) -> list[int]:
    return sorted(factory_indexes, key=lambda idx: _item_sort_key(factory_df.loc[idx].get("项次")))


def _item_sort_key(value: Any) -> tuple[int, str]:
    text = _norm_item(value)
    number = _number(text)
    return (int(number) if number is not None else 10**9, text)


def _block3_non_pp_price(factory_row: pd.Series) -> float | None:
    return _number(factory_row.get("单价"))


def _jingwang_select_factory_candidate(
    candidates: list[dict[str, Any]],
    customer_row: pd.Series,
    customer_price: Decimal | None,
) -> dict[str, Any]:
    sorted_candidates = sorted(candidates, key=lambda record: _number(record.get("合并表行号")) or 10**9)
    row_numbers = ",".join(_text(record.get("合并表行号")) for record in sorted_candidates if _text(record.get("合并表行号")))
    evaluated: list[dict[str, Any]] = []

    for record in sorted_candidates:
        factory_price, _, price_note = _jingwang_factory_check_price(
            record,
            {"物料名称": customer_row.get("物料名称")},
        )
        diff = customer_price - factory_price if customer_price is not None and factory_price is not None else None
        evaluated.append(
            {
                "record": record,
                "factory_price": factory_price,
                "price_diff": diff,
                "is_equal": diff == Decimal("0.00") if diff is not None else False,
                "price_note": price_note,
            }
        )

    equal_candidate = next((item for item in evaluated if item["is_equal"]), None)
    if equal_candidate:
        selected = equal_candidate["record"]
        if len(evaluated) > 1:
            note = f"多条候选{row_numbers}，取第一条价格一致行{selected.get('合并表行号')}；{equal_candidate['price_note']}"
        else:
            note = f"{equal_candidate['price_note']}；价格两位小数一致"
        return {
            "selected": selected,
            "row_numbers": row_numbers,
            "compare_result": "价格一致",
            "factory_check_price": _decimal_to_float(equal_candidate["factory_price"]),
            "price_diff": _decimal_to_float(equal_candidate["price_diff"]),
            "note": note,
        }

    if len(evaluated) > 1:
        return {
            "selected": None,
            "row_numbers": row_numbers,
            "compare_result": "价格不一致",
            "factory_check_price": None,
            "price_diff": None,
            "note": f"多条候选{row_numbers}均价格不一致",
        }

    selected_item = evaluated[0]
    note = f"{selected_item['price_note']}；价格两位小数不一致"
    if customer_price is None:
        note = f"{selected_item['price_note']}；客户价格为空，无法核对"
    elif selected_item["factory_price"] is None:
        note = f"{selected_item['price_note']}；厂内核对价格为空，无法核对"
    return {
        "selected": selected_item["record"],
        "row_numbers": row_numbers,
        "compare_result": "价格不一致",
        "factory_check_price": _decimal_to_float(selected_item["factory_price"]),
        "price_diff": _decimal_to_float(selected_item["price_diff"]),
        "note": note,
    }


def _jingwang_factory_check_price(
    factory_record: dict[str, Any],
    customer_record: dict[str, Any] | None = None,
) -> tuple[Decimal | None, float | None, str]:
    unit_price = _number(factory_record.get("单价"))
    if unit_price is None:
        return None, None, "厂内单价为空"

    specs = [
        (customer_record or {}).get("物料名称"),
        factory_record.get("客户规格"),
        factory_record.get("规格"),
        factory_record.get("品名"),
    ]
    is_pp = any(_is_pp_spec(spec) for spec in specs)
    length = _pp_length_from_specs(specs) if is_pp else None
    if is_pp and length is not None:
        if _jingwang_is_single_width_pp_roll(specs):
            price = _money_2(unit_price / 2 * length)
            return price, length, f"PP单幅卷料: 单价{unit_price:g}÷2×{length:g}"
        price = _money_2(unit_price * length)
        return price, length, f"PP卷料: 单价{unit_price:g}×{length:g}"

    price = _money_2(unit_price)
    if is_pp:
        return price, None, f"PP小片/未提取米数: 单价{unit_price:g}"
    return price, None, f"非PP: 单价{unit_price:g}"


def _jingwang_is_single_width_pp_roll(specs: list[Any]) -> bool:
    for spec in specs:
        text = _text(spec).upper().replace("ＩＮ", "IN").replace("″", "IN").replace('"', "IN")
        if not text:
            continue
        normalized = re.sub(r"\s+", "", text)
        if re.search(r"24[.．]65(?:IN|英寸)", normalized):
            return True
        if re.search(r"24[.．]41(?:IN|英寸)", normalized):
            return True
    return False


def _is_pp_spec(spec: Any) -> bool:
    text = _text(spec).upper().replace("×", "*").replace("Ｘ", "*")
    if not text:
        return False
    product_match = re.search(r"\b(NY[-A-Z0-9()]+)\b", text)
    if product_match:
        product = _norm_product(product_match.group(1))
        if _product_is_pp(product):
            return True
    return "RC" in text and _pp_length_from_spec(text) is not None


def _product_is_pp(product: str) -> bool:
    normalized = _norm_product(product)
    core = re.sub(r"\([^)]*\)$", "", normalized)
    return core.endswith("P")


def _pp_length_from_specs(specs: list[Any]) -> float | None:
    for spec in specs:
        length = _pp_length_from_spec(spec)
        if length is not None:
            return length
    return None


def _pp_length_from_spec(spec: Any) -> float | None:
    text = _text(spec).upper().replace("×", "*").replace("Ｘ", "*")
    if not text:
        return None
    parsed = parse_spec(text)
    if parsed.length_m is not None:
        return parsed.length_m
    patterns = [
        r"(\d+(?:\.\d+)?)\s*M\s*/?\s*卷",
        r"卷\s*(\d+(?:\.\d+)?)\s*M",
        r"(\d+(?:\.\d+)?)\s*M\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return float(match.group(1))
    return None


def read_business_sheet(path: Path, required_columns: set[str]) -> pd.DataFrame:
    workbook = pd.ExcelFile(path)
    for sheet_name in workbook.sheet_names:
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=object)
        header_idx = find_header_row(raw, required_columns)
        if header_idx is None:
            continue
        df = pd.read_excel(path, sheet_name=sheet_name, header=header_idx, dtype=object)
        df = df.dropna(how="all").reset_index(drop=True)
        df.columns = [str(col).strip() for col in df.columns]
        return df
    raise ValueError(f"{path.name} 未找到包含字段 {', '.join(sorted(required_columns))} 的业务 Sheet")


def read_block2_customer_sheet(path: Path) -> pd.DataFrame:
    workbook = pd.ExcelFile(path)
    for sheet_name in workbook.sheet_names:
        raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=object)
        for header_idx in range(min(20, len(raw))):
            headers = [_clean_header(value) for value in raw.iloc[header_idx].tolist()]
            spec_col = _find_col(headers, {"规格", "客户规格", "物料规格"})
            if spec_col is None:
                continue
            df = pd.read_excel(path, sheet_name=sheet_name, header=header_idx, dtype=object)
            df = df.dropna(how="all").reset_index(drop=True)
            df.columns = [str(col).strip() for col in df.columns]
            renamed = _rename_block2_customer_columns(df)
            if "规格" not in renamed.columns:
                continue
            return renamed
    raise ValueError(f"{path.name} 未找到包含字段 规格 的业务 Sheet")


def _rename_block2_customer_columns(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "采购订单号": {"采购订单号", "采购单号", "订购单号", "订单号"},
        "料件编号": {"料件编号", "物料编号", "料号", "客户产品编号"},
        "规格": {"规格", "客户规格", "物料规格"},
        "采购量": {"采购量", "订单行数量", "数量", "订购数量"},
        "采购单位": {"采购单位", "请购单位", "单位"},
        "含税单价": {"含税单价", "含税价格", "含税价", "单价含税"},
    }
    rename_map = {}
    used_targets = set(df.columns)
    for column in df.columns:
        clean = _clean_header(column)
        for target, names in aliases.items():
            if target in used_targets and target != column:
                continue
            if clean in {_clean_header(name) for name in names}:
                rename_map[column] = target
                used_targets.add(target)
                break
    return df.rename(columns=rename_map)


def find_header_row(raw: pd.DataFrame, required_columns: set[str]) -> int | None:
    required = {_clean_header(col) for col in required_columns}
    for idx in range(min(20, len(raw))):
        headers = {_clean_header(value) for value in raw.iloc[idx].tolist() if _text(value)}
        if required.issubset(headers):
            return idx
    return None


def save_result_workbook(sheets: list[tuple[str, pd.DataFrame]], output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="E8F0FE")
    for sheet_name, df in sheets:
        ws = wb.create_sheet(_safe_sheet_name(sheet_name))
        if df is None or df.empty:
            ws.append(["说明"])
            ws.append(["无数据"])
        else:
            headers = [str(col) for col in df.columns]
            ws.append(headers)
            for row in df.itertuples(index=False):
                ws.append([_excel_value(value) for value in row])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        for col_idx, column in enumerate(ws.columns, 1):
            values = [str(cell.value) for cell in list(column)[:80] if cell.value is not None]
            width = min(max([len(value) for value in values] + [10]) + 2, 36)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(output_path)


def summary_sheet(items: list[tuple[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame(items, columns=["项目", "数量"])


def _save_upload(file_obj, target_dir: Path, prefix: str) -> Path:
    filename = file_obj.filename or f"{prefix}.xlsx"
    if not filename.lower().endswith(EXCEL_EXTENSIONS):
        raise ValueError(f"{filename} 不是支持的 Excel 文件")
    safe_name = secure_filename(filename) or f"{prefix}.xlsx"
    path = target_dir / f"{prefix}_{safe_name}"
    file_obj.save(path)
    return path


def _is_ccl_header(headers: list[str]) -> bool:
    joined = "|".join(headers)
    return ("产品型号" in joined or "产品名称" in joined) and "厚度" in joined and ("叠构" in joined or "组合结构" in joined)


def _is_pp_header(headers: list[str]) -> bool:
    joined = "|".join(headers)
    return ("产品型号" in joined or "产品名称" in joined) and "布种" in joined and ("RC值" in joined or "树脂含量" in joined)


def _find_col(headers: list[str], aliases: set[str]) -> int | None:
    clean_aliases = {_clean_header(alias) for alias in aliases}
    for idx, header in enumerate(headers):
        if header in clean_aliases:
            return idx
    return None


def _looks_like_header(values: list[Any]) -> bool:
    text = "|".join(_clean_header(value) for value in values)
    return "产品型号" in text or "产品名称" in text


def _row_is_empty(values: list[Any]) -> bool:
    return not any(_text(value) for value in values)


def _prices_equal(left: Any, right: Any) -> bool:
    left_num = _number(left)
    right_num = _number(right)
    if left_num is None or right_num is None:
        return False
    return round(left_num + 1e-9, 2) == round(right_num + 1e-9, 2)


def _price_diff(left: Any, right: Any) -> float | None:
    left_num = _number(left)
    right_num = _number(right)
    if left_num is None or right_num is None:
        return None
    return round(left_num - right_num, 2)


def _jingwang_customer_po_column(df: pd.DataFrame) -> str:
    po_columns = [str(col) for col in df.columns if str(col).startswith("采购单号")]
    if "采购单号.1" in po_columns:
        return "采购单号.1"
    if len(po_columns) >= 2:
        return po_columns[1]
    if po_columns:
        return po_columns[0]
    raise ValueError("客户数据未找到采购单号列")


def _money_2(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, Decimal):
        number = value
    elif isinstance(value, (int, float)):
        number = Decimal(str(value))
    else:
        text = _text(value).replace(",", "").replace("￥", "").replace("¥", "")
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            number = Decimal(match.group(0))
        except InvalidOperation:
            return None
    try:
        return number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return None


def _decimal_to_float(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _quote_product_aliases(product: str, material_type: str) -> list[str]:
    primary = _norm_product(product)
    aliases = [primary]
    if material_type == "PP" and primary.endswith("P") and len(primary) > 1:
        aliases.append(primary[:-1])
    return list(dict.fromkeys(alias for alias in aliases if alias))


def _token_matches(rule_value: Any, target: str) -> bool:
    normalized_target = _norm_glass(target)
    if not normalized_target:
        return True
    return normalized_target in _token_alternatives(rule_value, _norm_glass)


def _token_alternatives(value: Any, normalizer) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [
        normalizer(part)
        for part in re.split(r"[/,，、\s]+", text)
        if normalizer(part)
    ]


def _rc_matches(rule_value: Any, target: float) -> bool:
    text = _text(rule_value).replace("％", "%")
    if not text:
        return False
    numbers = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", text)]
    if not numbers:
        return False
    if len(numbers) >= 2 and ("-" in text or "~" in text or "～" in text):
        low = _to_percent_value(numbers[0])
        high = _to_percent_value(numbers[1])
        return min(low, high) <= target <= max(low, high)
    value = _to_percent_value(numbers[0])
    return abs(value - target) <= 0.01


def _to_percent_value(value: float) -> float:
    return value * 100 if abs(value) <= 1 else value


def _float_equal(left: float | None, right: float | None, tolerance: float = 0.01) -> bool:
    return left is not None and right is not None and abs(left - right) <= tolerance


def _number_choice_matches(rule_value: Any, target: float, tolerance: float = 0.01) -> bool:
    numbers = rule_value if isinstance(rule_value, list) else _numbers(rule_value)
    if not numbers:
        return True
    return any(abs(target - number) <= tolerance for number in numbers)


def _number(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def _numbers(value: Any) -> list[float]:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return []
    if isinstance(value, (int, float)):
        return [float(value)]
    text = _text(value).replace(",", "")
    return [float(item) for item in re.findall(r"-?\d+(?:\.\d+)?", text)]


def _text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _norm(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value)).upper()


def _norm_identifier(value: Any) -> str:
    text = _norm(value)
    return text[:-2] if re.fullmatch(r"\d+\.0", text) else text


def _norm_item(value: Any) -> str:
    text = _norm(value)
    return text[:-2] if re.fullmatch(r"\d+\.0", text) else text


def _norm_product(value: Any) -> str:
    return _norm(value).replace("（", "(").replace("）", ")")


def _norm_glass(value: Any) -> str:
    return _norm(value)


def _norm_copper(value: Any) -> str:
    # T/T is a distinct foil specification from 1/1; only sort the two sides
    # so the confirmed H/1 and 1/H notation remains interchangeable.
    text = _norm(value).replace(" ", "")
    parts = [part for part in text.split("/") if part]
    if len(parts) == 2:
        return "/".join(sorted(parts, key=_copper_sort_key))
    return text


def _copper_sort_key(value: str) -> tuple[int, float | str]:
    if value == "H":
        return (0, value)
    number = _number(value)
    if number is not None:
        return (1, number)
    return (2, value)


def _norm_foil(value: Any) -> str:
    text = _norm(value)
    if "/" in text:
        text = text.split("/")[0]
    return text


def _norm_state(value: Any) -> str:
    text = _text(value)
    if "不含铜" in text:
        return "不含铜"
    if "含铜" in text:
        return "含铜"
    return _norm(value)


def _norm_state_candidates(value: Any) -> list[str]:
    text = _text(value)
    candidates = []
    if "含铜" in text:
        candidates.append("含铜")
    if "不含铜" in text:
        candidates.append("不含铜")
    normalized = _norm(value)
    if normalized and not candidates:
        candidates.append(normalized)
    return list(dict.fromkeys(candidates))


def _norm_size_key(value: Any) -> str:
    numbers = [_number(part) for part in re.split(r"[*Xx]", _text(value).upper())]
    if len(numbers) < 2 or numbers[0] is None or numbers[1] is None:
        return _norm(value).replace("X", "*")
    return f"{_format_size(numbers[0])}*{_format_size(numbers[1])}"


def _format_size(value: float) -> str:
    rounded = round(value)
    return str(int(rounded)) if abs(value - rounded) < 0.51 else f"{value:g}"


def _norm_stack(value: Any) -> str:
    text = _norm(value).replace("X", "*").replace("×", "*")
    parts: dict[str, int] = {}
    for token in re.split(r"[+/,，、]", text):
        token = token.strip()
        if not token:
            continue
        match = re.fullmatch(r"(?:(\d+)\*)?(\d{2,4})", token)
        if match:
            count = int(match.group(1)) if match.group(1) else 1
            glass = match.group(2)
        else:
            match = re.fullmatch(r"(\d{2,4})\*(\d+)", token)
            if not match:
                continue
            glass = match.group(1)
            count = int(match.group(2))
        parts[glass] = parts.get(glass, 0) + count
    if not parts:
        return text
    return "+".join(f"{glass}*{parts[glass]}" for glass in sorted(parts))


def _norm_stack_candidates(value: Any) -> list[str]:
    text = _norm(value).replace("X", "*").replace("×", "*")
    if not text:
        return []
    plus_groups = [group for group in re.split(r"[+，、]", text) if group]
    candidates = [""]
    for group in plus_groups:
        options = [_norm_stack(option) for option in group.split("/") if option]
        options = [option for option in options if option]
        if not options:
            continue
        candidates = [
            _combine_stack_candidate(prefix, option)
            for prefix in candidates
            for option in options
        ]
    return list(dict.fromkeys(candidate for candidate in candidates if candidate))


def _combine_stack_candidate(left: str, right: str) -> str:
    if not left:
        return right
    return _norm_stack(f"{left}+{right}")


def _stack_matches(row_values: dict[str, Any], target: str) -> bool:
    candidates = row_values.get("叠构候选") or []
    if candidates:
        return target in candidates
    rule_stack = row_values.get("叠构")
    return not rule_stack or target == rule_stack


def _state_matches(row_values: dict[str, Any], target: str) -> bool:
    candidates = row_values.get("板材类型候选") or []
    if candidates:
        return target in candidates
    rule_state = row_values.get("板材类型")
    return not rule_state or target == rule_state


def _clean_header(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value)).replace("（", "(").replace("）", ")")


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    if hasattr(value, "to_pytimedelta"):
        return str(value)
    return value


def _safe_sheet_name(value: str) -> str:
    return re.sub(r"[\[\]\:\*\?\/\\]", "_", value)[:31]
