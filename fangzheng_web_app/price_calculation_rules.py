from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from .db import get_setting, set_setting
from .excel_utils import excel_format, load_workbook_compat, normalized_xlsx_source
from .paths import DEFAULT_RULES_DIR, PRICE_CALCULATION_RULES_DIR
from .price_calculation_customers import PRICE_CALCULATION_CUSTOMERS, enabled_price_customer


PRICE_RULE_FILENAME = "price_rules.xlsx"
TEST_DATA_FILENAME = "test_data.xlsx"
GUANGHE_HUANGSHI_RULE_FILENAME = "guanghe_huangshi_price_rules.xlsx"
GUANGHE_NANYA_RULE_FILENAME = "guanghe_nanya_price_rules.xlsx"
SUHANG_PP_RULE_FILENAME = "suhang_pp_price_rules.xlsx"
SUHANG_CCL_RULE_FILENAME = "suhang_ccl_price_rules.xlsx"
ALLOWED_RULE_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}
JINGWANG_QUOTE_VARIANTS = {"new": "新报价单", "old": "旧报价单"}


def normalize_price_quote_variant(customer_key: str, quote_variant: str | None = None) -> str:
    if customer_key != "jingwang":
        return ""
    variant = (quote_variant or "new").strip().lower()
    return variant if variant in JINGWANG_QUOTE_VARIANTS else "new"


def _customer_dir(customer_key: str, quote_variant: str | None = None) -> Path:
    variant = normalize_price_quote_variant(customer_key, quote_variant)
    if customer_key == "jingwang" and variant == "old":
        return PRICE_CALCULATION_RULES_DIR / customer_key / variant
    return PRICE_CALCULATION_RULES_DIR / customer_key


def _versions_dir(customer_key: str, quote_variant: str | None = None) -> Path:
    return _customer_dir(customer_key, quote_variant) / "versions"


def _active_key(customer_key: str, quote_variant: str | None = None) -> str:
    variant = normalize_price_quote_variant(customer_key, quote_variant)
    return f"active_price_rule_version:{customer_key}:{variant}" if variant == "old" else f"active_price_rule_version:{customer_key}"


def _history_key(customer_key: str, quote_variant: str | None = None) -> str:
    variant = normalize_price_quote_variant(customer_key, quote_variant)
    return f"price_rule_history:{customer_key}:{variant}" if variant == "old" else f"price_rule_history:{customer_key}"


def _read_history(customer_key: str, quote_variant: str | None = None) -> list[dict]:
    raw = get_setting(_history_key(customer_key, quote_variant), "[]") or "[]"
    try:
        data = json.loads(raw)
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        return []


def _write_history(customer_key: str, history: list[dict], quote_variant: str | None = None) -> None:
    set_setting(_history_key(customer_key, quote_variant), json.dumps(history[:50], ensure_ascii=False))


def get_price_rule_history(customer_key: str, quote_variant: str | None = None) -> list[dict]:
    enabled_price_customer(customer_key)
    return _read_history(customer_key, quote_variant)


def append_price_rule_history(customer_key: str, entry: dict, quote_variant: str | None = None) -> None:
    history = _read_history(customer_key, quote_variant)
    history.insert(0, entry)
    _write_history(customer_key, history, quote_variant)


def activate_price_rule_version(customer_key: str, version: str, quote_variant: str | None = None) -> str:
    enabled_price_customer(customer_key)
    variant = normalize_price_quote_variant(customer_key, quote_variant)
    rule_version, _history = _validated_history_version(customer_key, version, variant)
    rule_path = get_price_rule_file_path(customer_key, rule_version, variant)
    if not rule_path.is_file():
        raise ValueError("该规则版本的报价单文件不存在，无法启用")
    test_path = get_price_test_data_file_path(customer_key, rule_version, variant)
    validate_price_rule_files(customer_key, rule_path, test_path if test_path.exists() else None)
    set_setting(_active_key(customer_key, variant), rule_version)
    return rule_version


def delete_price_rule_version(customer_key: str, version: str, quote_variant: str | None = None) -> str:
    enabled_price_customer(customer_key)
    variant = normalize_price_quote_variant(customer_key, quote_variant)
    rule_version, history = _validated_history_version(customer_key, version, variant)
    active_version = get_setting(_active_key(customer_key, variant), "") or ""
    if rule_version == active_version:
        raise ValueError("当前生效版本不能删除，请先启用其他版本")

    versions_root = _versions_dir(customer_key, variant).resolve()
    version_dir = (versions_root / rule_version).resolve()
    if version_dir.parent != versions_root:
        raise ValueError("规则版本目录不合法")
    if version_dir.exists():
        if not version_dir.is_dir():
            raise ValueError("规则版本路径不是目录，无法删除")
        shutil.rmtree(version_dir)

    _write_history(
        customer_key,
        [item for item in history if str(item.get("version") or "") != rule_version],
        variant,
    )
    return rule_version


def _validated_history_version(
    customer_key: str,
    version: str,
    quote_variant: str | None = None,
) -> tuple[str, list[dict]]:
    rule_version = (version or "").strip()
    if not rule_version or Path(rule_version).name != rule_version or rule_version in {".", ".."}:
        raise ValueError("规则版本名称不合法")
    history = _read_history(customer_key, quote_variant)
    if not any(str(item.get("version") or "") == rule_version for item in history):
        raise ValueError("规则版本不存在或已被删除")
    return rule_version, history


def get_price_rule_version_dir(customer_key: str, version: str | None = None, quote_variant: str | None = None) -> Path:
    enabled_price_customer(customer_key)
    rule_version = version or (get_setting(_active_key(customer_key, quote_variant), "") or "")
    return _versions_dir(customer_key, quote_variant) / rule_version


def get_price_rule_file_path(customer_key: str, version: str | None = None, quote_variant: str | None = None) -> Path:
    return get_price_rule_version_dir(customer_key, version, quote_variant) / PRICE_RULE_FILENAME


def get_price_test_data_file_path(customer_key: str, version: str | None = None, quote_variant: str | None = None) -> Path:
    return get_price_rule_version_dir(customer_key, version, quote_variant) / TEST_DATA_FILENAME


def get_active_price_rule_version(customer_key: str, quote_variant: str | None = None) -> str:
    enabled_price_customer(customer_key)
    version = get_setting(_active_key(customer_key, quote_variant), "") or ""
    if version and get_price_rule_file_path(customer_key, version, quote_variant).exists():
        return version
    return ensure_default_price_rule_version(customer_key, quote_variant)


def ensure_default_price_rule_versions() -> None:
    for customer in PRICE_CALCULATION_CUSTOMERS:
        if customer.get("enabled"):
            ensure_default_price_rule_version(customer["key"])


def ensure_default_price_rule_version(customer_key: str, quote_variant: str | None = None) -> str:
    customer = enabled_price_customer(customer_key)
    variant = normalize_price_quote_variant(customer_key, quote_variant)
    active_version = get_setting(_active_key(customer_key, variant), "") or ""
    if active_version and get_price_rule_file_path(customer_key, active_version, variant).exists():
        return active_version
    if customer_key == "jingwang" and variant == "old":
        return ""

    packaged_dir = DEFAULT_RULES_DIR / "price_calculation"
    if customer_key == "suhang":
        seed_pp = packaged_dir / "suhang" / "pp_price_rules.xlsx"
        seed_ccl = packaged_dir / "suhang" / "ccl_price_rules.xlsx"
        seed_test = packaged_dir / "suhang" / TEST_DATA_FILENAME
        if not seed_pp.exists() or not seed_ccl.exists():
            return ""
        version = datetime.now().strftime("suhang_bootstrap_%Y%m%d_%H%M%S")
        version_dir = _versions_dir(customer_key, variant) / version
        version_dir.mkdir(parents=True, exist_ok=True)
        _copy_price_rule(seed_pp, version_dir / SUHANG_PP_RULE_FILENAME)
        _copy_price_rule(seed_ccl, version_dir / SUHANG_CCL_RULE_FILENAME)
        _merge_excel_sources(
            [("苏杭PP报价", version_dir / SUHANG_PP_RULE_FILENAME), ("苏杭CCL报价", version_dir / SUHANG_CCL_RULE_FILENAME)],
            version_dir / PRICE_RULE_FILENAME,
        )
        if seed_test.exists():
            _copy_excel_as_xlsx(seed_test, version_dir / TEST_DATA_FILENAME)
        validate_price_rule_files(customer_key, version_dir / PRICE_RULE_FILENAME, version_dir / TEST_DATA_FILENAME)
        set_setting(_active_key(customer_key, variant), version)
        append_price_rule_history(
            customer_key,
            {
                "version": version,
                "customer_key": customer_key,
                "customer_label": customer["label"],
                "quote_variant": variant,
                "quote_variant_label": JINGWANG_QUOTE_VARIANTS.get(variant, ""),
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "updated_by": "system",
                "remark": "由内置苏杭PP和CCL报价表及测试数据初始化",
                "rule_file": f"{seed_pp.name}；{seed_ccl.name}",
                "test_file": seed_test.name,
            },
            variant,
        )
        return version

    seed_map = {
        "jingwang": (packaged_dir / "jingwang" / "new" / PRICE_RULE_FILENAME, packaged_dir / "jingwang" / "new" / TEST_DATA_FILENAME),
        "plin": (packaged_dir / "plin" / PRICE_RULE_FILENAME, packaged_dir / "plin" / TEST_DATA_FILENAME),
        "hanyu": (packaged_dir / "hanyu" / PRICE_RULE_FILENAME, packaged_dir / "hanyu" / TEST_DATA_FILENAME),
        "wutong": (packaged_dir / "wutong" / PRICE_RULE_FILENAME, packaged_dir / "wutong" / TEST_DATA_FILENAME),
        "eaton": (packaged_dir / "eaton" / PRICE_RULE_FILENAME, packaged_dir / "eaton" / TEST_DATA_FILENAME),
        "taixing": (packaged_dir / "taixing" / PRICE_RULE_FILENAME, packaged_dir / "taixing" / TEST_DATA_FILENAME),
        "aoshikang": (packaged_dir / "aoshikang" / PRICE_RULE_FILENAME, packaged_dir / "aoshikang" / TEST_DATA_FILENAME),
        "mingyang": (packaged_dir / "mingyang" / PRICE_RULE_FILENAME, packaged_dir / "mingyang" / TEST_DATA_FILENAME),
        "lejian": (packaged_dir / "lejian" / PRICE_RULE_FILENAME, packaged_dir / "lejian" / TEST_DATA_FILENAME),
        "guanghe": (packaged_dir / "guanghe" / PRICE_RULE_FILENAME, packaged_dir / "guanghe" / TEST_DATA_FILENAME),
        "shengyi": (packaged_dir / "shengyi" / PRICE_RULE_FILENAME, packaged_dir / "shengyi" / TEST_DATA_FILENAME),
        "guigu": (packaged_dir / "guigu" / PRICE_RULE_FILENAME, packaged_dir / "guigu" / TEST_DATA_FILENAME),
        "techuang": (packaged_dir / "techuang" / PRICE_RULE_FILENAME, packaged_dir / "techuang" / TEST_DATA_FILENAME),
        "zhongfu": (packaged_dir / "zhongfu" / PRICE_RULE_FILENAME, packaged_dir / "zhongfu" / TEST_DATA_FILENAME),
        "huaxingyu": (packaged_dir / "huaxingyu" / PRICE_RULE_FILENAME, packaged_dir / "huaxingyu" / TEST_DATA_FILENAME),
        "dongxun": (packaged_dir / "dongxun" / PRICE_RULE_FILENAME, packaged_dir / "dongxun" / TEST_DATA_FILENAME),
        "yingchuangli": (packaged_dir / "yingchuangli" / "price_rules.xls", packaged_dir / "yingchuangli" / "test_data.xls"),
        "zhongjing": (packaged_dir / "zhongjing" / PRICE_RULE_FILENAME, packaged_dir / "zhongjing" / TEST_DATA_FILENAME),
    }
    seed_files = seed_map.get(customer_key)
    if not seed_files:
        return ""

    seed_price = seed_files[0] if isinstance(seed_files[0], Path) else DEFAULT_RULES_DIR / seed_files[0]
    seed_test = seed_files[1] if isinstance(seed_files[1], Path) else DEFAULT_RULES_DIR / seed_files[1]
    if not seed_price.exists():
        return ""

    version_prefix = f"{customer_key}_{variant}_bootstrap" if variant else f"{customer_key}_bootstrap"
    version = datetime.now().strftime(f"{version_prefix}_%Y%m%d_%H%M%S")
    version_dir = _versions_dir(customer_key, variant) / version
    version_dir.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(seed_price, version_dir / PRICE_RULE_FILENAME)
        if seed_test.exists():
            _copy_excel_as_xlsx(seed_test, version_dir / TEST_DATA_FILENAME)
        validate_price_rule_files(customer_key, version_dir / PRICE_RULE_FILENAME, version_dir / TEST_DATA_FILENAME)
    except Exception:
        shutil.rmtree(version_dir, ignore_errors=True)
        raise

    set_setting(_active_key(customer_key, variant), version)
    append_price_rule_history(
        customer_key,
        {
            "version": version,
            "customer_key": customer_key,
            "customer_label": customer["label"],
            "quote_variant": variant,
            "quote_variant_label": JINGWANG_QUOTE_VARIANTS.get(variant, ""),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_by": "system",
            "remark": f"由内置{customer['label']}报价表和测试数据初始化",
            "rule_file": seed_price.name,
            "test_file": seed_test.name,
        },
        variant,
    )
    return version


def save_new_suhang_rule_version(
    pp_file: FileStorage | None = None,
    ccl_file: FileStorage | None = None,
    *,
    updated_by: str,
    remark: str,
) -> str:
    customer = enabled_price_customer("suhang")
    if not (pp_file and pp_file.filename) and not (ccl_file and ccl_file.filename):
        raise ValueError("请至少上传苏杭PP报价单或苏杭CCL报价单其中一份 Excel")

    for label, storage in (("苏杭PP报价单", pp_file), ("苏杭CCL报价单", ccl_file)):
        if storage and storage.filename:
            source_name = secure_filename(storage.filename) or PRICE_RULE_FILENAME
            if Path(source_name).suffix.lower() not in ALLOWED_RULE_EXTENSIONS:
                raise ValueError(f"{label}仅支持 .xlsx / .xls / .xlsm 文件")

    ensure_default_price_rule_version("suhang")
    version = datetime.now().strftime("suhang_rules_%Y%m%d_%H%M%S_%f")
    version_dir = _versions_dir("suhang") / version
    version_dir.mkdir(parents=True, exist_ok=True)

    uploaded_names: list[str] = []
    if pp_file and pp_file.filename:
        uploaded_name = pp_file.filename or SUHANG_PP_RULE_FILENAME
        uploaded_names.append(f"PP:{uploaded_name}")
        uploaded_path = version_dir / (secure_filename(uploaded_name) or SUHANG_PP_RULE_FILENAME)
        pp_file.save(uploaded_path)
        _copy_price_rule(uploaded_path, version_dir / SUHANG_PP_RULE_FILENAME)
    else:
        source_name = _copy_existing_suhang_component(SUHANG_PP_RULE_FILENAME, version_dir / SUHANG_PP_RULE_FILENAME)
        uploaded_names.append(f"PP沿用:{source_name}")

    if ccl_file and ccl_file.filename:
        uploaded_name = ccl_file.filename or SUHANG_CCL_RULE_FILENAME
        uploaded_names.append(f"CCL:{uploaded_name}")
        uploaded_path = version_dir / (secure_filename(uploaded_name) or SUHANG_CCL_RULE_FILENAME)
        ccl_file.save(uploaded_path)
        _copy_price_rule(uploaded_path, version_dir / SUHANG_CCL_RULE_FILENAME)
    else:
        source_name = _copy_existing_suhang_component(SUHANG_CCL_RULE_FILENAME, version_dir / SUHANG_CCL_RULE_FILENAME)
        uploaded_names.append(f"CCL沿用:{source_name}")

    _merge_excel_sources(
        [("苏杭PP报价", version_dir / SUHANG_PP_RULE_FILENAME), ("苏杭CCL报价", version_dir / SUHANG_CCL_RULE_FILENAME)],
        version_dir / PRICE_RULE_FILENAME,
    )
    test_name = _copy_existing_test_data("suhang", version_dir / TEST_DATA_FILENAME)
    validate_price_rule_files("suhang", version_dir / PRICE_RULE_FILENAME, version_dir / TEST_DATA_FILENAME)

    set_setting(_active_key("suhang"), version)
    append_price_rule_history(
        "suhang",
        {
            "version": version,
            "customer_key": "suhang",
            "customer_label": customer["label"],
            "quote_variant": "",
            "quote_variant_label": "",
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_by": updated_by,
            "remark": remark or "网页上传苏杭PP/CCL报价单并生效",
            "rule_file": "；".join(uploaded_names),
            "test_file": test_name,
        },
    )
    return version


def save_new_price_rule_version(
    customer_key: str,
    rule_file: FileStorage,
    test_file: FileStorage | None = None,
    *,
    updated_by: str,
    remark: str,
    quote_variant: str | None = None,
) -> str:
    customer = enabled_price_customer(customer_key)
    variant = normalize_price_quote_variant(customer_key, quote_variant)
    rule_name = secure_filename(rule_file.filename or PRICE_RULE_FILENAME) or PRICE_RULE_FILENAME
    if Path(rule_name).suffix.lower() not in ALLOWED_RULE_EXTENSIONS:
        raise ValueError("报价表仅支持 .xlsx / .xls / .xlsm 文件")
    test_name = ""
    if test_file and test_file.filename:
        test_name = secure_filename(test_file.filename or TEST_DATA_FILENAME) or TEST_DATA_FILENAME
        if Path(test_name).suffix.lower() not in ALLOWED_RULE_EXTENSIONS:
            raise ValueError("测试数据仅支持 .xlsx / .xls / .xlsm 文件")

    version_prefix = f"{customer_key}_{variant}_rules" if variant else f"{customer_key}_rules"
    version = datetime.now().strftime(f"{version_prefix}_%Y%m%d_%H%M%S")
    version_dir = _versions_dir(customer_key, variant) / version
    version_dir.mkdir(parents=True, exist_ok=True)

    uploaded_rule = version_dir / rule_name
    rule_file.save(uploaded_rule)

    _copy_price_rule(uploaded_rule, version_dir / PRICE_RULE_FILENAME)
    if test_file and test_file.filename:
        uploaded_test = version_dir / test_name
        test_file.save(uploaded_test)
        _copy_excel_as_xlsx(uploaded_test, version_dir / TEST_DATA_FILENAME)
    else:
        test_name = _copy_existing_test_data(customer_key, version_dir / TEST_DATA_FILENAME, variant)
    validate_price_rule_files(customer_key, version_dir / PRICE_RULE_FILENAME, version_dir / TEST_DATA_FILENAME)

    set_setting(_active_key(customer_key, variant), version)
    append_price_rule_history(
        customer_key,
        {
            "version": version,
            "customer_key": customer_key,
            "customer_label": customer["label"],
            "quote_variant": variant,
            "quote_variant_label": JINGWANG_QUOTE_VARIANTS.get(variant, ""),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_by": updated_by,
            "remark": remark or f"网页上传{customer['label']}价格计算规则",
            "rule_file": rule_file.filename or rule_name,
            "test_file": test_name,
        },
        variant,
    )
    return version


def save_new_guanghe_rule_version(
    huangshi_file: FileStorage,
    nanya_file: FileStorage,
    test_file: FileStorage | None = None,
    *,
    updated_by: str,
    remark: str,
) -> str:
    customer = enabled_price_customer("guanghe")
    source_files = [
        ("黄石广合单价", huangshi_file, GUANGHE_HUANGSHI_RULE_FILENAME),
        ("南亚新材价格更新", nanya_file, GUANGHE_NANYA_RULE_FILENAME),
    ]
    for label, storage, _target_name in source_files:
        if not storage or not storage.filename:
            raise ValueError(f"请上传{label} Excel")
        source_name = secure_filename(storage.filename) or _target_name
        if Path(source_name).suffix.lower() not in ALLOWED_RULE_EXTENSIONS:
            raise ValueError(f"{label}仅支持 .xlsx / .xls / .xlsm 文件")
    test_name = ""
    if test_file and test_file.filename:
        test_name = secure_filename(test_file.filename or TEST_DATA_FILENAME) or TEST_DATA_FILENAME
        if Path(test_name).suffix.lower() not in ALLOWED_RULE_EXTENSIONS:
            raise ValueError("测试数据仅支持 .xlsx / .xls / .xlsm 文件")

    version = datetime.now().strftime("guanghe_rules_%Y%m%d_%H%M%S")
    version_dir = _versions_dir("guanghe") / version
    version_dir.mkdir(parents=True, exist_ok=True)

    saved_sources: list[tuple[str, Path, str]] = []
    uploaded_names: list[str] = []
    for label, storage, target_name in source_files:
        uploaded_name = storage.filename or target_name
        uploaded_names.append(uploaded_name)
        target_path = version_dir / target_name
        storage.save(target_path)
        validate_price_rule_files("guanghe", target_path)
        saved_sources.append((label, target_path, uploaded_name))

    _merge_guanghe_rule_sources([(label, path) for label, path, _name in saved_sources], version_dir / PRICE_RULE_FILENAME)
    if test_file and test_file.filename:
        uploaded_test = version_dir / test_name
        test_file.save(uploaded_test)
        _copy_excel_as_xlsx(uploaded_test, version_dir / TEST_DATA_FILENAME)
    else:
        test_name = _copy_existing_test_data("guanghe", version_dir / TEST_DATA_FILENAME)
    validate_price_rule_files("guanghe", version_dir / PRICE_RULE_FILENAME, version_dir / TEST_DATA_FILENAME)

    set_setting(_active_key("guanghe"), version)
    append_price_rule_history(
        "guanghe",
        {
            "version": version,
            "customer_key": "guanghe",
            "customer_label": customer["label"],
            "quote_variant": "",
            "quote_variant_label": "",
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_by": updated_by,
            "remark": remark or "网页上传广合两份报价单并合并生效",
            "rule_file": "；".join(uploaded_names),
            "test_file": test_name,
        },
    )
    return version


def validate_price_rule_files(customer_key: str, rule_path: str | Path, test_data_path: str | Path | None = None) -> None:
    enabled_price_customer(customer_key)
    detected = excel_format(rule_path)
    if detected not in {"xlsx_zip", "xls_ole"}:
        raise ValueError("报价表必须是 .xlsx / .xls / .xlsm 格式")
    if customer_key == "plin":
        load_workbook_compat(rule_path, data_only=True)
        _validate_plin_rule_file(rule_path)
        return
    if customer_key in {"hanyu", "wutong", "eaton", "taixing", "aoshikang", "mingyang", "guanghe", "shengyi", "guigu", "techuang", "zhongfu", "huaxingyu", "dongxun", "suhang", "yingchuangli", "zhongjing"}:
        load_workbook_compat(rule_path, data_only=True)
        return
    if customer_key == "lejian":
        # 乐健历史报价单存在非标准样式 XML，计算逻辑会按单元格 XML 值兜底读取。
        return
    if customer_key == "jingwang" and detected == "xls_ole":
        load_workbook_compat(rule_path, data_only=True)
    if not test_data_path or not Path(test_data_path).exists():
        return
    workbook = load_workbook_compat(test_data_path, data_only=False)
    found_spec = False
    for ws in workbook.worksheets:
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        if "客户规格" in headers:
            found_spec = True
            break
    if not found_spec:
        raise ValueError("测试数据缺少“客户规格”列")


def _validate_plin_rule_file(rule_path: str | Path) -> None:
    workbook = load_workbook_compat(rule_path, data_only=True)
    has_ccl = has_pp = False
    for ws in workbook.worksheets:
        title = ws.title.upper()
        for row_idx in range(1, min(ws.max_row, 30) + 1):
            headers = {str(cell.value).strip() for cell in ws[row_idx] if cell.value is not None}
            if {"产品类别", "厚度mm", "铜厚", "铜箔类型", "组合叠构"}.issubset(headers):
                has_ccl = True
            if {"Products", "Glass type", "Resin Content", "Length (m)", "Per M"}.issubset(headers):
                has_pp = True
        if title.endswith("PP") or "PP" in title:
            has_pp = has_pp or False
        if "基板" in ws.title:
            has_ccl = has_ccl or False
    if not has_ccl or not has_pp:
        raise ValueError("普林报价表必须包含可识别的基板 Sheet 和 PP Sheet")


def _copy_existing_test_data(customer_key: str, target: Path, quote_variant: str | None = None) -> str:
    candidates: list[Path] = []
    active_version = get_setting(_active_key(customer_key, quote_variant), "") or ""
    if active_version:
        candidates.append(_versions_dir(customer_key, quote_variant) / active_version / TEST_DATA_FILENAME)
    packaged_dir = DEFAULT_RULES_DIR / "price_calculation"
    default_test_map = {
        "jingwang": packaged_dir / "jingwang" / "new" / TEST_DATA_FILENAME,
        "plin": packaged_dir / "plin" / TEST_DATA_FILENAME,
        "hanyu": packaged_dir / "hanyu" / TEST_DATA_FILENAME,
        "wutong": packaged_dir / "wutong" / TEST_DATA_FILENAME,
        "eaton": packaged_dir / "eaton" / TEST_DATA_FILENAME,
        "taixing": packaged_dir / "taixing" / TEST_DATA_FILENAME,
        "aoshikang": packaged_dir / "aoshikang" / TEST_DATA_FILENAME,
        "mingyang": packaged_dir / "mingyang" / TEST_DATA_FILENAME,
        "lejian": packaged_dir / "lejian" / TEST_DATA_FILENAME,
        "guanghe": packaged_dir / "guanghe" / TEST_DATA_FILENAME,
        "shengyi": packaged_dir / "shengyi" / TEST_DATA_FILENAME,
        "guigu": packaged_dir / "guigu" / TEST_DATA_FILENAME,
        "techuang": packaged_dir / "techuang" / TEST_DATA_FILENAME,
        "zhongfu": packaged_dir / "zhongfu" / TEST_DATA_FILENAME,
        "huaxingyu": packaged_dir / "huaxingyu" / TEST_DATA_FILENAME,
        "dongxun": packaged_dir / "dongxun" / TEST_DATA_FILENAME,
        "suhang": packaged_dir / "suhang" / TEST_DATA_FILENAME,
        "yingchuangli": packaged_dir / "yingchuangli" / "test_data.xls",
        "zhongjing": packaged_dir / "zhongjing" / TEST_DATA_FILENAME,
    }
    if customer_key in default_test_map:
        default_test = default_test_map[customer_key]
        candidates.append(default_test if isinstance(default_test, Path) else DEFAULT_RULES_DIR / default_test)

    for source in candidates:
        if not source.exists():
            continue
        if excel_format(source) == "xlsx_zip":
            shutil.copy2(source, target)
        else:
            _copy_excel_as_xlsx(source, target)
        return source.name
    return ""


def _copy_price_rule(source: Path, target: Path) -> None:
    detected = excel_format(source)
    if detected == "xlsx_zip":
        shutil.copy2(source, target)
        return
    _copy_excel_as_xlsx(source, target)


def _copy_excel_as_xlsx(source: Path, target: Path) -> None:
    workbook = load_workbook_compat(source, data_only=False)
    normalized = normalized_xlsx_source(source, workbook)
    if normalized == target:
        return
    if normalized.suffix.lower() == ".xlsx":
        shutil.copy2(normalized, target)
    else:
        workbook.save(target)


def _merge_guanghe_rule_sources(sources: list[tuple[str, Path]], target: Path) -> None:
    _merge_excel_sources(sources, target)


def _merge_excel_sources(sources: list[tuple[str, Path]], target: Path) -> None:
    from openpyxl import Workbook

    merged = Workbook()
    merged.remove(merged.active)
    used_titles: set[str] = set()
    for _label, source in sources:
        workbook = load_workbook_compat(source, data_only=True)
        for worksheet in workbook.worksheets:
            target_sheet = merged.create_sheet(_unique_sheet_title(worksheet.title, used_titles))
            for row in worksheet.iter_rows():
                target_sheet.append([cell.value for cell in row])
    if not merged.worksheets:
        raise ValueError("报价单未读取到任何 Sheet")
    merged.save(target)


def _copy_existing_suhang_component(filename: str, target: Path) -> str:
    active_version = get_setting(_active_key("suhang"), "") or ""
    candidates: list[Path] = []
    if active_version:
        candidates.append(_versions_dir("suhang") / active_version / filename)
    default_name = "pp_price_rules.xlsx" if filename == SUHANG_PP_RULE_FILENAME else "ccl_price_rules.xlsx"
    candidates.append(DEFAULT_RULES_DIR / "price_calculation" / "suhang" / default_name)
    for source in candidates:
        if source.exists():
            if source.resolve() == target.resolve():
                return source.name
            _copy_price_rule(source, target)
            return source.name
    raise ValueError(f"缺少苏杭规则文件：{filename}")


def _unique_sheet_title(title: str, used_titles: set[str]) -> str:
    base = (title or "Sheet").strip()[:31] or "Sheet"
    candidate = base
    index = 2
    while candidate in used_titles:
        suffix = f"_{index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(candidate)
    return candidate
