from __future__ import annotations

import importlib
import sys
import traceback
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.utils import secure_filename

from .db import append_job_log, create_job, prune_jobs_for_employee, update_job_status
from .excel_utils import load_workbook_compat, normalized_xlsx_source
from .file_utils import safe_unlink
from .job_control import launch_job_process
from .paths import JOBS_DIR
from .transcode_rules import get_active_transcode_rule_version, get_transcode_rule_file_path


TRANSCODE_MODULE_NAME = "fangzheng_web_app.transcode_engine"
NON_DATA_SPECS = {"物料描述", "规格", "客户规格", "材料描述", "物料规格"}


def load_transcode_module():
    """Reload the packaged transcode engine so jobs follow latest logic."""
    if TRANSCODE_MODULE_NAME in sys.modules:
        return importlib.reload(sys.modules[TRANSCODE_MODULE_NAME])
    return importlib.import_module(TRANSCODE_MODULE_NAME)


def is_effective_spec(value, engine) -> bool:
    text = engine._clean_cell(value)
    return bool(text and text.lower() not in {"nan", "none"} and text not in NON_DATA_SPECS)


def queue_transcode_job(employee_id: str, uploaded_file, source_filename: str) -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)

    safe_filename = secure_filename(source_filename) or f"transcode_{timestamp}.xlsx"
    input_path = employee_dir / f"{timestamp}_transcode_{safe_filename}"
    uploaded_file.save(input_path)

    rule_version = get_active_transcode_rule_version()
    job_id = create_job(
        employee_id,
        source_filename,
        str(input_path),
        rule_version,
        feature="transcode",
    )
    launch_job_process(job_id, "transcode", employee_id)
    return job_id


def calculate_transcode_quote(spec: str, *, customer: str = "", customer_code: str = "", order_text: str = "") -> dict:
    spec = str(spec or "").strip()
    if not spec:
        return {"status": "失败", "result": None, "error": "请输入客户规格"}

    engine = load_transcode_module()
    if engine.is_pp_or_rc_spec(spec):
        return {
            "status": "跳过",
            "result": "",
            "note": "PP/RC/% 规格按批量转码逻辑跳过",
            "material_type": "转码",
            "rule_version": get_active_transcode_rule_version(),
            "error": "",
        }

    rule_version = get_active_transcode_rule_version()
    rule_path = get_transcode_rule_file_path(rule_version)
    tables = engine.build_lookup_tables(engine.load_rule_sheets(str(rule_path)))
    code, steps, err = engine.transcode_row(
        engine._clean_cell(spec),
        str(order_text or "").strip(),
        str(customer or "").strip(),
        str(customer_code or "").strip(),
        tables,
        str(order_text or "").strip(),
    )
    explanation = _format_transcode_steps(steps)
    if err:
        return {
            "status": "失败",
            "result": code or "",
            "note": explanation,
            "material_type": "转码",
            "rule_version": rule_version,
            "error": err,
        }
    return {
        "status": "成功",
        "result": code,
        "note": explanation or "转码成功",
        "material_type": "转码",
        "rule_version": rule_version,
        "error": "",
    }


def run_transcode_job(job_id: int, employee_id: str) -> None:
    from .db import get_job

    update_job_status(job_id, status="running", log_text="")
    job = get_job(job_id)
    if not job:
        return

    append_job_log(job_id, f"开始转码任务，规则版本：{job['rule_version']}")

    try:
        engine = load_transcode_module()
        rule_path = get_transcode_rule_file_path(job["rule_version"])
        workbook = load_workbook_compat(job["stored_input_path"], data_only=True)
        source_for_result = normalized_xlsx_source(job["stored_input_path"], workbook)
        append_job_log(job_id, f"转码引擎已加载：{Path(engine.__file__).name}")
        append_job_log(job_id, f"规则文件：{rule_path.name}")

        sheets, tables = engine.load_transcode_inputs(str(source_for_result), str(rule_path))
        df_req = sheets["转码需求表"].copy()
        spec_col = engine.select_transcode_spec_column(df_req)
        customer_col = engine.detect_customer_column(df_req, spec_col)
        customer_code_col = engine.detect_customer_code_column(df_req)
        context_cols = engine.detect_transcode_context_columns(df_req, spec_col, customer_col, customer_code_col)
        df_req, result_col = engine.ensure_result_column(df_req)
        append_job_log(job_id, f"识别规格列：第 {spec_col + 1} 列；结果写入：第 {result_col + 1} 列")

        data_indices = [
            i for i in range(1, len(df_req))
            if is_effective_spec(df_req.iloc[i, spec_col], engine)
        ]
        total_rows = len(data_indices)
        update_job_status(job_id, status="running", total_rows=total_rows)
        append_job_log(job_id, f"检测到 {total_rows} 行有效规格数据", total_rows=total_rows)

        results: list[dict] = []
        success_count = 0
        fail_count = 0
        skip_count = 0
        for processed, i in enumerate(data_indices, start=1):
            row = df_req.iloc[i]
            a_val = (
                engine._clean_cell(row.iloc[customer_code_col])
                if customer_code_col is not None and len(row) > customer_code_col
                else ""
            )
            d_val = (
                str(row.iloc[customer_col]).strip()
                if customer_col is not None and len(row) > customer_col and pd.notna(row.iloc[customer_col])
                else ""
            )
            e_val = str(row.iloc[4]).strip() if len(row) > 4 and pd.notna(row.iloc[4]) else ""
            s_val = engine._clean_cell(row.iloc[spec_col])
            context_val = engine.build_context_text_from_row(row, context_cols)
            cust_spec_val = engine._clean_cell(row.iloc[6]) if len(row) > 6 else ""
            normalized_spec_val = engine._clean_cell(row.iloc[7]) if len(row) > 7 else ""
            pp_check_text = " ".join([s_val, cust_spec_val, normalized_spec_val])
            excel_row = i + 1

            if engine.is_pp_or_rc_spec(pp_check_text):
                skip_count += 1
                df_req.iloc[i, result_col] = ""
                append_job_log(
                    job_id,
                    f"第 {excel_row} 行跳过：PP/RC/%",
                    skip_count=skip_count,
                    current_row=processed,
                    total_rows=total_rows,
                )
                continue

            code, steps, err = engine.transcode_row(s_val, e_val, d_val, a_val, tables, context_val)
            explanation = _format_transcode_steps(steps)
            if err:
                fail_count += 1
                df_req.iloc[i, result_col] = f"未识别：{err}"
                results.append(
                    {
                        "行号": excel_row,
                        "客户": d_val,
                        "规格": s_val,
                        "编码": "",
                        "状态": "失败",
                        "错误": err,
                        "说明": explanation,
                    }
                )
                append_job_log(
                    job_id,
                    f"第 {excel_row} 行失败：{err}",
                    fail_count=fail_count,
                    current_row=processed,
                    total_rows=total_rows,
                )
            else:
                success_count += 1
                df_req.iloc[i, result_col] = code
                results.append(
                    {
                        "行号": excel_row,
                        "客户": d_val,
                        "规格": s_val,
                        "编码": code,
                        "状态": "成功",
                        "错误": "",
                        "说明": explanation,
                    }
                )
                append_job_log(
                    job_id,
                    f"第 {excel_row} 行成功：{code}",
                    success_count=success_count,
                    current_row=processed,
                    total_rows=total_rows,
                )

        input_path = Path(job["stored_input_path"])
        output_path = input_path.with_name(f"{input_path.stem}_转码结果.xlsx")
        _save_transcode_result(str(source_for_result), str(output_path), df_req, result_col, results)
        append_job_log(job_id, "结果文件已生成，任务完成", current_row=total_rows, total_rows=total_rows)

        update_job_status(
            job_id,
            status="completed",
            stored_result_path=str(output_path),
            success_count=success_count,
            fail_count=fail_count,
            skip_count=skip_count,
            current_row=total_rows,
            total_rows=total_rows,
            completed=True,
        )
    except Exception as exc:
        append_job_log(job_id, f"任务失败：{exc}")
        update_job_status(
            job_id,
            status="failed",
            error_message=f"{exc}\n{traceback.format_exc(limit=8)}",
            completed=True,
        )
    finally:
        stale_jobs = prune_jobs_for_employee(employee_id, keep_limit=500)
        for stale in stale_jobs:
            for key in ["stored_input_path", "stored_result_path"]:
                safe_unlink(stale[key])


def _format_transcode_steps(steps: dict) -> str:
    if not steps:
        return ""
    parts = [
        f"胶系={steps.get('glue_model', '')}->{steps.get('step1_glue_code', '')}",
        f"厚度={steps.get('thickness_raw', '')}->{steps.get('step2_thick_code', '')}",
        f"铜箔规格={steps.get('copper_spec_raw', '')}->{steps.get('step3_copper_code', '')}",
        f"尺寸={steps.get('size_w', '')}x{steps.get('size_h', '')}->{steps.get('step4_size_code', '')}",
        f"胶水类别={steps.get('glue_category', '')}->{steps.get('step5_glue_cat_code', '')}",
        f"铜箔类型={steps.get('step6_copper_type_code', '')}",
        f"基板级别={steps.get('step7_grade_code', '')}",
        f"总/芯厚={steps.get('order_type', '')}->{steps.get('step8_tc_code', '')}",
        f"结构={steps.get('step9_struct_code', '')}",
    ]
    errors = steps.get("errors") or []
    if steps.get("customer_order_rule"):
        parts.append(f"客户下单转换={steps.get('customer_order_rule')}")
    if steps.get("grade_note"):
        parts.append(steps.get("grade_note"))
    if steps.get("structured_special_rules"):
        parts.append(f"结构化特殊规则={'; '.join(steps.get('structured_special_rules'))}")
    if steps.get("structured_special_rule_conflicts"):
        parts.append(f"结构化规则冲突={'; '.join(steps.get('structured_special_rule_conflicts'))}")
    if errors:
        parts.append(f"未命中原因={'; '.join(errors)}")
    return " | ".join(parts)


def _save_transcode_result(source_path: str, output_path: str, df_req, result_col: int, results: list[dict]) -> None:
    wb = openpyxl.load_workbook(source_path)
    ws = wb["转码需求表"] if "转码需求表" in wb.sheetnames else wb[wb.sheetnames[0]]

    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    for i in range(1, len(df_req)):
        val = df_req.iloc[i, result_col]
        if pd.isna(val) or str(val).strip() in ("", "nan"):
            continue
        cell = ws.cell(row=i + 1, column=result_col + 1)
        cell.value = str(val)
        cell.fill = red_fill if str(val).startswith("未识别：") else green_fill

    if "转码说明" in wb.sheetnames:
        del wb["转码说明"]
    ws_note = wb.create_sheet("转码说明")

    headers = ["行号", "客户", "规格", "输出编码", "状态", "错误原因", "命中说明"]
    for col, header in enumerate(headers, 1):
        cell = ws_note.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(results, 2):
        ws_note.cell(row=row_idx, column=1, value=item.get("行号"))
        ws_note.cell(row=row_idx, column=2, value=str(item.get("客户", ""))[:80])
        ws_note.cell(row=row_idx, column=3, value=str(item.get("规格", ""))[:160])
        ws_note.cell(row=row_idx, column=4, value=item.get("编码", ""))
        ws_note.cell(row=row_idx, column=5, value=item.get("状态", ""))
        ws_note.cell(row=row_idx, column=6, value=str(item.get("错误", ""))[:200])
        ws_note.cell(row=row_idx, column=7, value=str(item.get("说明", ""))[:500])

        status_cell = ws_note.cell(row=row_idx, column=5)
        if item.get("状态") == "成功":
            status_cell.fill = green_fill
        elif item.get("状态") == "失败":
            status_cell.fill = red_fill

    widths = {"A": 8, "B": 18, "C": 70, "D": 34, "E": 10, "F": 40, "G": 100}
    for col, width in widths.items():
        ws_note.column_dimensions[col].width = width

    wb.save(output_path)
