from __future__ import annotations

import re
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .paths import STORAGE_DIR


SPECIAL_RULES_DIR = STORAGE_DIR / "rules" / "transcode_special"
SPECIAL_RULES_FILENAME = "structured_special_rules.xlsx"
SPECIAL_RULES_PATH = SPECIAL_RULES_DIR / SPECIAL_RULES_FILENAME
SPECIAL_RULES_SETTINGS_PATH = SPECIAL_RULES_DIR / "settings.json"
LATEST_ORIGINAL_IMPORT_PATH = SPECIAL_RULES_DIR / "latest_customer_special_requirements_original.xlsx"

STRUCTURED_HEADERS = [
    "规则ID", "启用", "来源行号", "客户代码", "客户简称", "物料类别", "规则大类", "是否参与转码", "优先级",
    "适用胶系", "厚度条件", "铜厚条件", "尺寸条件", "关键词条件", "排除关键词", "含不含铜条件", "CTI条件",
    "覆盖胶系代码", "覆盖厚度代码", "覆盖铜厚代码", "覆盖尺寸代码", "覆盖胶水类别", "覆盖铜箔类型", "覆盖基板级别",
    "覆盖总芯厚", "覆盖结构码", "非编码动作", "备注模板", "原始整段特殊要求", "原始特殊需求", "规则解释", "待确认",
    "匹配来源字段", "目标动作类型", "目标动作内容", "执行策略",
    "规则状态", "导入方式", "规则版本", "更新人", "更新时间",
]

RULE_TYPES = [
    "转码改写-胶系", "转码改写-厚度", "转码改写-铜厚", "转码改写-尺寸", "转码改写-胶水类别",
    "转码改写-铜箔类型", "转码改写-基板级别", "转码改写-总芯厚", "转码改写-结构码",
    "下单维护", "备注标签", "包装出货", "价格商务", "料号品号动作", "交期动作", "生产要求", "PP规则", "人工确认",
]

MATERIAL_TYPES = ["基板", "PP", "基板/PP", "全部", "其他"]
YES_NO = ["是", "否"]
MATCH_SOURCE_TYPES = ["客户规格", "备注", "订单备注", "客户物料描述", "客户特殊要求", "整行上下文", "备注/整行上下文"]
ACTION_TYPES = ["编码改写", "料号品号动作", "备注标签", "交期动作", "生产要求", "PP规则", "包装出货", "下单维护", "价格商务", "人工确认"]
EXECUTION_STRATEGIES = ["参与转码", "仅保存不转码", "人工确认"]
RULE_STATUSES = ["启用", "草稿", "归档"]
IMPORT_MODES = ["新增", "追加", "替换"]
TRANSCODE_FIELDS = [
    ("胶系", "覆盖胶系代码"),
    ("厚度", "覆盖厚度代码"),
    ("铜厚", "覆盖铜厚代码"),
    ("尺寸", "覆盖尺寸代码"),
    ("胶水类别", "覆盖胶水类别"),
    ("铜箔类型", "覆盖铜箔类型"),
    ("基板级别", "覆盖基板级别"),
    ("总/芯厚", "覆盖总芯厚"),
    ("结构码", "覆盖结构码"),
]
NON_TRANSCODE_TYPES = ["料号品号动作", "备注标签", "交期动作", "生产要求", "PP规则", "包装出货", "下单维护", "价格商务"]


def get_structured_special_rules_path() -> Path:
    return SPECIAL_RULES_PATH


def get_latest_original_import_path() -> Path:
    return LATEST_ORIGINAL_IMPORT_PATH


def get_structured_special_rule_settings() -> dict[str, Any]:
    default = {
        "structured_rules_enabled": False,
        "latest_original_filename": "",
        "latest_import_time": "",
    }
    if not SPECIAL_RULES_SETTINGS_PATH.exists():
        return default
    try:
        data = json.loads(SPECIAL_RULES_SETTINGS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return default
    default.update({key: data.get(key, value) for key, value in default.items()})
    return default


def is_structured_special_rules_enabled() -> bool:
    return bool(get_structured_special_rule_settings().get("structured_rules_enabled"))


def set_structured_special_rules_enabled(enabled: bool) -> dict[str, Any]:
    SPECIAL_RULES_DIR.mkdir(parents=True, exist_ok=True)
    settings = get_structured_special_rule_settings()
    settings["structured_rules_enabled"] = bool(enabled)
    settings["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    SPECIAL_RULES_SETTINGS_PATH.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
    return settings


def save_latest_original_import(file_bytes: bytes, original_filename: str = "") -> Path:
    SPECIAL_RULES_DIR.mkdir(parents=True, exist_ok=True)
    LATEST_ORIGINAL_IMPORT_PATH.write_bytes(file_bytes)
    settings = get_structured_special_rule_settings()
    settings["latest_original_filename"] = original_filename or "客户特殊要求原文件.xlsx"
    settings["latest_import_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    SPECIAL_RULES_SETTINGS_PATH.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
    return LATEST_ORIGINAL_IMPORT_PATH


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _blank_rule() -> dict[str, str]:
    return {header: "" for header in STRUCTURED_HEADERS}


def _base_rule(customer_code: str, customer_name: str, material_type: str,
               source_line: str, requirement_text: str, original_full_text: str = "") -> dict[str, str]:
    rule = _blank_rule()
    rule.update({
        "启用": "是",
        "来源行号": normalize_text(source_line),
        "客户代码": normalize_text(customer_code),
        "客户简称": normalize_text(customer_name),
        "物料类别": normalize_text(material_type) or "全部",
        "优先级": "50",
        "原始整段特殊要求": normalize_text(original_full_text) or requirement_text,
        "原始特殊需求": requirement_text,
        "待确认": "否",
        "匹配来源字段": "整行上下文",
        "执行策略": "仅保存不转码",
        "规则状态": "启用",
    })
    return rule


def _has_any(text: str, keywords: list[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def _split_requirement_items(text: str) -> list[str]:
    normalized = normalize_text(text)
    if not normalized:
        return []
    parts = re.split(r"(?:^|\s)(?:\d{1,2})[\.、．]\s*", normalized)
    items = [part.strip() for part in parts if part and part.strip()]
    return items or [normalized]


def _rule_with_execution(base: dict[str, str], *, participate: bool, action_type: str,
                         strategy: str | None = None) -> dict[str, str]:
    rule = dict(base)
    rule["是否参与转码"] = "是" if participate else "否"
    rule["目标动作类型"] = action_type
    rule["执行策略"] = strategy or ("参与转码" if participate else "仅保存不转码")
    return rule


def _manual_confirm_rule(base: dict[str, str], text: str, reason: str) -> dict[str, str]:
    rule = _rule_with_execution(base, participate=False, action_type="人工确认", strategy="人工确认")
    rule.update({
        "规则大类": "人工确认",
        "原始特殊需求": text,
        "待确认": "是",
        "目标动作内容": text,
        "规则解释": reason,
    })
    return rule


def _split_keywords(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[/,，、|；;]+", value or "") if item.strip()]


def _parse_hu_aoshikang_rule(base: dict[str, str], text: str) -> list[dict[str, str]]:
    rules: list[dict[str, str]] = []
    cust_code = normalize_text(base.get("客户代码"))
    cust_name = normalize_text(base.get("客户简称"))
    is_hu_aoshikang = cust_code == "105007" or "奥士康" in cust_name

    if ("A2TS" in text.upper()) or ("偏公差" in text):
        return [_manual_confirm_rule(base, text, "存在 A2TS/偏公差组合，无法稳定拆解到编码位段，需人工补充逻辑。")]

    car_keywords = ["X0A0", "X0B0", "car", "CAR", "汽车板", "458", "439"]
    if is_hu_aoshikang and _has_any(text, car_keywords):
        rule = _rule_with_execution(base, participate=True, action_type="编码改写")
        rule.update({
            "规则大类": "转码改写-基板级别",
            "优先级": "95",
            "关键词条件": "X0A0/X0B0/car/汽车板/458/439",
            "匹配来源字段": "备注/整行上下文",
            "覆盖基板级别": "AC",
            "目标动作内容": "下汽车板料号",
            "规则解释": "湖奥士康备注或上下文命中汽车板关键词时，基板级别覆盖为 AC。",
        })
        rules.append(rule)

    if "MINILED" in text.upper() or "MINIED" in text.upper():
        rule_grade = _rule_with_execution(base, participate=True, action_type="编码改写")
        rule_grade.update({
            "规则大类": "转码改写-基板级别",
            "优先级": "95",
            "关键词条件": "MINILED/MINIED",
            "匹配来源字段": "整行上下文",
            "覆盖基板级别": "AM",
            "目标动作内容": "下MINILED品号；下AM等级",
            "规则解释": "命中 MINILED/MINIED 时，基板级别覆盖为 AM。",
        })
        rules.append(rule_grade)
        rule_part = _rule_with_execution(base, participate=False, action_type="料号品号动作")
        rule_part.update({
            "规则大类": "料号品号动作",
            "关键词条件": "MINILED/MINIED",
            "目标动作内容": "下MINILED品号",
            "非编码动作": "下MINILED品号",
            "规则解释": "品号动作先沉淀，不直接参与当前品名编码。",
        })
        rules.append(rule_part)

    if re.search(r"R\s*/\s*R", text, re.IGNORECASE) and re.search(r"F\s*/\s*F|1\.5\s*/\s*1\.5", text, re.IGNORECASE):
        rule = _rule_with_execution(base, participate=True, action_type="编码改写")
        rule.update({
            "规则大类": "转码改写-铜厚",
            "优先级": "95",
            "铜厚条件": "R/R",
            "覆盖铜厚代码": "FF",
            "目标动作内容": "R/R按F/F，即1.5/1.5",
            "规则解释": "客户规格 R/R 对应我司 F/F，铜厚代码覆盖为 FF。",
        })
        rules.append(rule)

    if "NY2140L" in text.upper() and "A2" in text.upper():
        rule = _rule_with_execution(base, participate=True, action_type="编码改写")
        rule.update({
            "规则大类": "转码改写-基板级别",
            "优先级": "95",
            "适用胶系": "NY2140L",
            "覆盖基板级别": "A2",
            "目标动作内容": "NY2140L下A2级别",
            "规则解释": "湖奥士康 NY2140L 基板级别覆盖为 A2。",
        })
        rules.append(rule)

    non_code_patterns = [
        ("688料号", ["688", "华为", "HW"], "下688料号", "料号品号动作"),
        ("RC料号", ["NY-A1", "RC料号"], "NY-A1下RC料号", "料号品号动作"),
        ("耐CAF备注", ["耐CAF", "CAF"], "客户特殊要求栏备注耐CAF", "备注标签"),
        ("水印要求", ["水印"], "有水印要求需特殊留意", "备注标签"),
        ("T标志", ['"T"', "T标志", "T字母"], "标签及外包装需要注明T字母", "备注标签"),
        ("200米卷品号", ["3170M-2116", "200米/卷"], "3170M-2116的PP下200米/卷品号", "PP规则"),
        ("300米卷品号", ["300米/卷"], "其他胶系下300米/卷", "PP规则"),
        ("H10长秒规格", ["H10", "长秒"], "2150-1080 PP下H10长秒规格", "PP规则"),
        ("A1结构", ["3150HF", "A1结构"], "3150HF全部下A1结构", "PP规则"),
        ("交期动作", ["交期", "江西出货", "上海出货", "系统回交期", "指定交期"], text, "交期动作"),
        ("生产要求", ["华硕样品", "长春铜箔"], "订单备注华硕样品时，基板指定长春铜箔生产", "生产要求"),
    ]
    for label, keywords, action, action_type in non_code_patterns:
        if _has_any(text, keywords):
            rule = _rule_with_execution(base, participate=False, action_type=action_type)
            rule.update({
                "规则大类": action_type if action_type in RULE_TYPES else "下单维护",
                "关键词条件": "/".join(keywords),
                "目标动作内容": action,
                "非编码动作": action,
                "规则解释": f"{label}先结构化保存，不直接参与当前品名编码。",
            })
            if action_type == "PP规则":
                rule["物料类别"] = "PP"
            rules.append(rule)

    return rules


def _parse_size_rules(base: dict[str, str], text: str) -> list[dict[str, str]]:
    rules: list[dict[str, str]] = []
    pattern = re.compile(
        r"(\d+(?:\.\d+)?)\s*[*×Xx]\s*(\d+(?:\.\d+)?)"
        r"(?:\s*(?:下单|生产尺寸|下单生产尺寸|下单尺寸|厂内尺寸|尺寸))?"
        r"[^0-9]{0,12}(\d{8})"
    )
    seen: set[tuple[str, str]] = set()
    for match in pattern.finditer(text):
        size = f"{match.group(1)}*{match.group(2)}"
        code = match.group(3)
        key = (size, code)
        if key in seen:
            continue
        seen.add(key)
        rule = dict(base)
        rule.update({
            "规则大类": "转码改写-尺寸",
            "是否参与转码": "是",
            "优先级": "90",
            "尺寸条件": size,
            "覆盖尺寸代码": code,
            "目标动作类型": "编码改写",
            "执行策略": "参与转码",
            "规则解释": f"客户尺寸 {size} 时，转码尺寸段覆盖为 {code}。",
        })
        rules.append(rule)
    return rules


def _parse_thickness_mode_rules(base: dict[str, str], text: str) -> list[dict[str, str]]:
    rules: list[dict[str, str]] = []

    inclusive_core = re.search(
        r"(\d+(?:\.\d+)?)\s*(?:\(?含\)?|含)?\s*以下[^。；;，,]*?(?:芯厚|不含铜)",
        text,
    )
    above_total = re.search(
        r"(\d+(?:\.\d+)?)\s*(?:以上|以[上後后])[^。；;，,]*?(?:总厚|含铜)",
        text,
    )
    below_core = re.search(
        r"(\d+(?:\.\d+)?)\s*以下[^。；;，,]*?(?:芯厚|不含铜)",
        text,
    )
    at_above_total = re.search(
        r"(\d+(?:\.\d+)?)\s*(?:含)?\s*以上[^。；;，,]*?(?:总厚|含铜)",
        text,
    )

    core_match = inclusive_core or below_core
    total_match = above_total or at_above_total
    if core_match:
        op = "<=" if "含" in core_match.group(0) else "<"
        rule = dict(base)
        rule.update({
            "规则大类": "转码改写-总芯厚",
            "是否参与转码": "是",
            "优先级": "90",
            "厚度条件": f"{op}{core_match.group(1)}mm",
            "覆盖总芯厚": "C",
            "目标动作类型": "编码改写",
            "执行策略": "参与转码",
            "规则解释": f"厚度 {op}{core_match.group(1)}mm 时按芯厚/不含铜处理。",
            "待确认": "是",
        })
        rules.append(rule)
    if total_match:
        op = ">=" if "含" in total_match.group(0) else ">"
        rule = dict(base)
        rule.update({
            "规则大类": "转码改写-总芯厚",
            "是否参与转码": "是",
            "优先级": "90",
            "厚度条件": f"{op}{total_match.group(1)}mm",
            "覆盖总芯厚": "T",
            "目标动作类型": "编码改写",
            "执行策略": "参与转码",
            "规则解释": f"厚度 {op}{total_match.group(1)}mm 时按总厚/含铜处理。",
            "待确认": "是",
        })
        rules.append(rule)
    return rules


def _parse_non_transcode_rule(base: dict[str, str], text: str) -> dict[str, str] | None:
    order_keywords = ["SF", "面积", "数量", "小片", "PPM", "M数", "单位", "单价"]
    note_keywords = ["备注", "销货单", "合同", "标签", "COC", "报告", "二维码"]
    package_keywords = ["包装", "外箱", "打托", "托盘", "送货", "收货", "出货"]
    price_keywords = ["价格", "返利", "未税", "不含税", "一单一议"]

    rule_type = ""
    if _has_any(text, order_keywords):
        rule_type = "下单维护"
    elif _has_any(text, note_keywords):
        rule_type = "备注标签"
    elif _has_any(text, package_keywords):
        rule_type = "包装出货"
    elif _has_any(text, price_keywords):
        rule_type = "价格商务"
    if not rule_type:
        return None

    rule = dict(base)
    rule.update({
        "规则大类": rule_type,
        "是否参与转码": "否",
        "目标动作类型": rule_type,
        "执行策略": "仅保存不转码",
        "非编码动作": _summarize_non_code_action(text),
        "备注模板": _extract_note_template(text),
        "规则解释": "该规则不直接改品名编码，先作为非转码特殊需求归档。",
    })
    return rule


def _summarize_non_code_action(text: str) -> str:
    size_terms = re.findall(r"\d+(?:\.\d+)?\s*[*×Xx]\s*\d+(?:\.\d+)?\s*\([^)]*\)", text)
    ppm = "PPM数*13.12=面积" if "PPM" in text and "13.12" in text else ""
    parts = size_terms[:6]
    if ppm:
        parts.append(ppm)
    if parts:
        return "；".join(parts)
    return text[:120]


def _extract_note_template(text: str) -> str:
    if "面积" in text and "&" in text:
        return "面积:{面积}&"
    quote_match = re.search(r"[“\"]([^”\"]+)[”\"]", text)
    return quote_match.group(1) if quote_match else ""


def parse_special_requirement(customer_code: str, customer_name: str, material_type: str,
                              requirement_text: str, source_line: str = "") -> list[dict[str, str]]:
    text = normalize_text(requirement_text)
    if not text:
        return []
    rules: list[dict[str, str]] = []

    for item_idx, item in enumerate(_split_requirement_items(text), start=1):
        item_source_line = f"{source_line}.{item_idx}" if source_line and len(_split_requirement_items(text)) > 1 else source_line
        base = _base_rule(customer_code, customer_name, material_type, item_source_line, item, original_full_text=text)
        item_rules: list[dict[str, str]] = []

        item_rules.extend(_parse_hu_aoshikang_rule(base, item))
        if not any(rule.get("规则大类") == "人工确认" for rule in item_rules):
            item_rules.extend(_parse_size_rules(base, item))
            item_rules.extend(_parse_thickness_mode_rules(base, item))

            non_transcode = _parse_non_transcode_rule(base, item)
            if non_transcode and not item_rules:
                item_rules.insert(0, non_transcode)

        if not item_rules:
            item_rules.append(_manual_confirm_rule(base, item, "暂无法用确定性规则结构化，请人工判断是否影响转码。"))
        rules.extend(item_rules)
    return rules


def parse_bulk_special_requirement_workbook(file_bytes: bytes) -> tuple[list[dict[str, str]], dict[str, int]]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    return parse_bulk_special_requirement_workbook_object(wb)


def parse_bulk_special_requirement_workbook_object(wb, progress_callback=None) -> tuple[list[dict[str, str]], dict[str, int]]:
    all_rules: list[dict[str, str]] = []
    stats = {
        "sheets": 0,
        "source_rows": 0,
        "customers": 0,
        "rules": 0,
        "skipped_rows": 0,
    }
    seen_customers: set[str] = set()
    sheet_configs = []
    total_rows = 0

    for ws in wb.worksheets:
        stats["sheets"] += 1
        preview_rows = list(ws.iter_rows(max_row=5, values_only=True))
        if not preview_rows:
            continue
        header_idx, col_map = _detect_special_requirement_columns(preview_rows)
        first_data_row = header_idx + 2
        sheet_total = max(0, (ws.max_row or 0) - first_data_row + 1)
        total_rows += sheet_total
        sheet_configs.append((ws, first_data_row, col_map, sheet_total))
    stats["total_rows"] = total_rows

    processed = 0
    for ws, first_data_row, col_map, _sheet_total in sheet_configs:
        for excel_row_idx, row in enumerate(ws.iter_rows(min_row=first_data_row, values_only=True), start=first_data_row):
            processed += 1
            customer_code = normalize_text(_row_value(row, col_map.get("customer_code", 0)))
            customer_name = normalize_text(_row_value(row, col_map.get("customer_name", 1)))
            requirement_text = normalize_text(_row_value(row, col_map.get("requirement_text", 6)))
            if not customer_code or not _looks_like_requirement_text(requirement_text):
                if requirement_text or customer_name:
                    stats["skipped_rows"] += 1
                if progress_callback:
                    progress_callback(processed, total_rows, stats)
                continue
            stats["source_rows"] += 1
            seen_customers.add(customer_code)
            source_line = f"{ws.title}!{excel_row_idx}"
            parsed = parse_special_requirement(
                customer_code=customer_code,
                customer_name=customer_name,
                material_type="全部",
                requirement_text=requirement_text,
                source_line=source_line,
            )
            all_rules.extend(parsed)
            stats["rules"] = len(all_rules)
            if progress_callback:
                progress_callback(processed, total_rows, stats)

    stats["customers"] = len(seen_customers)
    stats["rules"] = len(all_rules)
    if progress_callback:
        progress_callback(processed, total_rows, stats)
    return all_rules, stats


def _row_value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if 0 <= index < len(row) else ""


def _detect_special_requirement_columns(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    best_idx = 0
    best_score = -1
    best_map: dict[str, int] = {}
    for idx, row in enumerate(rows[:5]):
        headers = [normalize_text(value) for value in row]
        col_map = _map_special_requirement_headers(headers)
        if "requirement_text" not in col_map:
            inferred = _infer_requirement_column(rows[idx + 1:])
            if inferred is not None:
                col_map["requirement_text"] = inferred
        score = len(col_map)
        if score > best_score:
            best_idx = idx
            best_score = score
            best_map = col_map
    if best_score <= 0:
        best_idx = -1
    return best_idx, {
        "customer_code": best_map.get("customer_code", 0),
        "customer_name": best_map.get("customer_name", 1),
        "requirement_text": best_map.get("requirement_text", 5),
    }


def _map_special_requirement_headers(headers: list[str]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    category_col: int | None = None
    for idx, header in enumerate(headers):
        compact = re.sub(r"\s+", "", header)
        if not compact:
            continue
        if "代码" in compact and "客户" in compact and "customer_code" not in col_map:
            col_map["customer_code"] = idx
        elif ("客户名称" in compact or "客户简称" in compact or compact == "客户") and "customer_name" not in col_map:
            col_map["customer_name"] = idx
        elif (
            "特殊" in compact
            and ("需求" in compact or "要求" in compact)
            and "分类" not in compact
            and "requirement_text" not in col_map
        ):
            col_map["requirement_text"] = idx
        elif compact == "类别":
            category_col = idx
    if "requirement_text" not in col_map and category_col is not None and category_col + 1 < len(headers):
        # Some source sheets put the free-text requirement in the blank column
        # immediately after "类别"; columns after that are Y/N maintenance flags.
        col_map["requirement_text"] = category_col + 1
    return col_map


def _looks_like_requirement_text(value: str) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    if text.upper() in {"Y", "N", "YES", "NO", "TRUE", "FALSE"}:
        return False
    return len(text) >= 2


def _infer_requirement_column(data_rows: list[tuple[Any, ...]]) -> int | None:
    scores: dict[int, int] = {}
    for row in data_rows[:20]:
        for idx, value in enumerate(row):
            text = normalize_text(value)
            if not _looks_like_requirement_text(text):
                continue
            score = len(text)
            if any(keyword in text for keyword in ["要求", "备注", "订单", "客户", "规格", "包装", "标签", "出货", "下单"]):
                score += 80
            scores[idx] = scores.get(idx, 0) + score
    if not scores:
        return None
    return max(scores.items(), key=lambda item: item[1])[0]


def ensure_structured_special_rules_workbook() -> Path:
    SPECIAL_RULES_DIR.mkdir(parents=True, exist_ok=True)
    if SPECIAL_RULES_PATH.exists():
        try:
            wb = load_workbook(SPECIAL_RULES_PATH)
        except Exception:
            corrupt_path = SPECIAL_RULES_PATH.with_name(
                f"{SPECIAL_RULES_PATH.stem}_corrupt_{datetime.now().strftime('%Y%m%d_%H%M%S')}{SPECIAL_RULES_PATH.suffix}"
            )
            SPECIAL_RULES_PATH.replace(corrupt_path)
            return ensure_structured_special_rules_workbook()
        changed = _ensure_workbook_schema(wb)
        if changed:
            _format_workbook(wb)
            wb.save(SPECIAL_RULES_PATH)
        return SPECIAL_RULES_PATH

    wb = Workbook()
    ws = wb.active
    ws.title = "结构化特殊规则"
    ws.append(STRUCTURED_HEADERS)
    ws_dict = wb.create_sheet("字段字典")
    ws_dict.append(["字段名", "填写说明", "建议填写/枚举", "是否直接参与转码"])
    for row in _field_dictionary_rows():
        ws_dict.append(row)
    ws_map = wb.create_sheet("原始需求映射")
    ws_map.append(["来源行号", "客户代码", "客户简称", "原始特殊需求", "结构化状态", "生成规则ID", "备注"])
    _format_workbook(wb)
    wb.save(SPECIAL_RULES_PATH)
    return SPECIAL_RULES_PATH


def _ensure_workbook_schema(wb) -> bool:
    changed = False
    if "结构化特殊规则" not in wb.sheetnames:
        ws = wb.active
        ws.title = "结构化特殊规则"
        ws.append(STRUCTURED_HEADERS)
        changed = True
    ws = wb["结构化特殊规则"]
    existing_headers = [normalize_text(cell.value) for cell in ws[1]]
    if not any(existing_headers):
        ws.append(STRUCTURED_HEADERS)
        changed = True
    else:
        for header in STRUCTURED_HEADERS:
            if header not in existing_headers:
                ws.cell(row=1, column=ws.max_column + 1, value=header)
                existing_headers.append(header)
                changed = True

    if "字段字典" not in wb.sheetnames:
        ws_dict = wb.create_sheet("字段字典")
        ws_dict.append(["字段名", "填写说明", "建议填写/枚举", "是否直接参与转码"])
        for row in _field_dictionary_rows():
            ws_dict.append(row)
        changed = True
    if "原始需求映射" not in wb.sheetnames:
        ws_map = wb.create_sheet("原始需求映射")
        ws_map.append(["来源行号", "客户代码", "客户简称", "原始特殊需求", "结构化状态", "生成规则ID", "备注"])
        changed = True
    return changed


def _field_dictionary_rows() -> list[tuple[str, str, str, str]]:
    return [
        ("规则ID", "唯一编号，保存时自动生成。", "SR-0001", "否"),
        ("启用", "是否启用该结构化规则。", "是 / 否", "是"),
        ("客户代码 / 客户简称", "规则匹配的客户身份，优先用客户代码。", "客户主数据", "是"),
        ("规则大类", "先判断是否为转码改写，再决定是否进入编码覆盖。", "见下拉选项", "是"),
        ("是否参与转码", "是：未来转码引擎可执行；否：只归档或给后续订单/备注模块使用。", "是 / 否", "是"),
        ("条件字段", "胶系、厚度、铜厚、尺寸、关键词、CTI 等匹配条件。", "如 <=0.71mm、1/1、28*48", "是"),
        ("覆盖字段", "只填需要覆盖的编码字段，其他留空。", "如 覆盖尺寸代码=28504900", "是"),
        ("原始整段特殊要求", "业务员原始单元格里的完整特殊要求，用于查询时先展示原文。", "原始整段文本", "否"),
        ("原始特殊需求", "解析拆分后的单条规则来源文本。", "拆分后的规则文本", "否"),
        ("匹配来源字段", "规则关键词应在哪类文本中匹配。", "客户规格 / 备注 / 订单备注 / 客户物料描述 / 客户特殊要求 / 整行上下文", "是"),
        ("目标动作类型", "区分编码改写、料号品号、备注、交期、生产等动作。", "见下拉选项", "是"),
        ("目标动作内容", "业务动作的可读描述。", "如下688料号、指定长春铜箔生产", "否"),
        ("执行策略", "决定规则是否可被转码引擎执行。", "参与转码 / 仅保存不转码 / 人工确认", "是"),
        ("规则状态", "规则生命周期状态，归档规则不再参与转码。", "启用 / 草稿 / 归档", "是"),
        ("导入方式", "记录本次保存是新增、追加还是替换。", "新增 / 追加 / 替换", "否"),
        ("规则版本", "客户规则版本号或批次号。", "例如 V20260530-153000", "否"),
        ("更新人 / 更新时间", "记录最后一次保存信息。", "系统自动写入", "否"),
        ("非编码动作", "不影响品名编码的动作，如数量换算、面积换算、订单备注。", "自由文本", "否"),
        ("待确认", "规则是否还需业务确认。", "是 / 否", "否"),
    ]


def read_structured_special_rules(include_archived: bool = False) -> list[dict[str, str]]:
    path = ensure_structured_special_rules_workbook()
    wb = load_workbook(path)
    ws = wb["结构化特殊规则"]
    headers = [normalize_text(cell.value) for cell in ws[1]]
    rules: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {header: normalize_text(row[idx] if idx < len(row) else "") for idx, header in enumerate(headers) if header}
        if not item.get("规则ID") and not item.get("客户代码") and not item.get("客户简称"):
            continue
        if not include_archived and item.get("规则状态") == "归档":
            continue
        rules.append(item)
    return rules


def search_structured_special_rules(query: str) -> list[dict[str, str]]:
    keyword = normalize_text(query)
    if not keyword:
        return []
    compact_keyword = re.sub(r"\s+", "", keyword).upper()
    results = []
    for rule in read_structured_special_rules():
        haystack = " ".join(str(rule.get(header, "")) for header in STRUCTURED_HEADERS)
        compact_haystack = re.sub(r"\s+", "", haystack).upper()
        if compact_keyword in compact_haystack:
            results.append(rule)
    return results


def customer_has_rules(customer_code: str, customer_name: str) -> bool:
    return bool(_rules_for_customer(read_structured_special_rules(), customer_code, customer_name))


def save_structured_special_rules(draft_rules: list[dict[str, str]], saved_by: str = "",
                                  import_mode: str = "追加") -> tuple[Path, list[str]]:
    path = ensure_structured_special_rules_workbook()
    wb = load_workbook(path)
    ws = wb["结构化特殊规则"]
    ws_map = wb["原始需求映射"]
    next_id = _next_rule_number(ws)
    saved_ids: list[str] = []
    saved_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    version = datetime.now().strftime("V%Y%m%d-%H%M%S")
    mode = import_mode if import_mode in IMPORT_MODES else "追加"
    customers: set[tuple[str, str]] = set()

    if mode == "替换":
        customers = {
            (normalize_text(draft.get("客户代码")), normalize_text(draft.get("客户简称")))
            for draft in draft_rules if normalize_text(draft.get("_delete")) != "on"
        }
        _archive_customer_rules(ws, customers, saved_by, saved_at)

    existing_rules = read_structured_special_rules()
    if mode == "替换" and customers:
        existing_rules = [
            rule for rule in existing_rules
            if not _rule_in_customer_set(rule, customers)
        ]

    for draft in draft_rules:
        if normalize_text(draft.get("_delete")) == "on":
            continue
        row = {header: normalize_text(draft.get(header, "")) for header in STRUCTURED_HEADERS}
        _apply_conflict_policy(row, existing_rules)
        if row.get("_skip_duplicate") == "1":
            continue
        row["规则ID"] = row.get("规则ID") or f"SR-{next_id:04d}"
        row["规则状态"] = row.get("规则状态") or "启用"
        row["导入方式"] = mode
        row["规则版本"] = row.get("规则版本") or version
        row["更新人"] = saved_by or "-"
        row["更新时间"] = saved_at
        next_id += 1
        saved_ids.append(row["规则ID"])
        ws.append([row.get(header, "") for header in STRUCTURED_HEADERS])
        ws_map.append([
            row.get("来源行号", ""),
            row.get("客户代码", ""),
            row.get("客户简称", ""),
            row.get("原始特殊需求", ""),
            "已保存",
            row["规则ID"],
            f"保存人：{saved_by or '-'}；保存时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        ])

    _format_workbook(wb)
    wb.save(path)
    return path, saved_ids


def _rule_in_customer_set(rule: dict[str, str], customers: set[tuple[str, str]]) -> bool:
    rule_code = normalize_text(rule.get("客户代码"))
    rule_name = normalize_text(rule.get("客户简称"))
    for cust_code, cust_name in customers:
        if cust_code and rule_code == cust_code:
            return True
        if cust_name and rule_name and (cust_name in rule_name or rule_name in cust_name):
            return True
    return False


def _archive_customer_rules(ws, customers: set[tuple[str, str]], saved_by: str, saved_at: str) -> None:
    if not customers:
        return
    headers = [normalize_text(cell.value) for cell in ws[1]]
    col = {header: idx + 1 for idx, header in enumerate(headers)}
    for row_idx in range(2, ws.max_row + 1):
        row_code = normalize_text(ws.cell(row=row_idx, column=col.get("客户代码", 0)).value)
        row_name = normalize_text(ws.cell(row=row_idx, column=col.get("客户简称", 0)).value)
        for cust_code, cust_name in customers:
            if cust_code and row_code == cust_code:
                break
            if cust_name and (cust_name in row_name or row_name in cust_name):
                break
        else:
            continue
        if "规则状态" in col:
            ws.cell(row=row_idx, column=col["规则状态"], value="归档")
        if "启用" in col:
            ws.cell(row=row_idx, column=col["启用"], value="否")
        if "导入方式" in col:
            ws.cell(row=row_idx, column=col["导入方式"], value="替换归档")
        if "更新人" in col:
            ws.cell(row=row_idx, column=col["更新人"], value=saved_by or "-")
        if "更新时间" in col:
            ws.cell(row=row_idx, column=col["更新时间"], value=saved_at)


def _rules_for_customer(rules: list[dict[str, str]], customer_code: str, customer_name: str) -> list[dict[str, str]]:
    code = normalize_text(customer_code)
    name = normalize_text(customer_name)
    matches = []
    for rule in rules:
        rule_code = normalize_text(rule.get("客户代码"))
        rule_name = normalize_text(rule.get("客户简称"))
        if code and rule_code == code:
            matches.append(rule)
        elif name and rule_name and (name in rule_name or rule_name in name):
            matches.append(rule)
    return matches


def _condition_signature(rule: dict[str, str]) -> tuple[str, ...]:
    return tuple(normalize_text(rule.get(key)) for key in [
        "适用胶系", "厚度条件", "铜厚条件", "尺寸条件", "关键词条件",
        "排除关键词", "含不含铜条件", "CTI条件", "匹配来源字段"
    ])


def _override_pairs(rule: dict[str, str]) -> list[tuple[str, str]]:
    return [(field, normalize_text(rule.get(field))) for _, field in TRANSCODE_FIELDS if normalize_text(rule.get(field))]


def _same_customer(a: dict[str, str], b: dict[str, str]) -> bool:
    a_code, b_code = normalize_text(a.get("客户代码")), normalize_text(b.get("客户代码"))
    if a_code and b_code and a_code == b_code:
        return True
    a_name, b_name = normalize_text(a.get("客户简称")), normalize_text(b.get("客户简称"))
    return bool(a_name and b_name and (a_name in b_name or b_name in a_name))


def _keyword_overlap(a: str, b: str) -> bool:
    a_set = set(_split_keywords(a))
    b_set = set(_split_keywords(b))
    return bool(a_set and b_set and (a_set & b_set))


def _append_rule_explanation(rule: dict[str, str], note: str) -> None:
    existing = normalize_text(rule.get("规则解释"))
    if note in existing:
        return
    rule["规则解释"] = f"{existing}；{note}" if existing else note


def _apply_conflict_policy(rule: dict[str, str], existing_rules: list[dict[str, str]]) -> None:
    if normalize_text(rule.get("是否参与转码")) != "是":
        return
    overrides = _override_pairs(rule)
    if not overrides:
        return
    signature = _condition_signature(rule)
    for old in existing_rules:
        if not _same_customer(rule, old):
            continue
        if signature == _condition_signature(old) and dict(_override_pairs(old)) == dict(overrides):
            rule["_skip_duplicate"] = "1"
            return
        old_overrides = dict(_override_pairs(old))
        for field, value in overrides:
            old_value = old_overrides.get(field)
            if not old_value:
                continue
            if signature == _condition_signature(old):
                if old_value == value:
                    _append_rule_explanation(rule, f"保存时发现与 {old.get('规则ID', '旧规则')} 完全重复")
                else:
                    rule["待确认"] = "是"
                    rule["是否参与转码"] = "否"
                    rule["执行策略"] = "人工确认"
                    _append_rule_explanation(rule, f"保存时发现与 {old.get('规则ID', '旧规则')} 冲突：{field} {old_value} vs {value}")
                continue
            if _keyword_overlap(rule.get("关键词条件", ""), old.get("关键词条件", "")) and old_value != value:
                _append_rule_explanation(rule, f"保存时发现与 {old.get('规则ID', '旧规则')} 潜在冲突：{field} {old_value} vs {value}")


def _next_rule_number(ws) -> int:
    max_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = normalize_text(row[0] if row else "")
        match = re.match(r"SR-(\d+)$", value)
        if match:
            max_num = max(max_num, int(match.group(1)))
    return max_num + 1


def get_structured_rule_count() -> int:
    path = ensure_structured_special_rules_workbook()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["结构化特殊规则"]
    count = max(0, ws.max_row - 1)
    wb.close()
    return count


def build_rule_workspace_view(rules: list[dict[str, str]]) -> dict:
    source_items = []
    seen_sources: set[tuple[str, str, str, str]] = set()
    for idx, rule in enumerate(rules):
        original = normalize_text(rule.get("原始整段特殊要求")) or normalize_text(rule.get("原始特殊需求"))
        if not original:
            continue
        source_line = _base_source_line(normalize_text(rule.get("来源行号")))
        key = (
            normalize_text(rule.get("客户代码")),
            normalize_text(rule.get("客户简称")),
            source_line,
            original,
        )
        if key in seen_sources:
            continue
        seen_sources.add(key)
        source_items.append({
            "index": idx,
            "客户代码": key[0],
            "客户简称": key[1],
            "来源行号": source_line,
            "原始特殊需求": original,
        })

    grid = []
    for label, field in TRANSCODE_FIELDS:
        items = []
        for idx, rule in enumerate(rules):
            value = normalize_text(rule.get(field))
            if not value:
                continue
            items.append({
                "index": idx,
                "rule_id": rule.get("规则ID", ""),
                "summary": _rule_summary(rule, label, value),
                "value": value,
                "field": field,
                "status": _display_status(rule),
                "rule": rule,
            })
        grid.append({
            "label": label,
            "field": field,
            "items": items,
            "state": "active" if items else "empty",
        })

    non_transcode_groups = []
    for action_type in NON_TRANSCODE_TYPES:
        items = [
            {"index": idx, "summary": _non_transcode_summary(rule), "rule": rule}
            for idx, rule in enumerate(rules)
            if normalize_text(rule.get("目标动作类型")) == action_type or normalize_text(rule.get("规则大类")) == action_type
        ]
        if items:
            non_transcode_groups.append({"label": action_type, "items": items})

    manual_items = [
        {"index": idx, "summary": _manual_summary(rule), "rule": rule}
        for idx, rule in enumerate(rules)
        if normalize_text(rule.get("待确认")) == "是"
        or normalize_text(rule.get("执行策略")) == "人工确认"
        or normalize_text(rule.get("规则大类")) == "人工确认"
    ]
    return {
        "rules": rules,
        "source_items": source_items,
        "grid": grid,
        "non_transcode_groups": non_transcode_groups,
        "manual_items": manual_items,
        "counts": {
            "total": len(rules),
            "transcode": sum(1 for rule in rules if normalize_text(rule.get("是否参与转码")) == "是"),
            "non_transcode": sum(1 for rule in rules if normalize_text(rule.get("是否参与转码")) != "是"),
            "manual": len(manual_items),
        },
    }


def _base_source_line(source_line: str) -> str:
    return re.sub(r"(\!\d+)\.\d+$", r"\1", normalize_text(source_line))


def build_customer_summaries(rules: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = {}
    for rule in rules:
        key = (normalize_text(rule.get("客户代码")), normalize_text(rule.get("客户简称")))
        grouped.setdefault(key, []).append(rule)
    summaries = []
    for (code, name), group in grouped.items():
        view = build_rule_workspace_view(group)
        summaries.append({
            "客户代码": code,
            "客户简称": name,
            "规则总数": str(view["counts"]["total"]),
            "参与转码": str(view["counts"]["transcode"]),
            "仅保存": str(view["counts"]["non_transcode"]),
            "人工确认": str(view["counts"]["manual"]),
            "影响转码字段": "；".join(
                f"{cell['label']}:{len(cell['items'])}" for cell in view["grid"] if cell["items"]
            ),
            "非编码动作": "；".join(group_item["label"] for group_item in view["non_transcode_groups"]),
        })
    return sorted(summaries, key=lambda item: (item.get("客户代码", ""), item.get("客户简称", "")))


def build_export_workbook(export_type: str = "full") -> BytesIO:
    rules = read_structured_special_rules(include_archived=True)
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    if export_type in ("full", "structured"):
        _append_dict_sheet(wb, "结构化规则总表", rules, STRUCTURED_HEADERS)
    if export_type in ("full", "customer_summary"):
        _append_dict_sheet(wb, "客户汇总视图", build_customer_summaries(rules), [
            "客户代码", "客户简称", "规则总数", "参与转码", "仅保存", "人工确认", "影响转码字段", "非编码动作"
        ])
    if export_type in ("full", "transcode"):
        transcode_rules = [rule for rule in rules if normalize_text(rule.get("是否参与转码")) == "是" and normalize_text(rule.get("待确认")) != "是"]
        _append_dict_sheet(wb, "参与转码规则", transcode_rules, STRUCTURED_HEADERS)
    if export_type == "full":
        non_transcode = [rule for rule in rules if normalize_text(rule.get("是否参与转码")) != "是" and normalize_text(rule.get("规则大类")) != "人工确认"]
        _append_dict_sheet(wb, "仅保存不转码", non_transcode, STRUCTURED_HEADERS)
    if export_type in ("full", "manual"):
        manual = [
            rule for rule in rules
            if normalize_text(rule.get("待确认")) == "是"
            or normalize_text(rule.get("执行策略")) == "人工确认"
            or normalize_text(rule.get("规则大类")) == "人工确认"
        ]
        _append_dict_sheet(wb, "人工确认", manual, STRUCTURED_HEADERS)
    if export_type == "full":
        path = ensure_structured_special_rules_workbook()
        try:
            src_wb = load_workbook(path, read_only=True, data_only=True)
            if "原始需求映射" in src_wb.sheetnames:
                ws_src = src_wb["原始需求映射"]
                ws_dst = wb.create_sheet("原文映射")
                for row in ws_src.iter_rows(values_only=True):
                    ws_dst.append(list(row))
        except Exception:
            pass

    _format_workbook(wb)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def _append_dict_sheet(wb, title: str, rows: list[dict[str, str]], headers: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def _rule_summary(rule: dict[str, str], label: str, value: str) -> str:
    if normalize_text(rule.get("目标动作内容")):
        return f"{rule.get('目标动作内容')} -> {label}{value}"
    if normalize_text(rule.get("关键词条件")):
        return f"{rule.get('关键词条件')} -> {value}"
    if normalize_text(rule.get("适用胶系")):
        return f"{rule.get('适用胶系')} -> {value}"
    if normalize_text(rule.get("铜厚条件")):
        return f"{rule.get('铜厚条件')} -> {value}"
    return f"{label} -> {value}"


def _non_transcode_summary(rule: dict[str, str]) -> str:
    return normalize_text(rule.get("目标动作内容")) or normalize_text(rule.get("非编码动作")) or normalize_text(rule.get("规则解释")) or normalize_text(rule.get("原始特殊需求"))


def _manual_summary(rule: dict[str, str]) -> str:
    return normalize_text(rule.get("规则解释")) or normalize_text(rule.get("目标动作内容")) or normalize_text(rule.get("原始特殊需求"))


def _display_status(rule: dict[str, str]) -> str:
    if normalize_text(rule.get("待确认")) == "是":
        return "待确认"
    if normalize_text(rule.get("是否参与转码")) == "是":
        return "参与转码"
    return "仅保存"


def _format_workbook(wb) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"
        ws.sheet_view.showGridLines = False

    if "结构化特殊规则" in wb.sheetnames:
        ws = wb["结构化特殊规则"]
        widths = {
            "A": 12, "B": 8, "C": 10, "D": 12, "E": 14, "F": 12, "G": 18, "H": 14, "I": 10,
            "J": 16, "K": 18, "L": 16, "M": 18, "N": 18, "O": 16, "P": 16, "Q": 12,
            "R": 14, "S": 14, "T": 14, "U": 16, "V": 14, "W": 14, "X": 14, "Y": 14, "Z": 14,
            "AA": 42, "AB": 18, "AC": 70, "AD": 44, "AE": 10, "AF": 18, "AG": 18, "AH": 36, "AI": 18,
            "AJ": 10, "AK": 12, "AL": 18, "AM": 14, "AN": 18,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        _set_validations(ws)

    if "字段字典" in wb.sheetnames:
        for col, width in {"A": 22, "B": 56, "C": 50, "D": 18}.items():
            wb["字段字典"].column_dimensions[col].width = width
    if "原始需求映射" in wb.sheetnames:
        for col, width in {"A": 10, "B": 12, "C": 14, "D": 90, "E": 18, "F": 20, "G": 44}.items():
            wb["原始需求映射"].column_dimensions[col].width = width


def _set_validations(ws) -> None:
    ws.data_validations.dataValidation = []
    yes_no_formula = '"' + ",".join(YES_NO) + '"'
    for col in ["B", "H", "AE"]:
        dv = DataValidation(type="list", formula1=yes_no_formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1000")
    dv_rule = DataValidation(type="list", formula1='"' + ",".join(RULE_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_rule)
    dv_rule.add("G2:G1000")
    dv_mat = DataValidation(type="list", formula1='"' + ",".join(MATERIAL_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_mat)
    dv_mat.add("F2:F1000")
    dv_source = DataValidation(type="list", formula1='"' + ",".join(MATCH_SOURCE_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_source)
    dv_source.add("AF2:AF1000")
    dv_action = DataValidation(type="list", formula1='"' + ",".join(ACTION_TYPES) + '"', allow_blank=True)
    ws.add_data_validation(dv_action)
    dv_action.add("AG2:AG1000")
    dv_strategy = DataValidation(type="list", formula1='"' + ",".join(EXECUTION_STRATEGIES) + '"', allow_blank=True)
    ws.add_data_validation(dv_strategy)
    dv_strategy.add("AI2:AI1000")
    dv_status = DataValidation(type="list", formula1='"' + ",".join(RULE_STATUSES) + '"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add("AJ2:AJ1000")
    dv_import = DataValidation(type="list", formula1='"' + ",".join(IMPORT_MODES) + '"', allow_blank=True)
    ws.add_data_validation(dv_import)
    dv_import.add("AK2:AK1000")
