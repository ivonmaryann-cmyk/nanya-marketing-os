from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .customer_archive_service import as_text, utcnow
from .database import automation_cursor as db_cursor


FORM_FIELDS = (
    "adhesive_code", "adhesive_name", "usage_category", "finance_category",
    "legacy_adhesive_code", "enabled",
)
WORKBOOK_HEADERS = {
    "胶系编号": "adhesive_code",
    "胶系名称": "adhesive_name",
    "用途类别": "usage_category",
    "财务类别": "finance_category",
    "旧胶系编号": "legacy_adhesive_code",
}


def _row(row: Any) -> dict[str, Any]:
    return dict(row) if row is not None else {}


def _enabled(value: Any) -> int:
    text = as_text(value).lower()
    if text in {"0", "n", "no", "否", "停用", "disabled"}:
        return 0
    if text in {"", "1", "y", "yes", "是", "启用", "active"}:
        return 1
    raise ValueError("状态仅支持启用/停用，或 1/0。")


def _validated(values: dict[str, Any]) -> dict[str, Any]:
    code = as_text(values.get("adhesive_code")).upper()
    if not code:
        raise ValueError("胶系编号不能为空。")
    return {
        "adhesive_code": code,
        "adhesive_name": as_text(values.get("adhesive_name")),
        "usage_category": as_text(values.get("usage_category")),
        "finance_category": as_text(values.get("finance_category")),
        "legacy_adhesive_code": as_text(values.get("legacy_adhesive_code")).upper(),
        "enabled": _enabled(values.get("enabled")),
    }


def list_adhesive_codes(
    *, keyword: str = "", usage_category: str = "all", finance_category: str = "all",
    status: str = "all",
) -> list[dict[str, Any]]:
    clauses = ["1=1"]
    params: list[Any] = []
    if usage_category != "all":
        clauses.append("usage_category=?")
        params.append(usage_category)
    if finance_category != "all":
        clauses.append("finance_category=?")
        params.append(finance_category)
    if status in {"active", "disabled"}:
        clauses.append("enabled=?")
        params.append(1 if status == "active" else 0)
    if keyword.strip():
        like = f"%{keyword.strip()}%"
        clauses.append("(adhesive_code LIKE ? OR adhesive_name LIKE ? OR legacy_adhesive_code LIKE ?)")
        params.extend([like, like, like])
    with db_cursor() as conn:
        rows = conn.execute(
            f"""SELECT * FROM automation_adhesive_codes
                 WHERE {' AND '.join(clauses)}
                 ORDER BY adhesive_code""",
            params,
        ).fetchall()
    return [_row(item) for item in rows]


def adhesive_code_filter_values() -> dict[str, list[str]]:
    with db_cursor() as conn:
        rows = conn.execute(
            """SELECT DISTINCT usage_category,finance_category
                 FROM automation_adhesive_codes
                 ORDER BY usage_category,finance_category"""
        ).fetchall()
    return {
        "usage_categories": sorted({as_text(item["usage_category"]) for item in rows if item["usage_category"]}),
        "finance_categories": sorted({as_text(item["finance_category"]) for item in rows if item["finance_category"]}),
    }


def find_adhesive_code_candidates(glue_system: Any) -> list[dict[str, Any]]:
    value = as_text(glue_system)
    if not value:
        return []
    like = f"%{value}%"
    with db_cursor() as conn:
        rows = conn.execute(
            """SELECT adhesive_code,adhesive_name
                 FROM automation_adhesive_codes
                 WHERE enabled=1
                   AND (adhesive_name LIKE ? OR adhesive_code LIKE ?)
                 ORDER BY adhesive_code""",
            (like, like),
        ).fetchall()
    return [_row(row) for row in rows]


def get_adhesive_code(record_id: int) -> dict[str, Any] | None:
    with db_cursor() as conn:
        row = conn.execute("SELECT * FROM automation_adhesive_codes WHERE id=?", (record_id,)).fetchone()
    return _row(row) if row else None


def save_adhesive_code(
    values: dict[str, Any], *, record_id: int | None = None, operated_by: str = "",
) -> int:
    data = _validated(values)
    columns = tuple(data)
    now = utcnow()
    with db_cursor() as conn:
        params: list[Any] = [data["adhesive_code"]]
        duplicate_sql = "SELECT id FROM automation_adhesive_codes WHERE adhesive_code=?"
        if record_id:
            duplicate_sql += " AND id<>?"
            params.append(record_id)
        if conn.execute(duplicate_sql, params).fetchone():
            raise ValueError("该胶系编号已存在。")
        if record_id:
            if not conn.execute("SELECT id FROM automation_adhesive_codes WHERE id=?", (record_id,)).fetchone():
                raise ValueError("胶系代码记录不存在。")
            conn.execute(
                f"UPDATE automation_adhesive_codes SET {','.join(f'{field}=?' for field in columns)},updated_by=?,updated_at=? WHERE id=?",
                (*[data[field] for field in columns], operated_by, now, record_id),
            )
            return record_id
        cursor = conn.execute(
            f"""INSERT INTO automation_adhesive_codes
                    ({','.join(columns)},source_json,updated_by,created_at,updated_at)
                VALUES ({','.join('?' for _ in columns)},?,?,?,?)""",
            (*[data[field] for field in columns], "{}", operated_by, now, now),
        )
        return int(cursor.lastrowid)


def set_adhesive_code_enabled(record_id: int, enabled: bool, *, operated_by: str = "") -> None:
    with db_cursor() as conn:
        result = conn.execute(
            "UPDATE automation_adhesive_codes SET enabled=?,updated_by=?,updated_at=? WHERE id=?",
            (1 if enabled else 0, operated_by, utcnow(), record_id),
        )
        if result.rowcount == 0:
            raise ValueError("胶系代码记录不存在。")


def import_adhesive_code_workbook(file_path: str | Path, *, operated_by: str = "") -> dict[str, Any]:
    book = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = book.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [as_text(value) for value in header_row]
        positions = {header: index for index, header in enumerate(headers) if header in WORKBOOK_HEADERS}
        if "胶系编号" not in positions:
            raise ValueError("未找到“胶系编号”表头，无法导入胶系主表。")
        parsed: list[tuple[dict[str, Any], dict[str, Any]]] = []
        errors: list[dict[str, Any]] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            source = {header: as_text(row[index]) if index < len(row) else "" for header, index in positions.items()}
            if not any(source.values()):
                continue
            try:
                parsed.append((_validated({
                    field: source.get(header, "") for header, field in WORKBOOK_HEADERS.items()
                }), source))
            except ValueError as exc:
                errors.append({"row": row_number, "error": str(exc)})
        imported = updated = 0
        columns = tuple(FORM_FIELDS[:-1]) + ("enabled",)
        now = utcnow()
        with db_cursor() as conn:
            for data, source in parsed:
                existing = conn.execute(
                    "SELECT id FROM automation_adhesive_codes WHERE adhesive_code=?", (data["adhesive_code"],)
                ).fetchone()
                if existing:
                    conn.execute(
                        f"UPDATE automation_adhesive_codes SET {','.join(f'{field}=?' for field in columns)},source_json=?,updated_by=?,updated_at=? WHERE id=?",
                        (*[data[field] for field in columns], json.dumps(source, ensure_ascii=False), operated_by, now, _row(existing)["id"]),
                    )
                    updated += 1
                else:
                    conn.execute(
                        f"""INSERT INTO automation_adhesive_codes
                                ({','.join(columns)},source_json,updated_by,created_at,updated_at)
                            VALUES ({','.join('?' for _ in columns)},?,?,?,?)""",
                        (*[data[field] for field in columns], json.dumps(source, ensure_ascii=False), operated_by, now, now),
                    )
                    imported += 1
        return {"imported": imported, "updated": updated, "skipped": len(errors), "errors": errors}
    finally:
        book.close()
