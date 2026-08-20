from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .purchase_factory_mapper import (
    FACTORY_DETAIL_HEADERS,
    FACTORY_MAIN_HEADERS,
    INTERNAL_SALES_DETAIL_HEADERS,
    INTERNAL_SALES_MAIN_HEADERS,
    _extract_roll_length,
    _project_detail_standard_fields,
    project_factory_document,
    project_internal_sales_document,
)
from .purchase_field_rules import STANDARD_HEADERS, clean_text, decimal_or_none, normalize_date


HEADER_FILL = "1F4E78"
SECTION_FILL = "D9EAF7"
LIGHT_FILL = "EEF3F8"
TITLE_FILL = "0B1F4D"
THIN_SIDE = Side(style="thin", color="A6A6A6")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _style_cell(cell, *, bold: bool = False, fill: str | None = None, color: str = "000000", align: str = "left") -> None:
    cell.font = Font(name="Microsoft YaHei", bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _style_row(ws, row: int, start_col: int, end_col: int, *, fill: str, color: str = "000000", bold: bool = True) -> None:
    for column in range(start_col, end_col + 1):
        _style_cell(ws.cell(row=row, column=column), bold=bold, fill=fill, color=color, align="center")


def _set_purchase_sheet_layout(ws) -> None:
    widths = {"A": 14, "B": 28, "C": 24, "D": 26, "E": 18, "F": 16, "G": 18, "H": 18, "I": 18, "J": 20, "K": 24, "L": 24}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for index in range(13, 31):
        ws.column_dimensions[get_column_letter(index)].width = 18
    ws.sheet_view.showGridLines = False


def _set_detail_sheet_layout(ws, headers: list[str]) -> None:
    descriptive_widths = {
        "物料编号": 26,
        "物料编码": 26,
        "物料名称": 28,
        "物料名/规格": 28,
        "型号/规格": 28,
        "说明": 28,
        "标准-物料编码": 26,
        "标准-物料名称": 28,
        "标准-说明": 28,
    }
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = descriptive_widths.get(header, 16)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


def _write_key_values(ws, row: int, header_info: dict[str, Any], max_col: int) -> int:
    pairs = [(key, value) for key, value in header_info.items() if clean_text(value)]
    if not pairs:
        return row
    title_row = row
    ws.cell(row=row, column=1, value="订单头信息")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _style_row(ws, row, 1, max_col, fill=SECTION_FILL)
    row += 1

    for index in range(0, len(pairs), 2):
        left_key, left_value = pairs[index]
        ws.cell(row=row, column=1, value=left_key)
        ws.cell(row=row, column=2, value=left_value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max(2, max_col // 2))
        _style_cell(ws.cell(row=row, column=1), bold=True, fill=LIGHT_FILL, align="center")
        _style_cell(ws.cell(row=row, column=2), align="left")
        if index + 1 < len(pairs):
            right_key, right_value = pairs[index + 1]
            right_label_col = max_col // 2 + 1
            right_value_col = right_label_col + 1
            ws.cell(row=row, column=right_label_col, value=right_key)
            ws.cell(row=row, column=right_value_col, value=right_value)
            ws.merge_cells(start_row=row, start_column=right_value_col, end_row=row, end_column=max_col)
            _style_cell(ws.cell(row=row, column=right_label_col), bold=True, fill=LIGHT_FILL, align="center")
            _style_cell(ws.cell(row=row, column=right_value_col), align="left")
        for column in range(1, max_col + 1):
            ws.cell(row=row, column=column).border = THIN_BORDER
        row += 1
    return row + 1


def _write_raw_table(ws, row: int, table: dict[str, Any], max_col_hint: int) -> tuple[int, int]:
    rows = table.get("rows") or []
    if not rows:
        return row, 0
    max_col = max(max((len(source_row) for source_row in rows), default=1), max_col_hint)
    title = f"明细表 - 第 {int(table.get('page_index') or 0) + 1} 页"
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _style_row(ws, row, 1, max_col, fill=SECTION_FILL)
    row += 1
    start = row
    for row_values in rows:
        for column in range(1, max_col + 1):
            value = row_values[column - 1] if column <= len(row_values) else ""
            cell = ws.cell(row=row, column=column, value=value)
            _style_cell(cell, align="left")
        row += 1
    if row > start:
        _style_row(ws, start, 1, max_col, fill=HEADER_FILL, color="FFFFFF")
    for row_index in range(start + 1, row):
        ws.row_dimensions[row_index].height = 36
    return row + 1, max(0, len(rows) - 1)


def _rebuilt_detail_rows(document: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        detail
        for detail in document.get("mapped_detail_rows") or []
        if "packed_text_rebuild" in str(detail.get("method") or "")
    ]


def _write_rebuilt_detail_table(ws, row: int, document: dict[str, Any], max_col_hint: int) -> tuple[int, int]:
    details = _rebuilt_detail_rows(document)
    if not details:
        return row, 0
    headers = list(STANDARD_HEADERS)
    max_col = max(max_col_hint, len(headers))
    page_indexes = sorted({int(detail.get("page_index") or 0) + 1 for detail in details})
    page_text = "、".join(str(index) for index in page_indexes) if page_indexes else "1"
    ws.cell(row=row, column=1, value=f"明细表 - 第 {page_text} 页（已重建分列）")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _style_row(ws, row, 1, max_col, fill=SECTION_FILL)
    row += 1

    for column, header in enumerate(headers, start=1):
        ws.cell(row=row, column=column, value=header)
    _style_row(ws, row, 1, max_col, fill=HEADER_FILL, color="FFFFFF")
    row += 1

    for detail in details:
        standard = detail.get("standard") or {}
        for column, header in enumerate(headers, start=1):
            value = _typed_value(header, standard.get(header, ""))
            cell = ws.cell(row=row, column=column, value=value)
            align = "right" if header in {"数量", "含税单价", "金额"} else "left"
            _style_cell(cell, align=align)
            number_format = _detail_number_format(f"标准-{header}")
            if number_format:
                cell.number_format = number_format
        ws.row_dimensions[row].height = 36
        row += 1
    return row + 1, len(details)


def _write_sections(ws, row: int, sections: dict[str, list[str]], max_col: int) -> int:
    for section in ["备注", "条款", "付款信息", "收货信息", "签核区"]:
        lines = [clean_text(line) for line in sections.get(section) or [] if clean_text(line)]
        if not lines:
            continue
        ws.cell(row=row, column=1, value=section)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        _style_row(ws, row, 1, max_col, fill=SECTION_FILL)
        row += 1
        ws.cell(row=row, column=1, value="\n".join(lines))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        _style_cell(ws.cell(row=row, column=1), align="left")
        for column in range(1, max_col + 1):
            ws.cell(row=row, column=column).border = THIN_BORDER
        ws.row_dimensions[row].height = min(max(36, len(lines) * 20), 160)
        row += 2
    return row


def _typed_value(header: str, value: Any):
    text = clean_text(value)
    if not text:
        return ""
    if header in {"数量", "含税单价", "金额"}:
        number = decimal_or_none(text)
        if number is not None:
            return float(number) if number % 1 else int(number)
    if header == "交货日期":
        date = normalize_date(text)
        return date or text
    return text


def _standard_header(name: str) -> str:
    return f"标准-{name}"


def _original_header_standard_field(header: str) -> str | None:
    compact = clean_text(header).replace(" ", "").lower()
    if not compact:
        return None
    aliases = [
        ("序号", ["序号", "序", "no", "item"]),
        ("物料编码", ["物料编码", "物料编号", "料件编号", "原料编码", "产品编码", "partno", "partnumber"]),
        ("物料名称", ["物料名称", "名称规格", "规格型号", "产品名称", "原料名称", "品名规格", "description"]),
        ("说明", ["说明"]),
        ("数量", ["数量", "订单数量", "采购量", "quantity", "qty"]),
        ("单位", ["单位", "unit"]),
        ("含税单价", ["含税单价", "单价", "unitprice", "price"]),
        ("金额", ["金额", "amount"]),
        ("交货日期", ["交货日期", "到货日期", "交期", "deliverydate"]),
        ("备注", ["备注", "comments", "remark"]),
    ]
    matches = [field for field, keys in aliases if any(key in compact for key in keys)]
    if len(matches) == 1:
        return matches[0]
    return None


def _is_noisy_original_header(header: str) -> bool:
    compact = clean_text(header).replace(" ", "").lower()
    if not compact:
        return True
    if compact.startswith("列") and compact[1:].isdigit():
        return True
    if ("金额" in compact or "amount" in compact) and any(key in compact for key in ["到货日期", "交货日期", "交期", "deliverydate"]):
        return True
    matched_fields = {
        field
        for field in STANDARD_HEADERS
        if field.replace(" ", "").lower() in compact
    }
    return len(matched_fields) > 1


def _normalized_original_values(original: dict[str, Any], standard: dict[str, Any]) -> dict[str, Any]:
    values = dict(original)
    for header in list(values.keys()):
        field = _original_header_standard_field(header)
        if field and clean_text(standard.get(field)):
            values[header] = standard.get(field)
    return values


def _detail_number_format(header: str) -> str | None:
    field = header.replace("标准-", "")
    if field == "数量":
        return "#,##0.###"
    if field == "含税单价":
        return "#,##0.000"
    if field == "金额":
        return "#,##0.00"
    if field == "交货日期":
        return "yyyy-mm-dd"
    return None


def _write_purchase_sheet(ws, documents: list[dict[str, Any]]) -> dict[str, int]:
    _set_purchase_sheet_layout(ws)
    row = 1
    detail_count = 0
    page_count = 0
    if not documents:
        ws.cell(row=1, column=1, value="未识别到可转换文件")
        return {"detail_count": 0, "page_count": 0}

    for index, document in enumerate(documents, start=1):
        raw_tables = document.get("raw_detail_tables") or []
        rebuilt_rows = _rebuilt_detail_rows(document)
        max_col = max([7, len(STANDARD_HEADERS) if rebuilt_rows else 0] + [max((len(r) for r in table.get("rows") or []), default=1) for table in raw_tables])
        title = clean_text(document.get("source_file")) or f"文件 {index}"
        ws.cell(row=row, column=1, value=f"{index}. {title}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        _style_row(ws, row, 1, max_col, fill=TITLE_FILL, color="FFFFFF")
        ws.row_dimensions[row].height = 24
        row += 1
        row = _write_key_values(ws, row, document.get("header_info") or {}, max_col)
        if rebuilt_rows:
            row, count = _write_rebuilt_detail_table(ws, row, document, max_col)
            detail_count += count
        elif raw_tables:
            for table in raw_tables:
                row, count = _write_raw_table(ws, row, table, max_col)
                detail_count += count
        else:
            ws.cell(row=row, column=1, value="未恢复出明细表")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
            _style_row(ws, row, 1, max_col, fill="FCE4D6")
            row += 2
        row = _write_sections(ws, row, document.get("sections") or {}, max_col)
        row += 1
        page_count += int(document.get("page_count") or 0)
    return {"detail_count": detail_count, "page_count": page_count}


def _detail_headers(documents: list[dict[str, Any]]) -> list[str]:
    original_headers: list[str] = []
    seen: set[str] = set()
    for document in documents:
        for row in document.get("mapped_detail_rows") or []:
            for header in (row.get("original") or {}).keys():
                header_text = clean_text(header)
                if _is_noisy_original_header(header_text):
                    continue
                key = header_text.lower()
                if header_text and key not in seen:
                    seen.add(key)
                    original_headers.append(header_text)
    headers = []
    if len(documents) > 1:
        headers.extend(["来源文件", "页码", "订单号"])
    headers.extend(original_headers)
    headers.extend([f"标准-{header}" for header in STANDARD_HEADERS])
    return headers


def _set_factory_sheet_layout(ws) -> None:
    widths = [16, 30, 18, 25, 18, 15, 18, 18, 24, 24, 20, 28]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def _style_factory_cell(cell, *, header: bool = False) -> None:
    cell.font = Font(name="宋体", size=11, bold=False, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=header)
    cell.border = THIN_BORDER


def _factory_cell_value(header: str, value: Any) -> Any:
    text = clean_text(value)
    if not text:
        return ""
    if header in {"数量（必填）", "税前单价（选填）", "单价（选填）"}:
        number = decimal_or_none(text)
        if number is not None:
            return float(number) if number % 1 else int(number)
    if header == "出货日期（选填）":
        normalized = normalize_date(text)
        if normalized:
            try:
                return datetime.strptime(normalized, "%Y-%m-%d").date()
            except ValueError:
                pass
    return text


def _write_detail_sheet(ws, documents: list[dict[str, Any]]) -> int:
    if len(documents) > 1:
        raise ValueError("厂内导入模板每个工作簿只能包含一份客户采购单。")
    document = documents[0] if documents else {}
    if document and not document.get("factory_import"):
        project_factory_document(document)
    factory_import = document.get("factory_import") or {
        "main_headers": FACTORY_MAIN_HEADERS,
        "main_values": ["220", "1", "1", "", "", "", "", "", "", "", "", ""],
        "detail_headers": FACTORY_DETAIL_HEADERS,
        "rows": [],
    }
    main_headers = list(factory_import.get("main_headers") or FACTORY_MAIN_HEADERS)
    main_values = list(factory_import.get("main_values") or [""] * len(main_headers))
    detail_headers = list(factory_import.get("detail_headers") or FACTORY_DETAIL_HEADERS)
    rows = list(factory_import.get("rows") or [])
    _set_factory_sheet_layout(ws)

    for column in range(1, 13):
        header = main_headers[column - 1] if column <= len(main_headers) else ""
        value = main_values[column - 1] if column <= len(main_values) else ""
        ws.cell(row=1, column=column, value=header or None)
        ws.cell(row=2, column=column, value=value or None)
        _style_factory_cell(ws.cell(row=1, column=column), header=True)
        _style_factory_cell(ws.cell(row=2, column=column))

    for column, header in enumerate(detail_headers, start=1):
        ws.cell(row=3, column=column, value=header)
        _style_factory_cell(ws.cell(row=3, column=column), header=True)

    for row_index, row in enumerate(rows, start=4):
        for column, header in enumerate(detail_headers, start=1):
            cell = ws.cell(row=row_index, column=column, value=_factory_cell_value(header, row.get(header, "")))
            _style_factory_cell(cell)
            if header in {"项次（必填）", "产品编号", "客户产品编号（必填）", "客户订单号"}:
                cell.number_format = "@"
            elif header == "出货日期（选填）":
                cell.number_format = "yyyy/m/d;@"
            elif header == "数量（必填）":
                cell.number_format = "0.###"
            elif header in {"税前单价（选填）", "单价（选填）"}:
                cell.number_format = "#,##0.00_ "

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 42
    for row_index in range(4, 4 + len(rows)):
        ws.row_dimensions[row_index].height = 24
    return len(rows)


def _write_issue_sheet(ws, documents: list[dict[str, Any]]) -> int:
    headers = ["文件", "页码", "区域", "字段", "原始识别值", "清洗后值", "置信度", "问题说明"]
    for column, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column, value=header)
    _style_row(ws, 1, 1, len(headers), fill="9E480E", color="FFFFFF")
    widths = [24, 8, 18, 18, 45, 30, 10, 45]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    row_index = 2
    for document in documents:
        for issue in document.get("issues") or []:
            values = [
                document.get("source_file", ""),
                "" if issue.get("page_index") in {"", None} else int(issue.get("page_index") or 0) + 1,
                issue.get("region", ""),
                issue.get("field", ""),
                issue.get("raw_value", ""),
                issue.get("clean_value", ""),
                issue.get("confidence", ""),
                issue.get("message", ""),
            ]
            for column, value in enumerate(values, start=1):
                _style_cell(ws.cell(row=row_index, column=column, value=value), align="left")
            ws.row_dimensions[row_index].height = 48
            row_index += 1
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    return max(0, row_index - 2)


def write_purchase_order_workbook(documents: list[dict[str, Any]], output_path: Path) -> dict[str, int]:
    if len(documents) > 1:
        raise ValueError("每份厂内订单导入工作簿只能写入一份采购单。")
    for document in documents:
        if not document.get("factory_import"):
            project_factory_document(document)

    wb = Workbook()
    purchase_ws = wb.active
    purchase_ws.title = "采购单"
    purchase_stats = _write_purchase_sheet(purchase_ws, documents)

    detail_ws = wb.create_sheet("明细数据")
    detail_count = _write_detail_sheet(detail_ws, documents)

    issue_count = sum(len(document.get("issues") or []) for document in documents)
    if issue_count:
        issue_ws = wb.create_sheet("识别日志")
        issue_count = _write_issue_sheet(issue_ws, documents)

    wb.save(output_path)
    return {
        "structured_count": detail_count,
        "cell_count": detail_count,
        "issue_count": issue_count,
        "page_count": purchase_stats.get("page_count", 0),
        "ready_count": sum(int((document.get("factory_mapping_summary") or {}).get("ready_rows") or 0) for document in documents),
        "review_count": sum(int((document.get("factory_mapping_summary") or {}).get("review_rows") or 0) for document in documents),
    }


def _set_internal_sales_sheet_layout(ws) -> None:
    widths = [16, 30, 18, 28, 52, 24, 18, 18, 18, 24, 24, 24, 22, 28]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def _style_internal_sales_cell(cell, *, header: bool = False) -> None:
    cell.font = Font(name="宋体", size=11, bold=False, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=header or cell.column in {4, 5, 14})
    cell.border = THIN_BORDER


def write_internal_sales_workbook(document: dict[str, Any], output_path: Path) -> int:
    internal_sales = project_internal_sales_document(document)
    main_headers = list(internal_sales.get("main_headers") or INTERNAL_SALES_MAIN_HEADERS)
    main_values = list(internal_sales.get("main_values") or [""] * len(main_headers))
    detail_headers = list(internal_sales.get("detail_headers") or INTERNAL_SALES_DETAIL_HEADERS)
    rows = list(internal_sales.get("rows") or [])

    wb = Workbook()
    ws = wb.active
    ws.title = "内销"
    _set_internal_sales_sheet_layout(ws)

    for column in range(1, 15):
        header = main_headers[column - 1] if column <= len(main_headers) else ""
        value = main_values[column - 1] if column <= len(main_values) else ""
        ws.cell(row=1, column=column, value=header or None)
        ws.cell(row=2, column=column, value=value or None)
        _style_internal_sales_cell(ws.cell(row=1, column=column), header=True)
        _style_internal_sales_cell(ws.cell(row=2, column=column))

    for column, header in enumerate(detail_headers, start=1):
        ws.cell(row=3, column=column, value=header)
        _style_internal_sales_cell(ws.cell(row=3, column=column), header=True)

    text_headers = {
        "项次（必填）",
        "产品编号",
        "客户产品编号（必填）",
        "客户订单序号（选填）",
        "客户订单号",
    }
    for row_index, row in enumerate(rows, start=4):
        for column, header in enumerate(detail_headers, start=1):
            cell = ws.cell(row=row_index, column=column, value=_factory_cell_value(header, row.get(header, "")))
            _style_internal_sales_cell(cell)
            if header in text_headers:
                cell.number_format = "@"
            elif header == "出货日期（选填）":
                cell.number_format = "yyyy/m/d;@"
            elif header == "数量（必填）":
                cell.number_format = "0.###"
            elif header in {"税前单价（选填）", "单价（选填）"}:
                cell.number_format = "#,##0.00_ "

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 42
    for row_index in range(4, 4 + len(rows)):
        ws.row_dimensions[row_index].height = 66

    wb.save(output_path)
    return len(rows)
