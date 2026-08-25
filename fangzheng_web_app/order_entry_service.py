from __future__ import annotations

import json
import re
import shutil
import subprocess
import sys
import tempfile
from copy import copy
from decimal import Decimal, InvalidOperation
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .database import automation_cursor as db_cursor
from .customer_archive_service import get_customer, get_enabled_extraction_maps
from .customer_spec_mapping_service import build_customer_spec_match
from .db import utcnow
from .excel_utils import load_workbook_compat
from .file_storage import resolve_attachment_path
from .order_intake_service import get_case
from .paths import PACKAGE_DIR, PROJECT_DIR
from .pdf_excel_domestic_export import build_domestic_template_data
from .order_document_sources import build_mail_html_purchase_document
from .order_interface_service import record_order_detail_event
from .purchase_field_rules import clean_text, normalize_date, normalize_number
from .purchase_factory_mapper import project_factory_document
from .pdf_excel_service import recognize_purchase_order_document


# The mail workspace and PDF/图片转Excel export intentionally share this
# exact workbook and field order.  The editable page remains the saved source
# of truth for downloads.
DOMESTIC_TEMPLATE_PATH = PACKAGE_DIR / "default_rules" / "order_entry" / "PDF转Excel内销录单模板.xlsx"

HEADER_FIELDS = (
    "order_type", "type_1", "type_2", "bill_to_customer_code",
    "ship_to_customer_code", "delivery_factory", "customer_order_number", "ledger",
)
LINE_FIELDS = (
    "line_no", "product_code", "product_name", "customer_product_code",
    "customer_spec", "customer_spec_match", "product_type", "delivery_date",
    "quantity", "price_before_tax", "unit_price", "origin",
    "customer_order_seq", "one_to_many", "remark",
)
HEADER_LABELS = {
    "order_type": "单别", "type_1": "类型1", "type_2": "类型2",
    "bill_to_customer_code": "账款客户编号", "ship_to_customer_code": "送货客户编号",
    "delivery_factory": "送货厂别", "customer_order_number": "客户订单号", "ledger": "账套",
}
LINE_LABELS = {
    "line_no": "项次", "product_code": "产品编号", "product_name": "品名",
    "customer_product_code": "客户产品编号", "customer_spec": "客户规格",
    "customer_spec_match": "客户规格匹配", "product_type": "产品类型（PP、基板）",
    "delivery_date": "出货日期", "quantity": "数量", "price_before_tax": "税前单价",
    "unit_price": "单价", "origin": "产地", "customer_order_seq": "客户订单序号",
    "one_to_many": "一对多", "remark": "备注",
}
REQUIRED_HEADER_FIELDS = {"order_type", "bill_to_customer_code", "ledger"}
REQUIRED_LINE_FIELDS = {"line_no", "customer_product_code", "quantity"}
DEFAULT_HEADER_VALUES = {
    "order_type": "220",
    "type_1": "1",
    "type_2": "1",
    "bill_to_customer_code": "",
    "ship_to_customer_code": "",
    "delivery_factory": "",
    "customer_order_number": "",
    "ledger": "KL01",
}
MAX_ORDER_ATTACHMENT_BYTES = 20 * 1024 * 1024
MAX_ORDER_ATTACHMENT_ROWS = 20_000

# These fields are deliberately not inferred from an order e-mail.  产品编号、品名
# will later come from Nyeos 151; 产地和一对多 are business decisions.  Leaving a
# value blank is safer than presenting an unverified guess as a usable result.
MANUAL_ONLY_LINE_FIELDS = {"product_code", "product_name", "origin", "one_to_many"}

# The source documents use different wording.  Keep this mapping here instead
# of forcing business users to normalise their customers' Excel files first.
_ATTACHMENT_HEADERS = {
    "line_no": {"序号", "项次", "项目", "行号", "item", "no"},
    "product_code": {"产品编号", "物料编号", "物料编码", "料号", "品号", "厂内料号"},
    "product_name": {"品名", "物料名称", "名称", "产品名称"},
    "customer_product_code": {"客户产品编号", "客户料号", "客户物料编号", "客户产品码", "part no", "p/n"},
    "customer_spec": {"客户规格", "规格", "型号", "名称规格", "物料规格", "物料描述"},
    "customer_spec_match": {"客户规格匹配", "规格匹配"},
    "product_type": {"产品类型", "产品类型（pp、基板）", "品类"},
    "delivery_date": {"出货日期", "交货日期", "交期", "到货日期", "delivery date"},
    "quantity": {"数量", "采购量", "订购数量", "订单数量", "qty", "quantity"},
    "_quantity_unit": {"单位", "计量单位", "数量单位", "uom"},
    "price_before_tax": {"税前单价", "未税单价", "不含税单价"},
    "unit_price": {"单价", "含税单价", "unit price", "price"},
    "origin": {"产地", "原产地"},
    "customer_order_seq": {"客户订单序号", "订单序号", "客户项次"},
    "one_to_many": {"一对多"},
    "remark": {"备注", "说明", "订单备注", "需方备注", "供方备注"},
}


def _compact_key(value: Any) -> str:
    return re.sub(r"[\s:：_\-（）()]+", "", str(value or "")).lower()


_ATTACHMENT_HEADER_INDEX = {
    _compact_key(alias): field
    for field, aliases in _ATTACHMENT_HEADERS.items()
    for alias in aliases
}


def _json(value: str, fallback: Any) -> Any:
    try:
        return json.loads(value or "")
    except (TypeError, ValueError):
        return fallback


def _case_for_template(case_id: int, employee_id: str) -> dict[str, Any]:
    case = get_case(case_id, employee_id)
    if not case:
        raise ValueError("录单邮件不存在或无权操作")
    if case.get("action_type") != "new_order":
        raise ValueError("只有已分流为“录单”的邮件才能提取到内销模板")
    return case


def _blank_line(line_no: int) -> dict[str, str]:
    return {field: str(line_no) if field == "line_no" else "" for field in LINE_FIELDS}


def _is_pp_spec(value: str) -> bool:
    """Only classify PP when the customer specification explicitly says so."""
    text = clean_text(value)
    return bool(re.search(r"(?:半固化片|(?<![A-Za-z0-9])PP(?![A-Za-z0-9]))", text, re.IGNORECASE))


def _meter_values(value: str) -> list[Decimal]:
    """Return unambiguous metre values, without mistaking 0.075MM for metres."""
    found = re.findall(r"(\d+(?:\.\d+)?)\s*(?:米|[mM])(?![A-Za-z])", clean_text(value))
    values: list[Decimal] = []
    for item in found:
        try:
            values.append(Decimal(item))
        except InvalidOperation:
            continue
    return values


def _decimal_text(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def _append_remark(remark: str, meter: Decimal | None) -> str:
    if meter is None:
        return remark
    meter_text = f"PP米数：{_decimal_text(meter)}米"
    return "；".join(item for item in (remark, meter_text) if item)


def _apply_auto_extraction_policy(
    values: dict[str, str], *, quantity_unit: Any = "", product_context: Any = "",
) -> dict[str, str]:
    """Apply the conservative domestic order-entry extraction rules.

    A value is kept only when its meaning is explicit in the source.  In
    particular, PP conversion requires an explicit PP spec, a roll unit, one
    unambiguous metre value and one numeric order quantity.  All other cases
    remain for business completion.
    """
    result = dict(values)
    for field in MANUAL_ONLY_LINE_FIELDS:
        result[field] = ""

    spec = result.get("customer_spec", "")
    # A PO often puts “半固化片” in the material-name column and the metre
    # value in the description column.  Both are explicit source evidence, but
    # only the description is shown as the customer specification.
    if not _is_pp_spec(f"{spec} {clean_text(product_context)}"):
        return result

    metres = _meter_values(spec)
    distinct_metres = {value.normalize() for value in metres}
    meter = next(iter(distinct_metres), None) if len(distinct_metres) == 1 else None
    result["remark"] = _append_remark(result.get("remark", ""), meter)

    unit = clean_text(quantity_unit)
    if "张" in unit:
        # PP 小片：客户明确以张计数，数量直接保留，不做米数换算。
        return result
    if "卷" not in unit or meter is None:
        # PP 的单位或米数不明确，不能猜测换算关系。
        result["quantity"] = ""
        return result
    try:
        quantity = Decimal(str(result.get("quantity") or "").replace(",", ""))
    except InvalidOperation:
        result["quantity"] = ""
        return result
    if quantity < 0:
        result["quantity"] = ""
        return result
    result["quantity"] = _decimal_text(quantity * meter)
    return result


def _split_body_order_rows(body_text: str) -> list[dict[str, Any]]:
    """Extract simple ERP-style rows from the line-oriented mail body.

    This is a safe first path for HTML mail tables. Attachment-specific parsers
    will feed the same structure in the next extraction layer.
    """
    lines = [" ".join(item.split()) for item in str(body_text or "").splitlines() if item.strip()]
    positions = [index for index, value in enumerate(lines) if re.fullmatch(r"(?:HJ\d{8,}|[A-Z]{1,4}\d{6,}[A-Z0-9_-]*)", value, re.I)]
    result: list[dict[str, Any]] = []
    for line_no, start in enumerate(positions, start=1):
        end = positions[line_no] if line_no < len(positions) else len(lines)
        chunk = lines[start:end]
        if len(chunk) < 2:
            continue
        customer_code = next((item for item in chunk[1:] if re.fullmatch(r"[A-Z0-9]{8,}", item, re.I)), "")
        quantity = next((item for item in reversed(chunk) if re.fullmatch(r"\d+(?:\.\d+)?", item.replace(",", ""))), "")
        material_index = next((index for index, item in enumerate(chunk) if re.fullmatch(r"NY\d+[A-Z0-9]*", item, re.I)), -1)
        spec = ""
        if material_index >= 0:
            spec = " ".join(chunk[material_index:min(len(chunk), material_index + 9)])
        if not customer_code and not spec:
            continue
        values = {
            **_blank_line(len(result) + 1),
            "customer_product_code": customer_code,
            "customer_spec": spec,
            "quantity": quantity.replace(",", ""),
            "customer_order_number": chunk[0],
        }
        result.append(_line_entry(values, label="邮件正文", reference=f"第 {start + 1} 行附近", line_no=len(result) + 1))
    return result


def _source(label: str, reference: str, values: dict[str, str]) -> dict[str, dict[str, str]]:
    return {field: {"label": label, "reference": reference} for field, value in values.items() if value and field != "line_no"}


def _line_entry(
    values: dict[str, Any], *, label: str, reference: str, line_no: int = 1, quantity_unit: Any = "", product_context: Any = "",
) -> dict[str, Any]:
    line = _blank_line(line_no)
    for field in LINE_FIELDS:
        value = values.get(field)
        if value not in (None, ""):
            line[field] = clean_text(value)
    line["line_no"] = str(line_no)
    if line["delivery_date"]:
        line["delivery_date"] = normalize_date(line["delivery_date"]) or line["delivery_date"]
    for field in {"quantity", "price_before_tax", "unit_price"}:
        if line[field]:
            line[field] = normalize_number(line[field]) or line[field]
    line = _apply_auto_extraction_policy(line, quantity_unit=quantity_unit, product_context=product_context)
    return {"values": line, "sources": _source(label, reference, line)}


def _value_by_alias(mapping: dict[str, Any], *aliases: str) -> Any:
    """Return an explicit source value even when a document uses bilingual headings."""
    if not mapping:
        return ""
    wanted = {_compact_key(alias) for alias in aliases}
    for key, value in mapping.items():
        if _compact_key(key) in wanted:
            return value
    for key, value in mapping.items():
        compact = _compact_key(key)
        if any(alias and alias in compact for alias in wanted):
            return value
    return ""


def _tax_inclusive_unit_price(mapping: dict[str, Any]) -> Any:
    excluded_markers = {
        "不含税",
        "未税",
        "税前",
        "nottaxinclusive",
        "taxexclusive",
        "excludingtax",
        "excltax",
        "withouttax",
    }
    eligible = {
        key: value
        for key, value in mapping.items()
        if not any(marker in _compact_key(key) for marker in excluded_markers)
    }
    return _value_by_alias(eligible, "含税单价", "单价", "Unit Price")


def _mapping_source_value(source: dict[str, Any], source_label: str, transform_type: str) -> str:
    """Read explicitly named attachment columns for one customer mapping.

    The mapping UI intentionally uses the customer's visible column names
    (for example ``Material Code``), not technical column indices.  Concatenation
    accepts ``+`` or Chinese/English commas.  Missing parts make the whole
    result blank: the system must never present a partially guessed value.
    """
    labels = [item.strip() for item in re.split(r"[+，,]", str(source_label or "")) if item.strip()]
    if not labels:
        return ""
    values = [clean_text(_value_by_alias(source, label)) for label in labels]
    if not all(values):
        return ""
    if transform_type == "concat":
        return " ".join(values)
    return values[0] if len(values) == 1 else ""


def _apply_customer_extraction_mappings(
    values: dict[str, Any], source: dict[str, Any], mappings: list[dict[str, Any]],
    *, source_kind: str = "attachment_table",
) -> dict[str, Any]:
    """Apply customer mappings registered for the current source adapter.

    Manual mappings explicitly blank the target.  For all other mapping source
    types we leave the universal result untouched.  ``source_kind`` is kept
    explicit so an eventual rule-maintenance page can expose different source
    adapters without making their mappings leak into one another.
    """
    if not mappings:
        return values
    result = dict(values)
    for mapping in mappings:
        if mapping.get("source_kind") != source_kind:
            continue
        target = str(mapping.get("target_field") or "")
        transform = str(mapping.get("transform_type") or "")
        if target not in LINE_FIELDS:
            continue
        if transform == "manual":
            result[target] = ""
            continue
        value = _mapping_source_value(source, str(mapping.get("source_label") or ""), transform)
        # A configured customer mapping takes precedence only when it produces
        # an explicit source value.  Otherwise retain an already verified
        # universal value; no fallback guess is introduced.
        if value:
            result[target] = value
    return result


def _canonical_order_number(header: dict[str, Any]) -> str:
    """Keep the actual PO token and discard surrounding document decorations."""
    for key in (
        "客户订单号",
        "订单号",
        "订单编号",
        "采购订单号",
        "合同编号",
        "PO号",
        "PO No",
    ):
        raw = clean_text(_value_by_alias(header, key))
        if not raw:
            continue
        candidates = re.findall(r"[A-Za-z]{1,10}[A-Za-z0-9_-]*\d{4,}[A-Za-z0-9_-]*", raw)
        if candidates:
            return max(candidates, key=len).upper()
        if re.fullmatch(r"[A-Za-z0-9_-]{6,}", raw):
            return raw
    return ""


def _within_order_attachment_size(path: Path) -> bool:
    """Reject real oversized files while keeping parser adapters mockable."""
    try:
        return path.stat().st_size <= MAX_ORDER_ATTACHMENT_BYTES
    except OSError:
        # The recognizer can operate on a virtual/test adapter path; its own
        # error handling remains responsible for unavailable source files.
        return True


def _rows_from_excel(path: Path, filename: str, customer_mappings: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    """Read customer Excel attachments as structured data without OCR."""
    if not _within_order_attachment_size(path):
        return []
    try:
        book = load_workbook_compat(path, data_only=True)
    except Exception:
        return []
    try:
        result: list[dict[str, Any]] = []
        scanned_rows = 0
        for sheet in book.worksheets:
            rows: list[tuple[Any, ...]] = []
            for row in sheet.iter_rows(values_only=True):
                scanned_rows += 1
                if scanned_rows > MAX_ORDER_ATTACHMENT_ROWS:
                    return []
                rows.append(row)
            header_index = -1
            mapping: dict[int, str] = {}
            for index, row in enumerate(rows[:40]):
                candidate = {
                    column: _ATTACHMENT_HEADER_INDEX.get(_compact_key(value))
                    for column, value in enumerate(row) if _ATTACHMENT_HEADER_INDEX.get(_compact_key(value))
                }
                # A single heading such as "规格" is too ambiguous; require a
                # usable set of columns before treating it as an order table.
                if len(set(candidate.values())) >= 2 and ("quantity" in candidate.values() or "customer_product_code" in candidate.values()):
                    header_index, mapping = index, candidate
                    break
            if header_index < 0:
                continue
            for index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                values = {field: row[column] if column < len(row) else "" for column, field in mapping.items()}
                if not any(clean_text(value) for value in values.values()):
                    continue
                text = " ".join(clean_text(value) for value in values.values())
                if any(token in text for token in ("合计", "总计", "小计")):
                    continue
                source_row = {
                    clean_text(rows[header_index][column]): row[column] if column < len(row) else ""
                    for column in range(len(rows[header_index])) if clean_text(rows[header_index][column])
                }
                values = _apply_customer_extraction_mappings(values, source_row, customer_mappings or [])
                result.append(_line_entry(
                    values,
                    label=f"附件：{filename}",
                    reference=f"{sheet.title} 第 {index} 行",
                    line_no=len(result) + 1,
                    quantity_unit=values.get("_quantity_unit", ""),
                ))
        return result
    finally:
        book.close()


def _line_from_pipeline_row(
    row: dict[str, Any], order_number: str, label: str, reference: str, line_no: int,
    customer_mappings: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    original = row.get("original") or {}
    standard = row.get("standard") or {}
    raw_spec = _value_by_alias(original, "物料描述", "Material Description", "名称规格", "客户规格", "物料规格", "规格", "型号")
    raw_before_tax_price = _value_by_alias(original, "不含税单价", "未税单价", "税前单价", "Not tax inclusive Unit Price")
    raw_unit_price = _tax_inclusive_unit_price(original)
    raw_quantity_unit = _value_by_alias(original, "单位", "计量单位", "Unit", "UOM")
    raw_material_name = _value_by_alias(original, "物料品名", "物料名称", "Material Name") or standard.get("物料名称") or ""
    values = {
        "line_no": standard.get("序号") or _value_by_alias(original, "序号", "No") or line_no,
        # 采购订单中的“物料编码 / Material Code”是客户提供的明确料号，
        # 可安全写入内销模板的“客户产品编号”；不从品名或规格推测编码。
        "customer_product_code": (
            _value_by_alias(
                original,
                "客户产品编号",
                "客户料号",
                "客户物料编号",
                "客户产品码",
                "物料编码",
                "Material Code",
            )
            or standard.get("物料编码")
            or ""
        ),
        "customer_spec": raw_spec or standard.get("说明") or standard.get("物料名称") or "",
        "delivery_date": standard.get("交货日期") or _value_by_alias(original, "交货日期", "出货日期", "交期", "Delivery Date") or "",
        "quantity": standard.get("数量") or _value_by_alias(original, "数量", "Quantity", "Qty") or "",
        "price_before_tax": raw_before_tax_price or standard.get("不含税单价") or "",
        "unit_price": raw_unit_price or "",
        "customer_order_number": order_number,
        "remark": standard.get("备注") or _value_by_alias(original, "备注", "说明", "订单备注") or "",
    }
    values = _apply_customer_extraction_mappings(values, original, customer_mappings or [])
    entry = _line_entry(
        values,
        label=label,
        reference=reference,
        line_no=line_no,
        quantity_unit=raw_quantity_unit or standard.get("单位") or "",
        product_context=raw_material_name,
    )
    # Similar PP rows may intentionally retain a blank quantity. Their source
    # row identity still makes them distinct order lines and prevents collapse.
    entry["_source_identity"] = "|".join(clean_text(value) for value in (
        label, reference,
        _value_by_alias(original, "PO项目号", "Project No", "项目号"),
        _value_by_alias(original, "物料编码", "Material Code"), raw_spec,
    ))
    return entry


def _rows_from_pdf_or_image(path: Path, filename: str, customer_mappings: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    if not _within_order_attachment_size(path):
        return []
    try:
        document = recognize_purchase_order_document(
            {"stored_path": str(path), "original_filename": filename}
        )
        return _rows_from_shared_purchase_document(
            document,
            label=f"附件：{filename}",
            reference_prefix="识别明细",
            customer_mappings=customer_mappings,
        )
    except Exception as exc:
        # A PDF/image order must never silently fall back to the lower-fidelity
        # mail-body parser.  That used to look like a successful refresh while
        # producing different rows from PDF/图片转Excel.
        raise ValueError(f"附件《{filename}》未能按 PDF/图片转Excel 规则提取：{exc}") from exc


def _rows_from_shared_purchase_document(
    document: dict[str, Any], *, label: str, reference_prefix: str,
    customer_mappings: list[dict[str, Any]] | None = None,
    source_kind: str = "attachment_table",
) -> list[dict[str, Any]]:
    """Project one canonical purchase document into editable domestic rows.

    Every source adapter (PDF/image, HTML mail tables, and later Excel) must
    pass this boundary.  It deliberately delegates to the same factory
    projection and domestic-template exporter used by PDF/图片转Excel.
    """
    project_factory_document(document)
    domestic_data = build_domestic_template_data(document)
    result: list[dict[str, Any]] = []
    source_rows = document.get("mapped_detail_rows") or []
    for index, (line, source_row) in enumerate(
        zip(domestic_data["lines"], source_rows), start=1
    ):
        values = {field: clean_text(line.get(field)) for field in LINE_FIELDS}
        values["line_no"] = str(index)
        values = _apply_customer_extraction_mappings(
            values,
            source_row.get("original") or {},
            customer_mappings or [],
            source_kind=source_kind,
        )
        reference = f"{reference_prefix}第 {index} 行"
        result.append(
            {
                "values": values,
                "sources": _source(label, reference, values),
                "_source_identity": "|".join(
                    clean_text(value)
                    for value in (
                        label,
                        reference,
                        values.get("customer_product_code"),
                        values.get("customer_spec"),
                    )
                ),
                "extracted_header": domestic_data["header"],
            }
        )
    return result


def _rows_from_mail_html(case: dict[str, Any]) -> list[dict[str, Any]]:
    """Use HTML mail order tables as another input to the common PO model.

    A non-table mail body intentionally returns no rows so the established
    line-oriented body fallback remains available.  Once a table is recognised
    as an order table, failures are surfaced instead of quietly substituting a
    lower-fidelity parser.
    """
    document = build_mail_html_purchase_document(
        str(case.get("body_html") or ""),
        str(case.get("body_text") or ""),
        source_name=f"邮件正文#{case.get('id') or ''}.html",
    )
    if not document:
        return []
    try:
        return _rows_from_shared_purchase_document(
            document,
            label="邮件正文表格",
            reference_prefix="表格明细",
            customer_mappings=get_enabled_extraction_maps(case.get("customer_id")),
            source_kind="mail_html_table",
        )
    except Exception as exc:
        raise ValueError(f"邮件正文表格未能按统一订单规则提取：{exc}") from exc


def _rows_from_word(path: Path, filename: str, customer_mappings: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    """Convert Word attachments locally, then use the same PDF/OCR pipeline."""
    if not _within_order_attachment_size(path):
        return []
    soffice = shutil.which("soffice")
    if not soffice:
        return []
    with tempfile.TemporaryDirectory(prefix="order-entry-word-") as output_dir:
        try:
            subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", output_dir, str(path)],
                check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=45,
            )
        except (OSError, subprocess.SubprocessError):
            return []
        pdf = Path(output_dir) / f"{path.stem}.pdf"
        return _rows_from_pdf_or_image(pdf, filename, customer_mappings=customer_mappings) if pdf.is_file() else []


def _attachment_rows(case: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    customer_mappings = get_enabled_extraction_maps(case.get("customer_id"))
    for attachment in case.get("attachments") or []:
        if attachment.get("is_inline"):
            continue
        try:
            path = resolve_attachment_path(str(attachment.get("stored_path") or ""))
        except FileNotFoundError:
            continue
        if not path.is_file():
            continue
        filename, suffix = str(attachment.get("filename") or path.name), path.suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xls"}:
            rows.extend(_rows_from_excel(path, filename, customer_mappings=customer_mappings))
        elif suffix in {".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}:
            rows.extend(_rows_from_pdf_or_image(path, filename, customer_mappings=customer_mappings))
        elif suffix in {".doc", ".docx"}:
            rows.extend(_rows_from_word(path, filename, customer_mappings=customer_mappings))
    return rows


def _line_signature(entry: dict[str, Any]) -> tuple[str, ...]:
    source_identity = _compact_key(entry.get("_source_identity"))
    if source_identity:
        return ("source", source_identity)
    values = entry["values"]
    return tuple(
        _compact_key(values.get(field))
        for field in ("customer_order_seq", "customer_product_code", "customer_spec", "quantity")
    )


def _merge_initial_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[tuple[str, ...]] = set()
    for entry in rows:
        signature = _line_signature(entry)
        # Do not collapse rows that have no identifying data: a blank source
        # row is still useful to the business user for manual completion.
        if any(signature) and signature in seen:
            continue
        if any(signature):
            seen.add(signature)
        entry["values"]["line_no"] = str(len(result) + 1)
        result.append(entry)
    return result


def _apply_customer_spec_matches(
    header: dict[str, Any], lines: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    customer_code = clean_text(header.get("bill_to_customer_code"))
    result: list[dict[str, Any]] = []
    for entry in lines:
        values = dict(entry.get("values") or {})
        sources = dict(entry.get("sources") or {})
        match_context = json.dumps([
            customer_code,
            clean_text(values.get("product_type")),
            clean_text(values.get("customer_spec")),
        ], ensure_ascii=False, separators=(",", ":"))
        existing_source = sources.get("customer_spec_match") or {}
        manual_match = (
            customer_code
            and existing_source.get("label") == "人工修改"
            and existing_source.get("context") == match_context
        )
        if manual_match:
            result.append({**entry, "values": values, "sources": sources})
            continue
        matched = build_customer_spec_match(
            customer_code,
            values.get("product_type"),
            values.get("customer_spec"),
        ) if customer_code else ""
        values["customer_spec_match"] = matched
        if matched:
            sources["customer_spec_match"] = {
                "label": "客户规格对照表",
                "reference": f"客户编号 {customer_code} / {values.get('product_type') or '未填写产品类型'}",
                "context": match_context,
            }
        else:
            sources.pop("customer_spec_match", None)
        result.append({**entry, "values": values, "sources": sources})
    return result


def _apply_matched_customer_code(
    header: dict[str, str], case: dict[str, Any], *, overwrite: bool,
) -> dict[str, str]:
    customer_id = case.get("customer_id")
    if not customer_id:
        return header
    customer = get_customer(int(customer_id))
    customer_code = clean_text((customer or {}).get("customer_code"))
    if not customer_code:
        return header
    bill_to_code = clean_text(header.get("bill_to_customer_code"))
    if overwrite or not bill_to_code:
        bill_to_code = customer_code
        header["bill_to_customer_code"] = bill_to_code
    if overwrite or not clean_text(header.get("ship_to_customer_code")):
        header["ship_to_customer_code"] = bill_to_code
    return header


def _initial_template_data(case: dict[str, Any]) -> tuple[dict[str, str], list[dict[str, Any]]]:
    attachment_rows = _attachment_rows(case)
    # An order attachment remains authoritative.  When there is no supported
    # attachment, an HTML mail table gets the *same* canonical mapping and
    # validation path; only then use the legacy line-oriented body fallback.
    rows = attachment_rows or _rows_from_mail_html(case) or _split_body_order_rows(str(case.get("body_text") or ""))
    rows = _merge_initial_rows(rows)
    header = dict(DEFAULT_HEADER_VALUES)
    extracted_header = next(
        (item.get("extracted_header") for item in rows if item.get("extracted_header")),
        {},
    )
    for field in HEADER_FIELDS:
        value = clean_text((extracted_header or {}).get(field))
        if value:
            header[field] = value
    _apply_matched_customer_code(header, case, overwrite=True)
    if not header["customer_order_number"]:
        header["customer_order_number"] = clean_text(
            (case.get("detected_fields") or {}).get("order_number")
        )
    if rows:
        return header, _apply_customer_spec_matches(header, rows)
    fields = case.get("detected_fields") or {}
    specs = fields.get("specs") or []
    header["customer_order_number"] = clean_text(fields.get("order_number"))
    lines = [
        {
            "values": {
                **_blank_line(index + 1),
                "customer_spec": str(spec),
            },
            "sources": {"customer_spec": {"label": "邮件正文", "reference": "自动识别"}},
        }
        for index, spec in enumerate(specs)
    ] or [{"values": _blank_line(1), "sources": {}}]
    return header, _apply_customer_spec_matches(header, lines)


def _initial_lines(case: dict[str, Any]) -> list[dict[str, Any]]:
    return _initial_template_data(case)[1]


def _serialize_template(conn, template_id: int) -> dict[str, Any]:
    template = conn.execute("SELECT * FROM order_entry_templates WHERE id=?", (template_id,)).fetchone()
    if not template:
        raise ValueError("录单模板不存在")
    rows = conn.execute(
        "SELECT * FROM order_entry_template_lines WHERE template_id=? ORDER BY line_no,id", (template_id,)
    ).fetchall()
    result = {
        **dict(template),
        "header": {**DEFAULT_HEADER_VALUES, **_json(template["header_json"], {})},
        "lines": [
            {"id": row["id"], "line_no": row["line_no"], "values": _json(row["values_json"], {}), "sources": _json(row["sources_json"], {})}
            for row in rows
        ],
    }
    result["lines"] = _apply_customer_spec_matches(result["header"], result["lines"])
    return result


def get_or_create_template(case_id: int, employee_id: str) -> tuple[dict[str, Any], dict[str, Any]]:
    case = _case_for_template(case_id, employee_id)
    with db_cursor() as conn:
        existing = conn.execute(
            "SELECT id FROM order_entry_templates WHERE case_id=? AND employee_id=?", (case_id, employee_id)
        ).fetchone()
        if existing:
            template_id = int(existing["id"])
            template = _serialize_template(conn, template_id)
            previous_header = dict(template["header"])
            _apply_matched_customer_code(template["header"], case, overwrite=False)
            if template["header"] != previous_header:
                conn.execute(
                    "UPDATE order_entry_templates SET header_json=?,updated_at=? WHERE id=?",
                    (json.dumps(template["header"], ensure_ascii=False), utcnow(), template_id),
                )
                template = _serialize_template(conn, template_id)
            return case, template
        now = utcnow()
        initial_header, initial_lines = _initial_template_data(case)
        cursor = conn.execute(
            "INSERT INTO order_entry_templates(case_id,employee_id,header_json,created_at,updated_at) VALUES (?,?,?,?,?)",
            (case_id, employee_id, json.dumps(initial_header, ensure_ascii=False), now, now),
        )
        template_id = int(cursor.lastrowid)
        for entry in initial_lines:
            values = entry["values"]
            conn.execute(
                "INSERT INTO order_entry_template_lines(template_id,line_no,values_json,sources_json,created_at,updated_at) VALUES (?,?,?,?,?,?)",
                (template_id, int(values["line_no"] or 0), json.dumps(values, ensure_ascii=False), json.dumps(entry["sources"], ensure_ascii=False), now, now),
            )
        record_order_detail_event(
            conn,
            case_id=case_id,
            template_id=template_id,
            employee_id=employee_id,
            event_type="template_extracted",
            title="已提取内销录单模板",
            detail={"line_count": len(initial_lines), "source": "邮件正文和附件"},
            operated_by=employee_id,
        )
        return case, _serialize_template(conn, template_id)


def _template_task_row(case_id: int, employee_id: str) -> dict[str, Any] | None:
    """Return the latest extraction task for one employee-owned mail case."""
    with db_cursor() as conn:
        row = conn.execute(
            """SELECT * FROM order_entry_template_tasks
               WHERE case_id=? AND employee_id=?
               ORDER BY id DESC LIMIT 1""",
            (case_id, employee_id),
        ).fetchone()
    return dict(row) if row else None


def queue_template_extraction(case_id: int, employee_id: str) -> dict[str, Any]:
    """Create one background extraction task without doing document work in HTTP.

    The original attachment can require native PDF parsing or OCR.  Keeping that
    work out of the request is important: a click should always return at once,
    while the case page can accurately show the durable task state.
    """
    _case_for_template(case_id, employee_id)
    with db_cursor() as conn:
        template = conn.execute(
            "SELECT id FROM order_entry_templates WHERE case_id=? AND employee_id=?",
            (case_id, employee_id),
        ).fetchone()
        if template:
            return {
                "status": "completed", "task_id": None, "template_id": int(template["id"]),
                "message": "内销模板已生成，可直接打开核对。",
            }
        active = conn.execute(
            """SELECT * FROM order_entry_template_tasks
               WHERE case_id=? AND employee_id=? AND status IN ('queued', 'running')
               ORDER BY id DESC LIMIT 1""",
            (case_id, employee_id),
        ).fetchone()
        if active:
            return {
                "status": str(active["status"]), "task_id": int(active["id"]),
                "template_id": None, "message": "订单信息正在后台提取，请稍候刷新。",
            }
        now = utcnow()
        cursor = conn.execute(
            """INSERT INTO order_entry_template_tasks
               (case_id,employee_id,status,message,started_at)
               VALUES (?,?,?,?,?)""",
            (case_id, employee_id, "queued", "等待后台提取订单附件", now),
        )
        task_id = int(cursor.lastrowid)

    command = [
        sys.executable, "-m", "fangzheng_web_app.order_entry_template_worker",
        "--task-id", str(task_id), "--case-id", str(case_id),
        "--employee-id", employee_id,
    ]
    try:
        subprocess.Popen(
            command,
            cwd=str(PROJECT_DIR),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            start_new_session=True,
        )
    except Exception as exc:
        _complete_template_extraction_task(
            task_id, status="error", message=f"无法启动后台提取：{exc}", template_id=None,
        )
        raise ValueError("无法启动后台订单提取，请稍后重试") from exc
    return {
        "status": "queued", "task_id": task_id, "template_id": None,
        "message": "已开始后台提取订单信息，完成后会自动显示内销模板。",
    }


def _complete_template_extraction_task(
    task_id: int, *, status: str, message: str, template_id: int | None,
) -> None:
    with db_cursor() as conn:
        conn.execute(
            """UPDATE order_entry_template_tasks
               SET status=?, message=?, template_id=?, completed_at=? WHERE id=?""",
            (status, message, template_id, utcnow(), task_id),
        )


def run_template_extraction_task(task_id: int, case_id: int, employee_id: str) -> None:
    """Worker entry point.  This process owns the expensive first extraction."""
    with db_cursor() as conn:
        updated = conn.execute(
            """UPDATE order_entry_template_tasks
               SET status='running', message='正在解析邮件正文和附件', started_at=?
               WHERE id=? AND case_id=? AND employee_id=? AND status='queued'""",
            (utcnow(), task_id, case_id, employee_id),
        )
    if not updated.rowcount:
        return
    try:
        _case, template = get_or_create_template(case_id, employee_id)
        _complete_template_extraction_task(
            task_id,
            status="completed",
            message="订单信息已提取，请核对并保存内销模板。",
            template_id=int(template["id"]),
        )
    except Exception as exc:
        _complete_template_extraction_task(
            task_id,
            status="error",
            message=f"订单提取失败：{str(exc)[:240]}",
            template_id=None,
        )


def _replace_backup(
    conn,
    *,
    template_id: int,
    header: dict[str, Any],
    lines: list[dict[str, Any]],
    employee_id: str,
    saved_at: str,
) -> None:
    """Keep exactly one recoverable snapshot before replacing a template."""
    conn.execute("DELETE FROM order_entry_template_versions WHERE template_id=?", (template_id,))
    conn.execute(
        "INSERT INTO order_entry_template_versions(template_id,version_number,header_json,lines_json,saved_by,saved_at) VALUES (?,?,?,?,?,?)",
        (
            template_id,
            1,
            json.dumps(header, ensure_ascii=False),
            json.dumps(lines, ensure_ascii=False),
            employee_id,
            saved_at,
        ),
    )


def reextract_template(case_id: int, employee_id: str) -> dict[str, Any]:
    """Rebuild one saved template's detail rows with the current extraction rules.

    The customer/header section is business-maintained and is intentionally left
    untouched. Before replacing the current detail rows, the immediately prior
    contents replace the single backup snapshot.
    """
    case = _case_for_template(case_id, employee_id)
    with db_cursor() as conn:
        row = conn.execute(
            "SELECT id,current_version FROM order_entry_templates WHERE case_id=? AND employee_id=?",
            (case_id, employee_id),
        ).fetchone()
        if not row:
            raise ValueError("请先打开录单模板")
        template_id = int(row["id"])
        previous = _serialize_template(conn, template_id)

    # Recognition can involve OCR and file conversion, so do it outside of the
    # database transaction.  It only reads the original mail and attachments.
    regenerated_header, regenerated_lines = _initial_template_data(case)
    now = utcnow()
    previous_lines = [
        {"values": line.get("values") or {}, "sources": line.get("sources") or {}}
        for line in previous.get("lines") or []
    ]
    previous_header = {**DEFAULT_HEADER_VALUES, **(previous.get("header") or {})}
    # Refresh adopts the PDF/图片转Excel template defaults and its extracted
    # order number, while retaining any customer information the business user
    # has already entered in this mail workspace.
    next_header = {
        **regenerated_header,
        **{field: value for field, value in previous_header.items() if clean_text(value)},
    }
    backup_version = 1
    current_version = 1

    with db_cursor() as conn:
        _replace_backup(
            conn,
            template_id=template_id,
            header=previous_header,
            lines=previous_lines,
            employee_id=employee_id,
            saved_at=now,
        )
        conn.execute(
            "UPDATE order_entry_templates SET header_json=?,current_version=?,updated_at=? WHERE id=?",
            (json.dumps(next_header, ensure_ascii=False), current_version, now, template_id),
        )
        conn.execute("DELETE FROM order_entry_template_lines WHERE template_id=?", (template_id,))
        for entry in regenerated_lines:
            values = entry["values"]
            conn.execute(
                "INSERT INTO order_entry_template_lines(template_id,line_no,values_json,sources_json,created_at,updated_at) VALUES (?,?,?,?,?,?)",
                (
                    template_id,
                    int(values["line_no"] or 0),
                    json.dumps(values, ensure_ascii=False),
                    json.dumps(entry["sources"], ensure_ascii=False),
                    now,
                    now,
                ),
            )
        record_order_detail_event(
            conn,
            case_id=case_id,
            template_id=template_id,
            employee_id=employee_id,
            event_type="template_reextracted",
            title="已重新提取订单明细",
            detail={"previous_line_count": len(previous_lines), "line_count": len(regenerated_lines)},
            operated_by=employee_id,
        )
        template = _serialize_template(conn, template_id)

    return {
        "case_id": case_id,
        "subject": str(case.get("subject") or ""),
        "previous_line_count": len(previous_lines),
        "line_count": len(template.get("lines") or []),
        "backup_version": backup_version,
        "current_version": current_version,
        "template": template,
    }


def reextract_all_templates(employee_id: str) -> dict[str, Any]:
    """Batch re-extract every existing domestic template owned by one user."""
    with db_cursor() as conn:
        rows = conn.execute(
            """SELECT t.case_id
               FROM order_entry_templates t
               JOIN order_intake_cases c ON c.id=t.case_id
               WHERE t.employee_id=? AND c.action_type='new_order'
               ORDER BY t.id""",
            (employee_id,),
        ).fetchall()
    results = [reextract_template(int(row["case_id"]), employee_id) for row in rows]
    return {
        "template_count": len(results),
        "previous_line_count": sum(item["previous_line_count"] for item in results),
        "line_count": sum(item["line_count"] for item in results),
        "results": results,
    }


def template_progress(case_id: int, employee_id: str) -> dict[str, Any]:
    """Return the single read-only workflow state used by every order view.

    A successful domestic-entry interface call is the terminal business fact.
    A completed order has a saved template too, but must not be shown as still
    waiting for an interface submission on another page.
    """
    with db_cursor() as conn:
        row = conn.execute(
            "SELECT current_version FROM order_entry_templates WHERE case_id=? AND employee_id=?",
            (case_id, employee_id),
        ).fetchone()
        completed = conn.execute(
            """SELECT 1 FROM order_interface_call_logs
               WHERE case_id=? AND employee_id=? AND interface_key='domestic_order_entry'
                 AND status='success'
               LIMIT 1""",
            (case_id, employee_id),
        ).fetchone()
    version = int(row["current_version"] or 0) if row else 0
    if completed:
        return {
            "created": bool(row), "saved": bool(row and version > 0), "completed": True,
            "version": version, "stage": "completed", "label": "已完成",
            "next_action": "录单已完成", "step": 5,
        }
    if not row:
        task = _template_task_row(case_id, employee_id)
        if task and task.get("status") in {"queued", "running"}:
            return {
                "created": False, "saved": False, "completed": False, "version": 0,
                "stage": "extracting", "label": "正在提取订单",
                "next_action": "正在后台解析附件…", "step": 2,
                "task_id": task["id"], "task_status": task["status"],
                "task_message": task.get("message") or "正在准备订单信息。",
            }
        if task and task.get("status") == "error":
            return {
                "created": False, "saved": False, "completed": False, "version": 0,
                "stage": "extraction_error", "label": "提取失败",
                "next_action": "重新提取订单", "step": 2,
                "task_id": task["id"], "task_status": "error",
                "task_message": task.get("message") or "请重新提取订单信息。",
            }
        return {
            "created": False, "saved": False, "completed": False, "version": 0,
            "stage": "pending_extraction", "label": "待提取订单",
            "next_action": "提取订单到内销模板", "step": 2,
        }
    if version <= 0:
        return {
            "created": True, "saved": False, "completed": False, "version": 0,
            "stage": "pending_template_save", "label": "待保存模板",
            "next_action": "核对并保存内销模板", "step": 3,
        }
    return {
        "created": True, "saved": True, "completed": False, "version": version,
        "stage": "pending_interface_submit", "label": "订单信息确认",
        "next_action": "订单信息确认", "step": 4,
    }


def _clean_values(values: dict[str, Any], fields: tuple[str, ...]) -> dict[str, str]:
    return {field: str(values.get(field) or "").strip() for field in fields}


def _template_changes(
    previous_header: dict[str, Any],
    previous_lines: list[dict[str, Any]],
    header: dict[str, Any],
    lines: list[dict[str, Any]],
) -> list[dict[str, str]]:
    changes: list[dict[str, str]] = []
    for field in HEADER_FIELDS:
        before = str(previous_header.get(field) or "")
        after = str(header.get(field) or "")
        if before != after:
            changes.append({"field": HEADER_LABELS[field], "before": before, "after": after, "scope": "表头"})
    old_by_line = {
        str((entry.get("values") or {}).get("line_no") or index): entry.get("values") or {}
        for index, entry in enumerate(previous_lines, start=1)
    }
    for index, entry in enumerate(lines, start=1):
        values = entry.get("values") or {}
        before_values = old_by_line.get(str(values.get("line_no") or index), {})
        for field in LINE_FIELDS:
            if field == "line_no":
                continue
            before = str(before_values.get(field) or "")
            after = str(values.get(field) or "")
            if before != after:
                changes.append({
                    "field": LINE_LABELS[field], "before": before, "after": after,
                    "scope": f"第 {values.get('line_no') or index} 行",
                })
    return changes[:120]


def save_template(case_id: int, employee_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    _case_for_template(case_id, employee_id)
    raw_header = payload.get("header") or {}
    if not isinstance(raw_header, dict):
        raise ValueError("表头格式无效")
    header = dict(DEFAULT_HEADER_VALUES)
    # The browser submits every header field, so an explicitly cleared input
    # remains blank.  Programmatic callers that omit a field keep the PDF
    # template default instead of silently replacing it with an empty value.
    for field in HEADER_FIELDS:
        if field in raw_header:
            header[field] = str(raw_header.get(field) or "").strip()
    raw_lines = payload.get("lines") or []
    if not isinstance(raw_lines, list):
        raise ValueError("明细行格式无效")
    lines: list[dict[str, Any]] = []
    for index, raw in enumerate(raw_lines, start=1):
        values = _clean_values((raw or {}).get("values") or raw or {}, LINE_FIELDS)
        if not any(values[field] for field in LINE_FIELDS if field != "line_no"):
            continue
        values["line_no"] = str(index)
        sources = (raw or {}).get("sources") or {}
        lines.append({"values": values, "sources": sources if isinstance(sources, dict) else {}})
    if not lines:
        lines = [{"values": _blank_line(1), "sources": {}}]
    lines = _apply_customer_spec_matches(header, lines)
    now = utcnow()
    with db_cursor() as conn:
        template = conn.execute(
            "SELECT id,current_version FROM order_entry_templates WHERE case_id=? AND employee_id=?", (case_id, employee_id)
        ).fetchone()
        if not template:
            raise ValueError("请先打开录单模板")
        template_id = int(template["id"])
        previous = _serialize_template(conn, template_id)
        previous_header = {**DEFAULT_HEADER_VALUES, **(previous.get("header") or {})}
        previous_lines = [
            {"values": line.get("values") or {}, "sources": line.get("sources") or {}}
            for line in previous.get("lines") or []
        ]
        changes = _template_changes(previous_header, previous_lines, header, lines)
        _replace_backup(
            conn,
            template_id=template_id,
            header=previous_header,
            lines=previous_lines,
            employee_id=employee_id,
            saved_at=now,
        )
        conn.execute("UPDATE order_entry_templates SET header_json=?,current_version=?,updated_at=? WHERE id=?", (json.dumps(header, ensure_ascii=False), 1, now, template_id))
        conn.execute("DELETE FROM order_entry_template_lines WHERE template_id=?", (template_id,))
        for entry in lines:
            values = entry["values"]
            conn.execute(
                "INSERT INTO order_entry_template_lines(template_id,line_no,values_json,sources_json,created_at,updated_at) VALUES (?,?,?,?,?,?)",
                (template_id, int(values["line_no"]), json.dumps(values, ensure_ascii=False), json.dumps(entry["sources"], ensure_ascii=False), now, now),
            )
            # A business user can explicitly complete a problematic material
            # number. That decision wins over a pending/failed external task,
            # and makes the later recording validation meaningful.
            if str(values.get("product_code") or "").strip() not in {"", "创建料号中"}:
                conn.execute(
                    """UPDATE order_material_resolution_tasks
                       SET status='manual_resolved',updated_at=?
                       WHERE template_id=? AND line_no=?
                         AND status IN ('waiting_callback','requerying','failed')""",
                    (now, template_id, int(values["line_no"])),
                )
        record_order_detail_event(
            conn,
            case_id=case_id,
            template_id=template_id,
            employee_id=employee_id,
            event_type="template_saved",
            title="保存内销录单模板",
            detail={"changes": changes, "line_count": len(lines)},
            operated_by=employee_id,
        )
        return _serialize_template(conn, template_id)


def validation_issues(template: dict[str, Any]) -> list[str]:
    issues = [f"{HEADER_LABELS[field]}未填写" for field in REQUIRED_HEADER_FIELDS if not str(template["header"].get(field) or "").strip()]
    for line in template.get("lines") or []:
        for field in REQUIRED_LINE_FIELDS:
            if not str((line.get("values") or {}).get(field) or "").strip():
                issues.append(f"第 {line.get('line_no')} 行{LINE_LABELS[field]}未填写")
    return issues


def build_domestic_export(case_id: int, employee_id: str) -> tuple[BytesIO, str]:
    _case_for_template(case_id, employee_id)
    with db_cursor() as conn:
        row = conn.execute(
            "SELECT id FROM order_entry_templates WHERE case_id=? AND employee_id=?", (case_id, employee_id)
        ).fetchone()
        if not row:
            raise ValueError("请先保存模板内容后再下载")
        template = _serialize_template(conn, int(row["id"]))
    if int(template.get("current_version") or 0) <= 0:
        raise ValueError("请先保存模板内容后再下载")
    if not DOMESTIC_TEMPLATE_PATH.is_file():
        raise ValueError("内销录单模板文件未配置")
    book = load_workbook(DOMESTIC_TEMPLATE_PATH)
    sheet = book["内销"]
    for index, field in enumerate(HEADER_FIELDS, start=1):
        sheet.cell(2, index).value = template["header"].get(field) or None
    required_rows = 3 + max(1, len(template["lines"]))
    style_source_row = 4
    while sheet.max_row < required_rows:
        target = sheet.max_row + 1
        for col in range(1, 16):
            source, cell = sheet.cell(style_source_row, col), sheet.cell(target, col)
            cell._style = copy(source._style)
            cell.number_format = source.number_format
    for row in range(4, max(sheet.max_row, required_rows) + 1):
        values = (template["lines"][row - 4]["values"] if row - 4 < len(template["lines"]) else {})
        for col, field in enumerate(LINE_FIELDS, start=1):
            value = values.get(field) or None
            if field in {"quantity", "price_before_tax", "unit_price"} and value not in (None, ""):
                try:
                    value = float(value)
                except ValueError:
                    pass
            if field == "delivery_date" and value:
                try:
                    value = datetime.fromisoformat(str(value)).date()
                except ValueError:
                    pass
            sheet.cell(row, col).value = value
    data = BytesIO()
    book.save(data)
    book.close()
    data.seek(0)
    return data, f"内销录单_邮件{case_id}_v{template['current_version']}.xlsx"
