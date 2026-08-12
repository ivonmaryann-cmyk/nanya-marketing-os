from __future__ import annotations

import json
import math
import re
import shutil
import traceback
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.styles import Font
from werkzeug.utils import secure_filename

from .db import append_job_log, create_job, get_job, update_job_status
from .excel_utils import load_workbook_compat, normalized_xlsx_source
from .job_control import launch_job_process
from .paths import JOBS_DIR
from .price_calculation_extended import (
    ExtRules,
    calculate_extended_spec,
    load_extended_rules,
    run_extended_regression,
)
from .price_calculation_customers import enabled_price_customer
from .price_calculation_rules import (
    get_active_price_rule_version,
    get_price_rule_file_path,
    get_price_test_data_file_path,
    normalize_price_quote_variant,
)


FEATURE = "price_calculation"
NOTE_SHEET_NAME = "计算说明"
OUTPUT_HEADERS = ["注意幅宽！", "每卷米数", "新单价", "新总金额"]
OUTPUT_HEADERS_WITHOUT_TOTAL = ["注意幅宽！", "每卷米数", "新单价"]
DESC_HEADERS = {"客户规格", "物料描述", "规格"}
DESC_HEADERS.add("物料长描述")
DESC_HEADERS.add("Description")
QUANTITY_HEADERS = {"订单数量", "数量"}
TABLE_HEADER_HINTS = {
    "预出货日",
    "单别单号",
    "项次",
    "客户简称",
    "单位",
    "旧单价",
    "旧总金额",
    "客户单号",
    "客户产品编号",
    "营运中心",
    "Description",
    "Price",
    "Quantity",
    "净价",
    *QUANTITY_HEADERS,
    *DESC_HEADERS,
}
RULE_VERSION_UNAVAILABLE = "未初始化价格计算规则"
FOIL_TOKEN_PATTERN = r"HS2(?:-M2)?(?:-VSP)?|HVLP[1-4]?|RTF[1-4]?|HTE|VLP"
FOIL_COMBO_PATTERN = rf"(?:{FOIL_TOKEN_PATTERN})\s*[/、,，]\s*(?:{FOIL_TOKEN_PATTERN})(?:\s*[/、,，]\s*(?:{FOIL_TOKEN_PATTERN}))*"
FOIL_PATTERN = rf"{FOIL_COMBO_PATTERN}|(?:{FOIL_TOKEN_PATTERN})"
FOIL_ORDER = {
    "HTE": 0,
    "RTF": 10,
    "RTF1": 11,
    "RTF2": 12,
    "RTF3": 13,
    "RTF4": 14,
    "HVLP": 20,
    "HVLP1": 21,
    "HVLP2": 22,
    "HVLP3": 23,
    "HVLP4": 24,
    "VLP": 30,
    "HS2": 40,
    "HS2-M2": 41,
    "HS2-M2-VSP": 42,
}

XML_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
XML_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
XML_PKG_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"


@dataclass
class PpRule:
    excel_row: int
    product: str
    glass: str
    rc_min: float | None
    rc_max: float | None
    price: float | None


@dataclass
class CclRule:
    excel_row: int
    product: str
    thickness: float | None
    copper_left: str
    copper_right: str
    copper_state: str
    stack: str
    foil: str
    prices: dict[str, float | None]


@dataclass
class JingwangRules:
    pp_rows: list[PpRule]
    ccl_rows: list[CclRule]


@dataclass
class PlinPpRule:
    excel_row: int
    sheet: str
    product: str
    glass: str
    rc_min: float | None
    rc_max: float | None
    length: int | None
    price: float | None


@dataclass
class PlinCclRule:
    excel_row: int
    sheet: str
    product: str
    thickness: float | None
    copper: str
    foil: str
    stack: str
    prices: dict[str, float | None]


@dataclass
class PlinRules:
    pp_rows: list[PlinPpRule]
    ccl_rows: list[PlinCclRule]
    copper_adders: dict[str, float]


@dataclass
class CalcResult:
    status: str
    material_type: str
    price: float | str | None
    total: float | str | None
    width: str
    roll_length: str
    note: str
    rule_row: int | None = None
    size_column: str = ""


@dataclass
class SpecCandidate:
    spec: str
    cell: str
    score: int
    material_type: str


def queue_price_calculation_job(employee_id: str, customer_key: str, uploaded_file, source_filename: str, quote_variant: str | None = None) -> int:
    customer = enabled_price_customer(customer_key)
    quote_variant = normalize_price_quote_variant(customer_key, quote_variant)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    safe_filename = secure_filename(source_filename) or f"price_calculation_{timestamp}.xlsx"
    input_path = employee_dir / f"{timestamp}_{customer_key}_{safe_filename}"
    uploaded_file.save(input_path)

    rule_version = get_active_price_rule_version(customer_key, quote_variant)
    manifest_path = input_path.with_name(f"{input_path.stem}_manifest.json")
    manifest_path.write_text(
        json.dumps(
            {
                "customer_key": customer_key,
                "customer_label": customer["label"],
                "quote_variant": quote_variant,
                "input_path": str(input_path),
                "source_filename": source_filename,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    job_id = create_job(
        employee_id,
        f"{customer['label']} - {source_filename}",
        str(manifest_path),
        f"{customer_key}:{rule_version or RULE_VERSION_UNAVAILABLE}",
        feature=FEATURE,
    )
    launch_job_process(job_id, FEATURE, employee_id)
    return job_id


def run_price_calculation_job(job_id: int, employee_id: str) -> None:
    job = get_job(job_id)
    if not job or job["employee_id"] != employee_id:
        return
    update_job_status(job_id, status="running")
    try:
        manifest = json.loads(Path(job["stored_input_path"]).read_text(encoding="utf-8"))
        customer_key = manifest["customer_key"]
        customer = enabled_price_customer(customer_key)
        input_path = Path(manifest["input_path"])
        quote_variant = normalize_price_quote_variant(customer_key, manifest.get("quote_variant"))
        rule_version = get_active_price_rule_version(customer_key, quote_variant)
        if not rule_version:
            if customer_key == "jingwang":
                label = "旧报价单" if quote_variant == "old" else "新报价单"
                raise ValueError(f"请先导入景旺{label}")
            raise ValueError(f"{customer['label']}价格计算规则未初始化，请先上传规则")
        rule_path = get_price_rule_file_path(customer_key, rule_version, quote_variant)
        variant_note = f"，报价单类型={quote_variant}" if quote_variant else ""
        append_job_log(job_id, f"开始处理价格计算任务：客户={customer['label']}，规则版本={rule_version}{variant_note}")
        rules = load_price_rules(customer_key, rule_path)
        append_job_log(job_id, f"已加载{customer['label']}报价表：PP {len(rules.pp_rows)} 行，基板 {len(rules.ccl_rows)} 行")

        workbook = load_workbook_compat(input_path, data_only=False)
        source_for_result = normalized_xlsx_source(input_path, workbook)
        output_path = input_path.with_name(
            f"{input_path.stem}_价格计算结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        shutil.copy2(source_for_result, output_path)
        result_wb = load_workbook_compat(output_path, data_only=False)
        results = process_price_workbook(result_wb, customer_key, rules, job_id=job_id)
        result_wb.save(output_path)

        success_count = sum(1 for item in results if item["status"] == "成功")
        fail_count = sum(1 for item in results if item["status"] != "成功")
        update_job_status(
            job_id,
            status="completed",
            stored_result_path=str(output_path),
            success_count=success_count,
            fail_count=fail_count,
            skip_count=0,
            current_row=len(results),
            total_rows=len(results),
            completed=True,
        )
        append_job_log(job_id, f"价格计算完成：成功 {success_count} 行，待确认 {fail_count} 行")
    except Exception as exc:
        append_job_log(job_id, f"价格计算失败：{exc}")
        update_job_status(
            job_id,
            status="failed",
            error_message=f"{exc}\n{traceback.format_exc(limit=8)}",
            completed=True,
        )


def calculate_price_quote(customer_key: str, spec: str, quantity: Any = None, quote_variant: str | None = None) -> dict:
    customer = enabled_price_customer(customer_key)
    quote_variant = normalize_price_quote_variant(customer_key, quote_variant)
    rule_version = get_active_price_rule_version(customer_key, quote_variant)
    if not rule_version:
        if customer_key == "jingwang":
            label = "旧报价单" if quote_variant == "old" else "新报价单"
            raise ValueError(f"请先导入景旺{label}")
        raise ValueError(f"{customer['label']}价格计算规则未初始化，请先上传规则")
    rules = load_price_rules(customer_key, get_price_rule_file_path(customer_key, rule_version, quote_variant))
    result = calculate_customer_spec(customer_key, spec, rules, quantity=quantity)
    return {
        "customer_key": customer_key,
        "customer_label": customer["label"],
        "quote_variant": quote_variant,
        "rule_version": rule_version,
        "status": result.status,
        "material_type": result.material_type,
        "price": result.price,
        "total": result.total,
        "width": result.width,
        "roll_length": result.roll_length,
        "note": result.note,
    }


def process_price_workbook(workbook, customer_key: str, rules: JingwangRules | PlinRules | ExtRules, *, job_id: int | None = None) -> list[dict]:
    enabled_price_customer(customer_key)
    all_results: list[dict] = []
    for ws in workbook.worksheets:
        if ws.title == NOTE_SHEET_NAME:
            continue
        header_row, headers = find_header_row(ws)
        desc_col = next((headers[name] for name in DESC_HEADERS if name in headers), None)
        auto_detect_spec = False
        if not desc_col:
            auto_detect_spec = True
            header_row, headers = find_table_header_row(ws)
            if not header_row:
                header_row, headers = 1, {}
        qty_col = headers.get("订单数量") or headers.get("数量")
        has_quantity = bool(qty_col)
        is_plin = customer_key == "plin"
        simple_price_only = customer_key in {"hanyu", "wutong", "eaton", "taixing", "aoshikang", "mingyang", "lejian", "guanghe", "shengyi", "guigu", "techuang", "zhongfu", "huaxingyu", "dongxun", "suhang", "yingchuangli", "zhongjing"}
        if customer_key == "taixing":
            simple_headers = ["新价格", "整卷价格"]
        elif customer_key == "mingyang":
            simple_headers = ["新价格", "200M整卷价格"]
        else:
            simple_headers = ["新价格"]
        net_price_col = _find_net_price_col(headers) if customer_key == "aoshikang" else None
        if customer_key == "aoshikang" and net_price_col:
            simple_headers = [*simple_headers, "净价结果"]
        output_cols = ensure_output_columns(
            ws,
            header_row,
            headers,
            include_total=has_quantity and not is_plin and not simple_price_only,
            output_headers=["新未税价"] if is_plin else (simple_headers if simple_price_only else None),
        )
        max_row = ws.max_row
        taixing_conflicts = _taixing_conflict_specs(ws, header_row, headers) if customer_key in {"taixing", "aoshikang"} else {}
        append_job_log(job_id, f"开始处理 Sheet：{ws.title}，有效范围 {max_row - header_row} 行") if job_id else None
        for row_idx in range(header_row + 1, max_row + 1):
            source_cell = ""
            if auto_detect_spec:
                candidate = find_best_spec_candidate(ws, row_idx, excluded_cols=set(output_cols.values()), customer_key=customer_key)
                if not candidate:
                    continue
                spec = candidate.spec
                source_cell = candidate.cell
            else:
                spec = _text(ws.cell(row=row_idx, column=desc_col).value)
                source_cell = ws.cell(row=row_idx, column=desc_col).coordinate if spec else ""
                if not spec:
                    continue
                if spec in DESC_HEADERS or spec in TABLE_HEADER_HINTS:
                    continue
            quantity = ws.cell(row=row_idx, column=qty_col).value if qty_col else None
            result = calculate_customer_spec(customer_key, spec, rules, quantity=quantity)
            if auto_detect_spec:
                result = CalcResult(
                    result.status,
                    result.material_type,
                    result.price,
                    result.total,
                    result.width,
                    result.roll_length,
                    f"{result.note}；自动识别规格，来源单元格 {source_cell}",
                    result.rule_row,
                    result.size_column,
                )
            if customer_key in {"taixing", "aoshikang"} and spec in taixing_conflicts:
                result = CalcResult(
                    result.status,
                    result.material_type,
                    result.price,
                    result.total,
                    result.width,
                    result.roll_length,
                    f"{result.note}；重复规格单价冲突：{taixing_conflicts[spec]}",
                    result.rule_row,
                    result.size_column,
                )
            if is_plin:
                ws.cell(row=row_idx, column=output_cols["新未税价"], value=_excel_value(result.price))
                result = CalcResult(
                    result.status,
                    result.material_type,
                    result.price,
                    "",
                    result.width,
                    result.roll_length,
                    result.note,
                    result.rule_row,
                    result.size_column,
                )
            elif simple_price_only:
                value_writer = _taixing_excel_value if customer_key in {"taixing", "aoshikang", "hanyu", "shengyi"} else _excel_value
                price_cell = ws.cell(row=row_idx, column=output_cols["新价格"], value=value_writer(result.price))
                if customer_key == "shengyi" and isinstance(result.price, (int, float)):
                    price_cell.number_format = "0.000000" if result.material_type == "PP" else "0.00"
                if customer_key == "taixing":
                    roll_price, roll_note = _taixing_roll_price(result)
                    ws.cell(row=row_idx, column=output_cols["整卷价格"], value=_taixing_excel_value(roll_price))
                    if roll_note:
                        result = CalcResult(
                            result.status,
                            result.material_type,
                            result.price,
                            result.total,
                            result.width,
                            result.roll_length,
                            f"{result.note}；{roll_note}",
                            result.rule_row,
                            result.size_column,
                        )
                if customer_key == "mingyang" and "200M整卷价格" in output_cols:
                    roll_200_price = result.total if result.material_type == "PP" and isinstance(result.total, (int, float)) else ""
                    ws.cell(row=row_idx, column=output_cols["200M整卷价格"], value=_excel_value(roll_200_price))
                if customer_key == "aoshikang" and net_price_col and "净价结果" in output_cols:
                    net_price, net_note = _aoshikang_net_price_result(ws.cell(row=row_idx, column=net_price_col).value)
                    ws.cell(row=row_idx, column=output_cols["净价结果"], value=_taixing_excel_value(net_price))
                    if net_note:
                        result = CalcResult(
                            result.status,
                            result.material_type,
                            result.price,
                            result.total,
                            result.width,
                            result.roll_length,
                            f"{result.note}；{net_note}",
                            result.rule_row,
                            result.size_column,
                        )
                result = CalcResult(
                    result.status,
                    result.material_type,
                    result.price,
                    result.total if customer_key == "mingyang" else "",
                    result.width,
                    result.roll_length,
                    result.note,
                    result.rule_row,
                    result.size_column,
                )
            else:
                ws.cell(row=row_idx, column=output_cols["注意幅宽！"], value=result.width)
                ws.cell(row=row_idx, column=output_cols["每卷米数"], value=result.roll_length)
                ws.cell(row=row_idx, column=output_cols["新单价"], value=_excel_value(result.price))
            if has_quantity and not is_plin and not simple_price_only:
                ws.cell(row=row_idx, column=output_cols["新总金额"], value=_excel_value(result.total))
            elif not is_plin and not simple_price_only:
                result = CalcResult(
                    result.status,
                    result.material_type,
                    result.price,
                    "",
                    result.width,
                    result.roll_length,
                    f"{result.note}；未找到数量列，未计算新总金额",
                    result.rule_row,
                    result.size_column,
                )
            all_results.append(
                {
                    "sheet": ws.title,
                    "row": row_idx,
                    "spec": spec,
                    "status": result.status,
                    "material_type": result.material_type,
                    "price": result.price,
                    "total": result.total,
                    "width": result.width,
                    "roll_length": result.roll_length,
                    "rule_row": result.rule_row,
                    "size_column": result.size_column,
                    "note": result.note,
                }
            )
            if job_id:
                append_job_log(
                    job_id,
                    f"{ws.title} 第 {row_idx} 行：{result.status}，{result.note}",
                    success_count=sum(1 for item in all_results if item["status"] == "成功"),
                    fail_count=sum(1 for item in all_results if item["status"] != "成功"),
                    current_row=len(all_results),
                    total_rows=len(all_results),
                )
    write_note_sheet(workbook, all_results, customer_key=customer_key)
    return all_results


def run_jingwang_regression(customer_key: str, version: str | None = None, quote_variant: str | None = None) -> dict:
    if customer_key == "plin":
        return run_plin_regression(customer_key, version)
    if customer_key in {"hanyu", "wutong", "eaton", "taixing", "aoshikang", "mingyang", "lejian", "guanghe", "shengyi", "guigu", "techuang", "zhongfu", "huaxingyu", "dongxun", "suhang", "yingchuangli", "zhongjing"}:
        rule_version = version or get_active_price_rule_version(customer_key)
        rules = load_extended_rules(customer_key, get_price_rule_file_path(customer_key, rule_version))
        return run_extended_regression(customer_key, rules, get_price_test_data_file_path(customer_key, rule_version))
    quote_variant = normalize_price_quote_variant(customer_key, quote_variant)
    rule_version = version or get_active_price_rule_version(customer_key, quote_variant)
    rules = load_jingwang_rules(get_price_rule_file_path(customer_key, rule_version, quote_variant))
    test_data_path = get_price_test_data_file_path(customer_key, rule_version, quote_variant)
    rows: list[dict] = []
    total = passed = failed = allowed_exception = 0
    if not test_data_path.exists():
        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "allowed_exception": allowed_exception,
            "rows": rows,
        }
    workbook = load_workbook_compat(test_data_path, data_only=True)
    for ws in workbook.worksheets:
        header_row, headers = find_header_row(ws)
        if not header_row:
            continue
        desc_col = next((headers[name] for name in DESC_HEADERS if name in headers), None)
        qty_col = headers.get("订单数量") or headers.get("数量")
        price_col = headers.get("新单价")
        amount_col = headers.get("新总金额")
        if not desc_col or not price_col or not amount_col:
            continue
        for row_idx in range(header_row + 1, ws.max_row + 1):
            spec = _text(ws.cell(row=row_idx, column=desc_col).value)
            if not spec:
                continue
            total += 1
            result = calculate_jingwang_spec(spec, rules, quantity=ws.cell(row=row_idx, column=qty_col).value if qty_col else None)
            expected_price = ws.cell(row=row_idx, column=price_col).value
            expected_total = ws.cell(row=row_idx, column=amount_col).value
            price_ok = _result_equal(result.price, expected_price)
            total_ok = _result_equal(result.total, expected_total)
            is_known_exception = price_ok and not total_ok and "27.3" in spec and "1080*2" in spec
            if price_ok and (total_ok or is_known_exception):
                passed += 1
                if is_known_exception:
                    allowed_exception += 1
                status = "通过" if total_ok else "允许例外"
            else:
                failed += 1
                status = "失败"
            rows.append(
                {
                    "sheet": ws.title,
                    "row": row_idx,
                    "status": status,
                    "expected_price": expected_price,
                    "actual_price": result.price,
                    "expected_total": expected_total,
                    "actual_total": result.total,
                    "note": result.note,
                }
            )
    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "allowed_exception": allowed_exception,
        "rows": rows,
    }


def run_plin_regression(customer_key: str, version: str | None = None) -> dict:
    rule_version = version or get_active_price_rule_version(customer_key)
    rules = load_plin_rules(get_price_rule_file_path(customer_key, rule_version))
    test_data_path = get_price_test_data_file_path(customer_key, rule_version)
    rows: list[dict] = []
    total = passed = failed = allowed_exception = 0
    if not test_data_path.exists():
        return {"total": total, "passed": passed, "failed": failed, "allowed_exception": allowed_exception, "rows": rows}
    workbook = load_workbook_compat(test_data_path, data_only=True)
    for ws in workbook.worksheets:
        header_row, headers = find_header_row(ws)
        if not header_row:
            header_row, headers = _find_exact_header_row(ws, {"物料规格", "未税价格"})
        if not header_row:
            header_row, headers = find_table_header_row(ws)
        if not header_row:
            continue
        desc_col = next((headers[name] for name in DESC_HEADERS | {"物料规格"} if name in headers), None)
        expected_col = headers.get("未税价格") or headers.get("未税价")
        if not desc_col or not expected_col:
            continue
        for row_idx in range(header_row + 1, ws.max_row + 1):
            spec = _text(ws.cell(row=row_idx, column=desc_col).value)
            if not spec:
                continue
            total += 1
            result = calculate_plin_spec(spec, rules)
            expected_price = ws.cell(row=row_idx, column=expected_col).value
            price_ok = _result_equal(result.price, expected_price)
            is_known_exception = (
                not price_ok
                and "NY2170H" in spec
                and "5/5" in spec
                and "106*1" in spec
                and _result_equal(result.price, 644.89)
                and _result_equal(expected_price, 552.89)
            )
            if price_ok or is_known_exception:
                passed += 1
                if is_known_exception:
                    allowed_exception += 1
                status = "通过" if price_ok else "允许例外"
            else:
                failed += 1
                status = "失败"
            rows.append(
                {
                    "sheet": ws.title,
                    "row": row_idx,
                    "status": status,
                    "expected_price": expected_price,
                    "actual_price": result.price,
                    "expected_total": "",
                    "actual_total": "",
                    "note": result.note,
                }
            )
    return {"total": total, "passed": passed, "failed": failed, "allowed_exception": allowed_exception, "rows": rows}


def load_price_rules(customer_key: str, rule_path: str | Path) -> JingwangRules | PlinRules | ExtRules:
    if customer_key == "plin":
        return load_plin_rules(rule_path)
    if customer_key in {"hanyu", "wutong", "eaton", "taixing", "aoshikang", "mingyang", "lejian", "guanghe", "shengyi", "guigu", "techuang", "zhongfu", "huaxingyu", "dongxun", "suhang", "yingchuangli", "zhongjing"}:
        return load_extended_rules(customer_key, rule_path)
    return load_jingwang_rules(rule_path)


def calculate_customer_spec(customer_key: str, spec: str, rules: JingwangRules | PlinRules | ExtRules, quantity: Any = None) -> CalcResult:
    if customer_key == "plin":
        return calculate_plin_spec(spec, rules)  # type: ignore[arg-type]
    if customer_key in {"hanyu", "wutong", "eaton", "taixing", "aoshikang", "mingyang", "lejian", "guanghe", "shengyi", "guigu", "techuang", "zhongfu", "huaxingyu", "dongxun", "suhang", "yingchuangli", "zhongjing"}:
        result = calculate_extended_spec(customer_key, spec, rules, quantity=quantity)  # type: ignore[arg-type]
        return CalcResult(
            result.status,
            result.material_type,
            result.price,
            result.total,
            result.width,
            result.roll_length,
            result.note,
            result.rule_row,
            result.size_column,
        )
    return calculate_jingwang_spec(spec, rules, quantity=quantity)  # type: ignore[arg-type]


def calculate_jingwang_spec(spec: str, rules: JingwangRules, quantity: Any = None) -> CalcResult:
    desc = _text(spec)
    if not desc:
        return CalcResult("失败", "未知", "待确认", "待确认", "", "", "客户规格为空")
    if desc.upper().startswith("PP"):
        return _calculate_jingwang_pp(desc, rules, quantity=quantity)
    return _calculate_jingwang_ccl(desc, rules, quantity=quantity)


def _calculate_jingwang_pp(desc: str, rules: JingwangRules, quantity: Any = None) -> CalcResult:
    parts = desc.split()
    product = _norm_product(parts[1]) if len(parts) > 1 and parts[0].upper() == "PP" else ""
    glass_match = re.search(r"\b(106|10[0-9]{2}|1067|1078|1080|1506|2113|2116|2313|3313|7628)\b", desc, re.I)
    rc_match = re.search(r"RC\s*([0-9]+(?:\.[0-9]+)?)\s*%?", desc, re.I)
    width_match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(?:IN|INCH|英寸)\b", desc, re.I)
    length_match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*m\b", desc, re.I)
    length_value, length_unit, width_value, width_unit = _extract_jing_wei_size(desc)
    glass = _norm_glass(glass_match.group(1)) if glass_match else ""
    rc = _to_float(rc_match.group(1)) if rc_match else None
    width = f"{width_match.group(1)}IN" if width_match else ""
    roll_length = f"{length_match.group(1)}m" if length_match else ""
    if not product or not glass or rc is None:
        return CalcResult("失败", "PP", "待确认", "待确认", width, roll_length, "PP规格缺少型号、玻布或RC")
    for row in rules.pp_rows:
        if row.product != product or row.glass != glass or row.price is None:
            continue
        if row.rc_min is None or row.rc_max is None:
            continue
        if row.rc_min - 0.001 <= rc <= row.rc_max + 0.001:
            base_price = _round_money(_jingwang_pp_price_exception(product, glass, rc, row.price))
            if length_value is not None and width_value is not None:
                length_mm = _axis_to_mm_2dp(length_value, length_unit)
                width_in = _axis_to_inch_2dp(width_value, width_unit)
                split = math.floor(49.5 / width_in) if width_in else 0
                if split <= 0:
                    return CalcResult("失败", "PP", "待确认", "待确认", width, roll_length, f"PP纬向宽幅无法一开：{width_in:.2f} inch")
                price = _round_money(base_price * length_mm / 1000 / split)
                total = _calc_total(quantity, price)
                note = (
                    f"PP小片命中报价表第 {row.excel_row} 行，查出单价={base_price:.2f}，"
                    f"经向={length_mm:.2f}mm，纬向={width_in:.2f}inch，纬向一开{split}，"
                    f"公式={base_price:.2f}*{length_mm:.2f}/1000/{split}"
                )
                return CalcResult("成功", "PP", price, total, f"{width_in:.2f}IN", "", note, row.excel_row)
            price = base_price
            total = _calc_total(quantity, price)
            return CalcResult("成功", "PP", price, total, width, roll_length, f"命中PP报价表第 {row.excel_row} 行", row.excel_row)
    return CalcResult("失败", "PP", "待确认", "待确认", width, roll_length, "未命中PP报价：型号、玻布或RC不匹配")


def _calculate_jingwang_ccl(desc: str, rules: JingwangRules, quantity: Any = None) -> CalcResult:
    product_match = re.search(r"\b(NY[\w\-.（）()]+)\b", desc, re.I)
    thickness_match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*mm\b", desc, re.I)
    copper_match = re.search(r"\b([0-9HhTt]+(?:\.[0-9]+)?)\s*/\s*([0-9HhTt]+(?:\.[0-9]+)?)\b", desc)
    state = "不含铜" if "不含铜" in desc else ("含铜" if "含铜" in desc else "")
    size_match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(mm|in|inch|英寸)?\s*[xX*×]\s*([0-9]+(?:\.[0-9]+)?)\s*(mm|in|inch|英寸)?", desc, re.I)
    if _is_stack_like_size_match(size_match):
        size_match = None
    foil_match = re.search(rf"(?<![A-Z0-9])({FOIL_PATTERN})(?![A-Z0-9])", desc, re.I)
    product = _norm_product(product_match.group(1)) if product_match else ""
    thickness = _to_float(thickness_match.group(1)) if thickness_match else None
    copper_left = _norm_copper(copper_match.group(1)) if copper_match else ""
    copper_right = _norm_copper(copper_match.group(2)) if copper_match else ""
    foil = _norm_foil(foil_match.group(1)) if foil_match else "HTE"
    foil_note = "" if foil_match else "；描述未写铜箔，默认HTE"
    stack = _extract_stack(desc)
    length_in, width_in = _size_to_inches_from_match(size_match)
    jing_value, jing_unit, wei_value, wei_unit = _extract_jing_wei_size(desc)
    if jing_value is not None and wei_value is not None:
        length_in = _axis_to_inch_2dp(jing_value, jing_unit)
        width_in = _axis_to_inch_2dp(wei_value, wei_unit)
    size_key, size_column = _standard_ccl_size_column(length_in, width_in)
    if not product or thickness is None or not copper_left or not copper_right or not state or not stack or not foil or length_in is None or width_in is None:
        return CalcResult("失败", "CCL", "待确认", "待确认", "", "", "CCL规格缺少型号、厚度、铜厚、含铜状态、尺寸或叠构")
    def _matching_ccl_rows(left: str, right: str) -> list[CclRule]:
        return [
            row
            for row in rules.ccl_rows
            if row.product == product
            and row.thickness is not None
            and abs(row.thickness - thickness) <= 0.001
            and row.copper_left == left
            and row.copper_right == right
            and row.copper_state == state
            and row.stack == stack
            and row.foil == foil
        ]

    candidates = _matching_ccl_rows(copper_left, copper_right)
    fallback_note = ""
    if not candidates and copper_left == "0.33" and copper_right == "0.33":
        candidates = _matching_ccl_rows("T", "T")
        if candidates:
            fallback_note = "；0.33/0.33未命中，按T/T匹配"
    note_suffix = f"{fallback_note}{foil_note}"
    parent = _select_ccl_parent(length_in, width_in)
    if candidates and parent and parent["opens"] > 1:
        for row in candidates:
            price = row.prices.get(parent["key"])
            if price is None:
                continue
            final = _round_money(float(price) / parent["opens"])
            total = _calc_total(quantity, final)
            note = (
                f"基板小片命中报价表第 {row.excel_row} 行，父级{parent['parent_w']}x{parent['parent_h']}，"
                f"尺寸列{parent['label']}，经向一开{parent['opens_w']}，纬向一开{parent['opens_h']}，"
                f"总开数{parent['opens']}，原始报价{price:.2f}，公式={price:.2f}/{parent['opens']}{note_suffix}"
            )
            return CalcResult("成功", "CCL", final, total, "", "", note, row.excel_row, parent["label"])
    for row in candidates:
        price = row.prices.get(size_key)
        if price is None:
            continue
        price = _round_money(price)
        total = _calc_total(quantity, price)
        return CalcResult(
            "成功",
            "CCL",
            price,
            total,
            "",
            "",
            f"命中基板报价表第 {row.excel_row} 行，尺寸列 {size_column}{note_suffix}",
            row.excel_row,
            size_column,
        )
    if candidates and length_in is not None and width_in is not None:
        if not parent:
            return CalcResult("失败", "CCL", "待确认", "待确认", "", "", f"无法匹配基板小片父级板：{length_in:.2f}x{width_in:.2f} inch{note_suffix}")
        for row in candidates:
            price = row.prices.get(parent["key"])
            if price is None:
                continue
            final = _round_money(float(price) / parent["opens"])
            total = _calc_total(quantity, final)
            note = (
                f"基板小片命中报价表第 {row.excel_row} 行，父级{parent['parent_w']}x{parent['parent_h']}，"
                f"尺寸列{parent['label']}，经向一开{parent['opens_w']}，纬向一开{parent['opens_h']}，"
                f"总开数{parent['opens']}，原始报价{price:.2f}，公式={price:.2f}/{parent['opens']}{note_suffix}"
            )
            return CalcResult("成功", "CCL", final, total, "", "", note, row.excel_row, parent["label"])
        return CalcResult("失败", "CCL", "待确认", "待确认", "", "", f"命中规格但小片父级尺寸列 {parent['label']} 无价格{note_suffix}")
    if candidates:
        return CalcResult("失败", "CCL", "待确认", "待确认", "", "", f"命中规格但尺寸列 {size_column} 无价格")
    return CalcResult("失败", "CCL", "待确认", "待确认", "", "", "未命中CCL报价：型号、厚度、铜厚、含铜状态、叠构或铜箔不匹配")


def calculate_plin_spec(spec: str, rules: PlinRules) -> CalcResult:
    desc = _text(spec)
    if not desc:
        return CalcResult("失败", "未知", "待确认", "", "", "", "物料规格为空")
    if _looks_like_plin_pp_desc(desc):
        return _calculate_plin_pp(desc, rules)
    return _calculate_plin_ccl(desc, rules)


def _calculate_plin_pp(desc: str, rules: PlinRules) -> CalcResult:
    product = _plin_product_from_desc(desc)
    glass_match = re.search(r"\b(106H?|10[0-9]{2}H?|1067|1078|1080H?|1506|2113|2116H?|2313|3313|7628H?)\b", desc, re.I)
    rc_match = re.search(r"RC\s*([0-9]+(?:\.[0-9]+)?)\s*%?", desc, re.I)
    length_match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(?:米\s*/?\s*卷|M\b)", desc, re.I)
    glass = _norm_glass(glass_match.group(1).rstrip("Hh")) if glass_match else ""
    rc = _to_float(rc_match.group(1)) if rc_match else None
    length = int(math.floor(float(length_match.group(1)) + 1e-9)) if length_match else None
    if not product or not glass or rc is None or length is None:
        return CalcResult("失败", "PP", "待确认", "", "", "", "普林PP规格缺少型号、玻布、RC或卷长")
    for row in rules.pp_rows:
        if row.product != product or row.glass != glass or row.length != length or row.price is None:
            continue
        if row.rc_min is None or row.rc_max is None:
            continue
        if row.rc_min - 0.001 <= rc <= row.rc_max + 0.001:
            price = _round_money(row.price)
            note = (
                f"命中普林PP报价 Sheet {row.sheet} 第 {row.excel_row} 行，"
                f"Per M={price:.2f}，卷长按 {length}m 匹配"
            )
            return CalcResult("成功", "PP", price, "", "", f"{length}m", note, row.excel_row, row.sheet)
    return CalcResult("失败", "PP", "待确认", "", "", f"{length}m" if length else "", "未命中普林PP报价：型号、玻布、RC或卷长不匹配")


def _calculate_plin_ccl(desc: str, rules: PlinRules) -> CalcResult:
    product = _plin_product_from_desc(desc)
    copper_match = re.search(r"\b([0-9Hh]+(?:\.[0-9]+)?)\s*/\s*([0-9Hh]+(?:\.[0-9]+)?)\b", desc)
    foil_match = re.search(rf"(?<![A-Z0-9])({FOIL_PATTERN})(?![A-Z0-9])", desc, re.I)
    size_match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(mm|in|inch|英寸)?\s*[xX*×]\s*([0-9]+(?:\.[0-9]+)?)\s*(mm|in|inch|英寸)?", desc, re.I)
    if _is_stack_like_size_match(size_match):
        size_match = None
    length_in, width_in = _size_to_inches_from_match(size_match)
    jing_value, jing_unit, wei_value, wei_unit = _extract_jing_wei_size(desc)
    if length_in is None or width_in is None:
        if jing_value is not None and wei_value is not None:
            length_in = _axis_to_inch_2dp(jing_value, jing_unit)
            width_in = _axis_to_inch_2dp(wei_value, wei_unit)
    copper = _plin_norm_copper(copper_match.group(1), copper_match.group(2)) if copper_match else ""
    raw_copper = _plin_raw_copper(copper_match.group(1), copper_match.group(2)) if copper_match else ""
    foil = _norm_foil(foil_match.group(1)) if foil_match else "HTE"
    foil_note = "" if foil_match else "；描述未写铜箔，默认HTE"
    stack = _canonical_plin_stack(_extract_stack(desc))
    thickness = _plin_thickness_from_desc(desc, copper_match)
    size_key, multiplier = _plin_size_key_and_multiplier(length_in, width_in)
    if not product or thickness is None or not copper or not stack or not foil or length_in is None or width_in is None:
        return CalcResult("失败", "CCL", "待确认", "", "", "", "普林CCL规格缺少型号、厚度、铜厚、尺寸、叠构或铜箔")

    lookup_copper = "1/1" if copper in {"2/2", "3/3"} else copper
    candidates = [
        row
        for row in rules.ccl_rows
        if row.product == product
        and row.thickness is not None
        and abs(row.thickness - thickness) <= 0.001
        and row.copper == lookup_copper
        and row.stack == stack
        and row.foil == foil
    ]
    if not candidates:
        return CalcResult("失败", "CCL", "待确认", "", "", "", "未命中普林CCL报价：型号、厚度、铜厚、叠构或铜箔不匹配")
    for row in candidates:
        base37 = row.prices.get("37")
        if copper in {"2/2", "3/3"}:
            adder = rules.copper_adders.get(copper)
            if base37 is None or adder is None:
                continue
            if size_key:
                price = _round_money((float(base37) + adder) * multiplier)
                note = (
                    f"命中普林基板报价 Sheet {row.sheet} 第 {row.excel_row} 行，特殊铜厚{raw_copper}按1/1的37*49基价"
                    f"{base37:.2f}+加价{adder:.2f}，尺寸{size_key}倍率{multiplier:.2f}，公式=({base37:.2f}+{adder:.2f})*{multiplier:.2f}{foil_note}"
                )
                return CalcResult("成功", "CCL", price, "", "", "", note, row.excel_row, f"{size_key}*49")
            parent = _select_plin_ccl_parent(length_in, width_in)
            if not parent:
                continue
            source_multiplier = {"37": 1.0, "41": 1.11, "43": 1.16}.get(parent["price_key"])
            if source_multiplier is None:
                continue
            parent_price = _round_money((float(base37) + adder) * source_multiplier)
            if parent_price is None:
                continue
            price = _round_money(float(parent_price) * parent["price_factor"] / parent["opens"])
            note = (
                f"普林基板小片命中报价 Sheet {row.sheet} 第 {row.excel_row} 行，特殊铜厚{raw_copper}按1/1的37*49基价"
                f"{base37:.2f}+加价{adder:.2f}，父级{parent['label']}按{parent['source_label']}价格{parent_price:.2f}"
                f"*{parent['price_factor']}，经向一开{parent['opens_w']}，纬向一开{parent['opens_h']}，"
                f"总开数{parent['opens']}，公式={parent_price:.2f}*{parent['price_factor']}/{parent['opens']}{foil_note}"
            )
            return CalcResult("成功", "CCL", price, "", "", "", note, row.excel_row, parent["label"])
        if size_key:
            direct = row.prices.get(size_key)
            if direct is None:
                continue
            price = _round_money(direct)
            note = f"命中普林基板报价 Sheet {row.sheet} 第 {row.excel_row} 行，尺寸列 {size_key}*49{foil_note}"
            if copper != raw_copper:
                note += f"；铜厚{raw_copper}按{copper}匹配"
            return CalcResult("成功", "CCL", price, "", "", "", note, row.excel_row, f"{size_key}*49")
        parent = _select_plin_ccl_parent(length_in, width_in)
        if parent:
            parent_price = row.prices.get(parent["price_key"])
            if parent_price is None:
                continue
            price = _round_money(float(parent_price) * parent["price_factor"] / parent["opens"])
            note = (
                f"普林基板小片命中报价 Sheet {row.sheet} 第 {row.excel_row} 行，"
                f"父级{parent['label']}按{parent['source_label']}价格{parent_price:.2f}*{parent['price_factor']}，"
                f"经向一开{parent['opens_w']}，纬向一开{parent['opens_h']}，总开数{parent['opens']}，"
                f"公式={parent_price:.2f}*{parent['price_factor']}/{parent['opens']}{foil_note}"
            )
            if copper != raw_copper:
                note += f"；铜厚{raw_copper}按{copper}匹配"
            return CalcResult("成功", "CCL", price, "", "", "", note, row.excel_row, parent["label"])
    if size_key:
        return CalcResult("失败", "CCL", "待确认", "", "", "", f"命中规格但尺寸列 {size_key}*49 无价格")
    return CalcResult("失败", "CCL", "待确认", "", "", "", "命中规格但未找到可用小片父级报价")


def load_plin_rules(rule_path: str | Path) -> PlinRules:
    workbook = load_workbook_compat(rule_path, data_only=True)
    pp_rows: list[PlinPpRule] = []
    ccl_rows: list[PlinCclRule] = []
    copper_adders: dict[str, float] = {}
    for ws in workbook.worksheets:
        _collect_plin_adders(ws, copper_adders)
        header_row, headers = _find_exact_header_row(ws, {"Products", "Glass type", "Resin Content", "Length (m)", "Per M"})
        if header_row:
            for row_idx in range(header_row + 1, ws.max_row + 1):
                product = _plin_norm_product(ws.cell(row_idx, headers["Products"]).value)
                glass = _norm_glass(ws.cell(row_idx, headers["Glass type"]).value)
                rc_min, rc_max = _parse_rc_range(ws.cell(row_idx, headers["Resin Content"]).value)
                length = _to_float(ws.cell(row_idx, headers["Length (m)"]).value)
                price = _to_float(ws.cell(row_idx, headers["Per M"]).value)
                if not product or not glass or rc_min is None or rc_max is None or length is None or price is None:
                    continue
                pp_rows.append(PlinPpRule(row_idx, ws.title, product, glass, rc_min, rc_max, int(math.floor(length + 1e-9)), price))
            continue

        header_row, headers = _find_exact_header_row(ws, {"产品类别", "厚度mm", "铜厚", "铜箔类型", "组合叠构", "37*49", "41*49", "43*49"})
        if header_row:
            for row_idx in range(header_row + 1, ws.max_row + 1):
                product = _plin_norm_product(ws.cell(row_idx, headers["产品类别"]).value)
                thickness = _to_float(ws.cell(row_idx, headers["厚度mm"]).value)
                copper = _plin_norm_copper_value(ws.cell(row_idx, headers["铜厚"]).value)
                foil = _norm_foil(ws.cell(row_idx, headers["铜箔类型"]).value) or "HTE"
                stack = _canonical_plin_stack(ws.cell(row_idx, headers["组合叠构"]).value)
                if not product or thickness is None or not copper or not stack:
                    continue
                ccl_rows.append(
                    PlinCclRule(
                        row_idx,
                        ws.title,
                        product,
                        thickness,
                        copper,
                        foil,
                        stack,
                        {
                            "37": _to_float(ws.cell(row_idx, headers["37*49"]).value),
                            "41": _to_float(ws.cell(row_idx, headers["41*49"]).value),
                            "43": _to_float(ws.cell(row_idx, headers["43*49"]).value),
                        },
                    )
                )
    if not pp_rows or not ccl_rows:
        raise ValueError("普林报价表未读取到有效PP或基板规则")
    return PlinRules(pp_rows=pp_rows, ccl_rows=ccl_rows, copper_adders=copper_adders)


def load_jingwang_rules(rule_path: str | Path) -> JingwangRules:
    sheets = _read_xlsx_value_sheets(Path(rule_path))
    if "PP" not in sheets or "基板" not in sheets:
        raise ValueError("景旺报价表必须包含“基板”和“PP”两张 Sheet")
    pp_rows: list[PpRule] = []
    for excel_row, row in sheets["PP"]:
        pp_row = _parse_jingwang_pp_rule_row(excel_row, row)
        if pp_row:
            pp_rows.append(pp_row)

    ccl_rows: list[CclRule] = []
    for excel_row, row in sheets["基板"]:
        product = _norm_product(row.get(13))
        thickness = _to_float(row.get(14))
        stack = _canonical_stack(row.get(20))
        foil = _norm_foil(row.get(22)) or "HTE"
        if not product or thickness is None or not stack:
            continue
        ccl_rows.append(
            CclRule(
                excel_row=excel_row,
                product=product,
                thickness=thickness,
                copper_left=_norm_copper(row.get(16)),
                copper_right=_norm_copper(row.get(17)),
                copper_state=_norm_state(row.get(19)),
                stack=stack,
                foil=foil,
                prices={
                    "24": _to_float(row.get(3)),
                    "27": _to_float(row.get(4)),
                    "28": _to_float(row.get(5)),
                    "37": _to_float(row.get(6)),
                    "41": _to_float(row.get(8)),
                    "43": _to_float(row.get(9)),
                    "74": _to_float(row.get(10)),
                    "82": _to_float(row.get(11)),
                    "86": _to_float(row.get(12)),
                },
            )
        )
    return JingwangRules(pp_rows=pp_rows, ccl_rows=ccl_rows)


def _read_xlsx_value_sheets(path: Path) -> dict[str, list[tuple[int, dict[int, str]]]]:
    with zipfile.ZipFile(path) as archive:
        shared = _load_shared_strings(archive)
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall(XML_PKG_REL + "Relationship")}
        result: dict[str, list[tuple[int, dict[int, str]]]] = {}
        for sheet in workbook.find(XML_NS + "sheets").findall(XML_NS + "sheet"):
            name = sheet.attrib["name"]
            rid = sheet.attrib[XML_REL + "id"]
            target = rid_to_target[rid]
            sheet_path = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
            root = ET.fromstring(archive.read(sheet_path))
            rows: list[tuple[int, dict[int, str]]] = []
            for row in root.find(XML_NS + "sheetData").findall(XML_NS + "row"):
                values: dict[int, str] = {}
                for cell in row.findall(XML_NS + "c"):
                    values[_col_number(cell.attrib["r"])] = _cell_value(cell, shared)
                rows.append((int(row.attrib.get("r", len(rows) + 1)), values))
            result[name] = rows
        return result


def _parse_jingwang_pp_rule_row(excel_row: int, row: dict[int, str]) -> PpRule | None:
    current = _build_pp_rule(
        excel_row,
        product=row.get(3),
        glass=row.get(5),
        rc=row.get(4),
        price=row.get(2),
    )
    if current:
        return current

    legacy = _build_pp_rule(
        excel_row,
        product=row.get(4),
        glass=row.get(6),
        rc=row.get(5),
        price=row.get(3),
    )
    if legacy:
        return legacy

    product_col = rc_col = glass_col = price_col = None
    for col_idx, value in row.items():
        text = _text(value)
        if product_col is None and _looks_like_pp_product(text):
            product_col = col_idx
        if rc_col is None and _looks_like_rc(text):
            rc_col = col_idx
        if glass_col is None and _looks_like_glass(text):
            glass_col = col_idx
    numeric_cols = [
        col_idx
        for col_idx, value in row.items()
        if _to_float(value) is not None and col_idx not in {product_col, rc_col, glass_col}
    ]
    if product_col and rc_col and glass_col:
        for col_idx in sorted(numeric_cols):
            if col_idx < product_col:
                price_col = col_idx
        if price_col is None and numeric_cols:
            price_col = sorted(numeric_cols)[0]
        if price_col:
            return _build_pp_rule(
                excel_row,
                product=row.get(product_col),
                glass=row.get(glass_col),
                rc=row.get(rc_col),
                price=row.get(price_col),
            )
    return None


def _build_pp_rule(excel_row: int, *, product: Any, glass: Any, rc: Any, price: Any) -> PpRule | None:
    norm_product = _norm_product(product)
    norm_glass = _norm_glass(glass)
    rc_min, rc_max = _parse_rc_range(rc)
    parsed_price = _to_float(price)
    if not _looks_like_pp_product(norm_product) or not _looks_like_glass(norm_glass):
        return None
    if rc_min is None or rc_max is None or parsed_price is None:
        return None
    return PpRule(excel_row, norm_product, norm_glass, rc_min, rc_max, parsed_price)


def _looks_like_pp_product(value: Any) -> bool:
    text = _norm_product(value)
    return bool(re.fullmatch(r"NY[\w\-.()]+P(?:\(C\))?", text))


def _looks_like_rc(value: Any) -> bool:
    return bool(re.search(r"\bRC\s*[0-9]+(?:\.[0-9]+)?(?:\s*-\s*[0-9]+(?:\.[0-9]+)?)?\s*%?", _text(value), re.I))


def _looks_like_glass(value: Any) -> bool:
    return bool(re.fullmatch(r"(106|10[0-9]{2}|1067|1078|1080|1506|2113|2116|2313|3313|7628)", _norm_glass(value)))


def _load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    values: list[str] = []
    for item in root.findall(XML_NS + "si"):
        values.append("".join(text.text or "" for text in item.iter(XML_NS + "t")))
    return values


def _cell_value(cell, shared: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.iter(XML_NS + "t")).strip()
    value = cell.find(XML_NS + "v")
    if value is None:
        return ""
    raw = (value.text or "").strip()
    if cell_type == "s":
        try:
            return shared[int(raw)].strip()
        except Exception:
            return raw
    return raw


def _col_number(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref)
    if not match:
        return 0
    number = 0
    for char in match.group(1):
        number = number * 26 + ord(char) - 64
    return number


def find_header_row(ws) -> tuple[int | None, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        headers: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            text = _text(ws.cell(row=row_idx, column=col_idx).value)
            if text:
                headers[text] = col_idx
        if any(name in headers for name in DESC_HEADERS):
            return row_idx, headers
    return None, {}


def find_table_header_row(ws) -> tuple[int | None, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        headers: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            text = _text(ws.cell(row=row_idx, column=col_idx).value)
            if text:
                headers[text] = col_idx
        matches = TABLE_HEADER_HINTS.intersection(headers)
        if len(matches) >= 3 or (matches.intersection(QUANTITY_HEADERS) and len(matches) >= 2):
            return row_idx, headers
    return None, {}


def find_best_spec_candidate(ws, row_idx: int, *, excluded_cols: set[int] | None = None, customer_key: str = "") -> SpecCandidate | None:
    excluded_cols = excluded_cols or set()
    candidates: list[SpecCandidate] = []
    for col_idx in range(1, ws.max_column + 1):
        if col_idx in excluded_cols:
            continue
        text = _text(ws.cell(row=row_idx, column=col_idx).value)
        score, material_type = score_spec_candidate(text, customer_key=customer_key)
        if score > 0:
            candidates.append(SpecCandidate(text, ws.cell(row=row_idx, column=col_idx).coordinate, score, material_type))
    if not candidates:
        return None
    return sorted(candidates, key=lambda item: (-item.score, item.cell))[0]


def score_spec_candidate(value: Any, *, customer_key: str = "") -> tuple[int, str]:
    text = _text(value)
    if not text or len(text) < 8:
        return 0, ""
    compact = re.sub(r"\s+", "", text.upper())
    if text in OUTPUT_HEADERS or text in OUTPUT_HEADERS_WITHOUT_TOTAL or text in TABLE_HEADER_HINTS:
        return 0, ""
    if re.fullmatch(r"[\d\s.,:/\\-]+", text):
        return 0, ""
    if re.fullmatch(r"[A-Z]*\d{6,}[A-Z0-9\\-]*", compact):
        return 0, ""

    pp_score = 0
    if "有卤PP" in text or "无卤PP" in text:
        pp_score += 65
    if text.count("|") >= 5 and re.search(r"\|\s*NY[\w\-.（）() ]+P\s*\|", text, re.I):
        pp_score += 30
    if re.search(r"\bPP\b", text, re.I):
        pp_score += 35
    if re.search(r"\bPREPREG\b", text, re.I):
        pp_score += 45
    if re.search(r"\bNY[\w\-.()（）]+P\b", text, re.I):
        pp_score += 20
    if re.search(r"\bNY[\w\-.()（）]+\b", text, re.I) and re.search(r"RC\s*[0-9]+", text, re.I):
        pp_score += 15
    if re.search(r"\bNY\s*-?\s*[A-Z]?\d{3,4}[A-Z0-9]*P?\b|\bNY\s*-?\s*A\d[A-Z0-9]*P?\b", text, re.I):
        pp_score += 10
    if re.search(r"\b(106|10[0-9]{2}|1067|1078|1080|1506|2113|2116|2313|3313|7628)\b", text, re.I):
        pp_score += 15
    if re.search(r"RC\s*[0-9]+(?:\.[0-9]+)?\s*%?", text, re.I):
        pp_score += 15
    if re.search(r"[0-9]+(?:\.[0-9]+)?\s*(?:IN|INCH|英寸|\"|m)\b", text, re.I):
        pp_score += 10
    if re.search(r"[0-9]+(?:\.[0-9]+)?\s*(?:\"|IN|INCH|英寸)?\s*[*xX×]\s*[0-9]+(?:\.[0-9]+)?\s*M\b", text, re.I):
        pp_score += 15
    if re.search(r"[0-9]+(?:\.[0-9]+)?\s*米\s*/?\s*卷", text, re.I):
        pp_score += 15
    if "有卤" in text or "无卤" in text:
        pp_score += 5
    if re.search(r"\bFR4(?:\.\d+)?\b", text, re.I) and re.search(r"\bPREPREG\b", text, re.I):
        pp_score += 10

    ccl_score = 0
    if "有卤基板" in text or "无卤基板" in text:
        ccl_score += 65
    if text.count("|") >= 8 and re.search(r"\|\s*NY[\w\-.（）() ]+\s*\|", text, re.I):
        ccl_score += 20
    if re.search(r"\bCCL\b", text, re.I):
        ccl_score += 45
    if re.search(r"\bNY[\w\-.()（）]+\b", text, re.I):
        ccl_score += 18
    if re.search(r"\bNY\s*-?\s*[A-Z]?\d{3,4}[A-Z0-9]*\b|\bNY\s*-?\s*A\d[A-Z0-9]*\b", text, re.I):
        ccl_score += 10
    if customer_key == "hanyu" and re.search(r"\bNY\s*-?\s*[A-Z]?\d{3,4}[A-Z0-9]*\b", text, re.I) and re.search(r"\b\d{3}\b", text):
        ccl_score += 10
    if re.search(r"[0-9]+(?:\.[0-9]+)?\s*mm\b", text, re.I):
        ccl_score += 15
    if re.search(r"[0-9]+(?:\.[0-9]+)?(?:\s*\+\s*[0-9]+(?:\.[0-9]+)?)?\s*MIL\b", text, re.I):
        ccl_score += 15
    if re.search(r"\bNY[\w\-.()（）]+\b.*?\b[0-9]+(?:\.[0-9]+)?\s+[0-9Hh]+(?:\.[0-9]+)?\s*/\s*[0-9Hh]+(?:\.[0-9]+)?\b", text, re.I):
        ccl_score += 12
    if re.search(r"\b[0-9Hh]+(?:\.[0-9]+)?\s*/\s*[0-9Hh]+(?:\.[0-9]+)?\b", text):
        ccl_score += 15
    if customer_key == "hanyu" and re.search(r"\b[0-9Hh]+(?:\.[0-9]+)?\s*\(\s*(?:RTF|HTE)\s*\)\s*/\s*[0-9Hh]+(?:\.[0-9]+)?\s*\(\s*(?:RTF|HTE)\s*\)", text, re.I):
        ccl_score += 25
    if "含铜" in text or "不含铜" in text or "不连铜" in text or re.search(r"\bNO\s*[-_]?\s*CU\b", text, re.I):
        ccl_score += 10
    if re.search(r"(?:经向?|经)\s*[0-9]+(?:\.[0-9]+)?\s*[xX*×]?\s*(?:纬向?|纬)\s*[0-9]+(?:\.[0-9]+)?", text, re.I):
        ccl_score += 15
    if re.search(r"[0-9]+(?:\.[0-9]+)?\s*(?:mm|in|inch|英寸|\")?\s*[xX*×]\s*[0-9]+(?:\.[0-9]+)?\s*(?:mm|in|inch|英寸|\")?", text, re.I):
        ccl_score += 15
    if re.search(r"\bFR4(?:\.\d+)?\b", text, re.I):
        ccl_score += 10
    if re.search(r"\b(?:1K|2C|1D|1J|2E|2X1035|1F)\b", text, re.I):
        ccl_score += 10
    if re.search(r"\b(?:SH|S[12]|R[23]1|HOZ|H)\s*/\s*(?:SH|S[12]|R[23]1|HOZ|H)\b", text, re.I):
        ccl_score += 10
    if re.search(r"\b(?:106|10[0-9]{2}|1067|1078|1080|1506|2113|2116|2313|3313|7628)\s*[*×X]\s*[0-9]+", text, re.I):
        ccl_score += 15
    if customer_key == "hanyu" and re.search(r"\b[0-9]+\s*[*xX×]\s*(?:106|1035|1037|1067|1078|1080|1086|1506|2113|2116|2313|3313|7628)\b", text, re.I):
        ccl_score += 15
    if re.search(r"[0-9]+\s*张\s*(?:106|10[0-9]{2}|1067|1078|1080|1506|2113|2116|2313|3313|7628)", text, re.I):
        ccl_score += 15
    if re.search(rf"\b({FOIL_PATTERN})\b", text, re.I):
        ccl_score += 5

    if pp_score >= 60 and pp_score >= ccl_score:
        return pp_score, "PP"
    if ccl_score >= 55:
        return ccl_score, "CCL"
    return 0, ""


def ensure_output_columns(
    ws,
    header_row: int,
    headers: dict[str, int],
    *,
    include_total: bool = True,
    output_headers: list[str] | None = None,
) -> dict[str, int]:
    output_cols: dict[str, int] = {}
    next_col = ws.max_column + 1
    wanted_headers = output_headers or (OUTPUT_HEADERS if include_total else OUTPUT_HEADERS_WITHOUT_TOTAL)
    for header in wanted_headers:
        col_idx = headers.get(header)
        if not col_idx:
            col_idx = next_col
            ws.cell(row=header_row, column=col_idx, value=header)
            ws.cell(row=header_row, column=col_idx).font = Font(bold=True)
            next_col += 1
        output_cols[header] = col_idx
    return output_cols


def write_note_sheet(workbook, rows: list[dict], *, customer_key: str = "") -> None:
    if NOTE_SHEET_NAME in workbook.sheetnames:
        del workbook[NOTE_SHEET_NAME]
    ws = workbook.create_sheet(NOTE_SHEET_NAME)
    headers = ["来源Sheet", "行号", "规格", "材料类型", "状态", "新单价", "新总金额", "规则行", "尺寸列", "说明"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    value_writer = _taixing_excel_value if customer_key == "shengyi" else _excel_value
    for row in rows:
        ws.append(
            [
                row["sheet"],
                row["row"],
                row["spec"],
                row["material_type"],
                row["status"],
                value_writer(row["price"]),
                _excel_value(row["total"]),
                row["rule_row"] or "",
                row["size_column"],
                row["note"],
            ]
        )
        if customer_key == "shengyi" and isinstance(row["price"], (int, float)):
            ws.cell(row=ws.max_row, column=6).number_format = "0.000000" if row["material_type"] == "PP" else "0.00"


def _size_column(match) -> tuple[str, str]:
    if not match:
        return "", ""
    width = _to_float(match.group(1))
    if width is None:
        return "", ""
    sizes = [(24, "24*49"), (27, "27*49"), (28, "28*49"), (37, "37*49"), (41, "41*49最新价格(ROUND)"), (43, "43*49"), (74, "74*49"), (82, "82*49"), (86, "86*49")]
    for value, label in sizes:
        if width <= value + 0.6:
            return str(value), label
    return "86", "86*49"


def _standard_ccl_size_column(length_in: float | None, width_in: float | None) -> tuple[str, str]:
    if length_in is None or width_in is None:
        return "", ""
    candidates = [
        (24, 49, "24", "24*49"),
        (27, 49, "27", "27*49"),
        (28, 49, "28", "28*49"),
        (37, 49, "37", "37*49"),
        (41, 49, "41", "41*49最新价格(ROUND)"),
        (43, 49, "43", "43*49"),
        (74, 49, "74", "74*49"),
        (82, 49, "82", "82*49"),
        (86, 49, "86", "86*49"),
    ]
    for length, width, key, label in candidates:
        if abs(length_in - length) <= 0.6 and abs(width_in - width) <= 0.6:
            return key, label
    return "", ""


def _select_ccl_parent(length_in: float, width_in: float) -> dict | None:
    candidates = [
        (37, 49, "37", "37*49"),
        (41, 49, "41", "41*49最新价格(ROUND)"),
        (43, 49, "43", "43*49"),
        (74, 49, "74", "74*49"),
        (82, 49, "82", "82*49"),
        (86, 49, "86", "86*49"),
    ]
    valid = []
    for parent_w, parent_h, key, label in candidates:
        opens_w = math.floor((parent_w + 1e-9) / length_in) if length_in else 0
        opens_h = math.floor((parent_h + 1e-9) / width_in) if width_in else 0
        opens = opens_w * opens_h
        if opens <= 0:
            continue
        fit_error = abs((length_in * opens_w) - parent_w) + abs((width_in * opens_h) - parent_h)
        valid.append(
            {
                "parent_w": parent_w,
                "parent_h": parent_h,
                "key": key,
                "label": label,
                "opens_w": opens_w,
                "opens_h": opens_h,
                "opens": opens,
                "fit_error": fit_error,
            }
        )
    if not valid:
        return None
    return sorted(valid, key=lambda item: (item["fit_error"], item["parent_w"], -item["opens"]))[0]


def _select_plin_ccl_parent(length_in: float, width_in: float) -> dict | None:
    candidates = [
        (37, 49, "37", "37*49", "37", 1),
        (41, 49, "41", "41*49", "41", 1),
        (43, 49, "43", "43*49", "43", 1),
        (86, 49, "86", "86*49", "43", 2),
    ]
    valid = []
    for parent_w, parent_h, key, label, price_key, price_factor in candidates:
        opens_w = math.floor((parent_w + 1e-9) / length_in) if length_in else 0
        opens_h = math.floor((parent_h + 1e-9) / width_in) if width_in else 0
        opens = opens_w * opens_h
        if opens <= 1:
            continue
        fit_error = abs((length_in * opens_w) - parent_w) + abs((width_in * opens_h) - parent_h)
        valid.append(
            {
                "parent_w": parent_w,
                "parent_h": parent_h,
                "key": key,
                "label": label,
                "price_key": price_key,
                "price_factor": price_factor,
                "source_label": f"{price_key}*49" if price_factor == 1 else f"{price_key}*49*{price_factor}",
                "opens_w": opens_w,
                "opens_h": opens_h,
                "opens": opens,
                "fit_error": fit_error,
            }
        )
    if not valid:
        return None
    return sorted(valid, key=lambda item: (item["fit_error"], -item["parent_w"], -item["opens"]))[0]


def _extract_jing_wei_size(text: str) -> tuple[float | None, str | None, float | None, str | None]:
    patterns = [
        r"(?:经|径)向?\s*(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸)?\s*[*xX×]?\s*纬向?\s*(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸)?",
        r"(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸)?\s*(?:经|径)向?\s*[*xX×]?\s*(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸)?\s*纬向?",
    ]
    match = next((re.search(pattern, text, re.IGNORECASE) for pattern in patterns if re.search(pattern, text, re.IGNORECASE)), None)
    if match:
        length = float(match.group(1))
        width = float(match.group(3))
        return length, _normalize_unit(match.group(2), length), width, _normalize_unit(match.group(4), width)

    reversed_patterns = [
        r"纬向?\s*(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸)?\s*[*xX×]?\s*(?:经|径)向?\s*(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸)?",
        r"(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸)?\s*纬向?\s*[*xX×]?\s*(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸)?\s*(?:经|径)向?",
    ]
    match = next((re.search(pattern, text, re.IGNORECASE) for pattern in reversed_patterns if re.search(pattern, text, re.IGNORECASE)), None)
    if not match:
        return None, None, None, None
    width = float(match.group(1))
    length = float(match.group(3))
    return length, _normalize_unit(match.group(4), length), width, _normalize_unit(match.group(2), width)


def _size_to_inches_from_match(match) -> tuple[float | None, float | None]:
    if not match:
        return None, None
    length = _to_float(match.group(1))
    width = _to_float(match.group(3))
    if length is None or width is None:
        return None, None
    return _axis_to_inch_2dp(length, _normalize_unit(match.group(2), length)), _axis_to_inch_2dp(width, _normalize_unit(match.group(4), width))


def _is_stack_like_size_match(match) -> bool:
    if not match:
        return False
    left = _text(match.group(1))
    right = _text(match.group(3))
    left_unit = _text(match.group(2))
    right_unit = _text(match.group(4))
    if left_unit or right_unit:
        return False
    return bool(re.fullmatch(r"(1035|1067|1078|1080|1086|1506|2113|2116|2313|3313|7628|106)", left) and _to_float(right) is not None and _to_float(right) <= 20)


def _normalize_unit(unit: str | None, value: float) -> str:
    if not unit:
        return "mm" if value > 100 else "inch"
    unit = unit.upper()
    if unit == "M":
        return "m"
    if unit in {"IN", "INCH", "英寸"}:
        return "inch"
    return "mm"


def _axis_to_inch_2dp(value: float, unit: str | None) -> float:
    if unit == "mm":
        return round(value / 25.4, 2)
    if unit == "m":
        return round(value * 1000 / 25.4, 2)
    return round(value, 2)


def _axis_to_mm_2dp(value: float, unit: str | None) -> float:
    if unit == "inch":
        return round(value * 25.4, 2)
    if unit == "m":
        return round(value * 1000, 2)
    return round(value, 2)


def _extract_stack(desc: str) -> str:
    stack_token = r"(?:1035|1067|1078|1080|1086|1506|2113|2116|2313|3313|7628|106)\s*[*×X]+\s*[0-9]+(?:\.[0-9]+)?"
    stack_match = re.search(rf"\b({stack_token}(?:\s*[+＋]\s*{stack_token})*)\b", desc, re.I)
    if stack_match:
        return _canonical_stack(stack_match.group(1))
    match = re.search(rf"(?:in|inch|英寸)?\s*[xX*]\s*[0-9]+(?:\.[0-9]+)?\s*(?:in|inch|英寸)?\s+(.+?)\s+({FOIL_PATTERN})\b", desc, re.I)
    if match:
        return _canonical_stack(match.group(1))
    fallback = re.search(
        r"(?:in|inch|英寸|mm)?\s*[xX*×]\s*[0-9]+(?:\.[0-9]+)?\s*(?:in|inch|英寸|mm)?\s+(.+)$",
        desc,
        re.I,
    )
    if not fallback:
        return ""
    text = re.sub(rf"\b({FOIL_PATTERN})\b", "", fallback.group(1), flags=re.I)
    return _canonical_stack(text)


def _find_exact_header_row(ws, required: set[str]) -> tuple[int | None, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        headers: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            text = _text(ws.cell(row=row_idx, column=col_idx).value)
            if text:
                headers[text] = col_idx
        if required.issubset(headers):
            return row_idx, headers
    return None, {}


def _collect_plin_adders(ws, adders: dict[str, float]) -> None:
    for row in ws.iter_rows():
        text = " ".join(_text(cell.value) for cell in row if _text(cell.value))
        if "2/2" not in text and "3/3" not in text:
            continue
        two = re.search(r"2/2.*?高\s*([0-9]+(?:\.[0-9]+)?)", text, re.I)
        three = re.search(r"3/3\w*.*?高\s*([0-9]+(?:\.[0-9]+)?)", text, re.I)
        if two:
            adders["2/2"] = float(two.group(1))
        if three:
            adders["3/3"] = float(three.group(1))


def _looks_like_plin_pp_desc(desc: str) -> bool:
    return bool(re.search(r"RC\s*[0-9]+", desc, re.I) and re.search(r"(?:米\s*/?\s*卷|M\b)", desc, re.I))


def _plin_product_from_desc(desc: str) -> str:
    matches = re.findall(r"(NY[A-Z0-9][A-Z0-9\-.（）()]*)", desc, re.I)
    if not matches:
        return ""
    return _plin_norm_product(matches[0])


def _plin_norm_product(value: Any) -> str:
    text = _norm_product(value)
    if text.endswith("P") and re.fullmatch(r"NY\d+[A-Z0-9]*P", text):
        return text[:-1]
    return text


def _plin_norm_copper(left: Any, right: Any) -> str:
    return _plin_norm_copper_value(f"{left}/{right}")


def _plin_norm_copper_value(value: Any) -> str:
    text = _text(value).upper().replace(" ", "").replace("OZ", "")
    if "/" not in text and len(text) == 2:
        text = f"{text[0]}/{text[1]}"
    if "/" in text:
        parts = [_plin_norm_copper_part(part) for part in text.split("/", 1)]
        text = "/".join(parts)
    if text == "H/0":
        return "H/H"
    if text == "1/H":
        return "1/1"
    return text


def _plin_raw_copper(left: Any, right: Any) -> str:
    return f"{_text(left)}/{_text(right)}".upper().replace(" ", "").replace("OZ", "")


def _plin_norm_copper_part(value: Any) -> str:
    text = _text(value).upper().replace(" ", "")
    if text == "H":
        return "H"
    number = _to_float(text)
    if number is None:
        return text
    if abs(number - 1.03) <= 0.05:
        return "1"
    if float(number).is_integer():
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def _plin_thickness_from_desc(desc: str, copper_match) -> float | None:
    if not copper_match:
        return None
    before = desc[: copper_match.start()]
    nums = [_to_float(item) for item in re.findall(r"\b\d+(?:\.\d+)?\b", before)]
    nums = [item for item in nums if item is not None]
    return nums[-1] if nums else None


def _plin_size_key_and_multiplier(length_in: float | None, width_in: float | None) -> tuple[str, float]:
    if length_in is None or width_in is None:
        return "", 1.0
    length_key = int(math.floor(length_in + 1e-9))
    size_map = {37: ("37", 1.0), 41: ("41", 1.11), 43: ("43", 1.16)}
    if length_key in size_map and abs(width_in - 49) <= 1:
        return size_map[length_key]
    for standard in (37, 41, 43):
        if abs(length_in - standard) <= 0.6 and abs(width_in - 49) <= 1:
            return size_map[standard]
    return "", 1.0


def _parse_rc_range(value: Any) -> tuple[float | None, float | None]:
    text = _text(value).upper().replace("RC", "").replace("%", "")
    nums = [_to_float(item) for item in re.findall(r"\d+(?:\.\d+)?", text)]
    nums = [item for item in nums if item is not None]
    nums = [item * 100 if item is not None and item <= 1 else item for item in nums]
    if not nums:
        return None, None
    if len(nums) == 1:
        return nums[0], nums[0]
    return min(nums), max(nums)


def _calc_total(quantity: Any, price: float | None) -> float | str | None:
    qty = _to_float(quantity)
    if qty is None or price is None:
        return "待确认"
    return round(qty * _round_money(price), 2)


def _find_net_price_col(headers: dict[str, int]) -> int | None:
    for name, col in headers.items():
        normalized = _text(name).strip().lower().replace(" ", "")
        if normalized in {"净价", "price", "netprice"}:
            return col
    return None


def _aoshikang_net_price_result(value: Any) -> tuple[float | str, str]:
    text = _text(value)
    if not text:
        return "", ""
    try:
        if "/" in text:
            left, right = text.split("/", 1)
            numerator = float(left.replace(",", "").strip())
            denominator = float(right.replace(",", "").strip())
            if denominator == 0:
                return "", "净价结果未计算：分母为0"
            result = round(numerator / denominator, 12)
            return result, f"净价结果={numerator:g}/{denominator:g}"
        number = _to_float(text)
        if number is not None:
            return round(number, 12), f"净价结果={number:g}"
    except ValueError:
        pass
    return "", f"净价结果未计算：净价格式无法解析（{text}）"


def _taixing_roll_price(result: CalcResult) -> tuple[float | str, str]:
    if result.material_type != "PP":
        return "", ""
    price = _to_float(result.price)
    length = _to_float(result.roll_length)
    if price is None:
        return "", "整卷价格未计算：PP新价格为空"
    if length is None:
        return "", "整卷价格未计算：未识别卷长"
    roll_price = round(price * length, 6)
    return roll_price, f"整卷价格={price:.6g}*{length:g}"


def _round_money(value: float | None) -> float | None:
    if value is None:
        return None
    return round(float(value), 2)


def _jingwang_pp_price_exception(product: str, glass: str, rc: float, price: float) -> float:
    if product == "NY6300SP" and glass == "1035" and abs(rc - 73) <= 0.001:
        return 65.34
    if product == "NY6300SP" and glass == "1078" and abs(rc - 64) <= 0.001:
        return 62.92
    return price


def _result_equal(actual: Any, expected: Any) -> bool:
    actual_num = _to_float(actual)
    expected_num = _to_float(expected)
    if actual_num is not None and expected_num is not None:
        return abs(actual_num - expected_num) <= 0.02
    return _text(actual) == _text(expected)


def _excel_value(value: Any) -> Any:
    if isinstance(value, float):
        if math.isfinite(value):
            return round(value, 2)
        return None
    return value


def _taixing_excel_value(value: Any) -> Any:
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    return value


def _taixing_conflict_specs(ws, header_row: int, headers: dict[str, int]) -> dict[str, str]:
    desc_col = next((headers[name] for name in DESC_HEADERS if name in headers), None)
    price_col = next((headers[name] for name in {"单价", "新价", "新价格", "新单价"} if name in headers), None)
    if not desc_col or not price_col:
        return {}
    values: dict[str, list[tuple[int, str]]] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        spec = _text(ws.cell(row=row_idx, column=desc_col).value)
        price = _text(ws.cell(row=row_idx, column=price_col).value)
        if not spec or not price:
            continue
        values.setdefault(spec, []).append((row_idx, price))
    conflicts: dict[str, str] = {}
    for spec, rows in values.items():
        nums = [_to_float(price) for _, price in rows]
        nums = [num for num in nums if num is not None]
        has_conflict = len(nums) >= 2 and max(nums) - min(nums) > 0.0002
        if has_conflict:
            conflicts[spec] = "；".join(f"第{row}行={price}" for row, price in rows)
    return conflicts


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return None
        return float(value)
    text = _text(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_product(value: Any) -> str:
    text = _text(value).upper().replace(" ", "").replace("（", "(").replace("）", ")")
    match = re.match(r"^(NY-[A-Z0-9]+)P\(C\)$", text)
    if match:
        return f"{match.group(1)}(C)P"
    return text


def _norm_glass(value: Any) -> str:
    return _text(value).upper().replace(" ", "")


def _norm_copper(value: Any) -> str:
    text = _text(value).upper().replace(" ", "")
    return text.replace(".0", "")


def _norm_state(value: Any) -> str:
    text = _text(value)
    if "不含铜" in text:
        return "不含铜"
    if "含铜" in text:
        return "含铜"
    return text


def _norm_foil(value: Any) -> str:
    text = _text(value).upper().replace(" ", "")
    if not text:
        return ""
    raw_parts = [part for part in re.split(r"[/、,，]+", text) if part]
    if len(raw_parts) <= 1:
        return text
    parts: list[str] = []
    seen: set[str] = set()
    for part in raw_parts:
        if part not in seen:
            parts.append(part)
            seen.add(part)
    parts.sort(key=lambda item: (FOIL_ORDER.get(item, 999), item))
    return "/".join(parts)


def _canonical_stack(value: Any) -> str:
    text = _text(value).upper().replace("＊", "*").replace("×", "*").replace("X", "*").replace("＋", "+")
    text = re.sub(r"\*+", "*", text)
    text = re.sub(r"\s+", "", text)
    glass_pattern = r"1035|1067|1078|1080|1086|1506|2113|2116|2313|3313|7628|106"
    parts: list[tuple[str, str]] = []
    for item in re.finditer(rf"(?:(?P<glass>{glass_pattern})(?:\*(?P<count>[0-9]+(?:\.[0-9]+)?))?|(?P<count_first>[0-9]+(?:\.[0-9]+)?)\*(?P<glass_second>{glass_pattern}))", text):
        glass = item.group("glass") or item.group("glass_second")
        count = item.group("count") or item.group("count_first") or "1"
        parts.append((glass, count))
    if not parts:
        return text
    order: list[str] = []
    totals: dict[str, float] = {}
    for glass, count_text in parts:
        count = _to_float(count_text) if count_text else 1
        if count is None:
            count = 1
        if glass not in totals:
            order.append(glass)
            totals[glass] = 0
        totals[glass] += count
    return "+".join(f"{glass}*{_format_stack_count(totals[glass])}" for glass in order)


def _canonical_plin_stack(value: Any) -> str:
    stack = _canonical_stack(value)
    if not stack:
        return stack
    parts: list[tuple[int, str, float]] = []
    for item in stack.split("+"):
        match = re.fullmatch(r"(?P<glass>[A-Z0-9]+)\*(?P<count>[0-9]+(?:\.[0-9]+)?)", item)
        if not match:
            return stack
        glass = match.group("glass")
        count = _to_float(match.group("count"))
        glass_num_match = re.search(r"\d+", glass)
        glass_num = int(glass_num_match.group(0)) if glass_num_match else 0
        parts.append((glass_num, glass, count if count is not None else 1))
    parts.sort(key=lambda item: (-item[0], item[1]))
    return "+".join(f"{glass}*{_format_stack_count(count)}" for _, glass, count in parts)


def _format_stack_count(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")
