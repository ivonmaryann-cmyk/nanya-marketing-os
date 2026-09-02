from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from .db import append_job_log, create_job, get_job, update_job_status
from .excel_utils import load_workbook_compat, normalized_xlsx_source
from .job_control import launch_job_process
from .paths import JOBS_DIR
from .pp_transcode_rules import (
    CUSTOMER_FIELDS,
    FIELD_META,
    list_base_rules,
    list_pp_confirmation_items,
    list_customer_rules,
    pp_confirmation_counts,
    replace_pp_confirmation_items,
    resolve_shared_pp_glue,
    seed_pp_transcode_rules,
)


PP_RESULT_HEADERS = (
    "PP待人工确认码值",
    "PP本次确认码值",
    "PP转码状态",
    "PP置信度",
    "PP系统分析",
    "PP人工确认状态",
)
PP_HEADER_ALIASES = {
    "customer_code": ("客户代码", "客户编号", "客户ID"),
    "customer_name": ("客户简称", "客户名称", "客户"),
    "spec": ("PP客户规格", "客户规格", "规格", "客户需求"),
    "order_remark": ("订单备注", "备注", "备注/整行上下文", "整行上下文"),
}


def _normalize(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def _stars(field_key: str) -> str:
    return "*" * FIELD_META[field_key]["width"]


def _empty_field(field_key: str, reason: str = "未命中已维护规则，使用占位符") -> dict[str, Any]:
    return {
        "field_key": field_key,
        "field_label": FIELD_META[field_key]["label"],
        "value": _stars(field_key),
        "confidence": 0,
        "source": "默认占位",
        "reason": reason,
    }


def _match_base_rule(field_key: str, normalized_spec: str) -> dict[str, Any] | None:
    matches = []
    for rule in list_base_rules(field_key=field_key, enabled="enabled"):
        input_value = _normalize(rule["input_value"])
        if input_value and input_value in normalized_spec:
            matches.append(rule)
    if not matches:
        return None
    # Avoid short aliases shadowing a more precise mapping.
    return max(matches, key=lambda item: len(_normalize(item["input_value"])))


def _parse_length(normalized_spec: str) -> str | None:
    # Whitespace is removed before parsing, so "玻布106 350M" becomes
    # "玻布106350M". The explicit M unit still gives PP length a stable boundary.
    match = re.search(r"(\d{1,3})M(?=RC|树脂|$|[^A-Z0-9])", normalized_spec)
    if not match:
        return None
    return f"{int(match.group(1)):03d}M"


def _parse_glass_style(normalized_spec: str) -> str | None:
    # Only accept explicit common glass-style patterns; arbitrary numbers must not be guessed.
    match = re.search(
        r"(?:玻布|玻璃布|布号|布规格)\s*[:：=]?\s*(\d{3,4}?)(?=\d{2,3}M|RC|树脂|$)"
        r"|(?:玻布|玻璃布|布号|布规格)\s*[:：=]?\s*(\d{3,4})(?!\d)"
        r"|(\d{3,4})布(?![\dA-Z])",
        normalized_spec,
    )
    if not match:
        return None
    return (match.group(1) or match.group(2) or match.group(3)).zfill(4)


def _parse_resin_content(normalized_spec: str) -> str | None:
    match = re.search(r"(?:树脂含量|树脂|RC|含胶量)\s*[:：=]?\s*(\d{2}(?:\.\d+)?)\s*%", normalized_spec)
    if not match:
        return None
    value = round(float(match.group(1)) * 10)
    if not 0 <= value <= 999:
        return None
    return f"{value:03d}"


def _resolve_base_field(field_key: str, normalized_spec: str) -> dict[str, Any]:
    if field_key == "glue_code":
        rule = resolve_shared_pp_glue(normalized_spec)
        if rule:
            confidence = 60 if rule.get("uncertain") else 100
            reason = rule.get("business_note") or "营销转码Agent统一胶系映射"
            if rule.get("conflict"):
                reason = f"{reason}；{rule['conflict']}"
            return {
                "field_key": field_key,
                "field_label": FIELD_META[field_key]["label"],
                "value": rule["output_value"],
                "confidence": confidence,
                "source": "营销转码Agent统一胶系映射",
                "reason": reason,
                "rule_id": rule["id"],
                "matched_value": rule["input_value"],
            }
        return _empty_field(field_key, "未命中营销转码Agent当前胶系主表或别名映射，使用占位符")
    rule = _match_base_rule(field_key, normalized_spec)
    if rule:
        return {
            "field_key": field_key,
            "field_label": FIELD_META[field_key]["label"],
            "value": rule["output_value"],
            "confidence": 100,
            "source": "已维护基础规则",
            "reason": rule.get("business_note") or f"规格命中 {rule['input_value']}",
            "rule_id": rule["id"],
            "matched_value": rule["input_value"],
        }
    direct_value = None
    if field_key == "pp_length":
        direct_value = _parse_length(normalized_spec)
    elif field_key == "glass_style":
        direct_value = _parse_glass_style(normalized_spec)
    elif field_key == "resin_content":
        direct_value = _parse_resin_content(normalized_spec)
    if direct_value:
        return {
            "field_key": field_key,
            "field_label": FIELD_META[field_key]["label"],
            "value": direct_value[: FIELD_META[field_key]["width"]].ljust(FIELD_META[field_key]["width"], "*"),
            "confidence": 100,
            "source": "规格直接识别",
            "reason": "已从客户规格按编码规范直接识别并完成格式校验",
            "matched_value": direct_value,
        }
    return _empty_field(field_key)


def _condition_value(condition_field: str, context: dict[str, str], fields: dict[str, dict[str, Any]]) -> str:
    if condition_field == "胶系":
        return fields["glue_code"].get("matched_value") or fields["glue_code"]["value"]
    if condition_field == "玻布规格":
        return fields["glass_style"].get("matched_value") or fields["glass_style"]["value"]
    if condition_field == "PP长度":
        return fields["pp_length"].get("matched_value") or fields["pp_length"]["value"]
    if condition_field == "树脂含量":
        return fields["resin_content"].get("matched_value") or fields["resin_content"]["value"]
    if condition_field == "订单备注":
        return context["order_remark"]
    return context["spec"]


def _rule_matches(rule: dict[str, Any], context: dict[str, str], fields: dict[str, dict[str, Any]]) -> bool:
    for condition in rule.get("conditions") or []:
        actual = _normalize(_condition_value(condition["field"], context, fields))
        expected = _normalize(condition["value"])
        operator = condition["operator"]
        if operator == "equals" and actual != expected:
            return False
        if operator == "contains" and expected not in actual:
            return False
        if operator == "not_contains" and expected in actual:
            return False
    return True


def _resolve_customer_fields(context: dict[str, str], fields: dict[str, dict[str, Any]]) -> None:
    rules = list_customer_rules(context["customer_code"], context["customer"], enabled="enabled")
    for field_key in CUSTOMER_FIELDS:
        matches = [rule for rule in rules if rule["target_field"] == field_key and _rule_matches(rule, context, fields)]
        if not matches:
            fields[field_key] = _empty_field(field_key, "尚未维护该客户的 PP 特殊规则，使用占位符")
            continue
        # Customer-code records take precedence over name/global records, then more conditions.
        selected = max(
            matches,
            key=lambda item: (
                2 if context["customer_code"] and item["customer_code"] == context["customer_code"] else 1 if context["customer"] and item["customer_name"] == context["customer"] else 0,
                len(item.get("conditions") or []),
                item["id"],
            ),
        )
        fields[field_key] = {
            "field_key": field_key,
            "field_label": FIELD_META[field_key]["label"],
            "value": selected["output_value"],
            "confidence": 100,
            "source": "客户特殊规则",
            "reason": selected.get("business_note") or "命中已维护客户特殊规则",
            "rule_id": selected["id"],
        }


def calculate_pp_transcode_quote(
    spec: str,
    *,
    customer: str = "",
    customer_code: str = "",
    order_remark: str = "",
) -> dict[str, Any]:
    """Generate a 27-position PP candidate code. First release never emits a formal code."""
    # The service is also callable outside the Flask request lifecycle.
    seed_pp_transcode_rules()
    spec = str(spec or "").strip()
    if not spec:
        return {"status": "失败", "formal_code": "", "pending_code": "", "error": "请输入 PP 客户规格"}
    context = {
        "spec": spec,
        "customer": str(customer or "").strip(),
        "customer_code": str(customer_code or "").strip(),
        "order_remark": str(order_remark or "").strip(),
    }
    normalized_spec = _normalize(spec)
    fields: dict[str, dict[str, Any]] = {}
    for field_key in ("glue_code", "glass_style", "pp_length", "formula_category", "resin_content"):
        # The shared Marketing Agent glue resolver relies on token boundaries.
        # Keep the original spacing for glue names such as "NY2150 1080";
        # compact normalization remains useful for PP numeric fields.
        field_source = spec if field_key == "glue_code" else normalized_spec
        fields[field_key] = _resolve_base_field(field_key, field_source)
    _resolve_customer_fields(context, fields)

    candidate_code = "".join(fields[key]["value"] for key in FIELD_META)
    unresolved = [item["field_label"] for item in fields.values() if item["source"] == "默认占位"]
    direct = [item["field_label"] for item in fields.values() if item["source"] == "规格直接识别"]
    note_parts = ["首期 PP 转码只生成待人工确认码值，不自动输出正式码。"]
    if unresolved:
        note_parts.append(f"待补充规则字段：{'、'.join(unresolved)}。")
    if direct:
        note_parts.append(f"规格直接识别字段：{'、'.join(direct)}，建议确认后维护为基础规则。")
    return {
        "status": "待人工确认",
        "formal_code": "",
        "candidate_code": candidate_code,
        "pending_code": candidate_code,
        "confidence": min((item["confidence"] for item in fields.values()), default=0),
        "summary": " ".join(note_parts),
        "note": " ".join(note_parts),
        "reason": "PP 转码首期统一进入人工确认。",
        "field_evidence": [fields[key] for key in FIELD_META],
        "rule_version": "PP页面规则库",
        "agent_rule_version": "PP-Agent-Phase-1",
        "requires_manual_confirmation": True,
        "unresolved_fields": unresolved,
    }


def queue_pp_transcode_job(employee_id: str, uploaded_file: FileStorage, source_filename: str) -> int:
    """Queue PP batch work. Its rules and result files stay independent from CCL Agent."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    safe_filename = secure_filename(source_filename) or f"pp_transcode_{timestamp}.xlsx"
    input_path = employee_dir / f"{timestamp}_pp_transcode_{safe_filename}"
    uploaded_file.save(input_path)
    job_id = create_job(
        employee_id,
        source_filename,
        str(input_path),
        "PP页面规则库；首期仅待人工确认码值",
        feature="pp_transcode_agent",
    )
    launch_job_process(job_id, "pp_transcode_agent", employee_id)
    return job_id


def queue_pp_transcode_single_job(
    employee_id: str,
    *,
    spec: str,
    customer: str = "",
    customer_code: str = "",
    order_remark: str = "",
) -> int:
    """Create a one-row workbook so PP single and batch inputs use one runtime."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    source_filename = f"PP单条转码_{timestamp}.xlsx"
    input_path = employee_dir / f"{timestamp}_pp_transcode_single.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "PP转码需求表"
    worksheet.append(["客户代码", "客户简称", "PP客户规格", "订单备注"])
    worksheet.append(
        [
            str(customer_code or "").strip(),
            str(customer or "").strip(),
            str(spec or "").strip(),
            str(order_remark or "").strip(),
        ]
    )
    workbook.save(input_path)
    workbook.close()

    job_id = create_job(
        employee_id,
        source_filename,
        str(input_path),
        "PP页面规则库；单条与批量共用处理链路；首期仅待人工确认码值",
        feature="pp_transcode_agent",
    )
    launch_job_process(job_id, "pp_transcode_agent", employee_id)
    return job_id


def _header_index(worksheet, aliases: tuple[str, ...]) -> int | None:
    normalized = {_normalize(cell.value): int(cell.column) for cell in worksheet[1] if cell.value is not None}
    for alias in aliases:
        matched = normalized.get(_normalize(alias))
        if matched:
            return matched
    return None


def _ensure_result_columns(worksheet) -> dict[str, int]:
    existing = {_normalize(cell.value): int(cell.column) for cell in worksheet[1] if cell.value is not None}
    columns: dict[str, int] = {}
    next_column = worksheet.max_column + 1
    for header in PP_RESULT_HEADERS:
        key = _normalize(header)
        column = existing.get(key)
        if not column:
            column = next_column
            next_column += 1
            worksheet.cell(row=1, column=column, value=header)
        columns[header] = column
    return columns


def _cell_text(worksheet, row: int, column: int | None) -> str:
    return str(worksheet.cell(row=row, column=column).value or "").strip() if column else ""


def _write_pp_result_row(worksheet, row: int, result_columns: dict[str, int], quote: dict[str, Any], confirmation_status: str) -> None:
    worksheet.cell(row=row, column=result_columns["PP待人工确认码值"], value=quote.get("pending_code") or "")
    worksheet.cell(row=row, column=result_columns["PP本次确认码值"], value="")
    worksheet.cell(row=row, column=result_columns["PP转码状态"], value=quote.get("status") or "待人工确认")
    worksheet.cell(row=row, column=result_columns["PP置信度"], value=int(quote.get("confidence") or 0))
    worksheet.cell(row=row, column=result_columns["PP系统分析"], value=quote.get("summary") or quote.get("reason") or "")
    worksheet.cell(row=row, column=result_columns["PP人工确认状态"], value=confirmation_status)


def run_pp_transcode_job(job_id: int, employee_id: str) -> None:
    update_job_status(job_id, status="running", log_text="")
    job = get_job(job_id)
    if not job:
        return
    append_job_log(job_id, "开始 PP 转码Agent任务：首期只生成待人工确认码值，不输出正式码。")
    workbook = None
    try:
        workbook = load_workbook_compat(job["stored_input_path"])
        source_path = normalized_xlsx_source(job["stored_input_path"], workbook)
        if source_path != Path(job["stored_input_path"]):
            workbook.close()
            workbook = load_workbook_compat(source_path)
        worksheet = workbook.active
        header_indexes = {key: _header_index(worksheet, aliases) for key, aliases in PP_HEADER_ALIASES.items()}
        if not header_indexes["spec"]:
            raise ValueError("未识别 PP 客户规格列，请使用“PP客户规格”“客户规格”或“规格”列。")
        result_columns = _ensure_result_columns(worksheet)
        effective_rows = [row for row in range(2, worksheet.max_row + 1) if _cell_text(worksheet, row, header_indexes["spec"])]
        total = len(effective_rows)
        update_job_status(job_id, status="running", total_rows=total, current_row=0)
        append_job_log(job_id, f"识别到 {total} 行 PP 客户规格，开始统一生成待人工确认码值。", total_rows=total)

        items: list[dict[str, Any]] = []
        failures = 0
        for processed, row in enumerate(effective_rows, start=1):
            spec = _cell_text(worksheet, row, header_indexes["spec"])
            customer_code = _cell_text(worksheet, row, header_indexes["customer_code"])
            customer_name = _cell_text(worksheet, row, header_indexes["customer_name"])
            order_remark = _cell_text(worksheet, row, header_indexes["order_remark"])
            quote = calculate_pp_transcode_quote(
                spec,
                customer_code=customer_code,
                customer=customer_name,
                order_remark=order_remark,
            )
            if quote.get("status") == "失败":
                failures += 1
            else:
                items.append(
                    {
                        "excel_row": row,
                        "customer_code": customer_code,
                        "customer_name": customer_name,
                        "spec": spec,
                        "order_remark": order_remark,
                        "pending_code": quote.get("pending_code") or "",
                        "confidence": quote.get("confidence") or 0,
                        "summary": quote.get("summary") or quote.get("reason") or "",
                        "field_evidence": quote.get("field_evidence") or [],
                    }
                )
            _write_pp_result_row(worksheet, row, result_columns, quote, "待人工确认")
            if processed == total or processed % 25 == 0:
                append_job_log(job_id, f"已处理 {processed}/{total} 行。", current_row=processed, total_rows=total)

        result_path = Path(source_path).with_name(f"{Path(source_path).stem}_PP转码结果.xlsx")
        workbook.save(result_path)
        replace_pp_confirmation_items(job_id, employee_id, items)
        counts = pp_confirmation_counts(job_id, employee_id)
        update_job_status(
            job_id,
            status="awaiting_confirmation" if counts["pending"] else "completed",
            stored_result_path=str(result_path),
            success_count=0,
            fail_count=failures,
            skip_count=0,
            confirm_count=counts["pending"],
            current_row=total,
            total_rows=total,
            completed=not counts["pending"],
        )
        append_job_log(job_id, f"PP 结果文件已生成：{counts['pending']} 行待人工确认，正式码保持为空。")
    finally:
        if workbook is not None:
            workbook.close()


def refresh_pp_result_file(job_id: int, employee_id: str) -> None:
    """Synchronize confirmation progress back to PP output; never write a formal code."""
    job = get_job(job_id)
    if not job or not job["stored_result_path"]:
        return
    result_path = Path(job["stored_result_path"])
    if not result_path.exists():
        return
    items = {item["excel_row"]: item for item in list_pp_confirmation_items(job_id, employee_id)}
    workbook = load_workbook_compat(result_path)
    try:
        worksheet = workbook.active
        result_columns = _ensure_result_columns(worksheet)
        status_labels = {"pending": "待人工确认", "confirmed": "本次已确认", "skipped": "暂不处理"}
        for row, item in items.items():
            confirmation_status = status_labels.get(item["confirmation_status"], "待人工确认")
            confirmed_code = item.get("confirmed_pending_code") or ""
            worksheet.cell(row=row, column=result_columns["PP本次确认码值"], value=confirmed_code)
            worksheet.cell(row=row, column=result_columns["PP人工确认状态"], value=confirmation_status)
            # PP 首期不产生正式码，但下载文件必须清楚展示当前行是否已完成业务确认。
            worksheet.cell(row=row, column=result_columns["PP转码状态"], value=confirmation_status)
        workbook.save(result_path)
    finally:
        workbook.close()
