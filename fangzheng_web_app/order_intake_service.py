from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta, timezone
from email.utils import parseaddr
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from .database import automation_cursor as db_cursor
from .db import utcnow
from .file_storage import resolve_attachment_path
from .customer_archive_service import (
    customer_routing_needs_attachment_content,
    identify_customer,
    match_customer_routing_rules,
)


ACTION_LABELS = {
    "unclassified": "暂不分流",
    "new_order": "录单",
    "order_change": "修改订单",
    "quotation": "报价",
}
ROUTABLE_ACTION_TYPES = {"new_order", "order_change", "quotation"}
STATUS_LABELS = {
    "pending_triage": "待处理",
    "pending_review": "处理中",
    "ready_for_erp": "待确认",
    "on_hold": "待补充",
    "archived": "已完成",
}
WORKFLOW_STAGE_LABELS = {
    "mail_triage": "1. 邮件分流",
    "order_identification": "2. 订单识别",
    "data_mapping": "3. 主数据与151映射",
    "erp_preparation": "4. 410录单准备",
    "completed": "5. 已提交，待回查",
}
MATCH_STATUS_LABELS = {
    "unmatched": "待匹配",
    "matched": "已匹配负责客户",
    "needs_assignment": "需调整归属",
}
DOCUMENT_STATUS_LABELS = {
    "pending": "待核对",
    "complete": "订单资料齐全",
    "missing": "缺少订单资料",
}
MAPPING_STATUS_LABELS = {
    "not_started": "尚未开始",
    "ready": "映射资料待核",
    "exception": "存在映射异常",
    "confirmed": "映射已确认",
}
ERP_PREPARE_STATUS_LABELS = {
    "not_started": "未准备",
    "waiting_interface": "等待接口/模板",
    "ready_for_entry": "可录入410",
    "submitted": "已提交待回查",
}

BUSINESS_TIMEZONE = ZoneInfo("Asia/Shanghai")
GLOBAL_RULE_OWNER = "__global__"
SYSTEM_ACTOR = "__system__"
MAX_ATTACHMENT_BYTES = 20 * 1024 * 1024
MAX_ATTACHMENT_PDF_PAGES = 20
MAX_ATTACHMENT_WORKBOOK_ROWS = 20_000


def _list_sender(sender: str) -> tuple[str, str]:
    """Return concise sender text for a list row without repeating an address."""
    raw = str(sender or "").strip()
    name, address = parseaddr(raw)
    name = str(name or "").strip().strip('"')
    address = str(address or "").strip()
    if not address:
        return raw or "发件人未提供", ""
    if not name or name.casefold() == address.casefold():
        return address, ""
    return name, address


def _list_summary(body_text: str, fallback: str = "") -> str:
    """Keep the list focused on one useful, readable business summary."""
    text = re.sub(r"\s+", " ", str(body_text or "")).strip()
    if not text:
        text = re.sub(r"\s+", " ", str(fallback or "")).strip()
    return text[:116] + ("…" if len(text) > 116 else "")


def _metadata_get(conn, key: str) -> str | None:
    """Read automation metadata without relying on SQL dialect rewrites.

    SQLite keeps the legacy ``settings`` table for compatibility.  PostgreSQL
    explicitly uses ``automation_metadata`` so this module does not depend on
    the broad regexp replacement in ``database.sql``.
    """
    if getattr(conn, "dialect", "sqlite") == "postgresql":
        row = conn.execute(
            "SELECT value FROM settings WHERE key=?",
            (key,),
            postgres_sql="SELECT value FROM automation_metadata WHERE key=?",
        ).fetchone()
    else:
        row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    return str(row["value"]) if row else None


def _metadata_set(conn, key: str, value: str) -> None:
    if getattr(conn, "dialect", "sqlite") == "postgresql":
        conn.execute(
            "INSERT INTO settings(key,value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value),
            postgres_sql=(
                "INSERT INTO automation_metadata(key,value,updated_at) VALUES (?, ?, CURRENT_TIMESTAMP) "
                "ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value,updated_at=EXCLUDED.updated_at"
            ),
        )
    else:
        conn.execute(
            "INSERT INTO settings(key,value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value),
        )


def business_today() -> date:
    """Calendar day used by the Shanghai-based operations team."""
    return datetime.now(BUSINESS_TIMEZONE).date()


def _business_date(timestamp: str) -> str:
    """Convert legacy UTC timestamps (stored without an offset) to business date."""
    if not timestamp:
        return ""
    try:
        value = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(BUSINESS_TIMEZONE).date().isoformat()
    except ValueError:
        return timestamp[:10]


def _as_dict(row) -> dict[str, Any] | None:
    return dict(row) if row else None


def _json(value: str | None, fallback: Any) -> Any:
    try:
        return json.loads(value or "")
    except (TypeError, ValueError):
        return fallback


def _rule_matches(rule: dict[str, Any], *, subject: str, sender: str, attachment_text: str = "") -> bool:
    """A business rule is an AND of the conditions it actually fills in."""
    conditions = (
        (str(rule.get("sender_contains") or ""), sender),
        (str(rule.get("subject_contains") or ""), subject),
        (str(rule.get("attachment_contains") or ""), attachment_text),
    )
    filled = [(needle.strip().lower(), haystack.lower()) for needle, haystack in conditions if needle.strip()]
    return bool(filled) and all(needle in haystack for needle, haystack in filled)


def _match_rule(rules: list[dict[str, Any]], *, subject: str, sender: str, attachment_text: str = "") -> dict[str, Any] | None:
    return next((rule for rule in rules if _rule_matches(rule, subject=subject, sender=sender, attachment_text=attachment_text)), None)


def classify_mail(subject: str, sender: str = "", attachment_text: str = "", rules: list[dict[str, Any]] | None = None) -> tuple[str, str, int | None, dict[str, Any] | None]:
    """Generate an explainable route that business users can override."""
    matched_rule = _match_rule(rules or [], subject=subject, sender=sender, attachment_text=attachment_text)
    if matched_rule:
        return (
            str(matched_rule["action_type"]),
            f"命中业务规则：{matched_rule['name']}",
            int(matched_rule["id"]),
            matched_rule,
        )
    text = f"{subject} {sender}".lower()
    rules = (
        ("order_change", r"修改|变更|改单|改期|取消|暂停|撤销|作废|revision|revise|amend|change|cancel", "主题包含订单变更信号"),
        ("quotation", r"rfq|rfx|询价|报价|竞价|议价|投标|中标|bidding", "主题包含询报价信号"),
        ("delivery", r"交期|交货|发货|送货|出货|催交|排产|shipment|delivery", "主题包含交期或发货信号"),
        ("new_order", r"采购订单|新增订单|订单发布|订单需求|新订单|purchase\s*order|\bpo\b|下单", "主题包含新订单信号"),
    )
    for action_type, pattern, reason in rules:
        if re.search(pattern, text, flags=re.IGNORECASE):
            return action_type, reason, None, None
    return "other", "暂未匹配到明确业务类型", None, None


def list_routing_rules(employee_id: str, *, include_disabled: bool = True) -> list[dict[str, Any]]:
    with db_cursor() as conn:
        sql = "SELECT * FROM order_mail_routing_rules WHERE employee_id = ?"
        if not include_disabled:
            sql += " AND enabled = 1"
        sql += " ORDER BY enabled DESC, id ASC"
        rows = conn.execute(sql, (employee_id,)).fetchall()
    return [dict(row) for row in rows]


def get_routing_rule(rule_id: int, employee_id: str) -> dict[str, Any] | None:
    with db_cursor() as conn:
        row = conn.execute(
            "SELECT * FROM order_mail_routing_rules WHERE id = ? AND employee_id = ?",
            (rule_id, employee_id),
        ).fetchone()
    return _as_dict(row)


def _rule_payload(payload: dict[str, Any]) -> dict[str, Any]:
    data = {
        "name": str(payload.get("name") or "").strip(),
        "enabled": 1 if str(payload.get("enabled") or "") in {"1", "on", "true"} else 0,
        "priority": 100,
        "sender_contains": str(payload.get("sender_contains") or "").strip(),
        "subject_contains": str(payload.get("subject_contains") or "").strip(),
        "attachment_contains": str(payload.get("attachment_contains") or "").strip(),
        "action_type": str(payload.get("action_type") or "").strip(),
        "customer_code": str(payload.get("customer_code") or "").strip(),
        "customer_name": str(payload.get("customer_name") or "").strip(),
        "note": str(payload.get("note") or "").strip(),
    }
    if not data["name"]:
        raise ValueError("请填写规则名称")
    if data["action_type"] not in ACTION_LABELS:
        raise ValueError("请选择规则要分流到的业务类型")
    if not (data["sender_contains"] or data["subject_contains"] or data["attachment_contains"]):
        raise ValueError("请至少填写一个匹配条件")
    return data


def save_routing_rule(employee_id: str, payload: dict[str, Any], *, rule_id: int | None = None) -> dict[str, Any]:
    data = _rule_payload(payload)
    now = utcnow()
    with db_cursor() as conn:
        before: dict[str, Any] = {}
        if rule_id:
            row = conn.execute("SELECT * FROM order_mail_routing_rules WHERE id = ? AND employee_id = ?", (rule_id, employee_id)).fetchone()
            if not row:
                raise ValueError("分流规则不存在或无权修改")
            before = dict(row)
            conn.execute(
                """UPDATE order_mail_routing_rules SET name=?, enabled=?, priority=?, sender_contains=?, subject_contains=?,
                   attachment_contains=?, action_type=?, customer_code=?, customer_name=?, note=?, updated_at=? WHERE id=?""",
                (*data.values(), now, rule_id),
            )
            saved_id = rule_id
            event = "updated"
        else:
            cursor = conn.execute(
                """INSERT INTO order_mail_routing_rules (employee_id, name, enabled, priority, sender_contains, subject_contains,
                   attachment_contains, action_type, customer_code, customer_name, note, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (employee_id, *data.values(), now, now),
            )
            saved_id = int(cursor.lastrowid)
            event = "created"
        conn.execute(
            """INSERT INTO order_mail_routing_rule_events (rule_id, employee_id, action, before_json, after_json, created_at)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (saved_id, employee_id, event, json.dumps(before, ensure_ascii=False), json.dumps(data, ensure_ascii=False), now),
        )
    return get_routing_rule(saved_id, employee_id) or {}


EXTRACTED_RULE_PATTERNS = (
    {
        "key": "bomin-po", "name": "博敏采购订单", "sender_contains": "bominelec.com",
        "subject_contains": "博敏电子采购订单", "attachment_contains": "", "action_type": "new_order",
        "customer_name": "博敏电子", "note": "从历史邮件中提炼：博敏采购订单主题。",
    },
    {
        "key": "bomin-new-order", "name": "博敏新增订单需求", "sender_contains": "bominelec.com",
        "subject_contains": "新增订单需求", "attachment_contains": "", "action_type": "new_order",
        "customer_name": "博敏电子", "note": "从历史邮件中提炼：博敏新增订单需求通知。",
    },
    {
        "key": "chinafast-order-release", "name": "方正订单发布", "sender_contains": "chinafastprint.com",
        "subject_contains": "订单发布消息通知", "attachment_contains": "", "action_type": "new_order",
        "customer_name": "方正", "note": "从历史邮件中提炼：SRM 订单发布通知。",
    },
    {
        "key": "chinafast-rfq", "name": "方正询报价通知", "sender_contains": "chinafastprint.com",
        "subject_contains": "RFQ", "attachment_contains": "", "action_type": "quotation",
        "customer_name": "方正", "note": "从历史邮件中提炼：SRM RFQ 通知。",
    },
    {
        "key": "chinafast-award", "name": "方正中标通知", "sender_contains": "chinafastprint.com",
        "subject_contains": "中标通知", "attachment_contains": "", "action_type": "quotation",
        "customer_name": "方正", "note": "从历史邮件中提炼：SRM 中标通知。",
    },
    {
        "key": "chinafast-rfx-adjust", "name": "方正 RFX 时间调整", "sender_contains": "chinafastprint.com",
        "subject_contains": "RFX时间调整通知", "attachment_contains": "", "action_type": "unclassified",
        "customer_name": "方正", "note": "从历史邮件中提炼：此类通知暂不进入订单处理。",
    },
    {
        "key": "bomin-price-change", "name": "博敏单价修改申请", "sender_contains": "bominelec.com",
        "subject_contains": "单价修改申请", "attachment_contains": "", "action_type": "order_change",
        "customer_name": "博敏电子", "note": "从历史邮件中提炼：单价修改申请。",
    },
)


def list_extracted_routing_suggestions(employee_id: str, account_id: int | None) -> list[dict[str, Any]]:
    if not account_id:
        return []
    active_conditions = {
        (rule["sender_contains"].lower(), rule["subject_contains"].lower(), rule["attachment_contains"].lower())
        for rule in list_routing_rules(employee_id)
    }
    suggestions: list[dict[str, Any]] = []
    with db_cursor() as conn:
        for pattern in EXTRACTED_RULE_PATTERNS:
            conditions = (pattern["sender_contains"].lower(), pattern["subject_contains"].lower(), pattern["attachment_contains"].lower())
            if conditions in active_conditions:
                continue
            where = ["m.account_id = ?"]
            params: list[Any] = [account_id]
            for column, key in (("m.sender", "sender_contains"), ("m.subject", "subject_contains")):
                if pattern[key]:
                    where.append(f"LOWER({column}) LIKE ?")
                    params.append(f"%{pattern[key].lower()}%")
            row = conn.execute(
                f"SELECT COUNT(*) AS total, MAX(COALESCE(m.sent_at, m.received_at, m.created_at)) AS latest_at "
                f"FROM mail_messages m WHERE {' AND '.join(where)}",
                params,
            ).fetchone()
            if row and int(row["total"]):
                item = dict(pattern)
                item["match_count"] = int(row["total"])
                item["latest_at"] = str(row["latest_at"] or "")
                suggestions.append(item)
    return sorted(suggestions, key=lambda item: item["match_count"], reverse=True)


def save_extracted_routing_suggestion(employee_id: str, suggestion_key: str, *, decision: str) -> dict[str, Any]:
    pattern = next((item for item in EXTRACTED_RULE_PATTERNS if item["key"] == suggestion_key), None)
    if not pattern:
        raise ValueError("分流建议不存在")
    if decision not in {"route", "ignore"}:
        raise ValueError("请选择明确分流或暂不分流")
    payload = dict(pattern)
    payload["enabled"] = "1"
    payload["action_type"] = pattern["action_type"] if decision == "route" else "unclassified"
    payload["name"] = pattern["name"] if decision == "route" else f"{pattern['name']}（暂不分流）"
    payload["note"] = f"{pattern['note']} 业务已确认：{'明确分流' if decision == 'route' else '暂不分流'}。"
    return save_routing_rule(employee_id, payload)


def toggle_routing_rule(rule_id: int, employee_id: str) -> dict[str, Any]:
    current = get_routing_rule(rule_id, employee_id)
    if not current:
        raise ValueError("分流规则不存在或无权操作")
    now = utcnow()
    enabled = 0 if current["enabled"] else 1
    with db_cursor() as conn:
        conn.execute("UPDATE order_mail_routing_rules SET enabled=?, updated_at=? WHERE id=?", (enabled, now, rule_id))
        conn.execute(
            "INSERT INTO order_mail_routing_rule_events (rule_id, employee_id, action, before_json, after_json, created_at) VALUES (?, ?, 'toggled', ?, ?, ?)",
            (rule_id, employee_id, json.dumps({"enabled": current["enabled"]}), json.dumps({"enabled": enabled}), now),
        )
    return get_routing_rule(rule_id, employee_id) or {}


def simulate_routing_rule(rule: dict[str, Any], employee_id: str, account_id: int, *, days: int = 30) -> dict[str, Any]:
    start_date = (business_today() - timedelta(days=max(1, days) - 1)).isoformat()
    with db_cursor() as conn:
        rows = conn.execute(
            """SELECT m.id, m.subject, m.sender, m.sent_at, m.received_at,
               COALESCE(GROUP_CONCAT(a.filename, ' '), '') AS attachment_text
               FROM mail_messages m LEFT JOIN mail_attachments a ON a.mail_id=m.id AND a.is_inline=0
               JOIN mail_accounts ma ON ma.id=m.account_id
               WHERE ma.owner_employee_id=? AND m.account_id=?
                 AND substr(COALESCE(NULLIF(m.sent_at, ''), NULLIF(m.received_at, ''), m.created_at), 1, 10) >= ?
               GROUP BY m.id ORDER BY COALESCE(m.sent_at, m.received_at, m.created_at) DESC""",
            (employee_id, account_id, start_date),
        ).fetchall()
    matches = [dict(row) for row in rows if _rule_matches(rule, subject=row["subject"], sender=row["sender"], attachment_text=row["attachment_text"])]
    return {"count": len(matches), "samples": matches[:5]}


SCOPE_LABELS = {"subject": "邮件主题", "body": "邮件正文", "attachment_name": "附件名称", "attachment_content": "附件内容"}
ROUTING_STATE_LABELS = {"routed": "已明确分流", "needs_business_routing": "待业务分流", "unrouted": "暂不分流", "business_routed": "业务已分流"}
CLEAR_ROUTING_SCOPES = {"subject", "attachment_name"}

DEFAULT_GROUPS = (
    ("录单", "new_order", (
        ("subject", "采购订单"), ("subject", "新增订单需求"), ("subject", "订单发布"), ("subject", "订单需求"),
        ("body", "采购订单"), ("body", "订单号"), ("body", "下单"),
        ("attachment_name", "PO"), ("attachment_name", "采购订单"), ("attachment_name", "订单"),
        ("attachment_content", "采购订单"), ("attachment_content", "订单号"),
    )),
    ("报价", "quotation", (
        ("subject", "RFQ"), ("subject", "询价"), ("subject", "报价"), ("subject", "中标"),
        ("body", "RFQ"), ("body", "询价"), ("body", "报价"), ("body", "请报价"), ("body", "议价"),
        ("attachment_name", "RFQ"), ("attachment_name", "报价单"),
        ("attachment_content", "RFQ"), ("attachment_content", "报价"),
    )),
    ("修改订单", "order_change", (
        ("subject", "变更"), ("subject", "改期"), ("subject", "取消订单"), ("subject", "单价修改"),
        ("subject", "交期提前"), ("subject", "交期协同提前"), ("subject", "交期加急"), ("subject", "提前交货"), ("subject", "提前交期"),
        ("body", "变更"), ("body", "改期"), ("body", "价格调整"), ("body", "单价修改"), ("body", "版本更新"), ("body", "取消"),
        ("body", "交期提前"), ("body", "交期协同提前"), ("body", "交期加急"), ("body", "提前交货"), ("body", "提前交期"),
        ("attachment_name", "变更"), ("attachment_name", "改单"), ("attachment_name", "改期"),
        ("attachment_name", "交期提前"), ("attachment_name", "交期协同提前"), ("attachment_name", "交期加急"), ("attachment_name", "提前交货"), ("attachment_name", "提前交期"),
        ("attachment_content", "变更"), ("attachment_content", "价格调整"), ("attachment_content", "交期修改"),
        ("attachment_content", "交期提前"), ("attachment_content", "交期协同提前"), ("attachment_content", "交期加急"), ("attachment_content", "提前交货"), ("attachment_content", "提前交期"),
    )),
)
DEFAULT_CHANGE_TAGS = (
    ("交期 / 发货日期", (
        ("subject", "交期修改"), ("subject", "交期变更"), ("subject", "改期"),
        ("subject", "交期提前"), ("subject", "交期协同提前"), ("subject", "交期加急"), ("subject", "提前交货"), ("subject", "提前交期"),
        ("body", "交期修改"), ("body", "交期变更"), ("body", "改期"),
        ("body", "发货日期修改"), ("body", "发货日期变更"),
        ("body", "交期提前"), ("body", "交期协同提前"), ("body", "交期加急"), ("body", "提前交货"), ("body", "提前交期"),
        ("attachment_name", "交期提前"), ("attachment_name", "交期协同提前"), ("attachment_name", "交期加急"), ("attachment_name", "提前交货"), ("attachment_name", "提前交期"),
        ("attachment_content", "交期提前"), ("attachment_content", "交期协同提前"), ("attachment_content", "交期加急"), ("attachment_content", "提前交货"), ("attachment_content", "提前交期"),
    )),
    ("价格调整", (("subject", "单价修改"), ("subject", "价格调整"), ("body", "价格调整"), ("body", "单价修改"), ("body", "单价调整"), ("body", "调价"))),
    ("数量调整", (("body", "数量调整"), ("body", "数量变更"))),
    ("规格 / 型号调整", (("subject", "规格调整"), ("subject", "规格变更"), ("subject", "型号调整"), ("subject", "型号变更"), ("body", "规格调整"), ("body", "规格变更"), ("body", "型号调整"), ("body", "型号变更"))),
    ("订单版本 / PO", (("body", "版本更新"), ("body", "修订"), ("body", "revision"))),
    ("取消 / 暂停", (("subject", "取消"), ("body", "暂停"), ("body", "撤销"))),
)


def _migrate_rule_groups_to_global(conn, now: str) -> None:
    """Copy existing personal rules into one shared, de-duplicated rule set.

    Legacy rows remain untouched as a rollback safety net.  All reads use the
    global owner after this migration, so stale per-user copies cannot alter a
    colleague's route.  The idempotency marker makes the migration safe on
    SQLite and PostgreSQL and safe to run on every application start.
    """
    marker = "order_mail_rule_global_migration_v1"
    if _metadata_get(conn, marker):
        return
    legacy_groups = conn.execute(
        "SELECT * FROM order_mail_rule_groups WHERE employee_id<>? AND action_type IN ('new_order','quotation','order_change') ORDER BY id",
        (GLOBAL_RULE_OWNER,),
    ).fetchall()
    for raw_group in legacy_groups:
        group = dict(raw_group)
        target = conn.execute(
            "SELECT id FROM order_mail_rule_groups WHERE employee_id=? AND name=? AND action_type=? ORDER BY id LIMIT 1",
            (GLOBAL_RULE_OWNER, group["name"], group["action_type"]),
        ).fetchone()
        if target is None:
            target_id = int(conn.execute(
                "INSERT INTO order_mail_rule_groups (employee_id,name,action_type,enabled,created_at,updated_at) VALUES (?,?,?,?,?,?)",
                (GLOBAL_RULE_OWNER, group["name"], group["action_type"], group["enabled"], now, now),
            ).lastrowid)
        else:
            target_id = int(target["id"])
        present = {(item["scope"], item["keyword"]) for item in conn.execute(
            "SELECT scope,keyword FROM order_mail_rule_keywords WHERE group_id=?", (target_id,)
        ).fetchall()}
        for keyword in conn.execute("SELECT scope,keyword FROM order_mail_rule_keywords WHERE group_id=?", (group["id"],)).fetchall():
            pair = (keyword["scope"], keyword["keyword"])
            if pair not in present:
                conn.execute(
                    "INSERT INTO order_mail_rule_keywords (group_id,scope,keyword,created_at) VALUES (?,?,?,?)",
                    (target_id, pair[0], pair[1], now),
                )
                present.add(pair)
    _metadata_set(conn, marker, now)


def ensure_universal_rules(employee_id: str) -> None:
    """Ensure shared universal rules and per-user change tags exist.

    Change tags deliberately stay personal until the product decision C2 is
    confirmed.  Universal routing groups are global as already agreed.
    """
    now = utcnow()
    with db_cursor() as conn:
        _migrate_rule_groups_to_global(conn, now)
        obsolete_ids = [row["id"] for row in conn.execute(
            "SELECT id FROM order_mail_rule_groups WHERE employee_id=? AND action_type='unclassified'", (GLOBAL_RULE_OWNER,)
        ).fetchall()]
        for group_id in obsolete_ids:
            conn.execute("DELETE FROM order_mail_rule_keywords WHERE group_id=?", (group_id,))
            conn.execute("DELETE FROM order_mail_rule_groups WHERE id=?", (group_id,))

        seed_key = "order_mail_rule_seed_v3:global"
        seeded = _metadata_get(conn, seed_key) is not None
        for name, action_type, keywords in DEFAULT_GROUPS:
            group = conn.execute(
                "SELECT id FROM order_mail_rule_groups WHERE employee_id=? AND action_type=? ORDER BY id LIMIT 1",
                (GLOBAL_RULE_OWNER, action_type),
            ).fetchone()
            if group is None:
                group_id = int(conn.execute(
                    "INSERT INTO order_mail_rule_groups (employee_id,name,action_type,enabled,created_at,updated_at) VALUES (?,?,?,?,?,?)",
                    (GLOBAL_RULE_OWNER, name, action_type, 1, now, now),
                ).lastrowid)
            else:
                group_id = int(group["id"])
            if not seeded:
                existing_pairs = {(row["scope"], row["keyword"]) for row in conn.execute(
                    "SELECT scope,keyword FROM order_mail_rule_keywords WHERE group_id=?", (group_id,)
                ).fetchall()}
                for scope, keyword in keywords:
                    if (scope, keyword) not in existing_pairs:
                        conn.execute("INSERT INTO order_mail_rule_keywords (group_id,scope,keyword,created_at) VALUES (?,?,?,?)", (group_id, scope, keyword, now))
        if not seeded:
            _metadata_set(conn, seed_key, now)

        # Change tags remain backward-compatible and per employee pending C2.
        tags_exist = conn.execute("SELECT 1 FROM order_change_tags WHERE employee_id=?", (employee_id,)).fetchone()
        if not tags_exist:
            for name, keywords in DEFAULT_CHANGE_TAGS:
                tag_id = int(conn.execute(
                    "INSERT INTO order_change_tags (employee_id,name,enabled,created_at,updated_at) VALUES (?,?,?,?,?)",
                    (employee_id, name, 1, now, now),
                ).lastrowid)
                for scope, keyword in keywords:
                    conn.execute("INSERT INTO order_change_tag_keywords (tag_id,scope,keyword,created_at) VALUES (?,?,?,?)", (tag_id, scope, keyword, now))

        # Preserve the existing conservative refinements during the global
        # migration: a bare “交期”/“规格”/“单价” in a normal PO is not a
        # change signal.  Rule keys are global; change-tag keys remain scoped.
        cleanup_key = "order_mail_rule_cleanup_v4:global"
        if _metadata_get(conn, cleanup_key) is None:
            change_group = conn.execute(
                "SELECT id FROM order_mail_rule_groups WHERE employee_id=? AND action_type='order_change' ORDER BY id LIMIT 1",
                (GLOBAL_RULE_OWNER,),
            ).fetchone()
            if change_group:
                conn.execute("DELETE FROM order_mail_rule_keywords WHERE group_id=? AND keyword='交期'", (change_group["id"],))
            delivery_tag = conn.execute("SELECT id FROM order_change_tags WHERE employee_id=? AND name='交期 / 发货日期'", (employee_id,)).fetchone()
            if delivery_tag:
                conn.execute("DELETE FROM order_change_tag_keywords WHERE tag_id=? AND keyword='交期'", (delivery_tag["id"],))
            _metadata_set(conn, cleanup_key, now)

        refinement_key = f"order_change_item_refinement_v5:{employee_id}"
        if _metadata_get(conn, refinement_key) is None:
            for tag_name, removed in (("价格调整", {"单价"}), ("规格 / 型号调整", {"规格", "型号"})):
                tag = conn.execute("SELECT id FROM order_change_tags WHERE employee_id=? AND name=?", (employee_id, tag_name)).fetchone()
                if tag:
                    for word in removed:
                        conn.execute("DELETE FROM order_change_tag_keywords WHERE tag_id=? AND keyword=?", (tag["id"], word))
            _metadata_set(conn, refinement_key, now)

        acceleration_key = "order_change_delivery_acceleration_v7:global"
        if _metadata_get(conn, acceleration_key) is None:
            words = ("交期提前", "交期协同提前", "交期加急", "提前交货", "提前交期")
            scopes = ("subject", "body", "attachment_name", "attachment_content")
            change_group = conn.execute(
                "SELECT id FROM order_mail_rule_groups WHERE employee_id=? AND action_type='order_change' ORDER BY id LIMIT 1",
                (GLOBAL_RULE_OWNER,),
            ).fetchone()
            if change_group:
                existing = {(row["scope"], row["keyword"]) for row in conn.execute(
                    "SELECT scope,keyword FROM order_mail_rule_keywords WHERE group_id=?", (change_group["id"],)
                ).fetchall()}
                for scope in scopes:
                    for word in words:
                        if (scope, word) not in existing:
                            conn.execute("INSERT INTO order_mail_rule_keywords (group_id,scope,keyword,created_at) VALUES (?,?,?,?)", (change_group["id"], scope, word, now))
            _metadata_set(conn, acceleration_key, now)


def list_universal_rules(employee_id: str) -> list[dict[str, Any]]:
    ensure_universal_rules(employee_id)
    with db_cursor() as conn:
        groups = [dict(row) for row in conn.execute("SELECT * FROM order_mail_rule_groups WHERE employee_id=? ORDER BY id", (GLOBAL_RULE_OWNER,)).fetchall()]
        for group in groups:
            group["keywords"] = [dict(row) for row in conn.execute("SELECT * FROM order_mail_rule_keywords WHERE group_id=? ORDER BY id", (group["id"],)).fetchall()]
    return groups


def list_change_tags(employee_id: str) -> list[dict[str, Any]]:
    ensure_universal_rules(employee_id)
    with db_cursor() as conn:
        tags = [dict(row) for row in conn.execute("SELECT * FROM order_change_tags WHERE employee_id=? ORDER BY id", (employee_id,)).fetchall()]
        for tag in tags:
            tag["keywords"] = [dict(row) for row in conn.execute("SELECT * FROM order_change_tag_keywords WHERE tag_id=? ORDER BY id", (tag["id"],)).fetchall()]
    return tags


def _keyword_pairs(conn, *, group_id: int | None = None, tag_id: int | None = None, scope: str | None = None) -> list[tuple[str, str]]:
    if group_id is not None:
        table, id_column, identifier = "order_mail_rule_keywords", "group_id", group_id
    elif tag_id is not None:
        table, id_column, identifier = "order_change_tag_keywords", "tag_id", tag_id
    else:
        return []
    sql = f"SELECT scope,keyword FROM {table} WHERE {id_column}=?"
    params: list[Any] = [identifier]
    if scope:
        sql += " AND scope=?"
        params.append(scope)
    return [(str(row["scope"]), str(row["keyword"])) for row in conn.execute(sql, params).fetchall()]


def _changed_terms(*collections: list[tuple[str, str]]) -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    for collection in collections:
        for scope, keyword in collection:
            cleaned = keyword.strip().lower()
            if cleaned:
                result.setdefault(scope, set()).add(cleaned)
    return result


def save_universal_rule(employee_id: str, payload: dict[str, Any], group_id: int | None = None) -> dict[str, Any]:
    name, action_type = str(payload.get("name") or "").strip(), str(payload.get("action_type") or "").strip()
    keywords = [(str(x.get("scope") or "").strip(), str(x.get("keyword") or "").strip()) for x in payload.get("keywords", []) if str(x.get("keyword") or "").strip()]
    if not name or action_type not in ROUTABLE_ACTION_TYPES or not keywords or any(scope not in SCOPE_LABELS for scope, _ in keywords):
        raise ValueError("请填写规则名称、明确分流结果和至少一个有效关键词")
    now = utcnow()
    with db_cursor() as conn:
        old_keywords: list[tuple[str, str]] = []
        if group_id:
            row = conn.execute("SELECT id FROM order_mail_rule_groups WHERE id=? AND employee_id=?", (group_id, GLOBAL_RULE_OWNER)).fetchone()
            if not row: raise ValueError("规则不存在")
            old_keywords = _keyword_pairs(conn, group_id=group_id)
            conn.execute("UPDATE order_mail_rule_groups SET name=?,action_type=?,enabled=?,updated_at=? WHERE id=?", (name, action_type, 1 if payload.get("enabled", True) else 0, now, group_id))
            conn.execute("DELETE FROM order_mail_rule_keywords WHERE group_id=?", (group_id,))
            saved_id = group_id
        else:
            cursor = conn.execute("INSERT INTO order_mail_rule_groups (employee_id,name,action_type,enabled,created_at,updated_at) VALUES (?,?,?,?,?,?)", (GLOBAL_RULE_OWNER, name, action_type, 1 if payload.get("enabled", True) else 0, now, now))
            saved_id = int(cursor.lastrowid)
        for scope, keyword in keywords:
            conn.execute("INSERT INTO order_mail_rule_keywords (group_id,scope,keyword,created_at) VALUES (?,?,?,?)", (saved_id, scope, keyword, now))
    _reclassify_all_cases(_changed_terms(old_keywords, keywords))
    return next(item for item in list_universal_rules(employee_id) if item["id"] == saved_id)


def save_universal_rule_scope(employee_id: str, group_id: int, scope: str, keywords: list[str]) -> dict[str, Any]:
    """Replace just one search area of a fixed top-level business category."""
    if scope not in SCOPE_LABELS:
        raise ValueError("无效的检索位置")
    cleaned: list[str] = []
    for value in keywords:
        word = str(value or "").strip()
        if word and word not in cleaned:
            cleaned.append(word)
    now = utcnow()
    with db_cursor() as conn:
        group = conn.execute(
            "SELECT id FROM order_mail_rule_groups WHERE id=? AND employee_id=? AND action_type IN ('new_order','quotation','order_change')",
            (group_id, GLOBAL_RULE_OWNER),
        ).fetchone()
        if not group:
            raise ValueError("规则分类不存在")
        old_keywords = _keyword_pairs(conn, group_id=group_id, scope=scope)
        conn.execute("DELETE FROM order_mail_rule_keywords WHERE group_id=? AND scope=?", (group_id, scope))
        for word in cleaned:
            conn.execute("INSERT INTO order_mail_rule_keywords (group_id,scope,keyword,created_at) VALUES (?,?,?,?)", (group_id, scope, word, now))
        conn.execute("UPDATE order_mail_rule_groups SET updated_at=? WHERE id=?", (now, group_id))
    _reclassify_all_cases(_changed_terms(old_keywords, [(scope, word) for word in cleaned]))
    return next(item for item in list_universal_rules(employee_id) if item["id"] == group_id)


def save_change_tag(employee_id: str, name: str, keywords: list[dict[str, Any]], tag_id: int | None = None) -> int:
    """Save one business-facing order change type and its recognition phrases.

    ``all`` is deliberately a UI convenience: a business phrase should work
    wherever it appears, while the database still stores the four explicit
    source scopes required by the routing engine.
    """
    pairs: list[tuple[str, str]] = []
    for item in keywords:
        scope = str(item.get("scope") or "").strip()
        word = str(item.get("keyword") or "").strip()
        if not word:
            continue
        if scope == "all":
            pairs.extend((source_scope, word) for source_scope in SCOPE_LABELS)
        else:
            pairs.append((scope, word))
    pairs = list(dict.fromkeys(pairs))
    if not name.strip() or not pairs or any(scope not in SCOPE_LABELS for scope, _ in pairs):
        raise ValueError("请填写变更类型名称和至少一个业务说法")
    now = utcnow()
    with db_cursor() as conn:
        old_keywords: list[tuple[str, str]] = []
        if tag_id:
            existing = conn.execute("SELECT id FROM order_change_tags WHERE id=? AND employee_id=?", (tag_id, employee_id)).fetchone()
            if not existing:
                raise ValueError("变更类型不存在或无权修改")
            old_keywords = _keyword_pairs(conn, tag_id=tag_id)
            conn.execute("UPDATE order_change_tags SET name=?,updated_at=? WHERE id=? AND employee_id=?", (name.strip(), now, tag_id, employee_id))
            conn.execute("DELETE FROM order_change_tag_keywords WHERE tag_id=?", (tag_id,)); saved_id = tag_id
        else:
            existing = conn.execute("SELECT id FROM order_change_tags WHERE employee_id=? AND name=?", (employee_id, name.strip())).fetchone()
            if existing:
                raise ValueError("该变更类型已存在")
            cursor=conn.execute("INSERT INTO order_change_tags (employee_id,name,enabled,created_at,updated_at) VALUES (?,?,?,?,?)", (employee_id,name.strip(),1,now,now)); saved_id=int(cursor.lastrowid)
        for scope, keyword in pairs: conn.execute("INSERT INTO order_change_tag_keywords (tag_id,scope,keyword,created_at) VALUES (?,?,?,?)", (saved_id,scope,keyword,now))
    reclassify_cases(employee_id, changed_terms=_changed_terms(old_keywords, pairs))
    return saved_id


def _mail_values(row: dict[str, Any]) -> dict[str, str]:
    return {"subject": str(row.get("subject") or "").lower(), "body": str(row.get("body_text") or "").lower(), "attachment_name": str(row.get("attachment_names") or "").lower(), "attachment_content": str(row.get("attachment_content") or "").lower()}


def _extract_attachment_text(path: str) -> tuple[str, str]:
    try:
        path = str(resolve_attachment_path(path))
    except FileNotFoundError:
        return "", "missing"
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if not file_path.is_file():
        return "", "missing"
    if file_path.stat().st_size > MAX_ATTACHMENT_BYTES:
        return "", "too_large"
    try:
        if suffix in {".txt", ".csv"}:
            return file_path.read_text(encoding="utf-8", errors="replace")[:120000], "parsed"
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook
            book = load_workbook(file_path, read_only=True, data_only=True)
            values: list[str] = []
            row_count = 0
            try:
                for sheet in book.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        row_count += 1
                        if row_count > MAX_ATTACHMENT_WORKBOOK_ROWS:
                            return "", "too_large"
                        values.extend(str(v) for v in row if v is not None)
            finally:
                book.close()
            return " ".join(values)[:120000], "parsed"
        if suffix == ".pdf":
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                if len(pdf.pages) > MAX_ATTACHMENT_PDF_PAGES:
                    return "", "too_large"
                return " ".join((page.extract_text() or "") for page in pdf.pages)[:120000], "parsed"
        return "", "unsupported"
    except Exception:
        return "", "failed"


def _attachment_content(conn, mail_id: int, *, parse_missing: bool = False) -> str:
    """Return cached attachment text; parsing is opt-in for background workers."""
    rows = conn.execute("SELECT id,stored_path FROM mail_attachments WHERE mail_id=? AND is_inline=0", (mail_id,)).fetchall(); parts=[]
    for row in rows:
        cached=conn.execute("SELECT text_content,parse_status FROM mail_attachment_texts WHERE attachment_id=?", (row["id"],)).fetchone()
        if cached is None:
            if not parse_missing:
                conn.execute("INSERT INTO mail_attachment_texts (attachment_id,text_content,parse_status,updated_at) VALUES (?,?,?,?)", (row["id"],"","pending",utcnow()))
                continue
            text,status=_extract_attachment_text(row["stored_path"])
            conn.execute("INSERT INTO mail_attachment_texts (attachment_id,text_content,parse_status,updated_at) VALUES (?,?,?,?)", (row["id"],text,status,utcnow()))
            conn.execute("UPDATE mail_attachments SET parse_status=? WHERE id=?", (status, row["id"]))
            parts.append(text)
        else: parts.append(cached["text_content"])
    return " ".join(parts)


def prepare_attachment_texts(mail_ids: list[int]) -> None:
    """Parse new attachments in a worker, never in a rule-save request."""
    if not mail_ids:
        return
    with db_cursor() as conn:
        for mail_id in mail_ids:
            _attachment_content(conn, int(mail_id), parse_missing=True)


def _rule_decision(groups: list[dict[str, Any]], tags: list[dict[str, Any]], values: dict[str, str]) -> tuple[str, str, str, list[dict[str, Any]], list[str]]:
    """Pure keyword-rule decision shared by production reclassification/tests."""
    matches: list[dict[str, Any]] = []
    clear_action_types: set[str] = set()
    assist_action_types: set[str] = set()
    tag_names: list[str] = []
    for group in groups:
        if not group["enabled"]:
            continue
        for keyword in group["keywords"]:
            if keyword["keyword"].lower() in values[keyword["scope"]]:
                strength = "clear" if keyword["scope"] in CLEAR_ROUTING_SCOPES else "assist"
                matches.append({"scope": keyword["scope"], "keyword": keyword["keyword"], "group": group["name"], "action_type": group["action_type"], "strength": strength})
                (clear_action_types if strength == "clear" else assist_action_types).add(group["action_type"])
    for tag in tags:
        tag_hits = [k for k in tag["keywords"] if tag["enabled"] and k["keyword"].lower() in values[k["scope"]]]
        if tag_hits:
            tag_names.append(tag["name"])
            for hit in tag_hits:
                strength = "clear" if hit["scope"] in CLEAR_ROUTING_SCOPES else "assist"
                matches.append({"scope": hit["scope"], "keyword": hit["keyword"], "group": f"订单变更事项：{tag['name']}", "action_type": "order_change", "source": "change_item", "strength": strength})
                (clear_action_types if strength == "clear" else assist_action_types).add("order_change")
    if len(clear_action_types) == 1:
        return next(iter(clear_action_types)), "routed", "明确分流依据匹配", matches, tag_names
    if len(clear_action_types) > 1:
        return "unclassified", "needs_business_routing", "多个明确分流依据同时命中", matches, tag_names
    if len(assist_action_types) == 1:
        return next(iter(assist_action_types)), "routed", "辅助识别线索匹配", matches, tag_names
    if len(assist_action_types) > 1:
        return "unclassified", "needs_business_routing", "多个辅助识别线索同时命中", matches, tag_names
    return "unclassified", "unrouted", "未命中通用规则", matches, tag_names


def _like_pattern(value: str) -> str:
    return "%" + value.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_") + "%"


def _changed_terms_clause(changed_terms: dict[str, set[str]] | None) -> tuple[str, list[str]]:
    """Return a safe SQL filter for cases affected by a keyword edit."""
    if not changed_terms:
        return "", []
    column_by_scope = {
        "subject": "LOWER(m.subject)",
        "body": "LOWER(m.body_text)",
        "attachment_name": "EXISTS (SELECT 1 FROM mail_attachments a WHERE a.mail_id=m.id AND a.is_inline=0 AND LOWER(a.filename) LIKE ? ESCAPE '\\')",
        "attachment_content": "EXISTS (SELECT 1 FROM mail_attachments a JOIN mail_attachment_texts t ON t.attachment_id=a.id WHERE a.mail_id=m.id AND a.is_inline=0 AND LOWER(t.text_content) LIKE ? ESCAPE '\\')",
    }
    clauses: list[str] = []
    params: list[str] = []
    for scope, terms in changed_terms.items():
        source = column_by_scope.get(scope)
        if not source:
            continue
        for term in sorted(terms):
            if scope in {"attachment_name", "attachment_content"}:
                clauses.append(source)
            else:
                clauses.append(f"{source} LIKE ? ESCAPE '\\'")
            params.append(_like_pattern(term))
    return (" AND (" + " OR ".join(clauses) + ")", params) if clauses else ("", [])


def reclassify_cases(
    employee_id: str,
    account_id: int | None = None,
    *,
    changed_terms: dict[str, set[str]] | None = None,
) -> None:
    ensure_universal_rules(employee_id)
    groups, tags = list_universal_rules(employee_id), list_change_tags(employee_id)
    # Archived cases are immutable history.  Rule saves should only recalculate
    # active work, which makes the operation proportional to the queue rather
    # than every email ever imported.
    clauses=["c.employee_id=?", "c.routing_source != 'manual'", "c.status != 'archived'"]; params: list[Any]=[employee_id]
    if account_id: clauses.append("m.account_id=?"); params.append(account_id)
    changed_sql, changed_params = _changed_terms_clause(changed_terms)
    with db_cursor() as conn:
        needs_attachment_content=any(k["scope"] == "attachment_content" for group in groups for k in group["keywords"]) or any(k["scope"] == "attachment_content" for tag in tags for k in tag["keywords"])
        rows=conn.execute(f"""SELECT c.id,c.mail_id,c.action_type,c.routing_state,c.routing_source,c.routing_reason,
                                  c.routing_matches_json,c.change_tags_json,c.customer_id,c.customer_code,c.customer_name,
                                  c.customer_match_status,c.customer_match_source,c.customer_match_detail,
                                  m.subject,m.sender,m.body_text,
                                  COALESCE((SELECT GROUP_CONCAT(a.filename,' ') FROM mail_attachments a WHERE a.mail_id=m.id AND a.is_inline=0),'') attachment_names
                           FROM order_intake_cases c JOIN mail_messages m ON m.id=c.mail_id
                           WHERE {' AND '.join(clauses)}{changed_sql}""", [*params, *changed_params]).fetchall()
        now=utcnow()
        for raw in rows:
            row=dict(raw)
            # Missing text is marked pending.  The mail-fetch worker fills the
            # cache, after which its normal bootstrap/reclassification exposes
            # the attachment rule matches without blocking this request.
            row["attachment_content"] = _attachment_content(conn,row["mail_id"]) if needs_attachment_content else ""
            values = _mail_values(row)
            customer = identify_customer(str(row.get("sender") or ""))
            if customer and not row["attachment_content"] and customer_routing_needs_attachment_content(int(customer["customer_id"])):
                row["attachment_content"] = _attachment_content(conn, row["mail_id"])
                values = _mail_values(row)
            customer_result = match_customer_routing_rules(int(customer["customer_id"]), values) if customer else None
            matches=[]; tag_names=[]
            if customer_result:
                action_type = str(customer_result["action_type"])
                state = str(customer_result["state"])
                reason = str(customer_result["reason"])
                matches = [
                    {"source": "customer_identity", "scope": "sender", "keyword": customer["match_detail"], "group": "客户识别", "action_type": ""},
                    *list(customer_result["matches"]),
                ]
                routing_source = "customer_rule"
            else:
                routing_source = "keyword_rule"

                action_type, state, reason, matches, tag_names = _rule_decision(groups, tags, values)
            before = {
                "action_type": row["action_type"], "routing_state": row["routing_state"],
                "routing_source": row["routing_source"], "routing_reason": row["routing_reason"],
                "routing_matches": _json(row["routing_matches_json"], []), "change_tags": _json(row["change_tags_json"], []),
                "customer_id": row.get("customer_id"), "customer_code": row.get("customer_code"),
                "customer_name": row.get("customer_name"), "customer_match_status": row.get("customer_match_status"),
            }
            after = {
                "action_type": action_type, "routing_state": state, "routing_source": routing_source,
                "routing_reason": reason, "routing_matches": matches, "change_tags": tag_names,
                "customer_id": customer["customer_id"] if customer else None,
                "customer_code": customer["customer_code"] if customer else row.get("customer_code", ""),
                "customer_name": customer["customer_name"] if customer else row.get("customer_name", ""),
                "customer_match_status": "matched" if customer else "unmatched",
            }
            if before == after:
                continue
            conn.execute(
                """UPDATE order_intake_cases
                      SET action_type=?,routing_state=?,routing_source=?,routing_reason=?,routing_matches_json=?,change_tags_json=?,
                          customer_id=?,customer_code=COALESCE(NULLIF(?,''),customer_code),customer_name=COALESCE(NULLIF(?,''),customer_name),
                          customer_match_status=?,customer_match_source=?,customer_match_detail=?,updated_at=?
                    WHERE id=?""",
                (
                    action_type, state, routing_source, reason, json.dumps(matches,ensure_ascii=False), json.dumps(tag_names,ensure_ascii=False),
                    customer["customer_id"] if customer else None,
                    customer["customer_code"] if customer else "", customer["customer_name"] if customer else "",
                    "matched" if customer else "unmatched", customer["match_source"] if customer else "", customer["match_detail"] if customer else "",
                    now, row["id"],
                ),
            )
            conn.execute(
                "INSERT INTO order_intake_case_events (case_id,employee_id,action,before_json,after_json,created_at) VALUES (?,?,?,?,?,?)",
                (row["id"], SYSTEM_ACTOR, "auto_reclassify", json.dumps(before, ensure_ascii=False), json.dumps(after, ensure_ascii=False), now),
            )


def _reclassify_all_cases(changed_terms: dict[str, set[str]] | None = None) -> None:
    """Apply a shared-rule change only to active cases it can affect."""
    with db_cursor() as conn:
        owners = [str(row["employee_id"]) for row in conn.execute(
            "SELECT DISTINCT employee_id FROM order_intake_cases ORDER BY employee_id"
        ).fetchall()]
    for owner in owners:
        reclassify_cases(owner, changed_terms=changed_terms)


def bootstrap_cases(employee_id: str, account_id: int | None = None) -> None:
    """Create cases only for mail IDs added after each mailbox's saved cursor."""
    now = utcnow()
    created_count = 0
    with db_cursor() as conn:
        accounts_sql = "SELECT id FROM mail_accounts WHERE owner_employee_id=?"
        account_params: list[Any] = [employee_id]
        if account_id is not None:
            accounts_sql += " AND id=?"
            account_params.append(account_id)
        accounts = conn.execute(accounts_sql, account_params).fetchall()
        for account in accounts:
            current_account_id = int(account["id"])
            cursor_key = f"intake_processed_max_mail_id:{employee_id}:{current_account_id}"
            last_mail_id = int(_metadata_get(conn, cursor_key) or "0")
            rows = conn.execute(
                """SELECT m.id mail_id,COALESCE(t.customer_code,'') customer_code,COALESCE(t.customer_name,'') customer_name,
                          COALESCE(t.order_number,'') order_number
                   FROM mail_messages m
                   LEFT JOIN mail_order_tasks t ON t.id=(SELECT id FROM mail_order_tasks WHERE mail_id=m.id ORDER BY id LIMIT 1)
                   WHERE m.account_id=? AND m.id>? ORDER BY m.id""",
                (current_account_id, last_mail_id),
            ).fetchall()
            max_mail_id = last_mail_id
            for row in rows:
                max_mail_id = max(max_mail_id, int(row["mail_id"]))
                inserted = conn.execute(
                    "INSERT OR IGNORE INTO order_intake_cases (employee_id,mail_id,action_type,customer_code,customer_name,order_number,routing_state,routing_source,routing_reason,created_at,updated_at) VALUES (?,?,?,?,?,?, 'unrouted','keyword_rule','未命中通用规则',?,?)",
                    (employee_id,row["mail_id"],"unclassified",row["customer_code"],row["customer_name"],row["order_number"],now,now),
                )
                created_count += int(inserted.rowcount > 0)
            if max_mail_id != last_mail_id:
                _metadata_set(conn, cursor_key, str(max_mail_id))
        revision_key = f"order_intake_rule_engine_v7_customer_archive:{employee_id}"
        revision_needed = _metadata_get(conn, revision_key) is None
    if created_count or revision_needed:
        reclassify_cases(employee_id, account_id)
        with db_cursor() as conn:
            _metadata_set(conn, revision_key, now)


def list_cases(
    employee_id: str,
    target_date: str | None = None,
    action_type: str = "all",
    account_id: int | None = None,
    fetch_task_id: int | None = None,
    prepare: bool = True,
) -> list[dict[str, Any]]:
    if prepare:
        bootstrap_cases(employee_id, account_id)
    clauses = ["c.employee_id = ?"]
    values: list[Any] = [employee_id]
    if account_id:
        clauses.append("m.account_id = ?")
        values.append(account_id)
    if fetch_task_id:
        clauses.append(
            "EXISTS (SELECT 1 FROM mail_fetch_task_messages x "
            "WHERE x.fetch_task_id = ? AND x.mail_id = m.id AND x.is_new = 1)"
        )
        values.append(fetch_task_id)
    if action_type == "needs_business_routing":
        clauses.append("c.routing_state = 'needs_business_routing'")
    elif action_type == "unclassified":
        clauses.append("c.routing_state = 'unrouted'")
    elif action_type in ACTION_LABELS:
        clauses.append("c.action_type = ?")
        values.append(action_type)
    if target_date and not fetch_task_id:
        clauses.append("substr(COALESCE(NULLIF(m.sent_at, ''), NULLIF(m.received_at, ''), m.created_at), 1, 10) = ?")
        values.append(target_date)
    with db_cursor() as conn:
        rows = conn.execute(
            f"""
            SELECT c.*, m.subject, m.sender, m.sent_at, m.received_at, m.body_text, a.email AS mailbox_email,
                   rr.name AS routing_rule_name,
                   (SELECT COUNT(*) FROM mail_attachments a WHERE a.mail_id = m.id AND a.is_inline = 0) AS attachment_count,
                   (SELECT a.filename FROM mail_attachments a WHERE a.mail_id = m.id AND a.is_inline = 0 ORDER BY a.id LIMIT 1) AS first_attachment_name,
                   (SELECT a.content_type FROM mail_attachments a WHERE a.mail_id = m.id AND a.is_inline = 0 ORDER BY a.id LIMIT 1) AS first_attachment_type,
                   (SELECT COUNT(*) FROM mail_order_tasks t WHERE t.mail_id = m.id) AS line_count
            FROM order_intake_cases c
            JOIN mail_messages m ON m.id = c.mail_id
            JOIN mail_accounts a ON a.id = m.account_id
            LEFT JOIN order_mail_routing_rules rr ON rr.id = c.routing_rule_id
            WHERE {' AND '.join(clauses)}
            ORDER BY CASE c.routing_state WHEN 'needs_business_routing' THEN 0 WHEN 'unrouted' THEN 4 ELSE 1 END,
                     COALESCE(m.sent_at, m.created_at) DESC, c.id DESC
            """,
            values,
        ).fetchall()
    today = business_today().isoformat()
    yesterday = (business_today() - timedelta(days=1)).isoformat()
    result = [dict(row) for row in rows]
    for item in result:
        item["received_day"] = str(item.get("sent_at") or item.get("received_at") or "")[:10]
        item["date_label"] = "今天" if item["received_day"] == today else "昨天" if item["received_day"] == yesterday else item["received_day"] or "日期待确认"
        item["routing_matches"] = json.loads(item.get("routing_matches_json") or "[]")
        item["change_tags"] = json.loads(item.get("change_tags_json") or "[]")
        item["sender_display"], item["sender_email"] = _list_sender(item.get("sender") or "")
        item["summary"] = _list_summary(item.get("body_text") or "", item.get("routing_reason") or "")
    return result


def list_date_counts(employee_id: str, account_id: int, days: int = 30, *, prepare: bool = True) -> list[dict[str, Any]]:
    if prepare:
        bootstrap_cases(employee_id, account_id)
    start_date = (business_today() - timedelta(days=max(1, days) - 1)).isoformat()
    with db_cursor() as conn:
        rows = conn.execute(
            """
            SELECT substr(COALESCE(NULLIF(m.sent_at, ''), NULLIF(m.received_at, ''), m.created_at), 1, 10) AS mail_date,
                   COUNT(*) AS total,
                   SUM(CASE WHEN c.action_type = 'new_order' THEN 1 ELSE 0 END) AS new_order_count,
                   SUM(CASE WHEN c.action_type = 'order_change' THEN 1 ELSE 0 END) AS change_count,
                   SUM(CASE WHEN c.action_type = 'quotation' THEN 1 ELSE 0 END) AS quotation_count
            FROM order_intake_cases c
            JOIN mail_messages m ON m.id = c.mail_id
            WHERE c.employee_id = ? AND m.account_id = ?
              AND substr(COALESCE(NULLIF(m.sent_at, ''), NULLIF(m.received_at, ''), m.created_at), 1, 10) >= ?
            GROUP BY mail_date ORDER BY mail_date DESC
            """,
            (employee_id, account_id, start_date),
        ).fetchall()
    return [dict(row) for row in rows]


def get_case(case_id: int, employee_id: str) -> dict[str, Any] | None:
    with db_cursor() as conn:
        row = conn.execute(
            """
            SELECT c.*, m.account_id, m.subject, m.sender, m.sent_at, m.received_at, m.body_html, m.body_text, m.eml_path,
                   rr.name AS routing_rule_name
            FROM order_intake_cases c
            JOIN mail_messages m ON m.id = c.mail_id
            LEFT JOIN order_mail_routing_rules rr ON rr.id = c.routing_rule_id
            WHERE c.id = ? AND c.employee_id = ?
            """,
            (case_id, employee_id),
        ).fetchone()
        if not row:
            return None
        data = dict(row)
        data["attachments"] = [dict(item) for item in conn.execute(
            "SELECT * FROM mail_attachments WHERE mail_id = ? ORDER BY is_inline, id", (data["mail_id"],)
        ).fetchall()]
        for attachment in data["attachments"]:
            attachment["previewable"] = Path(str(attachment.get("filename") or "")).suffix.lower() in {
                ".pdf", ".png", ".jpg", ".jpeg", ".gif", ".webp",
            }
        data["lines"] = [dict(item) for item in conn.execute(
            "SELECT * FROM mail_order_tasks WHERE mail_id = ? ORDER BY id", (data["mail_id"],)
        ).fetchall()]
        data["events"] = [dict(item) for item in conn.execute(
            "SELECT * FROM order_intake_case_events WHERE case_id = ? ORDER BY id DESC", (case_id,)
        ).fetchall()]
        data["routing_matches"] = json.loads(data.get("routing_matches_json") or "[]")
        data["change_tags"] = json.loads(data.get("change_tags_json") or "[]")
        from .mail_transcode_agent.mail_html_parser import extract_order_fields, safe_display_html
        data["display_html"] = safe_display_html(str(data.get("body_html") or ""), str(data.get("body_text") or ""))
        data["detected_fields"] = extract_order_fields(str(data.get("body_text") or ""), str(data.get("sender") or ""))
    return data


def get_attachment(attachment_id: int, employee_id: str) -> dict[str, Any] | None:
    with db_cursor() as conn:
        row = conn.execute(
            """
            SELECT a.* FROM mail_attachments a
            JOIN order_intake_cases c ON c.mail_id = a.mail_id
            WHERE a.id = ? AND c.employee_id = ?
            """,
            (attachment_id, employee_id),
        ).fetchone()
    return dict(row) if row else None


def update_case(case_id: int, employee_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    current = get_case(case_id, employee_id)
    if not current:
        raise ValueError("订单接入任务不存在或无权操作")
    action_type = str(payload.get("action_type") or current["action_type"]).strip()
    status = str(payload.get("status") or current["status"]).strip()
    if action_type not in ACTION_LABELS or status not in STATUS_LABELS:
        raise ValueError("订单动作或处理状态无效")
    workflow_stage = str(payload.get("workflow_stage") or current["workflow_stage"]).strip()
    customer_match_status = str(payload.get("customer_match_status") or current["customer_match_status"]).strip()
    source_document_status = str(payload.get("source_document_status") or current["source_document_status"]).strip()
    mapping_status = str(payload.get("mapping_status") or current["mapping_status"]).strip()
    erp_prepare_status = str(payload.get("erp_prepare_status") or current["erp_prepare_status"]).strip()
    if workflow_stage not in WORKFLOW_STAGE_LABELS:
        raise ValueError("录单流程阶段无效")
    if customer_match_status not in MATCH_STATUS_LABELS or source_document_status not in DOCUMENT_STATUS_LABELS:
        raise ValueError("客户归属或订单资料状态无效")
    if mapping_status not in MAPPING_STATUS_LABELS or erp_prepare_status not in ERP_PREPARE_STATUS_LABELS:
        raise ValueError("映射或录单准备状态无效")
    after = {
        "action_type": action_type,
        "status": status,
        "customer_code": str(payload.get("customer_code") or "").strip(),
        "customer_name": str(payload.get("customer_name") or "").strip(),
        "order_number": str(payload.get("order_number") or "").strip(),
        "order_version": str(payload.get("order_version") or "").strip(),
        "parent_order_number": str(payload.get("parent_order_number") or "").strip(),
        "workflow_stage": workflow_stage,
        "customer_match_status": customer_match_status,
        "source_document_status": source_document_status,
        "mapping_status": mapping_status,
        "erp_prepare_status": erp_prepare_status,
        "handling_note": str(payload.get("handling_note") or "").strip(),
    }
    if action_type == "new_order" and status == "ready_for_erp" and not (after["customer_name"] and after["order_number"]):
        raise ValueError("确认进入 ERP 前，请补齐客户和客户 PO 号")
    now = utcnow()
    with db_cursor() as conn:
        conn.execute(
            """
            UPDATE order_intake_cases SET action_type=?, status=?, customer_code=?, customer_name=?,
                order_number=?, order_version=?, parent_order_number=?, workflow_stage=?, customer_match_status=?,
                source_document_status=?, mapping_status=?, erp_prepare_status=?, handling_note=?,
                routing_source='manual', routing_state='business_routed', routing_reason='业务人员手工调整', routing_rule_id=NULL, routed_by=?, routed_at=?,
                confirmed_by=?, confirmed_at=?, completed_at=?, updated_at=?
            WHERE id=? AND employee_id=?
            """,
            (*after.values(), employee_id, now, employee_id, now,
             now if status == "archived" and current["status"] != "archived" else (current.get("completed_at") if status == "archived" else None),
             now, case_id, employee_id),
        )
        conn.execute(
            """
            INSERT INTO order_intake_case_events (case_id, employee_id, action, before_json, after_json, created_at)
            VALUES (?, ?, 'manual_confirmation', ?, ?, ?)
            """,
            (case_id, employee_id, json.dumps({key: current[key] for key in after}, ensure_ascii=False), json.dumps(after, ensure_ascii=False), now),
        )
    return get_case(case_id, employee_id) or {}


def update_routing(case_id: int, employee_id: str, action_type: str) -> dict[str, Any]:
    current = get_case(case_id, employee_id)
    if not current:
        raise ValueError("邮件不存在或无权操作")
    if action_type not in ACTION_LABELS:
        raise ValueError("分流类型无效")
    now = utcnow()
    with db_cursor() as conn:
        conn.execute(
            """
            UPDATE order_intake_cases
            SET action_type = ?, routing_source = 'manual', routing_state = 'business_routed', routing_reason = '业务人员手工调整', routing_rule_id = NULL,
                routed_by = ?, routed_at = ?, updated_at = ?
            WHERE id = ? AND employee_id = ?
            """,
            (action_type, employee_id, now, now, case_id, employee_id),
        )
        conn.execute(
            """
            INSERT INTO order_intake_case_events
                (case_id, employee_id, action, before_json, after_json, created_at)
            VALUES (?, ?, 'routing_adjusted', ?, ?, ?)
            """,
            (
                case_id,
                employee_id,
                json.dumps({"action_type": current["action_type"]}, ensure_ascii=False),
                json.dumps({"action_type": action_type}, ensure_ascii=False),
                now,
            ),
        )
    return get_case(case_id, employee_id) or {}


def case_summary(cases: list[dict[str, Any]]) -> dict[str, int]:
    return {
        "total": len(cases),
        **{key: sum(1 for item in cases if item["action_type"] == key) for key in ACTION_LABELS},
        "needs_business_routing": sum(1 for item in cases if item.get("routing_state") == "needs_business_routing"),
        "unrouted": sum(1 for item in cases if item.get("routing_state") == "unrouted"),
    }


def work_summary(
    employee_id: str,
    account_id: int | None = None,
    *,
    target_date: str | None = None,
) -> dict[str, Any]:
    """Return workload for all cases whose source mail belongs to ``target_date``.

    When no date is supplied this remains the cross-date summary used by older
    callers.  A supplied date deliberately scopes both open and completed work
    to the source-mail date, so a user can close the loop for one mail day.
    """
    clauses = ["c.employee_id=?"]
    params: list[Any] = [employee_id]
    if account_id:
        clauses.append("m.account_id=?")
        params.append(account_id)
    if target_date:
        clauses.append(
            "substr(COALESCE(NULLIF(m.sent_at, ''), NULLIF(m.received_at, ''), m.created_at), 1, 10)=?"
        )
        params.append(target_date)
    with db_cursor() as conn:
        rows = conn.execute(
            f"""SELECT c.action_type,c.status,c.routing_state,c.completed_at
                FROM order_intake_cases c JOIN mail_messages m ON m.id=c.mail_id
                WHERE {' AND '.join(clauses)}""",
            params,
        ).fetchall()
    result: dict[str, Any] = {
        "needs_routing": 0, "pending": 0, "in_progress": 0,
        "awaiting_confirmation": 0, "on_hold": 0, "active_total": 0,
        "completed_today": 0, "by_type": {key: 0 for key in ROUTABLE_ACTION_TYPES},
    }
    today = business_today().isoformat()
    for raw in rows:
        item = dict(raw)
        if item["status"] == "archived":
            if target_date or _business_date(str(item.get("completed_at") or "")) == today:
                result["completed_today"] += 1
            continue
        if item["routing_state"] == "needs_business_routing":
            result["needs_routing"] += 1
            continue
        if item["action_type"] not in ROUTABLE_ACTION_TYPES:
            continue
        result["active_total"] += 1
        result["by_type"][item["action_type"]] += 1
        if item["status"] == "pending_triage":
            result["pending"] += 1
        elif item["status"] == "pending_review":
            result["in_progress"] += 1
        elif item["status"] == "ready_for_erp":
            result["awaiting_confirmation"] += 1
        elif item["status"] == "on_hold":
            result["on_hold"] += 1
    return result
