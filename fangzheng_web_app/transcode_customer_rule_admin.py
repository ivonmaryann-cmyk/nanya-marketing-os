from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

import openpyxl

from .db import db_cursor
from .transcode_customer_identity import (
    CUSTOMER_ALIAS_GROUPS,
    customer_names_match,
    normalize_customer_name,
)
from .transcode_agent_standard import OFFICIAL_GRADE_CODES
from .transcode_semantic_rule_finalizer import validate_atomic_conditions
from .transcode_semantic_service import TARGET_FIELDS


BUSINESS_FIELD_TARGETS = {
    "胶系": ("glue",),
    "基板厚度": ("thickness",),
    "铜箔规格": ("copper",),
    "基板尺寸": ("size",),
    "胶水类别": ("glue_category",),
    "铜箔类型+印字/非印字": ("copper_type", "print_mark"),
    "基板级别": ("grade_intent",),
    "总/芯厚": ("total_core",),
}
BUSINESS_FIELDS = tuple(BUSINESS_FIELD_TARGETS)
TARGET_FIELD_LABELS = {
    "glue": "胶系代码",
    "thickness": "基板厚度",
    "copper": "铜箔规格",
    "size": "基板尺寸",
    "glue_category": "胶水类别",
    "copper_type": "铜箔类型",
    "print_mark": "印字/非印字",
    "grade_intent": "基板级别",
    "total_core": "总/芯厚",
}
CONDITION_FIELDS = (
    "订单备注",
    "订单规格/订单备注",
    "胶系",
    "基板厚度",
    "铜箔规格",
    "铜箔上面oz",
    "铜箔下面oz",
    "客户规格",
    "客户物料编码",
    "客户料品名称",
)
CONDITION_OPERATORS = (
    "contains_any",
    "contains_all",
    "not_contains",
    "equals",
    "not_equals",
    "in",
    "not_in",
    "lt",
    "lte",
    "gt",
    "gte",
    "missing",
    "present",
)
CONDITION_OPERATOR_LABELS = {
    "contains_any": "包含任一",
    "contains_all": "同时包含",
    "not_contains": "不包含",
    "equals": "等于",
    "not_equals": "不等于",
    "in": "属于",
    "not_in": "不属于",
    "lt": "小于",
    "lte": "小于等于",
    "gt": "大于",
    "gte": "大于等于",
    "missing": "为空",
    "present": "有值",
}
LIST_OPERATORS = {"contains_any", "contains_all", "not_contains", "in", "not_in"}
NUMBER_OPERATORS = {"lt", "lte", "gt", "gte"}
AGENT_ASSET_TYPE = "agent_deterministic"
SEMANTIC_ASSET_TYPE = "semantic"
MAPPING_ASSET_TYPE = "customer_mapping"
INFORMATION_ASSET_TYPE = "customer_rule_information"
CUSTOMER_METADATA_ASSET_TYPE = "customer_metadata"
LEGACY_CCL_ASSET_TYPE = "legacy_ccl_rule"
CUSTOMER_ORDER_ASSET_TYPE = "customer_order_rule"
CODE_MIGRATION_ASSET_TYPE = "code_rule_migration"
# 页面按业务用途分层。历史样本、外部资料和待接入项不能与实际参与转码的
# 规则混在默认客户清单中，否则容易被误认为已经生效的客户规则。
RULE_SCOPE_OPTIONS = {
    "active": "生效规则",
    "pending": "待完善",
}
RULE_SCOPE_LABELS = {
    "all": "全部资料",
    "global": "全客户规则",
    "customer": "客户确定规则",
    "pending": "待业务确认",
    "technical": "待技术支持",
    "migration": "后台运行规则（暂不可维护）",
    "reference": "外部资料",
    "history": "历史样本建议",
    **RULE_SCOPE_OPTIONS,
}
RULE_SCOPE_FILTERS = {
    "active": {"global", "customer", "migration"},
    "pending": {"pending", "technical"},
    "reference": {"reference", "history"},
}
AGENT_OVERRIDE_TO_BUSINESS_FIELD = {
    "glue_code": "胶系",
    "thickness_code": "基板厚度",
    "copper_code": "铜箔规格",
    "size_code": "基板尺寸",
    "glue_category_code": "胶水类别",
    "copper_type_code": "铜箔类型+印字/非印字",
    "grade_code": "基板级别",
    "tc_code": "总/芯厚",
}
AGENT_OVERRIDE_TO_TARGET = {
    "glue_code": "glue",
    "thickness_code": "thickness",
    "copper_code": "copper",
    "size_code": "size",
    "glue_category_code": "glue_category",
    "copper_type_code": "copper_type",
    "grade_code": "grade_intent",
    "tc_code": "total_core",
}

TARGET_VALUE_LABELS = {
    ("total_core", "core"): "芯厚（C）",
    ("total_core", "c"): "芯厚（C）",
    ("total_core", "芯厚"): "芯厚（C）",
    ("total_core", "total"): "总厚（T）",
    ("total_core", "t"): "总厚（T）",
    ("total_core", "总厚"): "总厚（T）",
    ("total_core", "core_after_total_to_core_conversion"): "按总芯厚转换表转为芯厚",
    ("size", "height_plus_0.3"): "单边尺寸加大0.3",
}

# 这些分支当前仍在 transcode_agent_engine.py 中运行。页面只负责让业务看见
# 其影响范围和迁移状态，不复制执行逻辑，也不把提示记录当成100分规则。
HARDCODED_CUSTOMER_RULE_NOTICES = (
    ("深圳安比", "胶系", "NY2140使用客户专属胶系代码；规则仍由代码执行"),
    ("深万基隆", "胶系", "NY2140系列使用客户专属胶系代码；规则仍由代码执行"),
    ("珠海益天", "胶系", "NY2140系列使用客户专属胶系代码；规则仍由代码执行"),
    ("广华升鑫", "胶系", "NY-A2使用客户专属胶系代码；规则仍由代码执行"),
    ("深华升鑫", "胶系", "NY-A2使用客户专属胶系代码；规则仍由代码执行"),
    ("湖奥士康/景旺", "胶系", "NY-A1使用客户专属胶系代码；规则仍由代码执行"),
    ("中富", "胶系", "NY2150H使用客户专属胶系代码；规则仍由代码执行"),
    ("方正F7", "胶系", "NY2150在客户下单转换多结果时优先采用已确认胶系代码2B"),
    ("深南", "胶系", "NY-P5Q考试板及90022-4存在客户专属胶系处理"),
    ("宜兴硅谷", "胶系", "NY2150存在客户专属胶系识别处理"),
    ("江苏瀚宇", "基板厚度", "客户规格中的三位数或P写法按mil解释"),
    ("泰兴电路", "基板厚度", "客户双厚度格式按已确认位置取值"),
    ("深南", "基板厚度", "客户双厚度格式按厚度边界选择芯厚或总厚"),
    ("兴森快捷", "总/芯厚", "T/C按芯厚、D/C按总厚处理"),
    ("方正F7", "总/芯厚", "默认芯厚并按0.8mm边界执行芯总厚转换"),
    ("深圳安比", "总/芯厚", "默认总厚并按0.8mm边界执行总芯厚转换"),
    ("健鼎/超颖", "基板厚度", "使用客户专属mil板厚换算表"),
    ("广合/依利安达/台湾敬鹏", "基板厚度", "指定客户英寸厚度按精确换算处理"),
    ("崇达", "基板厚度", "NY3150HC前芯厚后总厚规格存在客户专属取值"),
    ("广东依顿", "基板厚度", "mil厚度及31mil边界存在客户专属换算"),
    ("广东依顿", "铜箔规格", "1/HOZ按客户确认方向生成铜箔代码"),
    ("深万基隆", "铜箔规格", "H/H按J/J生成铜箔代码"),
    ("常熟斗山", "铜箔规格", "12um/12um RTF按T/T生成铜箔代码"),
    ("珠海超毅", "基板尺寸", "F/W毫米格式按客户尺寸规则换算"),
    ("惠州威健", "基板尺寸", "2184*1245使用客户专属尺寸结果"),
    ("特创", "基板尺寸", "客户尺寸按已确认单边尺寸映射处理"),
    ("深南", "铜箔类型+印字/非印字", "客户铜箔组合写法按专属规则识别"),
    ("惠州威健", "基板级别", "NY1600且CTI600使用客户专属基板级别"),
    ("中宝悦嘉", "基板级别", "NY1600使用客户专属基板级别"),
    ("广华升鑫/深华升鑫", "基板级别", "NY-A2使用客户专属基板级别"),
    ("湖奥士康/景旺", "基板级别", "NY-A1、NY-A2使用客户专属基板级别"),
    ("崇达", "基板级别", "NY3150HC使用客户专属基板级别"),
    ("深万基隆", "基板级别", "NY2140系列存在客户专属基板级别处理"),
    ("珠海益天", "基板级别", "NY2140系列存在客户专属基板级别处理"),
    ("世运", "基板级别", "NY3170HC使用客户专属基板级别"),
    ("江苏苏杭", "基板级别", "NY2140通用F1条件明确排除江苏苏杭"),
    ("深南", "基板级别", "NY-P5Q考试板、AE标识存在客户专属等级处理"),
    ("赣州超跃/赣州金顺/赣州逸豪/信丰共赢/信迅捷兴/中宝悦嘉/九江明阳/深圳明阳/益宝悦嘉/深三德盈", "基板级别", "NY2140且板厚不小于0.8mm、铜箔严格小于1oz时使用F1"),
)


class CustomerRuleMaintenanceError(ValueError):
    pass


def ensure_customer_rule_maintenance_tables() -> None:
    with db_cursor() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS transcode_customer_rule_overrides (
                rule_id TEXT PRIMARY KEY,
                rule_json TEXT,
                deleted INTEGER NOT NULL DEFAULT 0,
                updated_by TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS transcode_customer_rule_changes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rule_id TEXT NOT NULL,
                action TEXT NOT NULL,
                employee_id TEXT NOT NULL,
                before_json TEXT,
                after_json TEXT,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_transcode_customer_rule_changes_rule
            ON transcode_customer_rule_changes(rule_id, id DESC);

            CREATE TABLE IF NOT EXISTS transcode_agent_rule_overrides (
                rule_id TEXT PRIMARY KEY,
                rule_json TEXT,
                deleted INTEGER NOT NULL DEFAULT 0,
                updated_by TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            """
        )


def merge_customer_rule_overrides(base_rules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ensure_customer_rule_maintenance_tables()
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT rule_id, rule_json, deleted FROM transcode_customer_rule_overrides"
        ).fetchall()
    overrides = {str(row["rule_id"]): row for row in rows}
    base_ids = {str(rule.get("rule_id") or "") for rule in base_rules}
    merged: list[dict[str, Any]] = []
    for rule in base_rules:
        rule_id = str(rule.get("rule_id") or "")
        override = overrides.get(rule_id)
        if override is None:
            merged.append(rule)
        elif not int(override["deleted"] or 0):
            merged.append(_parse_rule_json(override["rule_json"], rule_id))
    for rule_id, row in overrides.items():
        if rule_id in base_ids or int(row["deleted"] or 0):
            continue
        merged.append(_parse_rule_json(row["rule_json"], rule_id))
    return merged


def merge_agent_rule_overrides(base_rules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Apply page-maintained overrides without modifying versioned Agent Excel assets."""
    ensure_customer_rule_maintenance_tables()
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT rule_id, rule_json, deleted FROM transcode_agent_rule_overrides"
        ).fetchall()
    overrides = {str(row["rule_id"]): row for row in rows}
    base_ids = {str(rule.get("规则ID") or "") for rule in base_rules}
    merged: list[dict[str, Any]] = []
    for rule in base_rules:
        rule_id = str(rule.get("规则ID") or "")
        override = overrides.get(rule_id)
        if override is None:
            merged.append(rule)
        elif not int(override["deleted"] or 0):
            merged.append(_parse_rule_json(override["rule_json"], rule_id))
    for rule_id, row in overrides.items():
        if rule_id in base_ids or int(row["deleted"] or 0):
            continue
        merged.append(_parse_rule_json(row["rule_json"], rule_id))
    return merged


def agent_rules_for_customer_workspace(rules: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return [_agent_rule_as_workspace_rule(rule) for rule in rules if _clean(rule.get("规则ID"))]


def project_customer_rule_assets_for_workspace(
    semantic_rules: Iterable[dict[str, Any]],
    agent_rules: Iterable[dict[str, Any]],
    mapping_tables: Mapping[str, list[dict[str, Any]]],
    *,
    base_workbook_path: Path | str | None = None,
    include_semantic_pending: bool = True,
) -> list[dict[str, Any]]:
    """Build the complete customer-special-rule projection for the route.

    ``base_workbook_path`` is optional. When omitted, the currently active
    ``transcode_rules.xlsx`` is loaded through ``transcode_rules``. A ``base_sheets``
    parameter is deliberately unnecessary, which keeps the route independent from
    pandas/openpyxl sheet representations.
    """
    semantic_assets = [_with_target_business_field(dict(rule)) for rule in semantic_rules]
    # 旧 Agent 表中同一业务规则常同时存在原始 TAR 行与已经整理过的
    # TSR 语义规则。页面只展示后者，避免业务看到两条相互重复的规则。
    semantic_signatures = {_workspace_rule_signature(rule) for rule in semantic_assets}
    agent_assets = [
        asset
        for asset in agent_rules_for_customer_workspace(agent_rules)
        if _workspace_rule_signature(asset) not in semantic_signatures
    ]
    projected = semantic_assets
    projected.extend(agent_assets)
    projected.extend(
        mapping_assets_for_customer_workspace(
            mapping_tables,
            include_semantic_pending=include_semantic_pending,
            include_runtime_legacy=False,
        )
    )
    projected.extend(legacy_rule_assets_for_customer_workspace(base_workbook_path))
    projected.extend(_hardcoded_customer_rule_assets())
    return _deduplicate_workspace_assets(
        _with_target_business_field(rule) for rule in projected
    )


def mapping_assets_for_customer_workspace(
    mapping_tables: Mapping[str, list[dict[str, Any]]],
    *,
    include_semantic_pending: bool = True,
    include_runtime_legacy: bool | None = None,
) -> list[dict[str, Any]]:
    group_meta = {
        "客户尺寸映射": ("基板尺寸", "目标size_code"),
        "客户单边尺寸映射": ("基板尺寸", "厂内单边尺寸"),
        "客户尺寸算法": ("基板尺寸", "算法类型"),
        "客户厚度映射": ("基板厚度", "厚度mm"),
        "客户物料编码口径": ("总/芯厚", "总芯厚口径"),
    }
    rules: list[dict[str, Any]] = []
    for group, (business_field, result_field) in group_meta.items():
        for row in mapping_tables.get(group, []):
            row_id = _clean(row.get("映射ID"))
            if not row_id:
                continue
            source_text = _clean(row.get("规则文本")) or _clean(row.get("客户厚度写法")) or group
            rules.append(
                {
                    "asset_type": MAPPING_ASSET_TYPE,
                    "rule_id": f"MAP::{group}::{row_id}",
                    "customer_code": _clean(row.get("客户代码")),
                    "customer_name": _clean(row.get("客户简称")),
                    "business_field": business_field,
                    "source_text": source_text,
                    "source_column": "客户规格",
                    "conditions": [{"field": "客户规格", "operator": "present", "value": ""}],
                    "target_fields": [result_field],
                    "normalized_values": [_clean(row.get(result_field)) or _clean(row.get("总芯厚口径"))],
                    "priority": 100,
                    "enabled": _clean(row.get("启用")) != "否",
                    "model": "客户专属转换规则",
                    "approval": {"basis": _clean(row.get("备注"))},
                    "mapping_group": group,
                    "mapping_row": dict(row),
                    "review_state": "pending" if _clean(row.get("待确认")) == "是" else "active",
                }
            )
    global_group_meta = {
        "Agent胶系选择规则": (("胶系", "输出胶系代码"),),
        "Agent基础条件规则": (
            ("胶系", "覆盖胶系代码"),
            ("胶水类别", "覆盖胶水类别"),
            ("基板级别", "覆盖基板级别"),
        ),
    }
    for group, result_specs in global_group_meta.items():
        for row in mapping_tables.get(group, []):
            row_id = _clean(row.get("映射ID"))
            if not row_id:
                continue
            for business_field, result_field in result_specs:
                result_value = _clean(row.get(result_field))
                if not result_value:
                    continue
                conditions = []
                for field, value in (
                    ("胶系", row.get("胶系名称") or row.get("条件胶系")),
                    ("客户规格", row.get("条件关键词")),
                ):
                    cleaned = _clean(value)
                    if cleaned:
                        conditions.append({"field": field, "operator": "contains_any", "value": cleaned})
                has_selection_condition = any(
                    _clean(row.get(field))
                    for field in ("条件客户代码", "条件客户简称", "条件关键词")
                )
                missing_selection_condition = (
                    group == "Agent胶系选择规则" and not has_selection_condition
                )
                if not conditions:
                    conditions.append({"field": "客户规格", "operator": "present", "value": ""})
                customer_code = _clean(row.get("条件客户代码"))
                customer_name = _clean(row.get("条件客户简称"))
                is_global = not customer_code and not customer_name
                rules.append(
                    {
                        "asset_type": MAPPING_ASSET_TYPE,
                        "rule_id": f"MAP::{group}::{row_id}::{result_field}",
                        "customer_code": customer_code,
                        "customer_name": customer_name or ("全部客户" if is_global else ""),
                        "global_scope": is_global,
                        "business_field": business_field,
                        "source_text": _clean(row.get("规则文本")) or _clean(row.get("胶系名称")) or group,
                        "source_column": "客户规格/订单备注",
                        "conditions": conditions,
                        "target_fields": [result_field],
                        "normalized_values": [result_value],
                        "priority": _to_int(row.get("优先级"), 100),
                        "enabled": _clean(row.get("启用")) != "否",
                        "model": "全客户特殊规则",
                        "approval": {"basis": _clean(row.get("备注"))},
                        "mapping_group": group,
                        "mapping_row": dict(row),
                        "review_state": (
                            "missing_selection_condition"
                            if missing_selection_condition
                            else ("pending" if _clean(row.get("待确认")) == "是" else "active")
                        ),
                        "review_reason": (
                            "同名胶系存在多个结果，但尚未维护客户、订单备注或用途选择条件。"
                            if missing_selection_condition
                            else ""
                        ),
                    }
                )

    rules.extend(_historical_mapping_assets(mapping_tables.get("客户字段映射", [])))
    rules.extend(_external_reference_assets(mapping_tables.get("外部尺寸表引用", [])))
    rules.extend(_technical_pending_assets(mapping_tables.get("待接入规则", [])))
    rules.extend(_customer_group_metadata(mapping_tables.get("客户规则组", [])))
    if include_semantic_pending:
        rules.extend(_load_pending_semantic_assets())
    if include_runtime_legacy is None:
        include_runtime_legacy = _is_full_active_mapping_bundle(mapping_tables)
    if include_runtime_legacy:
        rules.extend(legacy_rule_assets_for_customer_workspace())
        rules.extend(_hardcoded_customer_rule_assets())
    return _deduplicate_workspace_assets(rules)


def _is_full_active_mapping_bundle(
    mapping_tables: Mapping[str, list[dict[str, Any]]],
) -> bool:
    """Only auto-project runtime legacy assets for the complete active bundle.

    Unit tests and callers may intentionally pass a small synthetic mapping subset. In
    that case silently loading the application's active workbook would make the result
    depend on unrelated local state.
    """
    required_groups = {
        "Agent胶系主表",
        "客户字段映射",
        "客户尺寸映射",
        "待接入规则",
    }
    return required_groups.issubset(mapping_tables)


def legacy_rule_assets_for_customer_workspace(
    workbook_path: Path | str | None = None,
) -> list[dict[str, Any]]:
    """Project legacy runtime CCL assets into the customer maintenance workspace.

    These rows remain read-only migration records. The projection never promotes a
    historical or free-text row to a 100-point rule and does not duplicate execution
    logic from the transcode engine.
    """
    try:
        if workbook_path is None:
            from .transcode_rules import get_transcode_rule_file_path

            workbook_path = get_transcode_rule_file_path()
        path = Path(workbook_path)
        if not path.exists():
            return []
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (OSError, RuntimeError, ValueError, openpyxl.utils.exceptions.InvalidFileException):
        return []

    assets: list[dict[str, Any]] = []
    assets.extend(_legacy_special_requirement_assets(workbook))
    assets.extend(_legacy_customer_order_assets(workbook))
    return _deduplicate_workspace_assets(assets)


def _legacy_special_requirement_assets(workbook: Any) -> list[dict[str, Any]]:
    sheet_name = "特殊需求"
    if sheet_name not in workbook.sheetnames:
        return []
    rows = _worksheet_records(workbook[sheet_name])
    assets: list[dict[str, Any]] = []
    for row_number, row in rows:
        material_value = next(
            (value for key, value in row.items() if key.startswith("物料类别")),
            "",
        )
        material = _clean(material_value)
        if material not in {"2", "2.0", "CCL", "基板"}:
            continue
        text = _clean(row.get("特殊需求"))
        if not text:
            continue
        fields = _infer_legacy_business_fields(text)
        for business_field in fields:
            assets.append(
                _runtime_migration_asset(
                    asset_type=LEGACY_CCL_ASSET_TYPE,
                    rule_id=f"LEGACY::特殊需求::{row_number}::{business_field}",
                    customer_code=_clean_customer_identifier(row.get("客户代码")),
                    customer_name=_clean(row.get("客户简称")),
                    business_field=business_field,
                    source_text=text,
                    source_column="旧特殊需求（CCL有效规则）",
                    conditions=[{"field": "客户规格", "operator": "present", "value": ""}],
                    target_fields=list(BUSINESS_FIELD_TARGETS[business_field]),
                    normalized_values=["由现有转码引擎执行"],
                    model="旧CCL规则待迁移",
                    migration_hint="核对触发条件和标准结果后，迁移为该客户可编辑规则。",
                )
            )
    return assets


def _legacy_customer_order_assets(workbook: Any) -> list[dict[str, Any]]:
    sheet_name = "客户下单与胶系基板转换"
    if sheet_name not in workbook.sheetnames:
        return []
    assets: list[dict[str, Any]] = []
    for row_number, row in _worksheet_records(workbook[sheet_name]):
        customer_code = _clean_customer_identifier(row.get("客户编号") or row.get("客户代码"))
        customer_name = _clean(row.get("客户简称"))
        customer_glue = _clean(row.get("客户胶系"))
        if not (customer_code or customer_name) or not customer_glue:
            continue
        condition = [{"field": "胶系", "operator": "equals", "value": customer_glue}]
        glue_code = _clean(row.get("厂内胶系代码"))
        if glue_code:
            assets.append(
                _runtime_migration_asset(
                    asset_type=CUSTOMER_ORDER_ASSET_TYPE,
                    rule_id=f"LEGACY::客户下单转换::{row_number}::胶系",
                    customer_code=customer_code,
                    customer_name=customer_name,
                    business_field="胶系",
                    source_text=f"客户胶系{customer_glue}转换为厂内胶系{glue_code}",
                    source_column="客户下单与胶系基板转换",
                    conditions=condition,
                    target_fields=["glue"],
                    normalized_values=[glue_code],
                    model="客户下单转换表待迁移",
                    migration_hint="迁移后由客户特殊规则统一维护胶系结果。",
                )
            )
        grade_code = _clean(row.get("基板等级代码"))
        if grade_code:
            assets.append(
                _runtime_migration_asset(
                    asset_type=CUSTOMER_ORDER_ASSET_TYPE,
                    rule_id=f"LEGACY::客户下单转换::{row_number}::基板级别",
                    customer_code=customer_code,
                    customer_name=customer_name,
                    business_field="基板级别",
                    source_text=f"客户胶系{customer_glue}对应基板级别{grade_code}",
                    source_column="客户下单与胶系基板转换",
                    conditions=condition,
                    target_fields=["grade_intent"],
                    normalized_values=[grade_code],
                    model="客户下单转换表待迁移",
                    migration_hint="迁移后由客户特殊规则统一维护基板级别结果。",
                )
            )
    return assets


def _hardcoded_customer_rule_assets() -> list[dict[str, Any]]:
    assets: list[dict[str, Any]] = []
    sequence = 0
    for names, business_field, description in HARDCODED_CUSTOMER_RULE_NOTICES:
        for customer_name in (item.strip() for item in names.split("/") if item.strip()):
            sequence += 1
            assets.append(
                _runtime_migration_asset(
                    asset_type=CODE_MIGRATION_ASSET_TYPE,
                    rule_id=f"CODE::CUSTOMER::{sequence:03d}",
                    customer_code="",
                    customer_name=customer_name,
                    business_field=business_field,
                    source_text=description,
                    source_column="代码客户规则",
                    conditions=[{"field": "客户规格", "operator": "present", "value": ""}],
                    target_fields=list(BUSINESS_FIELD_TARGETS[business_field]),
                    normalized_values=["由现有代码执行"],
                    model="代码规则待迁移",
                    migration_hint="补齐精确条件和结果后迁移至客户特殊规则；迁移前不得删除代码逻辑。",
                )
            )
    return assets


def _runtime_migration_asset(
    *,
    asset_type: str,
    rule_id: str,
    customer_code: str,
    customer_name: str,
    business_field: str,
    source_text: str,
    source_column: str,
    conditions: list[dict[str, Any]],
    target_fields: list[str],
    normalized_values: list[str],
    model: str,
    migration_hint: str,
) -> dict[str, Any]:
    return {
        "asset_type": asset_type,
        "rule_id": rule_id,
        "customer_code": customer_code,
        "customer_name": customer_name or "未命名客户",
        "business_field": business_field,
        "source_text": source_text,
        "source_column": source_column,
        "conditions": conditions,
        "target_fields": target_fields,
        "normalized_values": normalized_values,
        "priority": 0,
        "enabled": True,
        "model": model,
        "approval": {"basis": "当前运行资产投影，不等同于业务确认规则"},
        "review_state": "migration",
        "runtime_active": True,
        "editable": False,
        "migration_hint": migration_hint,
        "status_detail": f"当前仍参与转码运行；{migration_hint}",
    }


def _worksheet_records(sheet: Any) -> list[tuple[int, dict[str, Any]]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = ["".join(_clean(value).split()) for value in next(rows)]
    except StopIteration:
        return []
    records: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(rows, start=2):
        row = {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }
        records.append((row_number, row))
    return records


def _infer_legacy_business_fields(text: str) -> list[str]:
    normalized = _normalize(text)
    keyword_groups = {
        "胶系": ("胶系", "ny"),
        "基板厚度": ("厚度", "板厚", "mil", "mm"),
        "铜箔规格": ("铜箔", "铜厚", "oz", "hoz", "0/0"),
        "基板尺寸": ("尺寸", "*", "×", "x", "长", "宽"),
        "胶水类别": ("胶水", "胶类"),
        "铜箔类型+印字/非印字": ("hte", "rtf", "hvlp", "vlp", "印字", "水印", "覆膜"),
        "基板级别": ("级别", "汽车板", "a级", "等级"),
        "总/芯厚": ("芯厚", "总厚", "含铜", "不含铜", "连铜"),
    }
    return [
        field
        for field in BUSINESS_FIELDS
        if any(_normalize(keyword) in normalized for keyword in keyword_groups[field])
    ]


def _clean_customer_identifier(value: Any) -> str:
    cleaned = _clean(value)
    return cleaned[:-2] if cleaned.endswith(".0") and cleaned[:-2].isdigit() else cleaned


def _deduplicate_workspace_assets(rules: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    seen_ids: set[str] = set()
    ordered: list[dict[str, Any]] = []
    signature_slots: dict[tuple[Any, ...], dict[str, Any]] = {}
    for rule in rules:
        rule_id = _clean(rule.get("rule_id"))
        if rule_id:
            if rule_id in seen_ids:
                continue
            seen_ids.add(rule_id)
        signature = (
            _workspace_rule_signature(rule)
            if rule.get("asset_type") == AGENT_ASSET_TYPE
            else None
        )
        if signature is None:
            ordered.append(rule)
            continue
        current = signature_slots.get(signature)
        if current is None:
            signature_slots[signature] = rule
        else:
            signature_slots[signature] = _better_workspace_duplicate(current, rule)
    return ordered + list(signature_slots.values())


def _better_workspace_duplicate(
    left: dict[str, Any],
    right: dict[str, Any],
) -> dict[str, Any]:
    """Keep the page-maintained or higher-priority duplicate for display."""
    left_editable = bool(left.get("editable"))
    right_editable = bool(right.get("editable"))
    if left_editable != right_editable:
        return left if left_editable else right
    left_priority = int(left.get("priority") or 0)
    right_priority = int(right.get("priority") or 0)
    if left_priority != right_priority:
        return left if left_priority > right_priority else right
    return left if _clean(left.get("rule_id")) <= _clean(right.get("rule_id")) else right


def _with_target_business_field(rule: dict[str, Any]) -> dict[str, Any]:
    """Classify workspace records by the field they actually change.

    Older imports retain the spreadsheet's source column in ``business_field``.
    That is useful evidence, but it is not the parameter that a business user
    needs to maintain. The management UI therefore follows the target field.
    """
    target_field = _first(rule.get("target_fields"))
    for business_field, target_fields in BUSINESS_FIELD_TARGETS.items():
        if target_field in target_fields:
            projected = dict(rule)
            projected["business_field"] = business_field
            return projected
    return rule


def _workspace_rule_signature(rule: Mapping[str, Any]) -> tuple[Any, ...]:
    """Stable business signature used only to suppress duplicate UI records."""
    conditions = []
    for condition in rule.get("conditions") or []:
        value = condition.get("value")
        if isinstance(value, (list, tuple)):
            value_key = tuple(sorted(_normalize(item) for item in value if _clean(item)))
        else:
            value_key = (_normalize(value),) if _clean(value) else ()
        conditions.append(
            (
                _normalize(condition.get("field")),
                _clean(condition.get("operator")),
                value_key,
            )
        )
    return (
        _clean(rule.get("customer_code")),
        _normalize(rule.get("customer_name")),
        _first(rule.get("target_fields")),
        _normalize(_first(rule.get("normalized_values"))),
        tuple(sorted(conditions)),
    )


def _historical_mapping_assets(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    rules: list[dict[str, Any]] = []
    for row in rows:
        row_id = _clean(row.get("映射ID"))
        override_field = _clean(row.get("覆盖字段"))
        business_field = AGENT_OVERRIDE_TO_BUSINESS_FIELD.get(override_field, "")
        target_field = AGENT_OVERRIDE_TO_TARGET.get(override_field, "")
        if not row_id or not business_field or not target_field:
            continue
        conditions: list[dict[str, Any]] = []
        glue = _clean(row.get("条件胶系"))
        keyword = _clean(row.get("条件关键词"))
        if glue:
            conditions.append({"field": "胶系", "operator": "equals", "value": glue})
        if keyword:
            conditions.append({"field": "客户规格", "operator": "contains_any", "value": keyword})
        if not conditions:
            conditions.append({"field": "客户规格", "operator": "present", "value": ""})
        rules.append(
            {
                "asset_type": INFORMATION_ASSET_TYPE,
                "rule_id": f"INFO::客户字段映射::{row_id}",
                "customer_code": _clean(row.get("客户代码")),
                "customer_name": _clean(row.get("客户简称")),
                "business_field": business_field,
                "source_text": _clean(row.get("规则文本")) or "历史正确码样本形成的建议",
                "source_column": "历史正确码样本",
                "conditions": conditions,
                "target_fields": [target_field],
                "normalized_values": [_clean(row.get("覆盖值"))],
                "priority": 0,
                "enabled": _clean(row.get("启用")) == "是",
                "model": "历史样本建议",
                "approval": {"basis": _clean(row.get("备注"))},
                "review_state": "history",
                "historical_suggestion": True,
                "editable": False,
                "mapping_group": "客户字段映射",
                "mapping_row": dict(row),
                "status_detail": "仅供核对和补充业务规则，不作为100分确定规则。",
            }
        )
    return rules


def _external_reference_assets(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    rules: list[dict[str, Any]] = []
    for row in rows:
        row_id = _clean(row.get("映射ID"))
        if not row_id:
            continue
        reference = _clean(row.get("引用文件"))
        sheet = _clean(row.get("引用Sheet"))
        result = reference + (f" / {sheet}" if sheet else "")
        rules.append(
            {
                "asset_type": INFORMATION_ASSET_TYPE,
                "rule_id": f"INFO::外部尺寸表引用::{row_id}",
                "customer_code": _clean(row.get("客户代码")),
                "customer_name": _clean(row.get("客户简称")),
                "business_field": "基板尺寸",
                "source_text": _clean(row.get("规则文本")) or "尺寸需要参考客户外部对照表",
                "source_column": "外部资料",
                "conditions": [{"field": "客户规格", "operator": "present", "value": ""}],
                "target_fields": ["size"],
                "normalized_values": [result or "外部尺寸表"],
                "priority": 0,
                "enabled": _clean(row.get("启用")) == "是",
                "model": "外部资料引用",
                "approval": {"basis": _clean(row.get("备注"))},
                "review_state": "reference",
                "editable": False,
                "mapping_group": "外部尺寸表引用",
                "mapping_row": dict(row),
                "status_detail": "该客户规则依赖已登记的外部尺寸资料。",
            }
        )
    return rules


def _technical_pending_assets(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    rules: list[dict[str, Any]] = []
    for row in rows:
        row_id = _clean(row.get("映射ID"))
        if not row_id:
            continue
        cleanup_marker = " ".join(
            _clean(row.get(key))
            for key in ("备注", "建议处理", "规则来源说明")
        )
        if any(marker in cleanup_marker for marker in ("已清理", "已解决", "已完成")):
            continue
        technical_type = _clean(row.get("技术类型"))
        business_field = "基板厚度" if "厚度" in technical_type else "基板尺寸"
        target_field = "thickness" if business_field == "基板厚度" else "size"
        rules.append(
            {
                "asset_type": INFORMATION_ASSET_TYPE,
                "rule_id": f"INFO::待接入规则::{row_id}",
                "customer_code": _clean(row.get("客户代码")),
                "customer_name": _clean(row.get("客户简称")),
                "business_field": business_field,
                "source_text": _clean(row.get("原始规则")) or technical_type or "待技术支持规则",
                "source_column": "待技术支持清单",
                "conditions": [{"field": "客户规格", "operator": "present", "value": ""}],
                "target_fields": [target_field],
                "normalized_values": [_clean(row.get("建议处理")) or "待技术支持"],
                "priority": 0,
                "enabled": False,
                "model": "待技术支持",
                "approval": {"basis": _clean(row.get("规则来源说明")) or _clean(row.get("备注"))},
                "review_state": "technical",
                "editable": False,
                "mapping_group": "待接入规则",
                "mapping_row": dict(row),
                "status_detail": f"{technical_type or '该规则'}尚未完成技术接入，不参与正式出码。",
            }
        )
    return rules


def _customer_group_metadata(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    metadata: list[dict[str, Any]] = []
    for row in rows:
        if _clean(row.get("启用")) != "是":
            continue
        row_id = _clean(row.get("映射ID"))
        if not row_id:
            continue
        metadata.append(
            {
                "asset_type": CUSTOMER_METADATA_ASSET_TYPE,
                "customer_metadata": True,
                "rule_id": f"META::客户规则组::{row_id}",
                "customer_code": _clean(row.get("客户代码")),
                "customer_name": _clean(row.get("客户简称")),
                "group_id": _clean(row.get("规则组ID")),
                "group_name": _clean(row.get("规则组名称")),
                "main_customer_code": _clean(row.get("主规则客户代码")),
                "main_customer_name": _clean(row.get("主规则客户简称")),
                "group_note": _clean(row.get("备注")),
            }
        )
    return metadata


def _load_pending_semantic_assets() -> list[dict[str, Any]]:
    """Load the active workbook's read-only pending sheet for the maintenance workspace."""
    try:
        from .transcode_semantic_rules import (
            PENDING_SHEET_NAME,
            get_active_transcode_semantic_rule_version,
            get_transcode_semantic_rule_workbook_path,
        )

        version = get_active_transcode_semantic_rule_version()
        workbook_path = get_transcode_semantic_rule_workbook_path(version)
        if not version or not Path(workbook_path).exists():
            return []
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        if PENDING_SHEET_NAME not in workbook.sheetnames:
            return []
        sheet = workbook[PENDING_SHEET_NAME]
        headers = [_clean(cell.value) for cell in sheet[1]]
        assets: list[dict[str, Any]] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
            candidate_id = _clean(row.get("候选规则ID"))
            if not candidate_id:
                continue
            business_field = _pending_business_field(row)
            assets.append(
                {
                    "asset_type": INFORMATION_ASSET_TYPE,
                    "rule_id": f"INFO::语义待业务确认::{candidate_id}",
                    "customer_code": _clean(row.get("客户代码")),
                    "customer_name": _clean(row.get("客户简称")),
                    "business_field": business_field,
                    "source_text": _clean(row.get("规则原文")) or "语义规则待业务确认",
                    "source_column": "订单备注语义待确认",
                    "conditions": [],
                    "target_fields": [],
                    "normalized_values": ["待业务确认"],
                    "priority": 0,
                    "enabled": False,
                    "model": "订单备注语义待确认",
                    "approval": {"basis": _clean(row.get("待确认问题"))},
                    "review_state": "pending",
                    "editable": False,
                    "status_detail": _clean(row.get("待确认问题")) or "需要业务补充选择条件。",
                    "pending_row": row,
                }
            )
        return assets
    except (OSError, ValueError, KeyError, openpyxl.utils.exceptions.InvalidFileException):
        return []


def customer_rule_workspace(
    rules: list[dict[str, Any]],
    *,
    search: str = "",
    customer_key: str = "",
    business_field: str = "",
    rule_id: str = "",
    rule_kind: str = "all",
    rule_scope: str = "all",
    status_filter: str = "all",
) -> dict[str, Any]:
    selected_kind = rule_kind if rule_kind in {"all", "deterministic", "semantic"} else "all"
    # 保留代码接口中的细分范围和旧链接；页面入口默认由路由传 active。
    selected_scope = rule_scope if rule_scope in RULE_SCOPE_LABELS else "active"
    # 旧转换表的补录行存在“????”这类占位客户简称。同一客户代码在其他
    # 已确认资料中已有名称时，配置页统一显示真实名称；不改动运行规则内容。
    customer_names_by_code = _customer_names_by_code(rules)
    customer_metadata = [
        _with_workspace_customer_name(rule, customer_names_by_code)
        for rule in rules
        if bool(rule.get("customer_metadata"))
    ]
    rules = [
        _with_workspace_customer_name(rule, customer_names_by_code)
        for rule in rules
        if not bool(rule.get("customer_metadata"))
    ]
    rules = [rule for rule in rules if _rule_matches_kind(rule, selected_kind)]
    # 细分计数保留给审计和兼容接口；页面分层计数另行提供，避免历史和外部资料
    # 被误当成实际生效规则。
    scope_counts = {
        "all": len(rules),
        "global": 0,
        "customer": 0,
        "pending": 0,
        "technical": 0,
        "migration": 0,
        "reference": 0,
        "history": 0,
    }
    for rule in rules:
        detail_scope = _rule_scope(rule)
        scope_counts[detail_scope] += 1
    display_scope_counts = {
        scope: sum(scope_counts.get(detail, 0) for detail in details)
        for scope, details in RULE_SCOPE_FILTERS.items()
    }
    # 已停用规则保留在修改记录里，不与业务正在使用的规则混在默认清单中。
    display_scope_counts["active"] = sum(
        1
        for rule in rules
        if _rule_scope(rule) in RULE_SCOPE_FILTERS["active"] and bool(rule.get("enabled"))
    )
    if selected_scope == "all":
        scoped_rules = rules
    elif selected_scope == "active":
        scoped_rules = [
            rule
            for rule in rules
            if _rule_scope(rule) in RULE_SCOPE_FILTERS["active"] and bool(rule.get("enabled"))
        ]
    elif selected_scope in RULE_SCOPE_FILTERS:
        scoped_rules = [
            rule for rule in rules
            if _rule_scope(rule) in RULE_SCOPE_FILTERS[selected_scope]
        ]
    else:
        scoped_rules = [rule for rule in rules if _rule_scope(rule) == selected_scope]
    status_options = [("all", "全部状态")]
    status_meta = (
        ("confirmed", "已确认生效"),
        ("migration", "后台运行规则（暂不可维护）"),
        ("pending", "待业务确认"),
        ("technical", "待技术支持"),
        ("disabled", "已停用"),
    )
    available_statuses = {_workspace_status_key(rule) for rule in scoped_rules}
    status_options.extend(
        (key, label) for key, label in status_meta if key in available_statuses
    )
    selected_status = status_filter if status_filter in dict(status_options) else "all"
    if selected_status != "all":
        scoped_rules = [
            rule for rule in scoped_rules
            if _workspace_status_key(rule) == selected_status
        ]
    override_ids = _override_ids()
    canonical_key_map = _canonical_customer_key_map(scoped_rules)
    customers_by_key: dict[str, dict[str, Any]] = {}
    for rule in scoped_rules:
        key = canonical_key_map.get(
            make_customer_key(rule.get("customer_code"), rule.get("customer_name")),
            make_customer_key(rule.get("customer_code"), rule.get("customer_name")),
        )
        customer = customers_by_key.setdefault(
            key,
            {
                "key": key,
                "code": _clean(rule.get("customer_code")),
                "name": _clean(rule.get("customer_name")) or "未命名客户",
                "rule_count": 0,
                "field_counts": {field: 0 for field in BUSINESS_FIELDS},
                "global_scope": _rule_scope(rule) == "global",
            },
        )
        customer["global_scope"] = bool(
            customer["global_scope"]
            or rule.get("global_scope")
            or _clean(rule.get("customer_name")) == "全部客户"
        )
        customer["rule_count"] += 1
        field = _clean(rule.get("business_field"))
        if field in customer["field_counts"]:
            customer["field_counts"][field] += 1

    groups_by_id: dict[str, list[dict[str, Any]]] = {}
    for item in customer_metadata:
        group_id = _clean(item.get("group_id"))
        if group_id:
            groups_by_id.setdefault(group_id, []).append(item)
    metadata_by_customer_key = {
        canonical_key_map.get(
            make_customer_key(item.get("customer_code"), item.get("customer_name")),
            make_customer_key(item.get("customer_code"), item.get("customer_name")),
        ): item
        for item in customer_metadata
    }
    if selected_scope == "all":
        for item in customer_metadata:
            key = canonical_key_map.get(
                make_customer_key(item.get("customer_code"), item.get("customer_name")),
                make_customer_key(item.get("customer_code"), item.get("customer_name")),
            )
            customers_by_key.setdefault(
                key,
                {
                    "key": key,
                    "code": _clean(item.get("customer_code")),
                    "name": _clean(item.get("customer_name")) or "未命名客户",
                    "rule_count": 0,
                    "field_counts": {field: 0 for field in BUSINESS_FIELDS},
                    "global_scope": False,
                },
            )
    for key, customer in customers_by_key.items():
        item = metadata_by_customer_key.get(key)
        customer["rule_group"] = ""
        customer["aliases"] = []
        customer["main_customer"] = ""
        customer["group_note"] = ""
        if not item:
            continue
        group_id = _clean(item.get("group_id"))
        members = groups_by_id.get(group_id, [])
        customer["rule_group"] = _clean(item.get("group_name"))
        customer["main_customer"] = _customer_display(
            item.get("main_customer_code"), item.get("main_customer_name")
        )
        customer["group_note"] = _clean(item.get("group_note"))
        customer["aliases"] = [
            _customer_display(member.get("customer_code"), member.get("customer_name"))
            for member in members
            if canonical_key_map.get(
                make_customer_key(member.get("customer_code"), member.get("customer_name")),
                make_customer_key(member.get("customer_code"), member.get("customer_name")),
            )
            != key
        ]

    query = _normalize(search)
    matching_customer_keys = {
        canonical_key_map.get(
            make_customer_key(rule.get("customer_code"), rule.get("customer_name")),
            make_customer_key(rule.get("customer_code"), rule.get("customer_name")),
        )
        for rule in scoped_rules
        if query and _rule_matches_search(rule, query)
    }
    customers = sorted(
        (
            customer
            for customer in customers_by_key.values()
            if not query
            or query in _normalize(customer["code"])
            or query in _normalize(customer["name"])
            or any(query in _normalize(alias) for alias in customer.get("aliases", []))
            or query in _normalize(customer.get("rule_group"))
            or customer["key"] in matching_customer_keys
        ),
        key=lambda item: (-int(item["rule_count"]), item["name"]),
    )
    visible_customer_keys = {customer["key"] for customer in customers}
    normalized_selected_key = (
        canonical_key_map.get(customer_key, customer_key) if customer_key else ""
    )
    selected_key = normalized_selected_key if normalized_selected_key in visible_customer_keys else ""
    if not selected_key and customers:
        selected_key = customers[0]["key"]
    selected_customer = customers_by_key.get(selected_key)
    selected_field = business_field if business_field in BUSINESS_FIELDS else "基板级别"
    if selected_customer and not selected_customer["field_counts"].get(selected_field):
        selected_field = next(
            (
                field
                for field, count in selected_customer["field_counts"].items()
                if count
            ),
            selected_field,
        )

    selected_rules: list[dict[str, Any]] = []
    if selected_customer:
        selected_rules = [
            _rule_view(rule, overridden=_clean(rule.get("rule_id")) in override_ids)
            for rule in scoped_rules
            if canonical_key_map.get(
                make_customer_key(rule.get("customer_code"), rule.get("customer_name")),
                make_customer_key(rule.get("customer_code"), rule.get("customer_name")),
            )
            == selected_key
            and _clean(rule.get("business_field")) == selected_field
        ]
        selected_rules.sort(key=lambda item: (-int(item["priority"]), item["rule_id"]))
    selected_rule = next(
        (item for item in selected_rules if item["rule_id"] == rule_id),
        selected_rules[0] if selected_rules else None,
    )
    return {
        "customers": customers,
        "selected_customer": selected_customer,
        "selected_customer_key": selected_key,
        "business_fields": BUSINESS_FIELDS,
        "selected_field": selected_field,
        "rules": selected_rules,
        "selected_rule": selected_rule,
        "rule_count": len(scoped_rules),
        "override_count": sum(
            1 for rule in scoped_rules if _clean(rule.get("rule_id")) in override_ids
        ),
        "rule_kind": selected_kind,
        "rule_scope": selected_scope,
        "status_filter": selected_status,
        "status_options": status_options,
        "scope_options": RULE_SCOPE_OPTIONS,
        "scope_counts": scope_counts,
        "display_scope_counts": display_scope_counts,
    }


def _workspace_status_key(rule: Mapping[str, Any]) -> str:
    scope = _rule_scope(rule)
    if scope in {"migration", "pending", "technical"}:
        return scope
    return "confirmed" if bool(rule.get("enabled")) else "disabled"


def _rule_scope(rule: Mapping[str, Any]) -> str:
    review_state = _clean(rule.get("review_state"))
    if review_state in {"pending", "missing_selection_condition"}:
        return "pending"
    if review_state == "technical":
        return "technical"
    if review_state == "reference":
        return "reference"
    if review_state == "history":
        return "history"
    if review_state == "migration":
        return "migration"
    approval = rule.get("approval") or {}
    if _clean(approval.get("status")).lower() in {"pending", "draft", "待确认"}:
        return "pending"
    if bool(rule.get("historical_suggestion")) or "历史建议" in (
        _clean(rule.get("model")) + _clean(rule.get("source_column"))
    ):
        return "history"
    if bool(rule.get("global_scope")) or _clean(rule.get("customer_name")) == "全部客户":
        return "global"
    return "customer"


def _rule_matches_search(rule: Mapping[str, Any], query: str) -> bool:
    values = [
        rule.get("source_text"),
        rule.get("customer_code"),
        rule.get("customer_name"),
        rule.get("business_field"),
        _first(rule.get("normalized_values")),
        _condition_summary(rule.get("conditions") or []),
        rule.get("status_detail"),
        rule.get("review_reason"),
        rule.get("source_column"),
        rule.get("model"),
        rule.get("migration_hint"),
    ]
    return any(query in _normalize(value) for value in values)


def _rule_matches_kind(rule: Mapping[str, Any], rule_kind: str) -> bool:
    if rule_kind == "all":
        return True
    source_fields = {
        _clean(item.get("field"))
        for item in rule.get("conditions") or []
        if isinstance(item, Mapping)
    }
    is_order_semantic = any("订单备注" in field for field in source_fields)
    return is_order_semantic if rule_kind == "semantic" else not is_order_semantic


def _customer_display(code: Any, name: Any) -> str:
    customer_code = _clean(code)
    customer_name = _clean(name)
    if customer_code and customer_name:
        return f"{customer_name}（{customer_code}）"
    return customer_name or customer_code


def _pending_business_field(row: Mapping[str, Any]) -> str:
    """Place unresolved items in a business tab without making them executable."""
    business_field = _clean(row.get("业务字段"))
    if business_field in BUSINESS_FIELDS:
        return business_field
    text = " ".join(
        _clean(row.get(field))
        for field in ("规则原文", "待确认问题")
    )
    if any(keyword in text for keyword in ("总/芯厚", "芯厚", "总厚", "第22码", "151")):
        return "总/芯厚"
    if any(keyword in text for keyword in ("基板级别", "HDI", "A级")):
        return "基板级别"
    return "基板级别"


def build_rule_from_form(
    form: Mapping[str, Any],
    *,
    existing_rule: dict[str, Any] | None = None,
) -> dict[str, Any]:
    rule_id = _clean(form.get("rule_id")) or _new_rule_id()
    customer_code = _clean(form.get("customer_code"))
    customer_name = _clean(form.get("customer_name"))
    business_field = _clean(form.get("business_field"))
    source_text = _clean(form.get("source_text"))
    target_field = _clean(form.get("target_field"))
    target_value = _clean(form.get("target_value"))
    if not customer_code and not customer_name:
        raise CustomerRuleMaintenanceError("客户代码和客户简称至少填写一个。")
    if business_field not in BUSINESS_FIELD_TARGETS:
        raise CustomerRuleMaintenanceError(f"不支持的维护参数：{business_field}")
    if target_field not in BUSINESS_FIELD_TARGETS[business_field]:
        raise CustomerRuleMaintenanceError("目标字段与维护参数不匹配。")
    if target_field not in TARGET_FIELDS or not target_value:
        raise CustomerRuleMaintenanceError("目标结果不能为空。")
    if not source_text:
        raise CustomerRuleMaintenanceError("业务触发条件不能为空。")

    condition_fields = _form_list(form, "condition_field")
    condition_operators = _form_list(form, "condition_operator")
    condition_values = _form_list(form, "condition_value")
    conditions = _build_conditions(condition_fields, condition_operators, condition_values)
    validate_atomic_conditions(conditions, context=f"规则{rule_id}")
    _validate_target_value(target_field, target_value)
    priority = _to_int(form.get("priority"), 100)
    enabled = _form_bool(form, "enabled", default=True)
    semantic_enabled = _form_bool(form, "semantic_enabled", default=False)
    if semantic_enabled and not any(item["field"] == "订单备注" for item in conditions):
        raise CustomerRuleMaintenanceError("启用模型语义标准化时，至少需要一条订单备注条件。")

    prior = existing_rule or {}
    semantic_type = _semantic_type(conditions)
    return {
        "rule_id": rule_id,
        "source_candidate_id": _clean(prior.get("source_candidate_id")) or f"PAGE-{rule_id}",
        "enabled": enabled,
        "customer_code": customer_code,
        "customer_name": customer_name,
        "source_row": int(prior.get("source_row") or 0),
        "source_column": "页面维护",
        "business_field": business_field,
        "source_text": source_text,
        "semantic_types": [semantic_type],
        "target_fields": [target_field],
        "normalized_values": [target_value],
        "stated_target_values": [target_value],
        "conditions": conditions,
        "required_input_fields": list(dict.fromkeys(item["field"] for item in conditions)),
        "execution_mode": "结构化后可确定性执行",
        "priority": priority,
        "model": _clean(prior.get("model")) or ("页面语义维护" if semantic_enabled else "页面确定性维护"),
        "prompt_sha256": _clean(prior.get("prompt_sha256")),
        "evidence_texts": [source_text],
        "approval": {
            "status": "confirmed",
            "basis": _clean(form.get("approval_basis")) or "客户特殊规则维护页面确认",
        },
        "note": "页面维护；保存后立即生效"
        + ("；允许订单备注模型语义标准化" if semantic_enabled else ""),
    }


def build_agent_rule_from_form(
    form: Mapping[str, Any],
    *,
    existing_rule: dict[str, Any] | None,
) -> dict[str, Any]:
    rule = dict(existing_rule or {})
    rule_id = _clean(rule.get("规则ID")) or _clean(form.get("rule_id"))
    if not rule_id:
        rule_id = f"PAGE-CUST-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    customer_code = _clean(form.get("customer_code"))
    customer_name = _clean(form.get("customer_name"))
    business_field = _clean(form.get("business_field"))
    override_field = _clean(form.get("agent_override_field"))
    override_value = _clean(form.get("target_value")).upper()
    source_text = _clean(form.get("source_text")) or _clean(rule.get("规则文本"))
    if not customer_code and not customer_name:
        raise CustomerRuleMaintenanceError("客户代码和客户简称至少填写一个。")
    if business_field not in BUSINESS_FIELDS:
        raise CustomerRuleMaintenanceError(f"不支持的维护参数：{business_field}")
    if override_field not in AGENT_OVERRIDE_TO_BUSINESS_FIELD:
        raise CustomerRuleMaintenanceError(f"Agent覆盖字段无效：{override_field}")
    if AGENT_OVERRIDE_TO_BUSINESS_FIELD[override_field] != business_field:
        raise CustomerRuleMaintenanceError("Agent覆盖字段与维护参数不匹配。")
    if not override_value:
        raise CustomerRuleMaintenanceError("目标结果不能为空。")
    _validate_agent_override_value(override_field, override_value)
    rule.update(
        {
            "_asset_type": AGENT_ASSET_TYPE,
            "_page_override": True,
            "规则ID": rule_id,
            "启用": "是" if _form_bool(form, "enabled", default=True) else "否",
            "客户代码": customer_code,
            "客户简称": customer_name,
            "原始字段": business_field,
            "规则文本": source_text,
            "条件胶系": _clean(form.get("agent_condition_glue")),
            "条件关键词": _clean(form.get("agent_condition_keyword")),
            "条件铜厚": _clean(form.get("agent_condition_copper")),
            "条件厚度": _clean(form.get("agent_condition_thickness")),
            "条件尺寸": _clean(form.get("agent_condition_size")),
            "覆盖字段": override_field,
            "覆盖值": override_value,
            "优先级": str(_to_int(form.get("priority"), _to_int(rule.get("优先级"), 100))),
            "强制执行": "是",
            "待确认": "否",
            "物料类别": "CCL",
            "来源字段": _clean(rule.get("来源字段")) or "客户规格",
            "条件文本": source_text or _clean(rule.get("条件文本")),
            "来源行号": _clean(rule.get("来源行号")) or "页面新增",
            "命中来源": "页面维护Agent确定性长期规则",
            "规则解释": _clean(form.get("approval_basis")) or "客户特殊规则维护页面确认",
            "跳过原因": "",
        }
    )
    if not rule["规则文本"]:
        rule["规则文本"] = _auto_agent_rule_text(rule)
        rule["条件文本"] = rule["规则文本"]
    return rule


def _auto_agent_rule_text(rule: dict[str, Any]) -> str:
    parts: list[str] = []
    customer_name = _clean(rule.get("客户简称"))
    customer_code = _clean(rule.get("客户代码"))
    if customer_name:
        parts.append(f"客户{customer_name}")
    elif customer_code:
        parts.append(f"客户代码{customer_code}")
    conditions: list[str] = []
    for key, label in (
        ("条件胶系", "胶系"),
        ("条件关键词", "关键词"),
        ("条件铜厚", "铜厚"),
        ("条件厚度", "厚度"),
        ("条件尺寸", "尺寸"),
    ):
        value = _clean(rule.get(key))
        if value:
            conditions.append(f"{label}={value}")
    if conditions:
        parts.append("；".join(conditions))
    field = _clean(rule.get("覆盖字段"))
    field_label = AGENT_OVERRIDE_TO_BUSINESS_FIELD.get(field, field)
    target = _clean(rule.get("覆盖值"))
    if target:
        parts.append(f"{field_label}={target}")
    return "；".join(parts)


def save_rule_override(
    rule: dict[str, Any],
    *,
    updated_by: str,
    previous_rule: dict[str, Any] | None,
) -> None:
    ensure_customer_rule_maintenance_tables()
    rule_id = _clean(rule.get("rule_id"))
    if not rule_id:
        raise CustomerRuleMaintenanceError("规则ID不能为空。")
    now = _now()
    rule_json = _dump(rule)
    with db_cursor() as conn:
        conn.execute(
            """
            INSERT INTO transcode_customer_rule_overrides
                (rule_id, rule_json, deleted, updated_by, updated_at)
            VALUES (?, ?, 0, ?, ?)
            ON CONFLICT(rule_id) DO UPDATE SET
                rule_json = excluded.rule_json,
                deleted = 0,
                updated_by = excluded.updated_by,
                updated_at = excluded.updated_at
            """,
            (rule_id, rule_json, updated_by, now),
        )
        _insert_rule_change(
            conn,
            rule_id=rule_id,
            action="新增" if previous_rule is None else "修改",
            employee_id=updated_by,
            before=previous_rule,
            after=rule,
            created_at=now,
        )


def save_agent_rule_override(
    rule: dict[str, Any],
    *,
    updated_by: str,
    previous_rule: dict[str, Any] | None,
) -> None:
    ensure_customer_rule_maintenance_tables()
    rule_id = _clean(rule.get("规则ID"))
    if not rule_id:
        raise CustomerRuleMaintenanceError("Agent规则ID不能为空。")
    now = _now()
    rule_json = _dump(rule)
    with db_cursor() as conn:
        conn.execute(
            """
            INSERT INTO transcode_agent_rule_overrides
                (rule_id, rule_json, deleted, updated_by, updated_at)
            VALUES (?, ?, 0, ?, ?)
            ON CONFLICT(rule_id) DO UPDATE SET
                rule_json = excluded.rule_json,
                deleted = 0,
                updated_by = excluded.updated_by,
                updated_at = excluded.updated_at
            """,
            (rule_id, rule_json, updated_by, now),
        )
        _insert_rule_change(
            conn,
            rule_id=rule_id,
            action="新增" if previous_rule is None else "修改",
            employee_id=updated_by,
            before=previous_rule,
            after=rule,
            created_at=now,
        )


def validate_customer_maintained_rule(rule: dict[str, Any]) -> None:
    from .transcode_semantic_rules import _validate_machine_rule

    _validate_machine_rule(rule, 0)


def delete_rule_override(
    rule: dict[str, Any],
    *,
    updated_by: str,
) -> None:
    ensure_customer_rule_maintenance_tables()
    rule_id = _clean(rule.get("rule_id"))
    if not rule_id:
        raise CustomerRuleMaintenanceError("规则ID不能为空。")
    now = _now()
    with db_cursor() as conn:
        conn.execute(
            """
            INSERT INTO transcode_customer_rule_overrides
                (rule_id, rule_json, deleted, updated_by, updated_at)
            VALUES (?, ?, 1, ?, ?)
            ON CONFLICT(rule_id) DO UPDATE SET
                rule_json = excluded.rule_json,
                deleted = 1,
                updated_by = excluded.updated_by,
                updated_at = excluded.updated_at
            """,
            (rule_id, _dump(rule), updated_by, now),
        )
        _insert_rule_change(
            conn,
            rule_id=rule_id,
            action="删除",
            employee_id=updated_by,
            before=rule,
            after=None,
            created_at=now,
        )


def delete_agent_rule_override(rule: dict[str, Any], *, updated_by: str) -> None:
    ensure_customer_rule_maintenance_tables()
    rule_id = _clean(rule.get("规则ID"))
    if not rule_id:
        raise CustomerRuleMaintenanceError("Agent规则ID不能为空。")
    now = _now()
    payload = dict(rule)
    payload["_asset_type"] = AGENT_ASSET_TYPE
    with db_cursor() as conn:
        conn.execute(
            """
            INSERT INTO transcode_agent_rule_overrides
                (rule_id, rule_json, deleted, updated_by, updated_at)
            VALUES (?, ?, 1, ?, ?)
            ON CONFLICT(rule_id) DO UPDATE SET
                rule_json = excluded.rule_json,
                deleted = 1,
                updated_by = excluded.updated_by,
                updated_at = excluded.updated_at
            """,
            (rule_id, _dump(payload), updated_by, now),
        )
        _insert_rule_change(
            conn,
            rule_id=rule_id,
            action="删除",
            employee_id=updated_by,
            before=payload,
            after=None,
            created_at=now,
        )


def list_customer_rule_changes(limit: int = 30) -> list[dict[str, Any]]:
    ensure_customer_rule_maintenance_tables()
    with db_cursor() as conn:
        rows = conn.execute(
            """
            SELECT id, rule_id, action, employee_id, before_json, after_json, created_at
            FROM transcode_customer_rule_changes
            ORDER BY id DESC
            LIMIT ?
            """,
            (max(1, min(int(limit), 100)),),
        ).fetchall()
    result = []
    for row in rows:
        payload = _optional_rule_json(row["after_json"]) or _optional_rule_json(row["before_json"]) or {}
        result.append(
            {
                "id": int(row["id"]),
                "rule_id": row["rule_id"],
                "action": row["action"],
                "employee_id": row["employee_id"],
                "created_at": row["created_at"],
                "customer": payload.get("customer_name") or payload.get("customer_code")
                or payload.get("客户简称") or payload.get("客户代码") or "",
                "business_field": payload.get("business_field") or payload.get("原始字段")
                or AGENT_OVERRIDE_TO_BUSINESS_FIELD.get(_clean(payload.get("覆盖字段")), ""),
                "source_text": payload.get("source_text") or payload.get("规则文本") or "",
            }
        )
    return result


def restore_customer_rule_change(change_id: int, *, updated_by: str) -> str:
    ensure_customer_rule_maintenance_tables()
    with db_cursor() as conn:
        row = conn.execute(
            """
            SELECT id, rule_id, before_json, after_json
            FROM transcode_customer_rule_changes
            WHERE id = ?
            """,
            (int(change_id),),
        ).fetchone()
        if row is None:
            raise CustomerRuleMaintenanceError("修改记录不存在。")
        rule_id = _clean(row["rule_id"])
        before = _optional_rule_json(row["before_json"])
        after = _optional_rule_json(row["after_json"])
        asset_payload = before or after or {}
        asset_type = _clean(asset_payload.get("_asset_type") or asset_payload.get("asset_type"))
        override_table = (
            "transcode_agent_rule_overrides"
            if asset_type == AGENT_ASSET_TYPE
            else "transcode_customer_rule_overrides"
        )
        current_row = conn.execute(
            f"SELECT rule_json, deleted FROM {override_table} WHERE rule_id = ?",
            (rule_id,),
        ).fetchone()
        current = None
        if current_row is not None and not int(current_row["deleted"] or 0):
            current = _optional_rule_json(current_row["rule_json"])
        now = _now()
        if before is None:
            conn.execute(
                f"""
                INSERT INTO {override_table}
                    (rule_id, rule_json, deleted, updated_by, updated_at)
                VALUES (?, NULL, 1, ?, ?)
                ON CONFLICT(rule_id) DO UPDATE SET
                    rule_json = NULL,
                    deleted = 1,
                    updated_by = excluded.updated_by,
                    updated_at = excluded.updated_at
                """,
                (rule_id, updated_by, now),
            )
        else:
            conn.execute(
                f"""
                INSERT INTO {override_table}
                    (rule_id, rule_json, deleted, updated_by, updated_at)
                VALUES (?, ?, 0, ?, ?)
                ON CONFLICT(rule_id) DO UPDATE SET
                    rule_json = excluded.rule_json,
                    deleted = 0,
                    updated_by = excluded.updated_by,
                    updated_at = excluded.updated_at
                """,
                (rule_id, _dump(before), updated_by, now),
            )
        _insert_rule_change(
            conn,
            rule_id=rule_id,
            action="恢复",
            employee_id=updated_by,
            before=current,
            after=before,
            created_at=now,
        )
    return rule_id


def find_rule(rules: Iterable[dict[str, Any]], rule_id: str) -> dict[str, Any] | None:
    target = _clean(rule_id)
    return next((rule for rule in rules if _clean(rule.get("rule_id")) == target), None)


def make_customer_key(code: Any, name: Any) -> str:
    customer_code = _clean(code)
    # 客户代码是唯一的稳定维度。旧资料的简称可能缺失或写成“????”，不能
    # 因显示名称不同而拆成多个客户规则组。
    if customer_code:
        return f"code:{customer_code}"
    return f"name:{_clean(name)}"


def _customer_code_tokens(value: Any) -> set[str]:
    return {
        token
        for token in re.findall(r"\d+", str(value or ""))
        if token
    }


def _canonical_customer_key_map(
    rules: Iterable[Mapping[str, Any]],
) -> dict[str, str]:
    """Merge customer groups that share a code or an exact/alias name."""
    entries = [
        {
            "key": make_customer_key(rule.get("customer_code"), rule.get("customer_name")),
            "codes": _customer_code_tokens(rule.get("customer_code")),
            "name": _clean(rule.get("customer_name")),
        }
        for rule in rules
        if rule.get("customer_code") or rule.get("customer_name")
    ]
    parent = {entry["key"]: entry["key"] for entry in entries}

    def find(key: str) -> str:
        while parent[key] != key:
            parent[key] = parent[parent[key]]
            key = parent[key]
        return key

    def union(left: str, right: str) -> None:
        left_root, right_root = find(left), find(right)
        if left_root != right_root:
            parent[right_root] = left_root

    for left_index, left in enumerate(entries):
        for right in entries[left_index + 1 :]:
            shares_code = bool(left["codes"] & right["codes"])
            shares_name = bool(
                left["name"]
                and right["name"]
                and customer_names_match(left["name"], right["name"])
            )
            if shares_code or shares_name:
                union(left["key"], right["key"])
    return {entry["key"]: find(entry["key"]) for entry in entries}


def _customer_names_by_code(rules: Iterable[Mapping[str, Any]]) -> dict[str, str]:
    names: dict[str, str] = {}
    for rule in rules:
        customer_code = _clean(rule.get("customer_code"))
        customer_name = _clean(rule.get("customer_name"))
        if customer_code and customer_name and not _is_placeholder_customer_name(customer_name):
            names.setdefault(customer_code, customer_name)
    return names


def _with_workspace_customer_name(
    rule: Mapping[str, Any], customer_names_by_code: Mapping[str, str]
) -> dict[str, Any]:
    projected = dict(rule)
    customer_code = _clean(projected.get("customer_code"))
    customer_name = _clean(projected.get("customer_name"))
    canonical_name = _clean(customer_names_by_code.get(customer_code))
    if customer_code and canonical_name and _is_placeholder_customer_name(customer_name):
        projected["customer_name"] = canonical_name
    return projected


def _is_placeholder_customer_name(value: Any) -> bool:
    customer_name = _clean(value)
    if not customer_name:
        return True
    return bool(customer_name) and all(char in {"?", "？", " ", "　"} for char in customer_name)


def resolve_customer_code_by_name(name: Any) -> str:
    """Find the stable customer code used by existing rules for a customer name."""
    target = _clean(name)
    if not target:
        return ""
    candidates: Counter[str] = Counter()

    def add(code: Any, candidate_name: Any) -> None:
        code_value = _clean(code)
        if code_value and _strict_customer_name_matches(target, candidate_name):
            candidates[code_value] += 1

    from .transcode_agent_rules import (
        load_transcode_agent_mapping_tables,
        load_transcode_agent_rules,
    )
    from .transcode_semantic_rules import (
        get_active_transcode_semantic_rule_version,
        load_transcode_semantic_rules,
    )

    for rule in load_transcode_agent_rules():
        add(rule.get("客户代码"), rule.get("客户简称"))
    semantic_version = get_active_transcode_semantic_rule_version()
    if semantic_version:
        for rule in load_transcode_semantic_rules(semantic_version):
            add(rule.get("customer_code"), rule.get("customer_name"))
    for rows in load_transcode_agent_mapping_tables().values():
        for row in rows:
            add(row.get("客户代码"), row.get("客户简称"))
    if not candidates:
        return ""
    return max(candidates, key=lambda code: (candidates[code], code))


def _strict_customer_name_matches(left: Any, right: Any) -> bool:
    left_name = normalize_customer_name(left)
    right_name = normalize_customer_name(right)
    if not left_name or not right_name:
        return False
    if left_name == right_name:
        return True
    return any(
        left_name in {normalize_customer_name(item) for item in group}
        and right_name in {normalize_customer_name(item) for item in group}
        for group in CUSTOMER_ALIAS_GROUPS
    )


def _agent_rule_as_workspace_rule(rule: dict[str, Any]) -> dict[str, Any]:
    override_field = _clean(rule.get("覆盖字段"))
    page_maintained = _is_page_maintained_agent_rule(rule)
    # 页面按实际输出参数分类，不能沿用原 Excel 的来源列。否则总/芯厚
    # 规则会被错误放入“基板厚度”等页面。
    business_field = AGENT_OVERRIDE_TO_BUSINESS_FIELD.get(override_field, "")
    if not business_field:
        business_field = _clean(rule.get("原始字段"))
    conditions: list[dict[str, Any]] = []
    condition_specs: list[tuple[str, str, Any]] = [("胶系", "equals", rule.get("条件胶系"))]
    keyword_field, keyword_value = _agent_workspace_keyword_condition(rule.get("条件关键词"))
    if keyword_value:
        condition_specs.append((keyword_field, "contains_any", keyword_value))
    condition_specs.extend(
        (
            ("铜箔规格", "equals", rule.get("条件铜厚")),
            ("基板厚度", "equals", rule.get("条件厚度")),
            ("客户规格", "contains_any", rule.get("条件尺寸")),
        )
    )
    for field, operator, value in condition_specs:
        cleaned = _clean(value)
        # 旧规则表用 0 表示“未设置条件”，并不表示厚度必须等于 0。
        if cleaned and not (field == "基板厚度" and cleaned == "0"):
            conditions.append(
                {
                    "field": field,
                    "operator": operator,
                    "value": list(value) if isinstance(value, (list, tuple)) else cleaned,
                }
            )
    if not conditions:
        conditions.append({"field": "客户规格", "operator": "present", "value": ""})
    source_text = _agent_workspace_source_text(
        rule,
        business_field=business_field,
        conditions=conditions,
    )
    customer_code = _clean(rule.get("客户代码"))
    customer_name = _clean(rule.get("客户简称"))
    is_global = not customer_code and not customer_name
    return {
        "asset_type": AGENT_ASSET_TYPE,
        "rule_id": _clean(rule.get("规则ID")),
        "customer_code": customer_code,
        "customer_name": customer_name or ("全部客户" if is_global else ""),
        "global_scope": is_global,
        "business_field": business_field,
        "source_text": source_text,
        "source_column": _clean(rule.get("来源字段")),
        "conditions": conditions,
        "target_fields": [AGENT_OVERRIDE_TO_TARGET.get(override_field, override_field)],
        "normalized_values": [_clean(rule.get("覆盖值"))],
        "priority": _to_int(rule.get("优先级"), 0),
        "enabled": _clean(rule.get("启用")) == "是",
        "model": "Agent运行规则",
        "approval": {"basis": _clean(rule.get("规则解释")) or _clean(rule.get("命中来源"))},
        "agent_rule": rule,
        "review_state": (
            "active"
            if page_maintained
            else ("pending" if _clean(rule.get("待确认")) == "是" else "migration")
        ),
        # 旧 TAR 运行规则仍继续参与引擎；本页先以可读、只读方式呈现，避免
        # 业务直接编辑底层列结构，待整理成正式客户规则后再维护。页面新增和
        # 页面维护的 Agent 规则则是正式可编辑规则。
        "editable": page_maintained,
        "runtime_active": _clean(rule.get("启用")) == "是",
        "migration_hint": (
            ""
            if page_maintained
            else "该规则仍在营销转码 Agent 中运行；请核对后整理为客户特殊规则进行维护。"
        ),
    }


def _is_page_maintained_agent_rule(rule: Mapping[str, Any]) -> bool:
    """Distinguish page-maintained agent rules from legacy Excel TAR rules."""
    if bool(rule.get("_page_override")):
        return True
    if _clean(rule.get("命中来源")) == "页面维护Agent确定性长期规则":
        return True
    if _clean(rule.get("来源行号")) == "页面新增":
        return True
    return False


def _agent_workspace_keyword_condition(value: Any) -> tuple[str, list[str]]:
    text = _clean(value)
    if not text:
        return "客户规格", []
    normalized = _normalize(text)
    field = "订单备注" if "备注" in normalized else "客户规格"
    for marker in ("当备注中有", "备注中有", "当订单备注有", "订单备注有"):
        if text.startswith(marker):
            text = text[len(marker):]
            break
    text = text.removesuffix("字样时").removesuffix("字样").removesuffix("时")
    values = [
        item.strip()
        for item in text.replace("/", "；").replace("、", "；").replace(",", "；").split("；")
        if item.strip()
    ]
    return field, values or [_clean(value)]


def _agent_workspace_source_text(
    rule: Mapping[str, Any],
    *,
    business_field: str,
    conditions: list[dict[str, Any]],
) -> str:
    source_text = _clean(rule.get("规则文本")) or _clean(rule.get("条件文本"))
    if source_text and not source_text.startswith("字段="):
        return source_text
    condition_text = _condition_summary(
        [
            {
                "field": _clean(condition.get("field")),
                "operator": _clean(condition.get("operator")),
                "value": _condition_value_text(condition.get("value")),
            }
            for condition in conditions
        ]
    )
    result = _clean(rule.get("覆盖值")) or "已维护结果"
    prefix = f"当{condition_text}时，" if condition_text else ""
    return f"{prefix}{business_field}取值为{result}"


def _rule_view(rule: dict[str, Any], *, overridden: bool) -> dict[str, Any]:
    conditions = list(rule.get("conditions") or [])
    source_fields = list(dict.fromkeys(_clean(item.get("field")) for item in conditions if _clean(item.get("field"))))
    editable_conditions = [
        {
            "field": _normalize_condition_field(item.get("field")),
            "operator": _clean(item.get("operator")),
            "value": _condition_value_text(item.get("value")),
        }
        for item in conditions
    ]
    target_field = _first(rule.get("target_fields"))
    source_column = _clean(rule.get("source_column"))
    model = _clean(rule.get("model"))
    asset_type = _clean(rule.get("asset_type")) or SEMANTIC_ASSET_TYPE
    if asset_type == INFORMATION_ASSET_TYPE:
        origin = model or "规则资料"
    elif asset_type == LEGACY_CCL_ASSET_TYPE:
        origin = "旧CCL有效规则"
    elif asset_type == CUSTOMER_ORDER_ASSET_TYPE:
        origin = "客户下单转换规则"
    elif asset_type == CODE_MIGRATION_ASSET_TYPE:
        origin = "代码客户规则"
    elif asset_type == MAPPING_ASSET_TYPE:
        origin = "客户专属转换规则"
    elif asset_type == AGENT_ASSET_TYPE:
        origin = "页面维护Agent长期规则" if overridden else "Agent确定性长期规则"
    elif source_column == "确认中心" or model == "确认中心人工规则":
        origin = "确认中心待生效规则"
    elif overridden:
        origin = "页面维护"
    else:
        origin = "正式规则表"
    target_value = _first(rule.get("normalized_values"))
    target_value_label = TARGET_VALUE_LABELS.get(
        (target_field, target_value.lower()),
        target_value,
    )
    scope_key = _rule_scope(rule)
    review_state = _clean(rule.get("review_state"))
    status_labels = {
        "pending": "待业务确认",
        "missing_selection_condition": "缺选择条件/待业务确认",
        "technical": "待技术支持",
        "reference": "外部资料引用",
        "history": "历史样本建议",
        "migration": "后台运行规则（暂不可维护）",
    }
    rule_family_labels = {
        LEGACY_CCL_ASSET_TYPE: "旧CCL有效规则",
        CUSTOMER_ORDER_ASSET_TYPE: "客户下单转换规则",
        CODE_MIGRATION_ASSET_TYPE: "代码客户规则",
        MAPPING_ASSET_TYPE: "客户转换规则",
        AGENT_ASSET_TYPE: "规格条件规则",
        INFORMATION_ASSET_TYPE: model or "规则资料",
    }
    if asset_type == SEMANTIC_ASSET_TYPE:
        rule_family_label = "订单备注语义规则" if "订单备注" in source_fields else "客户条件规则"
    else:
        rule_family_label = rule_family_labels.get(asset_type, origin)
    status_label = status_labels.get(
        review_state,
        "启用" if bool(rule.get("enabled")) else "停用",
    )
    return {
        "rule_id": _clean(rule.get("rule_id")),
        "asset_type": asset_type,
        "customer_code": _clean(rule.get("customer_code")),
        "customer_name": _clean(rule.get("customer_name")),
        "business_field": _clean(rule.get("business_field")),
        "source_text": _clean(rule.get("source_text")),
        "input_source": " + ".join(source_fields),
        "conditions": editable_conditions,
        "condition_summary": _condition_summary(editable_conditions),
        "target_field": target_field,
        "target_value": target_value,
        "target_value_label": target_value_label or "—",
        "priority": int(rule.get("priority") or 0),
        "enabled": bool(rule.get("enabled")),
        "semantic_enabled": "订单备注" in source_fields,
        "approval_basis": _clean((rule.get("approval") or {}).get("basis")),
        "origin": origin,
        "rule_family_label": rule_family_label,
        "machine_rule": rule,
        "agent_rule": dict(rule.get("agent_rule") or {}),
        "mapping_group": _clean(rule.get("mapping_group")),
        "mapping_row": dict(rule.get("mapping_row") or {}),
        "scope_key": scope_key,
        "scope_label": RULE_SCOPE_LABELS[scope_key],
        "status_label": status_label,
        "status_detail": _clean(rule.get("status_detail"))
        or _clean(rule.get("review_reason"))
        or _clean((rule.get("approval") or {}).get("basis")),
        "editable": bool(rule.get("editable", asset_type != INFORMATION_ASSET_TYPE)),
        "migration_hint": _clean(rule.get("migration_hint")),
        "runtime_active": bool(rule.get("runtime_active")),
    }


def _build_conditions(fields: list[str], operators: list[str], values: list[str]) -> list[dict[str, Any]]:
    conditions: list[dict[str, Any]] = []
    row_count = max(len(fields), len(operators), len(values))
    for index in range(row_count):
        field = _normalize_condition_field(fields[index] if index < len(fields) else "")
        operator = _clean(operators[index] if index < len(operators) else "")
        raw_value = _clean(values[index] if index < len(values) else "")
        if not field and not operator and not raw_value:
            continue
        if field not in CONDITION_FIELDS:
            raise CustomerRuleMaintenanceError(f"第{index + 1}条条件字段无效：{field}")
        if operator not in CONDITION_OPERATORS:
            raise CustomerRuleMaintenanceError(f"第{index + 1}条条件操作符无效：{operator}")
        if operator in {"missing", "present"}:
            value: Any = ""
        elif operator in LIST_OPERATORS:
            value = [item.strip() for item in raw_value.replace(",", "；").split("；") if item.strip()]
            if not value:
                raise CustomerRuleMaintenanceError(f"第{index + 1}条条件值不能为空。")
        elif operator in NUMBER_OPERATORS:
            try:
                value = float(raw_value)
            except ValueError as exc:
                raise CustomerRuleMaintenanceError(f"第{index + 1}条条件值必须是数字。") from exc
        else:
            if not raw_value:
                raise CustomerRuleMaintenanceError(f"第{index + 1}条条件值不能为空。")
            value = raw_value
        conditions.append(
            {
                "field": field,
                "operator": operator,
                "value": value,
                "source_scope": field,
            }
        )
    if not conditions:
        raise CustomerRuleMaintenanceError("至少维护一条触发条件。")
    return conditions


def _normalize_condition_field(value: Any) -> str:
    """Keep legacy condition fields editable after the page list was simplified."""
    field = _clean(value)
    if field == "订单规格":
        return "客户规格"
    if field == "品号/物料编号":
        return "客户物料编码"
    return field


def _validate_target_value(target_field: str, value: str) -> None:
    raw = _clean(value)
    upper = raw.upper()
    if target_field == "grade_intent" and upper not in {str(item).upper() for item in OFFICIAL_GRADE_CODES}:
        raise CustomerRuleMaintenanceError(f"基板级别代码无效：{raw}")
    if target_field == "glue_category" and upper not in {"Y", "R"}:
        raise CustomerRuleMaintenanceError("胶水类别只允许Y或R。")
    if target_field == "total_core" and raw.lower() not in {
        "core",
        "total",
        "c",
        "t",
        "芯厚",
        "总厚",
        "core_after_total_to_core_conversion",
    }:
        raise CustomerRuleMaintenanceError("总/芯厚结果无效。")
    if target_field == "thickness" and (not raw.isdigit() or len(raw) != 5):
        raise CustomerRuleMaintenanceError("厚度结果必须是5位厚度代码。")
    if target_field == "size" and not (
        (raw.isdigit() and len(raw) == 8) or raw == "height_plus_0.3"
    ):
        raise CustomerRuleMaintenanceError("尺寸结果必须是8位尺寸代码或height_plus_0.3。")


def _validate_agent_override_value(override_field: str, value: str) -> None:
    upper = _clean(value).upper()
    if override_field == "grade_code" and upper not in {str(item).upper() for item in OFFICIAL_GRADE_CODES}:
        raise CustomerRuleMaintenanceError(f"基板级别代码无效：{value}")
    if override_field == "glue_category_code" and upper not in {"Y", "R"}:
        raise CustomerRuleMaintenanceError("胶水类别只允许Y或R。")
    if override_field == "tc_code" and upper not in {"C", "T"}:
        raise CustomerRuleMaintenanceError("总/芯厚只允许C或T。")
    if override_field == "thickness_code" and (not upper.isdigit() or len(upper) != 5):
        raise CustomerRuleMaintenanceError("厚度结果必须是5位厚度代码。")
    if override_field == "size_code" and (not upper.isdigit() or len(upper) != 8):
        raise CustomerRuleMaintenanceError("尺寸结果必须是8位尺寸代码。")


def _semantic_type(conditions: list[dict[str, Any]]) -> str:
    if len(conditions) > 1:
        return "multi_condition"
    operator = conditions[0]["operator"]
    if operator in {"contains_any", "contains_all"}:
        return "keyword_present"
    if operator == "not_contains":
        return "keyword_absent"
    if operator == "missing":
        return "default_when_missing"
    if operator in NUMBER_OPERATORS:
        return "comparison"
    return "explicit_fact"


def _condition_summary(conditions: list[dict[str, str]]) -> str:
    return "；".join(
        f"{item['field']} {CONDITION_OPERATOR_LABELS.get(item['operator'], item['operator'])}"
        + (f" {item['value']}" if item["value"] else "")
        for item in conditions
    )


def _override_ids() -> set[str]:
    ensure_customer_rule_maintenance_tables()
    with db_cursor() as conn:
        rows = conn.execute(
            """
            SELECT rule_id FROM transcode_customer_rule_overrides WHERE deleted = 0
            UNION
            SELECT rule_id FROM transcode_agent_rule_overrides WHERE deleted = 0
            """
        ).fetchall()
    return {str(row["rule_id"]) for row in rows}


def _insert_rule_change(
    conn,
    *,
    rule_id: str,
    action: str,
    employee_id: str,
    before: dict[str, Any] | None,
    after: dict[str, Any] | None,
    created_at: str,
) -> None:
    conn.execute(
        """
        INSERT INTO transcode_customer_rule_changes
            (rule_id, action, employee_id, before_json, after_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            rule_id,
            action,
            employee_id,
            _dump(before) if before else None,
            _dump(after) if after else None,
            created_at,
        ),
    )


def _parse_rule_json(value: Any, rule_id: str) -> dict[str, Any]:
    parsed = _optional_rule_json(value)
    if not isinstance(parsed, dict):
        raise CustomerRuleMaintenanceError(f"页面维护规则JSON无效：{rule_id}")
    return parsed


def _optional_rule_json(value: Any) -> dict[str, Any] | None:
    if not value:
        return None
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, dict) else None


def _condition_value_text(value: Any) -> str:
    if isinstance(value, list):
        return "；".join(str(item) for item in value)
    return _clean(value)


def _form_list(form: Mapping[str, Any], key: str) -> list[str]:
    getter = getattr(form, "getlist", None)
    if callable(getter):
        return [str(item or "") for item in getter(key)]
    value = form.get(key)
    return list(value) if isinstance(value, (list, tuple)) else [str(value or "")]


def _form_bool(form: Mapping[str, Any], key: str, *, default: bool) -> bool:
    if key not in form:
        return False if hasattr(form, "getlist") else default
    return _clean(form.get(key)).lower() in {"1", "true", "yes", "on", "是"}


def _new_rule_id() -> str:
    return datetime.now().strftime("TCR-%Y%m%d-%H%M%S-%f")


def _now() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _dump(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _normalize(value: Any) -> str:
    return "".join(_clean(value).upper().split())


def _first(values: Any) -> str:
    if isinstance(values, (list, tuple)) and values:
        return _clean(values[0])
    return ""


def _to_int(value: Any, default: int) -> int:
    try:
        return int(float(_clean(value)))
    except ValueError:
        return default
