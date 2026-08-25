from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .customer_archive_service import as_text, utcnow
from .database import automation_cursor as db_cursor


PRODUCT_TYPE_LABELS = {"base": "基板", "pp": "PP"}
MATCH_STATUS_LABELS = {"matched": "已匹配客户档案", "unmatched": "客户档案待匹配"}
POSITION_FIELDS = (
    ("glue_system_position", "胶系位置", "TC_FEM04"),
    ("thickness_position", "厚度位置", "TC_FEM05"),
    ("core_thickness_position", "芯总厚位置", "TC_FEM06"),
    ("dimension_position", "尺寸位置", "TC_FEM07"),
    ("copper_foil_type_position", "铜箔类型", "TC_FEM08"),
    ("copper_thickness_position", "铜厚", "TC_FEM09"),
    ("structure_position", "结构", "TC_FEM10"),
    ("watermark_position", "有无水印", "TC_FEM11"),
    ("halogen_position", "有卤无卤", "TC_FEM12"),
    ("rc_position", "RC", "TC_FEM13"),
    ("cloth_type_position", "布种", "TC_FEM14"),
    ("size_position", "尺寸", "TC_FEM15"),
)
POSITION_FIELD_LABELS = {field: label for field, label, _header in POSITION_FIELDS}
FORM_FIELDS = (
    "customer_code", "customer_name", "product_type", "delimiter",
    *(field for field, _label, _header in POSITION_FIELDS),
    "note", "enabled",
)
_DIRECTIONAL_SIZE_PATTERN = re.compile(
    r"经\s*(\d+(?:\.\d+)?)\s*(inch|mm|毫米|英寸|m|米)?\s*(?:[xX*×]\s*)?"
    r"纬\s*(\d+(?:\.\d+)?)\s*(inch|mm|毫米|英寸|m|米)?",
    re.IGNORECASE,
)


def _row_dict(row: Any) -> dict[str, Any]:
    return dict(row) if row is not None else {}


def _product_type(value: Any) -> str:
    text = as_text(value).lower()
    aliases = {"1": "base", "base": "base", "基板": "base", "2": "pp", "pp": "pp"}
    if text not in aliases:
        raise ValueError("基板PP仅支持1（基板）或2（PP）。")
    return aliases[text]


def _enabled(value: Any) -> int:
    text = as_text(value).strip().lower()
    if text in {"0", "n", "no", "否", "停用", "disabled"}:
        return 0
    if text in {"", "1", "y", "yes", "是", "启用", "active"}:
        return 1
    raise ValueError("有效否仅支持空值、1/Y/是或0/N/否。")


def _position(value: Any, label: str) -> int | None:
    text = as_text(value)
    if not text:
        return None
    try:
        number = int(text)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label}必须是整数。") from exc
    if not 1 <= number <= 99999:
        raise ValueError(f"{label}必须在1到99999之间。")
    return number


def _validated(values: dict[str, Any]) -> dict[str, Any]:
    customer_code = as_text(values.get("customer_code"))
    if not customer_code:
        raise ValueError("客户编号不能为空。")
    data: dict[str, Any] = {
        "customer_code": customer_code,
        "customer_name": as_text(values.get("customer_name")),
        "product_type": _product_type(values.get("product_type")),
        "delimiter": as_text(values.get("delimiter")),
        "note": as_text(values.get("note")),
        "enabled": _enabled(values.get("enabled")),
    }
    for field, label, _header in POSITION_FIELDS:
        data[field] = _position(values.get(field), label)
    return data


def list_spec_mappings(
    *, keyword: str = "", product_type: str = "all", status: str = "all",
    match_status: str = "all",
) -> list[dict[str, Any]]:
    clauses = ["1=1"]
    params: list[Any] = []
    if product_type in PRODUCT_TYPE_LABELS:
        clauses.append("m.product_type=?")
        params.append(product_type)
    if status in {"active", "disabled"}:
        clauses.append("m.enabled=?")
        params.append(1 if status == "active" else 0)
    if match_status == "matched":
        clauses.append("c.id IS NOT NULL")
    elif match_status == "unmatched":
        clauses.append("c.id IS NULL")
    if keyword.strip():
        like = f"%{keyword.strip()}%"
        clauses.append(
            "(m.customer_code LIKE ? OR m.customer_name LIKE ? OR m.note LIKE ? "
            "OR c.customer_short_name LIKE ? OR c.customer_name LIKE ?)"
        )
        params.extend([like] * 5)
    with db_cursor() as conn:
        rows = conn.execute(
            f"""SELECT m.*, c.id AS customer_id,
                       c.customer_name AS master_customer_name,
                       c.customer_short_name AS master_customer_short_name,
                       CASE WHEN c.id IS NULL THEN 0 ELSE 1 END AS customer_matched
                  FROM automation_customer_spec_mappings m
             LEFT JOIN automation_customers c ON c.customer_code=m.customer_code
                 WHERE {' AND '.join(clauses)}
              ORDER BY m.customer_code,
                       CASE m.product_type WHEN 'base' THEN 0 ELSE 1 END""",
            params,
        ).fetchall()
    return [_row_dict(row) for row in rows]


def get_spec_mapping(mapping_id: int) -> dict[str, Any] | None:
    with db_cursor() as conn:
        row = conn.execute(
            """SELECT m.*, c.id AS customer_id,
                      CASE WHEN c.id IS NULL THEN 0 ELSE 1 END AS customer_matched
                 FROM automation_customer_spec_mappings m
            LEFT JOIN automation_customers c ON c.customer_code=m.customer_code
                WHERE m.id=?""",
            (mapping_id,),
        ).fetchone()
    return _row_dict(row) if row else None


def _matching_position(token: str, mapping: dict[str, Any], next_position: int) -> int | None:
    text = token.strip()
    upper = text.upper()
    compact = re.sub(r"\s+", "", upper)
    is_dimension = "经" in text and "纬" in text or bool(
        re.search(r"\d+(?:\.\d+)?(?:MM|INCH|英寸)?[X*×]\d+(?:\.\d+)?", compact)
    )
    checks = {
        "glue_system_position": bool(re.search(r"(?<![A-Z0-9])(?:NY|NPG?)[-A-Z0-9]+", upper)),
        "thickness_position": bool(
            not is_dimension
            and re.fullmatch(r"\d+(?:\.\d+)?\s*(?:MM|MIL)", upper)
        ),
        "core_thickness_position": any(value in text for value in ("芯厚", "总厚", "含铜", "不含铜")),
        "dimension_position": is_dimension,
        "copper_foil_type_position": any(value in upper for value in ("HVLP", "RTF", "VLP", "ED铜箔", "铜箔类型")),
        "copper_thickness_position": bool(
            re.fullmatch(r"(?:H|\d+(?:\.\d+)?)(?:/|\\)(?:H|\d+(?:\.\d+)?)(?:OZ)?", compact)
            or ("OZ" in upper and bool(re.search(r"\d", upper)))
        ),
        "structure_position": bool(
            re.fullmatch(r"\d+(?:\.\d+)?(?:[+*]\d+(?:\.\d+)?)+", compact)
        ),
        "watermark_position": any(value in upper for value in ("水印", "LOGO", "印字")),
        "halogen_position": any(value in text for value in ("有卤", "无卤")),
        "rc_position": bool(re.search(r"(?:RC|含量)\s*[=:]?\s*\d+(?:\.\d+)?%", upper)),
        "cloth_type_position": bool(
            re.fullmatch(r"(?:10[368]|10\d{2}|2\d{3}|7\d{3})", compact)
        ),
        "size_position": is_dimension or any(value in text for value in ("卷", "米")),
    }
    candidates = [
        int(mapping[field])
        for field, matched in checks.items()
        if matched and mapping.get(field) not in (None, "") and int(mapping[field]) >= next_position
    ]
    return min(candidates) if candidates else None


def _compact_directional_size(match: re.Match[str]) -> str:
    first_value, first_unit, second_value, second_unit = match.groups()
    first_unit = first_unit.lower() if first_unit and first_unit.isascii() else first_unit
    second_unit = second_unit.lower() if second_unit and second_unit.isascii() else second_unit
    suffix = second_unit or ""
    return f"经{first_value}{first_unit or ''}纬{second_value}{suffix}"


def _split_spec_parts(spec: str, delimiter: str) -> list[str]:
    # Normalize spacing first; field boundaries still come only from the configured delimiter.
    normalized = _DIRECTIONAL_SIZE_PATTERN.sub(_compact_directional_size, spec)
    if delimiter:
        return [part.strip() for part in normalized.split(delimiter)]
    return [part for part in re.split(r"\s+", normalized.strip()) if part]


def _formatted_spec_parts(
    customer_spec: Any, mapping: dict[str, Any] | None,
) -> tuple[list[str], str]:
    if not mapping or not mapping.get("enabled"):
        return [], " "
    spec = as_text(customer_spec)
    if not spec:
        return [], " "
    delimiter = as_text(mapping.get("delimiter"))
    parts = _split_spec_parts(spec, delimiter)
    if not any(mapping.get(field) not in (None, "") for field, _label, _header in POSITION_FIELDS):
        return [], delimiter or " "
    separator = delimiter or " "
    result: list[str] = []
    for part in parts:
        target_position = _matching_position(part, mapping, len(result) + 1)
        if target_position:
            result.extend("*" for _ in range(len(result) + 1, target_position))
        result.append(part or "*")
    return result, separator


def format_customer_spec_match(customer_spec: Any, mapping: dict[str, Any] | None) -> str:
    """Keep every source segment and insert wildcards to reach configured target positions."""
    parts, separator = _formatted_spec_parts(customer_spec, mapping)
    return separator.join(parts)


def format_customer_spec_match_detail(
    customer_spec: Any, mapping: dict[str, Any] | None,
) -> dict[str, Any]:
    parts, separator = _formatted_spec_parts(customer_spec, mapping)
    fields = []
    for field, label, header in POSITION_FIELDS:
        raw_position = mapping.get(field) if mapping else None
        position = int(raw_position) if raw_position not in (None, "") else None
        value = parts[position - 1] if position and position <= len(parts) else ""
        fields.append({
            "field": field,
            "label": label.removesuffix("位置"),
            "header": header,
            "position": position,
            "value": "" if value == "*" else value,
        })
    return {
        "customer_spec_match": separator.join(parts),
        "delimiter": separator,
        "segments": parts,
        "fields": fields,
        "mapping_found": bool(mapping and mapping.get("enabled")),
        "mapping_note": as_text(mapping.get("note")) if mapping else "",
    }


def build_customer_spec_match(customer_code: Any, product_type: Any, customer_spec: Any) -> str:
    return build_customer_spec_match_detail(
        customer_code, product_type, customer_spec,
    )["customer_spec_match"]


def build_customer_spec_match_detail(
    customer_code: Any, product_type: Any, customer_spec: Any,
) -> dict[str, Any]:
    code = as_text(customer_code)
    if not code or not as_text(customer_spec):
        return format_customer_spec_match_detail(customer_spec, None)
    try:
        normalized_type = _product_type(product_type)
    except ValueError:
        return format_customer_spec_match_detail(customer_spec, None)
    with db_cursor() as conn:
        row = conn.execute(
            """SELECT * FROM automation_customer_spec_mappings
               WHERE customer_code=? AND product_type=? AND enabled=1""",
            (code, normalized_type),
        ).fetchone()
    return format_customer_spec_match_detail(customer_spec, _row_dict(row) if row else None)


def save_spec_mapping(
    values: dict[str, Any], *, mapping_id: int | None = None, operated_by: str = "",
) -> int:
    data = _validated(values)
    now = utcnow()
    columns = [
        "customer_code", "customer_name", "product_type", "delimiter",
        *(field for field, _label, _header in POSITION_FIELDS),
        "note", "enabled",
    ]
    with db_cursor() as conn:
        duplicate_params: list[Any] = [data["customer_code"], data["product_type"]]
        duplicate_sql = (
            "SELECT id FROM automation_customer_spec_mappings "
            "WHERE customer_code=? AND product_type=?"
        )
        if mapping_id:
            duplicate_sql += " AND id<>?"
            duplicate_params.append(mapping_id)
        if conn.execute(duplicate_sql, duplicate_params).fetchone():
            raise ValueError("该客户的基板/PP规格对照已存在。")
        if mapping_id:
            if not conn.execute(
                "SELECT id FROM automation_customer_spec_mappings WHERE id=?", (mapping_id,)
            ).fetchone():
                raise ValueError("规格对照记录不存在。")
            assignments = ",".join(f"{column}=?" for column in columns)
            conn.execute(
                f"UPDATE automation_customer_spec_mappings SET {assignments},updated_by=?,updated_at=? WHERE id=?",
                (*[data[column] for column in columns], operated_by, now, mapping_id),
            )
            return mapping_id
        cursor = conn.execute(
            f"""INSERT INTO automation_customer_spec_mappings
                    ({','.join(columns)},source_json,updated_by,created_at,updated_at)
                VALUES ({','.join('?' for _ in columns)},?,?,?,?)""",
            (*[data[column] for column in columns], "{}", operated_by, now, now),
        )
        return int(cursor.lastrowid)


def set_spec_mapping_enabled(mapping_id: int, enabled: bool, *, operated_by: str = "") -> None:
    with db_cursor() as conn:
        result = conn.execute(
            "UPDATE automation_customer_spec_mappings SET enabled=?,updated_by=?,updated_at=? WHERE id=?",
            (1 if enabled else 0, operated_by, utcnow(), mapping_id),
        )
        if result.rowcount == 0:
            raise ValueError("规格对照记录不存在。")


def _normalized_headers(values: tuple[Any, ...]) -> list[str]:
    headers: list[str] = []
    for value in values:
        header = as_text(value).upper().replace(" ", "")
        headers.append("TC_FEM03" if header == "分隔符" else header)
    return headers


def import_spec_mapping_workbook(
    file_path: str | Path, *, operated_by: str = "",
) -> dict[str, Any]:
    book = load_workbook(file_path, read_only=True, data_only=True)
    sheet = book.active
    headers: list[str] = []
    header_index = 0
    try:
        for index, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True), start=1
        ):
            candidate = _normalized_headers(row)
            if "TC_FEM01" in candidate and "TC_FEM02" in candidate:
                headers, header_index = candidate, index
                break
        if not headers:
            raise ValueError("未找到TC_FEM01和TC_FEM02表头，无法导入规格对照表。")
        header_positions = {header: index for index, header in enumerate(headers) if header}
        required = {"TC_FEM01", "TC_FEM02"}
        missing = required - set(header_positions)
        if missing:
            raise ValueError(f"缺少必要表头：{','.join(sorted(missing))}。")

        parsed: list[tuple[int, dict[str, Any], dict[str, Any]]] = []
        errors: list[dict[str, Any]] = []
        position_headers = {field: header for field, _label, header in POSITION_FIELDS}
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_index + 1, values_only=True), start=header_index + 1
        ):
            source = {
                header: as_text(row[column]) if column < len(row) else ""
                for header, column in header_positions.items()
            }
            if not any(source.values()):
                continue
            values: dict[str, Any] = {
                "customer_code": source.get("TC_FEM01", ""),
                "customer_name": source.get("OCC02", ""),
                "product_type": source.get("TC_FEM02", ""),
                "delimiter": source.get("TC_FEM03", ""),
                "note": source.get("TC_FEM16", ""),
                "enabled": source.get("TC_FEM17", ""),
            }
            values.update({field: source.get(header, "") for field, header in position_headers.items()})
            try:
                parsed.append((row_number, _validated(values), source))
            except ValueError as exc:
                errors.append({"row": row_number, "error": str(exc)})

        imported = updated = 0
        columns = [
            "customer_code", "customer_name", "product_type", "delimiter",
            *(field for field, _label, _header in POSITION_FIELDS),
            "note", "enabled",
        ]
        now = utcnow()
        with db_cursor() as conn:
            for _row_number, data, source in parsed:
                existing = conn.execute(
                    "SELECT id FROM automation_customer_spec_mappings WHERE customer_code=? AND product_type=?",
                    (data["customer_code"], data["product_type"]),
                ).fetchone()
                source_json = json.dumps(source, ensure_ascii=False)
                if existing:
                    assignments = ",".join(f"{column}=?" for column in columns)
                    conn.execute(
                        f"UPDATE automation_customer_spec_mappings SET {assignments},source_json=?,updated_by=?,updated_at=? WHERE id=?",
                        (*[data[column] for column in columns], source_json, operated_by, now, _row_dict(existing)["id"]),
                    )
                    updated += 1
                else:
                    conn.execute(
                        f"""INSERT INTO automation_customer_spec_mappings
                                ({','.join(columns)},source_json,updated_by,created_at,updated_at)
                            VALUES ({','.join('?' for _ in columns)},?,?,?,?)""",
                        (*[data[column] for column in columns], source_json, operated_by, now, now),
                    )
                    imported += 1
        return {
            "imported": imported,
            "updated": updated,
            "skipped": len(errors),
            "errors": errors,
            "processed": imported + updated,
        }
    finally:
        book.close()
