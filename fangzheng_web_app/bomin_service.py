from __future__ import annotations

import math
import re
import shutil
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from werkzeug.utils import secure_filename

from .bomin_rules import get_active_bomin_rule_version, get_bomin_rule_file_path
from .db import append_job_log, create_job, get_job, update_job_status
from .excel_utils import load_workbook_compat, normalized_xlsx_source
from .job_control import launch_job_process
from .paths import JOBS_DIR


PRICE_OUTPUT_HEADER = "博敏计算价格"
NOTE_SHEET_NAME = "计算说明"
DESC_HEADERS = {"客户规格", "物料描述", "规格"}
ANSWER_HEADER = "单价（含税）"


@dataclass
class PriceRow:
    excel_row: int
    values: dict[str, Any]


@dataclass
class RuleBook:
    ccl_rows: list[PriceRow]
    pp_rows: list[PriceRow]


@dataclass
class ParsedSpec:
    row_idx: int
    desc: str
    material_type: str
    product: str | None = None
    thickness: float | None = None
    copper: str | None = None
    copper_state: str | None = None
    foil: str | None = None
    stack: str | None = None
    glass: str | None = None
    rc: float | None = None
    length_value: float | None = None
    length_unit: str | None = None
    width_value: float | None = None
    width_unit: str | None = None
    halogen: str | None = None
    tg: int | None = None


@dataclass
class CalcResult:
    row_idx: int
    desc: str
    status: str
    price: float | None
    note: str


def queue_bomin_job(employee_id: str, uploaded_file, source_filename: str) -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    safe_filename = secure_filename(source_filename) or f"bomin_upload_{timestamp}.xlsx"
    input_path = employee_dir / f"{timestamp}_{safe_filename}"
    uploaded_file.save(input_path)

    rule_version = get_active_bomin_rule_version()
    job_id = create_job(
        employee_id,
        source_filename,
        str(input_path),
        rule_version or "内置博敏价格表未初始化",
        feature="bomin",
    )
    launch_job_process(job_id, "bomin", employee_id)
    return job_id


def run_bomin_job(job_id: int, employee_id: str) -> None:
    job = get_job(job_id)
    if not job or job["employee_id"] != employee_id:
        return

    update_job_status(job_id, status="running")
    try:
        rule_version = get_active_bomin_rule_version()
        if not rule_version:
            raise ValueError("未找到博敏价格表规则，请先上传博敏价格表")
        rule_path = get_bomin_rule_file_path(rule_version)
        append_job_log(job_id, f"开始处理博敏价格计算任务，规则版本：{rule_version}")
        rules = load_bomin_rules(rule_path)
        append_job_log(job_id, f"已加载博敏价格表：CCL {len(rules.ccl_rows)} 行，PP {len(rules.pp_rows)} 行")

        workbook = load_workbook_compat(job["stored_input_path"], data_only=False)
        source_for_result = normalized_xlsx_source(job["stored_input_path"], workbook)
        output_path = Path(job["stored_input_path"]).with_name(
            f"{Path(job['stored_input_path']).stem}_博敏计算结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        shutil.copy2(source_for_result, output_path)

        result_wb = load_workbook_compat(output_path, data_only=False)
        all_results = process_bomin_workbook(result_wb, rules, job_id=job_id)
        result_wb.save(output_path)

        success_count = sum(1 for item in all_results if item.status == "成功")
        fail_count = sum(1 for item in all_results if item.status == "失败")
        update_job_status(
            job_id,
            status="completed",
            stored_result_path=str(output_path),
            success_count=success_count,
            fail_count=fail_count,
            skip_count=0,
            current_row=len(all_results),
            total_rows=len(all_results),
            completed=True,
        )
        append_job_log(job_id, f"博敏价格计算完成：成功 {success_count} 行，失败 {fail_count} 行")
    except Exception as exc:
        append_job_log(job_id, f"博敏价格计算失败：{exc}")
        update_job_status(
            job_id,
            status="failed",
            error_message=f"{exc}\n{traceback.format_exc(limit=8)}",
            completed=True,
        )


def load_bomin_rules(rule_path: str | Path) -> RuleBook:
    workbook = load_workbook_compat(rule_path, data_only=True)
    ccl_ws = workbook["CCL"]
    pp_ws = workbook["PP"]

    ccl_header_row, ccl_columns = _find_ccl_columns(ccl_ws)
    pp_header_row, pp_columns = _find_pp_columns(pp_ws)

    ccl_rows: list[PriceRow] = []
    first_data_row = ccl_header_row + 1
    for excel_row, row in enumerate(ccl_ws.iter_rows(min_row=first_data_row, values_only=True), start=first_data_row):
        if not row or not _row_value(row, ccl_columns["product"]):
            continue
        values = {
            "product": _norm_product(_row_value(row, ccl_columns["product"])),
            "thickness": _to_float(_row_value(row, ccl_columns["thickness"])),
            "copper": _norm_copper(_row_value(row, ccl_columns["copper"])),
            "copper_state": _norm_text(_row_value(row, ccl_columns["copper_state"])),
            "foil": _norm_token(_row_value(row, ccl_columns["foil"])),
            "stack": _canonical_stack(_row_value(row, ccl_columns["stack"])),
            "SQ.FT": _to_float(_row_value(row, ccl_columns["price_start"])),
            '37"X49"': _to_float(_row_value(row, ccl_columns["price_start"] + 1)),
            '41"X49"': _to_float(_row_value(row, ccl_columns["price_start"] + 2)),
            '43"X49"': _to_float(_row_value(row, ccl_columns["price_start"] + 3)),
            "37*43": _to_float(_row_value(row, ccl_columns["price_start"] + 4)),
            "41*43": _to_float(_row_value(row, ccl_columns["price_start"] + 5)),
        }
        if not (
            values["product"]
            and values["product"].startswith("NY")
            and values["thickness"] is not None
            and values["copper"]
            and values["copper_state"] in {"含铜", "不含铜"}
            and values["foil"]
            and values["stack"]
        ):
            continue
        ccl_rows.append(PriceRow(excel_row, values))

    pp_rows: list[PriceRow] = []
    first_pp_data_row = pp_header_row + 1
    for excel_row, row in enumerate(pp_ws.iter_rows(min_row=first_pp_data_row, values_only=True), start=first_pp_data_row):
        if not row or not _row_value(row, pp_columns["product"]):
            continue
        values = {
            "product": _norm_product(_row_value(row, pp_columns["product"])),
            "glass": _norm_glass(_row_value(row, pp_columns["glass"])),
            "rc": _to_rc(_row_value(row, pp_columns["rc"])),
            "length_m": _to_float(_row_value(row, pp_columns["length_m"])),
            "per_sf": _to_float(_row_value(row, pp_columns["per_sf"])),
            "per_m": _to_float(_row_value(row, pp_columns["per_m"])),
            "per_roll": _to_float(_row_value(row, pp_columns["per_roll"])),
        }
        if not (
            values["product"]
            and values["product"].startswith("NY")
            and values["glass"]
            and values["rc"] is not None
            and values["per_m"] is not None
        ):
            continue
        pp_rows.append(PriceRow(excel_row, values))

    return RuleBook(ccl_rows=ccl_rows, pp_rows=pp_rows)


def _find_ccl_columns(worksheet) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 50), values_only=True),
        start=1,
    ):
        headers = [_norm_token(value) for value in row]
        if "PRODUCT" not in headers or "STACKING" not in headers or "SQ.FT" not in headers:
            continue

        product = headers.index("PRODUCT")
        stack = headers.index("STACKING")
        price_start = headers.index("SQ.FT")
        thickness = next((index for index in range(product + 1, stack) if headers[index] == "MM"), None)
        copper_columns = [index for index in range(product + 1, stack) if headers[index] == "CU"]
        if thickness is None or len(copper_columns) < 2:
            continue

        copper, foil = copper_columns[:2]
        return row_number, {
            "product": product,
            "thickness": thickness,
            "copper": copper,
            "copper_state": copper + 1,
            "foil": foil,
            "stack": stack,
            "price_start": price_start,
        }

    return 2, {
        "product": 0,
        "thickness": 1,
        "copper": 2,
        "copper_state": 3,
        "foil": 4,
        "stack": 5,
        "price_start": 6,
    }


def _find_pp_columns(worksheet) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 50), values_only=True),
        start=1,
    ):
        headers = [_norm_token(value) for value in row]
        required = {"PRODUCTS", "GLASSTYPE", "LENGTH(M)", "PERSF", "PERM", "PERROLL"}
        if not required.issubset(headers):
            continue

        resin_columns = [index for index, header in enumerate(headers) if header == "RESINCONTENT"]
        if not resin_columns:
            continue
        return row_number, {
            "product": headers.index("PRODUCTS"),
            "glass": headers.index("GLASSTYPE"),
            "rc": resin_columns[-1],
            "length_m": headers.index("LENGTH(M)"),
            "per_sf": headers.index("PERSF"),
            "per_m": headers.index("PERM"),
            "per_roll": headers.index("PERROLL"),
        }

    return 2, {
        "product": 0,
        "glass": 1,
        "rc": 3,
        "length_m": 4,
        "per_sf": 5,
        "per_m": 6,
        "per_roll": 7,
    }


def _row_value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def calculate_bomin_quote(spec: str) -> dict[str, Any]:
    spec = _norm_text(spec)
    if not spec:
        return {
            "status": "失败",
            "price": None,
            "note": "",
            "material_type": "",
            "rule_version": "",
            "error": "请输入客户规格",
        }

    rule_version = get_active_bomin_rule_version()
    if not rule_version:
        return {
            "status": "失败",
            "price": None,
            "note": "",
            "material_type": "",
            "rule_version": "",
            "error": "未找到博敏价格表规则",
        }

    rules = load_bomin_rules(get_bomin_rule_file_path(rule_version))
    parsed = parse_spec(spec, row_idx=1)
    result = calculate_single(parsed, rules)
    return {
        "status": result.status,
        "price": result.price,
        "note": result.note,
        "material_type": parsed.material_type,
        "rule_version": rule_version,
        "error": "" if result.price is not None else result.note,
    }


def process_bomin_workbook(workbook: Workbook, rules: RuleBook, *, job_id: int | None = None) -> list[CalcResult]:
    all_results: list[CalcResult] = []
    worksheets = [ws for ws in workbook.worksheets if ws.title != NOTE_SHEET_NAME]
    total_rows = 0
    sheet_jobs: list[tuple[Any, int, int, list[int]]] = []
    for ws in worksheets:
        header_row, desc_col = _find_desc_column(ws)
        if not header_row or not desc_col:
            continue
        data_rows = [
            row_idx
            for row_idx in range(header_row + 1, ws.max_row + 1)
            if _norm_text(ws.cell(row=row_idx, column=desc_col).value)
        ]
        if data_rows:
            total_rows += len(data_rows)
            sheet_jobs.append((ws, header_row, desc_col, data_rows))

    processed = 0
    for ws, header_row, desc_col, data_rows in sheet_jobs:
        output_col = _next_output_column(ws, header_row)
        ws.cell(row=header_row, column=output_col, value=PRICE_OUTPUT_HEADER)
        ws.cell(row=header_row, column=output_col).font = Font(bold=True)

        sheet_results = calculate_sheet_rows(ws, desc_col, data_rows, rules)
        for result in sheet_results:
            if result.price is not None:
                ws.cell(row=result.row_idx, column=output_col, value=result.price)
            else:
                ws.cell(row=result.row_idx, column=output_col, value="")
            processed += 1
            if job_id:
                append_job_log(
                    job_id,
                    f"{ws.title} 第 {result.row_idx} 行{result.status}：{result.price if result.price is not None else result.note}",
                    success_count=sum(1 for item in all_results + sheet_results[: sheet_results.index(result) + 1] if item.status == "成功"),
                    fail_count=sum(1 for item in all_results + sheet_results[: sheet_results.index(result) + 1] if item.status == "失败"),
                    current_row=processed,
                    total_rows=total_rows,
                )
        all_results.extend(sheet_results)

    save_bomin_note_sheet(workbook, all_results)
    return all_results


def calculate_sheet_rows(ws, desc_col: int, data_rows: list[int], rules: RuleBook) -> list[CalcResult]:
    parsed_by_row: dict[int, ParsedSpec] = {}
    for row_idx in data_rows:
        desc = str(ws.cell(row=row_idx, column=desc_col).value or "")
        parsed_by_row[row_idx] = parse_spec(desc, row_idx)

    pair_overrides = _calculate_one_to_many_overrides(parsed_by_row, rules)
    results: list[CalcResult] = []
    for row_idx in data_rows:
        if row_idx in pair_overrides:
            results.append(pair_overrides[row_idx])
            continue
        parsed = parsed_by_row[row_idx]
        result = calculate_single(parsed, rules)
        results.append(result)
    return results


def parse_spec(desc: str, row_idx: int = 0) -> ParsedSpec:
    text = _norm_text(desc)
    product = _extract_product(text)
    material_type = "PP" if _looks_like_pp(text, product) else "CCL"
    parsed = ParsedSpec(
        row_idx=row_idx,
        desc=desc,
        material_type=material_type,
        product=product,
        halogen=_extract_halogen(text),
        tg=_extract_tg(text),
    )

    if material_type == "PP":
        parsed.glass = _extract_pp_glass(text)
        parsed.rc = _extract_rc(text)
        parsed.length_value, parsed.length_unit, parsed.width_value, parsed.width_unit = _extract_pp_size(text)
        return parsed

    parsed.thickness = _extract_thickness(text)
    parsed.copper = _extract_copper(text)
    parsed.foil = _extract_foil(text)
    if not parsed.foil:
        parsed.foil = "不连铜" if "不连铜" in text else "HTE"
    parsed.copper_state = "含铜" if "含铜" in text and "不含铜" not in text else "不含铜"
    parsed.stack = _canonical_stack(_extract_stack(text))
    parsed.length_value, parsed.length_unit, parsed.width_value, parsed.width_unit = _extract_jing_wei_size(text)
    return parsed


def calculate_single(parsed: ParsedSpec, rules: RuleBook) -> CalcResult:
    try:
        if parsed.material_type == "PP":
            return _calculate_pp(parsed, rules)
        return _calculate_ccl(parsed, rules)
    except Exception as exc:
        return CalcResult(parsed.row_idx, parsed.desc, "失败", None, str(exc))


def _calculate_ccl(parsed: ParsedSpec, rules: RuleBook) -> CalcResult:
    row = _match_ccl_row(parsed, rules)
    match_note = row.values.pop("_match_note", "")
    if parsed.length_value is None or parsed.width_value is None:
        return CalcResult(parsed.row_idx, parsed.desc, "失败", None, "无法解析 CCL 尺寸")
    length_in, width_in = _size_to_inches(parsed.length_value, parsed.length_unit, parsed.width_value, parsed.width_unit)
    col = _standard_ccl_col(length_in, width_in)
    if col:
        price = row.values.get(col)
        if price is None:
            return CalcResult(parsed.row_idx, parsed.desc, "失败", None, f"找到报价行，但对应价格字段为空：尺寸列 {col}")
        final = _round_price(price)
        return CalcResult(parsed.row_idx, parsed.desc, "成功", final, f"CCL 大板：报价表第 {row.excel_row} 行，列 {col}，价格 {price}{match_note}")

    parent = _select_ccl_parent(length_in, width_in)
    if not parent:
        return CalcResult(parsed.row_idx, parsed.desc, "失败", None, f"需要小片换算，但无法判断父级大板尺寸：{length_in:.2f}x{width_in:.2f} inch")
    price = row.values.get(parent["col"])
    if price is None:
        return CalcResult(parsed.row_idx, parsed.desc, "失败", None, f"找到报价行，但对应价格字段为空：尺寸列 {parent['col']}")
    final = _round_price(price * parent["multiplier"] / parent["opens"])
    note = (
        f"CCL 小片：报价表第 {row.excel_row} 行，父级 {parent['parent_w']}x{parent['parent_h']}，"
        f"列 {parent['col']}，开数 {parent['opens']}，倍率 {parent['multiplier']}，价格 {price}{match_note}"
    )
    return CalcResult(parsed.row_idx, parsed.desc, "成功", final, note)


def _calculate_pp(parsed: ParsedSpec, rules: RuleBook) -> CalcResult:
    row = _match_pp_row(parsed, rules)
    match_note = row.values.pop("_match_note", "")
    per_m = row.values.get("per_m")
    if per_m is None:
        return CalcResult(parsed.row_idx, parsed.desc, "失败", None, "找到报价行，但对应价格字段为空：Per M")
    if parsed.length_unit == "m":
        final = _round_price(per_m)
        return CalcResult(parsed.row_idx, parsed.desc, "成功", final, f"PP 整卷：报价表第 {row.excel_row} 行，取 Per M={per_m}{match_note}")
    if parsed.length_value is None or parsed.width_value is None:
        return CalcResult(parsed.row_idx, parsed.desc, "失败", None, "尺寸无法解析")
    length_m = _length_to_m(parsed.length_value, parsed.length_unit)
    width_in = _one_axis_to_inch(parsed.width_value, parsed.width_unit)
    split = math.floor(49.5 / width_in) if width_in else 0
    if split <= 0:
        return CalcResult(parsed.row_idx, parsed.desc, "失败", None, f"PP 纬向宽幅无法一开：{width_in:.2f} inch")
    final = _round_price(length_m / split * per_m)
    note = f"PP 小片：报价表第 {row.excel_row} 行，Per M={per_m}，经向 {length_m:.3f}m，纬向一开 {split}{match_note}"
    return CalcResult(parsed.row_idx, parsed.desc, "成功", final, note)


def _calculate_one_to_many_overrides(parsed_by_row: dict[int, ParsedSpec], rules: RuleBook) -> dict[int, CalcResult]:
    overrides: dict[int, CalcResult] = {}
    for material_type in ("PP", "CCL"):
        candidates = sorted(
            [item for item in parsed_by_row.values() if item.material_type == material_type and item.row_idx not in overrides],
            key=lambda item: item.row_idx,
        )
        for index, item in enumerate(candidates):
            if item.row_idx in overrides or item.length_value is None or item.width_value is None:
                continue
            key = _pair_key(item)
            if not key:
                continue

            group = [item]
            for other in candidates[index + 1 : index + 5]:
                if other.row_idx in overrides:
                    continue
                if other.row_idx - item.row_idx >= 5:
                    break
                if _pair_key(other) != key or not _same_length_for_pair(item, other):
                    break
                group.append(other)

            if len(group) < 2:
                continue
            total_width = sum(_one_axis_to_inch(member.width_value or 0, member.width_unit) for member in group)
            limit = 49.5 if material_type == "PP" else 49.0
            if total_width <= 0 or total_width > limit + 1e-9:
                continue
            try:
                if material_type == "PP":
                    pair_results = _calculate_pp_group(group, rules)
                else:
                    pair_results = _calculate_ccl_group(group, rules)
                overrides.update(pair_results)
            except Exception:
                continue
    return overrides


def _calculate_ccl_group(group: list[ParsedSpec], rules: RuleBook) -> dict[int, CalcResult]:
    anchor = group[0]
    row = _match_ccl_row(anchor, rules)
    length_in = _one_axis_to_inch(anchor.length_value or 0, anchor.length_unit)
    total_width = sum(_one_axis_to_inch(item.width_value or 0, item.width_unit) for item in group)
    parent = _select_ccl_parent(length_in, total_width)
    if not parent:
        raise ValueError("疑似一对多小片，但组合不完整")
    base_price = row.values.get(parent["col"])
    if base_price is None:
        raise ValueError(f"找到报价行，但对应价格字段为空：尺寸列 {parent['col']}")
    pool = base_price * parent["multiplier"] / parent["opens"]
    results = {}
    for item in group:
        width_in = _one_axis_to_inch(item.width_value or 0, item.width_unit)
        price = _round_price(width_in / total_width * pool)
        note = (
            f"CCL 一对多：报价表第 {row.excel_row} 行，列 {parent['col']}，"
            f"父级 {parent['parent_w']}x{parent['parent_h']}，开数 {parent['opens']}，"
            f"{len(group)} 行组合，分摊池 {pool:.6f}"
        )
        results[item.row_idx] = CalcResult(item.row_idx, item.desc, "成功", price, note)
    return results


def _calculate_pp_group(group: list[ParsedSpec], rules: RuleBook) -> dict[int, CalcResult]:
    anchor = group[0]
    row = _match_pp_row(anchor, rules)
    per_m = row.values.get("per_m")
    if per_m is None:
        raise ValueError("找到报价行，但对应价格字段为空：Per M")
    length_m = _length_to_m(anchor.length_value or 0, anchor.length_unit)
    total_width = sum(_one_axis_to_inch(item.width_value or 0, item.width_unit) for item in group)
    results = {}
    for item in group:
        width_in = _one_axis_to_inch(item.width_value or 0, item.width_unit)
        price = _round_price(width_in / total_width * length_m * per_m)
        note = f"PP 一对多：报价表第 {row.excel_row} 行，Per M={per_m}，经向 {length_m:.3f}m，{len(group)} 行按纬向占比分摊"
        results[item.row_idx] = CalcResult(item.row_idx, item.desc, "成功", price, note)
    return results


def _match_ccl_row(parsed: ParsedSpec, rules: RuleBook) -> PriceRow:
    if not parsed.product:
        raise ValueError("胶系未找到：无法解析胶系")
    if parsed.foil == "不连铜":
        raise ValueError("不连铜对应报价规则未找到")

    product_rows = [row for row in rules.ccl_rows if row.values["product"] == parsed.product]
    if not product_rows:
        raise ValueError(f"胶系未找到：产品={parsed.product}")
    if parsed.thickness is None:
        raise ValueError("胶系找到，但厚度未找到：无法解析厚度")
    thickness_rows = [row for row in product_rows if _float_equal(row.values["thickness"], parsed.thickness, 0.0005)]
    if not thickness_rows:
        raise ValueError(f"胶系找到，但厚度未找到：产品={parsed.product}, 厚度={parsed.thickness}")
    if not parsed.copper:
        raise ValueError("胶系和厚度找到，但铜箔规格未找到：无法解析铜厚")
    copper_rows = [row for row in thickness_rows if row.values["copper"] == parsed.copper]
    if not copper_rows:
        raise ValueError(f"胶系和厚度找到，但铜箔规格未找到：铜厚={parsed.copper}")

    foil_rows = [row for row in copper_rows if _foil_compatible(row.values["foil"], parsed.foil)]
    if not foil_rows:
        raise ValueError(f"胶系和厚度找到，但铜箔规格未找到：铜箔类型={parsed.foil}")

    stack_rows = [row for row in foil_rows if parsed.stack and row.values["stack"] == parsed.stack]
    if parsed.stack and stack_rows:
        candidates = stack_rows
    elif parsed.stack:
        raise ValueError(f"胶系、厚度、铜箔找到，但玻布结构未找到：叠构={parsed.stack}")
    else:
        raise ValueError("胶系、厚度、铜箔找到，但玻布结构未找到：无法解析叠构")

    scored = sorted(
        candidates,
        key=lambda row: (
            -_score_ccl_candidate(row, parsed),
            -_field_completeness(row),
            row.excel_row,
        ),
    )
    if not scored:
        raise ValueError("多候选无法确定")
    best_score = _score_ccl_candidate(scored[0], parsed)
    top = [row for row in scored if _score_ccl_candidate(row, parsed) == best_score]
    chosen = sorted(top, key=lambda row: (-_field_completeness(row), row.excel_row))[0]
    if len(top) > 1:
        chosen.values["_match_note"] = f"；多候选，取字段最完整的一行：报价表第 {chosen.excel_row} 行"
    else:
        chosen.values.pop("_match_note", None)
    return chosen


def _match_pp_row(parsed: ParsedSpec, rules: RuleBook) -> PriceRow:
    if not parsed.product:
        raise ValueError("胶系未找到：无法解析 PP 胶系")
    product_rows = [row for row in rules.pp_rows if row.values["product"] in _pp_product_aliases(parsed.product)]
    if not product_rows:
        raise ValueError(f"胶系未找到：产品={parsed.product}")
    if not parsed.glass:
        raise ValueError("PP 胶系找到，但玻布型号未找到：无法解析玻布型号")
    glass_rows = [row for row in product_rows if row.values["glass"] == parsed.glass]
    if not glass_rows:
        raise ValueError(f"PP 胶系找到，但玻布型号未找到：玻布={parsed.glass}")
    if parsed.rc is None:
        raise ValueError("PP 胶系和玻布找到，但 RC 未找到：无法解析 RC")
    candidates = [row for row in glass_rows if _float_equal(row.values["rc"], parsed.rc, 0.0005)]
    if not candidates:
        raise ValueError(f"PP 胶系和玻布找到，但 RC 未找到：RC={parsed.rc}")

    scored = sorted(
        candidates,
        key=lambda row: (
            -_score_pp_candidate(row, parsed),
            -_field_completeness(row),
            row.excel_row,
        ),
    )
    chosen = scored[0]
    best_score = _score_pp_candidate(chosen, parsed)
    top = [row for row in scored if _score_pp_candidate(row, parsed) == best_score]
    if len(top) > 1:
        chosen = sorted(top, key=lambda row: (-_field_completeness(row), row.excel_row))[0]
        chosen.values["_match_note"] = f"；多候选，取字段最完整的一行：报价表第 {chosen.excel_row} 行"
    else:
        chosen.values.pop("_match_note", None)
    return chosen


def save_bomin_note_sheet(workbook: Workbook, results: list[CalcResult]) -> None:
    if NOTE_SHEET_NAME in workbook.sheetnames:
        del workbook[NOTE_SHEET_NAME]
    ws = workbook.create_sheet(NOTE_SHEET_NAME)
    headers = ["行号", "物料描述/客户规格", "状态", "博敏计算价格", "说明"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for item in results:
        ws.append([item.row_idx, item.desc, item.status, item.price if item.price is not None else "", item.note])
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 110


def _find_desc_column(ws) -> tuple[int | None, int | None]:
    for row in range(1, min(ws.max_row, 30) + 1):
        for col in range(1, ws.max_column + 1):
            value = _norm_text(ws.cell(row=row, column=col).value)
            if value in DESC_HEADERS:
                return row, col
    return None, None


def _next_output_column(ws, header_row: int) -> int:
    max_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                max_col = max(max_col, cell.column)
    return max_col + 1


def _extract_product(text: str) -> str | None:
    upper = text.upper().replace("NY3150HF P", "NY3150HFP")
    match = re.search(r"\b(NY[-A-Z0-9()]+)\b", upper)
    return _norm_product(match.group(1)) if match else None


def _extract_thickness(text: str) -> float | None:
    match = re.search(r"\bNY[-A-Z0-9()]+\s+(\d+(?:\.\d+)?)\s*MM\b", text.upper())
    return float(match.group(1)) if match else None


def _extract_copper(text: str) -> str | None:
    for match in re.finditer(r"\b([A-Z0-9.]+/[A-Z0-9.]+)\s*(?:OZ)?\b", text.upper()):
        token = _norm_copper(match.group(1))
        if _foil_base(token.split("/", 1)[0]) in {"HTE", "RTF", "HVLP", "VLP"}:
            continue
        return {"1.5/1.5": "F/F"}.get(token, token)
    return None


def _extract_foil(text: str) -> str | None:
    match = re.search(
        r"\b(HTE2?|RTF[0-9]?|HVLP[0-9]?|VLP[0-9]?)(?:\([^)]*\))?\s*/\s*(HTE2?|RTF[0-9]?|HVLP[0-9]?|VLP[0-9]?)(?:\([^)]*\))?\b",
        text.upper(),
    )
    return _norm_token(match.group(1)) if match else None


def _extract_stack(text: str) -> str | None:
    patterns = [
        r"((?:\d{3,4}\s*[xX×*]\s*\d+|\d+\s*[xX×*]\s*\d{3,4}|\d+\s*张\s*\d{3,4}|\d{3,4}\s*\d+\s*张)(?:\s*\+\s*(?:\d{3,4}\s*[xX×*]\s*\d+|\d+\s*[xX×*]\s*\d{3,4}|\d+\s*张\s*\d{3,4}|\d{3,4}\s*\d+\s*张))*)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    match = re.search(r"\b(106|1080|2116|7628|1506|3313|1035|1078|2113)\b", text)
    if match:
        return f"1*{match.group(1)}"
    return None


def _extract_pp_glass(text: str) -> str | None:
    upper = text.upper()
    product = _extract_product(upper)
    if product:
        tail = upper.split(product, 1)[-1]
        match = re.search(r"\b(\d{2,4})\b(?=.*\bRC\b)", tail)
        if match:
            return _norm_glass(match.group(1))
    match = re.search(r"\b(\d{2,4})\b(?=.*\bRC\b)", upper)
    return _norm_glass(match.group(1)) if match else None


def _extract_rc(text: str) -> float | None:
    upper = text.upper()
    patterns = [
        r"\bRC\s*=?\s*(\d+(?:\.\d+)?)\s*%",
        r"\bRC\s+(\d+(?:\.\d+)?)\b",
        r"含胶量\s*(\d+(?:\.\d+)?)\s*%",
        r"\b(0\.\d+)\b",
        r"\b(\d+(?:\.\d+)?)\s*%",
    ]
    for pattern in patterns:
        match = re.search(pattern, upper)
        if match:
            number = float(match.group(1))
            return number / 100 if number > 1 else number
    return None


def _extract_pp_size(text: str) -> tuple[float | None, str | None, float | None, str | None]:
    length, length_unit, width, width_unit = _extract_jing_wei_size(text)
    if length is not None and width is not None:
        return length, length_unit, width, width_unit

    match = re.search(
        r"(\d+(?:\.\d+)?)\s*(?:\"|IN|INCH|英寸)\s*\*?\s*(\d+(?:\.\d+)?)\s*M\b",
        text.upper(),
        re.IGNORECASE,
    )
    if match:
        return float(match.group(2)), "m", float(match.group(1)), "inch"
    return None, None, None, None


def _extract_jing_wei_size(text: str) -> tuple[float | None, str | None, float | None, str | None]:
    match = re.search(
        r"经\s*(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸|\")?\s*\*?\s*纬\s*(\d+(?:\.\d+)?)\s*(MM|M|IN|INCH|英寸|\")?",
        text,
        re.IGNORECASE,
    )
    if not match:
        return None, None, None, None
    length = float(match.group(1))
    width = float(match.group(3))
    length_unit = _normalize_unit(match.group(2), length)
    width_unit = _normalize_unit(match.group(4), width)
    if not match.group(2) and match.group(4) and width > 100:
        length_unit = "mm"
        width_unit = "mm"
    return length, length_unit, width, width_unit


def _normalize_unit(unit: str | None, value: float) -> str:
    if not unit:
        return "mm" if value > 100 else "inch"
    unit = unit.upper()
    if unit == "M":
        return "m"
    if unit in {"IN", "INCH", "英寸", '"'}:
        return "mm" if value > 100 else "inch"
    return "mm"


def _looks_like_pp(text: str, product: str | None) -> bool:
    upper = text.upper()
    has_pp_glass = bool(re.search(r"\b\d{2,4}\b(?=.*\bRC\b)", upper))
    if re.search(r"\bRC\b|含胶量", upper) and has_pp_glass:
        return True
    if product and re.search(r"P$", product):
        return True
    return False


def _extract_halogen(text: str) -> str | None:
    if "无卤" in text:
        return "无卤"
    if "有卤" in text:
        return "有卤"
    return None


def _extract_tg(text: str) -> int | None:
    match = re.search(r"TG\s*(?:≥|>=|>|=)?\s*(\d{3})", text.upper())
    return int(match.group(1)) if match else None


def _pp_product_aliases(product: str) -> set[str]:
    aliases = {product}
    if product.endswith("P"):
        aliases.add(product[:-1])
    else:
        aliases.add(f"{product}P")
    if product.endswith("MP"):
        aliases.add(product[:-1])
    if product.endswith("HFP"):
        aliases.add(product[:-1])
    return aliases


def _foil_compatible(rule_foil: str | None, spec_foil: str | None) -> bool:
    if not rule_foil or not spec_foil:
        return False
    if rule_foil == spec_foil:
        return True
    if spec_foil == "HTE" and rule_foil == "HTE":
        return True
    return _foil_base(rule_foil) == _foil_base(spec_foil)


def _foil_base(value: str) -> str:
    return re.sub(r"\d+$", "", _norm_token(value))


def _score_ccl_candidate(row: PriceRow, parsed: ParsedSpec) -> int:
    values = row.values
    score = 0
    score += 100 if values.get("product") == parsed.product else 0
    score += 80 if _float_equal(values.get("thickness"), parsed.thickness, 0.0005) else 0
    score += 60 if values.get("copper") == parsed.copper else 0
    score += 40 if values.get("foil") == parsed.foil else 20 if _foil_compatible(values.get("foil"), parsed.foil) else 0
    score += 35 if parsed.stack and values.get("stack") == parsed.stack else 0
    score += 10 if parsed.copper_state and values.get("copper_state") == parsed.copper_state else 0
    return score


def _score_pp_candidate(row: PriceRow, parsed: ParsedSpec) -> int:
    values = row.values
    score = 100 if values.get("product") == parsed.product else 70
    score += 60 if values.get("glass") == parsed.glass else 0
    score += 50 if _float_equal(values.get("rc"), parsed.rc, 0.0005) else 0
    score += 10 if values.get("per_m") is not None else 0
    return score


def _field_completeness(row: PriceRow) -> int:
    return sum(1 for value in row.values.values() if value not in (None, ""))


def _pair_key(item: ParsedSpec) -> tuple | None:
    if item.material_type == "PP":
        if not item.product or not item.glass or item.rc is None:
            return None
        return ("PP", item.product, item.glass, round(item.rc, 4), _length_to_m(item.length_value or 0, item.length_unit))
    if not item.product or item.thickness is None or not item.copper or not item.foil:
        return None
    return (
        "CCL",
        item.product,
        round(item.thickness, 4),
        item.copper,
        item.copper_state,
        item.foil,
        item.stack,
        round(_one_axis_to_inch(item.length_value or 0, item.length_unit), 3),
    )


def _same_length_for_pair(a: ParsedSpec, b: ParsedSpec) -> bool:
    return abs(_one_axis_to_inch(a.length_value or 0, a.length_unit) - _one_axis_to_inch(b.length_value or 0, b.length_unit)) <= 0.05


def _standard_ccl_col(length_in: float, width_in: float) -> str | None:
    candidates = [
        (37, 49, '37"X49"'),
        (41, 49, '41"X49"'),
        (43, 49, '43"X49"'),
        (37, 43, "37*43"),
        (41, 43, "41*43"),
    ]
    for length, width, col in candidates:
        if abs(length_in - length) <= 0.35 and abs(width_in - width) <= 0.35:
            return col
    return None


def _select_ccl_parent(length_in: float, width_in: float) -> dict | None:
    candidates = [
        (37, 49, '37"X49"', 1),
        (41, 49, '41"X49"', 1),
        (43, 49, '43"X49"', 1),
        (74, 49, '37"X49"', 2),
        (82, 49, '41"X49"', 2),
        (86, 49, '43"X49"', 2),
    ]
    valid = []
    for parent_w, parent_h, col, multiplier in candidates:
        opens_w = math.floor((parent_w + 1e-9) / length_in) if length_in else 0
        opens_h = math.floor((parent_h + 1e-9) / width_in) if width_in else 0
        opens = opens_w * opens_h
        if opens <= 0:
            continue
        effective = opens / multiplier
        valid.append(
            {
                "parent_w": parent_w,
                "parent_h": parent_h,
                "col": col,
                "multiplier": multiplier,
                "opens": opens,
                "effective": effective,
            }
        )
    if not valid:
        return None
    return sorted(valid, key=lambda item: (-item["effective"], item["parent_w"], item["multiplier"]))[0]


def _size_to_inches(length: float, length_unit: str | None, width: float, width_unit: str | None) -> tuple[float, float]:
    return _one_axis_to_inch(length, length_unit), _one_axis_to_inch(width, width_unit)


def _one_axis_to_inch(value: float, unit: str | None) -> float:
    if unit == "mm":
        return value / 25.4
    if unit == "m":
        return value * 1000 / 25.4
    return value


def _length_to_m(value: float, unit: str | None) -> float:
    if unit == "mm":
        return value / 1000
    if unit == "inch":
        return value * 25.4 / 1000
    return value


def _canonical_stack(value: Any) -> str | None:
    text = _norm_text(value).upper().replace("×", "X")
    if not text:
        return None
    parts = []
    for part in re.split(r"\+", text):
        part = part.strip()
        match = re.fullmatch(r"(\d+)\s*张\s*(\d{3,4})", part)
        if match:
            parts.append(f"{int(match.group(1))}*{match.group(2)}")
            continue
        match = re.fullmatch(r"(\d{3,4})\s*(\d+)\s*张", part)
        if match:
            parts.append(f"{int(match.group(2))}*{match.group(1)}")
            continue
        match = re.fullmatch(r"(\d+)\s*[*X]\s*(\d{3,4})", part)
        if match:
            parts.append(f"{int(match.group(1))}*{match.group(2)}")
            continue
        match = re.fullmatch(r"(\d{3,4})\s*[*X]\s*(\d+)", part)
        if match:
            parts.append(f"{int(match.group(2))}*{match.group(1)}")
            continue
        parts.append(part.replace(" ", ""))
    return "+".join(sorted(parts))


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _norm_token(value: Any) -> str:
    return _norm_text(value).upper().replace(" ", "")


def _norm_copper(value: Any) -> str:
    text = _norm_token(value).replace("OZ", "")
    return {"1.5/1.5": "F/F"}.get(text, text)


def _norm_product(value: Any) -> str:
    return _norm_token(value).replace("－", "-")


def _norm_glass(value: Any) -> str:
    text = _norm_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text.lstrip("0") or text


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _norm_text(value).replace(",", "")
    if text.endswith("%"):
        return float(text[:-1]) / 100
    try:
        return float(text)
    except ValueError:
        return None


def _to_rc(value: Any) -> float | None:
    number = _to_float(value)
    if number is None:
        return None
    return number / 100 if number > 1 else number


def _float_equal(a: float | None, b: float | None, tolerance: float = 0.0001) -> bool:
    return a is not None and b is not None and abs(float(a) - float(b)) <= tolerance


def _round_price(value: float) -> float:
    return round(float(value) + 1e-9, 2)
