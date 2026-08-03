from __future__ import annotations

from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook


STANDARD_RULE_PATH = (
    Path(__file__).resolve().parent
    / "default_rules"
    / "transcode_agent"
    / "base_code_standard.xlsx"
)


def _enabled(value: object) -> bool:
    return str(value or "").strip().lower() in {"是", "y", "yes", "true", "1", "启用"}


def _rows_by_header(sheet) -> list[dict[str, object]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]


@lru_cache(maxsize=1)
def load_base_code_standard(path: str | Path | None = None) -> dict[str, object]:
    source = Path(path) if path else STANDARD_RULE_PATH
    if not source.exists():
        raise FileNotFoundError(f"营销转码Agent基础编码规范映射不存在：{source}")
    workbook = load_workbook(source, read_only=True, data_only=True)
    required = {"高频高速MIL换算", "标准毫米尺寸", "合法基板级别", "版本说明"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"基础编码规范映射缺少Sheet：{', '.join(sorted(missing))}")

    mil_map: dict[float, float] = {}
    for row in _rows_by_header(workbook["高频高速MIL换算"]):
        if _enabled(row.get("启用")):
            mil_map[float(row["mil"])] = float(row["厚度mm"])

    size_map: dict[float, float] = {}
    for row in _rows_by_header(workbook["标准毫米尺寸"]):
        if _enabled(row.get("启用")):
            size_map[float(row["毫米值"])] = float(row["英寸值"])

    grade_codes = {
        str(row.get("基板级别代码") or "").strip().upper()
        for row in _rows_by_header(workbook["合法基板级别"])
        if _enabled(row.get("启用")) and str(row.get("基板级别代码") or "").strip()
    }
    if not mil_map or not size_map or not grade_codes:
        raise ValueError("基础编码规范映射存在空的启用规则集")
    return {
        "path": source,
        "high_speed_mil_to_mm": mil_map,
        "standard_mm_size_aliases": size_map,
        "grade_codes": grade_codes,
    }


BASE_CODE_STANDARD = load_base_code_standard()
HIGH_SPEED_MIL_TO_MM = BASE_CODE_STANDARD["high_speed_mil_to_mm"]
STANDARD_MM_SIZE_ALIASES = BASE_CODE_STANDARD["standard_mm_size_aliases"]
OFFICIAL_GRADE_CODES = BASE_CODE_STANDARD["grade_codes"]
