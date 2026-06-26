from __future__ import annotations

import math
import re
import traceback
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from functools import lru_cache
from pathlib import Path
from typing import Any, Optional

from openpyxl.styles import Font, PatternFill
from werkzeug.utils import secure_filename

from .db import append_job_log, create_job, prune_jobs_for_employee, update_job_status
from .excel_utils import (
    is_ole_workbook as excel_is_ole_workbook,
    load_workbook_compat,
    load_xls_as_workbook as excel_load_xls_as_workbook,
)
from .file_utils import safe_unlink
from .hushi_rules import get_active_hushi_rule_version, get_hushi_rule_dir
from .job_control import launch_job_process
from .paths import JOBS_DIR


ALLOWED_INPUT_EXTENSIONS = (".xlsx", ".xlsm", ".xls")
NON_DATA_DESCRIPTIONS = {"物料描述", "规格", "客户规格", "材料描述", "物料规格"}


@dataclass
class HushiMatchResult:
    price: Optional[float]
    final_price: Optional[float]
    status: str
    reason: str
    rule_file: str = ""
    sheet_name: str = ""
    excel_row: Optional[int] = None
    product: str = ""
    material_type: str = ""
    area_step: Optional[float] = None
    width_inch_floor: Optional[float] = None
    height_inch_floor: Optional[float] = None
    normal_price: Any = None
    rebate_price: Any = None
    used_price_column: str = ""


def queue_hushi_job(employee_id: str, uploaded_file, source_filename: str) -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)

    safe_filename = secure_filename(source_filename) or f"hushi_{timestamp}.xlsx"
    input_path = employee_dir / f"{timestamp}_hushi_{safe_filename}"
    uploaded_file.save(input_path)

    rule_version = get_active_hushi_rule_version()
    job_id = create_job(
        employee_id,
        source_filename,
        str(input_path),
        rule_version or "未上传沪士规则",
        feature="hushi",
    )
    launch_job_process(job_id, "hushi", employee_id)
    return job_id


def calculate_hushi_quote(spec: str) -> dict:
    spec = normalize_text(spec)
    if not spec:
        return {"status": "失败", "price": None, "error": "请输入客户规格"}

    rule_version = get_active_hushi_rule_version()
    if not rule_version:
        return {"status": "失败", "price": None, "error": "当前没有沪士报价规则，请先上传规则包"}

    result = calculate_hushi_spec(spec, get_hushi_rule_dir(rule_version), rule_cache={})
    note_parts = [result.reason]
    if result.rule_file:
        note_parts.append(f"报价文件：{result.rule_file}")
    if result.sheet_name:
        note_parts.append(f"Sheet：{result.sheet_name}")
    if result.excel_row:
        note_parts.append(f"命中行：{result.excel_row}")
    note = " | ".join(part for part in note_parts if part)

    if result.final_price is None:
        return {
            "status": "失败",
            "price": None,
            "note": note,
            "material_type": result.material_type or "沪士价格",
            "rule_version": rule_version,
            "error": result.reason,
        }
    return {
        "status": "成功",
        "price": round(float(result.final_price), 2),
        "note": note or "计算成功",
        "material_type": result.material_type or "沪士价格",
        "rule_version": rule_version,
        "error": "",
    }


def run_hushi_job(job_id: int, employee_id: str) -> None:
    from .db import get_job

    update_job_status(job_id, status="running", log_text="")
    job = get_job(job_id)
    if not job:
        return

    try:
        rule_version = get_active_hushi_rule_version()
        if not rule_version:
            raise RuntimeError("当前没有沪士报价规则，请先在沪士规则管理上传 ZIP 规则包")
        rule_dir = get_hushi_rule_dir(rule_version)
        append_job_log(job_id, f"开始处理沪士价格计算任务，规则版本：{rule_version}")
        append_job_log(job_id, f"规则目录加载完成：{rule_dir}")

        output_path, total_rows, success_count, fail_count, skip_count = calculate_hushi_workbook(
            Path(job["stored_input_path"]),
            rule_dir,
            job_id=job_id,
        )
        append_job_log(job_id, "结果文件已生成，任务完成", current_row=total_rows, total_rows=total_rows)
        update_job_status(
            job_id,
            status="completed",
            stored_result_path=str(output_path),
            success_count=success_count,
            fail_count=fail_count,
            skip_count=skip_count,
            current_row=total_rows,
            total_rows=total_rows,
            completed=True,
        )
    except Exception as exc:
        append_job_log(job_id, f"任务失败：{exc}")
        update_job_status(
            job_id,
            status="failed",
            error_message=f"{exc}\n{traceback.format_exc(limit=8)}",
            completed=True,
        )
    finally:
        stale_jobs = prune_jobs_for_employee(employee_id, keep_limit=500)
        for stale in stale_jobs:
            for key in ["stored_input_path", "stored_result_path"]:
                safe_unlink(stale[key])


def calculate_hushi_workbook(input_path: Path, rule_dir: Path, *, job_id: int | None = None):
    if job_id is not None:
        append_job_log(job_id, f"开始读取上传文件：{input_path.name}")
    wb = load_input_workbook(input_path)
    ws = wb.worksheets[0]
    if job_id is not None:
        append_job_log(job_id, f"使用第一个工作表：{ws.title}")
        append_job_log(job_id, "已启用沪士报价文件缓存")
    rule_cache: dict[str, Optional[Path]] = {}
    desc_col = detect_description_column(ws)
    output_col = last_non_empty_column(ws) + 1
    ws.cell(row=1, column=output_col, value="沪士计算价格")

    note_sheet_name = "沪士计算说明"
    if note_sheet_name in wb.sheetnames:
        del wb[note_sheet_name]
    note = wb.create_sheet(note_sheet_name)
    headers = [
        "行号",
        "规格",
        "状态",
        "最终价格",
        "单价",
        "价格列",
        "面积",
        "宽inch取档",
        "高inch取档",
        "报价文件",
        "报价Sheet",
        "报价行",
        "类型",
        "原因",
    ]
    note.append(headers)
    for cell in note[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="185C43")

    data_rows = [
        row_idx
        for row_idx in range(1, ws.max_row + 1)
        if is_effective_description(ws.cell(row=row_idx, column=desc_col).value)
    ]
    total_rows = len(data_rows)
    if job_id is not None:
        update_job_status(job_id, status="running", total_rows=total_rows)
        append_job_log(job_id, f"识别到规格列：第 {desc_col} 列，共 {total_rows} 行待处理", total_rows=total_rows)

    success_count = 0
    fail_count = 0
    skip_count = 0
    for processed, row_idx in enumerate(data_rows, start=1):
        spec = ws.cell(row=row_idx, column=desc_col).value
        if not normalize_text(spec):
            skip_count += 1
            continue

        result = calculate_hushi_spec(str(spec), rule_dir, rule_cache=rule_cache)
        if result.final_price is not None:
            ws.cell(row=row_idx, column=output_col, value=result.final_price)
            success_count += 1
            log_msg = f"第 {row_idx} 行成功：{result.final_price}"
        else:
            ws.cell(row=row_idx, column=output_col, value=result.reason)
            fail_count += 1
            log_msg = f"第 {row_idx} 行失败：{result.reason}"

        note.append(
            [
                row_idx,
                normalize_text(spec),
                result.status,
                result.final_price,
                result.price,
                result.used_price_column,
                result.area_step,
                result.width_inch_floor,
                result.height_inch_floor,
                result.rule_file,
                result.sheet_name,
                result.excel_row,
                result.material_type,
                result.reason,
            ]
        )
        if job_id is not None:
            append_job_log(
                job_id,
                log_msg,
                success_count=success_count,
                fail_count=fail_count,
                skip_count=skip_count,
                current_row=processed,
                total_rows=total_rows,
            )

    output_path = input_path.with_name(f"{input_path.stem}_沪士计算结果.xlsx")
    wb.save(output_path)
    return output_path, total_rows, success_count, fail_count, skip_count


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return (
        text.replace("\u3000", " ")
        .replace("\xa0", " ")
        .replace("（", "(")
        .replace("）", ")")
        .replace("×", "X")
        .replace("*", "X")
        .strip()
    )


def normalize_key(value: Any) -> str:
    text = normalize_text(value).upper()
    return (
        text.replace(" ", "")
        .replace("-", "")
        .replace("(", "")
        .replace(")", "")
        .replace("（", "")
        .replace("）", "")
    )


def quote_product_key(value: Any) -> str:
    text = normalize_text(value).upper()
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("P(C)", "(C)")
    if text.endswith("P") and text.startswith("NY"):
        text = text[:-1]
    return normalize_key(text)


def extract_product(spec: str) -> Optional[str]:
    text = normalize_text(spec).upper()
    match = re.search(r"\bNY[\w\-]*\s*(?:\([A-Z0-9]+\))?P?(?:\([A-Z0-9]+\))?", text)
    if not match:
        return None
    product = match.group(0).strip()
    product = product.replace("P(C)", "(C)")
    if product.endswith("P") and not product.endswith("(C)"):
        product = product[:-1]
    return product


def extract_pp_fields(spec: str) -> tuple[Optional[str], Optional[float]]:
    text = normalize_text(spec).upper()
    glass = None
    rc = None
    match = re.search(r"\b(0?106|1035|1037|1078|1080|1086|1506|2116|3313|7628)\b", text)
    if match:
        glass = match.group(1)
        if glass == "0106":
            glass = "106"
    match = re.search(r"RC\s*=?\s*(\d+(?:\.\d+)?)\s*%", text)
    if match:
        rc = float(match.group(1))
    return glass, rc


def extract_ccl_fields(spec: str) -> dict[str, Optional[str | float]]:
    text = normalize_text(spec).upper()
    thickness = None
    copper = None
    structure = None
    foil = None

    match = re.search(r"\b(\d{3,4})\s*\"", text)
    if match:
        thickness = int(match.group(1)) / 10.0
    else:
        match = re.search(r"\b(\d+(?:\.\d+)?)\s*MIL\b", text)
        if match:
            thickness = float(match.group(1))

    match = re.search(r"\b(H|1|2|3)/(H|1|2|3)\b", text)
    if match:
        copper = f"{match.group(1)}/{match.group(2)}"

    match = re.search(r"\b(1067|0106|106|1035|1037|1078|1080|1086|1506|2116|3313|7628)\s*X\s*(\d+)\b", text)
    if match:
        glass = match.group(1)
        if glass == "0106":
            glass = "106"
        structure = f"{glass}X{match.group(2)}"

    foil_tokens = []
    if "H-VLP" in text or "HVLP" in text:
        foil_tokens.append("HVLP")
    if "PVLP" in text:
        foil_tokens.append("PVLP")
    if "FVLP" in text:
        foil_tokens.append("FVLP")
    if "RTF3" in text:
        foil_tokens.append("RTF3")
    if "RTF2" in text:
        foil_tokens.append("RTF2")
    if "RTF" in text and "RTF3" not in text and "RTF2" not in text:
        foil_tokens.append("RTF")
    if foil_tokens:
        foil = "/".join(dict.fromkeys(foil_tokens))

    return {"thickness": thickness, "copper": copper, "structure": structure, "foil": foil}


def extract_size(spec: str) -> Optional[tuple[float, float]]:
    text = normalize_text(spec).upper()
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*X\s*(\d+(?:\.\d+)?)", text)
    if not matches:
        return None
    pairs = [(float(a), float(b)) for a, b in matches]
    return max(pairs, key=lambda item: item[0] * item[1])


def half_inch_floor(value: float) -> float:
    return math.floor(value * 2) / 2.0


def round_half_up(value: float, digits: int = 2) -> float:
    quant = Decimal("1").scaleb(-digits)
    return float(Decimal(str(value)).quantize(quant, rounding=ROUND_HALF_UP))


def area_step_from_size(width_mm: float, height_mm: float) -> tuple[float, float, float]:
    width_inch = half_inch_floor(width_mm / 25.4)
    height_inch = half_inch_floor(height_mm / 25.4)
    area = width_inch * height_inch / 144.0
    return width_inch, height_inch, round_half_up(area, 2)


def cell_has_strike(cell) -> bool:
    return bool(cell.font and cell.font.strike)


def numeric_value(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value):
            return None
        return float(value)
    text = normalize_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_rule_file(product: str, rule_dir: Path, rule_cache: dict[str, Optional[Path]] | None = None) -> Optional[Path]:
    pkey = quote_product_key(product)
    if rule_cache is not None and pkey in rule_cache:
        return rule_cache[pkey]

    files = [
        path
        for path in rule_dir.iterdir()
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xls", ".xlsm"}
    ]
    candidates = []
    for path in files:
        fkey = quote_product_key(path.stem)
        if pkey in fkey:
            candidates.append(path)
    if candidates:
        result = sorted(candidates, key=_rule_file_sort_key)[0]
        if rule_cache is not None:
            rule_cache[pkey] = result
        return result

    if rule_cache is not None:
        rule_cache[pkey] = None
    return None


def _rule_file_sort_key(path: Path):
    dates = re.findall(r"\d{6,8}", path.stem)
    date_score = int(dates[-1]) if dates else 0
    return (-date_score, len(path.name), path.name)


def rule_file_contains_product(path: Path, product_key: str) -> bool:
    try:
        wb = load_excel_workbook(path)
        for ws in wb.worksheets:
            for row, product_idx in iter_data_rows(ws):
                if quote_product_key(row[product_idx].value) == product_key:
                    return True
    except Exception:
        return False
    return False


def iter_data_rows(ws):
    for row in ws.iter_rows(min_row=1):
        values = [cell.value for cell in row]
        for idx, value in enumerate(values[:3]):
            if value and "NY" in normalize_text(value).upper():
                yield row, idx
                break


def pick_price(row, normal_idx: int, rebate_idx: int) -> tuple[Optional[float], str, Any, Any, str]:
    rebate_cell = row[rebate_idx] if rebate_idx < len(row) else None
    normal_cell = row[normal_idx] if normal_idx < len(row) else None
    rebate = numeric_value(rebate_cell.value if rebate_cell else None)
    normal = numeric_value(normal_cell.value if normal_cell else None)

    if rebate is not None and not cell_has_strike(rebate_cell):
        return rebate, "Rebate", normal, rebate, ""
    if rebate is not None and cell_has_strike(rebate_cell):
        strike_note = "Rebate 有值但被删除线划掉，跳过；"
    else:
        strike_note = ""
    if normal is not None and not cell_has_strike(normal_cell):
        return normal, "Normal", normal, rebate, strike_note
    if normal is not None and cell_has_strike(normal_cell):
        strike_note += "Normal 有值但被删除线划掉，跳过；"
    return None, "", normal, rebate, strike_note or "Normal/Rebate 均为空"


def text_match(expected: Optional[str], actual: Any) -> bool:
    if expected is None:
        return False
    return normalize_key(expected) == normalize_key(actual)


def number_match(expected: Optional[float], actual: Any) -> bool:
    if expected is None:
        return False
    actual_num = numeric_value(actual)
    return actual_num is not None and abs(actual_num - expected) < 1e-6


def match_pp(spec: str, product: str, rule_file: Path, size_info: tuple[float, float, float]) -> HushiMatchResult:
    glass, rc = extract_pp_fields(spec)
    if not glass or rc is None:
        return HushiMatchResult(None, None, "failed", f"无法提取 PP 玻纤或 RC：glass={glass}, rc={rc}", product=product, material_type="PP")

    wb = load_excel_workbook(rule_file)
    pp_sheets = [name for name in wb.sheetnames if "PP" in name.upper()]
    if not pp_sheets:
        return HushiMatchResult(None, None, "failed", "报价单中找不到 PP 报价页", rule_file=rule_file.name, product=product, material_type="PP")

    pkey = quote_product_key(product)
    for sheet_name in pp_sheets:
        ws = wb[sheet_name]
        for row, product_idx in iter_data_rows(ws):
            product_cell = row[product_idx]
            if quote_product_key(product_cell.value) != pkey:
                continue
            glass_cell = row[product_idx + 1] if product_idx + 1 < len(row) else None
            rc_cell = row[product_idx + 2] if product_idx + 2 < len(row) else None
            if not text_match(glass, glass_cell.value if glass_cell else None):
                continue
            if not number_match(rc, rc_cell.value if rc_cell else None):
                continue
            price, col, normal, rebate, note = pick_price(row, product_idx + 3, product_idx + 4)
            w_floor, h_floor, area_step = size_info
            if price is None:
                return HushiMatchResult(None, None, "failed", note, rule_file.name, sheet_name, product_cell.row, product, "PP", area_step, w_floor, h_floor, normal, rebate, col)
            return HushiMatchResult(price, round(price * area_step, 2), "completed", "命中 PP 报价", rule_file.name, sheet_name, product_cell.row, product, "PP", area_step, w_floor, h_floor, normal, rebate, col)

    return HushiMatchResult(None, None, "failed", f"PP 报价未命中：产品={product}, 玻纤={glass}, RC={rc}", rule_file.name, ",".join(pp_sheets), product=product, material_type="PP")


def foil_match(expected: Optional[str], actual: Any) -> bool:
    if not expected:
        return False
    exp = normalize_key(expected).replace("H-", "H")
    act = normalize_key(actual).replace("H-", "H")
    if exp == act:
        return True
    if "/" in exp and act in exp.split("/"):
        return True
    return False


def match_ccl(spec: str, product: str, rule_file: Path, size_info: tuple[float, float, float]) -> HushiMatchResult:
    fields = extract_ccl_fields(spec)
    missing = [k for k, v in fields.items() if v is None]
    if missing:
        return HushiMatchResult(None, None, "failed", "无法提取 CCL 字段：" + ",".join(missing), rule_file.name, product=product, material_type="CCL")

    wb = load_excel_workbook(rule_file)
    ccl_sheets = [name for name in wb.sheetnames if any(key in name.upper() for key in ["CCL", "基板"])]
    if not ccl_sheets:
        return HushiMatchResult(None, None, "failed", "报价单中找不到 CCL/基板报价页", rule_file=rule_file.name, product=product, material_type="CCL")

    pkey = quote_product_key(product)
    for sheet_name in ccl_sheets:
        ws = wb[sheet_name]
        for row, product_idx in iter_data_rows(ws):
            product_cell = row[product_idx]
            if quote_product_key(product_cell.value) != pkey:
                continue
            if not number_match(fields["thickness"], row[product_idx + 1].value if product_idx + 1 < len(row) else None):
                continue
            if not text_match(fields["copper"], row[product_idx + 2].value if product_idx + 2 < len(row) else None):
                continue
            if not text_match(fields["structure"], row[product_idx + 3].value if product_idx + 3 < len(row) else None):
                continue
            if not foil_match(fields["foil"], row[product_idx + 4].value if product_idx + 4 < len(row) else None):
                continue
            price, col, normal, rebate, note = pick_price(row, product_idx + 5, product_idx + 6)
            w_floor, h_floor, area_step = size_info
            if price is None:
                return HushiMatchResult(None, None, "failed", note, rule_file.name, sheet_name, product_cell.row, product, "CCL", area_step, w_floor, h_floor, normal, rebate, col)
            return HushiMatchResult(price, round(price * area_step, 2), "completed", "命中 CCL 报价", rule_file.name, sheet_name, product_cell.row, product, "CCL", area_step, w_floor, h_floor, normal, rebate, col)

    return HushiMatchResult(None, None, "failed", f"CCL 报价未命中：{fields}", rule_file.name, ",".join(ccl_sheets), product=product, material_type="CCL")


def calculate_hushi_spec(spec: str, rule_dir: Path, *, rule_cache: dict[str, Optional[Path]] | None = None) -> HushiMatchResult:
    spec = normalize_text(spec)
    product = extract_product(spec)
    if not product:
        return HushiMatchResult(None, None, "failed", "无法识别胶系/产品")
    rule_file = find_rule_file(product, rule_dir, rule_cache=rule_cache)
    if not rule_file:
        return HushiMatchResult(None, None, "failed", f"找不到报价单：{product}", product=product)
    size = extract_size(spec)
    if not size:
        return HushiMatchResult(None, None, "failed", "无法识别尺寸", rule_file=rule_file.name, product=product)
    size_info = area_step_from_size(*size)
    if re.search(r"\bPP\b", spec.upper()):
        return match_pp(spec, product, rule_file, size_info)
    return match_ccl(spec, product, rule_file, size_info)


def last_non_empty_column(ws) -> int:
    max_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                max_col = max(max_col, cell.column)
    return max_col


def detect_description_column(ws) -> int:
    best_col = 3
    best_score = -1
    max_col = max(ws.max_column, 1)
    for col in range(1, max_col + 1):
        score = 0
        for row in range(1, ws.max_row + 1):
            value = normalize_text(ws.cell(row=row, column=col).value).upper()
            if "NY" in value and ("PP" in value or '"' in value or "RC" in value):
                score += 1
        if score > best_score:
            best_score = score
            best_col = col
    return best_col


def is_effective_description(value) -> bool:
    text = normalize_text(value)
    return bool(text and text.lower() not in {"nan", "none"} and text not in NON_DATA_DESCRIPTIONS)


def load_input_workbook(path: Path):
    return load_workbook_compat(path, keep_formatting=False)


def load_excel_workbook(path: Path):
    resolved = path.resolve()
    mtime_ns = resolved.stat().st_mtime_ns
    return _load_excel_workbook_cached(str(resolved), mtime_ns)


@lru_cache(maxsize=64)
def _load_excel_workbook_cached(path_text: str, mtime_ns: int):
    path = Path(path_text)
    return load_workbook_compat(path, data_only=True, keep_formatting=True)


def is_ole_workbook(path: Path) -> bool:
    return excel_is_ole_workbook(path)


def load_xls_as_workbook(path: Path, *, keep_formatting: bool):
    return excel_load_xls_as_workbook(path, keep_formatting=keep_formatting)
