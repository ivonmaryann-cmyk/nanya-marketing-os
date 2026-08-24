from __future__ import annotations

import json
import re
import zipfile
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from .paths import PACKAGE_DIR
from .purchase_factory_mapper import FACTORY_DETAIL_HEADERS, safe_result_stem
from .purchase_field_rules import clean_text, decimal_or_none, normalize_date


DOMESTIC_TEMPLATE_PATH = (
    PACKAGE_DIR / "default_rules" / "order_entry" / "PDF转Excel内销录单模板.xlsx"
)

DOMESTIC_DETAIL_HEADERS = [
    "项次（必填）",
    "产品编号",
    "品名（选填）",
    "客户产品编号（必填）",
    "客户规格（选填）",
    "客户规格匹配",
    "产品类型（PP、基板）",
    "出货日期（选填）",
    "数量（必填）",
    "税前单价（选填）",
    "单价（选填）",
    "产地（根据账套默认产地）",
    "客户订单序号（选填）",
    "一对多（选填）",
    "备注（选填）",
]


def _header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", clean_text(value).lower())


def _original_value(detail: dict[str, Any], *aliases: str) -> str:
    wanted = {_header_key(alias) for alias in aliases}
    original = detail.get("original") or {}
    for header, value in original.items():
        if _header_key(header) in wanted and clean_text(value):
            return clean_text(value)
    for header, value in original.items():
        key = _header_key(header)
        if any(alias and alias in key for alias in wanted) and clean_text(value):
            return clean_text(value)
    return ""


def _customer_spec(detail: dict[str, Any]) -> str:
    standard = detail.get("standard") or {}
    original_spec = _original_value(
        detail,
        "名称规格",
        "客户规格",
        "物料规格",
        "规格",
        "型号",
    )
    original_description = _original_value(
        detail,
        "物料描述",
        "材料描述",
        "Material Description",
        "Description",
    )
    original_name = _original_value(
        detail,
        "材料名称",
        "物料名称",
        "物料品名",
        "Material Name",
        "品名",
    )
    description = original_description
    if not description and not (original_spec or original_name):
        description = clean_text(standard.get("说明"))
    parts = [original_spec, description]
    if not description:
        parts.append(original_name or clean_text(standard.get("物料名称")))
    combined: list[str] = []
    for value in parts:
        text = clean_text(value)
        if not text:
            continue
        key = _header_key(text)
        existing_index = next(
            (index for index, existing in enumerate(combined) if key in _header_key(existing)),
            None,
        )
        if existing_index is not None:
            continue
        contained_indexes = [
            index for index, existing in enumerate(combined) if _header_key(existing) in key
        ]
        if contained_indexes:
            combined[contained_indexes[0]] = text
            continue
        combined.append(text)
    return " ".join(combined)


def _product_type(detail: dict[str, Any], customer_spec: str) -> str:
    standard = detail.get("standard") or {}
    context = " ".join(
        item
        for item in [
            customer_spec,
            clean_text(standard.get("物料名称")),
            clean_text(standard.get("说明")),
        ]
        if item
    )
    is_pp = bool(re.search(
        r"(?:半固化片|PREPREG|(?<![A-Za-z0-9])PP(?![A-Za-z0-9]))",
        context,
        flags=re.IGNORECASE,
    ))
    is_base = bool(re.search(
        r"(?:覆铜板|铜箔基板|基板|(?<![A-Za-z0-9])CCL(?![A-Za-z0-9])|(?<![A-Za-z0-9])FR\s*-\s*4(?![A-Za-z0-9]))",
        context,
        flags=re.IGNORECASE,
    ))
    if is_pp and not is_base:
        return "PP"
    if is_base and not is_pp:
        return "基板"
    return ""


def _clean_remark(value: Any) -> str:
    return clean_text(value).rstrip("&").rstrip()


def _factory_import(document: dict[str, Any]) -> dict[str, Any]:
    factory_import = document.get("factory_import")
    if not isinstance(factory_import, dict):
        raise ValueError("任务缺少厂内映射数据，请重新执行 PDF/图片转Excel 后再下载内销模板")
    return factory_import


def build_domestic_template_data(document: dict[str, Any]) -> dict[str, Any]:
    """Build shared business data for PDF export and mail order entry."""
    factory_import = _factory_import(document)
    mapped_details = list(document.get("mapped_detail_rows") or [])
    factory_rows = list(factory_import.get("rows") or [])
    if len(mapped_details) != len(factory_rows):
        raise ValueError("任务明细与厂内映射行数不一致，请重新执行 PDF/图片转Excel")

    main_values = list(factory_import.get("main_values") or [])
    order_number = clean_text(main_values[7] if len(main_values) > 7 else "")
    header = {
        "order_type": "220", "type_1": "1", "type_2": "1",
        "bill_to_customer_code": "", "ship_to_customer_code": "",
        "delivery_factory": "", "customer_order_number": order_number,
        "ledger": "KL01",
    }

    rows: list[dict[str, Any]] = []
    for index, (detail, factory_row) in enumerate(zip(mapped_details, factory_rows), start=1):
        customer_spec = _customer_spec(detail)
        rows.append(
            {
                "line_no": str(index),
                "product_code": "",
                "product_name": "",
                "customer_product_code": clean_text(
                    factory_row.get(FACTORY_DETAIL_HEADERS[3])
                ),
                "customer_spec": customer_spec,
                "customer_spec_match": "",
                "product_type": _product_type(detail, customer_spec),
                "delivery_date": clean_text(
                    factory_row.get(FACTORY_DETAIL_HEADERS[4])
                ),
                "quantity": clean_text(
                    factory_row.get(FACTORY_DETAIL_HEADERS[5])
                ),
                "price_before_tax": clean_text(
                    factory_row.get(FACTORY_DETAIL_HEADERS[6])
                ),
                "unit_price": clean_text(
                    factory_row.get(FACTORY_DETAIL_HEADERS[7])
                ),
                "origin": "",
                "customer_order_seq": clean_text(
                    factory_row.get(FACTORY_DETAIL_HEADERS[0])
                )
                or str(index),
                "customer_order_number": order_number,
                "one_to_many": "",
                "remark": _clean_remark(
                    factory_row.get(FACTORY_DETAIL_HEADERS[11])
                ),
            }
        )
    return {"header": header, "lines": rows}


def build_domestic_rows(document: dict[str, Any]) -> tuple[list[Any], list[dict[str, Any]]]:
    data = build_domestic_template_data(document)
    header = data["header"]
    header_values = [
        header["order_type"], header["type_1"], header["type_2"],
        header["bill_to_customer_code"], header["ship_to_customer_code"],
        header["delivery_factory"], header["customer_order_number"], header["ledger"],
    ]
    rows = [
        {
            DOMESTIC_DETAIL_HEADERS[0]: line["line_no"],
            DOMESTIC_DETAIL_HEADERS[1]: line["product_code"],
            DOMESTIC_DETAIL_HEADERS[2]: line["product_name"],
            DOMESTIC_DETAIL_HEADERS[3]: line["customer_product_code"],
            DOMESTIC_DETAIL_HEADERS[4]: line["customer_spec"],
            DOMESTIC_DETAIL_HEADERS[5]: line["customer_spec_match"],
            DOMESTIC_DETAIL_HEADERS[6]: line["product_type"],
            DOMESTIC_DETAIL_HEADERS[7]: line["delivery_date"],
            DOMESTIC_DETAIL_HEADERS[8]: line["quantity"],
            DOMESTIC_DETAIL_HEADERS[9]: line["price_before_tax"],
            DOMESTIC_DETAIL_HEADERS[10]: line["unit_price"],
            DOMESTIC_DETAIL_HEADERS[11]: line["origin"],
            DOMESTIC_DETAIL_HEADERS[12]: line["customer_order_seq"],
            DOMESTIC_DETAIL_HEADERS[13]: line["one_to_many"],
            DOMESTIC_DETAIL_HEADERS[14]: line["remark"],
        }
        for line in data["lines"]
    ]
    return header_values, rows


def _typed_detail_value(header: str, value: Any) -> Any:
    text = clean_text(value)
    if not text:
        return None
    if header in {"数量（必填）", "税前单价（选填）", "单价（选填）"}:
        number = decimal_or_none(text)
        if number is not None:
            return float(number) if number % 1 else int(number)
    if header == "出货日期（选填）":
        normalized = normalize_date(text)
        if normalized:
            try:
                return datetime.strptime(normalized, "%Y-%m-%d").date()
            except ValueError:
                pass
    return text


def build_domestic_workbook(document: dict[str, Any]) -> BytesIO:
    if not DOMESTIC_TEMPLATE_PATH.is_file():
        raise ValueError("PDF 转 Excel 内销模板文件未配置")
    header_values, rows = build_domestic_rows(document)
    book = load_workbook(DOMESTIC_TEMPLATE_PATH)
    sheet = book["内销"]

    for column in range(1, 16):
        sheet.cell(2, column).value = None
    for column, value in enumerate(header_values, start=1):
        sheet.cell(2, column).value = value or None

    required_last_row = 3 + max(1, len(rows))
    style_source_row = 4
    while sheet.max_row < required_last_row:
        target_row = sheet.max_row + 1
        for column in range(1, 16):
            source = sheet.cell(style_source_row, column)
            target = sheet.cell(target_row, column)
            target._style = copy(source._style)
            target.number_format = source.number_format
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[style_source_row].height

    for row_index in range(4, max(sheet.max_row, required_last_row) + 1):
        row = rows[row_index - 4] if row_index - 4 < len(rows) else {}
        for column, header in enumerate(DOMESTIC_DETAIL_HEADERS, start=1):
            cell = sheet.cell(row_index, column)
            cell.value = _typed_detail_value(header, row.get(header, ""))
            if header in {
                "项次（必填）",
                "产品编号",
                "客户产品编号（必填）",
                "客户订单序号（选填）",
            }:
                cell.number_format = "@"
            elif header == "出货日期（选填）":
                cell.number_format = "yyyy/m/d;@"
            elif header == "数量（必填）":
                cell.number_format = "0.###"
            elif header in {"税前单价（选填）", "单价（选填）"}:
                cell.number_format = "#,##0.00_ "

    output = BytesIO()
    book.save(output)
    book.close()
    output.seek(0)
    return output


def _load_job_documents(job: Mapping[str, Any]) -> list[dict[str, Any]]:
    manifest_path = Path(str(job["stored_input_path"] or ""))
    json_dir = manifest_path.parent / "json"
    if not json_dir.is_dir():
        raise ValueError("该任务没有可用的结构化结果，请重新执行 PDF/图片转Excel")
    documents: list[dict[str, Any]] = []
    for json_path in sorted(json_dir.glob("*.json")):
        try:
            document = json.loads(json_path.read_text(encoding="utf-8"))
        except (OSError, ValueError) as exc:
            raise ValueError(f"结构化结果读取失败：{json_path.name}") from exc
        if isinstance(document, dict):
            documents.append(document)
    if not documents:
        raise ValueError("该任务没有可用的结构化结果，请重新执行 PDF/图片转Excel")
    return documents


def build_job_domestic_export(job: Mapping[str, Any]) -> tuple[BytesIO, str, str]:
    documents = _load_job_documents(job)
    if len(documents) == 1:
        document = documents[0]
        filename = f"{safe_result_stem(document.get('source_file'), '采购单')}_内销模板.xlsx"
        return (
            build_domestic_workbook(document),
            filename,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    output = BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index, document in enumerate(documents, start=1):
            stem = safe_result_stem(document.get("source_file"), f"采购单_{index:03d}")
            workbook = build_domestic_workbook(document)
            archive.writestr(f"{index:03d}_{stem}_内销模板.xlsx", workbook.getvalue())
    output.seek(0)
    return output, f"PDF转Excel任务_{job['id']}_内销模板.zip", "application/zip"
