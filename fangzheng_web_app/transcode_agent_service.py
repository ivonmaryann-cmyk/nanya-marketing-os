from __future__ import annotations

import importlib
import json
import math
import os
import re
import sys
import tempfile
import traceback
import unicodedata
from collections import Counter
from functools import lru_cache
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

import openpyxl
import pandas as pd
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from .db import (
    append_job_log,
    create_job,
    get_job,
    get_transcode_agent_confirmation_item,
    list_transcode_agent_confirmation_events,
    list_transcode_agent_confirmation_items,
    prune_jobs_for_employee,
    refresh_transcode_agent_confirmation_item,
    replace_transcode_agent_confirmation_items,
    transcode_agent_confirmation_counts,
    update_job_status,
    update_transcode_agent_confirmation_item,
    update_transcode_agent_row_analysis,
)
from .transcode_customer_rule_admin import (
    CustomerRuleMaintenanceError,
    build_rule_from_form,
    resolve_customer_code_by_name,
    save_rule_override,
    validate_customer_maintained_rule,
)
from .excel_utils import load_workbook_compat, normalized_xlsx_source
from .transcode_evidence_scoring import (
    empty_evidence_score_shadow,
    evidence_gate_decision,
    evaluate_evidence_score_shadow,
    format_source_fields,
    get_evidence_score_runtime_mode,
    get_evidence_gate_mode,
    load_evidence_score_matrix,
)
from .transcode_evidence_model import (
    get_evidence_model_max_calls,
    load_evidence_model_runtime,
    review_evidence_shadow,
)
from .file_utils import safe_unlink
from .job_control import launch_job_process
from .paths import JOBS_DIR, PROJECT_DIR
from .transcode_agent_rules import (
    FEATURE_KEY,
    get_active_transcode_agent_rule_version,
    load_transcode_agent_mapping_tables,
    load_transcode_agent_rules,
)
from .transcode_agent_glue_resolver import (
    is_retired_agent_glue_mapping,
    resolve_agent_glue,
)
from .transcode_agent_standard import OFFICIAL_GRADE_CODES
from .transcode_rules import get_active_transcode_rule_version, get_transcode_rule_file_path
from .transcode_semantic_rules import (
    get_active_transcode_semantic_rule_version,
    load_transcode_semantic_rules,
)
from .transcode_semantic_shadow import (
    SHADOW_STATUS_ERROR,
    SHADOW_STATUS_MATCHED,
    SHADOW_STATUS_MISSING_INPUT,
    SHADOW_STATUS_NOT_MATCHED,
    evaluate_semantic_shadow,
    format_condition_results,
    format_observed_inputs,
    get_semantic_rule_runtime_mode,
)
from .transcode_semantic_overrides import (
    apply_confirmed_semantic_overrides,
    get_semantic_override_mode,
)
from .transcode_confirmation_policy import (
    apply_confirmation_rules_to_evidence,
    decide_confirmation,
    match_confirmation_policy_rules,
)
from .transcode_order_semantic_model import (
    build_model_rule_evaluations,
    build_order_semantic_cache_key,
    load_order_semantic_runtime,
    normalize_order_shadow,
    order_remark_source_fields,
    should_normalize_order,
)
from .transcode_customer_identity import customer_names_match


TRANSCODE_MODULE_NAME = "fangzheng_web_app.transcode_agent_engine"
FIELD_GATE_THRESHOLD = 100
FORMAL_RESULT_HEADER = "Agent转码结果"
PENDING_RESULT_HEADER = "待人工确认码值"
CODE_DIFFERENCE_HEADER = "22位码值差异"
OUTPUT_STATUS_HEADER = "结果对比"
TRANSCODE_STATUS_HEADER = "转码状态"
CONFIRMATION_HEADER = "人工确认提示"
SYSTEM_ANALYSIS_HEADER = "系统分析原因"


FIELD_DEFS = [
    ("glue", "胶系", "step1_glue_code", True),
    ("thickness", "厚度", "step2_thick_code", True),
    ("copper", "铜厚", "step3_copper_code", True),
    ("size", "尺寸", "step4_size_code", True),
    ("glue_category", "胶水类别", "step5_glue_cat_code", True),
    ("copper_type", "铜箔类型", "step6_copper_type_code", True),
    ("grade", "基板级别", "step7_grade_code", True),
    ("total_core", "总/芯厚", "step8_tc_code", True),
    ("structure", "结构码", "step9_struct_code", False),
]

OVERRIDE_STEP_MAP = {
    "glue_code": ("step1_glue_code", "无法识别胶系型号"),
    "thickness_code": ("step2_thick_code", "无法识别厚度"),
    "copper_code": ("step3_copper_code", "无法识别铜箔规格"),
    "size_code": ("step4_size_code", "无法识别尺寸"),
    "glue_category_code": ("step5_glue_cat_code", ""),
    "copper_type_code": ("step6_copper_type_code", ""),
    "grade_code": ("step7_grade_code", ""),
    "tc_code": ("step8_tc_code", ""),
    "struct_code": ("step9_struct_code", ""),
}

OVERRIDE_FIELD_LABELS = {
    "glue_code": "胶系",
    "thickness_code": "厚度",
    "copper_code": "铜厚",
    "size_code": "尺寸",
    "glue_category_code": "胶水类别",
    "copper_type_code": "铜箔类型",
    "grade_code": "基板级别",
    "tc_code": "总/芯厚",
    "struct_code": "结构码",
}

FIELD_KEY_TO_OVERRIDE = {
    "glue": "glue_code",
    "thickness": "thickness_code",
    "copper": "copper_code",
    "size": "size_code",
    "glue_category": "glue_category_code",
    "copper_type": "copper_type_code",
    "grade": "grade_code",
    "total_core": "tc_code",
    "structure": "struct_code",
}

FIELD_CODE_WIDTHS = {
    "glue": 2,
    "thickness": 5,
    "copper": 2,
    "size": 8,
    "glue_category": 1,
    "copper_type": 1,
    "grade": 2,
    "total_core": 1,
}

FIELD_STEP_KEYS = {
    key: step_key
    for key, _label, step_key, gate in FIELD_DEFS
    if gate
}

LONG_TERM_RULE_FIELDS = {
    "glue": ("胶系", "glue"),
    "thickness": ("基板厚度", "thickness"),
    "copper": ("铜箔规格", "copper"),
    "size": ("基板尺寸", "size"),
    "glue_category": ("胶水类别", "glue_category"),
    "copper_type": ("铜箔类型+印字/非印字", "copper_type"),
    "grade": ("基板级别", "grade_intent"),
    "total_core": ("总/芯厚", "total_core"),
}

AGENT_EXECUTABLE_OVERRIDE_FIELDS = {
    "glue_code",
    "thickness_code",
    "copper_code",
    "size_code",
    "glue_category_code",
    "copper_type_code",
    "grade_code",
    "tc_code",
}


def load_transcode_module():
    if TRANSCODE_MODULE_NAME in sys.modules:
        return importlib.reload(sys.modules[TRANSCODE_MODULE_NAME])
    return importlib.import_module(TRANSCODE_MODULE_NAME)


def queue_transcode_agent_job(employee_id: str, uploaded_file: FileStorage, source_filename: str) -> int:
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    safe_filename = secure_filename(source_filename) or f"transcode_agent_{timestamp}.xlsx"
    input_path = employee_dir / f"{timestamp}_transcode_agent_{safe_filename}"
    uploaded_file.save(input_path)
    base_version = get_active_transcode_rule_version()
    agent_version = get_active_transcode_agent_rule_version() or "未上传"
    semantic_version = get_active_transcode_semantic_rule_version() or "未发布"
    job_id = create_job(
        employee_id,
        source_filename,
        str(input_path),
        f"base:{base_version};agent:{agent_version};semantic:{semantic_version}",
        feature=FEATURE_KEY,
    )
    launch_job_process(job_id, FEATURE_KEY, employee_id)
    return job_id


def queue_transcode_agent_single_job(
    employee_id: str,
    *,
    spec: str,
    customer: str = "",
    customer_code: str = "",
    order_remark: str = "",
) -> int:
    """Create a one-row workbook so single and batch inputs use one runtime."""
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    source_filename = f"单条转码_{timestamp}.xlsx"
    input_path = employee_dir / f"{timestamp}_transcode_agent_single.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "转码需求表"
    worksheet.append(["客户代码", "客户简称", "客户规格", "订单备注"])
    worksheet.append(
        [
            str(customer_code or "").strip(),
            str(customer or "").strip(),
            str(spec or "").strip(),
            str(order_remark or "").strip(),
        ]
    )
    workbook.save(input_path)
    workbook.close()

    base_version = get_active_transcode_rule_version()
    agent_version = get_active_transcode_agent_rule_version() or "未上传"
    semantic_version = get_active_transcode_semantic_rule_version() or "未发布"
    job_id = create_job(
        employee_id,
        source_filename,
        str(input_path),
        f"base:{base_version};agent:{agent_version};semantic:{semantic_version}",
        feature=FEATURE_KEY,
    )
    launch_job_process(job_id, FEATURE_KEY, employee_id)
    return job_id


def calculate_transcode_agent_quote(
    spec: str,
    *,
    customer: str = "",
    customer_code: str = "",
    order_text: str = "",
    order_remark: str = "",
    employee_id: str = "",
) -> dict:
    spec = str(spec or "").strip()
    if not spec:
        return {
            "status": "失败",
            "result": None,
            "candidate_code": "",
            "pending_code": "",
            "note": "请输入客户规格",
            "error": "请输入客户规格",
            "confidence": 0,
            "field_evidence": [],
            "rule_version": "",
            "agent_rule_version": "",
            "requires_manual_completion": False,
            "incomplete_fields": [],
            "aps_query_ready": False,
        }
    analysis, base_version, agent_version = _calculate_transcode_agent_analysis(
        spec,
        customer=customer,
        customer_code=customer_code,
        order_text=order_text,
        order_remark=order_remark,
        employee_id=employee_id,
    )
    incomplete_fields = _incomplete_output_fields(analysis)
    return {
        "status": analysis["status"],
        "result": analysis["formal_code"],
        "candidate_code": analysis["candidate_code"],
        "pending_code": analysis["candidate_code"] if analysis["status"] == "待确认" else "",
        "note": analysis["summary"],
        "error": analysis["reason"] if analysis["status"] != "成功" else "",
        "confidence": analysis["overall_score"],
        "field_evidence": analysis["field_evidence"],
        "rule_version": base_version,
        "agent_rule_version": agent_version or "未上传",
        "order_semantic_model": analysis.get("order_semantic_model") or {},
        "requires_manual_completion": bool(incomplete_fields),
        "incomplete_fields": incomplete_fields,
        "aps_query_ready": bool(
            analysis["status"] == "成功"
            and analysis["formal_code"]
            and not incomplete_fields
        ),
    }


def _incomplete_output_fields(analysis: dict[str, Any]) -> list[str]:
    incomplete_fields: list[str] = []
    for item in analysis.get("field_evidence") or []:
        if str(item.get("field_key") or "").strip() != "structure":
            continue
        structure_code = str(item.get("code") or "").strip()
        hit_type = str(item.get("hit_type") or "").strip()
        if structure_code == "*" or hit_type == "占位符" or _is_placeholder(structure_code):
            incomplete_fields.append(str(item.get("field") or "结构码"))

    if not incomplete_fields:
        output_code = str(
            analysis.get("formal_code") or analysis.get("candidate_code") or ""
        ).strip()
        if output_code and "*" in output_code:
            incomplete_fields.append("结构码")
    return list(dict.fromkeys(incomplete_fields))


def _calculate_transcode_agent_analysis(
    spec: str,
    *,
    customer: str = "",
    customer_code: str = "",
    order_text: str = "",
    order_remark: str = "",
    employee_id: str = "",
    runtime: tuple[Any, ...] | None = None,
) -> tuple[dict[str, Any], str, str]:
    engine, tables, agent_rules, agent_mapping_tables, base_version, agent_version = (
        runtime or _load_runtime()
    )
    context_text = " ".join(
        dict.fromkeys(
            value
            for value in [str(order_text or "").strip(), str(order_remark or "").strip()]
            if value
        )
    )
    analysis = analyze_spec(
        engine,
        tables,
        agent_rules,
        spec,
        agent_mapping_tables=agent_mapping_tables,
        customer=customer,
        customer_code=customer_code,
        context_text=context_text,
    )
    _apply_quote_order_semantics(
        engine,
        tables,
        analysis,
        spec=spec,
        customer=customer,
        customer_code=customer_code,
        order_remark=order_remark,
        employee_id=employee_id,
    )
    _apply_explicit_order_instructions(analysis, order_remark, tables)
    confirmation_rules = match_confirmation_policy_rules(
        customer,
        spec,
        order_text,
        order_remark,
    )
    _apply_runtime_confirmation_rules(analysis, confirmation_rules)
    return analysis, base_version, agent_version or ""


def _apply_quote_order_semantics(
    engine,
    tables: dict[str, Any],
    analysis: dict[str, Any],
    *,
    spec: str,
    customer: str,
    customer_code: str,
    order_remark: str,
    employee_id: str,
) -> None:
    semantic_mode = get_semantic_rule_runtime_mode()
    semantic_override_mode = get_semantic_override_mode()
    semantic_version = get_active_transcode_semantic_rule_version()
    semantic_rules: list[dict[str, Any]] = []
    semantic_error = ""
    if semantic_mode == "shadow" and semantic_version:
        try:
            semantic_rules = load_transcode_semantic_rules(semantic_version)
        except Exception as exc:
            semantic_error = str(exc)

    steps = analysis.get("engine_steps") or {}
    clean_remark = str(order_remark or "").strip()
    observations = {
        "订单备注": {
            "available": True,
            "value": clean_remark,
            "sources": ["单条诊断/订单备注"] if clean_remark else [],
        },
        "客户规格": {"available": True, "value": spec, "sources": ["单条客户规格"]},
        "订单规格": {"available": True, "value": spec, "sources": ["单条客户规格"]},
        "胶系": _derived_observation(steps.get("glue_model"), "规格解析/胶系"),
        "基板厚度": _derived_observation(steps.get("thickness_mm"), "规格解析/基板厚度"),
        "铜箔规格": _derived_observation(steps.get("copper_spec_raw"), "规格解析/铜箔规格"),
    }
    copper_top_oz, copper_bottom_oz = _copper_oz_pair(steps.get("copper_spec_raw"))
    observations["铜箔上面oz"] = _derived_observation(copper_top_oz, "规格解析/铜箔上面oz")
    observations["铜箔下面oz"] = _derived_observation(copper_bottom_oz, "规格解析/铜箔下面oz")
    observations["订单规格/订单备注"] = {
        "available": True,
        "value": " ".join(value for value in [spec, clean_remark] if value),
        "sources": ["单条客户规格"] + (["单条诊断/订单备注"] if clean_remark else []),
    }
    evaluations: list[dict[str, Any]] = []
    if semantic_rules and not semantic_error:
        evaluations = evaluate_semantic_shadow(
            semantic_rules,
            customer_code=str(customer_code or ""),
            customer_name=str(customer or ""),
            observations=observations,
            spec=spec,
        )
    if semantic_override_mode == "enforce" and evaluations:
        applied, conflicts = apply_confirmed_semantic_overrides(
            engine,
            tables,
            analysis,
            evaluations,
            allow_order_remark_priority=True,
        )
        if applied or conflicts:
            _refresh_analysis_after_semantic_overrides(analysis, applied, conflicts)
    _attach_semantic_shadow_metadata(
        analysis,
        semantic_mode,
        semantic_version,
        len(semantic_rules),
        semantic_error,
        evaluations,
    )

    runtime = load_order_semantic_runtime(employee_id)
    source_fields = {"订单备注": clean_remark} if clean_remark else {}
    record: dict[str, Any] = {
        "mode": runtime.mode,
        "status": "未调用",
        "model": runtime.model,
        "source_fields": source_fields,
        "rule_ids": [str(item.get("rule_id") or "") for item in evaluations if item.get("rule_id")],
    }
    analysis["order_semantic_model"] = record
    if not clean_remark:
        record["reason"] = "订单备注为空，不调用模型"
        return
    if runtime.mode not in {"shadow", "active"} or runtime.client is None:
        record["reason"] = runtime.load_error or "当前用户未开启模型"
        return
    if not should_normalize_order(evaluations, clean_remark):
        record["reason"] = "该客户没有已批准的订单备注语义规则"
        return
    try:
        normalized = normalize_order_shadow(
            runtime,
            customer_code=str(customer_code or ""),
            customer_name=str(customer or ""),
            source_fields=source_fields,
            semantic_evaluations=evaluations,
        )
        model_evaluations, notes = build_model_rule_evaluations(normalized, evaluations)
        record.update(
            {
                "status": "成功",
                "result": normalized,
                "matched_rule_ids": [str(item.get("rule_id") or "") for item in model_evaluations],
                "notes": notes,
            }
        )
        if semantic_override_mode == "enforce" and model_evaluations:
            applied, conflicts = apply_confirmed_semantic_overrides(
                engine,
                tables,
                analysis,
                model_evaluations,
                allow_order_remark_priority=True,
            )
            if applied or conflicts:
                _refresh_analysis_after_semantic_overrides(analysis, applied, conflicts)
    except Exception as exc:
        record.update({"status": "失败", "error": str(exc)})


def run_transcode_agent_job(job_id: int, employee_id: str) -> None:
    update_job_status(job_id, status="running", log_text="")
    job = get_job(job_id)
    if not job:
        return
    append_job_log(job_id, f"开始营销转码Agent任务，规则版本：{job['rule_version']}")
    try:
        engine, tables, agent_rules, agent_mapping_tables, base_version, agent_version = _load_runtime()
        append_job_log(job_id, f"基础转码规则：{base_version}；Agent规则：{agent_version or '未上传'}")

        semantic_mode = get_semantic_rule_runtime_mode()
        semantic_override_mode = get_semantic_override_mode()
        semantic_version = get_active_transcode_semantic_rule_version()
        semantic_rules: list[dict] = []
        semantic_load_error = ""
        if semantic_mode == "shadow" and semantic_version:
            try:
                semantic_rules = load_transcode_semantic_rules(semantic_version)
            except Exception as exc:
                semantic_load_error = str(exc)
        append_job_log(
            job_id,
            "语义规则影子模式："
            f"{semantic_mode}；正式覆盖：{semantic_override_mode}；版本：{semantic_version or '未发布'}；"
            f"规则数：{len(semantic_rules)}"
            + (f"；加载失败：{semantic_load_error}" if semantic_load_error else ""),
        )
        evidence_score_mode = get_evidence_score_runtime_mode()
        evidence_gate_mode = get_evidence_gate_mode()
        evidence_score_matrix: dict[str, Any] = {}
        evidence_score_load_error = ""
        if evidence_score_mode == "shadow":
            try:
                evidence_score_matrix = load_evidence_score_matrix()
            except Exception as exc:
                evidence_score_load_error = str(exc)
        append_job_log(
            job_id,
            "证据评分影子模式："
            f"{evidence_score_mode}；正式证据门禁：{evidence_gate_mode}"
            + (f"；矩阵加载失败：{evidence_score_load_error}" if evidence_score_load_error else ""),
        )
        evidence_model_runtime = load_evidence_model_runtime()
        evidence_model_max_calls = get_evidence_model_max_calls()
        append_job_log(
            job_id,
            "模型证据审查模式："
            f"{evidence_model_runtime.mode}；模型：{evidence_model_runtime.model or '未启用'}；"
            f"单任务调用上限：{evidence_model_max_calls}"
            + (f"；加载失败：{evidence_model_runtime.load_error}" if evidence_model_runtime.load_error else ""),
        )
        order_semantic_runtime = load_order_semantic_runtime(employee_id)
        append_job_log(
            job_id,
            "订单实时语义标准化："
            f"{order_semantic_runtime.mode}；模型：{order_semantic_runtime.model or '未启用'}；"
            f"单任务调用上限：{order_semantic_runtime.max_calls}"
            + (f"；加载失败：{order_semantic_runtime.load_error}" if order_semantic_runtime.load_error else ""),
        )

        workbook = load_workbook_compat(job["stored_input_path"], data_only=True)
        source_for_result = normalized_xlsx_source(job["stored_input_path"], workbook)
        sheets, _ = engine.load_transcode_inputs(str(source_for_result), str(get_transcode_rule_file_path(base_version)))
        df_req = sheets["转码需求表"].copy()
        spec_col = engine.select_transcode_spec_column(df_req)
        customer_col = engine.detect_customer_column(df_req, spec_col)
        customer_code_col = engine.detect_customer_code_column(df_req)
        context_cols = engine.detect_transcode_context_columns(df_req, spec_col, customer_col, customer_code_col)
        parse_fallback_cols = _detect_parse_fallback_columns(df_req, spec_col)
        pp_context_cols = _detect_pp_context_columns(df_req, spec_col)
        semantic_input_columns = _detect_semantic_input_columns(df_req, spec_col)
        result_col = _ensure_agent_result_column(df_req)
        data_indices = [i for i in range(1, len(df_req)) if _is_effective_spec(df_req.iloc[i, spec_col], engine)]
        total_rows = len(data_indices)
        update_job_status(job_id, status="running", total_rows=total_rows)
        append_job_log(job_id, f"识别规格列：第 {spec_col + 1} 列；Agent结果写入：第 {result_col + 1} 列")
        append_job_log(job_id, f"检测到 {total_rows} 行有效规格数据", total_rows=total_rows)

        analyses: list[dict] = []
        success_count = fail_count = skip_count = confirm_count = 0
        evidence_model_call_count = 0
        order_semantic_call_count = 0
        order_semantic_cache: dict[str, dict[str, Any]] = {}
        for processed, i in enumerate(data_indices, start=1):
            row = df_req.iloc[i]
            customer_code = engine._clean_cell(row.iloc[customer_code_col]) if customer_code_col is not None and len(row) > customer_code_col else ""
            customer = str(row.iloc[customer_col]).strip() if customer_col is not None and len(row) > customer_col and pd.notna(row.iloc[customer_col]) else ""
            spec = engine._clean_cell(row.iloc[spec_col])
            context = engine.build_context_text_from_row(row, context_cols)
            parse_fallback_text = engine.build_context_text_from_row(row, parse_fallback_cols)
            pp_context_text = engine.build_context_text_from_row(row, pp_context_cols)
            cust_spec = engine._clean_cell(row.iloc[6]) if len(row) > 6 else ""
            normalized_spec = engine._clean_cell(row.iloc[7]) if len(row) > 7 else ""
            pp_check_text = " ".join([spec, cust_spec, normalized_spec, pp_context_text])
            excel_row = i + 1

            if engine.is_pp_or_rc_spec(pp_check_text):
                skip_count += 1
                result_text = "跳过：PP/RC/% 暂不输出CCL制造编码"
                df_req.iloc[i, result_col] = result_text
                skip_analysis = _skip_analysis(excel_row, customer_code, customer, spec, result_text)
                _attach_semantic_shadow_metadata(
                    skip_analysis,
                    semantic_mode,
                    semantic_version,
                    len(semantic_rules),
                    semantic_load_error,
                    [],
                )
                skip_analysis["evidence_score_shadow"] = empty_evidence_score_shadow(
                    reason="PP/RC/%跳过，不参与CCL证据评分",
                )
                skip_analysis["order_semantic_model"] = {
                    "mode": order_semantic_runtime.mode,
                    "status": "跳过",
                    "model": order_semantic_runtime.model,
                    "reason": "PP/RC/%不属于本轮CCL语义标准化范围",
                }
                analyses.append(skip_analysis)
                append_job_log(job_id, f"第 {excel_row} 行跳过：PP/RC/%", skip_count=skip_count, current_row=processed, total_rows=total_rows)
                continue

            analysis = analyze_spec(
                engine,
                tables,
                agent_rules,
                spec,
                agent_mapping_tables=agent_mapping_tables,
                customer=customer,
                customer_code=customer_code,
                context_text=context,
                parse_fallback_text=parse_fallback_text,
                excel_row=excel_row,
            )
            shadow_results: list[dict] = []
            observations: dict[str, dict[str, Any]] = {}
            if semantic_mode == "shadow" or evidence_score_mode == "shadow":
                observations = _build_semantic_observations(
                    row,
                    semantic_input_columns,
                    analysis.get("engine_steps") or {},
                )
            if semantic_mode == "shadow" and semantic_rules and not semantic_load_error:
                try:
                    shadow_results = evaluate_semantic_shadow(
                        semantic_rules,
                        customer_code=customer_code,
                        customer_name=customer,
                        observations=observations,
                        excel_row=excel_row,
                        spec=spec,
                    )
                except Exception as exc:
                    shadow_results = [_semantic_shadow_error_result(excel_row, customer_code, customer, spec, exc)]
            if semantic_override_mode == "enforce" and shadow_results:
                semantic_applied, semantic_conflicts = apply_confirmed_semantic_overrides(
                    engine,
                    tables,
                    analysis,
                    shadow_results,
                    allow_order_remark_priority=True,
                )
                if semantic_applied or semantic_conflicts:
                    _refresh_analysis_after_semantic_overrides(
                        analysis,
                        semantic_applied,
                        semantic_conflicts,
                    )
            _attach_semantic_shadow_metadata(
                analysis,
                semantic_mode,
                semantic_version,
                len(semantic_rules),
                semantic_load_error,
                shadow_results,
            )
            semantic_source_fields = order_remark_source_fields(observations)
            order_remark = semantic_source_fields.get("订单备注", "")
            analysis["order_semantic_model"] = {
                "mode": order_semantic_runtime.mode,
                "status": "未调用",
                "model": order_semantic_runtime.model,
                "source_fields": semantic_source_fields,
                "rule_ids": [
                    str(item.get("rule_id") or "")
                    for item in shadow_results
                    if item.get("rule_id")
                ],
            }
            if (
                order_semantic_runtime.mode in {"shadow", "active"}
                and order_semantic_runtime.client is not None
                and should_normalize_order(shadow_results, order_remark)
            ):
                cache_key = build_order_semantic_cache_key(
                    customer_code,
                    customer,
                    semantic_source_fields,
                    shadow_results,
                )
                cached = order_semantic_cache.get(cache_key)
                if cached is not None:
                    analysis["order_semantic_model"] = dict(cached, cached=True)
                elif order_semantic_call_count >= order_semantic_runtime.max_calls:
                    analysis["order_semantic_model"].update(
                        {"status": "限流跳过", "reason": "达到单任务DeepSeek调用上限"}
                    )
                else:
                    order_semantic_call_count += 1
                    try:
                        normalized = normalize_order_shadow(
                            order_semantic_runtime,
                            customer_code=customer_code,
                            customer_name=customer,
                            source_fields=semantic_source_fields,
                            semantic_evaluations=shadow_results,
                        )
                        model_record = {
                            "mode": order_semantic_runtime.mode,
                            "status": "成功",
                            "model": order_semantic_runtime.model,
                            "cached": False,
                            "source_fields": semantic_source_fields,
                            "rule_ids": analysis["order_semantic_model"]["rule_ids"],
                            "result": normalized,
                        }
                    except Exception as exc:
                        model_record = {
                            "mode": order_semantic_runtime.mode,
                            "status": "失败",
                            "model": order_semantic_runtime.model,
                            "cached": False,
                            "source_fields": semantic_source_fields,
                            "rule_ids": analysis["order_semantic_model"]["rule_ids"],
                            "error": str(exc),
                        }
                    order_semantic_cache[cache_key] = model_record
                    analysis["order_semantic_model"] = model_record
                model_result = (analysis.get("order_semantic_model") or {}).get("result")
                if model_result:
                    model_evaluations, model_notes = build_model_rule_evaluations(
                        model_result,
                        shadow_results,
                    )
                    analysis["order_semantic_model"]["matched_rule_ids"] = [
                        str(item.get("rule_id") or "") for item in model_evaluations
                    ]
                    analysis["order_semantic_model"]["notes"] = model_notes
                    analysis["semantic_model_evaluations"] = model_evaluations
                    if semantic_override_mode == "enforce" and model_evaluations:
                        model_applied, model_conflicts = apply_confirmed_semantic_overrides(
                            engine,
                            tables,
                            analysis,
                            model_evaluations,
                            allow_order_remark_priority=True,
                        )
                        if model_applied or model_conflicts:
                            _refresh_analysis_after_semantic_overrides(
                                analysis,
                                model_applied,
                                model_conflicts,
                            )
            elif not order_remark:
                analysis["order_semantic_model"]["reason"] = "订单备注为空，不调用模型"
            elif not any(
                "订单备注" in (item.get("observed_inputs") or {})
                for item in shadow_results
            ):
                analysis["order_semantic_model"]["reason"] = "该客户没有已批准的订单备注语义规则"
            _apply_explicit_order_instructions(analysis, order_remark, tables)
            confirmation_rules = match_confirmation_policy_rules(
                customer,
                spec,
                context,
                observations.get("客户规格", {}).get("value", ""),
                observations.get("客户料品名称", {}).get("value", ""),
            )
            _apply_runtime_confirmation_rules(analysis, confirmation_rules)
            if evidence_score_mode == "shadow" and evidence_score_matrix and not evidence_score_load_error:
                try:
                    analysis["evidence_score_shadow"] = evaluate_evidence_score_shadow(
                        analysis,
                        semantic_evaluations=shadow_results,
                        observations=observations,
                        matrix=evidence_score_matrix,
                    )
                except Exception as exc:
                    analysis["evidence_score_shadow"] = empty_evidence_score_shadow(
                        current_score=analysis.get("overall_score", 0),
                        reason=f"证据影子评分异常，不影响正式转码：{exc}",
                    )
            else:
                analysis["evidence_score_shadow"] = empty_evidence_score_shadow(
                    current_score=analysis.get("overall_score", 0),
                    reason=evidence_score_load_error or "证据影子评分未启用",
                )
            if (
                evidence_model_runtime.mode == "shadow"
                and evidence_model_runtime.client is not None
                and evidence_score_matrix
                and analysis["evidence_score_shadow"].get("field_reviews")
                and evidence_model_call_count < evidence_model_max_calls
            ):
                analysis["evidence_score_shadow"] = review_evidence_shadow(
                    analysis,
                    semantic_evaluations=shadow_results,
                    matrix=evidence_score_matrix,
                    client=evidence_model_runtime.client,
                )
                evidence_model_call_count += int(
                    analysis["evidence_score_shadow"].get("model_call_count") or 0
                )
            gate_result = evidence_gate_decision(analysis, mode=evidence_gate_mode)
            analysis["evidence_gate"] = gate_result
            if analysis.get("status") == "成功" and gate_result["blocked"]:
                analysis["status"] = "待确认"
                analysis["formal_code"] = ""
                blocked_fields = "、".join(gate_result["blockers"]) or "关键字段未达到100分"
                analysis["reason"] = f"100分正式码门禁拦截：{blocked_fields}"
                analysis["summary"] = _format_agent_summary(
                    analysis["status"],
                    analysis.get("candidate_code", ""),
                    gate_result["effective_score"],
                    analysis["reason"],
                    analysis.get("applied_rules") or [],
                )
            analyses.append(analysis)
            if analysis["status"] == "成功":
                success_count += 1
                df_req.iloc[i, result_col] = analysis["formal_code"]
                log_text = f"第 {excel_row} 行高置信出码：{analysis['formal_code']}"
            elif analysis["status"] == "待确认":
                confirm_count += 1
                analysis["formal_code"] = ""
                fail_count += 1
                df_req.iloc[i, result_col] = ""
                log_text = f"第 {excel_row} 行待确认：{analysis['reason']}"
            else:
                fail_count += 1
                df_req.iloc[i, result_col] = f"未识别：{analysis['reason']}"
                log_text = f"第 {excel_row} 行未识别：{analysis['reason']}"
            append_job_log(
                job_id,
                log_text,
                success_count=success_count,
                fail_count=fail_count,
                skip_count=skip_count,
                current_row=processed,
                total_rows=total_rows,
            )

        input_path = Path(job["stored_input_path"])
        output_path = input_path.with_name(f"{input_path.stem}_Agent转码结果.xlsx")
        _save_agent_result(
            str(source_for_result),
            str(output_path),
            df_req,
            result_col,
            analyses,
            agent_rules,
            agent_mapping_tables,
            confirm_count,
        )
        confirmation_items = _build_confirmation_items(analyses)
        replace_transcode_agent_confirmation_items(job_id, employee_id, confirmation_items)
        final_status = "awaiting_confirmation" if confirmation_items else "completed"
        update_job_status(
            job_id,
            status=final_status,
            stored_result_path=str(output_path),
            success_count=success_count,
            fail_count=fail_count,
            skip_count=skip_count,
            confirm_count=confirm_count,
            current_row=total_rows,
            total_rows=total_rows,
            completed=not confirmation_items,
        )
        if confirmation_items:
            append_job_log(
                job_id,
                f"第一遍解析完成，{confirm_count} 行进入待人工确认中心",
                confirm_count=confirm_count,
                current_row=total_rows,
                total_rows=total_rows,
            )
        else:
            append_job_log(
                job_id,
                "Agent结果文件已生成，任务完成",
                confirm_count=0,
                current_row=total_rows,
                total_rows=total_rows,
            )
    except Exception as exc:
        append_job_log(job_id, f"任务失败：{exc}")
        update_job_status(
            job_id,
            status="failed",
            error_message=f"{exc}\n{traceback.format_exc(limit=8)}",
            completed=True,
        )
    finally:
        stale_jobs = prune_jobs_for_employee(employee_id, keep_limit=500)
        for stale in stale_jobs:
            for key in ["stored_input_path", "stored_result_path"]:
                safe_unlink(stale[key])


def analyze_spec(
    engine,
    tables: dict,
    agent_rules: list[dict],
    spec: str,
    *,
    agent_mapping_tables: dict[str, list[dict]] | None = None,
    customer: str = "",
    customer_code: str = "",
    context_text: str = "",
    parse_fallback_text: str = "",
    excel_row: int | None = None,
) -> dict:
    runtime_tables = dict(tables)
    runtime_tables["structured_special_rules"] = []
    code, steps, err = engine.transcode_row(
        engine._clean_cell(spec),
        "",
        str(customer or "").strip(),
        str(customer_code or "").strip(),
        runtime_tables,
        str(context_text or "").strip(),
    )
    initial_errors = list((steps or {}).get("errors") or [])
    prefer_normalized_fallback = _prefer_normalized_spec_for_legacy_format(spec)
    if (initial_errors or prefer_normalized_fallback) and str(parse_fallback_text or "").strip():
        fallback_spec = str(parse_fallback_text).strip()
        fallback_code, fallback_steps, fallback_err = engine.transcode_row(
            fallback_spec,
            "",
            str(customer or "").strip(),
            str(customer_code or "").strip(),
            runtime_tables,
            str(context_text or "").strip(),
        )
        fallback_errors = list((fallback_steps or {}).get("errors") or [])
        if len(fallback_errors) < len(initial_errors) or (
            prefer_normalized_fallback and len(fallback_errors) <= len(initial_errors)
        ):
            code, steps, err = fallback_code, fallback_steps, fallback_err
            steps["context_fallback_used"] = True
            steps["context_fallback_note"] = (
                "旧格式客户规格存在歧义，优先使用同行标准规格列"
                if prefer_normalized_fallback and not initial_errors
                else "客户规格解析失败后，使用同一订单行的标准规格列补全缺失字段"
            )
    steps = dict(steps or {})
    steps["agent_input_text"] = " ".join(
        value for value in (str(spec or "").strip(), str(context_text or "").strip()) if value
    )
    errors = list(steps.get("errors") or [])
    match_customer_code, match_customer_name, customer_rule_group = _expand_customer_rule_identity(
        agent_mapping_tables or {}, customer_code, customer
    )
    if customer_rule_group:
        steps["customer_rule_group"] = customer_rule_group
    glue_master_rules, glue_master_conflicts = _apply_agent_glue_master(
        engine,
        tables,
        agent_mapping_tables or {},
        match_customer_code,
        match_customer_name,
        spec,
        context_text,
        steps,
        errors,
    )
    # Establish the latest global glue/category state before evaluating
    # customer rules whose conditions depend on the current glue code.
    global_special_rules = _apply_agent_global_special_rules(
        agent_mapping_tables or {}, spec, context_text, steps, errors
    )
    applied_rules, conflicts = _apply_agent_rules(
        agent_rules, match_customer_code, match_customer_name, spec, context_text, steps, errors
    )
    conflicts = glue_master_conflicts + conflicts
    applied_field_mappings = _apply_agent_field_mappings(
        engine,
        agent_mapping_tables or {},
        match_customer_code,
        match_customer_name,
        spec,
        context_text,
        steps,
        errors,
        applied_rules,
    )
    if applied_field_mappings:
        applied_rules.extend(applied_field_mappings)
    applied_mappings = _apply_agent_size_mappings(
        engine,
        agent_mapping_tables or {},
        match_customer_code,
        match_customer_name,
        spec,
        context_text,
        steps,
        errors,
        applied_rules,
    )
    if applied_mappings:
        applied_rules.extend(applied_mappings)
    applied_thickness_mappings = _apply_agent_thickness_mappings(
        engine,
        agent_mapping_tables or {},
        match_customer_code,
        match_customer_name,
        spec,
        context_text,
        steps,
        errors,
        applied_rules,
    )
    if applied_thickness_mappings:
        applied_rules.extend(applied_thickness_mappings)
    applied_material_mappings = _apply_agent_material_code_mappings(
        agent_mapping_tables or {},
        match_customer_code,
        match_customer_name,
        spec,
        context_text,
        steps,
        applied_rules,
    )
    if applied_material_mappings:
        applied_rules.extend(applied_material_mappings)
    # Global special conditions establish the shared baseline. Customer-scoped
    # rules are applied afterwards and therefore retain the higher precedence.
    if global_special_rules:
        applied_rules = global_special_rules + applied_rules
    if glue_master_rules:
        applied_rules = glue_master_rules + applied_rules
    _enforce_retired_glue_runtime_guard(steps, errors, conflicts)
    candidate_code = _build_code_from_steps(steps, errors)
    field_evidence = _build_field_evidence(steps, errors, applied_rules, conflicts)
    decision = decide_confirmation(
        errors=errors,
        conflicts=conflicts,
        candidate_code=candidate_code,
        field_evidence=field_evidence,
    )
    status = decision["status"]
    formal_code = decision["formal_code"]
    overall_score = decision["overall_score"]
    reason = decision["reason"]

    summary = _format_agent_summary(status, formal_code or candidate_code, overall_score, reason, applied_rules)
    return {
        "row": excel_row,
        "customer_code": customer_code,
        "customer": customer,
        "spec": spec,
        "context_text": context_text,
        "status": status,
        "formal_code": formal_code,
        "candidate_code": candidate_code,
        "overall_score": overall_score,
        "reason": reason,
        "summary": summary,
        "field_evidence": field_evidence,
        "applied_rules": applied_rules,
        "conflicts": conflicts,
        "engine_steps": steps,
        "decision_state": decision["decision_state"],
        "confirmation_triggers": decision["confirmation_triggers"],
    }


def _load_runtime():
    engine = load_transcode_module()
    base_version = get_active_transcode_rule_version()
    rule_path = get_transcode_rule_file_path(base_version)
    tables = engine.build_lookup_tables(engine.load_rule_sheets(str(rule_path)))
    from .transcode_rule_center import lookup_map_with_overrides, merge_lookup_overrides

    merge_lookup_overrides(tables)
    _remove_retired_glue_mappings(tables)
    engine.HIGH_SPEED_MIL_TO_MM = lookup_map_with_overrides(
        "high_speed_mil", engine.HIGH_SPEED_MIL_TO_MM
    )
    engine.STANDARD_MM_SIZE_ALIASES = lookup_map_with_overrides(
        "standard_size", engine.STANDARD_MM_SIZE_ALIASES
    )
    engine.MICRON_COPPER_MAP = lookup_map_with_overrides(
        "copper_micron", engine.MICRON_COPPER_MAP
    )
    copper_types = lookup_map_with_overrides(
        "copper_type", dict(engine.SPECIAL_COPPER_MAP)
    )
    engine.SPECIAL_COPPER_MAP = sorted(copper_types.items(), key=lambda item: len(item[0]), reverse=True)
    copper_valid = lookup_map_with_overrides(
        "copper_valid",
        {value: str(value).replace("/", "") for value in engine.VALID_COPPER_SPECS},
    )
    engine.VALID_COPPER_SPECS = set(copper_valid)
    base_size_ranges = {
        f"{w_min:g}-{w_max:g} × {h_min:g}-{h_max:g}": f"{std_w:g} × {std_h:g}"
        for w_min, w_max, h_min, h_max, std_w, std_h in engine.STANDARD_SIZE_RANGES
    }
    effective_size_ranges = lookup_map_with_overrides("size_range", base_size_ranges)
    parsed_size_ranges = []
    for source_range, target_size in effective_size_ranges.items():
        source_numbers = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", str(source_range))]
        target_numbers = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", str(target_size))]
        if len(source_numbers) == 4 and len(target_numbers) == 2:
            parsed_size_ranges.append((*source_numbers, *target_numbers))
    if parsed_size_ranges:
        engine.STANDARD_SIZE_RANGES = parsed_size_ranges
    grade_codes = set(
        lookup_map_with_overrides(
            "grade_code", {code: code for code in OFFICIAL_GRADE_CODES}
        )
    )
    # Agent合法等级可以由客户正式规则补充，不反向修改已停用的普通转码规则表。
    grade_code_map = tables.setdefault("grade_code_map", {})
    for grade_code in grade_codes:
        grade_code_map.setdefault(grade_code, f"Agent合法基板级别{grade_code}")
    tables["structured_special_rules"] = []
    agent_version = get_active_transcode_agent_rule_version()
    agent_rules = load_transcode_agent_rules(agent_version) if agent_version else []
    agent_mapping_tables = load_transcode_agent_mapping_tables(agent_version) if agent_version else {}
    return engine, tables, agent_rules, agent_mapping_tables, base_version, agent_version


def _remove_retired_glue_mappings(tables: dict[str, Any]) -> None:
    """Remove the retired NY-A1 -> 2Z mapping from legacy runtime lookup maps."""
    for map_name in ("glue_exact_map", "glue_model_map"):
        source = dict(tables.get(map_name) or {})
        tables[map_name] = {
            name: code
            for name, code in source.items()
            if not is_retired_agent_glue_mapping(
                {"胶系名称": name, "输出胶系代码": code}
            )
        }


def _enforce_retired_glue_runtime_guard(
    steps: dict[str, Any],
    errors: list[str],
    conflicts: list[str],
) -> None:
    """Prevent any later override path from restoring the retired NY-A1 -> 2Z."""
    current_code = str(steps.get("step1_glue_code") or "").strip().upper()
    if current_code != "2Z":
        return
    names = (
        steps.get("glue_model"),
        steps.get("agent_glue_name"),
        steps.get("raw_glue"),
    )
    if not any(
        is_retired_agent_glue_mapping(
            {"胶系名称": name, "输出胶系代码": current_code}
        )
        for name in names
        if str(name or "").strip()
    ):
        return

    message = "glue_code: NY-A1→2Z已废弃，运行时禁止正式出码"
    if message not in conflicts:
        conflicts.append(message)
    fallback = str(steps.get("agent_glue_resolved_code") or "").strip().upper()
    if fallback and fallback != "2Z":
        steps["step1_glue_code"] = fallback
        return

    steps["step1_glue_code"] = ""
    error = "无法识别胶系型号：NY-A1→2Z已废弃，需使用最新胶系主表"
    if error not in errors:
        errors.append(error)
    steps["errors"] = errors


def _apply_agent_rules(
    rules: list[dict],
    customer_code: str,
    customer_name: str,
    spec: str,
    context: str,
    steps: dict,
    errors: list[str],
) -> tuple[list[dict], list[str]]:
    matched_rules = [
        rule
        for rule in rules
        if _rule_executable(rule)
        and _rule_matches(rule, customer_code, customer_name, spec, context, steps)
    ]
    values: dict[str, str] = {}
    priorities: dict[str, int] = {}
    applied: list[dict] = []
    conflicts = _same_condition_rule_conflicts(matched_rules)
    for rule in sorted(matched_rules, key=lambda item: _rule_priority(item), reverse=True):
        field = rule.get("覆盖字段", "")
        value = str(rule.get("覆盖值", "") or "").strip().upper()
        if field == "glue_code" and is_retired_agent_glue_mapping(
            {
                "映射ID": rule.get("规则ID", ""),
                "胶系名称": rule.get("条件胶系", "") or rule.get("条件文本", ""),
                "输出胶系代码": value,
            }
        ):
            conflicts.append(
                f"glue_code: NY-A1→2Z已废弃，规则{rule.get('规则ID', '')}不得执行"
            )
            continue
        if _should_skip_agent_rule_override(rule, field, value, customer_name, spec, context, steps):
            continue
        if field in values and values[field] != value:
            if priorities.get(field, 0) > _rule_priority(rule):
                continue
            conflicts.append(f"{field}: {values[field]} vs {value} ({rule.get('规则ID')})")
            continue
        step_key, error_text = OVERRIDE_STEP_MAP.get(field, ("", ""))
        if not step_key:
            continue
        old_value = str(steps.get(step_key, "") or "")
        values[field] = value
        priorities[field] = _rule_priority(rule)
        steps[step_key] = value
        if field == "tc_code":
            steps["order_type"] = "芯厚" if value == "C" else "总厚"
        if error_text:
            errors[:] = [item for item in errors if item != error_text and not item.startswith(error_text)]
            steps["errors"] = errors
        applied.append(
            {
                "rule_id": rule.get("规则ID", ""),
                "field": field,
                "old": old_value,
                "new": value,
                "text": rule.get("规则文本", ""),
                "source": rule.get("命中来源", "Agent规则包"),
                "source_row": rule.get("来源行号", ""),
                "source_field": rule.get("来源字段", ""),
                "rule_type": "确认草稿机器规则",
            }
        )
        linked_glue_category = _linked_glue_category_override(field, value)
        if linked_glue_category and "glue_category_code" not in values:
            old_category = str(steps.get("step5_glue_cat_code", "") or "")
            values["glue_category_code"] = linked_glue_category
            priorities["glue_category_code"] = _rule_priority(rule)
            steps["step5_glue_cat_code"] = linked_glue_category
            steps["glue_category"] = "普通"
            applied.append(
                {
                    "rule_id": rule.get("规则ID", ""),
                    "field": "glue_category_code",
                    "old": old_category,
                    "new": linked_glue_category,
                    "text": rule.get("规则文本", ""),
                    "source": rule.get("命中来源", "Agent规则包"),
                    "source_row": rule.get("来源行号", ""),
                    "source_field": rule.get("来源字段", ""),
                    "rule_type": "确认草稿机器规则/胶系联动",
                }
            )
    if applied:
        steps["agent_rules"] = [f"{item['rule_id']} {item['field']}:{item['old']}->{item['new']}" for item in applied]
    if conflicts:
        conflicts = list(dict.fromkeys(conflicts))
        steps["agent_rule_conflicts"] = conflicts
    return applied, conflicts


def _same_condition_rule_conflicts(rules: list[dict]) -> list[str]:
    """Detect duplicate machine conditions before priority chooses a candidate."""
    grouped: dict[tuple[str, tuple[str, ...]], list[dict]] = {}
    for rule in rules:
        field = str(rule.get("覆盖字段") or "").strip()
        if not field:
            continue
        grouped.setdefault((field, _agent_rule_condition_signature(rule)), []).append(rule)

    conflicts: list[str] = []
    for (field, _signature), same_condition_rules in grouped.items():
        outputs = {
            str(rule.get("覆盖值") or "").strip().upper()
            for rule in same_condition_rules
            if str(rule.get("覆盖值") or "").strip()
        }
        if len(outputs) <= 1:
            continue
        rule_ids = "/".join(
            str(rule.get("规则ID") or "未编号") for rule in same_condition_rules
        )
        conflicts.append(
            f"{field}: 同一条件存在多个输出{'/'.join(sorted(outputs))}"
            f"（规则{rule_ids}）；不使用优先级消除冲突"
        )
    return conflicts


def _agent_rule_condition_signature(rule: dict) -> tuple[str, ...]:
    identity_fields = (
        "客户代码",
        "客户简称",
        "物料类别",
    )
    condition_fields = (
        "条件胶系",
        "条件铜厚",
        "条件厚度",
        "条件尺寸",
        "条件关键词",
    )
    conditions = tuple(_norm_match(rule.get(field, "")) for field in condition_fields)
    fallback_text = "" if any(conditions) else _norm_match(rule.get("条件文本", ""))
    return (
        "GLOBAL" if bool(rule.get("_global_rule")) else "CUSTOMER",
        *(str(_norm_match(rule.get(field, ""))) for field in identity_fields),
        *conditions,
        fallback_text,
    )


def _should_skip_agent_rule_override(
    rule: dict,
    field: str,
    value: str,
    customer_name: str,
    spec: str,
    context: str,
    steps: dict,
) -> bool:
    condition_text = str(rule.get("条件文本", "") or rule.get("规则文本", "") or "")
    glue_code_match = re.search(r"胶系代码\s*[:：=]\s*([A-Z0-9]{2})", condition_text, re.IGNORECASE)
    if glue_code_match:
        required_glue_code = glue_code_match.group(1).upper()
        current_glue_code = str(steps.get("step1_glue_code", "") or "").strip().upper()
        if current_glue_code != required_glue_code:
            return True
    if field == "grade_code" and value == "F1" and _is_shenwan_core_ny2140(customer_name, spec, context, steps):
        return True
    if (
        field == "grade_code"
        and value == "AT"
        and "惠州泰和" in f"{customer_name or ''} {spec or ''} {context or ''}"
        and str(steps.get("step8_tc_code", "") or "").upper() != "T"
    ):
        # 已确认口径是“总厚1.2mm（含）以下=AT”，芯厚规格不能套用。
        return True
    return False


def _is_shenwan_core_ny2140(customer_name: str, spec: str, context: str, steps: dict) -> bool:
    combined = f"{customer_name or ''} {spec or ''} {context or ''}"
    if "深万基隆" not in combined or "芯板" not in combined:
        return False
    glue_text = " ".join(
        str(steps.get(key, "") or "")
        for key in ("glue_model", "step1_glue_code", "glue_code", "raw_glue")
    ).upper()
    return "NY2140" in combined.upper() or "NY2140" in glue_text or str(steps.get("step1_glue_code", "")).upper() == "2A"


def _linked_glue_category_override(field: str, value: str) -> str:
    if field == "glue_code" and value in {"2A", "2B", "2C"}:
        return "Y"
    if field == "glue_code" and value == "AH":
        return "R"
    if field == "glue_code" and value == "2T":
        return "R"
    return ""


def _apply_agent_glue_master(
    engine,
    tables: dict,
    mapping_tables: dict[str, list[dict]],
    customer_code: str,
    customer_name: str,
    spec: str,
    context: str,
    steps: dict,
    errors: list[str],
) -> tuple[list[dict], list[str]]:
    resolution = resolve_agent_glue(
        mapping_tables,
        spec=spec,
        context=context,
        customer_code=customer_code,
        customer_name=customer_name,
        current_code=str(steps.get("step1_glue_code") or ""),
    )
    if not resolution:
        return [], []
    if resolution.get("status") == "conflict":
        return [], [str(resolution.get("conflict") or "胶系主表同名多码冲突")]

    code = str(resolution.get("code") or "").strip().upper()
    old_code = str(steps.get("step1_glue_code") or "").strip().upper()
    if not code:
        return [], []
    steps["step1_glue_code"] = code
    steps["agent_glue_resolved_code"] = code
    steps["agent_glue_id"] = str(resolution.get("glue_id") or "")
    steps["agent_glue_name"] = str(resolution.get("name") or "")
    steps["agent_glue_source"] = str(resolution.get("source") or "")
    if resolution.get("uncertain"):
        steps["agent_glue_uncertain"] = True
        steps["agent_glue_candidates"] = list(resolution.get("candidates") or [])
        steps["agent_glue_uncertain_reason"] = str(resolution.get("conflict") or "")
    errors[:] = [item for item in errors if "胶系" not in str(item)]
    steps["errors"] = errors

    category_model = str(resolution.get("name") or steps.get("glue_model") or "")
    category = engine.get_glue_category(
        category_model,
        tables.get("glue_cat_map") or {},
        code,
    )
    if category:
        steps["glue_category"] = category
        steps["step5_glue_cat_code"] = "Y" if category == "普通" else "R"

    applied = [
        {
            "rule_id": resolution.get("rule_id", ""),
            "field": "glue_code",
            "old": old_code,
            "new": code,
            "text": resolution.get("text", ""),
            "source": resolution.get("source", "Agent胶系主表"),
            "source_row": resolution.get("source_row", ""),
            "source_field": "Agent胶系主表",
            "rule_type": "Agent胶系主数据映射",
        }
    ]
    conflicts = []
    if resolution.get("uncertain"):
        conflicts.append(str(resolution.get("conflict") or "最新版胶系主表同名多码待补口径"))
    return applied, conflicts


def _apply_agent_field_mappings(
    engine,
    mapping_tables: dict[str, list[dict]],
    customer_code: str,
    customer_name: str,
    spec: str,
    context: str,
    steps: dict,
    errors: list[str],
    applied_rules: list[dict],
) -> list[dict]:
    """Apply exact customer+glue mappings derived from validated correct-code samples."""
    already_overridden = {str(item.get("field") or "") for item in applied_rules}
    glue_master_matched = bool(str(steps.get("agent_glue_name") or "").strip())
    combined = f"{spec or ''} {context or ''}".upper()
    current_glue = _norm_match(steps.get("glue_model", ""))
    applied: list[dict] = []

    for row in mapping_tables.get("客户字段映射", []):
        if not _mapping_row_enabled(row) or not _mapping_customer_matches(row, customer_code, customer_name):
            continue
        field = str(row.get("覆盖字段", "") or "").strip()
        value = str(row.get("覆盖值", "") or "").strip().upper()
        if field not in AGENT_EXECUTABLE_OVERRIDE_FIELDS or not value:
            continue
        glue_condition = _norm_match(row.get("条件胶系", ""))
        if glue_condition and glue_condition != current_glue:
            continue
        keyword_condition = str(row.get("条件关键词", "") or "").strip()
        if keyword_condition:
            # “2/2、H/H、T/T”中的斜杠是规格本身，不是多选分隔符。
            keywords = [item.strip() for item in re.split(r"[,，、;；]+", keyword_condition) if item.strip()]
            if keywords and not any(_keyword_condition_matches(item, combined) for item in keywords):
                continue

        step_key, error_text = OVERRIDE_STEP_MAP.get(field, ("", ""))
        if not step_key:
            continue
        old_value = str(steps.get(step_key, "") or "")
        if field in already_overridden or (
            glue_master_matched and field in {"glue_code", "glue_category_code"}
        ):
            if old_value.strip().upper() != value:
                applied.append(
                    {
                        "rule_id": row.get("映射ID", ""),
                        "field": field,
                        "old": old_value,
                        "new": old_value,
                        "historical_suggested": value,
                        "text": (
                            f"{row.get('规则文本', '')}；历史样本建议{value}，"
                            f"与当前正式规则结果{old_value}不一致"
                        ),
                        "source": "正确码回归客户字段映射/冲突证据",
                        "source_row": row.get("来源行号", ""),
                        "source_field": "客户字段映射",
                        "rule_type": "辅助客户字段映射",
                    }
                )
            continue
        steps[step_key] = value
        if field == "glue_category_code":
            steps["glue_category"] = "普通" if value == "Y" else "特殊"
        elif field == "tc_code":
            steps["order_type"] = "芯厚" if value == "C" else "总厚"
        if error_text:
            errors[:] = [item for item in errors if item != error_text and not item.startswith(error_text)]
            steps["errors"] = errors
        semantic_mapping = "模型语义" in str(row.get("来源批次", "") or "")
        applied.append(
            {
                "rule_id": row.get("映射ID", ""),
                "field": field,
                "old": old_value,
                "new": value,
                "text": row.get("规则文本", ""),
                "source": "已批准模型语义映射" if semantic_mapping else "正确码回归客户字段映射",
                "source_row": row.get("来源行号", ""),
                "source_field": "客户字段映射",
                "rule_type": "已批准模型语义规则/受控映射" if semantic_mapping else "辅助客户字段映射",
            }
        )
        already_overridden.add(field)
    return applied


def _apply_agent_size_mappings(
    engine,
    mapping_tables: dict[str, list[dict]],
    customer_code: str,
    customer_name: str,
    spec: str,
    context: str,
    steps: dict,
    errors: list[str],
    applied_rules: list[dict],
) -> list[dict]:
    if any(item.get("field") == "size_code" for item in applied_rules):
        return []
    combined_text = f"{spec or ''} {context or ''}"

    for row in mapping_tables.get("客户尺寸映射", []):
        if not _mapping_row_enabled(row) or not _mapping_customer_matches(row, customer_code, customer_name):
            continue
        customer_w = _to_float(row.get("客户尺寸W"))
        customer_h = _to_float(row.get("客户尺寸H"))
        factory_w = _to_float(row.get("厂内尺寸W"))
        factory_h = _to_float(row.get("厂内尺寸H"))
        if None in (customer_w, customer_h, factory_w, factory_h):
            continue
        if not (_steps_size_matches(steps, customer_w, customer_h) or _text_size_pair_matches(combined_text, customer_w, customer_h)):
            continue
        return [
            _apply_size_mapping_override(
                engine,
                steps,
                errors,
                row,
                factory_w,
                factory_h,
                "Agent尺寸映射/完整尺寸",
            )
        ]

    single_side_rows = [
        row for row in mapping_tables.get("客户单边尺寸映射", [])
        if _mapping_row_enabled(row) and _mapping_customer_matches(row, customer_code, customer_name)
    ]
    if single_side_rows:
        side_map = {
            _norm_size_token(row.get("客户单边尺寸")): _to_float(row.get("厂内单边尺寸"))
            for row in single_side_rows
            if _norm_size_token(row.get("客户单边尺寸")) and _to_float(row.get("厂内单边尺寸")) is not None
        }
        raw_pair = _extract_raw_size_pair(combined_text)
        if raw_pair:
            left_raw, right_raw = raw_pair
            left_mapped = side_map.get(_norm_size_token(left_raw))
            right_mapped = side_map.get(_norm_size_token(right_raw))
            if left_mapped is not None or right_mapped is not None:
                factory_w = left_mapped if left_mapped is not None else _to_float(steps.get("size_w"))
                factory_h = right_mapped if right_mapped is not None else _to_float(steps.get("size_h"))
                if factory_w is not None and factory_h is not None:
                    source_row = next(
                        row for row in single_side_rows
                        if _norm_size_token(row.get("客户单边尺寸")) in {_norm_size_token(left_raw), _norm_size_token(right_raw)}
                    )
                    return [
                        _apply_size_mapping_override(
                            engine,
                            steps,
                            errors,
                            source_row,
                            factory_w,
                            factory_h,
                            "Agent尺寸映射/单边尺寸",
                        )
                    ]

    for row in mapping_tables.get("客户尺寸算法", []):
        if not _mapping_row_enabled(row) or not _mapping_customer_matches(row, customer_code, customer_name):
            continue
        if row.get("算法类型") != "尺寸加大":
            continue
        current_w = _to_float(steps.get("size_w"))
        current_h = _to_float(steps.get("size_h"))
        delta_w = _to_float(row.get("加大W"))
        delta_h = _to_float(row.get("加大H"))
        if None in (current_w, current_h, delta_w, delta_h):
            continue
        # A size already ending in the configured enlargement must not be
        # enlarged a second time (e.g. 41.3*49.3 with a +0.3 rule).
        if _size_already_enlarged(current_w, delta_w) and _size_already_enlarged(current_h, delta_h):
            continue
        return [
            _apply_size_mapping_override(
                engine,
                steps,
                errors,
                row,
                current_w + delta_w,
                current_h + delta_h,
                "Agent尺寸映射/尺寸算法",
            )
        ]

    for row in mapping_tables.get("外部尺寸表引用", []):
        if not _mapping_row_enabled(row) or not _mapping_customer_matches(row, customer_code, customer_name):
            continue
        applied = _apply_external_size_reference(engine, row, combined_text, steps, errors)
        if applied:
            return [applied]
    return []


def _apply_size_mapping_override(engine, steps: dict, errors: list[str], row: dict, width: float, height: float, source: str) -> dict:
    old_value = str(steps.get("step4_size_code", "") or "")
    new_value = engine.size_to_code(width, height)
    steps["size_w"] = round(float(width), 2)
    steps["size_h"] = round(float(height), 2)
    steps["step4_size_code"] = new_value
    steps["size_note"] = f"{source}：{_format_size(width)}x{_format_size(height)} -> {new_value}"
    errors[:] = [item for item in errors if item != "无法识别尺寸" and not item.startswith("无法识别尺寸")]
    steps["errors"] = errors
    return {
        "rule_id": row.get("映射ID", ""),
        "field": "size_code",
        "old": old_value,
        "new": new_value,
        "text": row.get("规则文本", ""),
        "source": source,
        "source_row": row.get("来源行号", ""),
        "source_field": row.get("来源字段", ""),
        "rule_type": "辅助尺寸映射",
    }


def _apply_external_size_reference(engine, row: dict, combined_text: str, steps: dict, errors: list[str]) -> dict | None:
    if "新美亚" not in str(row.get("规则文本", "") + row.get("引用文件", "")):
        return None
    raw_pair = _extract_raw_size_pair(combined_text)
    if not raw_pair:
        return None
    raw_w = _to_float(raw_pair[0])
    raw_h = _to_float(raw_pair[1])
    if raw_w is None or raw_h is None:
        return None

    if _material_code_matches(combined_text, "631"):
        converted = _lookup_xinmeiya_631_size(row, raw_w, raw_h)
        source_label = "Agent尺寸映射/外部尺寸表-新美亚631"
    elif _material_code_matches(combined_text, "632"):
        converted = _convert_xinmeiya_632_size(raw_w, raw_h)
        source_label = "Agent尺寸映射/外部尺寸表-新美亚632"
    else:
        return None
    if not converted:
        return None

    mm_w, mm_h = converted
    width = mm_w / 25.4
    height = mm_h / 25.4
    applied = _apply_size_mapping_override(engine, steps, errors, row, width, height, source_label)
    applied["text"] = f"{raw_pair[0]}*{raw_pair[1]} -> {int(mm_w)}*{int(mm_h)}mm -> {_format_size(width)}*{_format_size(height)}"
    applied["rule_type"] = "辅助外部尺寸表映射"
    steps["size_note"] = f"{source_label}：{applied['text']} -> {applied['new']}"
    return applied


def _lookup_xinmeiya_631_size(row: dict, width: float, height: float) -> tuple[float, float] | None:
    reference_path = _resolve_external_reference_path(row.get("引用文件", ""))
    if not reference_path:
        return None
    for entry in _load_xinmeiya_size_table(str(reference_path)):
        if _near_size(entry["inch_w"], width) and _near_size(entry["inch_h"], height):
            return entry["mm_w"], entry["mm_h"]
    return None


def _convert_xinmeiya_632_size(width: float, height: float) -> tuple[float, float]:
    return _round_half_up(width * 25.4), _round_half_up(height * 25.4)


def _round_half_up(value: float) -> int:
    return int(math.floor(float(value) + 0.5))


def _resolve_external_reference_path(value: Any) -> Path | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    path = Path(raw)
    if not path.is_absolute():
        path = PROJECT_DIR / path
    return path if path.exists() else None


@lru_cache(maxsize=8)
def _load_xinmeiya_size_table(path: str) -> tuple[dict[str, float], ...]:
    workbook_rows = _read_xlsx_first_sheet_rows(Path(path))
    entries: list[dict[str, float]] = []
    in_631_section = False
    for row in workbook_rows:
        row_text = " ".join(str(value or "") for value in row)
        if "631." in row_text and "Thin Core" in row_text:
            in_631_section = True
            continue
        if in_631_section and ("620." in row_text or "Double Side" in row_text or "632." in row_text):
            break
        if not in_631_section or len(row) < 5:
            continue
        inch_w = _to_float(row[1])
        inch_h = _to_float(row[2])
        mm_w = _to_float(row[3])
        mm_h = _to_float(row[4])
        if None in (inch_w, inch_h, mm_w, mm_h):
            continue
        entries.append({"inch_w": inch_w, "inch_h": inch_h, "mm_w": mm_w, "mm_h": mm_h})
    return tuple(entries)


def _read_xlsx_first_sheet_rows(path: Path) -> list[list[str]]:
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path) as archive:
        shared_strings = _read_xlsx_shared_strings(archive, namespace)
        sheet_xml = _read_xlsx_member(archive, "xl/worksheets/sheet1.xml")
        root = ElementTree.fromstring(sheet_xml)
        rows: list[list[str]] = []
        for row in root.findall(".//x:row", namespace):
            values: dict[int, str] = {}
            for cell in row.findall("x:c", namespace):
                ref = cell.attrib.get("r", "")
                value_node = cell.find("x:v", namespace)
                value = ""
                if value_node is not None:
                    raw_value = value_node.text or ""
                    if cell.attrib.get("t") == "s" and raw_value.isdigit():
                        value = shared_strings[int(raw_value)] if int(raw_value) < len(shared_strings) else ""
                    else:
                        value = raw_value
                values[_column_index_from_cell_ref(ref)] = value
            if values:
                rows.append([values.get(idx, "") for idx in range(max(values) + 1)])
    return rows


def _read_xlsx_shared_strings(archive: ZipFile, namespace: dict[str, str]) -> list[str]:
    try:
        root = ElementTree.fromstring(_read_xlsx_member(archive, "xl/sharedStrings.xml"))
    except KeyError:
        return []
    values: list[str] = []
    for item in root.findall("x:si", namespace):
        values.append("".join(text_node.text or "" for text_node in item.findall(".//x:t", namespace)))
    return values


def _read_xlsx_member(archive: ZipFile, normalized_name: str) -> bytes:
    by_normalized_name = {name.replace("\\", "/"): name for name in archive.namelist()}
    actual_name = by_normalized_name.get(normalized_name)
    if not actual_name:
        raise KeyError(normalized_name)
    return archive.read(actual_name)


def _column_index_from_cell_ref(cell_ref: str) -> int:
    letters = "".join(ch for ch in str(cell_ref or "") if ch.isalpha())
    value = 0
    for letter in letters:
        value = value * 26 + ord(letter.upper()) - ord("A") + 1
    return max(value - 1, 0)


def _apply_agent_thickness_mappings(
    engine,
    mapping_tables: dict[str, list[dict]],
    customer_code: str,
    customer_name: str,
    spec: str,
    context: str,
    steps: dict,
    errors: list[str],
    applied_rules: list[dict],
) -> list[dict]:
    combined_text = f"{spec or ''} {context or ''}"
    applied: list[dict] = []
    has_thickness_override = any(item.get("field") == "thickness_code" for item in applied_rules)
    has_tc_override = any(item.get("field") == "tc_code" for item in applied_rules)

    for row in mapping_tables.get("客户厚度映射", []):
        if not _mapping_row_enabled(row) or not _mapping_customer_matches(row, customer_code, customer_name):
            continue
        alias = str(row.get("客户厚度写法", "") or "").strip()
        mm_value = _to_float(row.get("厚度mm"))
        mil_value = _to_float(row.get("厚度mil"))
        tc_code = str(row.get("总芯厚口径", "") or "").strip().upper()

        if mm_value is not None and alias and not has_thickness_override and _thickness_alias_matches(combined_text, alias):
            applied.append(_apply_thickness_mapping_override(engine, steps, errors, row, mm_value, mil_value))
            has_thickness_override = True
            if mil_value is not None and mil_value < 31 and not has_tc_override:
                core_row = dict(row)
                core_row["规则文本"] = f"{row.get('规则文本', '')}；{_format_size(mil_value)}mil<31mil按芯厚"
                applied.append(_apply_tc_mapping_override(steps, core_row, "C"))
                has_tc_override = True

        if tc_code in {"T", "C"} and not has_tc_override and _thickness_tc_condition_matches(combined_text, steps, mil_value, alias):
            applied.append(_apply_tc_mapping_override(steps, row, tc_code))
            has_tc_override = True
    return applied


def _apply_agent_material_code_mappings(
    mapping_tables: dict[str, list[dict]],
    customer_code: str,
    customer_name: str,
    spec: str,
    context: str,
    steps: dict,
    applied_rules: list[dict],
) -> list[dict]:
    if any(item.get("field") == "tc_code" for item in applied_rules):
        return []
    combined_text = f"{spec or ''} {context or ''}"
    for row in mapping_tables.get("客户物料编码口径", []):
        if not _mapping_row_enabled(row) or not _mapping_customer_matches(row, customer_code, customer_name):
            continue
        tc_code = str(row.get("总芯厚口径", "") or "").strip().upper()
        match_value = str(row.get("命中值", "") or "").strip()
        if tc_code not in {"T", "C"} or not match_value:
            continue
        if not _material_code_matches(combined_text, match_value):
            continue
        return [
            _apply_tc_mapping_override(
                steps,
                row,
                tc_code,
                source="Agent物料编码口径",
                rule_type="辅助物料编码口径",
            )
        ]
    return []


def _apply_thickness_mapping_override(engine, steps: dict, errors: list[str], row: dict, mm_value: float, mil_value: float | None) -> dict:
    old_value = str(steps.get("step2_thick_code", "") or "")
    new_value = engine.thickness_to_code(mm_value)
    steps["thickness_raw"] = row.get("客户厚度写法", "")
    steps["thickness_mm"] = mm_value
    steps["thickness_unit"] = "Agent客户厚度映射"
    steps["thickness_mode_source"] = "Agent客户厚度映射"
    note_parts = [str(row.get("规则文本", "") or "").strip()]
    if mil_value is not None:
        note_parts.append(f"{_format_size(mil_value)}mil")
    note_parts.append(f"{_format_size(mm_value)}mm")
    steps["thickness_mode_note"] = " -> ".join(item for item in note_parts if item)
    steps["order_mm"] = mm_value
    steps["step2_thick_code"] = new_value
    errors[:] = [item for item in errors if item != "无法识别厚度" and not item.startswith("无法识别厚度")]
    steps["errors"] = errors
    return {
        "rule_id": row.get("映射ID", ""),
        "field": "thickness_code",
        "old": old_value,
        "new": new_value,
        "text": row.get("规则文本", ""),
        "source": "Agent厚度映射",
        "source_row": row.get("来源行号", ""),
        "source_field": row.get("来源字段", ""),
        "rule_type": "辅助厚度映射",
    }


def _apply_tc_mapping_override(
    steps: dict,
    row: dict,
    tc_code: str,
    *,
    source: str = "Agent总芯厚映射",
    rule_type: str = "辅助厚度映射",
) -> dict:
    old_value = str(steps.get("step8_tc_code", "") or "")
    steps["step8_tc_code"] = tc_code
    steps["order_type"] = "总厚" if tc_code == "T" else "芯厚"
    steps["thickness_mode"] = steps["order_type"]
    steps["thickness_mode_source"] = "Agent客户厚度映射"
    if row.get("规则文本"):
        steps["thickness_mode_note"] = row.get("规则文本", "")
    return {
        "rule_id": row.get("映射ID", ""),
        "field": "tc_code",
        "old": old_value,
        "new": tc_code,
        "text": row.get("规则文本", ""),
        "source": source,
        "source_row": row.get("来源行号", ""),
        "source_field": row.get("来源字段", ""),
        "rule_type": rule_type,
    }


def _thickness_alias_matches(text: str, alias: str) -> bool:
    raw_alias = str(alias or "").strip()
    if not raw_alias:
        return False
    source = str(text or "")
    mil_alias = re.fullmatch(r"(\d+(?:\.\d+)?)\s*mil", raw_alias, flags=re.IGNORECASE)
    if mil_alias:
        value = re.escape(mil_alias.group(1))
        return bool(re.search(
            rf"(?<![\d.]){value}\s*(?:±\s*\d+(?:\.\d+)?\s*)?mil(?![A-Z0-9])",
            source,
            flags=re.IGNORECASE,
        ))
    if re.fullmatch(r"\d+(?:\.\d+)?", raw_alias):
        return bool(re.search(rf"(?<!\d){re.escape(raw_alias)}(?!\d)", source))
    return bool(re.search(rf"(?<![A-Z0-9]){re.escape(raw_alias)}(?![A-Z0-9])", source, flags=re.IGNORECASE))


def _size_already_enlarged(value: float, delta: float) -> bool:
    if delta <= 0:
        return False
    fractional = round(float(value) - int(float(value)), 2)
    return abs(fractional - round(float(delta), 2)) <= 0.01


def _material_code_matches(text: str, value: str) -> bool:
    raw_value = str(value or "").strip()
    if not raw_value:
        return False
    source = str(text or "")
    if re.fullmatch(r"\d+", raw_value):
        return bool(re.search(rf"(?<!\d){re.escape(raw_value)}(?!\d)", source))
    return bool(re.search(rf"(?<![A-Z0-9]){re.escape(raw_value)}(?![A-Z0-9])", source, flags=re.IGNORECASE))


def _thickness_tc_condition_matches(text: str, steps: dict, mil_value: float | None, alias: str) -> bool:
    actual_mil = _extract_mil_from_text(text)
    if actual_mil is None:
        current_mm = _to_float(steps.get("thickness_mm"))
        actual_mil = current_mm / 0.0254 if current_mm is not None else None
    if mil_value is not None and "含以上" in str(alias):
        return actual_mil is not None and actual_mil >= mil_value - 1e-9
    if mil_value is not None:
        return actual_mil is not None and abs(actual_mil - mil_value) <= 0.01
    return _thickness_alias_matches(text, alias)


def _extract_mil_from_text(text: str) -> float | None:
    matches = re.findall(r"(?<![A-Z0-9])(\d+(?:\.\d+)?)\s*MIL(?![A-Z0-9])", str(text or ""), flags=re.IGNORECASE)
    if not matches:
        return None
    try:
        return float(matches[-1])
    except ValueError:
        return None


def _mapping_row_enabled(row: dict) -> bool:
    value = str(row.get("启用", "")).strip().upper()
    return value in {"是", "Y", "YES", "TRUE", "1"}


def _expand_customer_rule_identity(
    mapping_tables: dict[str, list[dict]],
    customer_code: str,
    customer_name: str,
) -> tuple[str, str, str]:
    rows = [row for row in mapping_tables.get("客户规则组", []) if _mapping_row_enabled(row)]
    current_codes = set(_validated_customer_code_tokens(customer_code))
    current_name = _norm_match(customer_name)
    matched_groups: set[str] = set()
    for row in rows:
        row_codes = set(_customer_code_tokens(row.get("客户代码", "")))
        row_name = _norm_match(row.get("客户简称", ""))
        code_match = bool(current_codes and row_codes and not current_codes.isdisjoint(row_codes))
        name_match = bool(current_name and row_name and (current_name in row_name or row_name in current_name))
        if code_match or name_match:
            group_id = str(row.get("规则组ID", "") or "").strip()
            if group_id:
                matched_groups.add(group_id)
    if not matched_groups:
        return customer_code, customer_name, ""
    members = [
        row for row in rows
        if str(row.get("规则组ID", "") or "").strip() in matched_groups
    ]
    codes = list(dict.fromkeys(
        token for row in members for token in _customer_code_tokens(row.get("客户代码", ""))
    ))
    names = list(dict.fromkeys(
        str(row.get("客户简称", "") or "").strip() for row in members
        if str(row.get("客户简称", "") or "").strip()
    ))
    group_names = list(dict.fromkeys(
        str(row.get("规则组名称", "") or "").strip() for row in members
        if str(row.get("规则组名称", "") or "").strip()
    ))
    return "，".join(codes), " ".join(names), "，".join(group_names)


def _apply_agent_global_special_rules(
    mapping_tables: dict[str, list[dict]],
    spec: str,
    context: str,
    steps: dict,
    errors: list[str],
) -> list[dict]:
    """Apply approved special rules whose customer scope is all customers."""
    # Preserve separators so short business keywords such as TFT are matched
    # as independent tokens instead of being joined to adjacent model text.
    combined = f"{spec or ''} {context or ''}".upper()
    current_glue = _norm_match(steps.get("glue_model", ""))
    for row in mapping_tables.get("Agent基础条件规则", []):
        if not _mapping_row_enabled(row):
            continue
        if str(row.get("物料类别", "") or "").strip().upper() != "CCL":
            continue
        condition_glue = _norm_match(row.get("条件胶系", ""))
        # Use the exact model token in the uploaded specification as the
        # decisive condition. The parser may enrich glue_model with business
        # descriptors such as "汽车板"; that must not prevent an approved
        # NY3150HF condition from matching. The token boundary still prevents
        # NY3150HF from accidentally matching NY3150HFP/NY3150HFIST.
        if condition_glue and not _flexible_model_token_matches(condition_glue, combined):
            continue
        keyword = str(row.get("条件关键词", "") or "").strip().upper()
        mode = str(row.get("关键词模式", "") or "").strip()
        keyword_present = bool(keyword and _token_with_alnum_boundary_matches(keyword, combined))
        if mode == "包含" and not keyword_present:
            continue
        if mode == "不包含" and keyword_present:
            continue

        # Some historical customer-order rows encoded the old 3B choice and
        # its linked AT grade together. The latest rule makes AT conditional
        # on TFT, so clear only that stale pair before customer grade rules run.
        if (
            mode == "不包含"
            and str(steps.get("step1_glue_code", "") or "").upper() == "3B"
            and str(steps.get("step7_grade_code", "") or "").upper() == "AT"
        ):
            steps["step7_grade_code"] = "A1"

        applied: list[dict] = []
        overrides = (
            ("glue_code", "step1_glue_code", row.get("覆盖胶系代码", "")),
            ("glue_category_code", "step5_glue_cat_code", row.get("覆盖胶水类别", "")),
            ("grade_code", "step7_grade_code", row.get("覆盖基板级别", "")),
        )
        for field, step_key, raw_value in overrides:
            value = str(raw_value or "").strip().upper()
            if not value:
                continue
            old_value = str(steps.get(step_key, "") or "")
            steps[step_key] = value
            if field == "glue_category_code":
                steps["glue_category"] = "普通" if value == "Y" else "特殊"
            error_text = OVERRIDE_STEP_MAP[field][1]
            if error_text:
                errors[:] = [item for item in errors if item != error_text and not item.startswith(error_text)]
            applied.append({
                "rule_id": row.get("映射ID", ""),
                "field": field,
                "old": old_value,
                "new": value,
                "text": row.get("规则文本", ""),
                "source": "全客户特殊规则",
                "source_row": row.get("来源批次", ""),
                "source_field": "全客户特殊规则",
                "rule_type": "业务确认全客户特殊规则",
            })
        return applied
    return []


def _mapping_customer_matches(row: dict, customer_code: str, customer_name: str) -> bool:
    rule = {"客户代码": row.get("客户代码", ""), "客户简称": row.get("客户简称", "")}
    return _customer_matches(rule, customer_code, customer_name)


def _steps_size_matches(steps: dict, width: float, height: float) -> bool:
    current_w = _to_float(steps.get("size_w"))
    current_h = _to_float(steps.get("size_h"))
    return current_w is not None and current_h is not None and _near_size(current_w, width) and _near_size(current_h, height)


def _text_size_pair_matches(text: str, width: float, height: float) -> bool:
    for left, right in _extract_all_raw_size_pairs(text):
        left_num = _to_float(left)
        right_num = _to_float(right)
        if left_num is not None and right_num is not None and _near_size(left_num, width) and _near_size(right_num, height):
            return True
    return False


def _extract_raw_size_pair(text: str) -> tuple[str, str] | None:
    pairs = _extract_all_raw_size_pairs(text)
    return pairs[-1] if pairs else None


def _extract_all_raw_size_pairs(text: str) -> list[tuple[str, str]]:
    source = str(text or "").replace("×", "*").replace("x", "*").replace("X", "*")
    explicit = re.findall(r"(?<!\d)(\d{2,4}(?:\.\d+)?)\s*\*\s*(\d{2,4}(?:\.\d+)?)(?!\d)", source)
    loose = re.findall(r"(?<!\d)(\d{2,4}(?:\.\d+)?)\s+(\d{2,4}(?:\.\d+)?)(?!\d)", source)
    loose = [
        pair for pair in loose
        if pair not in explicit
        and not (len(pair[0].split(".")[0]) == 4 and len(pair[1].split(".")[0]) == 4)
    ]
    return explicit + loose


def _norm_size_token(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    if re.fullmatch(r"\d+(?:\.0+)?", raw):
        return str(int(float(raw)))
    return raw


def _to_float(value: Any) -> float | None:
    try:
        raw = str(value or "").strip()
        if not raw:
            return None
        return float(raw)
    except (TypeError, ValueError):
        return None


def _near_size(left: float, right: float) -> bool:
    return abs(float(left) - float(right)) <= 0.01


def _format_size(value: float) -> str:
    return f"{float(value):.2f}".rstrip("0").rstrip(".")


def _rule_priority(rule: dict) -> int:
    try:
        return int(float(rule.get("优先级") or 0))
    except (TypeError, ValueError):
        return 0


def _rule_executable(rule: dict) -> bool:
    return (
        str(rule.get("启用", "")).strip() == "是"
        and str(rule.get("待确认", "")).strip() != "是"
        and str(rule.get("强制执行", "")).strip() == "是"
        and str(rule.get("物料类别", "")).strip() == "CCL"
        and rule.get("覆盖字段") in OVERRIDE_STEP_MAP
        and rule.get("覆盖字段") in AGENT_EXECUTABLE_OVERRIDE_FIELDS
        and bool(str(rule.get("覆盖值", "") or "").strip())
    )


def _rule_matches(rule: dict, customer_code: str, customer_name: str, spec: str, context: str, steps: dict) -> bool:
    if not _customer_matches(rule, customer_code, customer_name):
        return False
    combined = _norm_match(f"{spec} {context}")
    spec_norm = _norm_match(spec)
    glue_norm = _norm_match(steps.get("glue_model", ""))
    copper_norm = _norm_match(steps.get("copper_spec_raw", ""))
    thickness_norm = _norm_match(steps.get("thickness_raw", ""))
    size_norm = _norm_match(f"{steps.get('size_w', '')}*{steps.get('size_h', '')}")
    raw_condition_text = str(rule.get("条件文本", "") or rule.get("规则文本", "") or "")
    has_positive_tolerance = bool(re.search(r"\+\s*\d|±|偏正", str(spec or "") + " " + str(context or "")))
    if "有偏正公差" in raw_condition_text and not has_positive_tolerance:
        return False
    if "无偏正公差" in raw_condition_text and has_positive_tolerance:
        return False

    glue_condition = _norm_match(rule.get("条件胶系", ""))
    if glue_condition:
        glue_tokens = [item for item in re.split(r"[/,，;；]+", glue_condition) if item]
        if glue_tokens and not any(_glue_condition_matches(token, glue_norm, spec_norm) for token in glue_tokens):
            return False

    copper_condition = _norm_match(rule.get("条件铜厚", ""))
    if copper_condition:
        copper_tokens = _copper_condition_tokens(copper_condition)
        if copper_tokens and not any(_copper_condition_matches(token, copper_norm, spec_norm) for token in copper_tokens):
            return False

    thickness_condition = str(rule.get("条件厚度", "") or "").strip()
    if thickness_condition:
        thickness_tokens = [item for item in re.split(r"[/,，;；]+", thickness_condition) if item.strip()]
        if thickness_tokens and not any(_thickness_condition_matches(token, steps, spec_norm, thickness_norm) for token in thickness_tokens):
            return False

    size_condition = _norm_match(rule.get("条件尺寸", ""))
    if size_condition:
        size_tokens = [item for item in re.split(r"[/,，;；]+", size_condition) if item]
        if size_tokens and not any(token in size_norm or token in spec_norm or token in combined for token in size_tokens):
            return False

    keyword_condition = str(rule.get("条件关键词", "") or "").strip()
    if keyword_condition:
        if _is_legacy_copper_ff_keyword_condition(rule, keyword_condition):
            if not _legacy_copper_ff_keyword_matches(combined, steps):
                return False
        else:
            keywords = [item for item in re.split(r"[/,，、;；]+", keyword_condition) if item.strip()]
            if keywords and not any(_keyword_condition_matches(keyword, combined) for keyword in keywords):
                return False

    condition_text = _norm_match(rule.get("条件文本", ""))
    if condition_text and not (glue_condition or copper_condition or thickness_condition or size_condition or keyword_condition):
        field_label = _norm_match(rule.get("原始字段", ""))
        if field_label and field_label in condition_text:
            return True
        relaxed = re.sub(r"^(当|如果)|时$", "", condition_text)
        if relaxed and relaxed not in combined:
            return False
    return True


def _customer_matches(rule: dict, customer_code: str, customer_name: str) -> bool:
    if bool(rule.get("_global_rule")):
        return True
    rule_codes = set(_customer_code_tokens(rule.get("客户代码", "")))
    current_codes = set(_validated_customer_code_tokens(customer_code))
    rule_name = _norm_match(rule.get("客户简称", ""))
    current_name = _norm_match(customer_name)
    if rule_codes:
        if current_codes and rule_codes.isdisjoint(current_codes):
            return False
        # Some order templates do not contain a customer-code column. In that
        # case an exact customer-name match is the only safe fallback; loose
        # aliases could apply one customer's rules to another customer.
        if not current_codes and (not rule_name or rule_name != current_name):
            return False
    if rule_name:
        if not current_name or not customer_names_match(rule.get("客户简称", ""), customer_name):
            return False
    return bool(rule_codes or rule_name)


def _build_field_evidence(steps: dict, errors: list[str], applied_rules: list[dict], conflicts: list[str]) -> list[dict]:
    by_field = {item["field"]: item for item in applied_rules}
    evidence = []
    for key, label, step_key, gate in FIELD_DEFS:
        code_value = str(steps.get(step_key, "") or "")
        override_field = FIELD_KEY_TO_OVERRIDE[key]
        score, hit_type, source, note = _score_field(key, code_value, steps, errors, by_field, conflicts)
        evidence.append(
            {
                "field_key": key,
                "field": label,
                "value": _field_raw_value(key, steps),
                "code": code_value,
                # 结构码是当前阶段的输出占位符，只留证据，不展示或参与置信度评分。
                "score": score if gate else None,
                "gate": gate,
                "hit_type": hit_type,
                "source": source,
                "evidence": note,
                "rule_id": by_field.get(override_field, {}).get("rule_id", ""),
                "rule_type": by_field.get(override_field, {}).get("rule_type", ""),
                "source_row": by_field.get(override_field, {}).get("source_row", ""),
            }
        )
    return evidence


def _score_field(key: str, code_value: str, steps: dict, errors: list[str], by_field: dict, conflicts: list[str]) -> tuple[int, str, str, str]:
    override_field = FIELD_KEY_TO_OVERRIDE[key]
    field_conflicts = _conflicts_for_field(key, conflicts)
    if field_conflicts:
        return 60, "规则冲突", "Agent规则包", "; ".join(field_conflicts)
    if override_field in by_field:
        rule = by_field[override_field]
        source = str(rule.get("source") or "")
        rule_type = str(rule.get("rule_type") or "")
        rule_id = str(rule.get("rule_id") or "")
        text = str(rule.get("text") or "")
        if (
            source.startswith("正确码回归客户字段映射")
            or "历史" in source
            or rule_type == "辅助客户字段映射"
        ):
            return 99, "历史样本建议", rule_id, (
                f"{text}；仅有历史正确码样本支持，未经业务正式规则确认不能单独正式出码"
            )
        if "模型" in source:
            return 98, "模型语义标准化", rule_id, (
                f"{text}；模型仅做语义标准化，不能单独构成100分正式出码依据"
            )
        if source.startswith("最新版胶系名称优先命中"):
            return 99, "胶系候选口径", rule_id, (
                f"{text}；优先候选不等于唯一业务结果，需确认同名多码口径"
            )
        formal_rule_types = {
            "确认草稿机器规则",
            "确认草稿机器规则/胶系联动",
            "Agent胶系主数据映射",
            "辅助尺寸映射",
            "辅助外部尺寸表映射",
            "辅助厚度映射",
            "辅助物料编码口径",
            "业务确认全客户特殊规则",
            "客户人工长期规则",
            "明确订单指令",
        }
        formal_source_prefixes = (
            "已确认人工长期规则",
            "已批准语义规则",
            "Agent尺寸映射",
            "Agent厚度映射",
            "Agent总芯厚映射",
            "Agent物料编码口径",
            "全客户特殊规则",
            "已确认明确订单指令",
            "新版胶系编号精确命中",
            "最新版胶系名称精确命中",
            "Agent胶系兼容别名",
        )
        if rule_id and (rule_type in formal_rule_types or source.startswith(formal_source_prefixes)):
            return 100, "业务正式规则", rule_id, text
        return 99, "规则来源待确认", rule_id, (
            f"{text}；该覆盖来源未列入业务正式规则，需人工确认"
        )
    if _field_has_error(key, errors) or _is_placeholder(code_value):
        return 0, "未识别", "基础解析", "; ".join(errors)
    if key == "structure" and code_value == "*":
        return 0, "占位符", "首版策略", "结构码为*占位符，不参与置信度评分和正式出码拦截"
    confirmed_base = _score_confirmed_base_field(key, code_value, steps)
    if confirmed_base:
        return confirmed_base
    return 99, "解析来源待确认", "未登记来源", (
        f"已解析为{code_value}，但缺少可追溯的业务基础映射或确定性算法来源"
    )


def _score_confirmed_base_field(
    key: str,
    code_value: str,
    steps: dict[str, Any],
) -> tuple[int, str, str, str] | None:
    glue_model = str(steps.get("agent_glue_name") or steps.get("glue_model") or "").strip()
    thickness_raw = str(steps.get("thickness_raw") or "").strip()
    copper_raw = str(steps.get("copper_spec_raw") or "").strip()
    has_size = steps.get("size_w") not in (None, "") and steps.get("size_h") not in (None, "")

    if key == "glue" and glue_model:
        source = str(steps.get("agent_glue_source") or "transcode_rules.xlsx/胶系代码")
        if "优先命中" in source or steps.get("agent_glue_uncertain"):
            return None
        return 100, "正式基础映射", source, f"{glue_model}→{code_value}"
    if key == "thickness" and thickness_raw and steps.get("thickness_mm") not in (None, ""):
        source = str(steps.get("thickness_mode_source") or steps.get("thickness_unit") or "")
        return 100, "确定性厚度算法", source or "NYG-ATD-002-A1/厚度编码", thickness_raw
    if key == "copper" and copper_raw:
        return 100, "正式基础映射", "transcode_rules.xlsx/铜箔规格", copper_raw
    if key == "size" and has_size:
        note = str(steps.get("size_note") or "标准尺寸")
        return 100, "确定性尺寸映射", "transcode_rules.xlsx/尺寸规则", (
            f"{steps.get('size_w')}x{steps.get('size_h')}；{note}"
        )
    if key == "glue_category" and glue_model and str(steps.get("glue_category") or "").strip():
        return 100, "正式基础映射", "transcode_rules.xlsx/胶水类别", (
            f"{steps.get('glue_category')}→{code_value}"
        )
    if key == "grade":
        grade_note = str(steps.get("grade_note") or "").strip()
        input_text = str(steps.get("agent_input_text") or "").upper()
        if grade_note:
            return 100, "业务确定性等级规则", "transcode_rules.xlsx/客户下单与胶系基板转换", grade_note
        if code_value == "AC" and (
            "汽车板" in glue_model
            or any(keyword in input_text for keyword in ("汽车板", "汽车专用"))
        ):
            return 100, "业务确定性等级规则", "BASE-GRADE-AUTOMOTIVE-AC", (
                "规格明确包含汽车板/汽车专用，按已确认等级AC"
            )
        return None
    if key == "total_core" and thickness_raw and str(steps.get("order_type") or "").strip():
        source = str(steps.get("thickness_mode_source") or "NYG-ATD-002-A1/总芯厚判断")
        return 100, "确定性总芯厚算法", source, str(steps.get("order_type"))
    if key == "copper_type" and (copper_raw or glue_model or thickness_raw or has_size):
        if code_value == "W":
            return 100, "已确认默认规则", "BASE-DEFAULT-COPPER-TYPE-W", (
                "未命中特殊铜箔类型时按已确认常规铜HTE/W默认"
            )
        return 100, "正式基础映射", "transcode_rules.xlsx/铜箔类型", code_value
    return None


def _conflicts_for_field(key: str, conflicts: list[str]) -> list[str]:
    override_field = FIELD_KEY_TO_OVERRIDE.get(key, "")
    labels = {
        "glue": ("胶系", "胶系代码", "glue_code", "最新版胶系主表"),
        "thickness": ("厚度", "基板厚度", "thickness_code"),
        "copper": ("铜厚", "铜箔规格", "copper_code"),
        "size": ("尺寸", "基板尺寸", "size_code"),
        "glue_category": ("胶水类别", "glue_category_code"),
        "copper_type": ("铜箔类型", "印字", "copper_type_code"),
        "grade": ("基板级别", "等级", "grade_code"),
        "total_core": ("总/芯厚", "总厚", "芯厚", "tc_code"),
        "structure": ("结构码", "struct_code"),
    }.get(key, ())
    matched: list[str] = []
    for conflict in conflicts:
        text = str(conflict or "")
        normalized = text.lower()
        if override_field and normalized.startswith(f"{override_field.lower()}:"):
            matched.append(text)
        elif any(label.lower() in normalized for label in labels):
            matched.append(text)
    return matched


def _refresh_analysis_after_semantic_overrides(
    analysis: dict[str, Any],
    semantic_applied: list[dict[str, Any]],
    semantic_conflicts: list[str],
) -> None:
    steps = analysis.get("engine_steps") or {}
    errors = list(steps.get("errors") or [])
    applied_rules = list(analysis.get("applied_rules") or []) + semantic_applied
    conflicts = list(analysis.get("conflicts") or []) + semantic_conflicts
    _enforce_retired_glue_runtime_guard(steps, errors, conflicts)
    candidate_code = _build_code_from_steps(steps, errors)
    field_evidence = _build_field_evidence(steps, errors, applied_rules, conflicts)
    decision = decide_confirmation(
        errors=errors,
        conflicts=conflicts,
        candidate_code=candidate_code,
        field_evidence=field_evidence,
    )
    status = decision["status"]
    formal_code = decision["formal_code"]
    overall_score = decision["overall_score"]
    reason = decision["reason"]
    analysis.update(
        {
            "status": status,
            "formal_code": formal_code,
            "candidate_code": candidate_code,
            "overall_score": overall_score,
            "reason": reason,
            "field_evidence": field_evidence,
            "applied_rules": applied_rules,
            "conflicts": conflicts,
            "decision_state": decision["decision_state"],
            "confirmation_triggers": decision["confirmation_triggers"],
            "summary": _format_agent_summary(status, formal_code or candidate_code, overall_score, reason, applied_rules),
        }
    )


def _apply_explicit_order_instructions(
    analysis: dict[str, Any],
    order_remark: str,
    tables: dict[str, Any],
) -> None:
    remark = str(order_remark or "").strip().upper()
    if not remark:
        return
    grade_codes = {
        match.group(1).upper()
        for match in re.finditer(
            r"(?:基板级别|板级|等级)\s*(?:下|为|=|：|:)?\s*"
            r"(A1|A2|AC|AD|AH|AL|AM|AP|AT|AY|F1)",
            remark,
        )
    }
    valid_codes = {
        str(code or "").strip().upper()
        for code in (tables.get("grade_code_map") or {}).keys()
    } | set(OFFICIAL_GRADE_CODES)
    grade_codes &= valid_codes
    if not grade_codes:
        return
    if len(grade_codes) > 1:
        conflicts = list(analysis.get("conflicts") or [])
        conflicts.append(f"基板级别: 订单备注同时指定多个等级 {'/'.join(sorted(grade_codes))}")
        _refresh_analysis_after_semantic_overrides(analysis, [], conflicts)
        return
    code = next(iter(grade_codes))
    steps = analysis.get("engine_steps") or {}
    old = str(steps.get("step7_grade_code") or "")
    steps["step7_grade_code"] = code
    applied = {
        "rule_id": "ORDER-EXPLICIT-GRADE-001",
        "field": "grade_code",
        "old": old,
        "new": code,
        "text": f"订单备注明确指定基板级别={code}",
        "source": "已确认明确订单指令",
        "source_row": "",
        "source_field": "基板级别",
        "source_column": "订单备注",
        "rule_type": "明确订单指令",
    }
    _refresh_analysis_after_semantic_overrides(analysis, [applied], [])


def _apply_runtime_confirmation_rules(
    analysis: dict[str, Any],
    rules: list[dict[str, Any]],
) -> None:
    if not rules or analysis.get("status") == "失败":
        return
    apply_confirmation_rules_to_evidence(analysis.get("field_evidence") or [], rules)
    decision = decide_confirmation(
        errors=list((analysis.get("engine_steps") or {}).get("errors") or []),
        conflicts=list(analysis.get("conflicts") or []),
        candidate_code=str(analysis.get("candidate_code") or ""),
        field_evidence=analysis.get("field_evidence") or [],
    )
    analysis.update(decision)
    analysis["confirmation_policy_rules"] = rules
    analysis["summary"] = _format_agent_summary(
        analysis["status"],
        analysis.get("candidate_code", ""),
        analysis.get("overall_score", 0),
        analysis["reason"],
        analysis.get("applied_rules") or [],
    )


def _field_raw_value(key: str, steps: dict) -> str:
    if key == "glue":
        return str(steps.get("glue_model", ""))
    if key == "thickness":
        return str(steps.get("thickness_raw", ""))
    if key == "copper":
        return str(steps.get("copper_spec_raw", ""))
    if key == "size":
        return f"{steps.get('size_w', '')}x{steps.get('size_h', '')}".strip("x")
    if key == "glue_category":
        return str(steps.get("glue_category", ""))
    if key == "total_core":
        return str(steps.get("order_type", ""))
    return ""


def _field_has_error(key: str, errors: list[str]) -> bool:
    text = ";".join(errors)
    return {
        "glue": "胶系" in text,
        "thickness": "厚度" in text,
        "copper": "铜箔规格" in text,
        "size": "尺寸" in text,
    }.get(key, False)


def _build_code_from_steps(steps: dict, errors: list[str]) -> str:
    if errors:
        return ""
    if str(steps.get("step1_glue_code") or "").strip().upper() == "2Z" and any(
        is_retired_agent_glue_mapping(
            {"胶系名称": steps.get(name_key), "输出胶系代码": "2Z"}
        )
        for name_key in ("glue_model", "agent_glue_name", "raw_glue")
    ):
        return ""
    parts = [
        str(steps.get("step1_glue_code", "") or ""),
        str(steps.get("step2_thick_code", "") or ""),
        str(steps.get("step3_copper_code", "") or ""),
        str(steps.get("step4_size_code", "") or ""),
        str(steps.get("step5_glue_cat_code", "") or ""),
        str(steps.get("step6_copper_type_code", "") or ""),
        str(steps.get("step7_grade_code", "") or ""),
        str(steps.get("step8_tc_code", "") or ""),
        str(steps.get("step9_struct_code", "") or ""),
    ]
    if any(_is_placeholder(part) for part in parts[:-1]):
        return ""
    suffix = "XXXXXX" if str(steps.get("step5_glue_cat_code", "")) == "R" else ""
    return "".join(parts) + suffix


def _build_confirmation_items(analyses: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for analysis in analyses:
        if analysis.get("status") != "待确认":
            continue
        low_evidence = [
            item
            for item in analysis.get("field_evidence") or []
            if item.get("gate") and int(item.get("score") or 0) < FIELD_GATE_THRESHOLD
        ]
        if not low_evidence:
            field_key, field_label = _infer_confirmation_field(analysis)
            low_evidence = [
                {
                    "field_key": field_key,
                    "field": field_label,
                    "code": _confirmation_current_code(analysis, field_key),
                    "score": int(analysis.get("overall_score") or 0),
                    "hit_type": "整行风险确认",
                    "source": "100分正式码门禁",
                    "evidence": analysis.get("reason") or "",
                    "gate": True,
                }
            ]
        for evidence in low_evidence:
            field_key = str(evidence.get("field_key") or "row_review")
            current_code = str(
                evidence.get("code")
                or _confirmation_current_code(analysis, field_key)
                or ""
            ).strip()
            items.append(
                {
                    "excel_row": int(analysis.get("row") or 0),
                    "customer_code": analysis.get("customer_code") or "",
                    "customer_name": analysis.get("customer") or "",
                    "spec": analysis.get("spec") or "",
                    "context_text": analysis.get("context_text") or "",
                    "field_key": field_key,
                    "field_label": evidence.get("field") or "整行码值",
                    "current_code": current_code,
                    "options": _confirmation_options(analysis, field_key, current_code),
                    "pending_code": analysis.get("candidate_code") or "",
                    "score": int(evidence.get("score") or 0),
                    "reason": evidence.get("evidence") or analysis.get("reason") or "",
                    "evidence": evidence,
                    "analysis": analysis,
                }
            )
    return items


def _infer_confirmation_field(analysis: dict[str, Any]) -> tuple[str, str]:
    text = "；".join(
        [
            str(analysis.get("reason") or ""),
            "；".join(str(item) for item in analysis.get("conflicts") or []),
            "；".join(
                str(item.get("field") or item.get("reason") or "")
                for item in analysis.get("confirmation_triggers") or []
            ),
        ]
    )
    candidates = (
        ("grade", "基板级别", ("基板级别", "等级", "A1", "AC", "AT", "AP", "AY")),
        ("glue", "胶系", ("胶系", "胶系代码", "型号")),
        ("thickness", "厚度", ("基板厚度", "厚度")),
        ("copper", "铜厚", ("铜箔规格", "铜厚")),
        ("size", "尺寸", ("基板尺寸", "尺寸")),
        ("glue_category", "胶水类别", ("胶水类别",)),
        ("copper_type", "铜箔类型", ("铜箔类型", "印字")),
        ("total_core", "总/芯厚", ("总/芯厚", "总厚", "芯厚")),
    )
    for field_key, label, keywords in candidates:
        if any(keyword in text for keyword in keywords):
            return field_key, label
    return "row_review", "整行码值"


def _confirmation_current_code(analysis: dict[str, Any], field_key: str) -> str:
    if field_key == "row_review":
        return str(analysis.get("candidate_code") or "")
    step_key = FIELD_STEP_KEYS.get(field_key)
    return str((analysis.get("engine_steps") or {}).get(step_key, "") or "")


def _confirmation_options(
    analysis: dict[str, Any],
    field_key: str,
    current_code: str,
) -> list[str]:
    if field_key == "row_review":
        return [current_code] if current_code else []
    width = FIELD_CODE_WIDTHS.get(field_key, 0)
    values: list[str] = []
    if current_code:
        values.append(current_code.upper())
    text = " ".join(
        [
            str(analysis.get("reason") or ""),
            " ".join(str(item) for item in analysis.get("conflicts") or []),
        ]
    ).upper()
    for token in re.findall(r"(?<![A-Z0-9])[A-Z0-9-]{1,8}(?![A-Z0-9])", text):
        if len(token) == width and token not in values:
            values.append(token)
    return values


def list_transcode_agent_confirmations(
    job_id: int,
    employee_id: str,
    *,
    record_scope: str = "all",
    record_page: int = 1,
    record_page_size: int = 200,
) -> dict[str, Any]:
    job = get_job(job_id)
    if not job or job["feature"] != FEATURE_KEY or job["employee_id"] != employee_id:
        raise LookupError("未找到营销转码Agent任务")
    rows = list_transcode_agent_confirmation_items(job_id, employee_id)
    items = []
    analyses_by_confirmation_row: dict[int, dict[str, Any]] = {}
    for row in rows:
        item = dict(row)
        excel_row = int(item["excel_row"])
        analysis = _json_loads(item.get("analysis_json"), {})
        analyses_by_confirmation_row[excel_row] = analysis
        item["decision_score"] = int(analysis.get("overall_score") or item.get("score") or 0)
        item["decision_state"] = str(analysis.get("decision_state") or "")
        item["decision_reason"] = str(analysis.get("reason") or item.get("reason") or "")
        candidate_codes: list[str] = []
        for candidate in (analysis.get("engine_steps") or {}).get("agent_glue_candidates") or []:
            code = str(candidate.get("code") or "").strip().upper()
            if code and code not in candidate_codes:
                candidate_codes.append(code)
        item["decision_options"] = candidate_codes
        item["options"] = _json_loads(item.pop("options_json", "[]"), [])
        item["evidence"] = _json_loads(item.pop("evidence_json", "{}"), {})
        item.pop("analysis_json", None)
        items.append(item)
    records = _load_transcode_agent_trace_records(str(job["stored_result_path"] or ""))
    item_statuses_by_row: dict[int, set[str]] = {}
    for item in items:
        item_statuses_by_row.setdefault(int(item["excel_row"]), set()).add(str(item["status"] or ""))
    records_by_row = {int(record["excel_row"]): record for record in records}
    workbook_refreshed = False
    for excel_row, statuses in item_statuses_by_row.items():
        record = records_by_row.get(excel_row, {})
        if (
            "confirmed" in statuses
            and "pending" not in statuses
            and str(record.get("transcode_status") or "") != "人工已确认"
        ):
            confirmed_fields = [
                f"{item['field_label']}={item.get('confirmed_code') or item.get('current_code') or ''}"
                for item in items
                if int(item["excel_row"]) == excel_row and item["status"] == "confirmed"
            ]
            _update_confirmation_workbook_row(
                str(job["stored_result_path"] or ""),
                excel_row,
                analyses_by_confirmation_row.get(excel_row, {}),
                confirmation_note="人工已确认：" + "、".join(confirmed_fields),
            )
            workbook_refreshed = True
    if workbook_refreshed:
        records = _load_transcode_agent_trace_records(str(job["stored_result_path"] or ""))
    for record in records:
        record["record_state"] = _trace_record_state(
            record,
            item_statuses_by_row.get(int(record["excel_row"]), set()),
        )
    state_order = {
        "pending": 0,
        "confirmed": 1,
        "automatic": 2,
        "deferred": 3,
        "failed": 4,
        "skipped": 5,
    }
    records.sort(key=lambda item: (state_order.get(str(item.get("record_state") or ""), 9), int(item["excel_row"])))
    scope_states = {
        "pending": {"pending"},
        "confirmed": {"confirmed"},
        "automatic": {"automatic"},
        "exceptions": {"deferred", "failed", "skipped"},
    }
    if record_scope in scope_states:
        scoped_records = [record for record in records if record.get("record_state") in scope_states[record_scope]]
    else:
        record_scope = "all"
        scoped_records = records
    record_page_size = max(20, min(int(record_page_size or 200), 500))
    record_page_count = max(1, math.ceil(len(scoped_records) / record_page_size))
    record_page = max(1, min(int(record_page or 1), record_page_count))
    start = (record_page - 1) * record_page_size
    record_counts = Counter(str(record.get("record_state") or "") for record in records)
    counts = transcode_agent_confirmation_counts(job_id)
    pending_rows = len({item["excel_row"] for item in items if item["status"] == "pending"})
    return {
        "job_id": job_id,
        "status": job["status"],
        "success_count": int(job["success_count"] or 0),
        "fail_count": int(job["fail_count"] or 0),
        "skip_count": int(job["skip_count"] or 0),
        "confirm_count": pending_rows,
        "counts": counts,
        "items": items,
        "records": scoped_records[start : start + record_page_size],
        "record_scope": record_scope,
        "record_page": record_page,
        "record_page_size": record_page_size,
        "record_page_count": record_page_count,
        "record_total": len(scoped_records),
        "record_counts": dict(record_counts),
    }


def _load_transcode_agent_trace_records(output_path: str) -> list[dict[str, Any]]:
    path = Path(output_path)
    if not output_path or not path.exists():
        return []
    stat = path.stat()
    try:
        return [
            dict(record)
            for record in _cached_transcode_agent_trace_records(
                str(path),
                stat.st_mtime_ns,
                stat.st_size,
            )
        ]
    except (BadZipFile, EOFError, OSError):
        _cached_transcode_agent_trace_records.cache_clear()
        return []


@lru_cache(maxsize=16)
def _cached_transcode_agent_trace_records(
    output_path: str,
    _modified_ns: int,
    _file_size: int,
) -> tuple[dict[str, Any], ...]:
    workbook = openpyxl.load_workbook(output_path, data_only=True, read_only=True)
    worksheet = workbook["转码需求表"] if "转码需求表" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    headers = {
        _normalize_semantic_header(cell.value): int(cell.column)
        for cell in worksheet[1]
        if _normalize_semantic_header(cell.value)
    }
    customer_code_col = _trace_column(headers, ("客户编号", "客户代码"))
    customer_name_col = _trace_column(headers, ("客户简称", "客户名称", "客户"))
    spec_col = _trace_column(headers, ("客户规格", "规格", "单条客户规格"))
    formal_col = headers.get(_normalize_semantic_header(FORMAL_RESULT_HEADER))
    pending_col = headers.get(_normalize_semantic_header(PENDING_RESULT_HEADER))
    comparison_col = headers.get(_normalize_semantic_header(OUTPUT_STATUS_HEADER))
    status_col = headers.get(_normalize_semantic_header(TRANSCODE_STATUS_HEADER))
    confirmation_col = headers.get(_normalize_semantic_header(CONFIRMATION_HEADER))
    system_col = headers.get(_normalize_semantic_header(SYSTEM_ANALYSIS_HEADER))
    context_cols = [
        column
        for header, column in headers.items()
        if header in {"订单备注", "备注", "整行上下文", "备注/整行上下文"}
    ]

    evidence_by_row = _trace_evidence_by_row(workbook)
    records: list[dict[str, Any]] = []
    for excel_row, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        customer_code = _trace_row_value(row_values, customer_code_col)
        customer_name = _trace_row_value(row_values, customer_name_col)
        spec = _trace_row_value(row_values, spec_col)
        formal_code = _trace_row_value(row_values, formal_col)
        pending_code = _trace_row_value(row_values, pending_col)
        transcode_status = _trace_row_value(row_values, status_col)
        if not any((customer_code, customer_name, spec, formal_code, pending_code, transcode_status)):
            continue
        context_text = "\n".join(
            value
            for value in (_trace_row_value(row_values, column) for column in context_cols)
            if value
        )
        trace_items = evidence_by_row.get(excel_row, [])
        scores = [
            int(item.get("score") or 0)
            for item in trace_items
            if item.get("score") is not None and str(item.get("score") or "").isdigit()
        ]
        records.append(
            {
                "excel_row": excel_row,
                "customer_code": customer_code,
                "customer_name": customer_name,
                "spec": spec,
                "context_text": context_text,
                "formal_code": formal_code,
                "pending_code": pending_code,
                "display_code": formal_code or pending_code,
                "comparison": _trace_row_value(row_values, comparison_col),
                "transcode_status": transcode_status,
                "confirmation_note": _trace_row_value(row_values, confirmation_col),
                "system_reason": _trace_row_value(row_values, system_col),
                "score": min(scores) if scores else 0,
                "trace_items": trace_items,
            }
        )
    return tuple(records)


def _trace_evidence_by_row(workbook) -> dict[int, list[dict[str, Any]]]:
    if "字段证据链" not in workbook.sheetnames:
        return {}
    worksheet = workbook["字段证据链"]
    headers = {
        _normalize_semantic_header(cell.value): int(cell.column)
        for cell in worksheet[1]
        if _normalize_semantic_header(cell.value)
    }
    row_col = headers.get("行号")
    evidence_by_row: dict[int, list[dict[str, Any]]] = {}
    for source_row, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            excel_row = int(float(_trace_row_value(row_values, row_col)))
        except (TypeError, ValueError):
            continue
        field_label = _trace_row_value(row_values, headers.get("字段")) or "转码证据"
        code = _trace_row_value(row_values, headers.get("编码片段"))
        score_text = _trace_row_value(row_values, headers.get("置信度"))
        try:
            score = int(float(score_text))
        except (TypeError, ValueError):
            score = 0
        hit_type = _trace_row_value(row_values, headers.get("命中方式"))
        source = _trace_row_value(row_values, headers.get("规则来源"))
        evidence = _trace_row_value(row_values, headers.get("证据"))
        rule_id = _trace_row_value(row_values, headers.get("规则ID"))
        is_structure = field_label == "结构码"
        evidence_by_row.setdefault(excel_row, []).append(
            {
                "id": f"trace-{excel_row}-{source_row}",
                "status": "record",
                "field_key": "structure" if is_structure else "trace",
                "field_label": f"{field_label}（首次解析）",
                "current_code": code,
                "confirmed_code": "",
                "score": None if is_structure else score,
                "reason": "；".join(value for value in (hit_type, source, evidence) if value),
                "evidence": {
                    "hit_type": hit_type,
                    "source": source,
                    "evidence": evidence,
                    "rule_id": rule_id,
                },
                "is_trace": True,
            }
        )
    return evidence_by_row


def _trace_column(headers: dict[str, int], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        column = headers.get(_normalize_semantic_header(alias))
        if column:
            return column
    return None


def _trace_row_value(row_values: tuple[Any, ...], column: int | None) -> str:
    if not column or column > len(row_values):
        return ""
    value = row_values[column - 1]
    return "" if value is None else str(value).strip()


def _trace_record_state(record: dict[str, Any], item_statuses: set[str]) -> str:
    if "pending" in item_statuses:
        return "pending"
    if "confirmed" in item_statuses or "人工已确认" in str(record.get("transcode_status") or ""):
        return "confirmed"
    if "skipped" in item_statuses:
        return "deferred"
    status = str(record.get("transcode_status") or "")
    if status == "跳过":
        return "skipped"
    if status == "未出码":
        return "failed"
    return "automatic"


def reevaluate_transcode_agent_confirmations(job_id: int, employee_id: str) -> dict[str, Any]:
    job = get_job(job_id)
    if not job or job["employee_id"] != employee_id or job["feature"] != FEATURE_KEY:
        raise LookupError("未找到营销转码Agent任务")
    pending_items = list_transcode_agent_confirmation_items(job_id, employee_id, status="pending")
    rows: dict[int, list[Any]] = {}
    for item in pending_items:
        rows.setdefault(int(item["excel_row"]), []).append(item)

    runtime = _load_runtime()
    resolved_rows = 0
    unresolved_rows = 0
    resolved_analyses: list[tuple[int, dict[str, Any]]] = []
    for excel_row, row_items in rows.items():
        first = row_items[0]
        analysis, _base_version, _agent_version = _calculate_transcode_agent_analysis(
            str(first["spec"] or ""),
            customer=str(first["customer_name"] or ""),
            customer_code=str(first["customer_code"] or ""),
            order_remark=str(first["context_text"] or ""),
            employee_id=employee_id,
            runtime=runtime,
        )
        analysis.update(
            {
                "row": excel_row,
                "customer_code": str(first["customer_code"] or ""),
                "customer": str(first["customer_name"] or ""),
                "spec": str(first["spec"] or ""),
                "context_text": str(first["context_text"] or ""),
            }
        )
        if analysis.get("status") == "成功" and int(analysis.get("overall_score") or 0) == 100:
            for item in row_items:
                field_key = str(item["field_key"] or "")
                current_code = _confirmation_current_code(analysis, field_key) or str(item["current_code"] or "")
                update_transcode_agent_confirmation_item(
                    int(item["id"]),
                    status="auto_resolved",
                    confirmed_code=current_code,
                    confirmation_basis="按当前活动规则重新评估达到100分",
                    confirmed_by="系统规则重评",
                    analysis=analysis,
                )
            update_transcode_agent_row_analysis(job_id, excel_row, analysis)
            resolved_analyses.append((excel_row, analysis))
            resolved_rows += 1
        else:
            evidence_by_key = {
                str(evidence.get("field_key") or ""): evidence
                for evidence in analysis.get("field_evidence") or []
            }
            for item in row_items:
                field_key = str(item["field_key"] or "")
                evidence = evidence_by_key.get(field_key, {})
                refresh_transcode_agent_confirmation_item(
                    int(item["id"]),
                    current_code=str(evidence.get("code") or item["current_code"] or ""),
                    pending_code=str(analysis.get("candidate_code") or ""),
                    score=int(evidence.get("score") or analysis.get("overall_score") or 0),
                    reason=str(evidence.get("evidence") or analysis.get("reason") or item["reason"] or ""),
                    evidence=evidence,
                    analysis=analysis,
                )
            unresolved_rows += 1

    if resolved_analyses:
        _update_automatic_reevaluation_workbook_rows(
            str(job["stored_result_path"] or ""),
            resolved_analyses,
        )
    _refresh_confirmation_job_status(job_id, newly_formal=resolved_rows)
    _refresh_confirmation_audit_sheet(job_id, str(job["stored_result_path"] or ""))
    append_job_log(
        job_id,
        f"按当前规则重新评估：{resolved_rows}行达到100分，{unresolved_rows}行仍待确认",
    )
    return {
        "job_id": job_id,
        "resolved_rows": resolved_rows,
        "unresolved_rows": unresolved_rows,
        "remaining": transcode_agent_confirmation_counts(job_id)["pending"],
        "status": get_job(job_id)["status"],
    }


def confirm_transcode_agent_item(
    item_id: int,
    employee_id: str,
    *,
    confirmed_code: str,
    basis: str = "",
    save_long_term: bool = False,
    long_term_rule: dict[str, Any] | None = None,
) -> dict[str, Any]:
    item = get_transcode_agent_confirmation_item(item_id)
    if not item or item["employee_id"] != employee_id:
        raise LookupError("未找到待确认项")
    if item["status"] != "pending":
        raise ValueError("该待确认项已处理")
    job = get_job(int(item["job_id"]))
    if not job or job["employee_id"] != employee_id or job["feature"] != FEATURE_KEY:
        raise LookupError("未找到营销转码Agent任务")

    analysis = _json_loads(item["analysis_json"], {})
    field_key = str(item["field_key"] or "")
    normalized_code = _validate_confirmation_code(
        field_key,
        confirmed_code or item["current_code"],
        analysis,
    )
    saved_rule: dict[str, Any] | None = None
    if save_long_term:
        saved_rule = _build_long_term_confirmation_rule(
            item,
            normalized_code,
            basis=basis,
            payload=long_term_rule or {},
        )
        validate_customer_maintained_rule(saved_rule)
        save_rule_override(
            saved_rule,
            updated_by=employee_id,
            previous_rule=None,
        )
    _apply_manual_confirmation_to_analysis(
        analysis,
        field_key,
        normalized_code,
        employee_id,
        basis,
    )
    update_transcode_agent_confirmation_item(
        item_id,
        status="confirmed",
        confirmed_code=normalized_code,
        confirmation_basis=basis,
        confirmed_by=employee_id,
        analysis=analysis,
        long_term_rule_id=(saved_rule or {}).get("rule_id") if saved_rule else None,
    )

    row_items = [
        row
        for row in list_transcode_agent_confirmation_items(int(item["job_id"]), employee_id)
        if int(row["excel_row"]) == int(item["excel_row"])
    ]
    pending_row_items = [row for row in row_items if row["status"] == "pending"]
    if pending_row_items:
        analysis["status"] = "待确认"
        analysis["formal_code"] = ""
        analysis["reason"] = "仍有字段待人工确认：" + "、".join(
            str(row["field_label"]) for row in pending_row_items
        )
    else:
        analysis["status"] = "成功"
        analysis["formal_code"] = analysis.get("candidate_code") or ""
        analysis["reason"] = ""
        analysis["overall_score"] = min(
            [
                int(evidence.get("score") or 0)
                for evidence in analysis.get("field_evidence") or []
                if evidence.get("gate")
            ]
            or [0]
        )
    analysis["summary"] = _format_agent_summary(
        analysis["status"],
        analysis.get("formal_code") or analysis.get("candidate_code") or "",
        int(analysis.get("overall_score") or 0),
        analysis.get("reason") or "",
        analysis.get("applied_rules") or [],
    )
    update_transcode_agent_row_analysis(
        int(item["job_id"]),
        int(item["excel_row"]),
        analysis,
    )
    newly_formal = _update_confirmation_workbook_row(
        str(job["stored_result_path"] or ""),
        int(item["excel_row"]),
        analysis,
        confirmation_note=f"当前任务人工确认：{item['field_label']}={normalized_code}"
        + (f"；依据：{basis}" if basis else ""),
    )
    _refresh_confirmation_job_status(int(item["job_id"]), newly_formal=newly_formal)
    if saved_rule:
        append_job_log(
            int(item["job_id"]),
            f"人工确认已保存为长期原子规则：{saved_rule['rule_id']}",
        )
    return {
        "item_id": item_id,
        "job_id": int(item["job_id"]),
        "excel_row": int(item["excel_row"]),
        "status": analysis["status"],
        "formal_code": analysis.get("formal_code") or "",
        "pending_code": "" if analysis["status"] == "成功" else analysis.get("candidate_code") or "",
        "overall_score": int(analysis.get("overall_score") or 0),
        "remaining": transcode_agent_confirmation_counts(int(item["job_id"]))["pending"],
        "long_term_rule_id": (saved_rule or {}).get("rule_id", ""),
    }


def finalize_transcode_agent_confirmations(job_id: int, employee_id: str) -> dict[str, Any]:
    job = get_job(job_id)
    if not job or job["employee_id"] != employee_id or job["feature"] != FEATURE_KEY:
        raise LookupError("未找到营销转码Agent任务")
    pending_items = list_transcode_agent_confirmation_items(job_id, employee_id, status="pending")
    rows: dict[int, dict[str, Any]] = {}
    for item in pending_items:
        analysis = _json_loads(item["analysis_json"], {})
        rows[int(item["excel_row"])] = analysis
        update_transcode_agent_confirmation_item(
            int(item["id"]),
            status="skipped",
            confirmed_code=None,
            confirmation_basis="暂不处理并生成结果",
            confirmed_by=employee_id,
            analysis=analysis,
        )
    for excel_row, analysis in rows.items():
        analysis["status"] = "待确认"
        analysis["formal_code"] = ""
        update_transcode_agent_row_analysis(job_id, excel_row, analysis)
        _update_confirmation_workbook_row(
            str(job["stored_result_path"] or ""),
            excel_row,
            analysis,
            confirmation_note="暂不处理：正式码保持为空，保留待人工确认码值",
        )
    _refresh_confirmation_audit_sheet(
        job_id,
        str(job["stored_result_path"] or ""),
    )
    update_job_status(
        job_id,
        status="completed",
        confirm_count=0,
        completed=True,
    )
    append_job_log(job_id, f"已暂不处理 {len(rows)} 行并生成结果", confirm_count=0)
    return {
        "job_id": job_id,
        "status": "completed",
        "skipped_rows": len(rows),
    }


def skip_transcode_agent_confirmation_row(
    item_id: int,
    employee_id: str,
) -> dict[str, Any]:
    item = get_transcode_agent_confirmation_item(item_id)
    if not item or item["employee_id"] != employee_id:
        raise LookupError("未找到待确认项")
    job_id = int(item["job_id"])
    excel_row = int(item["excel_row"])
    job = get_job(job_id)
    if not job or job["employee_id"] != employee_id or job["feature"] != FEATURE_KEY:
        raise LookupError("未找到营销转码Agent任务")
    row_items = [
        row
        for row in list_transcode_agent_confirmation_items(job_id, employee_id)
        if int(row["excel_row"]) == excel_row and row["status"] == "pending"
    ]
    analysis = _json_loads(item["analysis_json"], {})
    for row_item in row_items:
        update_transcode_agent_confirmation_item(
            int(row_item["id"]),
            status="skipped",
            confirmed_code=None,
            confirmation_basis="暂不处理当前行",
            confirmed_by=employee_id,
            analysis=analysis,
        )
    analysis["status"] = "待确认"
    analysis["formal_code"] = ""
    update_transcode_agent_row_analysis(job_id, excel_row, analysis)
    _update_confirmation_workbook_row(
        str(job["stored_result_path"] or ""),
        excel_row,
        analysis,
        confirmation_note="暂不处理：正式码保持为空，保留待人工确认码值",
    )
    _refresh_confirmation_job_status(job_id, newly_formal=False)
    return {
        "job_id": job_id,
        "excel_row": excel_row,
        "status": get_job(job_id)["status"],
        "remaining": transcode_agent_confirmation_counts(job_id)["pending"],
    }


def refresh_transcode_agent_audit_sheet(job_id: int) -> None:
    """Rebuild the confirmation audit sheet outside interactive confirmation."""
    job = get_job(job_id)
    if not job or job["feature"] != FEATURE_KEY:
        return
    _refresh_confirmation_audit_sheet(
        job_id,
        str(job["stored_result_path"] or ""),
    )


def _build_long_term_confirmation_rule(
    item: Any,
    confirmed_code: str,
    *,
    basis: str,
    payload: dict[str, Any],
) -> dict[str, Any]:
    field_key = str(item["field_key"] or "")
    if field_key not in LONG_TERM_RULE_FIELDS:
        raise CustomerRuleMaintenanceError("整行码值不能保存为长期规则，请按具体字段确认")
    if not bool(payload.get("second_confirmed")):
        raise CustomerRuleMaintenanceError("保存长期规则需要二次确认")
    if not str(basis or "").strip():
        raise CustomerRuleMaintenanceError("保存长期规则必须填写业务确认依据")
    condition_field = str(payload.get("condition_field") or "").strip()
    condition_operator = str(payload.get("condition_operator") or "").strip()
    condition_value = str(payload.get("condition_value") or "").strip()
    if not condition_field or not condition_operator:
        raise CustomerRuleMaintenanceError("保存长期规则必须填写适用条件")
    business_field, target_field = LONG_TERM_RULE_FIELDS[field_key]
    target_value = _long_term_target_value(field_key, confirmed_code)
    customer_code = str(item["customer_code"] or "").strip()
    customer_name = str(item["customer_name"] or "").strip()
    if not customer_code and customer_name:
        customer_code = resolve_customer_code_by_name(customer_name)
    source_text = str(payload.get("source_text") or "").strip() or (
        f"确认中心：{condition_field} {condition_operator} {condition_value}"
        f"时，{business_field}={target_value}"
    )
    rule = build_rule_from_form(
        {
            "rule_id": "",
            "customer_code": customer_code,
            "customer_name": customer_name,
            "business_field": business_field,
            "source_text": source_text,
            "target_field": target_field,
            "target_value": target_value,
            "condition_field": [condition_field],
            "condition_operator": [condition_operator],
            "condition_value": [condition_value],
            "priority": str(payload.get("priority") or "200"),
            "enabled": "1",
            "semantic_enabled": "1" if condition_field == "订单备注" else "0",
            "approval_basis": str(basis).strip(),
        }
    )
    rule["source_column"] = "确认中心"
    rule["model"] = "确认中心人工规则"
    rule["note"] = (
        f"确认中心长期规则；来源任务{item['job_id']}；Excel第{item['excel_row']}行；"
        "业务明确勾选并二次确认"
    )
    return rule


def _long_term_target_value(field_key: str, code: str) -> str:
    normalized = str(code or "").strip().upper()
    if field_key == "total_core":
        return "total" if normalized == "T" else "core"
    if field_key == "copper_type":
        return {"W": "HTE", "T": "RTF", "I": "IGAV"}.get(normalized, normalized)
    return normalized


def _validate_confirmation_code(
    field_key: str,
    code: str,
    analysis: dict[str, Any],
) -> str:
    normalized = re.sub(r"\s+", "", str(code or "").upper())
    if field_key == "row_review":
        candidate = str(analysis.get("candidate_code") or "")
        if not candidate or normalized != candidate.upper():
            raise ValueError("整行确认必须采用当前待人工确认码值")
        return candidate
    width = FIELD_CODE_WIDTHS.get(field_key)
    if not width:
        raise ValueError("不支持的确认字段")
    if len(normalized) != width or not re.fullmatch(r"[A-Z0-9-]+", normalized):
        label = next(
            (label for key, label, _step, _gate in FIELD_DEFS if key == field_key),
            field_key,
        )
        raise ValueError(f"{label}代码必须是{width}位有效代码")
    if field_key == "glue" and normalized == "2Z":
        steps = analysis.get("engine_steps") or {}
        if any(
            is_retired_agent_glue_mapping(
                {"胶系名称": steps.get(name_key), "输出胶系代码": normalized}
            )
            for name_key in ("glue_model", "agent_glue_name", "raw_glue")
        ):
            raise ValueError("NY-A1→2Z已废弃，不能作为人工确认胶系代码")
    return normalized


def _apply_manual_confirmation_to_analysis(
    analysis: dict[str, Any],
    field_key: str,
    confirmed_code: str,
    employee_id: str,
    basis: str,
) -> None:
    steps = analysis.get("engine_steps") or {}
    if field_key != "row_review":
        steps[FIELD_STEP_KEYS[field_key]] = confirmed_code
        for evidence in analysis.get("field_evidence") or []:
            if evidence.get("field_key") == field_key:
                evidence.update(
                    {
                        "code": confirmed_code,
                        "score": 100,
                        "hit_type": "当前任务人工确认",
                        "source": employee_id,
                        "evidence": basis or "业务在确认中心逐行确认",
                    }
                )
        analysis["conflicts"] = [
            conflict
            for conflict in analysis.get("conflicts") or []
            if conflict not in _conflicts_for_field(field_key, analysis.get("conflicts") or [])
        ]
    else:
        for evidence in analysis.get("field_evidence") or []:
            if evidence.get("gate"):
                evidence["score"] = 100
                evidence["hit_type"] = "当前任务整行人工确认"
                evidence["source"] = employee_id
                evidence["evidence"] = basis or "业务确认当前待人工确认码值"
        analysis["conflicts"] = []
        analysis.pop("confirmation_policy_rules", None)
        analysis["confirmation_triggers"] = []
    analysis["engine_steps"] = steps
    analysis["candidate_code"] = _build_code_from_steps(
        steps,
        list(steps.get("errors") or []),
    )
    analysis["overall_score"] = min(
        [
            int(evidence.get("score") or 0)
            for evidence in analysis.get("field_evidence") or []
            if evidence.get("gate")
        ]
        or [0]
    )


def _update_confirmation_workbook_row(
    output_path: str,
    excel_row: int,
    analysis: dict[str, Any],
    *,
    confirmation_note: str,
) -> bool:
    if not output_path or not Path(output_path).exists():
        raise FileNotFoundError("任务结果文件不存在")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["转码需求表"] if "转码需求表" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = {
        str(cell.value or "").strip(): int(cell.column)
        for cell in ws[1]
        if str(cell.value or "").strip()
    }
    result_col = headers.get(FORMAL_RESULT_HEADER)
    pending_col = headers.get(PENDING_RESULT_HEADER)
    difference_col = headers.get(CODE_DIFFERENCE_HEADER)
    comparison_col = headers.get(OUTPUT_STATUS_HEADER)
    status_col = headers.get(TRANSCODE_STATUS_HEADER)
    confirmation_col = headers.get(CONFIRMATION_HEADER)
    system_analysis_col = headers.get(SYSTEM_ANALYSIS_HEADER)
    if not all((result_col, pending_col, status_col, confirmation_col)):
        raise ValueError("结果文件缺少人工确认所需列")
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    formal_cell = ws.cell(row=excel_row, column=result_col)
    was_blank = not str(formal_cell.value or "").strip()
    if analysis.get("status") == "成功":
        formal_cell.value = analysis.get("formal_code") or analysis.get("candidate_code") or None
        formal_cell.fill = green_fill
        pending_cell = ws.cell(row=excel_row, column=pending_col)
        pending_cell.value = None
        pending_cell.fill = PatternFill(fill_type=None)
        if difference_col:
            ws.cell(row=excel_row, column=difference_col, value="人工已确认").fill = green_fill
        if comparison_col:
            ws.cell(row=excel_row, column=comparison_col, value="人工已确认").fill = green_fill
        ws.cell(row=excel_row, column=status_col, value="人工已确认").fill = green_fill
        if system_analysis_col:
            ws.cell(
                row=excel_row,
                column=system_analysis_col,
                value=confirmation_note,
            ).fill = green_fill
    else:
        formal_cell.value = None
        formal_cell.fill = red_fill
        ws.cell(
            row=excel_row,
            column=pending_col,
            value=analysis.get("candidate_code") or None,
        ).fill = red_fill
        if difference_col:
            ws.cell(row=excel_row, column=difference_col, value="待人工确认").fill = red_fill
        if comparison_col:
            ws.cell(row=excel_row, column=comparison_col, value="待人工确认").fill = red_fill
        ws.cell(row=excel_row, column=status_col, value="待人工确认").fill = red_fill
        if system_analysis_col:
            ws.cell(
                row=excel_row,
                column=system_analysis_col,
                value=confirmation_note,
            ).fill = red_fill
    ws.cell(
        row=excel_row,
        column=confirmation_col,
        value=confirmation_note,
    ).fill = red_fill if analysis.get("status") != "成功" else green_fill
    try:
        _atomic_save_workbook(wb, output_path)
    finally:
        wb.close()
    return bool(was_blank and analysis.get("status") == "成功")


def _update_automatic_reevaluation_workbook_row(
    output_path: str,
    excel_row: int,
    analysis: dict[str, Any],
) -> None:
    _update_automatic_reevaluation_workbook_rows(
        output_path,
        [(excel_row, analysis)],
    )


def _update_automatic_reevaluation_workbook_rows(
    output_path: str,
    rows: list[tuple[int, dict[str, Any]]],
) -> None:
    if not output_path or not Path(output_path).exists():
        raise FileNotFoundError("任务结果文件不存在")
    wb = openpyxl.load_workbook(output_path)
    ws = wb["转码需求表"] if "转码需求表" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = {
        str(cell.value or "").strip(): int(cell.column)
        for cell in ws[1]
        if str(cell.value or "").strip()
    }
    result_col = headers.get(FORMAL_RESULT_HEADER)
    pending_col = headers.get(PENDING_RESULT_HEADER)
    difference_col = headers.get(CODE_DIFFERENCE_HEADER)
    comparison_col = headers.get(OUTPUT_STATUS_HEADER)
    status_col = headers.get(TRANSCODE_STATUS_HEADER)
    confirmation_col = headers.get(CONFIRMATION_HEADER)
    system_analysis_col = headers.get(SYSTEM_ANALYSIS_HEADER)
    if not all((result_col, pending_col, status_col, confirmation_col)):
        raise ValueError("结果文件缺少规则重评所需列")
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    product_name_col = _find_product_name_column(ws)
    for excel_row, analysis in rows:
        product_name = (
            ws.cell(row=excel_row, column=product_name_col).value
            if product_name_col
            else ""
        )
        formal_code = str(analysis.get("formal_code") or analysis.get("candidate_code") or "")
        comparison = _export_result_comparison(analysis, product_name)
        comparison_failed = comparison is False
        fill = red_fill if comparison_failed else green_fill
        note = "系统按当前活动规则重新评估达到100分"
        ws.cell(row=excel_row, column=result_col, value=formal_code or None).fill = fill
        pending_cell = ws.cell(row=excel_row, column=pending_col)
        pending_cell.value = None
        pending_cell.fill = PatternFill(fill_type=None)
        if difference_col:
            ws.cell(
                row=excel_row,
                column=difference_col,
                value=_comparison_code_display(formal_code, product_name, comparison),
            ).fill = fill
        if comparison_col:
            ws.cell(row=excel_row, column=comparison_col, value=comparison).fill = fill
        ws.cell(
            row=excel_row,
            column=status_col,
            value="可直接采用",
        ).fill = green_fill
        ws.cell(
            row=excel_row,
            column=confirmation_col,
            value="无需人工确认：当前活动规则重评为100分",
        ).fill = green_fill
        if system_analysis_col:
            ws.cell(row=excel_row, column=system_analysis_col, value=note).fill = green_fill
    try:
        _atomic_save_workbook(wb, output_path)
    finally:
        wb.close()


def _refresh_confirmation_job_status(job_id: int, *, newly_formal: int | bool) -> None:
    job = get_job(job_id)
    if not job:
        return
    rows = list_transcode_agent_confirmation_items(job_id, job["employee_id"])
    pending_rows = {int(row["excel_row"]) for row in rows if row["status"] == "pending"}
    newly_formal_count = int(newly_formal or 0)
    success_count = int(job["success_count"] or 0) + newly_formal_count
    fail_count = max(0, int(job["fail_count"] or 0) - newly_formal_count)
    status = "awaiting_confirmation" if pending_rows else "completed"
    update_job_status(
        job_id,
        status=status,
        success_count=success_count,
        fail_count=fail_count,
        confirm_count=len(pending_rows),
        completed=not pending_rows,
    )
    append_job_log(
        job_id,
        f"人工确认后剩余 {len(pending_rows)} 行待处理",
        success_count=success_count,
        fail_count=fail_count,
        confirm_count=len(pending_rows),
    )


def _refresh_confirmation_audit_sheet(job_id: int, output_path: str) -> None:
    if not output_path or not Path(output_path).exists():
        return
    events = list_transcode_agent_confirmation_events(job_id)
    if not events:
        return
    workbook = openpyxl.load_workbook(output_path)
    sheet_name = "人工确认审计"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    headers = [
        "序号",
        "任务ID",
        "Excel行",
        "客户代码",
        "客户简称",
        "字段",
        "确认前代码",
        "确认后代码",
        "操作",
        "确认依据",
        "是否长期规则",
        "长期规则ID",
        "操作人",
        "操作时间",
    ]
    sheet.append(headers)
    for sequence, event in enumerate(events, start=1):
        before = _json_loads(event["before_json"], {})
        after = _json_loads(event["after_json"], {})
        item = after or before
        rule_id = str(after.get("long_term_rule_id") or "")
        action = str(event["action"] or "")
        sheet.append(
            [
                sequence,
                job_id,
                item.get("excel_row", ""),
                item.get("customer_code", ""),
                item.get("customer_name", ""),
                item.get("field_label", ""),
                before.get("current_code", ""),
                after.get("confirmed_code", "") or after.get("current_code", ""),
                "确认" if action == "confirmed" else "规则重评通过" if action == "auto_resolved" else "暂不处理",
                after.get("confirmation_basis", ""),
                "是" if rule_id else "否",
                rule_id,
                event["employee_id"],
                event["created_at"],
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [8, 10, 10, 14, 20, 16, 16, 16, 12, 38, 14, 32, 14, 24]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    try:
        _atomic_save_workbook(workbook, output_path)
    finally:
        workbook.close()


def _json_loads(value: Any, fallback: Any) -> Any:
    try:
        return json.loads(str(value or ""))
    except (TypeError, ValueError, json.JSONDecodeError):
        return fallback


def _is_placeholder(value: str) -> bool:
    return "?" in str(value or "")


def _format_agent_summary(status: str, code: str, score: int, reason: str, applied_rules: list[dict]) -> str:
    applied_text = "；".join(f"{item['rule_id']}:{item['field']}->{item['new']}" for item in applied_rules) or "无"
    if status == "成功":
        return f"高置信出码：{code}；最低关键字段分={score}；Agent规则={applied_text}"
    return f"{status}：{reason}；候选码={code or '无'}；最低关键字段分={score}；Agent规则={applied_text}"


def _detect_semantic_input_columns(df_req, spec_col: int) -> dict[str, dict[str, Any]]:
    headers = [str(value or "").strip() for value in df_req.iloc[0].tolist()]
    normalized = [_normalize_semantic_header(value) for value in headers]
    aliases = {
        "订单备注": {"订单备注", "备注", "订单说明", "订单附注"},
        "客户规格": {"客户规格"},
        "客户料品名称": {"客户料品名称", "客户料品", "料品名称", "品名"},
        "客户物料编码": {"客户物料编码", "客户产品编号", "客户料号"},
        "品号/物料编号": {"品号", "物料编号", "料号"},
    }
    result: dict[str, dict[str, Any]] = {}
    for field, names in aliases.items():
        normalized_names = {_normalize_semantic_header(name) for name in names}
        indices = [index for index, name in enumerate(normalized) if name in normalized_names]
        result[field] = {"indices": indices, "headers": [headers[index] for index in indices]}
    result["订单规格"] = {
        "indices": [spec_col],
        "headers": [headers[spec_col] if spec_col < len(headers) else "规格"],
    }
    return result


def _detect_parse_fallback_columns(df_req, spec_col: int) -> list[int]:
    """Find normalized spec columns used only after customer-spec parsing fails.

    A plain 品名 column is excluded because historical test files use it for the
    correct manufacturing code, which is not a parse input.
    """
    if df_req.empty:
        return []
    result: list[int] = []
    for row_idx in range(min(3, len(df_req))):
        for col_idx in range(len(df_req.columns)):
            if col_idx == spec_col:
                continue
            header = _normalize_semantic_header(df_req.iloc[row_idx, col_idx])
            if header in {"品名规格", "规格", "标准规格"} and col_idx not in result:
                result.append(col_idx)
    return result


def _detect_pp_context_columns(df_req, spec_col: int) -> list[int]:
    """Find normalized specification columns suitable for PP/RC classification.

    Correct-code/product-code columns are intentionally excluded because values such as
    RC can also be valid CCL glue codes and must not cause a false PP skip.
    """
    if df_req.empty:
        return []
    result: list[int] = []
    for row_idx in range(min(3, len(df_req))):
        for col_idx in range(len(df_req.columns)):
            if col_idx == spec_col:
                continue
            header = _normalize_semantic_header(df_req.iloc[row_idx, col_idx])
            if header in {"规格", "标准规格"} and col_idx not in result:
                result.append(col_idx)
    return result


def _build_semantic_observations(
    row,
    column_map: dict[str, dict[str, Any]],
    engine_steps: dict[str, Any],
) -> dict[str, dict[str, Any]]:
    observations = {
        field: _column_observation(row, metadata)
        for field, metadata in column_map.items()
    }
    observations["胶系"] = _derived_observation(
        engine_steps.get("glue_model"),
        "规格解析/胶系",
    )
    observations["基板厚度"] = _derived_observation(
        engine_steps.get("thickness_mm"),
        "规格解析/基板厚度",
    )
    observations["铜箔规格"] = _derived_observation(
        engine_steps.get("copper_spec_raw"),
        "规格解析/铜箔规格",
    )
    copper_top_oz, copper_bottom_oz = _copper_oz_pair(engine_steps.get("copper_spec_raw"))
    observations["铜箔上面oz"] = _derived_observation(
        copper_top_oz,
        "规格解析/铜箔上面oz",
    )
    observations["铜箔下面oz"] = _derived_observation(
        copper_bottom_oz,
        "规格解析/铜箔下面oz",
    )
    order_spec = observations.get("订单规格") or {"available": False, "value": "", "sources": []}
    order_remark = observations.get("订单备注") or {"available": False, "value": "", "sources": []}
    combined_values = [
        str(item.get("value") or "").strip()
        for item in [order_spec, order_remark]
        if item.get("available") and str(item.get("value") or "").strip()
    ]
    combined_sources = [
        source
        for item in [order_spec, order_remark]
        for source in item.get("sources") or []
    ]
    observations["订单规格/订单备注"] = {
        "available": bool(order_spec.get("available") or order_remark.get("available")),
        "value": " ".join(dict.fromkeys(combined_values)),
        "sources": list(dict.fromkeys(combined_sources)),
    }
    return observations


def _copper_oz_pair(value: Any) -> tuple[float | None, float | None]:
    raw = str(value or "").strip().upper().replace("OZ", "")
    match = re.search(r"(1\.5|0\.5|H|F|\d+(?:\.\d+)?)\s*/\s*(1\.5|0\.5|H|F|\d+(?:\.\d+)?)", raw)
    if not match:
        return None, None

    def to_oz(token: str) -> float | None:
        aliases = {"H": 0.5, "F": 1.5}
        if token in aliases:
            return aliases[token]
        try:
            return float(token)
        except ValueError:
            return None

    return to_oz(match.group(1)), to_oz(match.group(2))


def _column_observation(row, metadata: dict[str, Any]) -> dict[str, Any]:
    indices = list(metadata.get("indices") or [])
    selected_value = ""
    selected_source = ""
    headers = list(metadata.get("headers") or [])
    for index in indices:
        if index >= len(row):
            continue
        value = row.iloc[index]
        if pd.isna(value):
            continue
        text = str(value).strip()
        if text and text.lower() != "nan":
            selected_value = text
            position = indices.index(index)
            selected_source = headers[position] if position < len(headers) else ""
            break
    return {
        "available": bool(indices),
        "value": selected_value,
        "sources": [selected_source] if selected_source else [],
    }


def _derived_observation(value: Any, source: str) -> dict[str, Any]:
    available = value is not None and str(value).strip() not in {"", "nan", "None"}
    return {
        "available": available,
        "value": value if available else "",
        "sources": [source] if available else [],
    }


def _normalize_semantic_header(value: Any) -> str:
    return re.sub(r"[\s　]+", "", str(value or "")).strip().lower()


def _attach_semantic_shadow_metadata(
    analysis: dict,
    mode: str,
    version: str,
    rule_count: int,
    load_error: str,
    evaluations: list[dict],
) -> None:
    analysis["semantic_shadow_mode"] = mode
    analysis["semantic_rule_version"] = version
    analysis["semantic_rule_count"] = rule_count
    analysis["semantic_load_error"] = load_error
    analysis["semantic_shadow"] = evaluations


def _semantic_shadow_error_result(
    excel_row: int,
    customer_code: str,
    customer: str,
    spec: str,
    error: Exception,
) -> dict[str, Any]:
    return {
        "row": excel_row,
        "customer_code": customer_code,
        "customer": customer,
        "spec": spec,
        "rule_id": "",
        "source_candidate_id": "",
        "business_field": "",
        "target_fields": [],
        "normalized_values": [],
        "stated_target_values": [],
        "status": SHADOW_STATUS_ERROR,
        "missing_fields": [],
        "condition_results": [],
        "observed_inputs": {},
        "source_text": "",
        "evidence_texts": [],
        "model": "",
        "note": f"影子评估异常，不影响正式转码：{error}",
    }


def _save_agent_result(
    source_path: str,
    output_path: str,
    df_req,
    result_col: int,
    analyses: list[dict],
    agent_rules: list[dict],
    agent_mapping_tables: dict[str, list[dict]],
    confirm_count: int,
) -> None:
    wb = openpyxl.load_workbook(source_path)
    ws = wb["转码需求表"] if "转码需求表" in wb.sheetnames else wb[wb.sheetnames[0]]
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    result_excel_col = result_col + 1
    if ws.cell(row=1, column=result_excel_col).value == FORMAL_RESULT_HEADER and (
        ws.cell(row=1, column=result_excel_col + 1).value != PENDING_RESULT_HEADER
    ):
        ws.insert_cols(result_excel_col + 1, amount=1)

    pending_col = result_excel_col + 1
    difference_col = result_excel_col + 2
    comparison_col = result_excel_col + 3
    status_col = result_excel_col + 4
    confirmation_col = result_excel_col + 5
    ws.cell(row=1, column=result_excel_col, value=FORMAL_RESULT_HEADER)
    ws.cell(row=1, column=pending_col, value=PENDING_RESULT_HEADER)
    ws.cell(row=1, column=difference_col, value=CODE_DIFFERENCE_HEADER)
    ws.cell(row=1, column=comparison_col, value=OUTPUT_STATUS_HEADER)
    ws.cell(row=1, column=status_col, value=TRANSCODE_STATUS_HEADER)
    ws.cell(row=1, column=confirmation_col, value=CONFIRMATION_HEADER)
    ws.column_dimensions[get_column_letter(pending_col)].width = 30
    ws.column_dimensions[get_column_letter(difference_col)].width = 42
    system_analysis_col = _find_or_append_header_column(ws, SYSTEM_ANALYSIS_HEADER)
    ws.column_dimensions[get_column_letter(system_analysis_col)].width = 90
    product_name_col = _find_product_name_column(ws)
    analyses_by_row = {
        int(analysis["row"]): analysis
        for analysis in analyses
        if analysis.get("row") is not None
    }
    coverage_cache: dict[tuple[str, str], dict[str, Any]] = {}
    for i in range(1, len(df_req)):
        value = df_req.iloc[i, result_col]
        analysis = analyses_by_row.get(i + 1, {})
        product_name = ws.cell(row=i + 1, column=product_name_col).value if product_name_col else ""
        comparison = _export_result_comparison(
            analysis,
            product_name,
        )
        policy_difference_reason = _approved_policy_difference_reason(analysis, product_name)
        comparison_failed = comparison is False
        system_reason = _export_system_analysis_reason(
            analysis,
            product_name,
            comparison,
            agent_rules,
            agent_mapping_tables,
            coverage_cache,
        )
        reason_cell = ws.cell(row=i + 1, column=system_analysis_col, value=system_reason or None)
        if system_reason:
            reason_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(
            row=i + 1,
            column=comparison_col,
            value=comparison,
        )
        ws.cell(
            row=i + 1,
            column=status_col,
            value=_export_transcode_status(analysis, value),
        )
        if analysis.get("status") == "待确认":
            ws.cell(
                row=i + 1,
                column=confirmation_col,
                value=f"待确认：{analysis.get('reason', '')}",
            ).fill = red_fill
        formal_value = "" if pd.isna(value) or str(value).strip().lower() == "nan" else str(value).strip()
        pending_value = (
            str(analysis.get("candidate_code") or "").strip()
            if analysis.get("status") == "待确认"
            else ""
        )
        cell = ws.cell(row=i + 1, column=result_excel_col, value=formal_value or None)
        pending_cell = ws.cell(row=i + 1, column=pending_col, value=pending_value or None)
        if comparison_failed or policy_difference_reason or analysis.get("status") == "待确认":
            result_fill = red_fill
        elif formal_value.startswith("未识别"):
            result_fill = red_fill
        elif formal_value.startswith("跳过"):
            result_fill = yellow_fill
        else:
            result_fill = green_fill
        if formal_value:
            cell.fill = result_fill
        if pending_value:
            pending_cell.fill = red_fill
        difference_cell = ws.cell(
            row=i + 1,
            column=difference_col,
            value=_comparison_code_display(formal_value or pending_value, product_name, comparison),
        )
        difference_cell.fill = result_fill

    if not any(
        str(ws.cell(row=row, column=pending_col).value or "").strip()
        for row in range(2, ws.max_row + 1)
    ):
        ws.delete_cols(pending_col, amount=1)

    _format_agent_result_sheet(ws)

    for sheet_name in [
        "字段证据链",
        "待确认清单",
        "规则命中汇总",
        "问题分类明细",
        "问题聚合修复清单",
        "技术待支持清单",
        "模型语义影子证据",
        "模型实时语义标准化",
        "证据评分影子对比",
    ]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    _append_evidence_sheet(wb, analyses)
    _append_confirm_sheet(wb, analyses)
    _append_summary_sheet(wb, analyses, agent_rules, agent_mapping_tables, confirm_count)
    _append_issue_analysis_sheets(wb, ws, analyses, agent_rules, agent_mapping_tables)
    _append_technical_pending_sheet(wb, agent_mapping_tables)
    _append_semantic_shadow_sheet(wb, analyses)
    _append_order_semantic_model_sheet(wb, analyses)
    _append_evidence_score_shadow_sheet(wb, analyses)
    _remove_empty_sheets(wb, protected_sheet=ws.title)
    _atomic_save_workbook(wb, output_path)


def _atomic_save_workbook(workbook, output_path: str | Path) -> None:
    """Keep readers on a complete XLSX while an updated workbook is being saved."""
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{target.stem}.",
            suffix=".tmp.xlsx",
            dir=target.parent,
            delete=False,
        ) as temp_file:
            temp_path = Path(temp_file.name)
        workbook.save(temp_path)
        os.replace(temp_path, target)
        temp_path = None
        _cached_transcode_agent_trace_records.cache_clear()
    finally:
        if temp_path is not None:
            safe_unlink(temp_path)


def _export_transcode_status(
    analysis: dict[str, Any],
    value: Any,
) -> str:
    status = str(analysis.get("status") or "").strip()
    result_text = str(value or "").strip()
    if status == "成功":
        return "可直接采用"
    if status == "待确认":
        return "待人工确认"
    if status == "跳过" or result_text.startswith("跳过"):
        return "跳过"
    if status == "失败" or result_text.startswith("未识别"):
        return "未出码"
    return ""


def _find_product_name_column(ws) -> int | None:
    for cell in ws[1]:
        if _normalize_semantic_header(cell.value) in {"品名", "产品名称", "料品名称"}:
            return int(cell.column)
    return None


def _find_or_append_header_column(ws, header: str) -> int:
    for cell in ws[1]:
        if str(cell.value or "").strip() == header:
            return int(cell.column)
    last_header_col = max(
        (int(cell.column) for cell in ws[1] if str(cell.value or "").strip()),
        default=0,
    )
    column = last_header_col + 1
    ws.cell(row=1, column=column, value=header)
    return column


def _format_agent_result_sheet(ws) -> None:
    """Apply the result layout after optional columns have been removed."""
    headers = {
        str(cell.value or "").strip(): int(cell.column)
        for cell in ws[1]
        if str(cell.value or "").strip()
    }
    fixed_widths = {
        FORMAL_RESULT_HEADER: 34,
        PENDING_RESULT_HEADER: 34,
        CODE_DIFFERENCE_HEADER: 44,
        OUTPUT_STATUS_HEADER: 14,
        TRANSCODE_STATUS_HEADER: 16,
        CONFIRMATION_HEADER: 48,
    }
    for header, width in fixed_widths.items():
        column = headers.get(header)
        if column:
            ws.column_dimensions[get_column_letter(column)].width = width

    analysis_col = headers.get(SYSTEM_ANALYSIS_HEADER)
    if analysis_col:
        analysis_width = _content_based_column_width(
            (ws.cell(row=row, column=analysis_col).value for row in range(1, ws.max_row + 1)),
            minimum=36,
            maximum=64,
        )
        ws.column_dimensions[get_column_letter(analysis_col)].width = analysis_width

    wrapped_columns = [
        headers[header]
        for header in (
            FORMAL_RESULT_HEADER,
            PENDING_RESULT_HEADER,
            CODE_DIFFERENCE_HEADER,
            CONFIRMATION_HEADER,
            SYSTEM_ANALYSIS_HEADER,
        )
        if header in headers
    ]
    for row in range(2, ws.max_row + 1):
        required_lines = 1
        for column in wrapped_columns:
            cell = ws.cell(row=row, column=column)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            column_width = float(ws.column_dimensions[get_column_letter(column)].width or 13)
            required_lines = max(
                required_lines,
                _estimated_wrapped_lines(cell.value, column_width),
            )
        ws.row_dimensions[row].height = min(max(22, required_lines * 18 + 4), 108)


def _content_based_column_width(values, *, minimum: int, maximum: int) -> int:
    longest = max((_display_text_width(value) for value in values), default=0)
    return min(maximum, max(minimum, longest + 2))


def _estimated_wrapped_lines(value: Any, column_width: float) -> int:
    text = str(value or "")
    if not text:
        return 1
    usable_width = max(8, int(column_width) - 2)
    return sum(
        max(1, math.ceil(_display_text_width(line) / usable_width))
        for line in text.splitlines() or [""]
    )


def _display_text_width(value: Any) -> int:
    text = str(value or "")
    return max(
        (
            sum(2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1 for char in line)
            for line in text.splitlines() or [""]
        ),
        default=0,
    )


def _normalize_comparison_code(value: Any) -> str:
    text = re.sub(r"\s+", "", str(value or "").upper()).split("*", 1)[0]
    match = re.search(r"[A-Z0-9]{22,}", text)
    return match.group(0)[:22] if match else ""


COMPARISON_CODE_FIELDS = (
    ("胶系", 0, 2),
    ("基板厚度", 2, 7),
    ("铜箔规格", 7, 9),
    ("基板尺寸", 9, 17),
    ("胶水类别", 17, 18),
    ("铜箔类型", 18, 19),
    ("基板级别", 19, 21),
    ("总/芯厚", 21, 22),
)

COMPARISON_FIELD_OVERRIDE = {
    "胶系": "glue_code",
    "基板厚度": "thickness_code",
    "铜箔规格": "copper_code",
    "基板尺寸": "size_code",
    "胶水类别": "glue_category_code",
    "铜箔类型": "copper_type_code",
    "基板级别": "grade_code",
    "总/芯厚": "tc_code",
}

ERROR_FIELD_OVERRIDE = {
    "胶系": "glue_code",
    "厚度": "thickness_code",
    "铜箔": "copper_code",
    "尺寸": "size_code",
}


def _comparison_code_differences(actual_value: Any, expected_value: Any) -> list[tuple[str, str, str]]:
    actual_code = _normalize_comparison_code(actual_value)
    expected_code = _normalize_comparison_code(expected_value)
    if not actual_code or not expected_code:
        return []
    return [
        (field_name, actual_code[start:end], expected_code[start:end])
        for field_name, start, end in COMPARISON_CODE_FIELDS
        if actual_code[start:end] != expected_code[start:end]
    ]


def _comparison_code_display(
    actual_value: Any,
    expected_value: Any,
    comparison: bool | str,
) -> str | CellRichText:
    if comparison is not True and comparison is not False:
        return str(comparison)
    actual_code = _normalize_comparison_code(actual_value)
    if comparison is True:
        return actual_code
    expected_code = _normalize_comparison_code(expected_value)
    if not actual_code or not expected_code:
        return "无法对比"

    rich_text = CellRichText()
    error_font = InlineFont(color="FF9C0006")
    for _, start, end in COMPARISON_CODE_FIELDS:
        actual_part = actual_code[start:end]
        expected_part = expected_code[start:end]
        if actual_part == expected_part:
            rich_text.append(actual_part)
        else:
            rich_text.append(TextBlock(error_font, f"{actual_part}（{expected_part}）"))
    return rich_text


def _export_system_analysis_reason(
    analysis: dict[str, Any],
    product_name: Any,
    comparison: bool | str,
    agent_rules: list[dict],
    mapping_tables: dict[str, list[dict]],
    coverage_cache: dict[tuple[str, str], dict[str, Any]] | None = None,
) -> str:
    if str(analysis.get("status") or "").strip() == "待确认":
        reason = str(analysis.get("reason") or "").strip() or "存在未达到100分的关键字段"
        return f"待人工确认：{reason}。正式码保持为空；请在确认中心逐行核对后重新计算。"
    if comparison is not False and comparison != "未出码":
        return ""

    customer_code = str(analysis.get("customer_code") or "").strip()
    customer_name = str(analysis.get("customer") or "").strip()
    cache = coverage_cache if coverage_cache is not None else {}
    cache_key = (customer_code, customer_name)
    coverage = cache.get(cache_key)
    if coverage is None:
        coverage = _customer_rule_coverage(
            customer_code,
            customer_name,
            agent_rules,
            mapping_tables,
        )
        cache[cache_key] = coverage

    applied_by_field: dict[str, list[str]] = {}
    for rule in analysis.get("applied_rules", []) or []:
        field = str(rule.get("field") or "").strip()
        rule_id = str(rule.get("rule_id") or "").strip()
        if field:
            applied_by_field.setdefault(field, [])
            if rule_id and rule_id not in applied_by_field[field]:
                applied_by_field[field].append(rule_id)

    if comparison is False:
        differences = _comparison_code_differences(analysis.get("formal_code"), product_name)
        if not differences:
            return "码值不一致，但无法按22位字段定位；请检查Agent码和品名中的正确码是否完整。"
        diff_text = "；".join(
            f"{field_name} Agent={actual_code}、正确={expected_code}"
            for field_name, actual_code, expected_code in differences
        )
        override_fields = [COMPARISON_FIELD_OVERRIDE[field_name] for field_name, _, _ in differences]
        applied_ids = sorted({rule_id for field in override_fields for rule_id in applied_by_field.get(field, [])})
        if applied_ids:
            diagnosis = f"不一致字段已命中规则 {','.join(applied_ids)}，但规则输出与正确码冲突"
            action = "请核对正确码、规则条件和优先级，不应直接新增宽泛映射"
        elif any(field in coverage["fields"] for field in override_fields):
            diagnosis = "该客户已配置相关字段规则，但本行规格条件未命中，当前使用了基础结果"
            action = "请检查规则条件；如属新分支，需先补充客户特殊需求原表"
        elif coverage["has_customer_rules"]:
            diagnosis = "该客户存在其他规则，但未覆盖当前不一致字段，本行使用基础规则"
            action = "请先核对基础映射；如为客户特例，需补充客户特殊需求原表"
        else:
            diagnosis = "最终客户规则/辅助映射表未找到该客户的可执行规则，本行使用基础规则"
            action = "请先核对基础映射；若正确码依赖客户特例，需由业务补充客户特殊需求原表"
        return f"码值不一致：{diff_text}。定位：{diagnosis}。建议：{action}。"

    errors = list((analysis.get("engine_steps") or {}).get("errors") or [])
    reason = str(analysis.get("reason") or "").strip()
    if not errors and reason:
        errors = [item.strip() for item in re.split(r"[;；]", reason) if item.strip()]
    error_text = "；".join(errors) or "无法组成完整22位码"
    missing_fields = {
        override
        for keyword, override in ERROR_FIELD_OVERRIDE.items()
        if any(keyword in item for item in errors)
    }
    if missing_fields and any(field in coverage["fields"] for field in missing_fields):
        diagnosis = "已有该客户相关规则，但条件未命中或未能补齐基础字段"
    elif coverage["has_customer_rules"]:
        diagnosis = "该客户存在规则，但当前缺失字段没有可执行覆盖"
    else:
        diagnosis = "最终客户规则/辅助映射表未找到该客户的可执行规则"
    return (
        f"未出码：基础字段解析失败（{error_text}）。定位：{diagnosis}。"
        "建议：先补基础编码/解析规则；如为客户特殊写法，需先补充客户特殊需求原表。"
    )


def _customer_rule_coverage(
    customer_code: str,
    customer_name: str,
    agent_rules: list[dict],
    mapping_tables: dict[str, list[dict]],
) -> dict[str, Any]:
    fields: set[str] = set()
    candidates_by_field: dict[str, list[dict[str, str]]] = {}
    matched_count = 0
    for rule in agent_rules:
        if not _rule_executable(rule) or not _customer_matches(rule, customer_code, customer_name):
            continue
        matched_count += 1
        field = str(rule.get("覆盖字段") or "").strip()
        if field:
            fields.add(field)
            candidates_by_field.setdefault(field, []).append({
                "rule_id": str(rule.get("规则ID") or "").strip(),
                "condition": str(rule.get("条件") or rule.get("规则文本") or "").strip(),
            })

    sheet_fields = {
        "客户单边尺寸映射": {"size_code"},
        "客户尺寸映射": {"size_code"},
        "客户尺寸算法": {"size_code"},
        "客户厚度映射": {"thickness_code", "tc_code"},
        "客户物料编码口径": {"tc_code"},
        "外部尺寸表引用": {"size_code"},
    }
    for sheet_name, rows in mapping_tables.items():
        for row in rows:
            if not _mapping_row_enabled(row) or not _mapping_customer_matches(row, customer_code, customer_name):
                continue
            matched_count += 1
            field = str(row.get("覆盖字段") or "").strip()
            if field:
                fields.add(field)
                candidates_by_field.setdefault(field, []).append({
                    "rule_id": str(row.get("映射ID") or row.get("规则ID") or "").strip(),
                    "condition": str(row.get("条件") or row.get("规则文本") or sheet_name).strip(),
                })
            for mapped_field in sheet_fields.get(sheet_name, set()):
                fields.add(mapped_field)
                candidates_by_field.setdefault(mapped_field, []).append({
                    "rule_id": str(row.get("映射ID") or row.get("规则ID") or "").strip(),
                    "condition": str(row.get("条件") or row.get("规则文本") or sheet_name).strip(),
                })
    return {
        "has_customer_rules": matched_count > 0,
        "matched_count": matched_count,
        "fields": fields,
        "candidates_by_field": candidates_by_field,
    }


def _comparison_failure_prompt(analysis: dict[str, Any], product_name: Any) -> str:
    differences = _comparison_code_differences(analysis.get("formal_code"), product_name)
    if not differences:
        return "待确认：Agent转码结果与品名中的正确码不一致"
    details = "；".join(
        f"{field_name}不一致（Agent={actual_code}，正确码={expected_code}）"
        for field_name, actual_code, expected_code in differences
    )
    return f"待确认：{details}"


def _export_result_comparison(analysis: dict[str, Any], product_name: Any = "") -> bool | str:
    if str(analysis.get("status") or "").strip() == "跳过":
        return "跳过"
    if str(analysis.get("status") or "").strip() == "待确认" and analysis.get("candidate_code"):
        return "待人工确认"
    actual_code = _normalize_comparison_code(analysis.get("formal_code"))
    if not actual_code:
        return "未出码"
    expected_code = _normalize_comparison_code(product_name)
    if not expected_code:
        return "无法对比"
    return actual_code == expected_code or bool(_approved_policy_difference_reason(analysis, product_name))


def _approved_policy_difference_reason(analysis: dict[str, Any], product_name: Any) -> str:
    """Allow only field differences introduced by explicitly approved new policies."""
    actual_code = _normalize_comparison_code(analysis.get("formal_code"))
    expected_code = _normalize_comparison_code(product_name)
    if not actual_code or not expected_code or actual_code == expected_code:
        return ""

    differences = {
        field_name
        for field_name, start, end in COMPARISON_CODE_FIELDS
        if actual_code[start:end] != expected_code[start:end]
    }
    allowed_fields: set[str] = set()
    reasons: list[str] = []
    customer = str(analysis.get("customer") or "").strip()
    spec = f"{analysis.get('spec') or ''} {analysis.get('context_text') or ''}".upper()

    if re.search(r"(?<![A-Z0-9])NY3150HF(?![A-Z0-9])", spec):
        allowed_fields.update({"胶系", "胶水类别"})
        has_tft = bool(re.search(r"(?<![A-Z0-9])TFT(?![A-Z0-9])", spec))
        expected_grade = expected_code[19:21]
        actual_grade = actual_code[19:21]
        if has_tft or (expected_grade == "AT" and actual_grade == "A1"):
            allowed_fields.add("基板级别")
        reasons.append("NY3150HF按TFT最新口径")

    if customer == "珠海景旺" and re.search(r"(?<![A-Z0-9])NY2150(?![A-Z0-9])", spec):
        allowed_fields.update({"胶系", "胶水类别"})
        reasons.append("珠海景旺NY2150=2T/R最新口径")

    if customer == "江西景旺" and re.search(r"(?<![A-Z0-9])NY2150(?![A-Z0-9])", spec):
        allowed_fields.update({"胶系", "胶水类别"})
        reasons.append("江西景旺NY2150=2B/Y最新口径")

    if str((analysis.get("engine_steps") or {}).get("customer_rule_group") or "") == "深南集团":
        field_names = {override: name for name, override in COMPARISON_FIELD_OVERRIDE.items()}
        for applied in analysis.get("applied_rules") or []:
            field_name = field_names.get(str(applied.get("field") or ""))
            if field_name:
                allowed_fields.add(field_name)
        if allowed_fields:
            reasons.append("深南四客户共享特殊规则最新口径")

    if not differences or not differences.issubset(allowed_fields):
        return ""
    detail = "、".join(sorted(differences))
    return f"按已确认新业务口径判定TRUE，历史正确码待更新；口径：{'；'.join(reasons)}；差异字段：{detail}"


def _remove_empty_sheets(wb, *, protected_sheet: str) -> None:
    """Remove blank and header-only supporting sheets from the download."""
    for sheet_name in list(wb.sheetnames):
        if sheet_name == protected_sheet:
            continue
        ws = wb[sheet_name]
        populated_rows = [
            row
            for row in ws.iter_rows(values_only=True)
            if any(value is not None and str(value).strip() for value in row)
        ]
        if len(populated_rows) <= 1:
            del wb[sheet_name]


def _append_evidence_sheet(wb, analyses: list[dict]) -> None:
    ws = wb.create_sheet("字段证据链")
    headers = [
        "行号",
        "客户代码",
        "客户",
        "规格",
        "状态",
        "字段",
        "原始值",
        "编码片段",
        "置信度",
        "命中方式",
        "规则来源",
        "证据",
        "规则ID",
        "规则类型",
        "来源行号",
    ]
    ws.append(headers)
    for analysis in analyses:
        for item in analysis.get("field_evidence", []):
            ws.append([
                analysis.get("row"),
                analysis.get("customer_code", ""),
                analysis.get("customer", ""),
                analysis.get("spec", ""),
                analysis.get("status", ""),
                item.get("field", ""),
                item.get("value", ""),
                item.get("code", ""),
                item.get("score", ""),
                item.get("hit_type", ""),
                item.get("source", ""),
                item.get("evidence", ""),
                item.get("rule_id", ""),
                item.get("rule_type", ""),
                item.get("source_row", ""),
            ])
    _format_sheet(ws, [10, 14, 18, 46, 10, 12, 18, 12, 10, 18, 24, 54, 18, 18, 12])


def _append_confirm_sheet(wb, analyses: list[dict]) -> None:
    ws = wb.create_sheet("待确认清单")
    headers = ["行号", "客户代码", "客户", "规格", "状态", "正式码", "待人工确认码值", "最低分", "原因", "未达100分字段"]
    ws.append(headers)
    for analysis in analyses:
        if analysis.get("status") == "成功" or analysis.get("status") == "跳过":
            continue
        low_fields = [
            f"{item['field']}({item['score']})"
            for item in analysis.get("field_evidence", [])
            if item.get("gate") and int(item.get("score", 0)) < FIELD_GATE_THRESHOLD
        ]
        ws.append([
            analysis.get("row"),
            analysis.get("customer_code", ""),
            analysis.get("customer", ""),
            analysis.get("spec", ""),
            analysis.get("status", ""),
            analysis.get("formal_code", ""),
            analysis.get("candidate_code", ""),
            analysis.get("overall_score", ""),
            analysis.get("reason", ""),
            "；".join(low_fields),
        ])
    _format_sheet(ws, [10, 14, 18, 54, 10, 24, 24, 10, 44, 38])


def _append_summary_sheet(
    wb,
    analyses: list[dict],
    agent_rules: list[dict],
    agent_mapping_tables: dict[str, list[dict]],
    confirm_count: int,
) -> None:
    ws = wb.create_sheet("规则命中汇总")
    status_counter = Counter(analysis.get("status") for analysis in analyses)
    applied_counter = Counter()
    hit_type_counter = Counter()
    field_hit_counter = Counter()
    customer_agent_counter = Counter()
    rule_details: dict[str, dict] = {}
    for analysis in analyses:
        for item in analysis.get("field_evidence", []):
            hit_type = item.get("hit_type", "")
            field = item.get("field", "")
            if hit_type:
                hit_type_counter[hit_type] += 1
            if field and hit_type:
                field_hit_counter[(field, hit_type)] += 1
        for item in analysis.get("applied_rules", []):
            rule_id = item.get("rule_id", "")
            if not rule_id:
                continue
            hit_type = _applied_rule_hit_type(item)
            field_label = OVERRIDE_FIELD_LABELS.get(item.get("field", ""), item.get("field", ""))
            customer = analysis.get("customer", "") or analysis.get("customer_code", "")
            applied_counter[rule_id] += 1
            customer_agent_counter[(customer, hit_type)] += 1
            rule_details.setdefault(
                rule_id,
                {
                    "hit_type": hit_type,
                    "field": field_label,
                    "source_row": item.get("source_row", ""),
                    "rule_type": item.get("rule_type", ""),
                    "text": item.get("text", ""),
                },
            )
    ws.append(["指标", "数量/说明"])
    ws.append(["总行数", len(analyses)])
    ws.append(["高置信出码", status_counter.get("成功", 0)])
    ws.append(["待确认", confirm_count])
    ws.append(["未识别", status_counter.get("失败", 0)])
    ws.append(["跳过PP/RC/%", status_counter.get("跳过", 0)])
    ws.append(["当前Agent机器规则数", len(agent_rules)])
    ws.append(["当前Agent辅助映射总数", _mapping_total_count(agent_mapping_tables)])
    ws.append(["当前已接入辅助映射数", _mapping_enabled_count(agent_mapping_tables)])
    ws.append(["当前待确认/未接入技术项数", len(_technical_pending_rows(agent_mapping_tables))])
    semantic_evaluations = [
        item
        for analysis in analyses
        for item in analysis.get("semantic_shadow", [])
    ]
    semantic_counter = Counter(item.get("status") for item in semantic_evaluations)
    semantic_version = next(
        (analysis.get("semantic_rule_version") for analysis in analyses if analysis.get("semantic_rule_version")),
        "",
    )
    semantic_rule_count = max(
        [int(analysis.get("semantic_rule_count") or 0) for analysis in analyses] or [0]
    )
    semantic_load_error = next(
        (analysis.get("semantic_load_error") for analysis in analyses if analysis.get("semantic_load_error")),
        "",
    )
    ws.append(["模型语义影子版本", semantic_version or "未发布"])
    ws.append(["模型语义正式规则数", semantic_rule_count])
    ws.append(["模型语义影子评估数", len(semantic_evaluations)])
    ws.append(["模型语义影子命中", semantic_counter.get(SHADOW_STATUS_MATCHED, 0)])
    ws.append(["模型语义缺少输入", semantic_counter.get(SHADOW_STATUS_MISSING_INPUT, 0)])
    ws.append(["模型语义未命中", semantic_counter.get(SHADOW_STATUS_NOT_MATCHED, 0)])
    ws.append(["模型语义条件错误", semantic_counter.get(SHADOW_STATUS_ERROR, 0)])
    ws.append(["模型语义加载错误", semantic_load_error])
    ws.append(["模型语义运行时影响", "影子观察；不覆盖编码和评分"])
    order_model_records = [analysis.get("order_semantic_model") or {} for analysis in analyses]
    order_model_counter = Counter(item.get("status") for item in order_model_records)
    ws.append(["DeepSeek实时语义模式", next((item.get("mode") for item in order_model_records if item.get("mode")), "off")])
    ws.append(["DeepSeek实时语义模型", next((item.get("model") for item in order_model_records if item.get("model")), "未启用")])
    ws.append(["DeepSeek实时语义成功", order_model_counter.get("成功", 0)])
    ws.append(["DeepSeek实时语义失败", order_model_counter.get("失败", 0)])
    ws.append(["DeepSeek实时语义缓存命中", sum(1 for item in order_model_records if item.get("cached"))])
    ws.append(["DeepSeek实时语义限流跳过", order_model_counter.get("限流跳过", 0)])
    ws.append(["DeepSeek实时语义运行时影响", "影子观察；不覆盖制造码和正式评分"])
    score_shadows = [
        analysis.get("evidence_score_shadow") or {}
        for analysis in analyses
        if (analysis.get("evidence_score_shadow") or {}).get("field_reviews")
    ]
    score_decisions = Counter(item.get("shadow_decision") for item in score_shadows)
    ws.append(["证据影子评分行数", len(score_shadows)])
    ws.append(["证据影子评分通过", score_decisions.get("通过", 0)])
    ws.append(["证据影子评分需标注", score_decisions.get("需标注", 0)])
    ws.append(["证据影子模型调用", sum(int(item.get("model_call_count") or 0) for item in score_shadows)])
    ws.append(["证据影子模型成功", sum(1 for item in score_shadows if item.get("model_called") and not item.get("model_error"))])
    ws.append(["证据影子模型失败", sum(1 for item in score_shadows if item.get("model_error"))])
    evidence_gates = [analysis.get("evidence_gate") or {} for analysis in analyses if analysis.get("evidence_gate")]
    ws.append(["正式证据门禁模式", next((item.get("mode") for item in evidence_gates), "shadow")])
    ws.append(["正式证据门禁拦截", sum(1 for item in evidence_gates if item.get("blocked"))])
    ws.append(["证据影子运行时影响", "只对比；不覆盖当前确定性分和100分正式码门禁"])
    ws.append([])
    ws.append(["命中方式汇总", "命中次数"])
    for hit_type, count in hit_type_counter.most_common():
        ws.append([hit_type, count])
    ws.append([])
    ws.append(["字段命中汇总", "命中方式", "命中次数"])
    for (field, hit_type), count in field_hit_counter.most_common():
        ws.append([field, hit_type, count])
    ws.append([])
    ws.append(["客户Agent规则命中汇总", "命中方式", "命中次数"])
    for (customer, hit_type), count in customer_agent_counter.most_common():
        ws.append([customer, hit_type, count])
    ws.append([])
    ws.append(["规则ID", "命中次数", "命中方式", "字段", "规则类型", "来源行号", "规则文本"])
    for rule_id, count in applied_counter.most_common():
        detail = rule_details.get(rule_id, {})
        ws.append([
            rule_id,
            count,
            detail.get("hit_type", ""),
            detail.get("field", ""),
            detail.get("rule_type", ""),
            detail.get("source_row", ""),
            detail.get("text", ""),
        ])
    _format_sheet(ws, [26, 16, 18, 14, 18, 12, 60])


def _applied_rule_hit_type(item: dict) -> str:
    source = str(item.get("source", "") or "")
    if source.startswith("已批准模型语义映射"):
        return "已批准模型语义映射"
    if source.startswith("Agent尺寸映射"):
        return "Agent尺寸映射"
    if source.startswith("Agent厚度映射"):
        return "Agent厚度映射"
    if source.startswith("Agent总芯厚映射"):
        return "Agent总芯厚映射"
    if source.startswith("Agent物料编码口径"):
        return "Agent物料编码口径"
    return "Agent规则覆盖"


def _append_issue_analysis_sheets(
    wb,
    main_ws,
    analyses: list[dict],
    agent_rules: list[dict],
    mapping_tables: dict[str, list[dict]],
) -> None:
    detail_ws = wb.create_sheet("问题分类明细")
    detail_headers = [
        "行号", "客户代码", "客户", "规格", "结果对比", "问题分类", "错误字段",
        "Agent值", "正确值", "已命中规则ID", "候选未命中规则ID", "候选规则条件",
        "系统分析原因", "修复建议", "问题指纹",
    ]
    detail_ws.append(detail_headers)
    product_name_col = _find_product_name_column(main_ws)
    analyses_by_row = {int(item.get("row")): item for item in analyses if item.get("row") is not None}
    coverage_cache: dict[tuple[str, str], dict[str, Any]] = {}
    issue_rows: list[dict[str, Any]] = []
    for row_number, analysis in sorted(analyses_by_row.items()):
        product_name = main_ws.cell(row=row_number, column=product_name_col).value if product_name_col else ""
        comparison = _export_result_comparison(analysis, product_name)
        if comparison is not False and comparison != "未出码":
            continue
        issue_rows.extend(_build_issue_rows(
            analysis,
            product_name,
            comparison,
            agent_rules,
            mapping_tables,
            coverage_cache,
        ))
    for item in issue_rows:
        detail_ws.append([item.get(header, "") for header in detail_headers])
    _format_sheet(detail_ws, [10, 14, 18, 54, 12, 28, 14, 12, 12, 24, 28, 54, 72, 54, 72])

    aggregate_ws = wb.create_sheet("问题聚合修复清单")
    aggregate_headers = [
        "排名", "问题指纹", "聚合问题分类", "影响行数", "涉及客户数", "涉及客户",
        "错误字段", "Agent值→正确值", "已命中规则ID", "候选规则ID",
        "示例行号", "示例规格", "修复建议",
    ]
    aggregate_ws.append(aggregate_headers)
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in issue_rows:
        grouped.setdefault(str(item.get("问题指纹") or ""), []).append(item)
    ordered_groups = sorted(grouped.items(), key=lambda pair: (-len(pair[1]), pair[0]))
    for rank, (fingerprint, rows) in enumerate(ordered_groups, start=1):
        first = rows[0]
        customers = sorted({str(item.get("客户") or item.get("客户代码") or "").strip() for item in rows if str(item.get("客户") or item.get("客户代码") or "").strip()})
        category = str(first.get("问题分类") or "")
        if category == "基础规则映射/标准解析待核对" and len(customers) > 1:
            category = "疑似基础规则/解析共性问题"
        values = sorted({f"{item.get('Agent值', '')}→{item.get('正确值', '')}" for item in rows})
        hit_ids = sorted({rule_id for item in rows for rule_id in str(item.get("已命中规则ID") or "").split(",") if rule_id})
        candidate_ids = sorted({rule_id for item in rows for rule_id in str(item.get("候选未命中规则ID") or "").split(",") if rule_id})
        aggregate_ws.append([
            rank,
            fingerprint,
            category,
            len(rows),
            len(customers),
            "、".join(customers[:30]) + (f"等{len(customers)}个客户" if len(customers) > 30 else ""),
            first.get("错误字段", ""),
            "；".join(values[:20]),
            ",".join(hit_ids),
            ",".join(candidate_ids),
            "、".join(str(item.get("行号") or "") for item in rows[:10]),
            first.get("规格", ""),
            first.get("修复建议", ""),
        ])
    _format_sheet(aggregate_ws, [8, 72, 30, 12, 14, 54, 14, 38, 28, 28, 28, 60, 54])


def _build_issue_rows(
    analysis: dict[str, Any],
    product_name: Any,
    comparison: bool | str,
    agent_rules: list[dict],
    mapping_tables: dict[str, list[dict]],
    coverage_cache: dict[tuple[str, str], dict[str, Any]],
) -> list[dict[str, Any]]:
    customer_code = str(analysis.get("customer_code") or "").strip()
    customer_name = str(analysis.get("customer") or "").strip()
    cache_key = (customer_code, customer_name)
    coverage = coverage_cache.get(cache_key)
    if coverage is None:
        coverage = _customer_rule_coverage(customer_code, customer_name, agent_rules, mapping_tables)
        coverage_cache[cache_key] = coverage
    applied_by_field: dict[str, list[str]] = {}
    for rule in analysis.get("applied_rules", []) or []:
        field = str(rule.get("field") or "").strip()
        rule_id = str(rule.get("rule_id") or "").strip()
        if field and rule_id:
            applied_by_field.setdefault(field, [])
            if rule_id not in applied_by_field[field]:
                applied_by_field[field].append(rule_id)
    system_reason = _export_system_analysis_reason(
        analysis, product_name, comparison, agent_rules, mapping_tables, coverage_cache
    )
    common = {
        "行号": analysis.get("row", ""),
        "客户代码": customer_code,
        "客户": customer_name,
        "规格": analysis.get("spec", ""),
        "结果对比": comparison,
        "系统分析原因": system_reason,
    }
    rows: list[dict[str, Any]] = []
    if comparison is False:
        differences = _comparison_code_differences(analysis.get("formal_code"), product_name)
        if not differences:
            differences = [("无法定位", "", "")]
        for field_label, actual_value, expected_value in differences:
            override_field = COMPARISON_FIELD_OVERRIDE.get(field_label, "")
            applied_ids = sorted(applied_by_field.get(override_field, []))
            candidates = list((coverage.get("candidates_by_field") or {}).get(override_field, []))
            candidate_ids = sorted({str(item.get("rule_id") or "") for item in candidates if item.get("rule_id") and str(item.get("rule_id")) not in applied_ids})
            candidate_conditions = sorted({str(item.get("condition") or "") for item in candidates if item.get("condition")})
            if applied_ids:
                category = "已命中客户规则但输出错误"
                advice = "核对已命中规则的覆盖值、条件边界和优先级，防止宽泛规则覆盖更具体规则"
            elif candidate_ids:
                category = "已有客户规则未命中"
                advice = "检查候选规则的条件、边界、输入字段和客户匹配；如属新分支，先更新客户特殊需求原表"
            elif coverage.get("has_customer_rules"):
                category = "客户特殊规则缺失待核对"
                advice = "现有客户规则未覆盖该字段；先核对基础规则，需客户特例时由业务补充原表"
            else:
                category = "基础规则映射/标准解析待核对"
                advice = "先对照字段证据链核对原始值是否解析正确；解析正确则修基础映射，解析错误则修通用解析器"
            fingerprint_customer = customer_name or customer_code if category in {"已命中客户规则但输出错误", "已有客户规则未命中", "客户特殊规则缺失待核对"} else ""
            fingerprint = "|".join(filter(None, [category, field_label, f"{actual_value}→{expected_value}", ",".join(applied_ids or candidate_ids), fingerprint_customer]))
            rows.append({
                **common,
                "问题分类": category,
                "错误字段": field_label,
                "Agent值": actual_value,
                "正确值": expected_value,
                "已命中规则ID": ",".join(applied_ids),
                "候选未命中规则ID": ",".join(candidate_ids),
                "候选规则条件": "；".join(candidate_conditions),
                "修复建议": advice,
                "问题指纹": fingerprint,
            })
        return rows

    errors = list((analysis.get("engine_steps") or {}).get("errors") or [])
    reason = str(analysis.get("reason") or "").strip()
    if not errors and reason:
        errors = [part.strip() for part in re.split(r"[;；]", reason) if part.strip()]
    missing_fields = sorted({override for keyword, override in ERROR_FIELD_OVERRIDE.items() if any(keyword in item for item in errors)})
    candidates = [item for field in missing_fields for item in (coverage.get("candidates_by_field") or {}).get(field, [])]
    candidate_ids = sorted({str(item.get("rule_id") or "") for item in candidates if item.get("rule_id")})
    candidate_conditions = sorted({str(item.get("condition") or "") for item in candidates if item.get("condition")})
    if candidate_ids:
        category = "已有客户规则未命中"
        advice = "检查候选规则的输入条件和解析前置字段，确认是否应用规则补齐"
    elif coverage.get("has_customer_rules"):
        category = "客户特殊规则缺失待核对"
        advice = "先补通用基础解析；如为客户特殊写法，由业务补充客户特殊需求原表"
    else:
        category = "标准规格解析失败"
        advice = "核对缺失字段的单位、别名和复合格式，优先补充通用解析规则"
    field_labels = [OVERRIDE_FIELD_LABELS.get(field, field) for field in missing_fields] or ["未知基础字段"]
    fingerprint_customer = customer_name or customer_code if category != "标准规格解析失败" else ""
    fingerprint = "|".join(filter(None, [category, ",".join(field_labels), ",".join(candidate_ids), fingerprint_customer]))
    rows.append({
        **common,
        "问题分类": category,
        "错误字段": "、".join(field_labels),
        "Agent值": "未出码",
        "正确值": "",
        "已命中规则ID": "",
        "候选未命中规则ID": ",".join(candidate_ids),
        "候选规则条件": "；".join(candidate_conditions),
        "修复建议": advice,
        "问题指纹": fingerprint,
    })
    return rows


def _append_technical_pending_sheet(wb, agent_mapping_tables: dict[str, list[dict]]) -> None:
    ws = wb.create_sheet("技术待支持清单")
    headers = [
        "处理状态",
        "技术类型",
        "客户代码",
        "客户",
        "来源行号",
        "规则ID",
        "规则文本",
        "未执行原因",
        "建议确认/处理",
    ]
    ws.append(headers)
    rows = _technical_pending_rows(agent_mapping_tables)
    for row in rows:
        ws.append([
            row.get("处理状态", ""),
            row.get("技术类型", ""),
            row.get("客户代码", ""),
            row.get("客户简称", ""),
            row.get("来源行号", ""),
            row.get("规则ID", ""),
            row.get("规则文本", ""),
            row.get("未执行原因", ""),
            row.get("建议确认/处理", ""),
        ])
    _format_sheet(ws, [14, 18, 14, 18, 12, 18, 56, 40, 46])


def _append_semantic_shadow_sheet(wb, analyses: list[dict]) -> None:
    ws = wb.create_sheet("模型语义影子证据")
    headers = [
        "行号",
        "客户代码",
        "客户",
        "规格",
        "运行模式",
        "语义规则版本",
        "规则ID",
        "来源候选ID",
        "业务字段",
        "目标字段",
        "标准语义值",
        "原文目标值",
        "影子状态",
        "缺少输入字段",
        "条件校验JSON",
        "观察输入JSON",
        "规则原文",
        "业务证据",
        "模型",
        "说明",
    ]
    ws.append(headers)
    load_error_written = False
    for analysis in analyses:
        mode = analysis.get("semantic_shadow_mode", "")
        version = analysis.get("semantic_rule_version", "")
        load_error = str(analysis.get("semantic_load_error") or "")
        if load_error and not load_error_written:
            ws.append([
                "",
                "",
                "",
                "",
                mode,
                version,
                "",
                "",
                "",
                "",
                "",
                "",
                SHADOW_STATUS_ERROR,
                "",
                "[]",
                "{}",
                "",
                "",
                "",
                f"语义规则加载失败，不影响正式转码：{load_error}",
            ])
            load_error_written = True
        for item in analysis.get("semantic_shadow", []):
            ws.append([
                item.get("row", analysis.get("row")),
                item.get("customer_code", analysis.get("customer_code", "")),
                item.get("customer", analysis.get("customer", "")),
                item.get("spec", analysis.get("spec", "")),
                mode,
                version,
                item.get("rule_id", ""),
                item.get("source_candidate_id", ""),
                item.get("business_field", ""),
                "；".join(str(value) for value in item.get("target_fields", []) if value),
                "；".join(str(value) for value in item.get("normalized_values", []) if value),
                "；".join(str(value) for value in item.get("stated_target_values", []) if value),
                item.get("status", ""),
                "；".join(str(value) for value in item.get("missing_fields", []) if value),
                format_condition_results(item.get("condition_results", [])),
                format_observed_inputs(item.get("observed_inputs", {})),
                item.get("source_text", ""),
                "；".join(str(value) for value in item.get("evidence_texts", []) if value),
                item.get("model", ""),
                item.get("note", ""),
            ])
    _format_sheet(
        ws,
        [10, 14, 18, 42, 12, 30, 24, 24, 14, 20, 24, 24, 12, 22, 56, 56, 64, 52, 20, 42],
    )


def _append_order_semantic_model_sheet(wb, analyses: list[dict]) -> None:
    ws = wb.create_sheet("模型实时语义标准化")
    headers = [
        "行号",
        "客户代码",
        "客户",
        "规格",
        "运行模式",
        "调用状态",
        "模型",
        "缓存命中",
        "相关正式语义规则ID",
        "原始输入JSON",
        "模型置信度",
        "标准语义项JSON",
        "歧义JSON",
        "缺少输入JSON",
        "错误/说明",
        "运行时影响",
    ]
    ws.append(headers)
    for analysis in analyses:
        record = analysis.get("order_semantic_model") or {}
        if record.get("status") in {"未调用", "跳过", ""}:
            continue
        result = record.get("result") or {}
        matched_rule_ids = list(record.get("matched_rule_ids") or [])
        notes = list(record.get("notes") or [])
        ws.append([
            analysis.get("row"),
            analysis.get("customer_code", ""),
            analysis.get("customer", ""),
            analysis.get("spec", ""),
            record.get("mode", ""),
            record.get("status", ""),
            record.get("model", ""),
            "是" if record.get("cached") else "否",
            "；".join(matched_rule_ids or record.get("rule_ids") or []),
            json.dumps(record.get("source_fields") or {}, ensure_ascii=False, separators=(",", ":")),
            result.get("model_confidence", ""),
            json.dumps(result.get("semantic_items") or [], ensure_ascii=False, separators=(",", ":")),
            json.dumps(result.get("ambiguities") or [], ensure_ascii=False, separators=(",", ":")),
            json.dumps(result.get("missing_inputs") or [], ensure_ascii=False, separators=(",", ":")),
            record.get("error", "") or record.get("reason", "") or "；".join(notes),
            (
                "模型仅标准化；命中已批准规则后可覆盖允许字段并重新评分"
                if matched_rule_ids
                else "未绑定已批准规则，不影响制造码"
            ),
        ])
    _format_sheet(ws, [10, 14, 18, 42, 12, 12, 20, 10, 32, 60, 12, 72, 52, 52, 52, 28])


def _append_evidence_score_shadow_sheet(wb, analyses: list[dict]) -> None:
    ws = wb.create_sheet("证据评分影子对比")
    headers = [
        "行号",
        "客户代码",
        "客户",
        "规格",
        "当前状态",
        "正式结果",
        "候选码",
        "当前总分",
        "影子总分",
        "总分差",
        "影子决策",
        "字段",
        "候选原始值",
        "候选编码片段",
        "当前字段分",
        "影子字段分",
        "字段分差",
        "证据结论",
        "命中方式",
        "证据来源",
        "证据文本",
        "评分理由",
        "现有规则ID",
        "语义规则ID",
        "语义证据",
        "原始输入JSON",
        "模型调用",
        "程序结论",
        "程序影子分",
        "模型结论",
        "模型证据来源",
        "模型证据文本",
        "模型理由",
        "模型置信度",
        "模型结论采纳",
        "模型错误",
        "运行时影响",
        "正式证据门禁模式",
        "程序证据分",
        "正式有效分",
        "正式门禁拦截",
        "正式门禁原因",
    ]
    ws.append(headers)
    for analysis in analyses:
        shadow = analysis.get("evidence_score_shadow") or {}
        gate = analysis.get("evidence_gate") or {}
        source_fields_json = format_source_fields(shadow.get("source_fields") or {})
        for review in shadow.get("field_reviews") or []:
            ws.append([
                analysis.get("row"),
                analysis.get("customer_code", ""),
                analysis.get("customer", ""),
                analysis.get("spec", ""),
                analysis.get("status", ""),
                analysis.get("formal_code", ""),
                analysis.get("candidate_code", ""),
                shadow.get("current_score", ""),
                shadow.get("shadow_score", ""),
                shadow.get("score_delta", ""),
                shadow.get("shadow_decision", ""),
                review.get("field", ""),
                review.get("candidate_value", ""),
                review.get("candidate_code", ""),
                review.get("current_score", ""),
                review.get("shadow_score", ""),
                review.get("score_delta", ""),
                review.get("verdict", ""),
                review.get("hit_type", ""),
                review.get("source_field", ""),
                review.get("evidence_text", ""),
                review.get("reason", ""),
                review.get("rule_id", ""),
                "；".join(review.get("semantic_rule_ids") or []),
                "；".join(review.get("semantic_evidence") or []),
                source_fields_json,
                "否" if not review.get("model_called") else "是",
                review.get("program_verdict", ""),
                review.get("program_shadow_score", ""),
                review.get("model_verdict", ""),
                review.get("model_source_field", ""),
                review.get("model_evidence_text", ""),
                review.get("model_reason", ""),
                review.get("model_confidence", ""),
                "是" if review.get("model_accepted") else ("否" if review.get("model_called") else ""),
                shadow.get("model_error", ""),
                shadow.get("runtime_effect", ""),
                gate.get("mode", ""),
                gate.get("program_evidence_score", ""),
                gate.get("effective_score", ""),
                "是" if gate.get("blocked") else "否",
                "；".join(gate.get("blockers") or []) or ("证据有效分低于90" if gate.get("blocked") else ""),
            ])
    _format_sheet(
        ws,
        [10, 14, 18, 42, 12, 24, 24, 12, 12, 10, 12, 12, 18, 18, 12, 12, 10, 18, 18, 24, 44, 52, 20, 34, 52, 64, 12, 18, 14, 18, 24, 44, 52, 14, 14, 52, 48, 18, 14, 14, 14, 42],
    )


def _technical_pending_rows(agent_mapping_tables: dict[str, list[dict]]) -> list[dict]:
    pending: list[dict] = []
    for row in agent_mapping_tables.get("客户物料编码口径", []):
        if _mapping_row_enabled(row):
            continue
        pending.append(
            {
                "处理状态": "待确认后接入",
                "技术类型": "客户物料编码口径",
                "客户代码": row.get("客户代码", ""),
                "客户简称": row.get("客户简称", ""),
                "来源行号": row.get("来源行号", ""),
                "规则ID": row.get("映射ID", ""),
                "规则文本": row.get("规则文本", ""),
                "未执行原因": "物料编码来源字段未确认，避免把规格文本片段误判为订单物料编码",
                "建议确认/处理": "确认字段来源：品号、客户产品编号或其他订单字段；确认命中值是完整编码还是片段",
            }
        )
    for row in agent_mapping_tables.get("外部尺寸表引用", []):
        if _mapping_row_enabled(row):
            continue
        pending.append(
            {
                "处理状态": "待确认后接入",
                "技术类型": "外部尺寸表引用",
                "客户代码": row.get("客户代码", ""),
                "客户简称": row.get("客户简称", ""),
                "来源行号": row.get("来源行号", ""),
                "规则ID": row.get("映射ID", ""),
                "规则文本": row.get("规则文本", ""),
                "未执行原因": "外部尺寸表读取方式和命中字段未确认",
                "建议确认/处理": "确认是否允许运行时读取引用文件，以及命中后是否统一用基础 size_to_code 生成 size_code",
            }
        )
    for row in agent_mapping_tables.get("待接入规则", []):
        technical_type = str(row.get("技术类型", "") or "")
        if "物料" in technical_type:
            continue
        if "物料" in technical_type or "外部" in technical_type or "订单" in technical_type:
            pending.append(
                {
                    "处理状态": "待确认后接入",
                    "技术类型": technical_type,
                    "客户代码": row.get("客户代码", ""),
                    "客户简称": row.get("客户简称", ""),
                    "来源行号": row.get("来源行号", ""),
                    "规则ID": row.get("映射ID", ""),
                    "规则文本": row.get("原始规则", ""),
                    "未执行原因": "待接入规则总览项，本阶段不执行不确定语义",
                    "建议确认/处理": row.get("建议处理", ""),
                }
            )
    return pending


def _mapping_total_count(agent_mapping_tables: dict[str, list[dict]]) -> int:
    return sum(len(rows) for rows in agent_mapping_tables.values())


def _mapping_enabled_count(agent_mapping_tables: dict[str, list[dict]]) -> int:
    executable_sheets = {"客户尺寸映射", "客户单边尺寸映射", "客户尺寸算法", "客户厚度映射", "客户物料编码口径", "外部尺寸表引用"}
    return sum(
        1
        for sheet_name, rows in agent_mapping_tables.items()
        if sheet_name in executable_sheets
        for row in rows
        if _mapping_row_enabled(row)
    )


def _format_sheet(ws, widths: list[int]) -> None:
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width


def _skip_analysis(row: int, customer_code: str, customer: str, spec: str, reason: str) -> dict:
    return {
        "row": row,
        "customer_code": customer_code,
        "customer": customer,
        "spec": spec,
        "status": "跳过",
        "formal_code": "",
        "candidate_code": "",
        "overall_score": 0,
        "reason": reason,
        "summary": reason,
        "field_evidence": [],
        "applied_rules": [],
        "conflicts": [],
    }


def _ensure_agent_result_column(df_req) -> int:
    if len(df_req) > 0:
        header_matches = [
            index
            for index in range(len(df_req.columns))
            if str(df_req.iloc[0, index]).strip() == FORMAL_RESULT_HEADER
        ]
        if header_matches:
            index = header_matches[-1]
            df_req.iloc[:, index] = ""
            return index
    for index, col in enumerate(df_req.columns):
        if str(col).strip() == FORMAL_RESULT_HEADER:
            df_req.iloc[:, index] = ""
            return index
    result_col = len(df_req.columns)
    df_req[FORMAL_RESULT_HEADER] = ""
    return result_col


def _is_effective_spec(value: Any, engine) -> bool:
    text = engine._clean_cell(value)
    if not text:
        return False
    lower = text.lower()
    return lower not in ("nan", "客户规格", "规格", "品名")


def _prefer_normalized_spec_for_legacy_format(spec: str) -> bool:
    """Prefer an existing normalized-spec column for ambiguous legacy layouts."""
    text = str(spec or "").upper()
    patterns = (
        r'FR4\.[01]\s+\d+(?:\.\d+)?\s+[A-Z0-9.]+\s*/\s*[A-Z0-9.]+\s+\d+(?:\.\d+)?',
        r'\d+\.\d+\s*[*×X]\s*\d{3,4}(?:\.\d+)?\s*[*×X]\s*\d{3,4}(?:\.\d+)?\s*MM',
        r'\d+(?:\.\d+)?\s*IN(?:CH)?\s*\d+(?:\.\d+)?\s*IN(?:CH)?',
        r'[经纬]\s*\d+(?:\.\d+)?\s*[*×X]\s*[经纬]\s*\d+(?:\.\d+)?',
        r"\d'{2}\s*[*×X]\s*\d+'{2}",
        r'(?<![\d.])0{2,3}\d{1,2}\s*(?:"|″|”)',
        r'(?:FR|RF)\s*-?\s*4\s*-{2,}\s*\d+(?:\.\d+)?',
    )
    return any(re.search(pattern, text, re.IGNORECASE) for pattern in patterns)


def _norm_match(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").upper())


def _keyword_condition_matches(keyword: str, combined: str) -> bool:
    norm = _norm_match(keyword)
    if not norm:
        return False
    if norm == "__NO_TG__":
        return not re.search(r"TG\d+", combined)
    if "客户规格没有Q" in norm or "规格没有Q" in norm:
        return not re.search(r"(?<![A-Z0-9])Q(?![A-Z0-9])", combined)
    if "AT板" in norm and "字样" in norm:
        return "AT板" in combined
    parts = [part for part in re.split(r"&|\+|和|且", norm) if part]
    if len(parts) > 1:
        return all(_keyword_condition_matches(part, combined) for part in parts)
    if re.fullmatch(r"\d+(?:\.\d+)?", norm):
        return bool(re.search(rf"(?<![A-Z0-9.]){re.escape(norm)}(?![A-Z0-9.])", combined))
    if _is_model_like_token(norm):
        return _model_token_matches(norm, combined)
    return norm in combined


def _glue_condition_matches(token: str, glue_norm: str, spec_norm: str) -> bool:
    token_norm = _norm_match(token)
    if not token_norm:
        return False
    if token_norm == glue_norm:
        return True
    if _is_model_like_token(token_norm):
        return _model_token_matches(token_norm, spec_norm)
    return _token_with_alnum_boundary_matches(token_norm, spec_norm)


def _copper_condition_matches(token: str, copper_norm: str, spec_norm: str) -> bool:
    token_norm = _norm_match(token)
    if not token_norm:
        return False
    comparison = re.fullmatch(r"(>=|<=|>|<|≥|≤)(\d+(?:\.\d+)?)OZ", token_norm)
    if comparison:
        actual = _max_copper_oz_from_norm(copper_norm)
        if actual is None:
            return False
        operator, expected_text = comparison.groups()
        expected = float(expected_text)
        if operator in (">=", "≥"):
            return actual >= expected
        if operator in ("<=", "≤"):
            return actual <= expected
        if operator == ">":
            return actual > expected
        return actual < expected
    if token_norm == copper_norm:
        return True
    return _token_with_alnum_boundary_matches(token_norm, spec_norm)


_COPPER_OZ_BY_CODE = {
    "0": 0.0,
    "J": 15 / 35,
    "H": 0.5,
    "K": 28 / 35,
    "1": 1.0,
    "F": 1.5,
    "R": 1.5,
    "2": 2.0,
}


def _max_copper_oz_from_norm(copper_norm: str) -> float | None:
    values = [
        _COPPER_OZ_BY_CODE[part]
        for part in str(copper_norm or "").upper().split("/")
        if part in _COPPER_OZ_BY_CODE
    ]
    return max(values) if values else None


def _is_model_like_token(token: str) -> bool:
    return bool(re.fullmatch(r"(?:NY|DS|TG|FR)[A-Z0-9().-]*\d[A-Z0-9().-]*", token))


def _model_token_matches(token: str, text: str) -> bool:
    if not token or not text:
        return False
    return bool(re.search(rf"(?<![A-Z0-9]){re.escape(token)}(?![A-Z-])", text))


def _token_with_alnum_boundary_matches(token: str, text: str) -> bool:
    if not token or not text:
        return False
    return bool(re.search(rf"(?<![A-Z0-9]){re.escape(token)}(?![A-Z0-9])", text))


def _flexible_model_token_matches(token: str, text: str) -> bool:
    """Match the same glue model across harmless spacing and separator variants."""
    normalized = _norm_match(token)
    if not normalized or not text:
        return False
    characters = [re.escape(char) for char in normalized]
    pattern = r"[\s_\-]*".join(characters)
    return bool(re.search(rf"(?<![A-Z0-9]){pattern}(?![A-Z0-9])", str(text), re.I))


def _is_legacy_copper_ff_keyword_condition(rule: dict, keyword_condition: str) -> bool:
    keywords = {_norm_match(item) for item in re.split(r"[/,，、;；]+", keyword_condition) if item.strip()}
    return (
        str(rule.get("覆盖字段", "") or "") == "copper_code"
        and str(rule.get("覆盖值", "") or "").strip().upper() == "FF"
        and keywords == {"5"}
    )


def _legacy_copper_ff_keyword_matches(combined: str, steps: dict) -> bool:
    source = f"{combined} {_norm_match(steps.get('copper_spec_raw', ''))}"
    return any(token in source for token in ("1.5/1.5", "R/R", "F/F"))


def _copper_condition_tokens(value: str) -> list[str]:
    source = _norm_match(value)
    comparisons = re.findall(r"(?:>=|<=|>|<|≥|≤)\d+(?:\.\d+)?OZ", source)
    if comparisons:
        return comparisons
    known_tokens = ["1.5/1.5", "0.5/0.5", "H/H", "R/R", "F/F", "J/J", "K/K", "W/W", "1/1", "2/2"]
    tokens = [token for token in known_tokens if token in source]
    if tokens:
        return tokens
    return [item for item in re.split(r"[,，;；]+", source) if item]


def _thickness_condition_matches(token: str, steps: dict, spec_norm: str, thickness_norm: str) -> bool:
    raw = str(token or "").strip()
    if not raw:
        return False
    norm = _norm_match(raw)
    if "MIL" in norm:
        return norm in spec_norm or norm in thickness_norm
    match = re.fullmatch(r"(>=|<=|>|<|≥|≤)?(\d+(?:\.\d+)?)", norm)
    if not match:
        return norm in spec_norm or norm in thickness_norm
    operator, number_text = match.groups()
    try:
        expected = float(number_text)
    except ValueError:
        return False
    actual = steps.get("thickness_mm")
    try:
        actual_value = float(actual)
    except (TypeError, ValueError):
        return number_text in thickness_norm
    if operator in (">=", "≥"):
        return actual_value >= expected
    if operator in ("<=", "≤"):
        return actual_value <= expected
    if operator == ">":
        return actual_value > expected
    if operator == "<":
        return actual_value < expected
    return abs(actual_value - expected) < 0.0001


def _norm_customer_code(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def _customer_code_tokens(value: Any) -> list[str]:
    return re.findall(r"\d+", str(value or ""))


def _validated_customer_code_tokens(value: Any) -> list[str]:
    raw = str(value or "").strip()
    if not raw or not re.fullmatch(r"\d{5,8}(?:\s*[,，、/;；]\s*\d{5,8})*", raw):
        return []
    return _customer_code_tokens(raw)
