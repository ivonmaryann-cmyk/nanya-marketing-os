from __future__ import annotations

"""Customer archive and customer-level automation configuration.

This module deliberately keeps customer identification deterministic.  A mail is
only assigned to a customer when an enabled identity has a single active owner;
otherwise it remains unmatched for business confirmation.
"""

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .database import automation_cursor as db_cursor


CUSTOMER_FIELDS = (
    "customer_code", "customer_short_name", "quick_code", "customer_name", "group_name", "customer_type",
    "sales_name", "service_name", "internal_clerk_name", "internal_clerk_employee_id", "technical_support_name",
    "insurer_days", "credit_amount", "payment_terms", "grace_days", "invoice_address", "delivery_address",
    "contact_name", "contact_phone", "settlement_day", "transit_days", "first_trade_at", "last_order_at",
    "last_delivery_at", "last_payment_at", "status", "note",
)

CUSTOMER_FIELD_LABELS = {
    "customer_code": "客户编号", "customer_short_name": "客户简称", "quick_code": "快捷码", "customer_name": "客户全称",
    "group_name": "所属集团", "customer_type": "客户类型", "sales_name": "业务员", "service_name": "客服人员",
    "internal_clerk_name": "内勤人员", "internal_clerk_employee_id": "内勤工号", "technical_support_name": "技服人员",
    "insurer_days": "中信保天数", "credit_amount": "额度", "payment_terms": "账期", "grace_days": "宽限天数",
    "invoice_address": "发票地址", "delivery_address": "送货地址", "contact_name": "联系人", "contact_phone": "联系电话",
    "settlement_day": "月结日", "transit_days": "运输天数", "first_trade_at": "初次交易", "last_order_at": "最近接单",
    "last_delivery_at": "最近出货", "last_payment_at": "最近收款", "status": "状态", "note": "备注",
}

CONTACT_TYPE_LABELS = {
    "sender_email": "发件人邮箱", "alias_email": "别名邮箱", "sender_domain": "发件邮箱域名",
}
ROUTING_SCOPE_LABELS = {
    "subject": "邮件主题", "body": "邮件正文", "attachment_name": "附件名称", "attachment_content": "附件内容",
}
EXTRACTION_SOURCE_KIND_LABELS = {
    # 客户级字段映射当前仅对已解析出的附件订单表格生效。不要在页面上
    # 暴露尚未执行的正文或附件名称映射，避免业务维护了无效配置。
    "attachment_table": "附件订单表格字段",
}
EXTRACTION_TARGET_LABELS = {
    "customer_product_code": "客户产品编号", "customer_spec": "客户规格", "quantity": "数量", "delivery_date": "出货日期",
    "price_before_tax": "税前单价", "unit_price": "单价", "customer_order_seq": "客户订单序号",
    "customer_order_number": "客户订单号", "remark": "备注",
}
TRANSFORM_LABELS = {
    "direct": "直接带入", "concat": "拼接字段", "ccl_quantity": "CCL 数量", "pp_meter_quantity": "PP 米数换算", "manual": "不确定时留空，人工填写",
}

_EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)


def utcnow() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat()


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_dict(row: Any) -> dict[str, Any]:
    return dict(row) if row is not None else {}


def _record_event(conn, customer_id: int, action: str, *, before: dict[str, Any] | None = None,
                  after: dict[str, Any] | None = None, operated_by: str = "") -> None:
    conn.execute(
        "INSERT INTO automation_customer_events(customer_id,action,before_json,after_json,operated_by,created_at) VALUES (?,?,?,?,?,?)",
        (customer_id, action, json.dumps(before or {}, ensure_ascii=False), json.dumps(after or {}, ensure_ascii=False), operated_by, utcnow()),
    )


def list_customers(
    *,
    keyword: str = "",
    status: str = "all",
    clerk_name: str = "",
) -> list[dict[str, Any]]:
    clauses = ["1=1"]
    params: list[Any] = []
    if status in {"active", "disabled"}:
        clauses.append("c.status=?")
        params.append(status)
    if clerk_name.strip():
        clauses.append("c.internal_clerk_name LIKE ?")
        params.append(f"%{clerk_name.strip()}%")
    if keyword.strip():
        like = f"%{keyword.strip()}%"
        clauses.append(
            "(c.customer_code LIKE ? OR c.customer_short_name LIKE ? OR c.quick_code LIKE ? "
            "OR c.customer_name LIKE ? OR c.group_name LIKE ? OR c.contact_name LIKE ? "
            "OR c.contact_phone LIKE ? OR c.internal_clerk_name LIKE ? OR c.sales_name LIKE ? "
            "OR c.technical_support_name LIKE ?)"
        )
        params.extend([like] * 10)
    with db_cursor() as conn:
        rows = conn.execute(
            f"""SELECT c.*, COUNT(DISTINCT ct.id) AS contact_count,
                       COUNT(DISTINCT rr.id) AS routing_rule_count, COUNT(DISTINCT em.id) AS extraction_map_count
                  FROM automation_customers c
             LEFT JOIN automation_customer_contacts ct ON ct.customer_id=c.id AND ct.enabled=1
             LEFT JOIN automation_customer_routing_rules rr ON rr.customer_id=c.id AND rr.enabled=1
             LEFT JOIN automation_customer_extraction_maps em ON em.customer_id=c.id AND em.enabled=1
                 WHERE {' AND '.join(clauses)}
              GROUP BY c.id
              ORDER BY CASE c.status WHEN 'active' THEN 0 ELSE 1 END, c.customer_code""",
            params,
        ).fetchall()
    return [_row_dict(row) for row in rows]


def customer_summary() -> dict[str, int]:
    with db_cursor() as conn:
        row = conn.execute(
            "SELECT COUNT(*) AS total, SUM(CASE WHEN status='active' THEN 1 ELSE 0 END) AS active, "
            "SUM(CASE WHEN status='disabled' THEN 1 ELSE 0 END) AS disabled FROM automation_customers"
        ).fetchone()
    data = _row_dict(row)
    return {key: int(data.get(key) or 0) for key in ("total", "active", "disabled")}


def list_clerks() -> list[str]:
    with db_cursor() as conn:
        rows = conn.execute(
            "SELECT DISTINCT internal_clerk_name FROM automation_customers "
            "WHERE TRIM(internal_clerk_name)<>'' ORDER BY internal_clerk_name"
        ).fetchall()
    return [str(_row_dict(row).get("internal_clerk_name") or "") for row in rows]


def list_customer_choices() -> list[dict[str, Any]]:
    with db_cursor() as conn:
        rows = conn.execute(
            """SELECT customer_code,customer_short_name
                 FROM automation_customers
                WHERE status='active' AND TRIM(customer_code)<>''
                ORDER BY customer_code"""
        ).fetchall()
    return [_row_dict(row) for row in rows]


def get_customer(customer_id: int) -> dict[str, Any] | None:
    with db_cursor() as conn:
        row = conn.execute("SELECT * FROM automation_customers WHERE id=?", (customer_id,)).fetchone()
    return _row_dict(row) if row else None


def get_customer_workspace(customer_id: int) -> dict[str, Any] | None:
    customer = get_customer(customer_id)
    if not customer:
        return None
    with db_cursor() as conn:
        contacts = [_row_dict(row) for row in conn.execute(
            "SELECT * FROM automation_customer_contacts WHERE customer_id=? ORDER BY contact_type, id", (customer_id,)
        ).fetchall()]
        rules = [_row_dict(row) for row in conn.execute(
            "SELECT * FROM automation_customer_routing_rules WHERE customer_id=? ORDER BY priority,id", (customer_id,)
        ).fetchall()]
        for rule in rules:
            rule["conditions"] = [_row_dict(row) for row in conn.execute(
                "SELECT * FROM automation_customer_routing_conditions WHERE rule_id=? ORDER BY id", (rule["id"],)
            ).fetchall()]
        mappings = [_row_dict(row) for row in conn.execute(
            "SELECT * FROM automation_customer_extraction_maps WHERE customer_id=? ORDER BY sort_order,id", (customer_id,)
        ).fetchall()]
    return {"customer": customer, "contacts": contacts, "rules": rules, "mappings": mappings}


def save_customer(values: dict[str, Any], *, customer_id: int | None = None, operated_by: str = "") -> int:
    data = {key: as_text(values.get(key)) for key in CUSTOMER_FIELDS}
    if not data["customer_code"]:
        raise ValueError("客户编号不能为空。")
    if data["status"] not in {"active", "disabled"}:
        data["status"] = "active"
    now = utcnow()
    with db_cursor() as conn:
        if customer_id:
            before = _row_dict(conn.execute("SELECT * FROM automation_customers WHERE id=?", (customer_id,)).fetchone())
            if not before:
                raise ValueError("客户档案不存在。")
            exists = conn.execute("SELECT id FROM automation_customers WHERE customer_code=? AND id<>?", (data["customer_code"], customer_id)).fetchone()
            if exists:
                raise ValueError("客户编号已存在。")
            sets = ",".join(f"{key}=?" for key in CUSTOMER_FIELDS)
            conn.execute(f"UPDATE automation_customers SET {sets},updated_at=? WHERE id=?", (*[data[key] for key in CUSTOMER_FIELDS], now, customer_id))
            _record_event(conn, customer_id, "update_customer", before=before, after=data, operated_by=operated_by)
            return customer_id
        exists = conn.execute("SELECT id FROM automation_customers WHERE customer_code=?", (data["customer_code"],)).fetchone()
        if exists:
            raise ValueError("客户编号已存在，请在列表中编辑。")
        fields = ",".join(CUSTOMER_FIELDS)
        marks = ",".join("?" for _ in CUSTOMER_FIELDS)
        cursor = conn.execute(
            f"INSERT INTO automation_customers({fields},created_at,updated_at) VALUES ({marks},?,?)",
            (*[data[key] for key in CUSTOMER_FIELDS], now, now),
        )
        new_id = cursor.lastrowid
        if not new_id:
            row = conn.execute("SELECT id FROM automation_customers WHERE customer_code=?", (data["customer_code"],)).fetchone()
            new_id = _row_dict(row).get("id")
        _record_event(conn, int(new_id), "create_customer", after=data, operated_by=operated_by)
    return int(new_id)


def set_customer_status(customer_id: int, status: str, *, operated_by: str = "") -> None:
    if status not in {"active", "disabled"}:
        raise ValueError("客户状态不正确。")
    with db_cursor() as conn:
        before = _row_dict(conn.execute("SELECT * FROM automation_customers WHERE id=?", (customer_id,)).fetchone())
        if not before:
            raise ValueError("客户档案不存在。")
        conn.execute("UPDATE automation_customers SET status=?,updated_at=? WHERE id=?", (status, utcnow(), customer_id))
        _record_event(conn, customer_id, "set_customer_status", before=before, after={"status": status}, operated_by=operated_by)


def save_contact(customer_id: int, *, contact_type: str, contact_value: str, note: str = "", contact_id: int | None = None,
                 enabled: bool = True, operated_by: str = "") -> int:
    value = as_text(contact_value).lower()
    if contact_type not in CONTACT_TYPE_LABELS or not value:
        raise ValueError("请选择邮件身份类型并填写匹配值。")
    if contact_type in {"sender_email", "alias_email"} and not _EMAIL_RE.fullmatch(value):
        raise ValueError("邮箱格式不正确。")
    if contact_type == "sender_domain":
        value = value.removeprefix("@").strip()
        if "." not in value or "@" in value:
            raise ValueError("请填写邮箱域名，例如 customer.com。")
    now = utcnow()
    with db_cursor() as conn:
        if contact_id:
            before = _row_dict(conn.execute("SELECT * FROM automation_customer_contacts WHERE id=? AND customer_id=?", (contact_id, customer_id)).fetchone())
            if not before:
                raise ValueError("邮件身份不存在。")
            conn.execute("UPDATE automation_customer_contacts SET contact_type=?,contact_value=?,enabled=?,note=?,updated_at=? WHERE id=?", (contact_type, value, int(enabled), as_text(note), now, contact_id))
            saved_id = contact_id
        else:
            existing = conn.execute("SELECT id FROM automation_customer_contacts WHERE customer_id=? AND contact_type=? AND contact_value=?", (customer_id, contact_type, value)).fetchone()
            if existing:
                saved_id = int(_row_dict(existing)["id"])
                conn.execute("UPDATE automation_customer_contacts SET enabled=?,note=?,updated_at=? WHERE id=?", (int(enabled), as_text(note), now, saved_id))
            else:
                cursor = conn.execute("INSERT INTO automation_customer_contacts(customer_id,contact_type,contact_value,enabled,note,created_at,updated_at) VALUES (?,?,?,?,?,?,?)", (customer_id, contact_type, value, int(enabled), as_text(note), now, now))
                saved_id = int(cursor.lastrowid)
        _record_event(conn, customer_id, "save_mail_identity", after={"type": contact_type, "value": value}, operated_by=operated_by)
    return saved_id


def delete_contact(customer_id: int, contact_id: int, *, operated_by: str = "") -> None:
    with db_cursor() as conn:
        before = _row_dict(conn.execute("SELECT * FROM automation_customer_contacts WHERE id=? AND customer_id=?", (contact_id, customer_id)).fetchone())
        if before:
            conn.execute("DELETE FROM automation_customer_contacts WHERE id=?", (contact_id,))
            _record_event(conn, customer_id, "delete_mail_identity", before=before, operated_by=operated_by)


def save_routing_rule(customer_id: int, *, name: str, action_type: str, scope: str, keywords: Iterable[str], note: str = "",
                      priority: int = 100, rule_id: int | None = None, operated_by: str = "") -> int:
    words = list(dict.fromkeys(as_text(word) for word in keywords if as_text(word)))
    if not as_text(name) or action_type not in {"new_order", "order_change", "quotation"}:
        raise ValueError("请填写规则名称并选择录单、修改订单或报价。")
    if scope not in ROUTING_SCOPE_LABELS or not words:
        raise ValueError("请选择检索位置，并至少填写一个关键词。")
    now = utcnow()
    with db_cursor() as conn:
        if rule_id:
            before = _row_dict(conn.execute("SELECT * FROM automation_customer_routing_rules WHERE id=? AND customer_id=?", (rule_id, customer_id)).fetchone())
            if not before:
                raise ValueError("客户分流规则不存在。")
            conn.execute("UPDATE automation_customer_routing_rules SET name=?,action_type=?,priority=?,note=?,updated_at=? WHERE id=?", (as_text(name), action_type, max(1, int(priority or 100)), as_text(note), now, rule_id))
            conn.execute("DELETE FROM automation_customer_routing_conditions WHERE rule_id=?", (rule_id,))
            saved_id = rule_id
        else:
            cursor = conn.execute("INSERT INTO automation_customer_routing_rules(customer_id,name,action_type,enabled,priority,note,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?)", (customer_id, as_text(name), action_type, 1, max(1, int(priority or 100)), as_text(note), now, now))
            saved_id = int(cursor.lastrowid)
        conn.executemany("INSERT INTO automation_customer_routing_conditions(rule_id,scope,keyword,created_at) VALUES (?,?,?,?)", [(saved_id, scope, word, now) for word in words])
        _record_event(conn, customer_id, "save_customer_routing_rule", after={"rule_id": saved_id, "name": as_text(name), "scope": scope, "keywords": words}, operated_by=operated_by)
    return saved_id


def set_routing_rule_enabled(customer_id: int, rule_id: int, enabled: bool, *, operated_by: str = "") -> None:
    with db_cursor() as conn:
        row = _row_dict(conn.execute("SELECT * FROM automation_customer_routing_rules WHERE id=? AND customer_id=?", (rule_id, customer_id)).fetchone())
        if not row:
            raise ValueError("客户分流规则不存在。")
        conn.execute("UPDATE automation_customer_routing_rules SET enabled=?,updated_at=? WHERE id=?", (int(enabled), utcnow(), rule_id))
        _record_event(conn, customer_id, "set_customer_routing_rule", before=row, after={"enabled": enabled}, operated_by=operated_by)


def delete_routing_rule(customer_id: int, rule_id: int, *, operated_by: str = "") -> None:
    with db_cursor() as conn:
        row = _row_dict(conn.execute("SELECT * FROM automation_customer_routing_rules WHERE id=? AND customer_id=?", (rule_id, customer_id)).fetchone())
        if row:
            conn.execute("DELETE FROM automation_customer_routing_conditions WHERE rule_id=?", (rule_id,))
            conn.execute("DELETE FROM automation_customer_routing_rules WHERE id=?", (rule_id,))
            _record_event(conn, customer_id, "delete_customer_routing_rule", before=row, operated_by=operated_by)


def save_extraction_map(customer_id: int, *, target_field: str, source_kind: str, source_label: str, transform_type: str,
                        note: str = "", sort_order: int = 100, map_id: int | None = None, operated_by: str = "") -> int:
    if target_field not in EXTRACTION_TARGET_LABELS or source_kind not in EXTRACTION_SOURCE_KIND_LABELS or transform_type not in TRANSFORM_LABELS:
        raise ValueError("提取映射参数不正确。")
    if transform_type != "manual" and not as_text(source_label):
        raise ValueError("请填写客户订单中对应的字段名称。")
    now = utcnow()
    data = (
        target_field,
        source_kind,
        as_text(source_label),
        transform_type,
        "{}",
        max(1, int(sort_order or 100)),
        as_text(note),
        now,
    )
    with db_cursor() as conn:
        if map_id:
            before = _row_dict(conn.execute("SELECT * FROM automation_customer_extraction_maps WHERE id=? AND customer_id=?", (map_id, customer_id)).fetchone())
            if not before:
                raise ValueError("订单提取映射不存在。")
            conn.execute("UPDATE automation_customer_extraction_maps SET target_field=?,source_kind=?,source_label=?,transform_type=?,transform_config_json=?,sort_order=?,note=?,updated_at=? WHERE id=?", (*data, map_id))
            saved_id = map_id
        else:
            cursor = conn.execute(
                "INSERT INTO automation_customer_extraction_maps(customer_id,target_field,source_kind,source_label,transform_type,transform_config_json,enabled,sort_order,note,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (
                    customer_id,
                    target_field,
                    source_kind,
                    as_text(source_label),
                    transform_type,
                    "{}",
                    1,
                    max(1, int(sort_order or 100)),
                    as_text(note),
                    now,
                    now,
                ),
            )
            saved_id = int(cursor.lastrowid)
        _record_event(conn, customer_id, "save_order_extraction_mapping", after={"mapping_id": saved_id, "target": target_field, "source": source_label, "transform": transform_type}, operated_by=operated_by)
    return saved_id


def delete_extraction_map(customer_id: int, map_id: int, *, operated_by: str = "") -> None:
    with db_cursor() as conn:
        row = _row_dict(conn.execute("SELECT * FROM automation_customer_extraction_maps WHERE id=? AND customer_id=?", (map_id, customer_id)).fetchone())
        if row:
            conn.execute("DELETE FROM automation_customer_extraction_maps WHERE id=?", (map_id,))
            _record_event(conn, customer_id, "delete_order_extraction_mapping", before=row, operated_by=operated_by)


def get_enabled_extraction_maps(customer_id: int | None) -> list[dict[str, Any]]:
    """Return the active mappings in their explicit business-maintained order.

    A missing customer identity deliberately returns no mappings.  The caller
    must then use only the conservative universal extraction path instead of
    applying another customer's configuration by accident.
    """
    if not customer_id:
        return []
    with db_cursor() as conn:
        rows = conn.execute(
            """SELECT * FROM automation_customer_extraction_maps
                 WHERE customer_id=? AND enabled=1
              ORDER BY sort_order,id""",
            (customer_id,),
        ).fetchall()
    return [_row_dict(row) for row in rows]


def _sender_email(sender: str) -> str:
    match = _EMAIL_RE.search(sender or "")
    return match.group(0).lower() if match else ""


def identify_customer(sender: str) -> dict[str, Any] | None:
    """Return an unambiguous active customer for sender, otherwise None."""
    email = _sender_email(sender)
    if not email:
        return None
    domain = email.rsplit("@", 1)[-1]
    with db_cursor() as conn:
        exact = conn.execute(
            """SELECT c.id,c.customer_code,c.customer_name,ct.contact_type,ct.contact_value
                 FROM automation_customer_contacts ct JOIN automation_customers c ON c.id=ct.customer_id
                WHERE c.status='active' AND ct.enabled=1 AND ct.contact_type IN ('sender_email','alias_email') AND ct.contact_value=?""",
            (email,),
        ).fetchall()
        exact_rows = [_row_dict(row) for row in exact]
        ids = {row["id"] for row in exact_rows}
        if len(ids) == 1:
            row = exact_rows[0]
            return {"customer_id": row["id"], "customer_code": row["customer_code"], "customer_name": row["customer_name"], "match_source": row["contact_type"], "match_detail": email}
        if ids:
            return None
        domains = conn.execute(
            """SELECT c.id,c.customer_code,c.customer_name
                 FROM automation_customer_contacts ct JOIN automation_customers c ON c.id=ct.customer_id
                WHERE c.status='active' AND ct.enabled=1 AND ct.contact_type='sender_domain' AND ct.contact_value=?""",
            (domain,),
        ).fetchall()
    domain_rows = [_row_dict(row) for row in domains]
    if len({row["id"] for row in domain_rows}) != 1:
        return None
    row = domain_rows[0]
    return {"customer_id": row["id"], "customer_code": row["customer_code"], "customer_name": row["customer_name"], "match_source": "sender_domain", "match_detail": domain}


def match_customer_routing_rules(customer_id: int, values: dict[str, str]) -> dict[str, Any] | None:
    """Customer rules are used only after deterministic customer identification.

    All conditions inside one rule must match.  More than one target business
    category is deliberately returned as a conflict instead of guessing.
    """
    with db_cursor() as conn:
        rules = [_row_dict(row) for row in conn.execute(
            "SELECT * FROM automation_customer_routing_rules WHERE customer_id=? AND enabled=1 ORDER BY priority,id", (customer_id,)
        ).fetchall()]
        hits: list[dict[str, Any]] = []
        for rule in rules:
            conditions = [_row_dict(row) for row in conn.execute(
                "SELECT scope,keyword FROM automation_customer_routing_conditions WHERE rule_id=? ORDER BY id", (rule["id"],)
            ).fetchall()]
            if conditions and all(str(item["keyword"]).lower() in str(values.get(item["scope"], "")).lower() for item in conditions):
                hits.append({"rule_id": rule["id"], "name": rule["name"], "action_type": rule["action_type"], "conditions": conditions})
    if not hits:
        return None
    actions = {item["action_type"] for item in hits}
    if len(actions) != 1:
        return {"action_type": "unclassified", "state": "needs_business_routing", "reason": "同一客户的多个业务规则同时命中", "matches": hits}
    return {"action_type": next(iter(actions)), "state": "routed", "reason": "客户专属规则匹配", "matches": hits}


def customer_routing_needs_attachment_content(customer_id: int) -> bool:
    """Avoid parsing attachments unless this customer's enabled rules need them."""
    with db_cursor() as conn:
        row = conn.execute(
            """SELECT 1
                 FROM automation_customer_routing_conditions c
                 JOIN automation_customer_routing_rules r ON r.id=c.rule_id
                WHERE r.customer_id=? AND r.enabled=1 AND c.scope='attachment_content'
                LIMIT 1""",
            (customer_id,),
        ).fetchone()
    return row is not None


def import_customer_workbook(file_path: str | Path, *, operated_by: str = "") -> dict[str, int]:
    """Upsert the supplied HB002 workbook without requiring an employee number."""
    book = load_workbook(file_path, read_only=True, data_only=True)
    sheet = book.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5), values_only=True), None)
    # The supplied baseline has a blank first row; locate the row containing 客户编号.
    headers: list[str] = []
    header_index = 0
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True), start=1):
        values = [as_text(value) for value in row]
        if "客户编号" in values:
            headers, header_index = values, index
            break
    if not headers:
        raise ValueError("未找到“客户编号”表头，无法导入客户档案。")
    label_to_key = {label: key for key, label in CUSTOMER_FIELD_LABELS.items() if key not in {"status", "note", "internal_clerk_employee_id"}}
    imported = updated = skipped = 0
    try:
        with db_cursor() as conn:
            for row in sheet.iter_rows(min_row=header_index + 1, values_only=True):
                source = {headers[index]: as_text(value) for index, value in enumerate(row) if index < len(headers)}
                code = source.get("客户编号", "")
                if not code:
                    skipped += 1
                    continue
                data = {key: source.get(label, "") for label, key in label_to_key.items()}
                data.update({"customer_code": code, "status": "active", "note": ""})
                existing = _row_dict(conn.execute("SELECT id FROM automation_customers WHERE customer_code=?", (code,)).fetchone())
                now = utcnow()
                if existing:
                    fields = [key for key in CUSTOMER_FIELDS if key not in {"internal_clerk_employee_id", "status", "note"}]
                    conn.execute(f"UPDATE automation_customers SET {','.join(f'{key}=?' for key in fields)},updated_at=? WHERE id=?", (*[data.get(key, "") for key in fields], now, existing["id"]))
                    updated += 1
                else:
                    fields = list(CUSTOMER_FIELDS)
                    payload = {key: data.get(key, "") for key in fields}
                    cursor = conn.execute(f"INSERT INTO automation_customers({','.join(fields)},created_at,updated_at) VALUES ({','.join('?' for _ in fields)},?,?)", (*[payload[key] for key in fields], now, now))
                    _record_event(conn, int(cursor.lastrowid), "import_customer_workbook", after={"customer_code": code}, operated_by=operated_by)
                    imported += 1
    finally:
        book.close()
    return {"imported": imported, "updated": updated, "skipped": skipped}
