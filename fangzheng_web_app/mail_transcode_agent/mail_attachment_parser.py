from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..purchase_order_pipeline import run_purchase_order_pipeline


EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}
SPEC_HEADERS = {"规格", "客户规格", "品名", "物料名称", "名称规格", "物料规格", "物料描述"}
REMARK_HEADERS = {"备注", "订单备注", "需方备注", "供方备注"}
ORDER_HEADERS = {"订单号", "采购订单号", "订单编号", "客户订单号"}


def parse_attachment(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in EXCEL_SUFFIXES:
        return _parse_excel(path)
    if suffix in IMAGE_SUFFIXES or suffix == ".pdf":
        return _parse_pdf_image(path)
    return []


def _parse_excel(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    header_index = -1
    header_row: list[str] = []
    for index, row in enumerate(rows[:20]):
        values = [str(value or "").strip() for value in row]
        if any(value in SPEC_HEADERS for value in values):
            header_index = index
            header_row = values
            break
    if header_index < 0:
        return []

    def find_col(names: set[str]) -> int:
        for col, value in enumerate(header_row):
            if value in names:
                return col
        return -1

    spec_col = find_col(SPEC_HEADERS)
    remark_col = find_col(REMARK_HEADERS)
    order_col = find_col(ORDER_HEADERS)
    result: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        spec = str(row[spec_col]).strip() if spec_col >= 0 and spec_col < len(row) else ""
        if not spec:
            continue
        result.append(
            {
                "spec": spec,
                "remark": str(row[remark_col]).strip() if remark_col >= 0 and remark_col < len(row) else "",
                "order_number": str(row[order_col]).strip() if order_col >= 0 and order_col < len(row) else "",
                "customer": "",
            }
        )
    return result


def _parse_pdf_image(path: Path) -> list[dict[str, Any]]:
    document = run_purchase_order_pipeline(
        {
            "stored_path": str(path),
            "original_filename": path.name,
        }
    )
    header = document.get("header_info") or {}
    order_number = str(header.get("订单号") or "").strip()
    supplier = str(header.get("供应商") or "").strip()
    result: list[dict[str, Any]] = []
    for row in document.get("mapped_detail_rows") or []:
        original = row.get("original") or {}
        standard = row.get("standard") or {}
        spec = str(standard.get("物料名称") or original.get("名称规格") or "").strip()
        remark = str(standard.get("备注") or original.get("备注") or "").strip()
        if not spec:
            continue
        result.append(
            {
                "spec": spec,
                "remark": remark,
                "order_number": order_number,
                "customer": supplier,
            }
        )
    if not result and order_number:
        result.append({"spec": "", "remark": "", "order_number": order_number, "customer": supplier})
    return result
