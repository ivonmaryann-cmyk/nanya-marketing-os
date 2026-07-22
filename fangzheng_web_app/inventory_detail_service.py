from __future__ import annotations

import json
import math
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Iterator

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename

from .db import append_job_log, create_job, get_job, update_job_status
from .file_utils import safe_unlink
from .job_control import launch_job_process
from .paths import JOBS_DIR


FEATURE = "inventory_detail"
RULE_VERSION = "库存明细内置规则 v1"
PLAN_A_RULE_VERSION = "计划A级库存明细内置规则 v1"
ALLOWED_EXTENSIONS = {".xls", ".xlsx"}
WAREHOUSE_MODE = "warehouse"
PLAN_A_MODE = "plan_a"
INVENTORY_MODES = {WAREHOUSE_MODE, PLAN_A_MODE}

NO_ORDER_PREFIX = "无订单"
SCRAP_TEXT = "报废"

A_HEADERS = [
    "品名",
    "品号",
    "规格",
    "上海_江西",
    "胶系",
    "厚度",
    "铜箔",
    "单重",
    "数量",
    "折合大板",
    "是否光板",
    "规格胶系",
    "厚度mm",
    "厚度类型",
    "铜厚规格",
    "尺寸长",
    "尺寸宽",
    "尺寸",
    "铜箔类型",
    "基板等级",
    "颜色",
    "水印",
    "特殊备注",
    "解析状态",
]

PLAN_A_HEADERS = [
    "品名",
    "品号",
    "规格",
    "排版结构",
    "上海_江西",
    "胶系",
    "厚度",
    "铜箔",
    "单重",
    "数量",
    "折合大板",
    "是否光板",
    "规格胶系",
    "厚度mm",
    "厚度类型",
    "铜厚规格",
    "尺寸长",
    "尺寸宽",
    "尺寸",
    "铜箔类型",
    "基板等级",
    "颜色",
    "水印",
    "特殊备注",
    "解析状态",
]

B_HEADERS = [
    "规格",
    "上海_江西",
    "胶系",
    "厚度",
    "铜箔",
    "数量",
    "折合大板",
    "是否光板",
    "规格胶系",
    "厚度mm",
    "厚度类型",
    "铜厚规格",
    "尺寸长",
    "尺寸宽",
    "尺寸",
    "铜箔类型",
    "基板等级",
    "颜色",
    "水印",
    "特殊备注",
    "解析状态",
]

REQUIRED_HEADERS = {
    "品号",
    "品名",
    "数量",
    "折合大板数量",
    "规格",
    "类别",
    "厚度",
    "铜箔",
    "尺寸",
    "颜色",
    "水印",
    "等级",
    "库位信息",
    "芯/总厚",
    "单重",
}

PLAN_A_REQUIRED_HEADERS = {
    "品名",
    "品号",
    "规格",
    "排版结构",
    "上海_江西",
    "胶系",
    "胶系名称",
    "厚度",
    "铜箔",
    "单重",
    "库龄",
    "数量",
    "折合大板",
}

NAVY = "1F4E78"
DARK_NAVY = "17365D"
TEAL = "0F766E"
LIGHT_BLUE = "EAF2F8"
LIGHT_TEAL = "E2F2EF"
LIGHT_GRAY = "F4F7FB"
BORDER_BLUE = "C7D6E6"
TEXT_DARK = "17365D"
MUTED = "667085"
WHITE = "FFFFFF"
LINK_BLUE = "2F80ED"
ERROR_FILL = "FFF2CC"


@dataclass
class AggregateRow:
    grade: str
    plant: str
    spec: str
    product_name: str = ""
    product_no: str = ""
    layout: str = ""
    glue: str = ""
    thickness_raw: Any = ""
    copper_raw: Any = ""
    size_class: str = ""
    color_raw: str = ""
    watermark_raw: str = ""
    thickness_type_raw: str = ""
    unit_weight: Any = ""
    quantity: float = 0.0
    folded_large_quantity: float = 0.0
    source_row_count: int = 0
    issue_notes: list[str] = field(default_factory=list)
    age_values: list[float] = field(default_factory=list)
    prefer_source_glue: bool = False


@dataclass
class InventoryRow:
    grade: str
    plant: str
    product_name: str
    product_no: str
    layout: str
    spec: str
    glue: str
    thickness_raw: Any
    copper_raw: str
    unit_weight: Any
    quantity: float
    folded_large_quantity: float
    is_light: str
    spec_glue: str
    thickness_mm: float | None
    thickness_type: str
    copper_spec: str
    length: float | None
    width: float | None
    size_text: str
    copper_type: str
    spec_grade: str
    color: str
    watermark: str
    special_note: str
    parse_status: str


def normalize_inventory_mode(value: Any) -> str:
    mode = _text(value).lower()
    return mode if mode in INVENTORY_MODES else WAREHOUSE_MODE


def queue_inventory_detail_job(
    employee_id: str,
    shanghai_file=None,
    jiangxi_file=None,
    *,
    inventory_mode: str = WAREHOUSE_MODE,
    plan_a_file=None,
) -> int:
    inventory_mode = normalize_inventory_mode(inventory_mode)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    job_dir = employee_dir / f"{timestamp}_inventory_detail"
    job_dir.mkdir(parents=True, exist_ok=True)

    manifest: dict[str, Any] = {"feature": FEATURE, "inventory_mode": inventory_mode}
    if inventory_mode == PLAN_A_MODE:
        if not plan_a_file or not plan_a_file.filename:
            raise ValueError("请先上传计划A级库存表")
        plan_a_path = _save_upload(plan_a_file, job_dir, "plan_a")
        manifest.update(
            {
                "plan_a_path": str(plan_a_path),
                "plan_a_original_name": plan_a_file.filename,
            }
        )
        source_filename = f"计划A级库存明细：{plan_a_file.filename}"
        rule_version = PLAN_A_RULE_VERSION
    else:
        if not shanghai_file or not shanghai_file.filename or not jiangxi_file or not jiangxi_file.filename:
            raise ValueError("请同时上传上海厂和江西厂库存表")
        shanghai_path = _save_upload(shanghai_file, job_dir, "shanghai")
        jiangxi_path = _save_upload(jiangxi_file, job_dir, "jiangxi")
        manifest.update(
            {
                "shanghai_path": str(shanghai_path),
                "jiangxi_path": str(jiangxi_path),
                "shanghai_original_name": shanghai_file.filename,
                "jiangxi_original_name": jiangxi_file.filename,
            }
        )
        source_filename = f"库存明细：{shanghai_file.filename} + {jiangxi_file.filename}"
        rule_version = RULE_VERSION
    manifest_path = job_dir / "inventory_detail_input.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    job_id = create_job(
        employee_id,
        source_filename,
        str(manifest_path),
        rule_version,
        feature=FEATURE,
    )
    launch_job_process(job_id, FEATURE, employee_id)
    return job_id


def _save_upload(file_obj, job_dir: Path, prefix: str) -> Path:
    original_name = (file_obj.filename or "").strip()
    suffix = Path(original_name).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("库存明细仅支持 .xls / .xlsx 文件")
    safe_name = secure_filename(Path(original_name).stem) or prefix
    target = job_dir / f"{prefix}_{safe_name}{suffix}"
    file_obj.save(target)
    return target


def run_inventory_detail_job(job_id: int, employee_id: str) -> None:
    job = get_job(job_id)
    if not job or job["employee_id"] != employee_id:
        return

    update_job_status(job_id, status="running", log_text="")
    try:
        input_manifest_path = Path(job["stored_input_path"])
        manifest = json.loads(input_manifest_path.read_text(encoding="utf-8"))
        inventory_mode = normalize_inventory_mode(manifest.get("inventory_mode"))
        job_dir = input_manifest_path.parent
        date_text = datetime.now().strftime("%Y%m%d")
        append_job_log(job_id, "开始处理库存明细任务。")

        if inventory_mode == PLAN_A_MODE:
            plan_a_path = Path(manifest["plan_a_path"])
            append_job_log(job_id, f"计划A级输入：{plan_a_path.name}")
            plan_rows, stats = load_plan_a_rows(plan_a_path, job_id=job_id)
            output_path = job_dir / f"计划A级库存明细_{date_text}_A级_胶系分类导航版.xlsx"
            append_job_log(job_id, f"开始生成计划A级导航版，共 {len(plan_rows)} 行。")
            build_inventory_workbook(plan_rows, "A", output_path, workbook_mode=PLAN_A_MODE)
            append_job_log(job_id, f"计划A级导航版已生成：{output_path.name}")
            result_files = {"plan-a": str(output_path)}
            total_rows = len(plan_rows)
            issue_count = sum(row.parse_status != "正常" for row in plan_rows)
            filtered_count = stats["plan_a"]["filtered"]
            completion_log = f"库存明细处理完成：计划A级 {total_rows} 行。"
        else:
            shanghai_path = Path(manifest["shanghai_path"])
            jiangxi_path = Path(manifest["jiangxi_path"])
            append_job_log(job_id, f"上海厂输入：{shanghai_path.name}")
            append_job_log(job_id, f"江西厂输入：{jiangxi_path.name}")
            rows_by_grade, stats = load_inventory_rows(shanghai_path, jiangxi_path, job_id=job_id)
            a_path = job_dir / f"无订单库存明细_{date_text}_A级_胶系分类导航版.xlsx"
            b_path = job_dir / f"无订单库存明细_{date_text}_B级_胶系分类导航版.xlsx"
            append_job_log(job_id, f"开始生成A级导航版，共 {len(rows_by_grade['A'])} 行。")
            build_inventory_workbook(rows_by_grade["A"], "A", a_path)
            append_job_log(job_id, f"A级导航版已生成：{a_path.name}")
            append_job_log(job_id, f"开始生成B级导航版，共 {len(rows_by_grade['B'])} 行。")
            build_inventory_workbook(rows_by_grade["B"], "B", b_path)
            append_job_log(job_id, f"B级导航版已生成：{b_path.name}")
            result_files = {"a": str(a_path), "b": str(b_path)}
            total_rows = len(rows_by_grade["A"]) + len(rows_by_grade["B"])
            issue_count = sum(1 for rows in rows_by_grade.values() for row in rows if row.parse_status != "正常")
            filtered_count = stats["shanghai"]["filtered"] + stats["jiangxi"]["filtered"]
            completion_log = f"库存明细处理完成：A级 {len(rows_by_grade['A'])} 行，B级 {len(rows_by_grade['B'])} 行。"

        result_manifest = {
            "feature": FEATURE,
            "inventory_mode": inventory_mode,
            "files": result_files,
            "stats": stats,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        }
        result_manifest_path = job_dir / "inventory_detail_results.json"
        result_manifest_path.write_text(
            json.dumps(result_manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        update_job_status(
            job_id,
            status="completed",
            stored_result_path=str(result_manifest_path),
            success_count=total_rows - issue_count,
            fail_count=issue_count,
            skip_count=filtered_count,
            current_row=total_rows,
            total_rows=total_rows,
            completed=True,
        )
        append_job_log(job_id, completion_log)
    except Exception as exc:
        append_job_log(job_id, f"库存明细处理失败：{exc}")
        update_job_status(job_id, status="failed", error_message=str(exc), completed=True)
        raise


def load_inventory_result_manifest(job: Any) -> dict[str, Any]:
    if not job or job["feature"] != FEATURE:
        return {}
    path_value = job["stored_result_path"]
    if not path_value:
        return {}
    path = Path(path_value)
    if not path.exists() or path.suffix.lower() != ".json":
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def load_inventory_input_manifest(job: Any) -> dict[str, Any]:
    if not job or job["feature"] != FEATURE:
        return {}
    path_value = job["stored_input_path"]
    if not path_value:
        return {}
    path = Path(path_value)
    if not path.exists() or path.suffix.lower() != ".json":
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def get_inventory_result_path(job: Any, grade: str) -> Path | None:
    grade_key = grade.strip().lower()
    if grade_key not in {"a", "b", "plan-a"}:
        return None
    manifest = load_inventory_result_manifest(job)
    path_value = (manifest.get("files") or {}).get(grade_key)
    if not path_value:
        return None
    path = Path(path_value)
    return path if path.exists() and path.suffix.lower() == ".xlsx" else None


def cleanup_inventory_detail_job_files(job: Any) -> None:
    paths: set[Path] = set()
    for key in ("stored_input_path", "stored_result_path"):
        value = job[key] if job else None
        if value:
            paths.add(Path(value))

    input_path = Path(job["stored_input_path"]) if job and job["stored_input_path"] else None
    if input_path and input_path.exists() and input_path.suffix.lower() == ".json":
        try:
            data = json.loads(input_path.read_text(encoding="utf-8"))
            for key in ("shanghai_path", "jiangxi_path", "plan_a_path"):
                if data.get(key):
                    paths.add(Path(data[key]))
        except (OSError, json.JSONDecodeError):
            pass

    result_path = Path(job["stored_result_path"]) if job and job["stored_result_path"] else None
    if result_path and result_path.exists() and result_path.suffix.lower() == ".json":
        try:
            data = json.loads(result_path.read_text(encoding="utf-8"))
            for value in (data.get("files") or {}).values():
                if value:
                    paths.add(Path(value))
        except (OSError, json.JSONDecodeError):
            pass

    parent_dirs = {path.parent for path in paths}
    for path in paths:
        safe_unlink(str(path))
    for directory in sorted(parent_dirs, key=lambda item: len(str(item)), reverse=True):
        try:
            directory.rmdir()
        except OSError:
            pass


def load_inventory_rows(
    shanghai_path: str | Path,
    jiangxi_path: str | Path,
    *,
    job_id: int | None = None,
) -> tuple[dict[str, list[InventoryRow]], dict[str, Any]]:
    aggregates: OrderedDict[tuple[str, ...], AggregateRow] = OrderedDict()
    stats = {
        "shanghai": {"read": 0, "kept": 0, "filtered": 0},
        "jiangxi": {"read": 0, "kept": 0, "filtered": 0},
    }

    _consume_inventory_file(Path(shanghai_path), "上海", "shanghai", aggregates, stats["shanghai"], job_id)
    _consume_inventory_file(Path(jiangxi_path), "江西", "jiangxi", aggregates, stats["jiangxi"], job_id)

    rows_by_grade: dict[str, list[InventoryRow]] = {"A": [], "B": []}
    for aggregate in aggregates.values():
        row = _finalize_inventory_row(aggregate)
        rows_by_grade[row.grade].append(row)
    for rows in rows_by_grade.values():
        rows.sort(key=_inventory_sort_key)

    stats["merged"] = {
        "a_rows": len(rows_by_grade["A"]),
        "b_rows": len(rows_by_grade["B"]),
        "total_rows": len(rows_by_grade["A"]) + len(rows_by_grade["B"]),
    }
    return rows_by_grade, stats


def load_plan_a_rows(
    plan_a_path: str | Path,
    *,
    job_id: int | None = None,
) -> tuple[list[InventoryRow], dict[str, Any]]:
    path = Path(plan_a_path)
    if path.suffix.lower() not in ALLOWED_EXTENSIONS:
        raise ValueError(f"不支持的计划库存文件格式：{path.name}")

    headers, source_rows = _iter_sheet_rows(path)
    indexes = _header_indexes(headers)
    missing = sorted(PLAN_A_REQUIRED_HEADERS - set(indexes))
    if missing:
        close_iterator = getattr(source_rows, "close", None)
        if callable(close_iterator):
            close_iterator()
        raise ValueError(f"{path.name} 缺少必要字段：{', '.join(missing)}")

    aggregates: OrderedDict[tuple[str, ...], AggregateRow] = OrderedDict()
    source_stats = {"read": 0, "kept": 0, "filtered": 0}
    for excel_row, values in source_rows:
        source_stats["read"] += 1
        spec = _normalize_spec(_get(values, indexes, "规格"))
        if re.search(r"B3\s*$", spec, re.IGNORECASE):
            source_stats["filtered"] += 1
            continue

        plant = _text(_get(values, indexes, "上海_江西"))
        category_code = _get(values, indexes, "胶系")
        glue_name = _get(values, indexes, "胶系名称")
        thickness_value = _get(values, indexes, "厚度")
        copper_value = _get(values, indexes, "铜箔")
        length, width = _extract_dimensions(spec)
        watermark = _extract_watermark(spec, "")
        quantity, quantity_ok = _number(_get(values, indexes, "数量"))
        folded, folded_ok = _number(_get(values, indexes, "折合大板"))
        issue_notes: list[str] = []
        if not quantity_ok:
            issue_notes.append(f"原始第{excel_row}行数量无效，按0处理")
        if not folded_ok:
            issue_notes.append(f"原始第{excel_row}行折合大板无效，按0处理")
        if not spec:
            issue_notes.append(f"原始第{excel_row}行规格为空")

        dimension_key = (
            f"{length:.12g}x{width:.12g}"
            if length is not None and width is not None
            else ""
        )
        key = (
            "A",
            _normalize_merge_text(plant),
            _normalize_merge_text(spec, upper=True),
            _normalize_merge_text(category_code, compact=True, upper=True),
            _normalize_merge_number(thickness_value),
            _normalize_merge_text(copper_value, compact=True, upper=True),
            dimension_key,
            _normalize_merge_text(watermark),
        )
        row = aggregates.get(key)
        if row is None:
            row = AggregateRow(
                grade="A",
                plant=plant,
                spec=spec,
                product_name=_text(_get(values, indexes, "品名")),
                product_no=_text(_get(values, indexes, "品号")),
                layout=_text(_get(values, indexes, "排版结构")),
                glue=_text(glue_name),
                thickness_raw=thickness_value,
                copper_raw=copper_value,
                unit_weight=_get(values, indexes, "单重"),
                prefer_source_glue=True,
            )
            aggregates[key] = row
        else:
            _fill_first_nonblank(row, "product_name", _get(values, indexes, "品名"))
            _fill_first_nonblank(row, "product_no", _get(values, indexes, "品号"))
            _fill_first_nonblank(row, "layout", _get(values, indexes, "排版结构"))
            _fill_first_nonblank(row, "glue", glue_name)
            _fill_first_nonblank(row, "thickness_raw", thickness_value)
            _fill_first_nonblank(row, "copper_raw", copper_value)
            _fill_first_nonblank(row, "unit_weight", _get(values, indexes, "单重"))

        age = _optional_number(_get(values, indexes, "库龄"))
        if age is not None:
            row.age_values.append(age)
        row.quantity += quantity
        row.folded_large_quantity += folded
        row.source_row_count += 1
        row.issue_notes.extend(issue_notes)
        source_stats["kept"] += 1
        if job_id and source_stats["read"] % 5000 == 0:
            append_job_log(job_id, f"计划A级库存已读取 {source_stats['read']} 行。")

    rows = [_finalize_inventory_row(aggregate) for aggregate in aggregates.values()]
    rows.sort(key=_inventory_sort_key)
    stats = {
        "plan_a": source_stats,
        "merged": {"a_rows": len(rows), "total_rows": len(rows)},
    }
    if job_id:
        append_job_log(
            job_id,
            f"计划A级清洗完成：读取 {source_stats['read']} 行，排除B3 {source_stats['filtered']} 行，合并后 {len(rows)} 行。",
        )
    return rows, stats


def _consume_inventory_file(
    path: Path,
    plant: str,
    source_key: str,
    aggregates: OrderedDict[tuple[str, ...], AggregateRow],
    stats: dict[str, int],
    job_id: int | None,
) -> None:
    suffix = path.suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError(f"不支持的库存文件格式：{path.name}")

    headers, rows = _iter_sheet_rows(path)
    indexes = _header_indexes(headers)
    missing = sorted(REQUIRED_HEADERS - set(indexes))
    if missing:
        close_iterator = getattr(rows, "close", None)
        if callable(close_iterator):
            close_iterator()
        raise ValueError(f"{path.name} 缺少必要字段：{', '.join(missing)}")

    for excel_row, values in rows:
        stats["read"] += 1
        grade_value = _get(values, indexes, "等级")
        grade_raw = _text(grade_value)
        spec = _normalize_spec(_get(values, indexes, "规格"))
        location_info = _text(_get(values, indexes, "库位信息"))
        category_value = _get(values, indexes, "类别")
        thickness_value = _get(values, indexes, "厚度")
        copper_value = _get(values, indexes, "铜箔")
        size_value = _get(values, indexes, "尺寸")
        watermark_value = _get(values, indexes, "水印")

        if source_key == "shanghai":
            if not grade_raw or _is_percentage_grade(grade_raw) or _is_numeric_percentage_grade(grade_value):
                stats["filtered"] += 1
                continue
        else:
            if not location_info.startswith(NO_ORDER_PREFIX) or SCRAP_TEXT in spec:
                stats["filtered"] += 1
                continue

        grade = "B" if grade_raw.upper() == "B3" else "A"
        quantity, quantity_ok = _number(_get(values, indexes, "数量"))
        folded_large_quantity, _ = _number(_get(values, indexes, "折合大板数量"))
        issue_notes: list[str] = []
        if not quantity_ok:
            issue_notes.append(f"原始第{excel_row}行数量无效，按0处理")
        if not spec:
            issue_notes.append(f"原始第{excel_row}行规格为空")

        key = (
            grade,
            plant,
            spec,
            _normalize_merge_text(category_value),
            _normalize_merge_number(thickness_value),
            _normalize_merge_text(copper_value, compact=True, upper=True),
            _normalize_merge_text(size_value, compact=True, upper=True),
            _normalize_merge_text(watermark_value),
        )
        row = aggregates.get(key)
        if row is None:
            row = AggregateRow(
                grade=grade,
                plant=plant,
                spec=spec,
                product_name=_text(_get(values, indexes, "品名")),
                product_no=_text(_get(values, indexes, "品号")),
                glue=_text(category_value),
                thickness_raw=thickness_value,
                copper_raw=copper_value,
                size_class=_text(size_value),
                color_raw=_text(_get(values, indexes, "颜色")),
                watermark_raw=_text(watermark_value),
                thickness_type_raw=_text(_get(values, indexes, "芯/总厚")),
                unit_weight=_get(values, indexes, "单重"),
            )
            aggregates[key] = row
        else:
            _fill_first_nonblank(row, "product_name", _get(values, indexes, "品名"))
            _fill_first_nonblank(row, "product_no", _get(values, indexes, "品号"))
            _fill_first_nonblank(row, "glue", _get(values, indexes, "类别"))
            _fill_first_nonblank(row, "thickness_raw", _get(values, indexes, "厚度"))
            _fill_first_nonblank(row, "copper_raw", _get(values, indexes, "铜箔"))
            _fill_first_nonblank(row, "size_class", _get(values, indexes, "尺寸"))
            _fill_first_nonblank(row, "color_raw", _get(values, indexes, "颜色"))
            _fill_first_nonblank(row, "watermark_raw", _get(values, indexes, "水印"))
            _fill_first_nonblank(row, "thickness_type_raw", _get(values, indexes, "芯/总厚"))
            _fill_first_nonblank(row, "unit_weight", _get(values, indexes, "单重"))

        row.quantity += quantity
        row.folded_large_quantity += folded_large_quantity
        row.source_row_count += 1
        row.issue_notes.extend(issue_notes)

        if job_id and stats["read"] % 5000 == 0:
            append_job_log(job_id, f"{plant}厂已读取 {stats['read']} 行。")

    stats["kept"] = sum(row.source_row_count for row in aggregates.values() if row.plant == plant)
    if job_id:
        append_job_log(
            job_id,
            f"{plant}厂清洗完成：读取 {stats['read']} 行，过滤 {stats['filtered']} 行。",
        )


def _iter_sheet_rows(path: Path) -> tuple[list[Any], Iterator[tuple[int, list[Any]]]]:
    if path.suffix.lower() == ".xls":
        workbook = xlrd.open_workbook(str(path), on_demand=True, formatting_info=False)
        sheet = workbook.sheet_by_index(0)
        headers = sheet.row_values(0)

        def iterator() -> Iterator[tuple[int, list[Any]]]:
            try:
                for index in range(1, sheet.nrows):
                    yield index + 1, sheet.row_values(index)
            finally:
                workbook.release_resources()

        return headers, iterator()

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    value_rows = sheet.iter_rows(values_only=True)
    headers = list(next(value_rows, ()))

    def iterator() -> Iterator[tuple[int, list[Any]]]:
        try:
            for index, values in enumerate(value_rows, start=2):
                yield index, list(values)
        finally:
            workbook.close()

    return headers, iterator()


def _header_indexes(headers: Iterable[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(headers):
        name = _text(value)
        if name and name not in result:
            result[name] = index
    return result


def _get(values: list[Any], indexes: dict[str, int], name: str) -> Any:
    index = indexes[name]
    return values[index] if index < len(values) else None


def _fill_first_nonblank(row: AggregateRow, field_name: str, value: Any) -> None:
    current = getattr(row, field_name)
    if current not in (None, ""):
        return
    if value not in (None, ""):
        setattr(row, field_name, value)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_spec(value: Any) -> str:
    return " ".join(_text(value).split())


def _normalize_merge_text(value: Any, *, compact: bool = False, upper: bool = False) -> str:
    text = " ".join(_text(value).split())
    if compact:
        text = text.replace(" ", "")
    return text.upper() if upper else text


def _normalize_merge_number(value: Any) -> str:
    number, ok = _number(value)
    return f"{number:.12g}" if ok else _normalize_merge_text(value, compact=True, upper=True)


def _is_percentage_grade(value: str) -> bool:
    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?\s*%", value.strip()))


def _is_numeric_percentage_grade(value: Any) -> bool:
    return not isinstance(value, bool) and isinstance(value, (int, float)) and 0 <= float(value) <= 1


def _number(value: Any) -> tuple[float, bool]:
    if value in (None, ""):
        return 0.0, False
    if isinstance(value, bool):
        return 0.0, False
    if isinstance(value, (int, float)):
        if math.isfinite(float(value)):
            return float(value), True
        return 0.0, False
    text = str(value).strip().replace(",", "")
    try:
        number = float(text)
        return (number, True) if math.isfinite(number) else (0.0, False)
    except ValueError:
        return 0.0, False


def _optional_number(value: Any) -> float | None:
    number, ok = _number(value)
    return number if ok else None


def _finalize_inventory_row(row: AggregateRow) -> InventoryRow:
    spec = row.spec
    thickness_mm = _optional_number(row.thickness_raw)
    if thickness_mm is None:
        match = re.search(r"(\d+(?:\.\d+)?)\s*mm", spec, re.IGNORECASE)
        thickness_mm = float(match.group(1)) if match else None

    thickness_type = _normalize_thickness_type(row.thickness_type_raw, spec)
    copper_spec = _text(row.copper_raw)
    if not copper_spec:
        copper_spec = _extract_copper_spec(spec)
    length, width = _extract_dimensions(spec)
    size_text = f"{length:.2f}x{width:.2f}" if length is not None and width is not None else ""
    # The exported "类别" field is not consistently a glue-system name (some
    # rows contain numeric classifications), while the specification begins
    # with the business glue-system label used by the navigation workbooks.
    parsed_glue = _extract_glue(spec)
    spec_glue = (row.glue or parsed_glue) if row.prefer_source_glue else (parsed_glue or row.glue)
    copper_type = _extract_copper_type(spec, row.watermark_raw)
    spec_grade = _extract_spec_grade(spec)
    color = _extract_color(spec, row.color_raw)
    watermark = _extract_watermark(spec, row.watermark_raw)
    is_light = "是" if _is_light_board(spec, copper_spec) else "否"

    notes: list[str] = []
    if row.size_class:
        notes.append(f"尺寸分类：{row.size_class}")
    if row.age_values:
        minimum_age = min(row.age_values)
        maximum_age = max(row.age_values)
        if math.isclose(minimum_age, maximum_age):
            notes.append(f"库龄：{_format_compact_number(minimum_age)}个月")
        else:
            notes.append(
                f"库龄：{_format_compact_number(minimum_age)}–{_format_compact_number(maximum_age)}个月"
            )
    notes.extend(row.issue_notes)

    missing: list[str] = []
    if not spec_glue:
        missing.append("胶系未识别")
    if thickness_mm is None:
        missing.append("厚度未识别")
    if not thickness_type:
        missing.append("厚度类型未识别")
    if not copper_spec:
        missing.append("铜厚未识别")
    if length is None or width is None:
        missing.append("尺寸未识别")
    if row.issue_notes:
        missing.extend(row.issue_notes)
    parse_status = "正常" if not missing else "待确认：" + "；".join(dict.fromkeys(missing))

    return InventoryRow(
        grade=row.grade,
        plant=row.plant,
        product_name=row.product_name,
        product_no=row.product_no,
        layout=row.layout,
        spec=spec,
        glue=spec_glue,
        thickness_raw=_excel_number(thickness_mm) if thickness_mm is not None else _text(row.thickness_raw),
        copper_raw=copper_spec,
        unit_weight=_number_or_text(row.unit_weight),
        quantity=row.quantity,
        folded_large_quantity=row.folded_large_quantity,
        is_light=is_light,
        spec_glue=spec_glue,
        thickness_mm=thickness_mm,
        thickness_type=thickness_type,
        copper_spec=copper_spec,
        length=length,
        width=width,
        size_text=size_text,
        copper_type=copper_type,
        spec_grade=spec_grade,
        color=color,
        watermark=watermark,
        special_note="；".join(dict.fromkeys(notes)),
        parse_status=parse_status,
    )


def _normalize_thickness_type(raw_value: Any, spec: str) -> str:
    raw = _text(raw_value)
    if "芯" in raw:
        return "芯厚"
    if "总" in raw:
        return "总厚"
    if "芯厚" in spec:
        return "芯厚"
    if "总厚" in spec:
        return "总厚"
    return ""


def _extract_copper_spec(spec: str) -> str:
    patterns = [
        r"(?<!\w)(\d+(?:\.\d+)?(?:um)?(?:\+\d+(?:\.\d+)?um)?/[A-Za-z0-9.]+(?:\+\d+(?:\.\d+)?um)?)(?!\w)",
        r"(?<!\w)([A-Z0-9.]+/[A-Z0-9.]+)(?!\w)",
        r"(?<!\w)(00)(?!\w)",
    ]
    for pattern in patterns:
        match = re.search(pattern, spec, re.IGNORECASE)
        if match:
            value = match.group(1)
            return "0/0" if value == "00" else value
    return ""


def _extract_dimensions(spec: str) -> tuple[float | None, float | None]:
    matches = list(re.finditer(r"(?<!\d)(\d+(?:\.\d+)?)\s*[xX×*]\s*(\d+(?:\.\d+)?)(?!\d)", spec))
    if not matches:
        return None, None
    match = matches[-1]
    return float(match.group(1)), float(match.group(2))


def _extract_glue(spec: str) -> str:
    match = re.match(r"\s*([A-Za-z0-9][A-Za-z0-9\- ()（）]+?)\s+(?:\d|[A-Za-z]/)", spec)
    return match.group(1).strip() if match else ""


def _extract_copper_type(spec: str, watermark_raw: str) -> str:
    text = f"{spec} {watermark_raw}"
    found = re.findall(r"载体铜箔|HVLP\d*|RTF|HTE|VLP|DSTF", text, re.IGNORECASE)
    unique: list[str] = []
    for item in found:
        normalized = item.upper() if item != "载体铜箔" else item
        if normalized not in unique:
            unique.append(normalized)
    return "+".join(unique[:2])


def _extract_spec_grade(spec: str) -> str:
    match = re.search(r"([A-D])级", spec, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    match = re.search(r"\b([A-D])\b(?=\s+(?:有水印|无水印|无印字))", spec, re.IGNORECASE)
    return match.group(1).upper() if match else ""


def _extract_color(spec: str, raw: str) -> str:
    for color in ("黄色", "自然色", "白色", "黑色", "棕色", "红色", "蓝色"):
        if color in spec:
            return color
    return raw if any("\u4e00" <= char <= "\u9fff" for char in raw) else ""


def _extract_watermark(spec: str, raw: str) -> str:
    text = f"{spec} {raw}"
    if "有水印" in text:
        return "有水印"
    if "无水印" in text or "无印字" in text:
        return "无水印"
    return ""


def _is_light_board(spec: str, copper_spec: str) -> bool:
    normalized = copper_spec.replace(" ", "").upper()
    return "光板" in spec or normalized in {"0/0", "00", "0-0"}


def _number_or_text(value: Any) -> Any:
    number, ok = _number(value)
    return _excel_number(number) if ok else _text(value)


def _excel_number(value: float) -> int | float:
    return int(value) if float(value).is_integer() else round(float(value), 6)


def _format_compact_number(value: float) -> str:
    return f"{value:.12g}"


def _inventory_sort_key(row: InventoryRow) -> tuple[Any, ...]:
    return (
        (row.spec_glue or row.glue).upper(),
        _copper_rank(row.copper_spec),
        row.thickness_mm if row.thickness_mm is not None else float("inf"),
        0 if row.thickness_type == "芯厚" else 1 if row.thickness_type == "总厚" else 2,
        row.spec,
        row.plant,
    )


def _copper_rank(value: str) -> tuple[float, float, str]:
    text = value.strip().upper().replace("μ", "U")
    parts = re.split(r"[/\\]", text, maxsplit=1)
    ranks = [_copper_side_rank(part) for part in parts]
    if len(ranks) == 1:
        ranks.append(ranks[0])
    return max(ranks), min(ranks), text


def _copper_side_rank(value: str) -> float:
    token = value.strip().upper()
    letter_map = {"W": 3, "E": 5, "F": 9, "J": 15, "H": 18, "I": 25, "T": 30, "K": 35}
    if token in letter_map:
        return float(letter_map[token])
    if "+" in token:
        return sum(_copper_side_rank(part) for part in token.split("+"))
    match = re.search(r"(\d+(?:\.\d+)?)\s*U?M", token)
    if match:
        return float(match.group(1))
    try:
        number = float(token)
        return number * 35 if number <= 10 else number
    except ValueError:
        return 9999.0


def build_inventory_workbook(
    rows: list[InventoryRow],
    grade: str,
    output_path: str | Path,
    *,
    workbook_mode: str = WAREHOUSE_MODE,
) -> Path:
    grade = grade.upper()
    if grade not in {"A", "B"}:
        raise ValueError("库存等级只能是 A 或 B")
    workbook_mode = normalize_inventory_mode(workbook_mode)
    profile = "plan_a" if workbook_mode == PLAN_A_MODE else "a" if grade == "A" else "b"
    headers = PLAN_A_HEADERS if profile == "plan_a" else A_HEADERS if profile == "a" else B_HEADERS

    workbook = Workbook()
    workbook.remove(workbook.active)
    _register_named_styles(workbook, profile)
    dashboard = workbook.create_sheet("导航仪表盘")

    grouped: dict[str, list[InventoryRow]] = defaultdict(list)
    for row in rows:
        grouped[row.spec_glue or row.glue or "未识别胶系"].append(row)
    glue_items = sorted(grouped.items(), key=lambda item: (-sum(row.quantity for row in item[1]), item[0]))

    used_sheet_names = {"导航仪表盘", "全部明细", "胶系厚度汇总", "光板明细", "异常待确认"}
    sheet_map: dict[str, str] = {}
    table_index = 1
    for glue, glue_rows in glue_items:
        sheet_name = _safe_sheet_name(glue, used_sheet_names)
        sheet_map[glue] = sheet_name
        sheet = workbook.create_sheet(sheet_name)
        _write_glue_sheet(sheet, glue, glue_rows, headers, profile, table_index)
        table_index += 1

    all_sheet = workbook.create_sheet("全部明细")
    _write_all_sheet(all_sheet, rows, headers, profile, table_index)
    table_index += 1

    summary_sheet = workbook.create_sheet("胶系厚度汇总")
    _write_thickness_summary(summary_sheet, rows, profile, table_index)
    table_index += 1

    light_rows = [row for row in rows if row.is_light == "是"]
    light_sheet = workbook.create_sheet("光板明细")
    _write_special_sheet(light_sheet, "光板明细", light_rows, headers, profile, table_index)
    table_index += 1

    exception_rows = [row for row in rows if row.parse_status != "正常"]
    exception_sheet = workbook.create_sheet("异常待确认")
    _write_special_sheet(exception_sheet, "异常待确认", exception_rows, headers, profile, table_index)

    _write_dashboard(dashboard, rows, glue_items, sheet_map, profile, len(light_rows), len(exception_rows))
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


def _register_named_styles(workbook: Workbook, profile: str) -> None:
    primary = NAVY if profile != "b" else TEAL
    thin = Side(style="thin", color=BORDER_BLUE)
    styles = [
        NamedStyle(
            name="inv_header",
            font=Font(name="等线", size=10, bold=True, color=WHITE),
            fill=PatternFill("solid", fgColor=primary),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=Border(bottom=thin),
        ),
        NamedStyle(
            name="inv_body",
            font=Font(name="等线", size=10, color="000000"),
            alignment=Alignment(vertical="center"),
        ),
        NamedStyle(
            name="inv_text",
            font=Font(name="等线", size=10, color="000000"),
            alignment=Alignment(vertical="center"),
            number_format="@",
        ),
        NamedStyle(
            name="inv_number",
            font=Font(name="等线", size=10, color="000000"),
            alignment=Alignment(horizontal="right", vertical="center"),
            number_format="#,##0.##",
        ),
        NamedStyle(
            name="inv_link",
            font=Font(name="等线", size=10, bold=True, color=LINK_BLUE, underline="single"),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
    ]
    for style in styles:
        workbook.add_named_style(style)


def _write_glue_sheet(
    sheet,
    glue: str,
    rows: list[InventoryRow],
    headers: list[str],
    profile: str,
    table_index: int,
) -> None:
    last_col = len(headers)
    primary = NAVY if profile != "b" else TEAL
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D5" if profile != "b" else "B5"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    sheet["A1"] = '=HYPERLINK("#\'导航仪表盘\'!A1","← 返回仪表盘")'
    sheet["A1"].style = "inv_link"
    sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=last_col)
    sheet.cell(1, 3, f"{glue} 库存明细")
    sheet.cell(1, 3).font = Font(name="等线", size=16, bold=True, color=WHITE)
    sheet.cell(1, 3).fill = PatternFill("solid", fgColor=primary)
    sheet.cell(1, 3).alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, last_col + 1):
        sheet.cell(1, col).fill = PatternFill("solid", fgColor=primary)
    sheet.row_dimensions[1].height = 28

    summary_values = [
        ("明细行数", len(rows)),
        ("库存数量", sum(row.quantity for row in rows)),
        ("折合大板", sum(row.folded_large_quantity for row in rows)),
        ("厚度档数", len({row.thickness_mm for row in rows if row.thickness_mm is not None})),
        ("光板行数", sum(row.is_light == "是" for row in rows)),
        ("分类依据", "类别/胶系"),
    ]
    cursor = 1
    for label, value in summary_values:
        if cursor + 1 > last_col:
            break
        sheet.cell(2, cursor, label)
        sheet.cell(2, cursor).font = Font(name="等线", size=10, bold=True, color=primary)
        sheet.cell(2, cursor + 1, _excel_number(value) if isinstance(value, (int, float)) else value)
        sheet.cell(2, cursor + 1).number_format = "#,##0.##"
        cursor += 2

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    sheet.cell(3, 1, "本页按铜厚实际厚度、板厚mm、芯厚/总厚顺序排列；上海、江西保持分行显示。")
    sheet.cell(3, 1).font = Font(name="等线", size=10, color=MUTED)
    sheet.cell(3, 1).alignment = Alignment(vertical="center")

    _write_header(sheet, 4, headers)
    for output_row in rows:
        sheet.append(_row_values(output_row, profile))
    _style_data_rows(sheet, 5, sheet.max_row, headers)
    _add_table(sheet, f"GlueDetailTable{table_index}", 4, sheet.max_row, last_col)
    _set_detail_widths(sheet, profile)


def _write_all_sheet(sheet, rows: list[InventoryRow], headers: list[str], profile: str, table_index: int) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D2" if profile != "b" else "B2"
    _write_header(sheet, 1, headers)
    for output_row in rows:
        sheet.append(_row_values(output_row, profile))
    _style_data_rows(sheet, 2, sheet.max_row, headers)
    _add_table(sheet, f"InventoryAllTable{table_index}", 1, sheet.max_row, len(headers))
    _set_detail_widths(sheet, profile)


def _write_special_sheet(
    sheet,
    title: str,
    rows: list[InventoryRow],
    headers: list[str],
    profile: str,
    table_index: int,
) -> None:
    last_col = len(headers)
    primary = NAVY if profile != "b" else TEAL
    sheet.sheet_view.showGridLines = False
    header_row = 4 if profile != "b" else 2
    sheet.freeze_panes = "D5" if profile != "b" else "B3"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    sheet.cell(1, 1, title)
    sheet.cell(1, 1).font = Font(name="等线", size=14, bold=True, color=WHITE)
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor=primary)
    sheet.cell(1, 1).alignment = Alignment(vertical="center")
    if profile != "b":
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        sheet.cell(2, 1, f"共 {len(rows)} 行；可通过表头筛选继续复核。")
        sheet.cell(2, 1).font = Font(name="等线", size=10, color=MUTED)
    _write_header(sheet, header_row, headers)
    for output_row in rows:
        sheet.append(_row_values(output_row, profile))
    _style_data_rows(sheet, header_row + 1, sheet.max_row, headers)
    _add_table(sheet, f"InventorySpecialTable{table_index}", header_row, sheet.max_row, last_col)
    _set_detail_widths(sheet, profile)


def _write_thickness_summary(sheet, rows: list[InventoryRow], profile: str, table_index: int) -> None:
    primary = NAVY if profile != "b" else TEAL
    headers = ["胶系", "厚度类型", "厚度mm", "明细行数", "数量合计", "折合大板合计"]
    grouped: dict[tuple[str, str, float | None], list[InventoryRow]] = defaultdict(list)
    for row in rows:
        grouped[(row.spec_glue or row.glue, row.thickness_type, row.thickness_mm)].append(row)

    summary_rows = []
    for (glue, thickness_type, thickness_mm), members in grouped.items():
        summary_rows.append(
            [
                glue,
                thickness_type,
                _excel_number(thickness_mm) if thickness_mm is not None else "",
                len(members),
                _excel_number(sum(row.quantity for row in members)),
                _excel_number(sum(row.folded_large_quantity for row in members)),
            ]
        )
    summary_rows.sort(key=lambda values: (str(values[0]).upper(), str(values[1]), values[2] if values[2] != "" else float("inf")))

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D4" if profile != "b" else "A4"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = f"{'计划A' if profile == 'plan_a' else 'A' if profile == 'a' else 'B'}级胶系厚度汇总"
    sheet["A1"].font = Font(name="等线", size=14, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=primary)
    sheet["A1"].alignment = Alignment(vertical="center")
    _write_header(sheet, 3, headers)
    for values in summary_rows:
        sheet.append(values)
    _style_data_rows(sheet, 4, sheet.max_row, headers)
    _add_table(sheet, f"InventorySummaryTable{table_index}", 3, sheet.max_row, len(headers))
    for letter, width in {"A": 20, "B": 12, "C": 12, "D": 12, "E": 14, "F": 18}.items():
        sheet.column_dimensions[letter].width = width


def _write_dashboard(
    sheet,
    rows: list[InventoryRow],
    glue_items: list[tuple[str, list[InventoryRow]]],
    sheet_map: dict[str, str],
    profile: str,
    light_count: int,
    exception_count: int,
) -> None:
    primary = DARK_NAVY if profile != "b" else NAVY
    accent = TEAL
    grade_label = "计划A级" if profile == "plan_a" else "A级" if profile == "a" else "B级"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85
    sheet.merge_cells("A1:M2")
    sheet["A1"] = (
        "计划A级库存 胶系分类导航"
        if profile == "plan_a"
        else f"无订单库存 {grade_label}分类导航"
    )
    sheet["A1"].font = Font(name="等线", size=20, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=primary)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells("A3:M3")
    sheet["A3"] = "按胶系进入独立库存明细；分类页按铜厚实际厚度、板厚mm、芯厚/总厚顺序排列。"
    sheet["A3"].font = Font(name="等线", size=10, color=MUTED)

    kpis = [
        ("库存明细", len(rows)),
        ("胶系数量", len(glue_items)),
        ("光板行数", light_count),
        ("待确认", exception_count),
    ]
    blocks = [(1, 3), (4, 6), (7, 9), (10, 13)]
    for (label, value), (start_col, end_col) in zip(kpis, blocks):
        sheet.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
        sheet.merge_cells(start_row=6, start_column=start_col, end_row=7, end_column=end_col)
        label_cell = sheet.cell(5, start_col, label)
        label_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        label_cell.font = Font(name="等线", size=10, bold=True, color=NAVY)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell = sheet.cell(6, start_col, _excel_number(value))
        value_cell.font = Font(name="等线", size=18 if profile == "b" else 20, bold=True, color=primary)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = "#,##0"

    default_glue = glue_items[0][0] if glue_items else ""
    sheet.merge_cells("A9:B10")
    sheet["A9"] = "选择胶系"
    sheet["A9"].fill = PatternFill("solid", fgColor=accent)
    sheet["A9"].font = Font(name="等线", size=10, bold=True, color=WHITE)
    sheet["A9"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells("C9:F10")
    sheet["C9"] = default_glue
    sheet["C9"].font = Font(name="等线", size=10, bold=True, color=TEXT_DARK)
    sheet["C9"].alignment = Alignment(vertical="center")
    sheet.merge_cells("G9:I10")
    sheet["G9"] = '=IF(C9="","请选择胶系",HYPERLINK("#\'"&C9&"\'!A1","打开 "&C9&" →"))'
    sheet["G9"].fill = PatternFill("solid", fgColor=LINK_BLUE)
    sheet["G9"].font = Font(name="等线", size=11, bold=True, color=WHITE)
    sheet["G9"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells("J9:M10")
    sheet["J9"] = "普通 .xlsx\n无需启用宏"
    sheet["J9"].fill = PatternFill("solid", fgColor=LIGHT_TEAL)
    sheet["J9"].font = Font(name="等线", size=10, color=TEAL)
    sheet["J9"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index, (glue, _) in enumerate(glue_items, start=2):
        sheet.cell(index, 16, glue)
        sheet.cell(index, 17, sheet_map[glue])
    if glue_items:
        validation = DataValidation(type="list", formula1=f"=$P$2:$P${len(glue_items)+1}", allow_blank=False)
        sheet.add_data_validation(validation)
        validation.add(sheet["C9"])
    sheet.column_dimensions["P"].hidden = True
    sheet.column_dimensions["Q"].hidden = True
    if glue_items:
        sheet["G9"] = (
            f'=IF(C9="","请选择胶系",HYPERLINK("#\'"&'
            f'VLOOKUP(C9,$P$2:$Q${len(glue_items)+1},2,FALSE)&"\'!A1","打开 "&C9&" →"))'
        )

    sheet.merge_cells("A12:M12")
    sheet["A12"] = "胶系快捷导航"
    sheet["A12"].fill = PatternFill("solid", fgColor=primary)
    sheet["A12"].font = Font(name="等线", size=11, bold=True, color=WHITE)
    nav_start = 14
    card_blocks = [(1, 3), (4, 6), (7, 9), (10, 13)]
    for index, (glue, glue_rows) in enumerate(glue_items):
        block = card_blocks[index % 4]
        start_row = nav_start + (index // 4) * 2
        sheet.merge_cells(start_row=start_row, start_column=block[0], end_row=start_row + 1, end_column=block[1])
        target_sheet = sheet_map[glue].replace("'", "''")
        formula = f'=HYPERLINK("#\'{target_sheet}\'!A1","{glue}"&CHAR(10)&"{len(glue_rows)} 行")'
        cell = sheet.cell(start_row, block[0], formula)
        cell.font = Font(name="等线", size=10, bold=True, color=TEAL)
        cell.fill = PatternFill("solid", fgColor=LIGHT_TEAL if index < 12 else WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        _set_merged_border(sheet, start_row, block[0], start_row + 1, block[1])

    nav_rows = max(2, math.ceil(max(len(glue_items), 1) / 4) * 2)
    section_row = nav_start + nav_rows + 1
    sheet.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=9)
    sheet.cell(section_row, 1, "库存规模概览")
    sheet.cell(section_row, 1).fill = PatternFill("solid", fgColor=primary)
    sheet.cell(section_row, 1).font = Font(name="等线", size=11, bold=True, color=WHITE)
    sheet.merge_cells(start_row=section_row, start_column=10, end_row=section_row, end_column=13)
    sheet.cell(section_row, 10, "常用入口")
    sheet.cell(section_row, 10).fill = PatternFill("solid", fgColor=accent)
    sheet.cell(section_row, 10).font = Font(name="等线", size=11, bold=True, color=WHITE)
    sheet.cell(section_row, 10).alignment = Alignment(horizontal="center")

    top_items = sorted(glue_items, key=lambda item: sum(row.quantity for row in item[1]), reverse=True)[:10]
    sheet.cell(1, 14, "胶系")
    sheet.cell(1, 15, "数量")
    for index, (glue, glue_rows) in enumerate(top_items, start=2):
        sheet.cell(index, 14, glue)
        sheet.cell(index, 15, _excel_number(sum(row.quantity for row in glue_rows)))
    sheet.column_dimensions["N"].hidden = True
    sheet.column_dimensions["O"].hidden = True
    if top_items:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "库存数量最多的 10 个胶系"
        chart.y_axis.title = "数量"
        chart.x_axis.title = "胶系"
        chart.height = 7.6
        chart.width = 14.2
        data = Reference(sheet, min_col=15, min_row=1, max_row=len(top_items) + 1)
        categories = Reference(sheet, min_col=14, min_row=2, max_row=len(top_items) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend = None
        chart.visible_cells_only = False
        sheet.add_chart(chart, f"A{section_row + 2}")

    entry_names = ["全部明细", "光板明细", "胶系厚度汇总", "异常待确认"]
    for index, name in enumerate(entry_names):
        start_row = section_row + 2 + index * 4
        sheet.merge_cells(start_row=start_row, start_column=10, end_row=start_row + 2, end_column=13)
        cell = sheet.cell(start_row, 10, f'=HYPERLINK("#\'{name}\'!A1","{name} →")')
        cell.font = Font(name="等线", size=11, bold=True, color=LINK_BLUE)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if index % 2 == 0 else WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        _set_merged_border(sheet, start_row, 10, start_row + 2, 13)

    for letter, width in {
        "A": 13,
        "B": 12,
        "C": 13,
        "D": 13,
        "E": 12,
        "F": 12,
        "G": 13,
        "H": 12,
        "I": 12,
        "J": 13,
        "K": 12,
        "L": 12,
        "M": 13,
    }.items():
        sheet.column_dimensions[letter].width = width
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 24
    sheet.print_area = f"A1:M{section_row + 18}"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _row_values(row: InventoryRow, profile: str) -> list[Any]:
    common = [
        row.spec,
        row.plant,
        row.glue,
        row.thickness_raw,
        row.copper_raw,
        _excel_number(row.quantity),
        _excel_number(row.folded_large_quantity),
        row.is_light,
        row.spec_glue,
        _excel_number(row.thickness_mm) if row.thickness_mm is not None else "",
        row.thickness_type,
        row.copper_spec,
        _excel_number(row.length) if row.length is not None else "",
        _excel_number(row.width) if row.width is not None else "",
        row.size_text,
        row.copper_type,
        row.spec_grade,
        row.color,
        row.watermark,
        row.special_note,
        row.parse_status,
    ]
    if profile == "b":
        return common
    if profile == "plan_a":
        return [
            row.product_name,
            row.product_no,
            row.spec,
            row.layout,
            row.plant,
            row.glue,
            row.thickness_raw,
            row.copper_raw,
            row.unit_weight,
            _excel_number(row.quantity),
            _excel_number(row.folded_large_quantity),
            row.is_light,
            row.spec_glue,
            _excel_number(row.thickness_mm) if row.thickness_mm is not None else "",
            row.thickness_type,
            row.copper_spec,
            _excel_number(row.length) if row.length is not None else "",
            _excel_number(row.width) if row.width is not None else "",
            row.size_text,
            row.copper_type,
            row.spec_grade,
            row.color,
            row.watermark,
            row.special_note,
            row.parse_status,
        ]
    return [
        row.product_name,
        row.product_no,
        row.spec,
        row.plant,
        row.glue,
        row.thickness_raw,
        row.copper_raw,
        row.unit_weight,
        _excel_number(row.quantity),
        _excel_number(row.folded_large_quantity),
        row.is_light,
        row.spec_glue,
        _excel_number(row.thickness_mm) if row.thickness_mm is not None else "",
        row.thickness_type,
        row.copper_spec,
        _excel_number(row.length) if row.length is not None else "",
        _excel_number(row.width) if row.width is not None else "",
        row.size_text,
        row.copper_type,
        row.spec_grade,
        row.color,
        row.watermark,
        row.special_note,
        row.parse_status,
    ]


def _write_header(sheet, row_index: int, headers: list[str]) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row_index, column, header)
        cell.style = "inv_header"
    sheet.row_dimensions[row_index].height = 28


def _style_data_rows(sheet, start_row: int, end_row: int, headers: list[str]) -> None:
    if end_row < start_row:
        return
    text_headers = {"品名", "品号", "规格", "排版结构", "上海_江西", "胶系", "铜箔", "铜厚规格", "尺寸", "解析状态"}
    number_headers = {
        "厚度",
        "单重",
        "数量",
        "折合大板",
        "厚度mm",
        "尺寸长",
        "尺寸宽",
        "明细行数",
        "数量合计",
        "折合大板合计",
    }
    for column, header in enumerate(headers, start=1):
        style_name = "inv_text" if header in text_headers else "inv_number" if header in number_headers else "inv_body"
        for row in range(start_row, end_row + 1):
            sheet.cell(row, column).style = style_name
    if "解析状态" in headers:
        status_column = _column_letter(headers.index("解析状态") + 1)
        sheet.conditional_formatting.add(
            f"{status_column}{start_row}:{status_column}{end_row}",
            FormulaRule(
                formula=[f'LEFT({status_column}{start_row},3)="待确认"'],
                fill=PatternFill("solid", fgColor=ERROR_FILL),
            ),
        )


def _add_table(sheet, name: str, header_row: int, max_row: int, max_col: int) -> None:
    # Excel 不接受只有表头、没有数据行的原生 Table。
    if max_row <= header_row:
        return
    reference = f"A{header_row}:{_column_letter(max_col)}{max_row}"
    table = Table(displayName=re.sub(r"[^A-Za-z0-9_]", "", name), ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _set_detail_widths(sheet, profile: str) -> None:
    widths_a = [24, 15, 58, 12, 16, 10, 12, 10, 12, 16, 10, 15, 11, 12, 18, 11, 11, 18, 14, 11, 11, 12, 30, 38]
    widths_plan_a = [24, 15, 58, 18, 12, 16, 10, 12, 10, 12, 16, 10, 15, 11, 12, 18, 11, 11, 18, 14, 11, 11, 12, 30, 38]
    widths_b = [58, 12, 16, 10, 12, 12, 16, 10, 15, 11, 12, 18, 11, 11, 18, 14, 11, 11, 12, 30, 38]
    widths = widths_plan_a if profile == "plan_a" else widths_a if profile == "a" else widths_b
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[_column_letter(index)].width = width


def _safe_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", " ", value).strip() or "未识别胶系"
    base = re.sub(r"\s+", " ", base)[:31]
    candidate = base
    suffix = 2
    while candidate in used:
        tail = f" ({suffix})"
        candidate = f"{base[:31-len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def _column_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _set_merged_border(sheet, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    thin = Side(style="thin", color=BORDER_BLUE)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            cell.border = Border(
                left=thin if col == min_col else Side(style=None),
                right=thin if col == max_col else Side(style=None),
                top=thin if row == min_row else Side(style=None),
                bottom=thin if row == max_row else Side(style=None),
            )
