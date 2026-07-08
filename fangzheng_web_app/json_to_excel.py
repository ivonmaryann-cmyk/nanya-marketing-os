from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DETAIL_HEADERS = ["序号", "物料编码", "名称规格", "单位", "数量", "单价", "金额", "到货日期", "备注"]
REVIEW_HEADERS = ["文件名", "页码", "来源", "原始文本", "疑似字段", "失败原因", "建议处理"]

HEADER_LABELS = {
    "source_file": "文件名",
    "customer": "客户/采购方",
    "supplier": "供应商",
    "order_number": "订单号",
    "contract_number": "合同编号",
    "order_date": "日期",
    "currency": "币种",
    "tax_rate": "税率",
    "contact": "联系方式",
    "parser_mode": "识别方式",
}

HEADER_ALIASES = {
    "seq": ["序号", "项次", "行号", "item no", "item", "no", "no."],
    "material_code": ["物料编码", "物料代码", "料号", "品号", "合约编号", "物料号", "料件编号", "原料编码", "part no", "part no.", "part"],
    "spec": ["物料名称规格", "名称规格", "物料说明", "物料名称", "原料名称", "规格型号", "规格", "品名", "描述", "description"],
    "unit": ["单位", "计量单位", "unit"],
    "qty": ["数量", "采购量", "订购数量", "订单数量", "quantity", "qty"],
    "price": ["单价", "单价RMB", "含税单价", "价格", "unit price", "price"],
    "amount": ["金额", "价税合计", "合计金额", "总价", "amount"],
    "arrival_date": ["到货日期", "交货日期", "预交货日", "交期", "delivery date", "delivery"],
    "remark": ["备注", "说明", "comments", "remark"],
}

SECTION_KEYWORDS = {
    "payment_info": ["付款", "结算", "月结", "账期", "发票", "PaymentTerms", "Payment Terms"],
    "shipping_info": ["收货", "送货", "地址", "联系人", "电话", "传真", "Shipto", "VendorAddress", "Telephone"],
    "terms": ["条款", "交货", "验收", "质量", "合同", "违约", "承担", "通知", "包装"],
    "notes": ["备注", "说明", "注意", "合计", "总计", "总金额"],
}

DATE_PATTERN = r"20\d{2}(?:[年./-]\s*\d{1,2}[月./-]\s*\d{1,2}日?|[-/]\d{4})"
MONEY_PATTERN = r"-?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?"


def _clean_text(value: Any) -> str:
    text = str(value or "").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _compact(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _normalize_numeric_text(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r"(?<=\d),\s+(?=\d{3}(?:\D|$))", ",", text)
    text = re.sub(r"(?<=\d)\.\s+(?=\d)", ".", text)
    return text


def _first_number(text: str) -> str:
    match = re.search(MONEY_PATTERN, _normalize_numeric_text(text))
    return match.group(0).strip(",") if match else ""


def _first_date(text: str) -> str:
    match = re.search(DATE_PATTERN, text or "")
    if not match:
        return ""
    raw = match.group(0).strip()
    short_match = re.match(r"(20\d{2})[-/](\d{2})(\d{2})$", raw)
    if short_match:
        return f"{short_match.group(1)}-{short_match.group(2)}-{short_match.group(3)}"
    raw = raw.replace("年", "-").replace("月", "-").replace("日", "")
    raw = raw.replace("/", "-").replace(".", "-")
    parts = [part.strip() for part in raw.split("-") if part.strip()]
    if len(parts) >= 3:
        return f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
    return raw


def _to_decimal(text: Any) -> Decimal | None:
    value = _normalize_numeric_text(text)
    if not value:
        return None
    match = re.search(MONEY_PATTERN, value)
    if not match:
        return None
    try:
        return Decimal(match.group(0).replace(",", ""))
    except InvalidOperation:
        return None


def _money_text_ok(text: str) -> bool:
    value = _normalize_numeric_text(text)
    if not value:
        return False
    return bool(re.fullmatch(MONEY_PATTERN, value.replace(" ", "")))


def _dedupe(items: list[str], *, limit: int = 20) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for item in items:
        text = _clean_text(item)
        key = _compact(text)
        if not text or key in seen:
            continue
        seen.add(key)
        result.append(text)
        if len(result) >= limit:
            break
    return result


def _grid_for_table(cells: list[dict[str, Any]]) -> dict[int, dict[int, dict[str, Any]]]:
    grid: dict[int, dict[int, dict[str, Any]]] = {}
    for cell in cells:
        try:
            row_index = int(cell.get("row_index", 0))
            column_index = int(cell.get("column_index", 0))
        except (TypeError, ValueError):
            continue
        grid.setdefault(row_index, {})[column_index] = cell
    return grid


def _row_text(row_cells: dict[int, dict[str, Any]]) -> str:
    return " ".join(_clean_text(row_cells[col].get("text")) for col in sorted(row_cells) if _clean_text(row_cells[col].get("text")))


def _header_map(header_values: dict[int, str]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for column_index, value in header_values.items():
            normalized = _compact(value)
            if not normalized:
                continue
            if any(_compact(alias) in normalized for alias in aliases):
                mapped[field] = column_index
                break
    return mapped


def _cell_confidence(row_cells: dict[int, dict[str, Any]]) -> str:
    values = [
        float(cell.get("confidence") or 0)
        for cell in row_cells.values()
        if cell.get("method") != "manual_empty" and (cell.get("text") or "").strip()
    ]
    return str(round(sum(values) / len(values), 4)) if values else ""


def _cell_methods(row_cells: dict[int, dict[str, Any]]) -> str:
    return ",".join(sorted({str(cell.get("method") or "") for cell in row_cells.values() if cell.get("method")}))


def _is_total_row(text: str) -> bool:
    compact = _compact(text)
    return any(keyword in compact for keyword in ["合计", "总计", "小计", "subtotal", "total", "总金额"])


def _looks_like_detail_line(text: str) -> bool:
    compact = _compact(text)
    has_code = bool(re.search(r"[a-z]{1,5}\d{5,}|[a-z0-9]{10,}", compact, flags=re.I))
    has_amount = bool(re.search(r"\d[\d,]*\.\d{2,}", compact))
    has_date = bool(re.search(DATE_PATTERN, compact))
    return has_code and (has_amount or has_date)


def _looks_like_table_header_line(text: str) -> bool:
    compact = _compact(text)
    has_material = any(_compact(alias) in compact for alias in HEADER_ALIASES["material_code"] + HEADER_ALIASES["spec"])
    has_amount = any(_compact(alias) in compact for alias in HEADER_ALIASES["qty"] + HEADER_ALIASES["amount"])
    return has_material and has_amount


def _looks_like_summary_header_line(text: str) -> bool:
    compact = _compact(text)
    return any(keyword in compact for keyword in ["大写金额", "合计金额", "总金额", "税前总金额", "含税总金额"])


def _line_items(document: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for page in document.get("pages", []):
        page_index = int(page.get("page_index", 0))
        text = _clean_text(page.get("text"))
        if text:
            for line in text.splitlines():
                line = line.strip()
                if line:
                    items.append({"text": line, "page_index": page_index, "source": "page_text"})

        words = [word for word in page.get("words") or [] if _clean_text(word.get("text")) and word.get("bbox")]
        for line in _group_words_into_lines(words):
            items.append({"text": line, "page_index": page_index, "source": "word_line"})

        for table in page.get("tables", []):
            grid = _grid_for_table(table.get("cells") or [])
            for row_index in sorted(grid):
                text = _row_text(grid[row_index])
                if text:
                    items.append({"text": text, "page_index": page_index, "source": f"table_row:{row_index}"})

    for line in document.get("docling_lines") or []:
        text = _clean_text(line)
        if text:
            items.append({"text": text, "page_index": 0, "source": "docling_markdown"})

    for table in document.get("docling_classified_tables") or []:
        if table.get("table_type") == "detail_table":
            continue
        for row in table.get("rows") or []:
            text = " ".join(_clean_text(value) for value in row if _clean_text(value))
            if text:
                items.append({"text": text, "page_index": 0, "source": f"docling_{table.get('table_type') or 'unknown'}"})

    markdown = _clean_text(document.get("markdown"))
    if markdown and not document.get("docling_lines"):
        for line in markdown.splitlines():
            line = line.strip(" |#*")
            if line:
                items.append({"text": line, "page_index": 0, "source": "markdown"})
    return _dedupe_line_items(items)


def _dedupe_line_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[tuple[int, str]] = set()
    for item in items:
        text = _clean_text(item.get("text"))
        key = (int(item.get("page_index", 0)), _compact(text))
        if not text or key in seen:
            continue
        seen.add(key)
        result.append({**item, "text": text})
    return result


def _group_words_into_lines(words: list[dict[str, Any]]) -> list[str]:
    sortable = []
    heights: list[float] = []
    for word in words:
        try:
            x0, y0, _x1, y1 = [float(value) for value in word["bbox"]]
        except (TypeError, ValueError):
            continue
        sortable.append((y0, x0, y1, _clean_text(word.get("text"))))
        heights.append(max(y1 - y0, 1.0))
    if not sortable:
        return []

    tolerance = max(6.0, (median(heights) if heights else 10.0) * 0.7)
    rows: list[dict[str, Any]] = []
    for y0, x0, _y1, text in sorted(sortable):
        target = None
        for row in rows:
            if abs(y0 - row["y"]) <= tolerance:
                target = row
                break
        if target is None:
            rows.append({"y": y0, "items": [(x0, text)]})
        else:
            target["items"].append((x0, text))
            target["y"] = (target["y"] + y0) / 2

    result: list[str] = []
    for row in rows:
        parts = [text for _x, text in sorted(row["items"]) if text]
        if parts:
            result.append(" ".join(parts))
    return result


def _extract_after_label(text: str, labels: list[str]) -> str:
    for label in labels:
        pattern = rf"{re.escape(label)}\s*(?:\([^)）]*\)|（[^)）]*）)?\s*[:：]\s*(.+?)(?:\s{{2,}}|\n|$)"
        match = re.search(pattern, text, flags=re.I)
        if match:
            value = match.group(1).strip()
            return re.split(
                r"\s+(?:供应商名称|供应商|Vendor|供应商代码|VendorCode|订购日期|P\.ODate|交易条件|付款条件|联系人|币种|税种|税率)\b",
                value,
                flags=re.I,
            )[0].strip(" ：:")
    return ""


def _extract_order_number(full_text: str) -> str:
    patterns = [
        r"(?:订单号|采购单号|采购订单号|订单编号)\s*(?:\([^)）]*\)|（[^)）]*）)?\s*[:：]\s*([A-Za-z0-9][A-Za-z0-9_-]{4,})",
        r"(?:P\.?\s*O\.?\s*NO|PONO)\s*(?:\([^)）]*\)|（[^)）]*）)?\s*[:：]\s*([A-Za-z0-9][A-Za-z0-9_-]{4,})",
        r"\b(GA\d{3,}-\d{6,})\b",
        r"\b(P\d{6,})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, full_text, flags=re.I)
        if match and re.search(r"\d{5,}", match.group(1)):
            return match.group(1)
    return ""


def _extract_contract_number(full_text: str) -> str:
    patterns = [
        r"(?:合同编号|合同号|Contract\s*No\.?)\s*[:：]\s*([A-Za-z0-9][A-Za-z0-9_-]{4,})",
        r"\b(TL\d{6,})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, full_text, flags=re.I)
        if match:
            return match.group(1).strip()
    return ""


def _extract_tax_rate(lines: list[str]) -> str:
    for line in lines:
        if not re.search(r"税率|税种|税码|TaxRate|Tax Rate|VAT", line, flags=re.I):
            continue
        vat_match = re.search(r"VAT\s*([0-9]{1,2})", line, flags=re.I)
        if vat_match:
            return f"{vat_match.group(1)}%"
        percent_match = re.search(r"(?:税率|税种|税码|TaxRate|Tax Rate)[^0-9%]{0,20}(\d+(?:\.\d+)?%)", line, flags=re.I)
        if percent_match:
            return percent_match.group(1)
    return ""


def _extract_currency(lines: list[str], full_text: str) -> str:
    for line in lines:
        match = re.search(r"(?:币种|币别|Currency)\s*(?:\([^)）]*\)|（[^)）]*）)?\s*[:：]\s*([A-Za-z\u4e00-\u9fff]+)", line, flags=re.I)
        if match:
            value = match.group(1).strip()
            return "人民币" if value.upper() in {"RMB", "CNY"} else value
    if re.search(r"\bRMB\b|人民币|CNY|￥", full_text, flags=re.I):
        return "人民币"
    if re.search(r"\bUSD\b|美元|\$", full_text, flags=re.I):
        return "USD"
    return ""


def _extract_header_info(document: dict[str, Any], items: list[dict[str, Any]]) -> tuple[dict[str, str], list[dict[str, str]]]:
    lines = [item["text"] for item in items]
    full_text = "\n".join(lines)
    header = {
        "source_file": _clean_text(document.get("source_file")),
        "customer": "",
        "supplier": _clean_text(document.get("supplier")),
        "order_number": _clean_text(document.get("order_number")),
        "contract_number": "",
        "order_date": "",
        "currency": "",
        "tax_rate": "",
        "contact": "",
        "parser_mode": _clean_text(document.get("parser_mode")),
    }
    review_rows: list[dict[str, str]] = []

    if not header["order_number"] or not re.search(r"\d{5,}", header["order_number"]):
        header["order_number"] = _extract_order_number(full_text)
    header["contract_number"] = _extract_contract_number(full_text)
    if header["order_number"] and not re.search(r"\d{5,}", header["order_number"]):
        review_rows.append(_make_review_row(document, 0, "订单头", full_text[:500], f"订单号={header['order_number']}", "订单号不包含连续数字", "请人工核对订单号"))
        header["order_number"] = ""
    elif not header["order_number"] and not header["contract_number"]:
        review_rows.append(_make_review_row(document, 0, "订单头", full_text[:500], "订单号=", "未识别到可靠订单号", "请人工核对订单号"))

    if not header["supplier"]:
        header["supplier"] = _extract_after_label(full_text, ["供应商名称", "供应商", "Vendor"])

    customer = _extract_after_label(full_text, ["公司名称", "采购方", "买方", "客户"])
    if customer:
        header["customer"] = customer
    else:
        company_line = next((line for line in lines[:18] if "公司" in line and not any(key in line for key in ["供应商", "供方", "Vendor"])), "")
        header["customer"] = company_line[:80]

    date_line = next((line for line in lines if any(key in line for key in ["订单日期", "采购日期", "订购日期", "P.ODate", "日期"])), "")
    header["order_date"] = _first_date(date_line) or _first_date(full_text)
    header["tax_rate"] = _extract_tax_rate(lines)
    header["currency"] = _extract_currency(lines, full_text)

    contacts = [
        line
        for line in lines
        if len(line) <= 120
        and not re.search(r"PaymentTerms|付款方式|付款条件", line, flags=re.I)
        and re.search(r"(电话|传真|联系人|收货人|ContactPerson|Telephone)\s*(?:\([^)）]*\)|（[^)）]*）)?\s*[:：]", line, flags=re.I)
    ]
    header["contact"] = "；".join(_dedupe(contacts, limit=3))
    return header, review_rows


def _make_detail_row(
    *,
    seq: str,
    material_code: str,
    spec: str,
    unit: str,
    qty: str,
    price: str,
    amount: str,
    arrival_date: str,
    remark: str = "",
    source_file: str = "",
    page_index: int = 0,
    raw_text: str = "",
    method: str = "",
    confidence: str = "",
    table_group: str = "",
) -> dict[str, str]:
    return {
        "序号": _clean_text(seq),
        "物料编码": _clean_text(material_code),
        "名称规格": _clean_text(spec),
        "单位": _clean_text(unit),
        "数量": _first_number(qty) or _clean_text(qty),
        "单价": _first_number(price) or _clean_text(price),
        "金额": _first_number(amount) or _clean_text(amount),
        "到货日期": _first_date(arrival_date) or _clean_text(arrival_date),
        "备注": _clean_text(remark),
        "_source_file": source_file,
        "_page_index": str(page_index),
        "_raw_text": _clean_text(raw_text),
        "_method": method,
        "_confidence": confidence,
        "_table_group": _clean_text(table_group),
    }


def _collect_following_text(items: list[dict[str, Any]], start_index: int, *, template_id: str) -> str:
    parts: list[str] = []
    start_page = int(items[start_index].get("page_index", 0))
    for item in items[start_index + 1 : start_index + 6]:
        if int(item.get("page_index", 0)) != start_page:
            break
        text = item["text"]
        if _is_total_row(text) or _looks_like_table_header_line(text):
            break
        if _line_starts_detail(text, template_id):
            break
        if re.search(r"^(备注[:：]?)?$", text.strip()):
            continue
        if any(keyword in text for keyword in ["交易条件", "付款条件", "送货地址", "供应商编号", "公司名称", "采购员"]):
            break
        parts.append(text)
    return "\n".join(parts)


def _line_starts_detail(text: str, template_id: str) -> bool:
    if template_id == "jingwang_purchase_order":
        return bool(re.match(r"^\s*\d+(?:\.\d+)?\s+[A-Za-z0-9_-]{6,}\s+", text))
    if template_id == "talian_purchase_order":
        return _talian_detail_match(text) is not None
    return bool(re.match(r"^\s*\d{1,3}\s+[A-Za-z0-9_-]{8,}\s+\d[\d,]*(?:\.\d+)?\s*[\u4e00-\u9fffA-Za-z/]+\s+", text))


def _extract_chaoyue_text_rows(document: dict[str, Any], items: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    source_file = _clean_text(document.get("source_file"))
    pattern = re.compile(
        r"^\s*(?P<seq>\d{1,3})\s+"
        r"(?P<code>[A-Za-z0-9_-]{8,})\s+"
        r"(?P<qty>\d[\d,]*(?:\.\d+)?)\s*(?P<unit>[\u4e00-\u9fffA-Za-z/]+)\s+"
        r"(?P<price>\d[\d,]*(?:\.\d+)?)\s+"
        r"(?P<amount>\d[\d,]*(?:\.\d+)?)\s+"
        rf"(?P<date>{DATE_PATTERN})"
    )
    for index, item in enumerate(items):
        text = item["text"]
        match = pattern.search(text)
        if not match:
            continue
        spec = _collect_following_text(items, index, template_id="ganzhou_chaoyue_purchase_order")
        rows.append(
            _make_detail_row(
                seq=match.group("seq"),
                material_code=match.group("code"),
                spec=spec,
                unit=match.group("unit"),
                qty=match.group("qty"),
                price=match.group("price"),
                amount=match.group("amount"),
                arrival_date=match.group("date"),
                source_file=source_file,
                page_index=int(item.get("page_index", 0)),
                raw_text="\n".join(part for part in [text, spec] if part),
                method="template_text_line",
                table_group=f"第 {int(item.get('page_index', 0)) + 1} 页明细",
            )
        )
    return _dedupe_detail_rows(rows)


def _extract_jingwang_text_rows(document: dict[str, Any], items: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    source_file = _clean_text(document.get("source_file"))
    pattern = re.compile(
        r"^\s*(?P<seq>\d+(?:\.\d+)?)\s+"
        r"(?P<code>[A-Za-z0-9_-]{6,})\s+"
        r"(?P<body>.+?)\s+"
        r"(?P<qty>\d[\d,]*(?:\.\d+)?)\s+"
        r"(?P<unit>[\u4e00-\u9fffA-Za-z/]+)\s+"
        r"(?P<price>-?\d[\d,]*(?:\.\d+)?)\s+"
        r"(?P<amount>-?\d[\d,]*(?:\.\d+)?)\s+"
        rf"(?P<date>{DATE_PATTERN})"
    )
    for index, item in enumerate(items):
        text = item["text"]
        match = pattern.search(text)
        if not match:
            continue
        extra = _collect_following_text(items, index, template_id="jingwang_purchase_order")
        spec = "\n".join(part for part in [match.group("body"), extra] if part)
        rows.append(
            _make_detail_row(
                seq=match.group("seq"),
                material_code=match.group("code"),
                spec=spec,
                unit=match.group("unit"),
                qty=match.group("qty"),
                price=match.group("price"),
                amount=match.group("amount"),
                arrival_date=match.group("date"),
                source_file=source_file,
                page_index=int(item.get("page_index", 0)),
                raw_text="\n".join(part for part in [text, extra] if part),
                method="template_text_line",
                table_group=f"第 {int(item.get('page_index', 0)) + 1} 页明细",
            )
        )
    return _dedupe_detail_rows(rows)


def _talian_detail_match(text: str) -> re.Match[str] | None:
    pattern = re.compile(
        r"^\s*(?P<seq>\d{1,3})\s+"
        r"(?P<code>[A-Za-z0-9_-]{4,})\s+"
        r"(?P<qty>\d[\d,]*(?:\.\d+)?)\s*"
        r"(?P<unit>[\u4e00-\u9fffA-Za-z/]+)\s+"
        r"(?P<price>-?\d[\d,]*(?:\.\d+)?)\s+"
        r"(?P<tax>\d+(?:\.\d+)?)\s+"
        r"(?P<amount>-?\d[\d,]*(?:,\d{3})*(?:\.\d+)?)\s+"
        rf"(?P<date>{DATE_PATTERN})"
        r"(?:\s+(?P<remark>.*))?$"
    )
    return pattern.search(text)


def _is_talian_spec_line(text: str) -> bool:
    clean = _clean_text(text)
    if not clean:
        return False
    if _talian_detail_match(clean) or _is_total_row(clean) or _looks_like_table_header_line(clean):
        return False
    if any(keyword in clean for keyword in ["付款方式", "交货期", "环保要求", "产品质量", "验收标准", "买方", "卖方", "盖章"]):
        return False
    return bool(
        re.search(
            r"南亚|FR-?4|半固化片|NY\d|RC\d+%|RTF|HTE|HVLP|TG|Tg|mm|\b\d+/\d+\b|\(\s*[A-Za-z0-9]",
            clean,
            flags=re.I,
        )
    )


def _collect_talian_after_specs(items: list[dict[str, Any]], start_index: int) -> tuple[str, set[int]]:
    parts: list[str] = []
    consumed_indexes: set[int] = set()
    start_page = int(items[start_index].get("page_index", 0))
    for offset, item in enumerate(items[start_index + 1 : start_index + 5], start=start_index + 1):
        if int(item.get("page_index", 0)) != start_page:
            break
        text = item["text"]
        if _talian_detail_match(text) or _is_total_row(text) or _looks_like_table_header_line(text):
            break
        if any(keyword in text for keyword in ["付款方式", "交货期", "环保要求", "产品质量", "验收标准", "买方", "卖方"]):
            break
        if re.match(r"^\s*\(", text) or re.search(r"RC\d+%|\b\d+/\d+\b|\d{3,4}\s*(?:\*|x|X)", text, flags=re.I):
            parts.append(text)
            consumed_indexes.add(offset)
            continue
        break
    return "\n".join(parts), consumed_indexes


def _extract_talian_text_rows(document: dict[str, Any], items: list[dict[str, Any]]) -> list[dict[str, str]]:
    source_file = _clean_text(document.get("source_file"))
    ordered_items = [item for item in items if item.get("source") == "page_text"] or items
    rows: list[dict[str, str]] = []
    spec_buffer: list[str] = []
    consumed_indexes: set[int] = set()
    last_page = None

    for index, item in enumerate(ordered_items):
        if index in consumed_indexes:
            continue
        text = item["text"]
        page_index = int(item.get("page_index", 0))
        if last_page is not None and page_index != last_page:
            spec_buffer = []
        last_page = page_index

        match = _talian_detail_match(text)
        if match:
            after_spec, consumed_after = _collect_talian_after_specs(ordered_items, index)
            consumed_indexes.update(consumed_after)
            spec_parts = [*spec_buffer, after_spec]
            spec = "\n".join(part for part in spec_parts if part)
            remark = _clean_text(match.group("remark") or "")
            tax = _clean_text(match.group("tax") or "")
            if tax and tax not in {"13", "13.0", "13.00"}:
                remark = "；".join(part for part in [remark, f"税率：{tax}%"] if part)
            rows.append(
                _make_detail_row(
                    seq=match.group("seq"),
                    material_code=match.group("code"),
                    spec=spec,
                    unit=match.group("unit"),
                    qty=match.group("qty"),
                    price=match.group("price"),
                    amount=match.group("amount"),
                    arrival_date=match.group("date"),
                    remark=remark,
                    source_file=source_file,
                    page_index=page_index,
                    raw_text="\n".join(part for part in [*spec_buffer, text, after_spec] if part),
                    method="talian_text_line",
                    table_group=f"第 {page_index + 1} 页明细",
                )
            )
            spec_buffer = []
            continue

        if _is_talian_spec_line(text):
            spec_buffer.append(text)
            spec_buffer = spec_buffer[-4:]
        elif _looks_like_table_header_line(text) or _is_total_row(text) or any(keyword in text for keyword in ["合同编号", "采购订单", "付款方式", "交货期", "买方", "卖方"]):
            spec_buffer = []

    return _dedupe_detail_rows(rows)


def _extract_yihao_grid_rows(document: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    source_file = _clean_text(document.get("source_file"))
    for page in document.get("pages", []):
        page_index = int(page.get("page_index", 0))
        for table in page.get("tables", []):
            grid = _grid_for_table(table.get("cells") or [])
            for row_index in sorted(grid):
                row_cells = grid[row_index]
                seq = _first_number(_row_value(row_cells, 0))
                if not seq or len(seq) > 3 or _is_total_row(_row_text(row_cells)):
                    continue
                material_code = _clean_material_code(_row_value(row_cells, 1))
                spec = _row_value(row_cells, 2)
                qty = _row_value(row_cells, 4)
                amount = _row_value(row_cells, 6)
                arrival_date = _row_value(row_cells, 7)
                if not (material_code and spec and _first_number(qty)):
                    continue
                rows.append(
                    _make_detail_row(
                        seq=seq,
                        material_code=material_code,
                        spec=spec,
                        unit=_row_value(row_cells, 3),
                        qty=qty,
                        price=_row_value(row_cells, 5),
                        amount=amount,
                        arrival_date=arrival_date,
                        remark=_row_value(row_cells, 8),
                        source_file=source_file,
                        page_index=page_index,
                        raw_text=_row_text(row_cells),
                        method=_cell_methods(row_cells) or "template_image_grid",
                        confidence=_cell_confidence(row_cells),
                        table_group=f"第 {page_index + 1} 页 表 {int(table.get('table_index') or 0) + 1}",
                    )
                )
    return _dedupe_detail_rows(rows)


def _extract_generic_grid_rows(document: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    source_file = _clean_text(document.get("source_file"))
    for page in document.get("pages", []):
        page_index = int(page.get("page_index", 0))
        for table in page.get("tables", []):
            grid = _grid_for_table(table.get("cells") or [])
            if not grid:
                continue
            header_row_index = None
            header: dict[str, int] = {}
            for row_index in sorted(grid):
                row_text = _row_text(grid[row_index])
                if _is_total_row(row_text) or _looks_like_summary_header_line(row_text):
                    continue
                values = {col: _clean_text(cell.get("text")) for col, cell in grid[row_index].items()}
                mapped = _header_map(values)
                business_fields = {"material_code", "spec", "qty", "unit", "price", "amount", "arrival_date"}
                mapped_fields = business_fields.intersection(mapped)
                has_identity = bool({"material_code", "spec"}.intersection(mapped))
                has_values = bool({"qty", "unit"}.intersection(mapped)) and bool({"price", "amount", "arrival_date"}.intersection(mapped))
                if len(mapped_fields) >= 5 and has_identity and has_values:
                    header_row_index = row_index
                    header = mapped
                    break
            if header_row_index is None:
                continue
            row_number = 1
            for row_index in sorted(idx for idx in grid if idx > header_row_index):
                row_cells = grid[row_index]
                row_text = _row_text(row_cells)
                if not row_text or _is_total_row(row_text) or _looks_like_summary_header_line(row_text):
                    continue
                detail = _detail_row_from_map(
                    row_cells,
                    header,
                    row_number=row_number,
                    source_file=source_file,
                    page_index=page_index,
                    table_index=int(table.get("table_index") or 0),
                )
                if detail:
                    rows.append(detail)
                    row_number += 1
    return _dedupe_detail_rows(rows)


def _detail_row_from_map(
    row_cells: dict[int, dict[str, Any]],
    header: dict[str, int],
    *,
    row_number: int,
    source_file: str,
    page_index: int,
    table_index: int = 0,
) -> dict[str, str] | None:
    def value(field: str) -> str:
        return _row_value(row_cells, header.get(field))

    row = _make_detail_row(
        seq=value("seq") or str(row_number),
        material_code=value("material_code"),
        spec=value("spec"),
        unit=value("unit"),
        qty=value("qty"),
        price=value("price"),
        amount=value("amount"),
        arrival_date=value("arrival_date"),
        remark=value("remark"),
        source_file=source_file,
        page_index=page_index,
        raw_text=_row_text(row_cells),
        method=_cell_methods(row_cells) or "header_grid",
        confidence=_cell_confidence(row_cells),
        table_group=f"第 {page_index + 1} 页 表 {table_index + 1}",
    )
    if not any(row[field] for field in ["物料编码", "名称规格", "数量", "金额"]):
        return None
    return row


def _extract_generic_line_rows(document: dict[str, Any], items: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    source_file = _clean_text(document.get("source_file"))
    pattern = re.compile(
        r"^\s*(?:(?P<seq>\d+(?:\.\d+)?)\s+)?"
        r"(?P<code>[A-Za-z0-9_-]{6,})\s+"
        r"(?P<body>.+?)\s+"
        r"(?P<qty>\d[\d,]*(?:\.\d+)?)\s*(?P<unit>[\u4e00-\u9fffA-Za-z/]+)\s+"
        r"(?P<price>-?\d[\d,]*(?:\.\d+)?)\s+"
        r"(?:(?P<tax>\d+(?:\.\d+)?)\s+)?"
        r"(?P<amount>-?\d[\d,]*(?:\.\d+)?)\s+"
        rf"(?P<date>{DATE_PATTERN})"
    )
    for item in items:
        text = item["text"]
        if _is_total_row(text) or _looks_like_table_header_line(text):
            continue
        match = pattern.search(text)
        if not match:
            continue
        rows.append(
            _make_detail_row(
                seq=match.group("seq") or str(len(rows) + 1),
                material_code=match.group("code"),
                spec=match.group("body"),
                unit=match.group("unit"),
                qty=match.group("qty"),
                price=match.group("price"),
                amount=match.group("amount"),
                arrival_date=match.group("date"),
                source_file=source_file,
                page_index=int(item.get("page_index", 0)),
                raw_text=text,
                method="generic_text_line",
                table_group=f"第 {int(item.get('page_index', 0)) + 1} 页明细",
            )
        )
    return _dedupe_detail_rows(rows)


def _extract_detail_candidates(document: dict[str, Any], items: list[dict[str, Any]]) -> list[dict[str, str]]:
    template_id = document.get("template_id")
    if template_id == "ganzhou_chaoyue_purchase_order":
        rows = _extract_chaoyue_text_rows(document, items)
        if rows:
            return rows
    if template_id == "jingwang_purchase_order":
        rows = _extract_jingwang_text_rows(document, items)
        if rows:
            return rows
    if template_id == "talian_purchase_order":
        rows = _extract_talian_text_rows(document, items)
        if rows:
            return rows
    if template_id == "ganzhou_yihao_purchase_order":
        rows = _extract_yihao_grid_rows(document)
        if rows:
            return rows

    rows = _extract_generic_grid_rows(document)
    if rows:
        return rows
    return _extract_generic_line_rows(document, items)


def _row_value(row_cells: dict[int, dict[str, Any]], column_index: int | None) -> str:
    if column_index is None:
        return ""
    return _clean_text(row_cells.get(column_index, {}).get("text"))


def _clean_material_code(text: str) -> str:
    tokens = re.findall(r"[A-Za-z0-9_-]{4,}", text or "")
    if tokens:
        return " ".join(tokens)
    return _clean_text(text)


def _dedupe_detail_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    for row in rows:
        key = (
            _compact(row.get("序号")),
            _compact(row.get("物料编码")),
            _compact(row.get("数量")),
            _compact(row.get("金额")),
            _compact(row.get("到货日期")),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def _validate_detail_row(row: dict[str, str], document: dict[str, Any]) -> list[str]:
    reasons: list[str] = []
    file_type = str(document.get("file_type") or "")
    seq = _clean_text(row.get("序号"))
    material_code = _clean_text(row.get("物料编码"))
    unit = _clean_text(row.get("单位"))
    qty = _to_decimal(row.get("数量"))
    price = _to_decimal(row.get("单价"))
    amount = _to_decimal(row.get("金额"))
    date = _clean_text(row.get("到货日期"))

    if not re.fullmatch(r"\d{1,4}(?:\.\d{1,4})?", seq):
        reasons.append("序号不像正常行号，可能把物料编码放进了序号列")
    if not material_code or len(_compact(material_code)) < 4:
        reasons.append("物料编码缺失或过短")
    if qty is None:
        reasons.append("数量不是数字")
    if unit and re.fullmatch(MONEY_PATTERN, unit):
        reasons.append("单位列是数字，疑似列错位")
    if amount is None or not _money_text_ok(row.get("金额", "")):
        reasons.append("金额格式不可靠")
    if not date:
        reasons.append("到货/交货日期缺失或格式不可靠")

    if qty is not None and price is not None and amount is not None:
        expected = qty * price
        tolerance = max(abs(expected) * Decimal("0.02"), Decimal("0.05"))
        if abs(expected - amount) > tolerance:
            reasons.append(f"金额与数量×单价不一致，应约为 {expected:,.2f}")
        if file_type == "image" and amount != 0 and amount < Decimal("1000") and abs(expected) >= Decimal("1000"):
            reasons.append("图片 OCR 金额疑似截断")
    return reasons


def _make_review_row(document: dict[str, Any], page_index: int, source: str, raw_text: str, fields: str, reason: str, suggestion: str) -> dict[str, str]:
    return {
        "文件名": _clean_text(document.get("source_file")),
        "页码": str(int(page_index) + 1) if str(page_index).isdigit() else str(page_index or ""),
        "来源": source,
        "原始文本": _clean_text(raw_text),
        "疑似字段": _clean_text(fields),
        "失败原因": _clean_text(reason),
        "建议处理": _clean_text(suggestion),
    }


def _review_from_detail(document: dict[str, Any], row: dict[str, str], reasons: list[str]) -> dict[str, str]:
    fields = "；".join(f"{header}={row.get(header, '')}" for header in DETAIL_HEADERS if row.get(header, ""))
    return _make_review_row(
        document,
        int(row.get("_page_index") or 0),
        row.get("_method") or "detail_candidate",
        row.get("_raw_text") or fields,
        fields,
        "；".join(reasons),
        "请按原文件核对该明细行后手工修正或补录",
    )


def _split_valid_and_review(document: dict[str, Any], rows: list[dict[str, str]], header_review_rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    valid_rows: list[dict[str, str]] = []
    review_rows = list(header_review_rows)
    for row in rows:
        reasons = _validate_detail_row(row, document)
        if reasons:
            review_rows.append(_review_from_detail(document, row, reasons))
        else:
            valid_rows.append(row)
    if not rows:
        review_rows.append(
            _make_review_row(
                document,
                0,
                "明细抽取",
                "",
                "",
                "未识别到可靠明细行",
                "请检查原文件表头、文字层或 OCR 质量",
            )
        )
    return valid_rows, review_rows


def _validation_notes(document: dict[str, Any], rows: list[dict[str, str]], header_review_rows: list[dict[str, str]]) -> list[str]:
    notes = [f"{row.get('来源', '订单头')}：{row.get('失败原因', '')}" for row in header_review_rows if row.get("失败原因")]
    issue_count = 0
    for row in rows:
        reasons = _validate_detail_row(row, document)
        if not reasons:
            continue
        issue_count += 1
        if row.get("备注"):
            row["备注"] = f"{row['备注']}；系统提示：{'；'.join(reasons)}"
        else:
            row["备注"] = f"系统提示：{'；'.join(reasons)}"
    if issue_count:
        notes.append(f"有 {issue_count} 行明细存在格式或金额提示，已保留在客户结果表，不再移入复核表。")
    return notes


def _group_detail_rows(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    groups: list[dict[str, Any]] = []
    group_map: dict[str, dict[str, Any]] = {}
    for row in rows:
        title = row.get("_table_group") or "明细表"
        group = group_map.get(title)
        if group is None:
            group = {"title": title, "rows": []}
            group_map[title] = group
            groups.append(group)
        group["rows"].append(row)
    return groups


def _detail_headers_for_rows(rows: list[dict[str, str]]) -> list[str]:
    headers = list(DETAIL_HEADERS)
    for row in rows:
        for key in row:
            if key.startswith("_") or key in headers:
                continue
            headers.append(key)
    return headers


def _matrix_from_cells(cells: list[dict[str, Any]]) -> list[list[str]]:
    grid = _grid_for_table(cells)
    if not grid:
        return []
    columns = sorted({column for row_cells in grid.values() for column in row_cells})
    rows: list[list[str]] = []
    for row_index in sorted(grid):
        row = [_clean_text(grid[row_index].get(column, {}).get("text")) for column in columns]
        if any(row):
            rows.append(row)
    if not rows:
        return []

    first_col = 0
    last_col = len(columns) - 1
    while first_col <= last_col and all(not row[first_col] for row in rows):
        first_col += 1
    while last_col >= first_col and all(not row[last_col] for row in rows):
        last_col -= 1
    if first_col > last_col:
        return []
    return [row[first_col : last_col + 1] for row in rows]


def _find_source_table_start(rows: list[list[str]]) -> int | None:
    for index, row in enumerate(rows[:15]):
        row_text = " ".join(row)
        if _looks_like_summary_header_line(row_text):
            continue
        mapped = _header_map({column: value for column, value in enumerate(row)})
        business_fields = {"material_code", "spec", "qty", "unit", "price", "amount", "arrival_date"}
        if len(business_fields.intersection(mapped)) >= 4 and ({"material_code", "spec"} & set(mapped)):
            return index
        if _looks_like_table_header_line(row_text):
            return index
    return None


def _extract_source_detail_tables(document: dict[str, Any]) -> list[dict[str, Any]]:
    document_tables: list[dict[str, Any]] = []
    for table in document.get("source_tables") or []:
        rows = table.get("rows") or []
        if not rows:
            continue
        if table.get("table_type") and table.get("table_type") != "detail_table":
            continue
        document_tables.append(
            {
                "title": table.get("title") or f"明细表 {len(document_tables) + 1}",
                "rows": rows,
                "method": table.get("method") or "source_table",
                "table_type": table.get("table_type") or "detail_table",
            }
        )
    if document_tables:
        return document_tables

    tables: list[dict[str, Any]] = []
    for page in document.get("pages", []):
        page_index = int(page.get("page_index", 0))
        for table in page.get("tables", []):
            rows = _matrix_from_cells(table.get("cells") or [])
            if not rows:
                continue
            start_index = _find_source_table_start(rows)
            if start_index is None:
                non_empty_cols = max((sum(1 for value in row if value) for row in rows), default=0)
                if len(rows) < 2 or non_empty_cols < 3:
                    continue
                start_index = 0
            source_rows = rows[start_index:]
            if len(source_rows) < 2:
                continue
            max_cols = min(max((len(row) for row in source_rows), default=0), 24)
            source_rows = [(row + [""] * max_cols)[:max_cols] for row in source_rows]
            title = f"第 {page_index + 1} 页 表 {int(table.get('table_index') or 0) + 1}"
            tables.append({"title": title, "rows": source_rows})
    return tables


def _combine_source_tables(source_tables: list[dict[str, Any]]) -> dict[str, Any]:
    rows: list[list[str]] = []
    max_cols = 0
    for table in source_tables:
        for source_row in table.get("rows") or []:
            rows.append([_clean_text(value) for value in source_row])
            max_cols = max(max_cols, len(source_row))
    if not rows:
        return {"title": "明细表", "rows": []}
    max_cols = max(1, min(max_cols, 24))
    return {"title": "明细表", "rows": [(row + [""] * max_cols)[:max_cols] for row in rows]}


def _classify_sections(items: list[dict[str, Any]], detail_rows: list[dict[str, str]], warnings: list[str]) -> dict[str, list[str]]:
    detail_signatures = {_compact(row.get("物料编码")) for row in detail_rows if row.get("物料编码")}
    sections = {"notes": [], "terms": [], "payment_info": [], "shipping_info": []}
    for item in items:
        line = item["text"]
        compact = _compact(line)
        if not compact or _looks_like_detail_line(line) or _looks_like_table_header_line(line) or _looks_like_summary_header_line(line):
            continue
        if any(signature and signature in compact for signature in detail_signatures):
            continue
        target = ""
        for section, keywords in SECTION_KEYWORDS.items():
            if any(keyword.lower() in line.lower() for keyword in keywords):
                target = section
                break
        if target:
            sections[target].append(line)

    if warnings:
        sections["notes"].extend(f"系统提示：{warning}" for warning in warnings)
    return {key: _dedupe(value, limit=12) for key, value in sections.items()}


def _docling_table_lines(table: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    title = _clean_text(table.get("title"))
    if title:
        lines.append(title)
    for row in table.get("rows") or []:
        text = "；".join(_clean_text(value) for value in row if _clean_text(value))
        if text:
            lines.append(text)
    return lines


def _append_docling_sections(sections: dict[str, list[str]], classified_tables: list[dict[str, Any]]) -> None:
    for table in classified_tables:
        table_type = table.get("table_type")
        if table_type in {"detail_table", "order_header"}:
            continue
        lines = _docling_table_lines(table)
        if not lines:
            continue
        if table_type == "payment_shipping":
            payment_lines = [line for line in lines if any(keyword in line for keyword in ["付款", "结算", "月结", "账期", "发票", "Payment", "payment"])]
            shipping_lines = [line for line in lines if any(keyword in line for keyword in ["收货", "送货", "地址", "联系人", "电话", "传真", "Ship", "ship"])]
            if payment_lines:
                sections["payment_info"].extend(payment_lines)
            if shipping_lines:
                sections["shipping_info"].extend(shipping_lines)
            if not payment_lines and not shipping_lines:
                sections["payment_info"].extend(lines)
            continue
        if table_type == "terms_notes":
            note_lines = [line for line in lines if any(keyword in line for keyword in ["备注", "合计", "总金额", "大写金额"])]
            term_lines = [line for line in lines if line not in note_lines]
            sections["notes"].extend(note_lines)
            sections["terms"].extend(term_lines or lines)
            continue
        if table_type == "unknown":
            title = _clean_text(table.get("title")) or f"Docling 表 {table.get('table_index', '')}"
            sections["notes"].append(f"Docling 未分类表格已跳过明细输出：{title}")


def _find_total_amount(items: list[dict[str, Any]]) -> Decimal | None:
    for item in items:
        text = item["text"]
        match = re.search(r"(?:总计|合计金额|合计|总金额)\s*[:：]\s*(.+)", text)
        if not match:
            continue
        numbers = re.findall(MONEY_PATTERN, match.group(1))
        for number in reversed(numbers):
            value = _to_decimal(number)
            if value is not None and abs(value) >= 1:
                return value
    return None


def build_business_document(document: dict[str, Any]) -> dict[str, Any]:
    items = _line_items(document)
    header_info, header_review_rows = _extract_header_info(document, items)
    candidates = _extract_detail_candidates(document, items)
    source_tables = _extract_source_detail_tables(document)
    has_docling_source = any(str(table.get("method") or "") == "docling_markdown" for table in source_tables)
    validation_notes = [] if has_docling_source else _validation_notes(document, candidates, header_review_rows)

    total_amount = _find_total_amount(items)
    needs_total_check = str(document.get("file_type") or "") == "image" or document.get("template_id") == "talian_purchase_order"
    if not has_docling_source and total_amount is not None and candidates and needs_total_check:
        valid_sum = sum((_to_decimal(row.get("金额")) or Decimal("0")) for row in candidates)
        tolerance = max(abs(total_amount) * Decimal("0.02"), Decimal("0.05"))
        if abs(valid_sum - total_amount) > tolerance:
            validation_notes.append(
                f"明细金额合计与原文合计不一致：明细合计={valid_sum:,.2f}；原文合计={total_amount:,.2f}。"
            )

    sections = _classify_sections(items, candidates, document.get("warnings") or [])
    _append_docling_sections(sections, document.get("docling_classified_tables") or [])
    sections["notes"].extend(validation_notes)
    sections = {key: _dedupe(value, limit=16) for key, value in sections.items()}
    use_source_tables = bool(source_tables) and (has_docling_source or not candidates or not document.get("template_id"))
    if not candidates and not source_tables:
        sections["notes"].append("未生成结构化明细；当前文件未检测到可写入的明细表。")
    return {
        "header_info": header_info,
        "detail_rows": candidates,
        "detail_groups": _group_detail_rows(candidates),
        "source_tables": source_tables if use_source_tables else [],
        "review_rows": [],
        "notes": sections["notes"],
        "terms": sections["terms"],
        "payment_info": sections["payment_info"],
        "shipping_info": sections["shipping_info"],
        "page_count": int(document.get("page_count") or len(document.get("pages") or [])),
    }


def _set_customer_sheet_layout(ws) -> None:
    widths = {"A": 8, "B": 22, "C": 50, "D": 10, "E": 14, "F": 14, "G": 16, "H": 16, "I": 34}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for column_index in range(10, 25):
        ws.column_dimensions[get_column_letter(column_index)].width = 18
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A1"


def _set_review_sheet_layout(ws) -> None:
    widths = {"A": 22, "B": 8, "C": 18, "D": 58, "E": 42, "F": 36, "G": 32}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


def _style_range(ws, row: int, start_col: int, end_col: int, *, fill: str, font_color: str = "000000", bold: bool = False) -> None:
    for column in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=column)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Microsoft YaHei", color=font_color, bold=bold)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_border(ws, start_row: int, end_row: int, start_col: int = 1, end_col: int = 9) -> None:
    thin = Side(style="thin", color="D9E2F3")
    for row in range(start_row, end_row + 1):
        for column in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=column)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_header_info(ws, row: int, header_info: dict[str, str]) -> int:
    pairs = [(HEADER_LABELS[key], value) for key, value in header_info.items() if value and key in HEADER_LABELS]
    if not pairs:
        return row
    start_row = row
    for index in range(0, len(pairs), 2):
        left_label, left_value = pairs[index]
        ws.cell(row=row, column=1, value=left_label)
        ws.cell(row=row, column=2, value=left_value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        if index + 1 < len(pairs):
            right_label, right_value = pairs[index + 1]
            ws.cell(row=row, column=5, value=right_label)
            ws.cell(row=row, column=6, value=right_value)
            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
        for label_col in (1, 5):
            cell = ws.cell(row=row, column=label_col)
            cell.fill = PatternFill("solid", fgColor="EEF3F8")
            cell.font = Font(name="Microsoft YaHei", bold=True, color="44546A")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for value_col in (2, 6):
            cell = ws.cell(row=row, column=value_col)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1
    _apply_border(ws, start_row, row - 1)
    return row + 1


def _write_detail_table(ws, row: int, detail_rows: list[dict[str, str]], title: str = "明细表") -> tuple[int, int]:
    headers = _detail_headers_for_rows(detail_rows)
    max_cols = max(1, len(headers))
    title_row = row
    ws.cell(row=title_row, column=1, value=title)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max_cols)
    _style_range(ws, title_row, 1, max_cols, fill="D9EAF7", bold=True)
    row += 1
    for column, header in enumerate(headers, start=1):
        ws.cell(row=row, column=column, value=header)
    _style_range(ws, row, 1, max_cols, fill="1F4E78", font_color="FFFFFF", bold=True)
    header_row = row
    row += 1

    if not detail_rows:
        ws.cell(row=row, column=1, value="未生成明细")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
    else:
        for detail in detail_rows:
            for column, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=column, value=detail.get(header, ""))
                horizontal = "right" if header in {"数量", "单价", "金额"} else "left"
                if header == "序号":
                    horizontal = "center"
                cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
            row += 1
    _apply_border(ws, title_row, row - 1, 1, max_cols)
    for row_idx in range(header_row + 1, row):
        ws.row_dimensions[row_idx].height = 36
    return row + 1, len(detail_rows)


def _write_raw_detail_table(ws, row: int, table: dict[str, Any]) -> tuple[int, int]:
    rows = table.get("rows") or []
    if not rows:
        return row, 0
    max_cols = max((len(source_row) for source_row in rows), default=1)
    max_cols = max(1, min(max_cols, 24))
    title_row = row
    table_title = _clean_text(table.get("title")) or "明细表"
    ws.cell(row=title_row, column=1, value=table_title if table_title == "明细表" else f"明细表 - {table_title}")
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max_cols)
    _style_range(ws, title_row, 1, max_cols, fill="D9EAF7", bold=True)
    row += 1

    data_start = row
    for source_row in rows:
        padded = (source_row + [""] * max_cols)[:max_cols]
        for column, value in enumerate(padded, start=1):
            cell = ws.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1
    if row > data_start:
        _style_range(ws, data_start, 1, max_cols, fill="1F4E78", font_color="FFFFFF", bold=True)
    _apply_border(ws, title_row, row - 1, 1, max_cols)
    for row_idx in range(data_start + 1, row):
        ws.row_dimensions[row_idx].height = 36
    return row + 1, max(0, len(rows) - 1)


def _write_section(ws, row: int, title: str, lines: list[str]) -> int:
    if not lines:
        return row
    ws.cell(row=row, column=1, value=title)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    _style_range(ws, row, 1, 9, fill="E2F0D9", bold=True)
    row += 1
    ws.cell(row=row, column=1, value="\n".join(lines))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = min(max(24, 18 * len(lines)), 160)
    _apply_border(ws, row - 1, row)
    return row + 2


def _write_document(ws, row: int, document: dict[str, Any], section_index: int) -> tuple[int, int, int]:
    business_doc = build_business_document(document)
    header_info = business_doc["header_info"]
    detail_rows = business_doc["detail_rows"]
    title = header_info.get("source_file") or document.get("source_file") or f"文件 {section_index}"
    ws.cell(row=row, column=1, value=f"{section_index}. {title}")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    _style_range(ws, row, 1, 9, fill="0B1F4D", font_color="FFFFFF", bold=True)
    ws.row_dimensions[row].height = 24
    row += 1
    row = _write_header_info(ws, row, header_info)
    detail_count = 0
    if business_doc["source_tables"]:
        row, detail_count = _write_raw_detail_table(ws, row, _combine_source_tables(business_doc["source_tables"]))
    elif detail_rows:
        row, detail_count = _write_detail_table(ws, row, detail_rows, title="明细表")
    else:
        row, detail_count = _write_detail_table(ws, row, detail_rows)
    row = _write_section(ws, row, "备注", business_doc["notes"])
    row = _write_section(ws, row, "条款", business_doc["terms"])
    row = _write_section(ws, row, "付款信息", business_doc["payment_info"])
    row = _write_section(ws, row, "收货信息", business_doc["shipping_info"])
    return row + 1, detail_count, business_doc["page_count"]


def _write_review_sheet(ws, review_rows: list[dict[str, str]]) -> None:
    _set_review_sheet_layout(ws)
    for column, header in enumerate(REVIEW_HEADERS, start=1):
        ws.cell(row=1, column=column, value=header)
    _style_range(ws, 1, 1, len(REVIEW_HEADERS), fill="9E480E", font_color="FFFFFF", bold=True)

    if not review_rows:
        ws.cell(row=2, column=1, value="暂无需复核明细")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(REVIEW_HEADERS))
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        _apply_border(ws, 1, 2, 1, len(REVIEW_HEADERS))
        return

    row_index = 2
    for review in review_rows:
        for column, header in enumerate(REVIEW_HEADERS, start=1):
            cell = ws.cell(row=row_index, column=column, value=review.get(header, ""))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row_index].height = 58
        row_index += 1
    _apply_border(ws, 1, row_index - 1, 1, len(REVIEW_HEADERS))


def write_conversion_workbook(documents: list[dict[str, Any]], output_path: Path, *, restore_mode: bool = False) -> dict[str, int]:
    wb = Workbook()
    customer_ws = wb.active
    customer_ws.title = "客户结果表"
    _set_customer_sheet_layout(customer_ws)

    row = 1
    total_detail_rows = 0
    total_pages = 0
    if not documents:
        customer_ws.cell(row=row, column=1, value="未解析到可转换文件")
        customer_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        _style_range(customer_ws, row, 1, 9, fill="FCE4D6", bold=True)
    else:
        for section_index, document in enumerate(documents, start=1):
            row, detail_count, page_count = _write_document(customer_ws, row, document, section_index)
            total_detail_rows += detail_count
            total_pages += page_count

    wb.save(output_path)
    return {
        "structured_count": total_detail_rows,
        "cell_count": total_detail_rows,
        "review_count": 0,
        "page_count": total_pages,
    }
