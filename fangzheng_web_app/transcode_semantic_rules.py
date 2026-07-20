from __future__ import annotations

import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from .db import get_setting, set_setting
from .paths import DEFAULT_RULES_DIR, TRANSCODE_SEMANTIC_RULES_DIR, TRANSCODE_SEMANTIC_RULES_VERSIONS_DIR
from .transcode_semantic_rule_finalizer import validate_atomic_conditions
from .transcode_semantic_service import TARGET_FIELDS


RULE_WORKBOOK_FILENAME = "transcode_semantic_rules.xlsx"
MACHINE_RULE_FILENAME = "transcode_semantic_rules.json"
MANIFEST_FILENAME = "manifest.json"
RULE_SHEET_NAME = "模型语义规则"
PENDING_SHEET_NAME = "待业务确认"
SCHEMA_VERSION = "1.0"
DEFAULT_SEMANTIC_RULE_WORKBOOK = DEFAULT_RULES_DIR / "transcode_semantic" / RULE_WORKBOOK_FILENAME

REQUIRED_HEADERS = {
    "规则ID",
    "来源候选ID",
    "启用",
    "客户代码",
    "客户简称",
    "来源行号",
    "业务字段",
    "规则原文",
    "语义类型",
    "目标字段",
    "标准语义值",
    "原文目标值",
    "条件JSON",
    "所需订单字段",
    "执行方式",
    "优先级",
    "模型版本",
    "提示词SHA256",
    "原文证据",
    "业务确认",
    "确认依据",
    "备注",
}

ALLOWED_BUSINESS_FIELDS = {
    "胶系",
    "基板厚度",
    "铜箔规格",
    "基板尺寸",
    "胶水类别",
    "铜箔类型+印字/非印字",
    "基板级别",
    "总/芯厚",
}

ALLOWED_EXECUTION_MODES = {"结构化后可确定性执行"}


class SemanticRuleAssetError(ValueError):
    pass


def _active_key() -> str:
    return "active_transcode_semantic_rule_version"


def _history_key() -> str:
    return "transcode_semantic_rule_history"


def get_active_transcode_semantic_rule_version() -> str:
    return get_setting(_active_key(), "") or ""


def ensure_default_transcode_semantic_rule_version() -> str:
    active = get_active_transcode_semantic_rule_version()
    if active:
        try:
            validate_transcode_semantic_rule_version(active)
            return active
        except (FileNotFoundError, SemanticRuleAssetError):
            pass
    if not DEFAULT_SEMANTIC_RULE_WORKBOOK.exists():
        raise FileNotFoundError(f"内置模型语义规则表不存在：{DEFAULT_SEMANTIC_RULE_WORKBOOK}")
    return publish_transcode_semantic_rule_version(
        DEFAULT_SEMANTIC_RULE_WORKBOOK,
        updated_by="system",
        approval_basis="项目内置已确认模型语义规则",
        remark="首次启动自动发布项目内置模型语义规则",
    )


def get_transcode_semantic_rule_history() -> list[dict[str, Any]]:
    raw = get_setting(_history_key(), "[]") or "[]"
    try:
        value = json.loads(raw)
    except json.JSONDecodeError:
        return []
    return value if isinstance(value, list) else []


def get_transcode_semantic_rule_dir(version: str | None = None) -> Path:
    selected = version or get_active_transcode_semantic_rule_version()
    return TRANSCODE_SEMANTIC_RULES_VERSIONS_DIR / selected if selected else TRANSCODE_SEMANTIC_RULES_DIR


def get_transcode_semantic_rule_workbook_path(version: str | None = None) -> Path:
    return get_transcode_semantic_rule_dir(version) / RULE_WORKBOOK_FILENAME


def get_transcode_semantic_machine_rule_path(version: str | None = None) -> Path:
    return get_transcode_semantic_rule_dir(version) / MACHINE_RULE_FILENAME


def parse_semantic_rule_workbook(path: str | Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise SemanticRuleAssetError(f"模型语义规则表不存在：{workbook_path}")
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if RULE_SHEET_NAME not in workbook.sheetnames:
        raise SemanticRuleAssetError(f"模型语义规则表缺少Sheet：{RULE_SHEET_NAME}")
    sheet = workbook[RULE_SHEET_NAME]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    missing = sorted(REQUIRED_HEADERS - set(headers))
    if missing:
        raise SemanticRuleAssetError(f"模型语义规则表缺少列：{', '.join(missing)}")

    rules: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
        rule_id = _clean(row.get("规则ID"))
        if not rule_id:
            continue
        if rule_id in seen_ids:
            raise SemanticRuleAssetError(f"模型语义规则ID重复：{rule_id}")
        seen_ids.add(rule_id)
        rules.append(_parse_rule_row(row, row_number))

    pending_count = 0
    if PENDING_SHEET_NAME in workbook.sheetnames:
        pending_sheet = workbook[PENDING_SHEET_NAME]
        pending_count = sum(
            1
            for value in pending_sheet.iter_rows(min_row=2, max_col=1, values_only=True)
            if _clean(value[0])
        )
    if not rules:
        raise SemanticRuleAssetError("模型语义规则表没有可发布规则")
    return rules, {
        "rule_count": len(rules),
        "pending_count": pending_count,
        "models": sorted({rule["model"] for rule in rules}),
        "prompt_sha256": sorted({rule["prompt_sha256"] for rule in rules}),
    }


def publish_transcode_semantic_rule_version(
    source_workbook: str | Path,
    *,
    updated_by: str,
    approval_basis: str,
    remark: str = "",
    activate: bool = True,
) -> str:
    source_path = Path(source_workbook)
    rules, summary = parse_semantic_rule_workbook(source_path)
    version = _new_version_name()
    version_dir = TRANSCODE_SEMANTIC_RULES_VERSIONS_DIR / version
    version_dir.mkdir(parents=True, exist_ok=False)

    workbook_path = version_dir / RULE_WORKBOOK_FILENAME
    machine_path = version_dir / MACHINE_RULE_FILENAME
    manifest_path = version_dir / MANIFEST_FILENAME
    shutil.copy2(source_path, workbook_path)

    machine_payload = {
        "schema_version": SCHEMA_VERSION,
        "version": version,
        "rules": rules,
    }
    _write_json_atomic(machine_path, machine_payload)
    manifest = {
        "schema_version": SCHEMA_VERSION,
        "version": version,
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "updated_by": updated_by,
        "approval_basis": approval_basis,
        "remark": remark or "发布业务已确认的模型语义规则",
        "source_file": source_path.name,
        "rule_count": summary["rule_count"],
        "pending_count": summary["pending_count"],
        "models": summary["models"],
        "prompt_sha256": summary["prompt_sha256"],
        "files": {
            RULE_WORKBOOK_FILENAME: _sha256(workbook_path),
            MACHINE_RULE_FILENAME: _sha256(machine_path),
        },
    }
    _write_json_atomic(manifest_path, manifest)
    validate_transcode_semantic_rule_version(version)

    history = get_transcode_semantic_rule_history()
    history.insert(
        0,
        {
            "version": version,
            "updated_at": manifest["created_at"],
            "updated_by": updated_by,
            "approval_basis": approval_basis,
            "remark": manifest["remark"],
            "rule_count": summary["rule_count"],
            "pending_count": summary["pending_count"],
        },
    )
    set_setting(_history_key(), json.dumps(history[:50], ensure_ascii=False))
    if activate:
        activate_transcode_semantic_rule_version(version)
    return version


def validate_transcode_semantic_rule_version(version: str) -> dict[str, Any]:
    version_dir = TRANSCODE_SEMANTIC_RULES_VERSIONS_DIR / str(version or "").strip()
    manifest_path = version_dir / MANIFEST_FILENAME
    if not manifest_path.exists():
        raise SemanticRuleAssetError(f"模型语义规则版本不存在：{version}")
    manifest = _read_json(manifest_path)
    if manifest.get("schema_version") != SCHEMA_VERSION or manifest.get("version") != version:
        raise SemanticRuleAssetError(f"模型语义规则manifest不匹配：{version}")
    for filename in [RULE_WORKBOOK_FILENAME, MACHINE_RULE_FILENAME]:
        path = version_dir / filename
        if not path.exists():
            raise SemanticRuleAssetError(f"模型语义规则版本缺少文件：{filename}")
        expected_hash = str((manifest.get("files") or {}).get(filename) or "")
        if not expected_hash or _sha256(path) != expected_hash:
            raise SemanticRuleAssetError(f"模型语义规则文件哈希不匹配：{filename}")
    rules = load_transcode_semantic_rules(version, validate_manifest=False)
    if len(rules) != int(manifest.get("rule_count") or 0):
        raise SemanticRuleAssetError("模型语义规则数量与manifest不一致")
    return manifest


def activate_transcode_semantic_rule_version(version: str) -> None:
    validate_transcode_semantic_rule_version(version)
    set_setting(_active_key(), version)


def load_transcode_semantic_rules(
    version: str | None = None,
    *,
    validate_manifest: bool = True,
) -> list[dict[str, Any]]:
    selected = version or get_active_transcode_semantic_rule_version()
    if not selected:
        return []
    if validate_manifest:
        validate_transcode_semantic_rule_version(selected)
    payload = _read_json(TRANSCODE_SEMANTIC_RULES_VERSIONS_DIR / selected / MACHINE_RULE_FILENAME)
    if payload.get("schema_version") != SCHEMA_VERSION or payload.get("version") != selected:
        raise SemanticRuleAssetError("模型语义机器规则版本信息不一致")
    rules = payload.get("rules")
    if not isinstance(rules, list):
        raise SemanticRuleAssetError("模型语义机器规则rules不是数组")
    for index, rule in enumerate(rules):
        _validate_machine_rule(rule, index)
    return rules


def _parse_rule_row(row: dict[str, Any], row_number: int) -> dict[str, Any]:
    if _clean(row.get("启用")) != "是":
        raise SemanticRuleAssetError(f"第{row_number}行正式规则未启用")
    if _clean(row.get("业务确认")) != "确认":
        raise SemanticRuleAssetError(f"第{row_number}行正式规则未经业务确认")
    business_field = _clean(row.get("业务字段"))
    if business_field not in ALLOWED_BUSINESS_FIELDS:
        raise SemanticRuleAssetError(f"第{row_number}行业务字段不在当前CCL范围：{business_field}")
    execution_mode = _clean(row.get("执行方式"))
    if execution_mode not in ALLOWED_EXECUTION_MODES:
        raise SemanticRuleAssetError(f"第{row_number}行执行方式不可发布：{execution_mode}")
    conditions = _parse_json_list(row.get("条件JSON"), row_number, "条件JSON")
    if not conditions:
        raise SemanticRuleAssetError(f"第{row_number}行没有可执行条件")
    target_fields = _split_values(row.get("目标字段"))
    if not target_fields or any(field not in TARGET_FIELDS for field in target_fields):
        raise SemanticRuleAssetError(f"第{row_number}行目标字段无效：{target_fields}")
    priority = _to_int(row.get("优先级"), default=100)
    rule = {
        "rule_id": _clean(row.get("规则ID")),
        "source_candidate_id": _clean(row.get("来源候选ID")),
        "enabled": True,
        "customer_code": _clean(row.get("客户代码")),
        "customer_name": _clean(row.get("客户简称")),
        "source_row": _to_int(row.get("来源行号")),
        "business_field": business_field,
        "source_text": _clean(row.get("规则原文")),
        "semantic_types": _split_values(row.get("语义类型")),
        "target_fields": target_fields,
        "normalized_values": _split_values(row.get("标准语义值")),
        "stated_target_values": _split_values(row.get("原文目标值")),
        "conditions": conditions,
        "required_input_fields": _split_values(row.get("所需订单字段")),
        "execution_mode": execution_mode,
        "priority": priority,
        "model": _clean(row.get("模型版本")),
        "prompt_sha256": _clean(row.get("提示词SHA256")),
        "evidence_texts": _split_values(row.get("原文证据")),
        "approval": {
            "status": "confirmed",
            "basis": _clean(row.get("确认依据")),
        },
        "note": _clean(row.get("备注")),
    }
    _validate_machine_rule(rule, row_number - 2)
    return rule


def _validate_machine_rule(rule: Any, index: int) -> None:
    if not isinstance(rule, dict):
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项不是对象")
    required = {
        "rule_id",
        "source_candidate_id",
        "enabled",
        "customer_code",
        "customer_name",
        "source_row",
        "business_field",
        "source_text",
        "semantic_types",
        "target_fields",
        "normalized_values",
        "stated_target_values",
        "conditions",
        "required_input_fields",
        "execution_mode",
        "priority",
        "model",
        "prompt_sha256",
        "evidence_texts",
        "approval",
        "note",
    }
    if set(rule) != required:
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项字段不完整")
    if not rule["rule_id"] or not rule["source_candidate_id"] or not rule["enabled"]:
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项未启用或缺少ID")
    if rule["business_field"] not in ALLOWED_BUSINESS_FIELDS:
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项业务字段无效")
    if rule["execution_mode"] not in ALLOWED_EXECUTION_MODES:
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项执行方式无效")
    if not isinstance(rule["conditions"], list) or not rule["conditions"]:
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项没有条件")
    if not all(field in TARGET_FIELDS for field in rule["target_fields"]):
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项目标字段无效")
    if (rule.get("approval") or {}).get("status") != "confirmed":
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项未经确认")
    if len(rule.get("semantic_types") or []) != 1:
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项必须是单一语义类型")
    if len(rule.get("target_fields") or []) != 1:
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项必须是单一目标字段")
    if len(rule.get("normalized_values") or []) != 1:
        raise SemanticRuleAssetError(f"机器规则第{index + 1}项必须有且仅有一个标准语义值")
    try:
        validate_atomic_conditions(rule["conditions"], context=f"机器规则第{index + 1}项")
    except ValueError as exc:
        raise SemanticRuleAssetError(str(exc)) from exc


def _new_version_name() -> str:
    return datetime.now().strftime("transcode_semantic_rules_%Y%m%d_%H%M%S_%f")


def _read_json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SemanticRuleAssetError(f"模型语义JSON文件无效：{path}") from exc
    if not isinstance(value, dict):
        raise SemanticRuleAssetError(f"模型语义JSON根节点不是对象：{path}")
    return value


def _write_json_atomic(path: Path, value: dict[str, Any]) -> None:
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)


def _parse_json_list(value: Any, row_number: int, field: str) -> list[Any]:
    try:
        parsed = json.loads(str(value or "[]"))
    except json.JSONDecodeError as exc:
        raise SemanticRuleAssetError(f"第{row_number}行{field}不是有效JSON") from exc
    if not isinstance(parsed, list):
        raise SemanticRuleAssetError(f"第{row_number}行{field}必须是数组")
    return parsed


def _split_values(value: Any) -> list[str]:
    text = _clean(value)
    return [item.strip() for item in text.replace(",", "；").split("；") if item.strip()]


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _to_int(value: Any, *, default: int = 0) -> int:
    try:
        return int(float(str(value or default).strip()))
    except ValueError:
        return default


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()
