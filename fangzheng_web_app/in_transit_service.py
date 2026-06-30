from __future__ import annotations

import importlib
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.utils import secure_filename

from .db import append_job_log, create_job, get_job, update_job_status
from .excel_utils import load_workbook_compat
from .job_control import launch_job_process
from .paths import JOBS_DIR
from .transcode_rules import get_active_transcode_rule_version, get_transcode_rule_file_path


FEATURE_KEY = "in_transit"
RULE_VERSION = "in_transit_v1"
ALLOWED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
INTERNAL_DETAIL_SHEET = "厂内明细"
CUSTOMER_DETAIL_SHEET = "客户明细"

SYSTEM_KEEP_COLUMNS = [
    "预交货日",
    "单别单号",
    "项次",
    "厂别",
    "订单日期",
    "客户编号",
    "客户简称",
    "所属集团公司",
    "品号",
    "品名",
    "规格",
    "双幅",
    "订单数量",
    "单位",
    "数量",
    "单位",
    "客户单号",
    "客户产品编号",
    "客户规格",
    "到货日期",
]

SYSTEM_REQUIRED_COLUMNS = {
    "预交货日",
    "单别单号",
    "项次",
    "厂别",
    "订单数量",
    "客户简称",
    "客户单号",
    "客户产品编号",
    "品名",
}

CUSTOMER_REQUIRED_COLUMNS = {
    "单据编号",
    "物料编码",
    "单据行号",
    "订单承诺日期",
    "物料数量",
    "物料单位",
    "物料说明",
}

ARRIVAL_DIRECT_CUSTOMERS = {"无锡深南", "南通深南"}
ARRIVAL_OFFSET_CUSTOMERS = {"深南电路"}


@dataclass
class SystemRow:
    excel_row: int
    values: dict[str, Any]
    output_values: list[Any]
    key: tuple[str, str, str, str]
    order_qty: float | None
    product_name: str
    match_status: str = ""
    customer_qty: float | None = None
    qty_judgement: str = ""


@dataclass
class CustomerRow:
    excel_row: int
    values: dict[str, Any]
    output_values: list[Any]
    key: tuple[str, str, str, str]
    material_qty: float | None
    material_desc: str
    order_qty: float | None = None
    qty_judgement: str = ""
    product_name: str = ""
    converted_name: str = ""
    name_judgement: str = ""
    system_arrival_date: date | None = None
    customer_commitment_date: date | None = None
    date_judgement: str = ""
    match_status: str = ""


def _safe_excel_filename(original_filename: str, fallback_stem: str) -> str:
    original_name = str(original_filename or "").strip()
    original_path = Path(original_name)
    suffix = original_path.suffix.lower()
    if suffix not in ALLOWED_EXCEL_EXTENSIONS:
        suffix = ".xlsx"

    if original_path.suffix:
        stem = original_path.name[: -len(original_path.suffix)]
    else:
        stem = original_path.name
    safe_stem = secure_filename(stem).strip("._-")
    if not safe_stem or safe_stem.lower() in {"xlsx", "xlsm", "xls"}:
        safe_stem = fallback_stem
    return f"{safe_stem}{suffix}"


def queue_in_transit_job(employee_id: str, uploaded_file, original_filename: str) -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)

    safe_name = _safe_excel_filename(original_filename, "in_transit")
    input_path = employee_dir / f"{timestamp}_in_transit_{safe_name}"
    uploaded_file.save(input_path)

    display_name = f"深南在途核对：{original_filename}"
    job_id = create_job(
        employee_id,
        display_name,
        str(input_path),
        RULE_VERSION,
        feature=FEATURE_KEY,
    )
    launch_job_process(job_id, FEATURE_KEY, employee_id)
    return job_id


def run_in_transit_job(job_id: int, employee_id: str) -> None:
    update_job_status(job_id, status="running", log_text="")
    job = get_job(job_id)
    if not job:
        return

    append_job_log(job_id, "开始深南在途核对任务")
    try:
        input_path = Path(job["stored_input_path"])

        append_job_log(job_id, f"输入文件：{input_path.name}")
        append_job_log(job_id, f"读取 Sheet：{INTERNAL_DETAIL_SHEET} / {CUSTOMER_DETAIL_SHEET}")

        result = calculate_in_transit_workbook(input_path, job_id=job_id)
        output_path = input_path.with_name(f"{input_path.stem}_深南在途核对结果.xlsx")
        save_result_workbook(result, output_path)

        update_job_status(
            job_id,
            status="completed",
            stored_result_path=str(output_path),
            success_count=result["summary"]["matched_pairs"],
            fail_count=result["summary"]["qty_error_count"] + result["summary"]["name_error_count"],
            skip_count=result["summary"]["unmatched_or_duplicate_count"] + result["summary"]["name_unknown_count"],
            current_row=result["summary"]["processed_total"],
            total_rows=result["summary"]["processed_total"],
            completed=True,
        )
        append_job_log(job_id, f"核对完成，结果文件：{output_path.name}")
    except Exception as exc:
        append_job_log(job_id, f"深南在途核对失败：{exc}")
        update_job_status(job_id, status="failed", error_message=str(exc), completed=True)
        raise


def calculate_in_transit_workbook(
    input_path: Path,
    customer_path: Path | None = None,
    *,
    job_id: int | None = None,
    base_date: date | None = None,
) -> dict:
    system_sheet = INTERNAL_DETAIL_SHEET if customer_path is None else None
    customer_sheet = CUSTOMER_DETAIL_SHEET if customer_path is None else None
    customer_source = input_path if customer_path is None else customer_path

    if job_id:
        append_job_log(job_id, "开始清洗系统数据。", current_row=0, total_rows=0)
    system_rows, system_removed = load_system_rows(input_path, job_id=job_id, sheet_name=system_sheet)
    if job_id:
        append_job_log(job_id, f"系统数据清洗完成：保留 {len(system_rows)} 行。")

    if job_id:
        append_job_log(job_id, "开始清洗客户数据。")
    customer_rows, customer_stats = load_customer_rows(customer_source, job_id=job_id, sheet_name=customer_sheet)
    if job_id:
        append_job_log(job_id, f"客户数据清洗完成：保留 {len(customer_rows)} 行。")

    system_counts = Counter(row.key for row in system_rows)
    customer_counts = Counter(row.key for row in customer_rows)
    customer_by_key = _unique_map(customer_rows, customer_counts)
    system_by_key = _unique_map(system_rows, system_counts)

    matched_pairs = 0
    qty_ok_count = 0
    qty_error_count = 0
    unmatched_or_duplicate_count = 0
    name_ok_count = 0
    name_error_count = 0
    name_unknown_count = 0
    date_ok_count = 0
    date_error_count = 0
    date_unknown_count = 0

    if job_id:
        append_job_log(job_id, "开始匹配并核对数量。", current_row=0, total_rows=len(system_rows) + len(customer_rows))

    for index, row in enumerate(system_rows, start=1):
        match = _match_partner(row.key, system_counts, customer_counts, customer_by_key)
        row.match_status = match[0]
        customer_row = match[1]
        if customer_row:
            matched_pairs += 1
            row.customer_qty = customer_row.material_qty
            row.qty_judgement = _qty_judgement(row.order_qty, customer_row.material_qty)
            if row.qty_judgement == "正确":
                qty_ok_count += 1
            elif row.qty_judgement == "错误":
                qty_error_count += 1
        else:
            unmatched_or_duplicate_count += 1
        if job_id and index % 20 == 0:
            append_job_log(
                job_id,
                f"已匹配系统数据 {index}/{len(system_rows)}",
                current_row=index,
                total_rows=len(system_rows) + len(customer_rows),
            )

    matched_customer_rows: list[CustomerRow] = []
    for index, row in enumerate(customer_rows, start=1):
        match = _match_partner(row.key, customer_counts, system_counts, system_by_key)
        row.match_status = match[0]
        system_row = match[1]
        if system_row:
            row.order_qty = system_row.order_qty
            row.qty_judgement = _qty_judgement(system_row.order_qty, row.material_qty)
            row.product_name = system_row.product_name
            row.system_arrival_date = _as_date(system_row.values.get("到货日期"))
            row.customer_commitment_date = _as_date(row.values.get("订单承诺日期"))
            row.date_judgement = _date_judgement(row.system_arrival_date, row.customer_commitment_date)
            if row.date_judgement == "正确":
                date_ok_count += 1
            elif row.date_judgement == "错误":
                date_error_count += 1
            elif row.date_judgement == "无法判断":
                date_unknown_count += 1
            matched_customer_rows.append(row)
        else:
            row.name_judgement = ""
            unmatched_or_duplicate_count += 1
        if job_id and index % 20 == 0:
            append_job_log(
                job_id,
                f"已匹配客户数据 {index}/{len(customer_rows)}",
                current_row=len(system_rows) + index,
                total_rows=len(system_rows) + len(customer_rows),
            )

    if job_id:
        append_job_log(job_id, f"开始客户物料说明转品名：{len(matched_customer_rows)} 行。")
    transcode_tables = _load_transcode_tables() if matched_customer_rows else None
    for index, row in enumerate(matched_customer_rows, start=1):
        row.converted_name = convert_material_desc_to_name(row.material_desc, transcode_tables)
        row.name_judgement = _name_judgement(row.product_name, row.converted_name)
        if row.name_judgement == "正确":
            name_ok_count += 1
        elif row.name_judgement == "错误":
            name_error_count += 1
        else:
            name_unknown_count += 1
        if job_id and index % 20 == 0:
            append_job_log(job_id, f"已转换客户品名 {index}/{len(matched_customer_rows)}")

    delivery_base_date = base_date or _delivery_base_date_from_customer_rows(customer_rows) or date.today()
    delivery_rows = _delivery_detail_rows(system_rows, delivery_base_date)
    processed_total = len(system_rows) + len(customer_rows)
    summary = {
        "system_source_rows": system_removed["source_rows"],
        "system_clean_rows": len(system_rows),
        "system_removed_341": system_removed["removed_341"],
        "system_removed_blank_qty": system_removed["removed_blank_qty"],
        "system_removed_blank_factory": system_removed["removed_blank_factory"],
        "customer_rows": len(customer_rows),
        "customer_roll_rows": customer_stats["roll_rows"],
        "customer_roll_missing_length": customer_stats["roll_missing_length"],
        "matched_pairs": matched_pairs,
        "qty_ok_count": qty_ok_count,
        "qty_error_count": qty_error_count,
        "name_ok_count": name_ok_count,
        "name_error_count": name_error_count,
        "name_unknown_count": name_unknown_count,
        "date_ok_count": date_ok_count,
        "date_error_count": date_error_count,
        "date_unknown_count": date_unknown_count,
        "delivery_detail_rows": len(delivery_rows),
        "delivery_filter_base_date": delivery_base_date,
        "unmatched_or_duplicate_count": unmatched_or_duplicate_count,
        "processed_total": processed_total,
    }

    if job_id:
        append_job_log(job_id, "开始生成结果 Excel。")
    return {"system_rows": system_rows, "delivery_rows": delivery_rows, "customer_rows": customer_rows, "summary": summary}


def load_system_rows(path: Path, *, job_id: int | None = None, sheet_name: str | None = None) -> tuple[list[SystemRow], Counter]:
    workbook = load_workbook_compat(path, data_only=True)
    worksheet = (
        _select_named_sheet_with_columns(workbook, sheet_name, SYSTEM_REQUIRED_COLUMNS)
        if sheet_name
        else _select_sheet_with_columns(workbook, SYSTEM_REQUIRED_COLUMNS)
    )
    header_row = _find_header_row(worksheet, SYSTEM_REQUIRED_COLUMNS)
    headers = _header_occurrences(worksheet, header_row)
    output_specs = _system_output_specs(headers)
    output_columns = _output_columns_from_specs(headers, output_specs)

    stats = Counter(source_rows=max(worksheet.max_row - header_row, 0))
    rows: list[SystemRow] = []
    for excel_row in range(header_row + 1, worksheet.max_row + 1):
        values = _row_values_by_header(worksheet, excel_row, headers)
        order_no = _text(values.get("单别单号"))
        order_qty = _number(values.get("订单数量"))
        factory = _text(values.get("厂别"))
        if "341-" in order_no or "219-" in order_no:
            stats["removed_341"] += 1
            continue
        if order_qty is None or order_qty == 0:
            stats["removed_blank_qty"] += 1
            continue
        if not factory:
            stats["removed_blank_factory"] += 1
            continue

        arrival_date = _arrival_date(values.get("预交货日"), factory, _text(values.get("客户简称")))
        values["到货日期"] = arrival_date
        output_values = [_value_for_output_spec(worksheet, excel_row, values, spec, output_columns) for spec in output_specs]
        key = (
            _text(values.get("客户单号")),
            _text(values.get("客户产品编号")),
            _text(values.get("项次")),
            _date_key(arrival_date),
        )
        rows.append(
            SystemRow(
                excel_row=excel_row,
                values=values,
                output_values=output_values,
                key=key,
                order_qty=order_qty,
                product_name=_text(values.get("品名")),
            )
        )
        processed = excel_row - header_row
        if job_id and processed % 500 == 0:
            append_job_log(
                job_id,
                f"系统数据清洗中：{processed}/{stats['source_rows']}，保留 {len(rows)} 行。",
                current_row=processed,
                total_rows=stats["source_rows"],
            )
    return rows, stats


def load_customer_rows(path: Path, *, job_id: int | None = None, sheet_name: str | None = None) -> tuple[list[CustomerRow], Counter]:
    workbook = load_workbook_compat(path, data_only=True)
    worksheet = (
        _select_named_sheet_with_columns(workbook, sheet_name, CUSTOMER_REQUIRED_COLUMNS)
        if sheet_name
        else _select_sheet_with_columns(workbook, CUSTOMER_REQUIRED_COLUMNS)
    )
    header_row = _find_header_row(worksheet, CUSTOMER_REQUIRED_COLUMNS)
    headers = _header_occurrences(worksheet, header_row)
    header_names = _first_header_names(headers)

    stats = Counter()
    rows: list[CustomerRow] = []
    for excel_row in range(header_row + 1, worksheet.max_row + 1):
        values = _row_values_by_header(worksheet, excel_row, headers)
        if not any(_text(values.get(name)) for name in CUSTOMER_REQUIRED_COLUMNS):
            continue
        material_qty = _number(values.get("物料数量"))
        unit = _text(values.get("物料单位"))
        desc = _text(values.get("物料说明"))
        adjusted_qty = material_qty
        if unit == "卷":
            stats["roll_rows"] += 1
            length = extract_pp_roll_length(desc)
            if length is None:
                stats["roll_missing_length"] += 1
            elif material_qty is not None:
                adjusted_qty = material_qty * length
            values["物料数量"] = adjusted_qty

        key = (
            _text(values.get("单据编号")),
            _text(values.get("物料编码")),
            _text(values.get("单据行号")),
            _date_key(values.get("订单承诺日期")),
        )
        output_values = [values.get(name) for name in header_names]
        rows.append(
            CustomerRow(
                excel_row=excel_row,
                values=values,
                output_values=output_values,
                key=key,
                material_qty=adjusted_qty,
                material_desc=desc,
            )
        )
        processed = excel_row - header_row
        if job_id and processed % 500 == 0:
            append_job_log(
                job_id,
                f"客户数据清洗中：{processed}/{max(worksheet.max_row - header_row, 0)}，保留 {len(rows)} 行。",
                current_row=processed,
                total_rows=max(worksheet.max_row - header_row, 0),
            )
    return rows, stats


def extract_pp_roll_length(text: str) -> float | None:
    # The leading negative lookbehind prevents matching model text such as NHY3150MP / NY3170MP.
    pattern = re.compile(r"(?<![A-Za-z0-9])(\d+(?:\.\d+)?)\s*[Mm](?=\d*[A-Za-z]?\b|[^A-Za-z0-9]|$)")
    match = pattern.search(str(text or ""))
    return float(match.group(1)) if match else None


def convert_material_desc_to_name(desc: str, transcode_tables: dict | None) -> str:
    text = _text(desc)
    if not text:
        return "无法判断"
    if "半固化片" in text or _looks_like_pp_desc(text):
        return convert_pp_desc_to_name(text)
    if "覆铜板" not in text and not _looks_like_ccl_desc(text):
        return "无法判断"
    if not transcode_tables:
        return "无法判断"
    try:
        engine = _load_transcode_module()
        code, steps, err = engine.transcode_row(text, "", "", "", transcode_tables)
        final_code = _text(steps.get("final_code") if isinstance(steps, dict) else code)
        return final_code if final_code and not err else "无法判断"
    except Exception:
        return "无法判断"


def convert_pp_desc_to_name(desc: str) -> str:
    text = _normalize_compare_text(desc)
    glass = _extract_first(r"\b(10[368]0|106|1078|1080|2116|3313|7628)\b", text)
    rc = _extract_first(r"\bRC\s*(\d+(?:\.\d+)?)", text)
    product = _extract_first(r"\b(NHY?\d+[A-Z0-9-]*|NY[-]?\d+[A-Z0-9-]*)\b", text)
    if not (glass and rc and product):
        return "无法判断"
    # PP/半固化片的客户描述可提取出关键规格，但缺少已确认的内部品名编码公式。
    # 第一版宁可标记无法判断，避免把规则不足误判成品名错误。
    return "无法判断"


def save_result_workbook(result: dict, output_path: Path) -> None:
    workbook = openpyxl.Workbook()
    summary_ws = workbook.active
    summary_ws.title = "核对汇总"
    system_ws = workbook.create_sheet("系统数据核对")
    delivery_ws = workbook.create_sheet("待出货明细")
    customer_ws = workbook.create_sheet("客户数据核对")

    _write_summary_sheet(summary_ws, result["summary"])
    _write_system_sheet(system_ws, result["system_rows"])
    _write_system_sheet(delivery_ws, result["delivery_rows"])
    _write_customer_sheet(customer_ws, result["customer_rows"])
    for ws in workbook.worksheets:
        _style_sheet(ws)
    workbook.save(output_path)


def _write_summary_sheet(ws, summary: dict) -> None:
    rows = [
        ("系统原始数据行", summary["system_source_rows"]),
        ("系统清洗后数据行", summary["system_clean_rows"]),
        ("删除单别单号含341-/219-行", summary["system_removed_341"]),
        ("删除订单数量为空或0行", summary["system_removed_blank_qty"]),
        ("删除厂别为空行", summary["system_removed_blank_factory"]),
        ("客户数据行", summary["customer_rows"]),
        ("客户卷单位行", summary["customer_roll_rows"]),
        ("卷单位未识别长度行", summary["customer_roll_missing_length"]),
        ("匹配成功对数", summary["matched_pairs"]),
        ("数量正确", summary["qty_ok_count"]),
        ("数量错误", summary["qty_error_count"]),
        ("品名正确", summary["name_ok_count"]),
        ("品名错误", summary["name_error_count"]),
        ("品名无法判断", summary["name_unknown_count"]),
        ("日期正确", summary["date_ok_count"]),
        ("日期错误", summary["date_error_count"]),
        ("日期无法判断", summary["date_unknown_count"]),
        ("待出货明细行数", summary["delivery_detail_rows"]),
        ("待出货筛选基准日", summary["delivery_filter_base_date"]),
        ("未匹配或重复键", summary["unmatched_or_duplicate_count"]),
    ]
    ws.append(["指标", "数量"])
    for row in rows:
        ws.append(row)


def _write_system_sheet(ws, rows: list[SystemRow]) -> None:
    ws.append(SYSTEM_KEEP_COLUMNS + ["客户物料数量", "匹配状态", "数量判断"])
    for row in rows:
        ws.append(row.output_values + [_format_number(row.customer_qty), row.match_status, row.qty_judgement])


def _write_customer_sheet(ws, rows: list[CustomerRow]) -> None:
    customer_headers = list(rows[0].values.keys()) if rows else []
    output_headers = []
    product_name_inserted = False
    for header in customer_headers:
        output_headers.append(header)
        if header == "物料说明":
            output_headers.append("品名")
            product_name_inserted = True
    if not product_name_inserted:
        output_headers.append("品名")
    ws.append(
        output_headers
        + [
            "订单数量",
            "数量判断",
            "物料说明转品名",
            "品名判断",
            "系统到货日期",
            "订单承诺日期(年月日)",
            "日期判断",
        ]
    )
    for row in rows:
        source_values = []
        for header in customer_headers:
            source_values.append(row.values.get(header))
            if header == "物料说明":
                source_values.append(row.product_name)
        if not product_name_inserted:
            source_values.append(row.product_name)
        ws.append(
            source_values
            + [
                _format_number(row.order_qty),
                row.qty_judgement,
                row.converted_name,
                row.name_judgement,
                row.system_arrival_date or "",
                row.customer_commitment_date or "",
                row.date_judgement,
            ]
        )


def _delivery_detail_rows(system_rows: list[SystemRow], base_date: date) -> list[SystemRow]:
    direct_dates = {base_date + timedelta(days=1)}
    offset_dates = {base_date + timedelta(days=3)}
    if base_date.weekday() == 4:
        direct_dates = {base_date + timedelta(days=days) for days in (1, 2, 3)}
        offset_dates = {base_date + timedelta(days=days) for days in (3, 4, 5)}

    delivery_rows: list[SystemRow] = []
    for row in system_rows:
        customer_short = _text(row.values.get("客户简称"))
        arrival_date = _as_date(row.values.get("到货日期"))
        if customer_short in ARRIVAL_DIRECT_CUSTOMERS and arrival_date in direct_dates:
            delivery_rows.append(row)
        elif customer_short in ARRIVAL_OFFSET_CUSTOMERS and arrival_date in offset_dates:
            delivery_rows.append(row)
    return delivery_rows


def _delivery_base_date_from_customer_rows(customer_rows: list[CustomerRow]) -> date | None:
    commitment_dates = [
        parsed
        for row in customer_rows
        if (parsed := _as_date(row.values.get("订单承诺日期"))) is not None
    ]
    if not commitment_dates:
        return None
    return min(commitment_dates) - timedelta(days=1)


def _select_named_sheet_with_columns(workbook, sheet_name: str, required_columns: set[str]):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"上传文件必须包含 Sheet：{INTERNAL_DETAIL_SHEET} 和 {CUSTOMER_DETAIL_SHEET}。当前缺少：{sheet_name}")
    worksheet = workbook[sheet_name]
    _find_header_row(worksheet, required_columns)
    return worksheet


def _select_sheet_with_columns(workbook, required_columns: set[str]):
    for worksheet in workbook.worksheets:
        if _find_header_row(worksheet, required_columns, raise_missing=False):
            return worksheet
    raise ValueError(f"未找到包含必要字段的 Sheet：{', '.join(sorted(required_columns))}")


def _find_header_row(worksheet, required_columns: set[str], *, raise_missing: bool = True) -> int | None:
    max_scan = min(worksheet.max_row, 30)
    for row_idx in range(1, max_scan + 1):
        names = {_text(cell.value) for cell in worksheet[row_idx] if _text(cell.value)}
        if required_columns.issubset(names):
            return row_idx
    if raise_missing:
        raise ValueError(f"{worksheet.title} 未找到必要表头：{', '.join(sorted(required_columns))}")
    return None


def _header_occurrences(worksheet, header_row: int) -> dict[str, list[int]]:
    headers: dict[str, list[int]] = {}
    for cell in worksheet[header_row]:
        name = _text(cell.value)
        if name:
            headers.setdefault(name, []).append(cell.column)
    return headers


def _first_header_names(headers: dict[str, list[int]]) -> list[str]:
    ordered = sorted(((cols[0], name) for name, cols in headers.items()), key=lambda item: item[0])
    return [name for _, name in ordered]


def _system_output_specs(headers: dict[str, list[int]]) -> list[tuple[str, int]]:
    occurrences: Counter[str] = Counter()
    specs: list[tuple[str, int]] = []
    for name in SYSTEM_KEEP_COLUMNS:
        if name == "到货日期":
            specs.append((name, 0))
            continue
        occurrences[name] += 1
        if name not in headers or len(headers[name]) < occurrences[name]:
            raise ValueError(f"系统数据缺少必要输出字段：{name}")
        specs.append((name, occurrences[name]))
    return specs


def _output_columns_from_specs(headers: dict[str, list[int]], specs: list[tuple[str, int]]) -> dict[tuple[str, int], int]:
    columns: dict[tuple[str, int], int] = {}
    for name, occurrence in specs:
        if name == "到货日期":
            continue
        columns[(name, occurrence)] = headers[name][occurrence - 1]
    return columns


def _row_values_by_header(worksheet, row_idx: int, headers: dict[str, list[int]]) -> dict[str, Any]:
    values = {}
    for name, columns in headers.items():
        values[name] = worksheet.cell(row=row_idx, column=columns[0]).value
    return values


def _value_for_output_spec(
    worksheet,
    row_idx: int,
    values: dict[str, Any],
    spec: tuple[str, int],
    output_columns: dict[tuple[str, int], int],
) -> Any:
    name, occurrence = spec
    if name == "到货日期":
        return values.get(name)
    column = output_columns[(name, occurrence)]
    return worksheet.cell(row=row_idx, column=column).value


def _arrival_date(pre_date: Any, factory: str, customer_short: str) -> date | str:
    base_date = _as_date(pre_date)
    if not base_date:
        return ""
    offset = None
    if factory == "N3":
        if customer_short in ARRIVAL_DIRECT_CUSTOMERS:
            offset = 0
        elif customer_short in ARRIVAL_OFFSET_CUSTOMERS:
            offset = 2
    elif factory in {"N4", "N5", "N6"}:
        if customer_short in ARRIVAL_DIRECT_CUSTOMERS:
            offset = 2
        elif customer_short in ARRIVAL_OFFSET_CUSTOMERS:
            offset = 1
    if offset is None:
        return ""
    return base_date + timedelta(days=offset)


def _match_partner(key, own_counts, partner_counts, partner_by_key):
    if own_counts[key] > 1:
        return "本表匹配键重复", None
    if partner_counts[key] > 1:
        return "对方匹配键重复", None
    partner = partner_by_key.get(key)
    if not partner:
        return "未匹配", None
    return "匹配成功", partner


def _unique_map(rows, counts):
    return {row.key: row for row in rows if counts[row.key] == 1}


def _qty_judgement(left: float | None, right: float | None) -> str:
    if left is None or right is None:
        return "无法判断"
    return "正确" if abs(float(left) - float(right)) < 0.000001 else "错误"


def _date_judgement(system_date: date | None, customer_date: date | None) -> str:
    if system_date is None or customer_date is None:
        return "无法判断"
    return "正确" if system_date == customer_date else "错误"


def _name_judgement(product_name: str, converted_name: str) -> str:
    if not converted_name or converted_name == "无法判断":
        return "无法判断"
    return "正确" if _normalize_compare_text(product_name) == _normalize_compare_text(converted_name) else "错误"


def _load_transcode_tables() -> dict | None:
    try:
        engine = _load_transcode_module()
        rule_path = get_transcode_rule_file_path(get_active_transcode_rule_version())
        return engine.build_lookup_tables(engine.load_rule_sheets(str(rule_path)))
    except Exception:
        return None


def _load_transcode_module():
    module_name = "fangzheng_web_app.transcode_engine"
    if module_name in sys.modules:
        return importlib.reload(sys.modules[module_name])
    return importlib.import_module(module_name)


def _looks_like_pp_desc(text: str) -> bool:
    normalized = text.upper()
    return " RC" in normalized or "RC" in normalized and any(glass in normalized for glass in ("1080", "2116", "3313", "7628"))


def _looks_like_ccl_desc(text: str) -> bool:
    normalized = text.upper()
    return bool(re.search(r"\b\d+(?:\.\d+)?\s*(?:MM)?\s+[A-Z0-9]+/[A-Z0-9]+", normalized))


def _extract_first(pattern: str, text: str) -> str:
    match = re.search(pattern, text, flags=re.IGNORECASE)
    return match.group(1).upper() if match else ""


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            if 20000 <= float(value) <= 80000:
                parsed = from_excel(value)
                return parsed.date() if isinstance(parsed, datetime) else parsed
        except Exception:
            pass
    text = _text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19] if "%H" in fmt else text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _date_key(value: Any) -> str:
    parsed = _as_date(value)
    return parsed.isoformat() if parsed else _text(value)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> float | None:
    text = _text(value)
    if not text:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def _format_number(value: float | None) -> Any:
    if value is None:
        return ""
    return int(value) if float(value).is_integer() else value


def _normalize_compare_text(value: str) -> str:
    return re.sub(r"\s+", "", _text(value).upper())


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="173F35")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = 8
        column = column_cells[0].column_letter
        for cell in column_cells[:80]:
            value = _text(cell.value)
            if value:
                max_len = max(max_len, min(len(value), 36))
        ws.column_dimensions[column].width = min(max_len + 2, 38)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (date, datetime)):
                cell.number_format = "yyyy-mm-dd"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
