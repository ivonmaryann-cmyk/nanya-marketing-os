from __future__ import annotations

import json
import math
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Iterator

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.utils import secure_filename

from .db import append_job_log, create_job, get_job, update_job_status
from .file_utils import safe_unlink
from .job_control import launch_job_process
from .paths import JOBS_DIR


FEATURE = "inventory_bid"
RULE_VERSION = "库存竞标内置规则 v2"
ALLOWED_EXTENSIONS = {".xls", ".xlsx"}
SUMMARY_MODE = "summary"
MAX_MODE = "max"

REQUIRED_HEADERS = {
    "规格",
    "类别",
    "厚度",
    "铜箔",
    "尺寸",
    "水印",
    "数量",
    "单重",
}

OUTPUT_HEADERS = [
    "规格",
    "类别",
    "厚度",
    "铜箔",
    "尺寸",
    "铜箔类型",
    "单重",
    "江西",
    "上海",
    "总计",
    "竞标含税单价",
]


@dataclass
class BidAggregate:
    spec: str
    category: Any
    thickness: Any
    copper: Any
    size: Any
    copper_type: Any
    unit_weight: Any
    jiangxi_quantity: float = 0.0
    shanghai_quantity: float = 0.0


def queue_inventory_bid_job(employee_id: str, shanghai_file, jiangxi_file) -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    job_dir = employee_dir / f"{timestamp}_inventory_bid"
    job_dir.mkdir(parents=True, exist_ok=True)

    if not shanghai_file or not shanghai_file.filename:
        raise ValueError("请先上传上海库存表")
    if not jiangxi_file or not jiangxi_file.filename:
        raise ValueError("请先上传江西库存表")

    shanghai_path = _save_upload(shanghai_file, job_dir, "shanghai")
    jiangxi_path = _save_upload(jiangxi_file, job_dir, "jiangxi")
    manifest = {
        "feature": FEATURE,
        "mode": SUMMARY_MODE,
        "shanghai_path": str(shanghai_path),
        "jiangxi_path": str(jiangxi_path),
        "shanghai_original_name": shanghai_file.filename,
        "jiangxi_original_name": jiangxi_file.filename,
    }
    manifest_path = job_dir / "inventory_bid_input.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    job_id = create_job(
        employee_id,
        f"库存竞标汇总：{shanghai_file.filename} + {jiangxi_file.filename}",
        str(manifest_path),
        RULE_VERSION,
        feature=FEATURE,
    )
    launch_job_process(job_id, FEATURE, employee_id)
    return job_id


def queue_inventory_bid_max_job(employee_id: str, bid_file) -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)
    job_dir = employee_dir / f"{timestamp}_inventory_bid_max"
    job_dir.mkdir(parents=True, exist_ok=True)

    if not bid_file or not bid_file.filename:
        raise ValueError("请先上传已填写报价的库存竞标汇总表")

    bid_path = _save_upload(bid_file, job_dir, "bid")
    manifest = {
        "feature": FEATURE,
        "mode": MAX_MODE,
        "bid_path": str(bid_path),
        "bid_original_name": bid_file.filename,
    }
    manifest_path = job_dir / "inventory_bid_input.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    job_id = create_job(
        employee_id,
        f"库存竞标取最大值：{bid_file.filename}",
        str(manifest_path),
        RULE_VERSION,
        feature=FEATURE,
    )
    launch_job_process(job_id, FEATURE, employee_id)
    return job_id


def run_inventory_bid_job(job_id: int, employee_id: str) -> None:
    job = get_job(job_id)
    if not job or job["employee_id"] != employee_id:
        return

    update_job_status(job_id, status="running", log_text="")
    try:
        input_manifest_path = Path(job["stored_input_path"])
        manifest = json.loads(input_manifest_path.read_text(encoding="utf-8"))
        job_dir = input_manifest_path.parent
        mode = manifest.get("mode") or SUMMARY_MODE

        if mode == MAX_MODE:
            _run_inventory_bid_max_job(job_id, manifest, job_dir)
        else:
            _run_inventory_bid_summary_job(job_id, manifest, job_dir)
    except Exception as exc:
        append_job_log(job_id, f"库存竞标处理失败：{exc}")
        update_job_status(job_id, status="failed", error_message=str(exc), completed=True)
        raise


def _run_inventory_bid_summary_job(job_id: int, manifest: dict[str, Any], job_dir: Path) -> None:
    shanghai_path = Path(manifest["shanghai_path"])
    jiangxi_path = Path(manifest["jiangxi_path"])
    append_job_log(job_id, "开始处理库存竞标汇总任务。")
    append_job_log(job_id, f"上海库存表：{shanghai_path.name}")
    append_job_log(job_id, f"江西库存表：{jiangxi_path.name}")

    rows, stats = build_inventory_bid_rows(shanghai_path, jiangxi_path, job_id=job_id)
    output_path = job_dir / f"库存竞标汇总_{datetime.now().strftime('%Y%m%d')}.xlsx"
    write_inventory_bid_workbook(rows, output_path)

    update_job_status(
        job_id,
        status="completed",
        stored_result_path=str(output_path),
        success_count=len(rows),
        fail_count=0,
        skip_count=0,
        current_row=stats["read"],
        total_rows=stats["read"],
        completed=True,
    )
    append_job_log(
        job_id,
        f"库存竞标汇总完成：读取 {stats['read']} 行，汇总 {len(rows)} 行，结果文件 {output_path.name}",
    )


def _run_inventory_bid_max_job(job_id: int, manifest: dict[str, Any], job_dir: Path) -> None:
    bid_path = Path(manifest["bid_path"])
    append_job_log(job_id, "开始处理库存竞标报价最大值任务。")
    append_job_log(job_id, f"报价汇总表：{bid_path.name}")
    output_path = job_dir / f"库存竞标最大值_{datetime.now().strftime('%Y%m%d')}.xlsx"
    stats = write_inventory_bid_max_workbook(bid_path, output_path)

    update_job_status(
        job_id,
        status="completed",
        stored_result_path=str(output_path),
        success_count=stats["rows"],
        fail_count=0,
        skip_count=stats["blank_rows"],
        current_row=stats["rows"],
        total_rows=stats["rows"],
        completed=True,
    )
    append_job_log(
        job_id,
        f"库存竞标报价最大值完成：处理 {stats['rows']} 行，空报价 {stats['blank_rows']} 行，结果文件 {output_path.name}",
    )


def build_inventory_bid_rows(
    shanghai_path: str | Path,
    jiangxi_path: str | Path,
    *,
    job_id: int | None = None,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    aggregates: OrderedDict[tuple[str, str, str, str, str, str, str], BidAggregate] = OrderedDict()
    stats = {"read": 0}
    _consume_inventory_file(Path(shanghai_path), "上海", aggregates, stats, job_id)
    _consume_inventory_file(Path(jiangxi_path), "江西", aggregates, stats, job_id)

    rows: list[dict[str, Any]] = []
    for item in aggregates.values():
        jiangxi = _excel_number(item.jiangxi_quantity) if item.jiangxi_quantity else None
        shanghai = _excel_number(item.shanghai_quantity) if item.shanghai_quantity else None
        total = item.jiangxi_quantity + item.shanghai_quantity
        rows.append(
            {
                "规格": item.spec,
                "类别": _number_or_text(item.category),
                "厚度": _number_or_text(item.thickness),
                "铜箔": _number_or_text(item.copper),
                "尺寸": _number_or_text(item.size),
                "铜箔类型": _number_or_text(item.copper_type),
                "单重": _number_or_text(item.unit_weight),
                "江西": jiangxi,
                "上海": shanghai,
                "总计": _excel_number(total) if total else None,
                "竞标含税单价": None,
            }
        )
    rows.sort(key=_bid_sort_key)
    return rows, stats


def write_inventory_bid_workbook(rows: list[dict[str, Any]], output_path: str | Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "B级竞标"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(OUTPUT_HEADERS)

    for row in rows:
        sheet.append([row.get(header) for header in OUTPUT_HEADERS])

    _style_bid_sheet(sheet, len(OUTPUT_HEADERS))

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target


def write_inventory_bid_max_workbook(input_path: str | Path, output_path: str | Path) -> dict[str, int]:
    source = load_workbook(input_path, read_only=True, data_only=True)
    source_sheet = source.worksheets[0]
    rows_iter = source_sheet.iter_rows(values_only=True)
    raw_headers = list(next(rows_iter, ()))
    headers = [_text(header) for header in raw_headers]
    if "竞标含税单价" not in headers:
        source.close()
        raise ValueError("报价汇总表缺少字段：竞标含税单价")
    if "类别" not in headers:
        source.close()
        raise ValueError("报价汇总表缺少字段：类别")

    bid_price_index = headers.index("竞标含税单价")
    quote_headers = [
        header
        for header in headers[bid_price_index + 1 :]
        if header and header not in {"业务员", "最大值"}
    ]
    cleaned_headers = [header for header in headers if header not in {"业务员", "最大值"}]
    category_index = cleaned_headers.index("类别")
    output_headers = cleaned_headers[: category_index + 1] + ["业务员"] + cleaned_headers[category_index + 1 :] + ["最大值"]

    workbook = Workbook()
    output_sheet = workbook.active
    output_sheet.title = source_sheet.title[:31] if source_sheet.title else "B级竞标"
    output_sheet.sheet_view.showGridLines = False
    output_sheet.freeze_panes = "A2"
    output_sheet.append(output_headers)

    input_indexes = {header: index for index, header in enumerate(headers) if header}
    blank_rows = 0
    row_count = 0
    for values in rows_iter:
        row_count += 1
        row_data = {
            header: values[index] if index < len(values) else None
            for header, index in input_indexes.items()
            if header not in {"业务员", "最大值"}
        }
        salesperson, max_value = _max_quote_value(row_data, quote_headers)
        if max_value is None:
            blank_rows += 1
        output_row = []
        for header in output_headers:
            if header == "业务员":
                output_row.append(salesperson)
            elif header == "最大值":
                output_row.append(max_value)
            else:
                output_row.append(row_data.get(header))
        output_sheet.append(output_row)

    source.close()
    _style_bid_sheet(output_sheet, len(output_headers))

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return {"rows": row_count, "blank_rows": blank_rows}


def cleanup_inventory_bid_job_files(job: Any) -> None:
    paths: set[Path] = set()
    for key in ("stored_input_path", "stored_result_path"):
        value = job[key] if job else None
        if value:
            paths.add(Path(value))

    input_path = Path(job["stored_input_path"]) if job and job["stored_input_path"] else None
    if input_path and input_path.exists() and input_path.suffix.lower() == ".json":
        try:
            data = json.loads(input_path.read_text(encoding="utf-8"))
            for key in ("shanghai_path", "jiangxi_path", "bid_path"):
                if data.get(key):
                    paths.add(Path(data[key]))
        except (OSError, json.JSONDecodeError):
            pass

    parent_dirs = {path.parent for path in paths}
    for path in paths:
        safe_unlink(str(path))
    for directory in sorted(parent_dirs, key=lambda item: len(str(item)), reverse=True):
        try:
            directory.rmdir()
        except OSError:
            pass


def _consume_inventory_file(
    path: Path,
    plant: str,
    aggregates: OrderedDict[tuple[str, str, str, str, str, str, str], BidAggregate],
    stats: dict[str, int],
    job_id: int | None,
) -> None:
    headers, source_rows = _iter_sheet_rows(path)
    indexes = _header_indexes(headers)
    missing = sorted(REQUIRED_HEADERS - set(indexes))
    if missing:
        close_iterator = getattr(source_rows, "close", None)
        if callable(close_iterator):
            close_iterator()
        raise ValueError(f"{path.name} 缺少必要字段：{', '.join(missing)}")

    for excel_row, values in source_rows:
        stats["read"] += 1
        spec = _normalize_text(_get(values, indexes, "规格"))
        category = _get(values, indexes, "类别")
        thickness = _get(values, indexes, "厚度")
        copper = _get(values, indexes, "铜箔")
        size = _get(values, indexes, "尺寸")
        copper_type = _get(values, indexes, "水印")
        unit_weight = _get(values, indexes, "单重")
        key = (
            _normalize_text(spec, upper=True),
            _normalize_text(category, compact=True, upper=True),
            _normalize_number(thickness),
            _normalize_text(copper, compact=True, upper=True),
            _normalize_text(size, compact=True, upper=True),
            _normalize_text(copper_type, compact=True, upper=True),
            _normalize_number(unit_weight),
        )
        item = aggregates.get(key)
        if item is None:
            item = BidAggregate(
                spec=spec,
                category=category,
                thickness=thickness,
                copper=copper,
                size=size,
                copper_type=copper_type,
                unit_weight=unit_weight,
            )
            aggregates[key] = item

        quantity = _number(_get(values, indexes, "数量"))
        if plant == "上海":
            item.shanghai_quantity += quantity
        else:
            item.jiangxi_quantity += quantity
        if job_id and stats["read"] % 5000 == 0:
            append_job_log(job_id, f"库存竞标已读取 {stats['read']} 行。")


def _save_upload(file_obj, job_dir: Path, prefix: str) -> Path:
    original_name = (file_obj.filename or "").strip()
    suffix = Path(original_name).suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("库存竞标仅支持 .xls / .xlsx 文件")
    safe_name = secure_filename(Path(original_name).stem) or prefix
    target = job_dir / f"{prefix}_{safe_name}{suffix}"
    file_obj.save(target)
    return target


def _iter_sheet_rows(path: Path) -> tuple[list[Any], Iterator[tuple[int, list[Any]]]]:
    if path.suffix.lower() == ".xls":
        workbook = xlrd.open_workbook(str(path), on_demand=True, formatting_info=False)
        sheet = workbook.sheet_by_index(0)
        headers = sheet.row_values(0)

        def iterator() -> Iterator[tuple[int, list[Any]]]:
            try:
                for index in range(1, sheet.nrows):
                    yield index + 1, sheet.row_values(index)
            finally:
                workbook.release_resources()

        return headers, iterator()

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    value_rows = sheet.iter_rows(values_only=True)
    headers = list(next(value_rows, ()))

    def iterator() -> Iterator[tuple[int, list[Any]]]:
        try:
            for index, values in enumerate(value_rows, start=2):
                yield index, list(values)
        finally:
            workbook.close()

    return headers, iterator()


def _style_bid_sheet(sheet, column_count: int) -> None:
    thin_black = Side(style="thin", color="000000")
    border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
    font = Font(name="宋体", size=11, color="000000")
    header_font = Font(name="宋体", size=11, color="000000")
    white_fill = PatternFill(fill_type=None)
    for row in sheet.iter_rows(min_row=1, max_row=max(sheet.max_row, 1), max_col=column_count):
        for cell in row:
            cell.fill = white_fill
            cell.font = header_font if cell.row == 1 else font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=False)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.######"

    widths = [72, 16, 10, 12, 10, 24, 12, 12, 12, 12, 16]
    for index in range(1, column_count + 1):
        width = widths[index - 1] if index <= len(widths) else 14
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    if sheet.max_row >= 1 and column_count >= 1:
        sheet.auto_filter.ref = f"A1:{sheet.cell(1, column_count).column_letter}{sheet.max_row}"
    sheet.row_dimensions[1].height = 22


def _max_quote_value(row_data: dict[str, Any], quote_headers: list[str]) -> tuple[str | None, int | float | None]:
    best_header: str | None = None
    best_value: float | None = None
    for header in quote_headers:
        number = _number_or_none(row_data.get(header))
        if number is None:
            continue
        if best_value is None or number > best_value:
            best_header = header
            best_value = number
    if best_value is None:
        return None, None
    return best_header, _excel_number(best_value)


def _header_indexes(headers: Iterable[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(headers):
        name = _text(value)
        if name and name not in result:
            result[name] = index
    return result


def _get(values: list[Any], indexes: dict[str, int], name: str) -> Any:
    index = indexes[name]
    return values[index] if index < len(values) else None


def _bid_sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
    thickness_number = _number_or_none(row.get("厚度"))
    return (
        _normalize_text(row.get("类别"), upper=True),
        0 if thickness_number is not None else 1,
        thickness_number if thickness_number is not None else _normalize_text(row.get("厚度"), upper=True),
        _normalize_text(row.get("规格"), upper=True),
        _normalize_number(row.get("单重")),
    )


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_text(value: Any, *, compact: bool = False, upper: bool = False) -> str:
    text = " ".join(_text(value).split())
    if compact:
        text = text.replace(" ", "")
    return text.upper() if upper else text


def _normalize_number(value: Any) -> str:
    number = _number_or_none(value)
    return f"{number:.12g}" if number is not None else _normalize_text(value, compact=True, upper=True)


def _number(value: Any) -> float:
    number = _number_or_none(value)
    return number if number is not None else 0.0


def _number_or_none(value: Any) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    try:
        number = float(str(value).strip().replace(",", ""))
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _number_or_text(value: Any) -> Any:
    number = _number_or_none(value)
    return _excel_number(number) if number is not None else _text(value)


def _excel_number(value: float) -> int | float:
    return int(value) if float(value).is_integer() else round(float(value), 6)
