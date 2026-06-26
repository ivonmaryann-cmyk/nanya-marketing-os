from __future__ import annotations

import importlib
import re
import shutil
import sys
import traceback
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.utils import secure_filename

from .db import append_job_log, create_job, prune_jobs_for_employee, update_job_status
from .excel_utils import load_workbook_compat, normalized_xlsx_source
from .file_utils import safe_unlink
from .job_control import launch_job_process
from .paths import JOBS_DIR
from .rules import get_active_rule_version, load_rule_dataframes
from .shennan_rules import (
    get_active_shennan_rule_version,
    get_shennan_rule_file_path,
    load_shennan_price_dataframe,
    load_shennan_surcharge_rules,
)


CALCULATOR_MODULE_NAME = "fangzheng_web_app.price_calculator_v3"
NON_DATA_DESCRIPTIONS = {"物料描述", "规格", "客户规格", "材料描述", "物料规格"}
SKIP_SHEET_NAMES = {"计算说明", "深南计算说明", "方正价格", "基板对账"}
PRICE_OUTPUT_HEADER = "深南计算价格"
NOTE_SHEET_NAME = "计算说明"


def load_calculator_module():
    """Reload the shared calculator so shennan jobs follow latest formula logic."""
    if CALCULATOR_MODULE_NAME in sys.modules:
        return importlib.reload(sys.modules[CALCULATOR_MODULE_NAME])
    return importlib.import_module(CALCULATOR_MODULE_NAME)


def is_effective_description(value) -> bool:
    text = str(value).strip() if value is not None else ""
    return bool(text and text.lower() not in {"nan", "none"} and text not in NON_DATA_DESCRIPTIONS)


def _iter_shennan_business_sheets(workbook, calculator):
    for sheet in workbook.worksheets:
        if sheet.title in SKIP_SHEET_NAMES:
            continue
        desc_col = calculator.detect_description_column_openpyxl(sheet)
        desc_index = desc_col - 1
        data_rows = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            desc = str(row[desc_index]).strip() if len(row) > desc_index and row[desc_index] is not None else ""
            if is_effective_description(desc):
                data_rows.append((row_index, desc))
        if data_rows:
            yield sheet.title, desc_col, data_rows


def _add_normalized_thickness_aliases(price_df: pd.DataFrame) -> pd.DataFrame:
    """Add in-memory CCL thickness aliases such as 0.80 -> 0.8 for the shared engine."""
    if "不含铜板厚/（mm)" not in price_df.columns:
        return _add_shennan_model_aliases(price_df)
    rows_to_add = []
    key_columns = [col for col in price_df.columns if col != "不含铜板厚/（mm)"]
    existing = {
        tuple(str(row[col]).strip() for col in key_columns) + (str(row["不含铜板厚/（mm)"]).strip(),)
        for _, row in price_df.iterrows()
    }
    ccl_rows = price_df[price_df["CCL"].astype(str).str.strip() == "CCL"]
    for _, row in ccl_rows.iterrows():
        raw = str(row["不含铜板厚/（mm)"]).strip()
        normalized = _format_numeric_text(raw)
        key = tuple(str(row[col]).strip() for col in key_columns) + (normalized,)
        if normalized and normalized != raw and key not in existing:
            alias = row.copy()
            alias["不含铜板厚/（mm)"] = normalized
            rows_to_add.append(alias)
            existing.add(key)
    if not rows_to_add:
        return _add_shennan_model_aliases(price_df)
    return _add_shennan_model_aliases(pd.concat([price_df, pd.DataFrame(rows_to_add)], ignore_index=True))


def _add_shennan_model_aliases(price_df: pd.DataFrame) -> pd.DataFrame:
    if "型号" not in price_df.columns or "CCL" not in price_df.columns:
        return price_df
    rows_to_add = []
    key_columns = list(price_df.columns)
    existing = {tuple(str(row[col]).strip() for col in key_columns) for _, row in price_df.iterrows()}
    for _, row in price_df.iterrows():
        product = str(row["CCL"]).strip()
        if product not in {"CCL", "PP"}:
            continue
        raw_model = str(row["型号"]).strip()
        for alias_model in _shennan_model_aliases(raw_model, product=product):
            if not alias_model or alias_model == raw_model:
                continue
            alias = row.copy()
            alias["型号"] = alias_model
            key = tuple(str(alias[col]).strip() for col in key_columns)
            if key in existing:
                continue
            rows_to_add.append(alias)
            existing.add(key)
    if not rows_to_add:
        return price_df
    return pd.concat([price_df, pd.DataFrame(rows_to_add)], ignore_index=True)


def queue_shennan_job(employee_id: str, uploaded_file, source_filename: str) -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    employee_dir = JOBS_DIR / employee_id
    employee_dir.mkdir(parents=True, exist_ok=True)

    safe_filename = secure_filename(source_filename) or f"shennan_{timestamp}.xlsx"
    input_path = employee_dir / f"{timestamp}_shennan_{safe_filename}"
    uploaded_file.save(input_path)

    rule_version = get_active_shennan_rule_version()
    job_id = create_job(
        employee_id,
        source_filename,
        str(input_path),
        rule_version,
        feature="shennan",
    )
    launch_job_process(job_id, "shennan", employee_id)
    return job_id


def calculate_shennan_quote(spec: str) -> dict:
    spec = str(spec or "").strip()
    if not spec:
        return {"status": "失败", "price": None, "error": "请输入客户规格"}

    calculator = load_calculator_module()
    rule_version = get_active_shennan_rule_version()
    rule_path = get_shennan_rule_file_path(rule_version)
    price_df = _add_normalized_thickness_aliases(load_shennan_price_dataframe(rule_path))
    surcharge_rules = load_shennan_surcharge_rules(rule_path)
    _, account_df = load_rule_dataframes(get_active_rule_version())
    price, note, err, calc_desc = calculate_shennan_price(
        spec,
        price_df,
        account_df,
        calculator,
        surcharge_rules,
    )
    if calc_desc != spec:
        note = f"标准化：{calc_desc} | {note}" if note else f"标准化：{calc_desc}"
    if err:
        return {
            "status": "失败",
            "price": None,
            "note": note or "",
            "material_type": "深南价格",
            "rule_version": rule_version,
            "error": err,
        }
    return {
        "status": "成功",
        "price": round(float(price), 2) if price is not None else None,
        "note": note or "计算成功",
        "material_type": "深南价格",
        "rule_version": rule_version,
        "error": "",
    }


def run_shennan_job(job_id: int, employee_id: str) -> None:
    from .db import get_job

    update_job_status(job_id, status="running", log_text="")
    job = get_job(job_id)
    if not job:
        return

    append_job_log(job_id, f"开始处理深南计算任务，规则版本：{job['rule_version']}")

    try:
        calculator = load_calculator_module()
        calculator_path = Path(calculator.__file__).resolve()
        rule_path = get_shennan_rule_file_path(job["rule_version"])
        price_df = load_shennan_price_dataframe(rule_path)
        raw_price_rows = len(price_df)
        price_df = _add_normalized_thickness_aliases(price_df)
        surcharge_rules = load_shennan_surcharge_rules(rule_path)
        _, account_df = load_rule_dataframes(get_active_rule_version())

        append_job_log(job_id, f"深南报价单加载完成：{rule_path.name}，转换后 {raw_price_rows} 行")
        append_job_log(job_id, f"基板对照表沿用方正规则版本：{get_active_rule_version()}，{len(account_df)} 行")
        append_job_log(
            job_id,
            f"计算引擎已加载：{calculator_path.name}（{datetime.fromtimestamp(calculator_path.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}）",
        )

        workbook = load_workbook_compat(job["stored_input_path"], data_only=True)
        source_for_result = normalized_xlsx_source(job["stored_input_path"], workbook)
        sheet_jobs = list(_iter_shennan_business_sheets(workbook, calculator))
        if not sheet_jobs:
            fallback = calculator.select_calculation_sheet_name(workbook.sheetnames)
            sheet_jobs = [(fallback, calculator.detect_description_column_openpyxl(workbook[fallback]), [])]
        total_rows = sum(len(data_rows) for _, _, data_rows in sheet_jobs)
        update_job_status(job_id, status="running", total_rows=total_rows)
        business_sheet_count = len([item for item in sheet_jobs if item[2]])
        append_job_log(job_id, f"检测到 {business_sheet_count} 个业务 Sheet，共 {total_rows} 行有效数据", total_rows=total_rows)

        results = []
        success_count = 0
        fail_count = 0
        skip_count = 0
        processed = 0

        for sheet_name, desc_col, data_rows in sheet_jobs:
            if not data_rows:
                append_job_log(job_id, f"跳过 Sheet：{sheet_name}，未检测到有效物料描述")
                continue
            append_job_log(job_id, f"开始处理 Sheet：{sheet_name}，物料描述列：第 {desc_col} 列，共 {len(data_rows)} 行")
            for index, desc in data_rows:
                processed += 1
                price, note, err, calc_desc = calculate_shennan_price(
                    desc,
                    price_df,
                    account_df,
                    calculator,
                    surcharge_rules,
                )
                if err:
                    fail_count += 1
                    explain = f"{err}（标准化后：{calc_desc}）" if calc_desc != desc else err
                    results.append({"Sheet": sheet_name, "行号": index, "物料描述": desc, "价格": "", "说明": explain, "状态": "失败"})
                    append_job_log(
                        job_id,
                        f"{sheet_name} 第 {index} 行失败：{explain}",
                        fail_count=fail_count,
                        current_row=processed,
                        total_rows=total_rows,
                    )
                else:
                    success_count += 1
                    explain = f"标准化：{calc_desc} | {note}" if calc_desc != desc else note
                    results.append({"Sheet": sheet_name, "行号": index, "物料描述": desc, "价格": price, "说明": explain, "状态": "成功"})
                    append_job_log(
                        job_id,
                        f"{sheet_name} 第 {index} 行成功：{price}",
                        success_count=success_count,
                        current_row=processed,
                        total_rows=total_rows,
                    )

        input_path = Path(job["stored_input_path"])
        output_path = input_path.with_name(f"{input_path.stem}_深南计算结果.xlsx")
        save_shennan_multi_sheet_result(str(source_for_result), results, str(output_path))
        append_job_log(job_id, "结果文件已生成，任务完成", current_row=total_rows, total_rows=total_rows)

        update_job_status(
            job_id,
            status="completed",
            stored_result_path=str(output_path),
            success_count=success_count,
            fail_count=fail_count,
            skip_count=skip_count,
            current_row=total_rows,
            total_rows=total_rows,
            completed=True,
        )
    except Exception as exc:
        append_job_log(job_id, f"任务失败：{exc}")
        update_job_status(
            job_id,
            status="failed",
            error_message=f"{exc}\n{traceback.format_exc(limit=8)}",
            completed=True,
        )
    finally:
        stale_jobs = prune_jobs_for_employee(employee_id, keep_limit=500)
        for stale in stale_jobs:
            for key in ["stored_input_path", "stored_result_path"]:
                safe_unlink(stale[key])


def calculate_shennan_price(
    desc: str,
    price_df: pd.DataFrame,
    account_df: pd.DataFrame,
    calculator,
    surcharge_rules: dict[str, dict] | None = None,
):
    text = _clean_desc(desc)
    if text.startswith("半固化片") and _is_shennan_roll_pp(text):
        price, note, err, calc_desc = _calculate_shennan_roll_pp(text, price_df)
        return price, note, err, calc_desc

    if text.startswith("覆铜板"):
        price, note, err, calc_desc = _calculate_shennan_ccl(
            text,
            price_df,
            account_df,
            calculator,
            surcharge_rules or {},
        )
        return price, note, err, calc_desc

    calc_desc, normalize_note = _normalize_shennan_description_detail(text, price_df)
    price, note, err = calculator.calculate_price(calc_desc, price_df, account_df)
    if err:
        return price, note, err, calc_desc

    if normalize_note:
        note = f"{normalize_note} | {note}" if note else normalize_note

    surcharge = _calculate_ccl_surcharge(text, calc_desc, surcharge_rules or {})
    if surcharge:
        price = round(float(price) + surcharge["amount"], 2)
        note = (
            f"{note} | 深南加价：{surcharge['foil']} {surcharge['copper_label']}"
            f"{surcharge['side_label']} +{surcharge['per_sf']}/SF × {surcharge['area_sf']}SF"
            f" = {surcharge['amount']:.2f}，合计={price:.2f}"
        )
    return price, note, None, calc_desc


def save_shennan_result(source_path: str, results: list[dict], output_path: str, *, sheet_name: str) -> None:
    """Save Shennan results without overwriting the detected material description column."""
    shutil.copy2(source_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws_calc = wb[sheet_name]

    price_col = _last_value_column(ws_calc) + 1
    header_row = _detect_header_row(ws_calc)
    ws_calc.cell(row=header_row, column=price_col, value="深南计算价格")
    ws_calc.cell(row=header_row, column=price_col).font = Font(bold=True)

    success_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fail_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for item in results:
        if item["状态"] == "跳过":
            continue
        row_num = item["行号"]
        if item["状态"] == "成功" and item["价格"] != "":
            cell = ws_calc.cell(row=row_num, column=price_col, value=item["价格"])
            cell.fill = success_fill
        elif item["状态"] == "失败":
            cell = ws_calc.cell(row=row_num, column=price_col, value="未找到")
            cell.fill = fail_fill

    if "计算说明" in wb.sheetnames:
        del wb["计算说明"]
    ws_note = wb.create_sheet("计算说明")

    headers = ["行号", "物料描述", "计算价格", "计算说明", "状态"]
    for col, header in enumerate(headers, 1):
        cell = ws_note.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    data_row = 2
    for item in results:
        if item["状态"] == "跳过":
            continue
        ws_note.cell(row=data_row, column=1, value=item["行号"])
        ws_note.cell(row=data_row, column=2, value=str(item["物料描述"])[:120])
        ws_note.cell(row=data_row, column=3, value=item["价格"])
        ws_note.cell(row=data_row, column=4, value=str(item["说明"])[:200])
        ws_note.cell(row=data_row, column=5, value=item["状态"])
        if item["状态"] == "成功":
            ws_note.cell(row=data_row, column=5).fill = PatternFill(
                start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"
            )
        elif item["状态"] == "失败":
            ws_note.cell(row=data_row, column=5).fill = PatternFill(
                start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
            )
        data_row += 1

    ws_note.column_dimensions["A"].width = 8
    ws_note.column_dimensions["B"].width = 65
    ws_note.column_dimensions["C"].width = 12
    ws_note.column_dimensions["D"].width = 90
    ws_note.column_dimensions["E"].width = 10
    wb.save(output_path)


def save_shennan_multi_sheet_result(source_path: str, results: list[dict], output_path: str) -> None:
    """Save Shennan results for every detected business sheet."""
    shutil.copy2(source_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    success_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fail_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    by_sheet: dict[str, list[dict]] = {}
    for item in results:
        by_sheet.setdefault(item["Sheet"], []).append(item)

    for sheet_name, sheet_results in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws_calc = wb[sheet_name]
        price_col = _last_value_column(ws_calc) + 1
        header_row = _detect_header_row(ws_calc)
        ws_calc.cell(row=header_row, column=price_col, value=PRICE_OUTPUT_HEADER)
        ws_calc.cell(row=header_row, column=price_col).font = Font(bold=True)
        for item in sheet_results:
            row_num = item["行号"]
            if item["状态"] == "成功" and item["价格"] != "":
                cell = ws_calc.cell(row=row_num, column=price_col, value=item["价格"])
                cell.fill = success_fill
            elif item["状态"] == "失败":
                cell = ws_calc.cell(row=row_num, column=price_col, value="未找到")
                cell.fill = fail_fill

    if NOTE_SHEET_NAME in wb.sheetnames:
        del wb[NOTE_SHEET_NAME]
    ws_note = wb.create_sheet(NOTE_SHEET_NAME)
    headers = ["Sheet", "行号", "物料描述", "计算价格", "计算说明", "状态"]
    for col, header in enumerate(headers, 1):
        cell = ws_note.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    data_row = 2
    for item in results:
        ws_note.cell(row=data_row, column=1, value=item["Sheet"])
        ws_note.cell(row=data_row, column=2, value=item["行号"])
        ws_note.cell(row=data_row, column=3, value=str(item["物料描述"])[:120])
        ws_note.cell(row=data_row, column=4, value=item["价格"])
        ws_note.cell(row=data_row, column=5, value=str(item["说明"])[:300])
        ws_note.cell(row=data_row, column=6, value=item["状态"])
        if item["状态"] == "成功":
            ws_note.cell(row=data_row, column=6).fill = success_fill
        elif item["状态"] == "失败":
            ws_note.cell(row=data_row, column=6).fill = fail_fill
        data_row += 1

    widths = {"A": 18, "B": 8, "C": 65, "D": 12, "E": 90, "F": 10}
    for col, width in widths.items():
        ws_note.column_dimensions[col].width = width
    wb.save(output_path)


def _last_value_column(ws) -> int:
    last_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                last_col = max(last_col, cell.column)
    return last_col


def _detect_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 30) + 1):
        values = {str(cell.value).strip() for cell in ws[row] if cell.value not in (None, "")}
        if "物料描述" in values:
            return row
    return 1


def normalize_shennan_description(desc: str, price_df: pd.DataFrame) -> str:
    return _normalize_shennan_description_detail(desc, price_df)[0]


def _normalize_shennan_description_detail(desc: str, price_df: pd.DataFrame) -> tuple[str, str]:
    text = _clean_desc(desc)
    if text.startswith("半固化片"):
        return _normalize_shennan_pp(text, price_df)
    if text.startswith("覆铜板"):
        return _normalize_shennan_ccl(text, price_df)
    return desc, ""


def _normalize_shennan_pp(desc: str, price_df: pd.DataFrame) -> tuple[str, str]:
    glue = _extract_shennan_glue(desc)
    glass = _match_text(r"\s(\d{3,4})\s+RC\s*\d+", desc)
    rc = _match_text(r"\bRC\s*(\d+)\b", desc)
    if not glue or not glass or not rc:
        return desc, ""
    original_glue = glue
    glue = _resolve_pp_glue(price_df, glue, glass, int(rc))
    rc_value = int(rc)
    pp_rows = price_df[price_df["CCL"].astype(str).str.strip() == "PP"]
    candidates = _filter_pp_candidates(pp_rows, glue, glass)
    matched, rc_note = _match_rc_rows_with_note(candidates, rc_value)
    rc_for_engine = rc_value
    if not matched.empty and rc_note:
        numeric_rc = pd.to_numeric(matched["铜厚"], errors="coerce").dropna()
        if not numeric_rc.empty:
            rc_for_engine = int(numeric_rc.iloc[0])
    alias_note = f"深南PP型号别名：{original_glue}→{glue}" if glue != original_glue else ""

    roll_match = re.search(r"\b(\d+)M\s*(\d+(?:\.\d+)?)I\b", desc, re.IGNORECASE)
    if roll_match:
        length = roll_match.group(1)
        width = roll_match.group(2)
        if width == "49":
            width = "49.5"
        notes = " | ".join(note for note in [alias_note, rc_note] if note)
        return f"PP {glue} {glass} RC{rc_for_engine}% {width}\"*{length}M/Roll", notes

    size = _extract_size(desc)
    if not size:
        return desc, ""
    notes = " | ".join(note for note in [alias_note, rc_note] if note)
    return f"{glue} {glass} RC{rc_for_engine}% {size}", notes


def _is_shennan_roll_pp(desc: str) -> bool:
    return re.search(r"\b\d+M\s*\d+(?:\.\d+)?I\b", desc, re.IGNORECASE) is not None


def _calculate_shennan_roll_pp(desc: str, price_df: pd.DataFrame):
    glue = _extract_shennan_glue(desc)
    glass = _match_text(r"\s(\d{3,4})\s+RC\s*\d+", desc)
    rc_text = _match_text(r"\bRC\s*(\d+)\b", desc)
    roll_match = re.search(r"\b(\d+)M\s*(\d+(?:\.\d+)?)I\b", desc, re.IGNORECASE)
    if not glue or not glass or not rc_text or not roll_match:
        return None, "", "深南卷料 PP 无法提取胶系、玻纤、RC 或卷料规格", desc

    rc = int(rc_text)
    original_glue = glue
    resolved_glue = _resolve_pp_glue(price_df, glue, glass, rc)
    calc_desc = f"PP {resolved_glue} {glass} RC{rc}% {roll_match.group(2)}\"*{roll_match.group(1)}M/Roll"
    pp_rows = price_df[price_df["CCL"].astype(str).str.strip() == "PP"]
    candidates = _filter_pp_candidates(pp_rows, resolved_glue, glass)
    matched, rc_note = _match_rc_rows_with_note(candidates, rc)
    if matched.empty:
        return None, "", f"深南卷料 PP 未找到匹配：胶系={resolved_glue}, 叠构={glass}, RC%={rc}", calc_desc
    numeric_rc = pd.to_numeric(matched["铜厚"], errors="coerce").dropna()
    if not numeric_rc.empty:
        calc_desc = f"PP {resolved_glue} {glass} RC{int(numeric_rc.iloc[0])}% {roll_match.group(2)}\"*{roll_match.group(1)}M/Roll"

    roll_price = matched.iloc[0].get("每卷单价")
    if pd.isna(roll_price) or roll_price in ("", None):
        return None, "", f"深南卷料 PP 匹配行没有每卷单价：胶系={resolved_glue}, 叠构={glass}, RC%={rc}", calc_desc
    price = round(float(roll_price), 2)
    note = (
        f"[深南卷料/PP] 原始胶系={original_glue}→匹配={resolved_glue} | 叠构={glass} | RC%={rc} | "
        f"卷料={roll_match.group(1)}M{roll_match.group(2)}I | 直接取报价单每卷单价={price:.2f}"
    )
    if rc_note:
        note = f"{note} | {rc_note}"
    return price, note, None, calc_desc


def _calculate_shennan_ccl(
    desc: str,
    price_df: pd.DataFrame,
    account_df: pd.DataFrame,
    calculator,
    surcharge_rules: dict[str, dict],
):
    parsed = _parse_shennan_ccl(desc)
    if not parsed:
        return None, "", f"深南CCL无法提取芯厚、铜厚、总厚、尺寸或叠构：{desc}", desc

    original_glue = parsed["glue"]
    glue = _normalize_shennan_glue(original_glue, product="CCL")
    copper = _copper_from_shennan_token(parsed["copper_token"])
    if not copper:
        return None, "", f"深南CCL无法提取铜厚：{desc}", desc

    foil = _base_foil_for_copper_token(parsed["copper_token"]) or _foil_for_glue(price_df, glue) or "HTE"
    requested_laminate = _laminate_from_structure(parsed["structure"])
    if not requested_laminate:
        return None, "", f"深南CCL无法解析叠构：{parsed['structure']}", desc

    selected = _select_shennan_ccl_row(
        price_df,
        glue=glue,
        thickness=parsed["thickness"],
        copper=copper,
        foil=foil,
        requested_laminate=requested_laminate,
    )
    laminate_pattern = requested_laminate.replace("x", "*")
    calc_desc = (
        f"{glue} {parsed['thickness']}mm {copper} ({foil}) "
        f"({laminate_pattern}) {_size_for_engine(parsed['size'], ccl_standard=True)}"
    )
    notes = []
    if glue != original_glue:
        notes.append(f"深南CCL型号别名：{original_glue}→{glue}")
    notes.append(
        f"深南厚度规则：只用芯厚{parsed['core_thickness']}mm查价，总厚{parsed['total_thickness']}mm不参与匹配"
    )

    if not selected:
        detail = " | ".join(notes)
        err = "深南CCL同叠构下未找到可用厚度档"
        return None, detail, f"{err}：{detail}" if detail else err, calc_desc

    laminate = selected["laminate"]
    laminate_pattern = laminate.replace("x", "*")
    calc_desc = (
        f"{glue} {selected['thickness']}mm {selected['copper']} ({selected['foil']}) "
        f"({laminate_pattern}) {_size_for_engine(parsed['size'], ccl_standard=True)}"
    )
    if selected["thickness_note"]:
        notes.append(selected["thickness_note"])

    size = _parse_size_pair(parsed["size"])
    if not size:
        return None, " | ".join(notes), f"深南CCL无法提取尺寸：{parsed['size']}", calc_desc
    width, height = size

    price, size_note, size_ctx, err = _calculate_shennan_ccl_base_price(
        selected,
        width,
        height,
        account_df,
        calculator,
    )
    if err:
        detail = " | ".join(notes)
        return None, detail, f"{err}：{detail}" if detail else err, calc_desc
    notes.append(size_note)

    surcharge = _calculate_ccl_surcharge_for_context(
        parsed,
        glue,
        selected,
        surcharge_rules,
        size_ctx,
        price,
    )
    if surcharge:
        price = round(float(price) + surcharge["amount"], 2)
        notes.append(
            f"深南加价：{surcharge['foil']} {surcharge['copper_label']}{surcharge['side_label']} "
            f"{surcharge['rule_text']} = {surcharge['amount']:.2f}，合计={price:.2f}"
        )

    return round(float(price), 2), " | ".join(notes), None, calc_desc


def _calculate_shennan_ccl_base_price(
    selected: dict,
    width: float,
    height: float,
    account_df: pd.DataFrame,
    calculator,
):
    direct_size_col = _shennan_direct_standard_size_col(width, height)
    if direct_size_col:
        unit = selected["prices"].get(direct_size_col)
        if pd.isna(unit) or unit in ("", None):
            return None, "", {}, f"深南CCL报价行缺少价格列：{direct_size_col}"
        price = round(float(unit), 2)
        ctx = {"size_col": direct_size_col, "multiplier": 1, "qty": 1, "tail_factor": 1.0}
        note = (
            f"[深南CCL报价表标准板尺寸] 胶系={selected['glue']} | 厚度={selected['thickness']}mm | "
            f"铜厚={selected['copper']} | 铜箔={selected['foil']} | 叠构={selected['laminate']} | "
            f"尺寸={_format_size_number(width)}x{_format_size_number(height)} | "
            f"直接取列={direct_size_col} | 价格={price:.2f}"
        )
        return price, note, ctx, None

    narrow_ctx = _shennan_narrow_size_context(width, height)
    if narrow_ctx:
        size_col, factor, narrow_type, numerator, denominator = narrow_ctx
        unit = selected["prices"].get(size_col)
        if pd.isna(unit) or unit in ("", None):
            return None, "", {}, f"深南CCL报价行缺少价格列：{size_col}"
        unit_price = float(unit)
        price = round(unit_price * factor, 2)
        ctx = {"size_col": size_col, "multiplier": 1, "qty": 1, "tail_factor": factor}
        note = (
            f"[深南CCL窄板直算] 胶系={selected['glue']} | 厚度={selected['thickness']}mm | "
            f"铜厚={selected['copper']} | 铜箔={selected['foil']} | 叠构={selected['laminate']} | "
            f"窄板类型={narrow_type} | 尺寸={_format_size_number(width)}x{_format_size_number(height)} | "
            f"取列={size_col} | 标准板单价={unit_price} | "
            f"公式={_format_size_number(numerator)}/{_format_size_number(denominator)}×1.07×{unit_price}={price:.2f}"
        )
        return price, note, ctx, None

    if calculator.is_standard_size(width, height):
        size_col = calculator.get_standard_size_col(width, height)
        unit = selected["prices"].get(size_col)
        if pd.isna(unit) or unit in ("", None):
            return None, "", {}, f"深南CCL报价行缺少价格列：{size_col}"
        price = round(float(unit), 2)
        ctx = {"size_col": size_col, "multiplier": 1, "qty": 1, "tail_factor": 1.0}
        note = (
            f"[深南CCL标准尺寸] 胶系={selected['glue']} | 厚度={selected['thickness']}mm | "
            f"铜厚={selected['copper']} | 铜箔={selected['foil']} | 叠构={selected['laminate']} | "
            f"尺寸={_format_size_number(width)}x{_format_size_number(height)} | 列={size_col} | 价格={price:.2f}"
        )
        return price, note, ctx, None

    big_w, big_h, qty, err = calculator.query_nonstandard_size(account_df, width, height)
    if err:
        return None, "", {}, err
    size_col, multiplier = _shennan_price_col_from_big_width(big_w)
    if not size_col:
        return None, "", {}, f"大板 {big_w}x{big_h} 无法按宽度确定价格列"

    unit = selected["prices"].get(size_col)
    if pd.isna(unit) or unit in ("", None):
        return None, "", {}, f"深南CCL报价行缺少价格列：{size_col}"

    tail_factor, tail_note = _shennan_tail_factor(height, big_h)
    base_price = float(unit) * multiplier / float(qty)
    price = round(base_price * tail_factor, 2)
    ctx = {
        "size_col": size_col,
        "multiplier": multiplier,
        "qty": float(qty),
        "tail_factor": tail_factor,
    }
    note = (
        f"[深南CCL非标准尺寸] 胶系={selected['glue']} | 厚度={selected['thickness']}mm | "
        f"铜厚={selected['copper']} | 铜箔={selected['foil']} | 叠构={selected['laminate']} | "
        f"小片{_format_size_number(width)}x{_format_size_number(height)} | 大板{big_w}x{big_h}(共{qty}片) | "
        f"按大板宽度取列={size_col} | 单价={float(unit)} | 倍率={multiplier} | "
        f"{multiplier}×{float(unit)}/{qty}={round(base_price, 2)}"
    )
    if tail_note:
        note += f" | {tail_note} | 修正后={price:.2f}"
    return price, note, ctx, None


def _normalize_shennan_ccl(desc: str, price_df: pd.DataFrame) -> tuple[str, str]:
    parsed = _parse_shennan_ccl(desc)
    if not parsed:
        return desc, ""

    original_glue = parsed["glue"]
    glue = _normalize_shennan_glue(original_glue, product="CCL")
    copper = _copper_from_shennan_token(parsed["copper_token"])
    if not copper:
        return desc, ""

    foil = _base_foil_for_copper_token(parsed["copper_token"]) or _foil_for_glue(price_df, glue)
    requested_laminate = _laminate_from_structure(parsed["structure"])
    selected = _select_shennan_ccl_row(
        price_df,
        glue=glue,
        thickness=parsed["thickness"],
        copper=copper,
        foil=foil,
        requested_laminate=requested_laminate,
    )
    if not selected:
        if not requested_laminate:
            return desc, ""
        laminate_pattern = requested_laminate.replace("x", "*")
        calc_desc = (
            f"{glue} {parsed['thickness']}mm {copper} ({foil}) "
            f"({laminate_pattern}) {_size_for_engine(parsed['size'], ccl_standard=True)}"
        )
        notes = []
        if glue != original_glue:
            notes.append(f"深南CCL型号别名：{original_glue}→{glue}")
        notes.append("深南CCL已完成厚度/铜厚/叠构解析，但报价表未找到可用厚度档")
        return calc_desc, " | ".join(notes)

    laminate = selected["laminate"]
    laminate_pattern = laminate.replace("x", "*")
    notes = []
    if glue != original_glue:
        notes.append(f"深南CCL型号别名：{original_glue}→{glue}")
    notes.append(
        f"深南厚度规则：只用芯厚{parsed['core_thickness']}mm查价，总厚{parsed['total_thickness']}mm不参与匹配"
    )
    if selected["thickness_note"]:
        notes.append(selected["thickness_note"])
    if requested_laminate and laminate != requested_laminate:
        notes.append(f"叠构按报价表匹配：{requested_laminate}→{laminate}")
    calc_desc = (
        f"{glue} {selected['thickness']}mm {selected['copper']} ({selected['foil']}) "
        f"({laminate_pattern}) {_size_for_engine(parsed['size'], ccl_standard=True)}"
    )
    return calc_desc, " | ".join(notes)


def _extract_shennan_glue(desc: str) -> str:
    matches = re.findall(r"\bNY[\w\-.]+(?:\([A-Za-z]\))?", desc)
    return matches[-1] if matches else ""


def _extract_size(desc: str) -> str:
    match = re.search(r"(\d+(?:\.\d+)?)\s*[xX×]\s*(\d+(?:\.\d+)?)", desc)
    if not match:
        return ""
    return f"{match.group(1)}\"X{match.group(2)}\""


def _match_text(pattern: str, desc: str) -> str:
    match = re.search(pattern, desc, re.IGNORECASE)
    return match.group(1) if match else ""


def _clean_desc(desc: str) -> str:
    text = re.sub(r"\s+", " ", str(desc).replace("\xa0", " ").replace("\u3000", " ")).strip()
    text = re.sub(r"^(?:CCL)\b", "覆铜板", text, flags=re.IGNORECASE)
    text = re.sub(r"^(?:Prepreg|PP)\b", "半固化片", text, flags=re.IGNORECASE)
    return text


def _parse_shennan_ccl(desc: str) -> dict | None:
    glue = _extract_shennan_glue(desc)
    match = re.search(
        r"覆铜板(?:[（(][^）)]*[）)])?\s+\S+\s+"
        r"(?P<core>\d+(?:\.\d+)?)\s+"
        r"(?P<copper>\S+)\s+"
        r"(?P<total>\d+(?:\.\d+)?)\s+"
        r"(?P<size>\d+(?:\.\d+)?\s*[xX×]\s*\d+(?:\.\d+)?)(?:\([^)]*\))?\s+"
        r"(?P<structure>\S+)",
        desc,
        re.IGNORECASE,
    )
    if not glue or not match:
        return None
    core = float(match.group("core"))
    total = float(match.group("total"))
    return {
        "glue": glue,
        "core_thickness": _format_size_number(core),
        "total_thickness": _format_size_number(total),
        "thickness": _format_size_number(core),
        "thickness_source": "core",
        "copper_token": match.group("copper"),
        "size": match.group("size"),
        "structure": match.group("structure"),
    }


def _normalize_shennan_glue(glue: str, *, product: str) -> str:
    return _canonical_shennan_model(glue, product=product)


def _clean_shennan_model_text(glue: str) -> str:
    text = str(glue).strip().upper().replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", "", text)


def _strip_pp_suffix_for_model_key(model: str, *, product: str) -> str:
    if product != "PP":
        return model
    return model[:-1] if model.endswith("P") else model


def _canonical_shennan_model(glue: str, *, product: str) -> str:
    clean = _clean_shennan_model_text(glue)
    if product == "PP":
        base = re.sub(r"\(C\)$", "", clean)
        base = _strip_pp_suffix_for_model_key(base, product=product)
        base = re.sub(r"\(C\)$", "", base)
    else:
        base = re.sub(r"\(C\)$", "", clean)
    canonical_base = {
        "NY-A1": "NY2150",
        "NY2150": "NY2150",
        "NY6300": "NY6300",
        "NY-P1": "NY-P1",
        "NY-P2": "NY-P2",
        "NY-P3": "NY-P3",
        "NY-P4": "NY-P4",
    }.get(base, base)
    return f"{canonical_base}P" if product == "PP" else canonical_base


def _shennan_model_aliases(glue: str, *, product: str) -> set[str]:
    clean = _clean_shennan_model_text(glue)
    aliases = {str(glue).strip(), clean}
    canonical = _canonical_shennan_model(glue, product=product)
    aliases.add(canonical)
    if product == "PP":
        for base in _equivalent_shennan_base_models(canonical[:-1] if canonical.endswith("P") else canonical):
            aliases.add(f"{base}P")
            if base.endswith("(C)"):
                aliases.add(f"{base[:-3]}P(C)")
    else:
        aliases.update(_equivalent_shennan_base_models(canonical))
    return {alias for alias in aliases if alias}


def _equivalent_shennan_base_models(canonical_base: str) -> set[str]:
    groups = [
        {"NY6300", "NY6300(C)"},
        {"NY-P1", "NY-P1(C)"},
        {"NY-P2", "NY-P2(C)"},
        {"NY-P3", "NY-P3(C)"},
        {"NY-P4", "NY-P4(C)"},
        {"NY2150", "NY-A1"},
    ]
    for group in groups:
        if canonical_base in group:
            return set(group)
    return {canonical_base}


def _filter_shennan_model_rows(rows: pd.DataFrame, glue: str, *, product: str) -> pd.DataFrame:
    model_key = _canonical_shennan_model(glue, product=product)
    model_values = rows["型号"].astype(str).map(lambda value: _canonical_shennan_model(value, product=product))
    return rows[model_values == model_key].copy()


def _select_display_glue(rows: pd.DataFrame, requested_glue: str, *, product: str) -> str:
    if rows.empty:
        return _normalize_shennan_glue(requested_glue, product=product)
    requested_clean = _clean_shennan_model_text(requested_glue)
    model_values = rows["型号"].astype(str).str.strip()
    exact = model_values[model_values.map(_clean_shennan_model_text) == requested_clean]
    if not exact.empty:
        return str(exact.iloc[0]).strip()
    return str(model_values.iloc[0]).strip()


def _select_shennan_ccl_row(
    price_df: pd.DataFrame,
    *,
    glue: str,
    thickness: str,
    copper: str,
    foil: str,
    requested_laminate: str,
) -> dict | None:
    ccl_rows = price_df[(price_df["CCL"].astype(str).str.strip() == "CCL")].copy()
    ccl_rows = _filter_shennan_model_rows(ccl_rows, glue, product="CCL")
    if ccl_rows.empty:
        return None

    copper_candidates = {copper, _copper_fallback(copper)}
    rows = ccl_rows[
        ccl_rows["铜厚"].astype(str).str.strip().isin(copper_candidates)
        & (ccl_rows["铜箔"].astype(str).str.strip() == foil)
    ].copy()
    if rows.empty and foil != "HTE":
        rows = ccl_rows[
            ccl_rows["铜厚"].astype(str).str.strip().isin(copper_candidates)
            & (ccl_rows["铜箔"].astype(str).str.strip() == "HTE")
        ].copy()
    if rows.empty:
        return None

    if requested_laminate:
        laminate_values = rows["叠构"].astype(str).map(_normalize_laminate_key)
        rows = rows[laminate_values == _normalize_laminate_key(requested_laminate)].copy()
        if rows.empty:
            return None

    selected_thickness, thickness_note = _select_shennan_numeric_slot(
        rows["不含铜板厚/（mm)"],
        float(thickness),
        label="厚度",
        unit="mm",
    )
    if not selected_thickness:
        return None

    thickness_rows = rows[rows["不含铜板厚/（mm)"].astype(str).str.strip() == selected_thickness]
    if thickness_rows.empty:
        return None

    row = thickness_rows.iloc[0]
    return {
        "glue": str(row["型号"]).strip(),
        "thickness": str(row["不含铜板厚/（mm)"]).strip(),
        "copper": str(row["铜厚"]).strip(),
        "foil": str(row["铜箔"]).strip(),
        "laminate": str(row["叠构"]).strip(),
        "thickness_note": thickness_note,
        "prices": {
            '36"*48"': row.get('36"*48"'),
            '40"*48"': row.get('40"*48"'),
            '42"*48"': row.get('42"*48"'),
        },
    }


def _select_shennan_numeric_slot(values: pd.Series, target: float, *, label: str, unit: str = "") -> tuple[str, str]:
    numeric = pd.to_numeric(values, errors="coerce").dropna()
    if numeric.empty:
        return "", ""
    exact = numeric[numeric == target]
    if not exact.empty:
        idx = exact.index[0]
        return str(values.loc[idx]).strip(), ""
    upper = numeric[numeric >= target]
    if not upper.empty:
        idx = upper.sort_values().index[0]
        chosen = float(numeric.loc[idx])
        suffix = unit if unit else ""
        return str(values.loc[idx]).strip(), f"深南{label}取档：{_format_numeric_text(target)}{suffix}未精确命中，向上取{_format_numeric_text(chosen)}{suffix}"
    suffix = unit if unit else ""
    return "", f"深南{label}取档：{_format_numeric_text(target)}{suffix}无更大档，按深南严格规则失败"


def _format_numeric_text(value) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    return f"{number:.4f}".rstrip("0").rstrip(".")


def _normalize_laminate_key(value: str) -> str:
    text = str(value).strip().replace("×", "x").replace("*", "x").replace("X", "x")
    text = re.sub(r"\s+", "", text).lower()
    canonical = _canonical_laminate_key(text)
    return canonical or text


def _canonical_laminate_key(text: str) -> str:
    counts: dict[str, int] = {}
    for term in str(text).split("+"):
        if not term:
            return ""
        match = re.fullmatch(r"(\d+)\s*x\s*(\d+)", term)
        if not match:
            return ""
        left, right = match.group(1), match.group(2)
        if len(left) >= 3:
            glass, count = left, right
        elif len(right) >= 3:
            glass, count = right, left
        else:
            return ""
        counts[glass] = counts.get(glass, 0) + int(count)
    if not counts:
        return ""
    return "+".join(f"{glass}x{counts[glass]}" for glass in sorted(counts, key=lambda item: int(item)))


def _copper_fallback(copper: str) -> str:
    return {"H/1": "1/1", "H/2": "2/2", "1/2": "2/2"}.get(copper, copper)


def _resolve_pp_glue(price_df: pd.DataFrame, glue: str, glass: str, rc: int) -> str:
    pp_rows = price_df[price_df["CCL"].astype(str).str.strip() == "PP"]

    def matched_rows(candidate_glue: str) -> pd.DataFrame:
        candidates = _filter_pp_candidates(pp_rows, candidate_glue, glass)
        return _match_rc_rows(candidates, rc)

    aliased = _normalize_shennan_glue(glue, product="PP")
    if aliased != glue:
        matched = matched_rows(aliased)
        if not matched.empty:
            return _select_display_glue(matched, glue, product="PP")
    matched = matched_rows(glue)
    if not matched.empty:
        return _select_display_glue(matched, glue, product="PP")
    if glue.endswith("P"):
        stripped = glue[:-1]
        matched = matched_rows(stripped)
        if not matched.empty:
            return _select_display_glue(matched, glue, product="PP")
    if aliased != glue:
        return aliased
    return glue


def _filter_pp_candidates(pp_rows: pd.DataFrame, glue: str, glass: str) -> pd.DataFrame:
    rows = _filter_shennan_model_rows(pp_rows, glue, product="PP")
    if rows.empty:
        return rows
    glass_text = str(glass).strip()
    glass_values = rows["不含铜板厚/（mm)"].astype(str).str.strip()
    exact = rows[glass_values == glass_text]
    if not exact.empty:
        return exact
    return rows[glass_values.apply(lambda value: _glass_matches(value, glass_text))]


def _glass_matches(rule_glass: str, requested_glass: str) -> bool:
    parts = [part.strip() for part in re.split(r"[/、,，]", str(rule_glass)) if part.strip()]
    return requested_glass in parts


def _match_rc_rows(candidates: pd.DataFrame, rc_value: int) -> pd.DataFrame:
    return _match_rc_rows_with_note(candidates, rc_value)[0]


def _match_rc_rows_with_note(candidates: pd.DataFrame, rc_value: int) -> tuple[pd.DataFrame, str]:
    if candidates.empty:
        return candidates, ""
    numeric_rc = pd.to_numeric(candidates["铜厚"], errors="coerce")
    exact = candidates[numeric_rc == rc_value]
    if not exact.empty:
        return exact, ""
    for idx, row in candidates.iterrows():
        rc_text = str(row["铜厚"]).strip()
        if _rc_in_range(rc_text, rc_value):
            return candidates.loc[[idx]], ""

    valid = numeric_rc.dropna()
    if valid.empty:
        return candidates.iloc[0:0], ""
    upper = valid[valid >= rc_value]
    if not upper.empty:
        chosen = upper.sort_values().index[0]
        chosen_rc = int(numeric_rc.loc[chosen])
        return candidates.loc[[chosen]], f"深南RC取档：RC{rc_value}未精确命中，向上取RC{chosen_rc}"
    return candidates.iloc[0:0], f"深南RC取档：RC{rc_value}无更大档，按深南严格规则失败"


def _rc_in_range(rc_text: str, rc_value: int) -> bool:
    text = rc_text.replace("＞", ">").replace("＜", "<").replace("％", "%").replace("%", "").strip()
    match = re.match(r"(\d+)\s*-\s*(\d+)", text)
    if match:
        return int(match.group(1)) <= rc_value <= int(match.group(2))
    match = re.match(r"[≥≧]+\s*(\d+)|>=\s*(\d+)", text)
    if match:
        return rc_value >= int(match.group(1) or match.group(2))
    match = re.match(r"[≤≦]+\s*(\d+)|<=\s*(\d+)", text)
    if match:
        return rc_value <= int(match.group(1) or match.group(2))
    match = re.match(r">\s*(\d+)", text)
    if match:
        return rc_value > int(match.group(1))
    match = re.match(r"<\s*(\d+)", text)
    if match:
        return rc_value < int(match.group(1))
    return False


def _size_for_engine(size: str, *, ccl_standard: bool = False) -> str:
    match = re.match(r"(\d+(?:\.\d+)?)\s*[xX×]\s*(\d+(?:\.\d+)?)", size)
    if not match:
        return size
    width = float(match.group(1))
    height = float(match.group(2))
    if ccl_standard and round(height) == 48 and round(width) in {36, 40, 42}:
        width += 1
        height += 1
    return f"{_format_size_number(width)}\"X{_format_size_number(height)}\""


def _format_size_number(value: float) -> str:
    return f"{value:.4f}".rstrip("0").rstrip(".")


def _copper_from_shennan_token(token: str) -> str:
    parts = str(token).upper().split("/")
    if len(parts) != 2:
        side = _copper_side(str(token))
        return side if side else ""
    front = _copper_side(parts[0])
    back = _copper_side(parts[1])
    return f"{front}/{back}" if front and back else ""


def _base_foil_for_copper_token(token: str) -> str:
    details = [_foil_detail_from_side(part) for part in str(token).upper().split("/")]
    if any(detail.startswith("RTF") for detail in details):
        return "RTF"
    if any(detail.startswith("HVLP") for detail in details):
        return "HTE"
    if any(detail == "VLP" for detail in details):
        return "HTE"
    return ""


def _foil_detail_from_shennan_token(token: str) -> str:
    details = [_foil_detail_from_side(part) for part in str(token).upper().split("/")]
    details = [detail for detail in details if detail]
    if not details:
        return ""
    if "RTF3" in details:
        return "RTF3"
    if "RTF2" in details:
        return "RTF2"
    if "RTF" in details:
        return "RTF"
    if any(detail.startswith("HVLP") for detail in details):
        return sorted(detail for detail in details if detail.startswith("HVLP"))[-1]
    return details[0]


def _foil_detail_from_side(side: str) -> str:
    clean = re.sub(r"[^A-Z0-9]", "", side.upper())
    if clean.startswith("R"):
        number = re.search(r"R(\d)", clean)
        if not number or number.group(1) == "1":
            return "RTF"
        return f"RTF{number.group(1)}"
    if clean.startswith("HV"):
        number = re.search(r"HV(\d)", clean)
        return f"HVLP{number.group(1)}" if number else "HVLP1"
    if clean.startswith("V"):
        return "VLP"
    if clean.startswith("S"):
        return "RTF"
    return ""


def _copper_side(part: str) -> str:
    clean = re.sub(r"[^A-Z0-9]", "", part.upper())
    if clean in {"H", "1", "2", "3"}:
        return clean
    prefixed = re.match(r"^(?:R\d?|S\d?|HV\d?|V)(H|[123])$", clean)
    if prefixed:
        return prefixed.group(1)
    if clean.endswith("H"):
        return "H"
    digit = re.search(r"([123])$", clean)
    if digit:
        return digit.group(1)
    return ""


def _foil_lookup_key(foil_detail: str) -> str:
    if foil_detail == "RTF1":
        return "RTF"
    return foil_detail


def _foil_for_glue(price_df: pd.DataFrame, glue: str) -> str:
    rows = price_df[price_df["CCL"].astype(str).str.strip() == "CCL"].copy()
    rows = _filter_shennan_model_rows(rows, glue, product="CCL")
    foils = [str(value).strip() for value in rows["铜箔"].dropna().unique()]
    if len(foils) == 1:
        return foils[0]
    if "RTF" in foils and "HTE" not in foils:
        return "RTF"
    return "HTE"


def _laminate_from_structure(structure: str) -> str:
    match = re.match(r"(\d+)\s*[xX×]\s*(\d{3,4})", structure)
    if match:
        count = match.group(1)
        glass = match.group(2)
        return f"{glass}x{count}"

    parsed = []
    text = structure.upper().strip()
    for count, code in re.findall(r"(\d*)(LVW|[A-Z])", text):
        glass = GLASS_CODE_MAP.get(code)
        if not glass:
            return ""
        parsed.append(f"{glass}x{int(count) if count else 1}")
    if not parsed:
        return ""
    return "+".join(parsed)


def _parse_size_pair(size: str) -> tuple[float, float] | None:
    match = re.match(r"(\d+(?:\.\d+)?)\s*[xX×]\s*(\d+(?:\.\d+)?)", str(size))
    if not match:
        return None
    return float(match.group(1)), float(match.group(2))


def _shennan_direct_standard_size_col(width: float, height: float) -> str | None:
    if abs(float(height) - 48) > 0.25:
        return None
    for target_width, column in [(36, '36"*48"'), (40, '40"*48"'), (42, '42"*48"')]:
        if abs(float(width) - target_width) <= 0.25:
            return column
    return None


def _shennan_narrow_size_context(width: float, height: float) -> tuple[str, float, str, float, float] | None:
    width = float(width)
    height = float(height)
    if abs(width - 32) <= 0.25 and abs(height - 48) <= 0.25:
        return '36"*48"', width / 36 * 1.07, "经向窄板", width, 36
    if abs(height - 42) <= 0.25:
        if abs(width - 36) <= 0.25:
            return '36"*48"', height / 48 * 1.07, "纬向窄板", height, 48
        if abs(width - 40) <= 0.25:
            return '40"*48"', height / 48 * 1.07, "纬向窄板", height, 48
    return None


def _shennan_price_col_from_big_width(big_w: float) -> tuple[str | None, int | None]:
    for target, column, multiplier in [
        (37, '36"*48"', 1),
        (41, '40"*48"', 1),
        (43, '42"*48"', 1),
        (74, '36"*48"', 2),
        (82, '40"*48"', 2),
        (86, '42"*48"', 2),
    ]:
        if abs(float(big_w) - target) <= 1.5:
            return column, multiplier
    return None, None


def _shennan_tail_factor(piece_h: float, big_h: float) -> tuple[float, str]:
    count = max(1, int(float(big_h) / float(piece_h) + 0.5))
    packed_h = float(piece_h) * count
    if packed_h >= 47.999:
        return 1.0, ""
    factor = packed_h / 48 * 1.07
    note = (
        f"尾板修正：拼板高度={_format_size_number(piece_h)}×{count}="
        f"{_format_size_number(packed_h)}，系数={_format_size_number(packed_h)}/48×1.07={factor:.6f}"
    )
    return factor, note


def _allocated_area_sf(size_col: str, multiplier: float, qty: float, tail_factor: float) -> float:
    base_area = {'36"*48"': 12.0, '40"*48"': 40 * 48 / 144, '42"*48"': 14.0}.get(size_col, 0.0)
    if not base_area or not qty:
        return 0.0
    return base_area * float(multiplier) / float(qty) * float(tail_factor)


def _calculate_ccl_surcharge_for_context(
    parsed: dict,
    glue: str,
    selected: dict,
    surcharge_rules: dict[str, dict],
    size_ctx: dict,
    base_price: float,
):
    foil_detail = _foil_detail_from_shennan_token(parsed["copper_token"])
    rule = surcharge_rules.get(glue, {}).get(foil_detail)
    if not rule and foil_detail == "RTF":
        rule = surcharge_rules.get(glue, {}).get("RTF1")
    if not rule and foil_detail == "RTF1":
        rule = surcharge_rules.get(glue, {}).get("RTF")
    if not rule:
        return None

    copper_label = _surcharge_copper_label(selected["copper"])
    side_label = _surcharge_side_label(selected["copper"])
    if rule.get("type") == "percent":
        percent = float(rule.get("percent", 0))
        amount = round(float(base_price) * percent, 4)
        return {
            "foil": foil_detail,
            "copper_label": _display_copper_label(copper_label),
            "side_label": "单面" if side_label == "single" else "双面",
            "rule_text": f"+{percent:.0%} × {float(base_price):.2f}",
            "amount": amount,
        }

    per_sf = rule.get(copper_label, {}).get(side_label)
    if per_sf is None:
        return None
    area_sf = _allocated_area_sf(
        size_ctx.get("size_col", ""),
        size_ctx.get("multiplier", 1),
        size_ctx.get("qty", 1),
        size_ctx.get("tail_factor", 1.0),
    )
    amount = round(float(per_sf) * area_sf, 4)
    return {
        "foil": foil_detail,
        "copper_label": _display_copper_label(copper_label),
        "side_label": "单面" if side_label == "single" else "双面",
        "rule_text": f"+{float(per_sf)}/SF × {area_sf:.4f}SF",
        "amount": amount,
    }


def _display_copper_label(copper_label: str) -> str:
    return {"H": "HOZ", "1": "1OZ", "2": "2OZ"}.get(copper_label, copper_label)


GLASS_CODE_MAP = {
    "A": "106",
    "B": "1065",
    "C": "1067",
    "D": "1078",
    "E": "1080",
    "F": "1086",
    "G": "2112",
    "H": "2113",
    "I": "2313",
    "J": "3313",
    "K": "2116",
    "L": "2165",
    "M": "1500",
    "N": "1501",
    "O": "1504",
    "P": "1506",
    "Q": "1652",
    "R": "6700",
    "S": "7627",
    "T": "7628",
    "U": "7629",
    "V": "7630",
    "Z": "1037",
    "LVW": "1035",
}


def _calculate_ccl_surcharge(desc: str, calc_desc: str, surcharge_rules: dict[str, dict]):
    if not desc.startswith("覆铜板"):
        return None
    parsed = _parse_shennan_ccl(desc)
    if not parsed:
        return None

    glue = _normalize_shennan_glue(parsed["glue"], product="CCL")
    copper_token = parsed["copper_token"]
    original_size = parsed["size"]
    foil_detail = _foil_detail_from_shennan_token(copper_token)
    rule = surcharge_rules.get(glue, {}).get(foil_detail)
    if not rule and foil_detail == "RTF":
        rule = surcharge_rules.get(glue, {}).get("RTF1")
    if not rule and foil_detail == "RTF1":
        rule = surcharge_rules.get(glue, {}).get("RTF")
    if not rule:
        return None

    copper = _copper_from_shennan_token(copper_token)
    copper_label = _surcharge_copper_label(copper)
    side_label = _surcharge_side_label(copper)
    per_sf = rule.get(copper_label, {}).get(side_label)
    if per_sf is None:
        return None

    area_sf = _standard_area_sf(original_size)
    amount = round(float(per_sf) * area_sf, 4)
    return {
        "foil": foil_detail,
        "copper_label": "HOZ" if copper_label == "H" else "1OZ",
        "side_label": "单面" if side_label == "single" else "双面",
        "per_sf": float(per_sf),
        "area_sf": area_sf,
        "amount": amount,
    }


def _surcharge_copper_label(copper: str) -> str:
    sides = [side for side in str(copper).upper().split("/") if side and side != "0"]
    if any(side == "2" for side in sides):
        return "2"
    if any(side == "1" for side in sides):
        return "1"
    return "H"


def _surcharge_side_label(copper: str) -> str:
    sides = [side for side in str(copper).upper().split("/") if side]
    non_zero = [side for side in sides if side != "0"]
    return "single" if len(non_zero) == 1 else "double"


def _standard_area_sf(size: str) -> float:
    match = re.match(r"(\d+(?:\.\d+)?)\s*[xX×]\s*(\d+(?:\.\d+)?)", size)
    if not match:
        return 0.0
    width = round(float(match.group(1)))
    height = round(float(match.group(2)))
    if height == 48 and width == 36:
        return 12.0
    if height == 48 and width == 40:
        return 13.33
    if height == 48 and width == 42:
        return 14.0
    normalized_width = width + 1 if height == 48 and width in {36, 40, 42} else width
    normalized_height = height + 1 if height == 48 and width in {36, 40, 42} else height
    return round(normalized_width * normalized_height / 144, 2)


def _choose_laminate(price_df: pd.DataFrame, glue: str, thickness: str, copper: str, foil: str) -> str:
    ccl_rows = price_df[price_df["CCL"].astype(str).str.strip() == "CCL"].copy()
    ccl_rows = _filter_shennan_model_rows(ccl_rows, glue, product="CCL")
    if ccl_rows.empty:
        return ""

    copper_candidates = {copper}
    copper_candidates.update({"H/1": "1/1", "H/2": "2/2", "1/2": "2/2"}.get(copper, copper) for _ in [0])
    candidates = ccl_rows[
        ccl_rows["铜厚"].astype(str).str.strip().isin(copper_candidates)
        & (ccl_rows["铜箔"].astype(str).str.strip() == foil)
    ].copy()
    if candidates.empty:
        candidates = ccl_rows[ccl_rows["铜厚"].astype(str).str.strip().isin(copper_candidates)].copy()
    if candidates.empty:
        candidates = ccl_rows.copy()

    thickness_values = pd.to_numeric(candidates["不含铜板厚/（mm)"], errors="coerce")
    target = float(thickness)
    exact = candidates[thickness_values == target]
    if not exact.empty:
        return str(exact.iloc[0]["叠构"]).strip()

    valid = thickness_values.dropna()
    if valid.empty:
        return str(candidates.iloc[0]["叠构"]).strip()
    nearest_index = (valid - target).abs().sort_values().index[0]
    return str(candidates.loc[nearest_index, "叠构"]).strip()
