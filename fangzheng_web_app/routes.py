from __future__ import annotations

import hmac
import json
import os
import secrets
import sqlite3
from datetime import date
from pathlib import Path

from flask import Blueprint, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import load_workbook

PDF_EXCEL_FEATURE = "pdf_excel"

from .bomin_rules import (
    get_active_bomin_rule_version,
    get_bomin_rule_file_path,
    get_bomin_rule_history,
    save_new_bomin_rule_version,
)
from .bomin_service import calculate_bomin_quote, queue_bomin_job
from .calculator_service import calculate_fangzheng_quote, queue_job
from .db import (
    change_user_password,
    create_feedback,
    create_personal_task,
    create_task_category,
    create_user,
    delete_personal_task,
    delete_task_category,
    delete_job,
    ensure_bootstrap_user,
    get_feedback,
    get_active_job,
    get_job,
    get_personal_task,
    get_rule_history,
    get_task_category,
    get_user,
    is_admin_user,
    list_feedback,
    list_jobs,
    list_personal_tasks,
    list_task_categories,
    list_users,
    reorder_personal_tasks,
    reorder_personal_tasks_by_priority,
    reset_user_password,
    restore_personal_task,
    archive_personal_task,
    update_admin_password,
    update_feedback_status,
    update_personal_task,
    update_task_category,
    verify_admin_password,
    verify_user_password,
)
from .file_utils import safe_unlink
from .hushi_rules import (
    get_active_hushi_rule_version,
    get_hushi_rule_dir,
    get_hushi_rule_history,
    save_new_hushi_rule_version,
)
from .hushi_service import calculate_hushi_quote, queue_hushi_job
from .in_transit_service import queue_in_transit_job
from .inventory_detail_service import (
    FEATURE as INVENTORY_DETAIL_FEATURE,
    PLAN_A_MODE,
    cleanup_inventory_detail_job_files,
    get_inventory_result_path,
    load_inventory_input_manifest,
    load_inventory_result_manifest,
    normalize_inventory_mode,
    queue_inventory_detail_job,
)
from .inventory_bid_service import (
    FEATURE as INVENTORY_BID_FEATURE,
    cleanup_inventory_bid_job_files,
    queue_inventory_bid_job,
    queue_inventory_bid_max_job,
)
from .job_control import cancel_job_process, reconcile_interrupted_jobs
from .order_reprice_service import MODE_LABELS as ORDER_REPRICE_MODE_LABELS
from .order_reprice_service import queue_order_reprice_job
from .price_calculation_customers import (
    PRICE_CALCULATION_CUSTOMERS,
    default_price_customer_key,
    enabled_price_customer,
)
from .price_calculation_rule_docs import get_price_calculation_rule_doc
from .price_calculation_rules import (
    JINGWANG_QUOTE_VARIANTS,
    activate_price_rule_version,
    delete_price_rule_version,
    get_active_price_rule_version,
    get_price_rule_history,
    normalize_price_quote_variant,
    save_new_guanghe_rule_version,
    save_new_suhang_rule_version,
    save_new_price_rule_version,
)
from .price_calculation_service import calculate_price_quote, queue_price_calculation_job, run_jingwang_regression
from .rules import get_active_rule_version, get_rule_file_paths, save_new_rule_version
from .shennan_rules import (
    get_active_shennan_rule_version,
    get_shennan_rule_file_path,
    get_shennan_rule_history,
    save_new_shennan_rule_version,
)
from .shennan_service import calculate_shennan_quote, queue_shennan_job
from .task_backup import get_task_backup_status, restore_task_backup, save_task_backup
from .transcode_rules import (
    TRANSCODE_RULE_SHEETS,
    get_active_transcode_rule_version,
    get_transcode_rule_file_path,
    get_transcode_rule_history,
    save_new_transcode_rule_version_from_sheets,
)
from .transcode_agent_rules import (
    FEATURE_KEY as TRANSCODE_AGENT_FEATURE,
    export_transcode_agent_rules,
    get_active_transcode_agent_rule_version,
    get_transcode_agent_rule_count,
    get_transcode_agent_rule_file_path,
    get_transcode_agent_rule_history,
    load_transcode_agent_rules,
    save_new_transcode_agent_rule_version,
)
from .transcode_agent_service import (
    activate_transcode_agent_pending_rule,
    calculate_transcode_agent_quote,
    confirm_transcode_agent_item,
    delete_transcode_agent_pending_rule,
    finalize_transcode_agent_confirmations,
    get_transcode_agent_pending_rule,
    list_transcode_agent_confirmations,
    list_transcode_agent_pending_rules,
    load_transcode_module,
    queue_transcode_agent_job,
    queue_transcode_agent_single_job,
    reevaluate_transcode_agent_confirmations,
    refresh_transcode_agent_audit_sheet,
    skip_transcode_agent_confirmation_row,
    update_transcode_agent_pending_rule,
    verify_all_transcode_agent_rows,
    verify_transcode_agent_row,
)
from .transcode_customer_rule_admin import (
    AGENT_ASSET_TYPE,
    MAPPING_ASSET_TYPE,
    AGENT_OVERRIDE_TO_BUSINESS_FIELD,
    BUSINESS_FIELDS as CUSTOMER_RULE_BUSINESS_FIELDS,
    BUSINESS_FIELD_TARGETS as CUSTOMER_RULE_FIELD_TARGETS,
    TARGET_FIELD_LABELS as CUSTOMER_RULE_TARGET_FIELD_LABELS,
    CONDITION_FIELDS as CUSTOMER_RULE_CONDITION_FIELDS,
    CONDITION_OPERATOR_LABELS as CUSTOMER_RULE_CONDITION_OPERATOR_LABELS,
    CONDITION_OPERATORS as CUSTOMER_RULE_CONDITION_OPERATORS,
    CustomerRuleMaintenanceError,
    agent_rules_for_customer_workspace,
    build_agent_rule_from_form,
    build_rule_from_form,
    customer_rule_workspace,
    delete_agent_rule_override,
    delete_rule_override,
    find_rule,
    list_customer_rule_changes,
    make_customer_key,
    project_customer_rule_assets_for_workspace,
    restore_customer_rule_change,
    save_agent_rule_override,
    save_rule_override,
    validate_customer_maintained_rule,
    _rule_view,
)
from .transcode_model_config import (
    load_user_model_config,
    test_user_model_connection,
    update_user_model_config,
)
from .transcode_semantic_rules import (
    get_active_transcode_semantic_rule_version,
    load_transcode_semantic_rules,
)
from .transcode_rule_center import (
    BUSINESS_FIELDS as RULE_CENTER_BUSINESS_FIELDS,
    RuleCenterError,
    build_base_rule_from_form,
    build_asset_row_from_form,
    build_rule_center_lookup_tables,
    business_rule_category_meta,
    list_business_rule_rows,
    build_confirmation_policy_from_form,
    confirmation_field_meta,
    create_backup as create_rule_center_backup,
    delete_base_override,
    delete_asset_override,
    delete_confirmation_policy,
    delete_lookup_override,
    ensure_daily_backup as ensure_rule_center_daily_backup,
    find_base_override,
    find_asset_row,
    list_backups as list_rule_center_backups,
    list_base_overrides,
    list_asset_rows,
    list_confirmation_policy_views,
    list_lookup_rows,
    list_rule_center_changes,
    load_score_config,
    lookup_group_meta,
    asset_group_meta,
    restore_backup as restore_rule_center_backup,
    rule_center_summary,
    save_base_override,
    save_asset_override,
    save_confirmation_policy,
    save_lookup_override,
    save_score_config,
)
from .transcode_agent_standard import (
    HIGH_SPEED_MIL_TO_MM,
    OFFICIAL_GRADE_CODES,
    STANDARD_MM_SIZE_ALIASES,
)
from .transcode_service import calculate_transcode_quote, queue_transcode_job
from .transcode_special_import_service import FEATURE_NAME as SPECIAL_IMPORT_FEATURE, queue_transcode_special_import_job
from .transcode_special_rules import (
    STRUCTURED_HEADERS,
    build_export_workbook,
    build_rule_workspace_view,
    customer_has_rules,
    get_latest_original_import_path,
    get_structured_rule_count,
    get_structured_special_rule_settings,
    get_structured_special_rules_path,
    is_structured_special_rules_enabled,
    parse_bulk_special_requirement_workbook,
    parse_special_requirement,
    save_latest_original_import,
    search_structured_special_rules,
    set_structured_special_rules_enabled,
    save_structured_special_rules,
)


bp = Blueprint("main", __name__)
PLATFORM_NAME = "南亚营销自动化平台"
PLATFORM_VERSION = "v1.10.0"

STAGE_META = {
    "online": {"label": "已上线", "desc": "功能已完成主要验证，可作为正式工具使用。"},
    "beta": {"label": "Beta", "desc": "功能已完整可用，但仍需业务继续验证，使用时需复核关键结果。"},
    "test": {"label": "Test", "desc": "功能仍在测试和调试阶段，可试用，但不建议作为正式报价/交付依据。"},
    "planned": {"label": "未上线", "desc": "功能已规划或开发中，暂不可使用。"},
}

FUNCTION_CARDS = [
    {
        "key": "transcode",
        "title": "营销自动化转码",
        "desc": "上传转码需求 Excel，按当前规则自动生成内部编码并输出结果文件。",
        "route": "main.transcode",
        "stage": "test",
    },
    {
        "key": "transcode_agent",
        "title": "营销转码Agent",
        "desc": "按字段证据链和置信度评分进行可信转码，低置信结果自动拦截待确认。",
        "route": "main.transcode_agent",
        "stage": "test",
    },
    {
        "key": "price_calculation",
        "title": "价格计算",
        "desc": "统一使用方正、博敏、深南、沪士及各客户价格表，支持 Excel 批量计算和单条规格即时计算。",
        "route": "main.price_calculation",
        "stage": "test",
    },
    {
        "key": "in_transit",
        "title": "在途核对",
        "desc": "上传包含厂内明细和客户明细的 Excel，自动核对数量、品名、日期和待出货明细。",
        "route": "main.in_transit",
        "stage": "test",
    },
    {
        "key": "inventory_detail",
        "title": "库存明细",
        "desc": "上传上海厂和江西厂库存表，自动筛选、分级并生成A级与B级胶系分类导航版。",
        "route": "main.inventory_detail",
        "stage": "test",
    },
    {
        "key": "inventory_bid",
        "title": "库存竞标",
        "desc": "上传上海和江西两份库存表，按规格、类别、厚度、铜箔、尺寸和铜箔类型汇总竞标库存。",
        "route": "main.inventory_bid",
        "stage": "test",
    },
    {
        "key": "order_reprice",
        "title": "订单改价",
        "desc": "上传胜宏客户明细、厂内明细和报价单，自动完成订单匹配、价格核对与改价结果校验。",
        "route": "main.order_reprice",
        "stage": "test",
    },
    {
        "key": "pdf_excel",
        "title": "PDF/图片转Excel",
        "desc": "批量上传 PDF 或图片，通用识别采购单版式和明细表，输出采购单与明细数据 Excel。",
        "route": "main.pdf_excel",
        "stage": "test",
    },
    {
        "key": "work_planning",
        "title": "工作规划",
        "desc": "按自定义任务类型管理个人私有待办，支持优先级、进展、截止日期和任务描述。",
        "route": "main.work_planning",
        "stage": "online",
    },
    {
        "key": "audit151",
        "title": "151审核",
        "desc": "已进入开发排期，后续用于自动化审核与规则校验。",
        "route": None,
        "stage": "planned",
    },
]

FEATURE_LABELS = {
    "pdf_excel": "PDF/图片转Excel",
    "fangzheng": "方正价格计算",
    "transcode": "营销自动化转码",
    "transcode_agent": "营销转码Agent",
    "shennan": "深南价格计算",
    "bomin": "博敏价格计算",
    "hushi": "沪士价格计算",
    "price_calculation": "价格计算",
    "in_transit": "深南在途核对",
    "inventory_detail": "库存明细",
    "inventory_bid": "库存竞标",
    "order_reprice": "订单改价",
    "work_planning": "工作规划",
}

SPECIAL_PRICE_CALCULATORS = [
    {
        "key": "fangzheng",
        "label": "方正",
        "feature": "fangzheng",
        "upload_endpoint": "main.create_job_view",
        "quote_endpoint": "main.api_fangzheng_quote",
        "admin_endpoint": "main.admin_rules",
        "doc_feature": "fangzheng",
        "panel_tag": "BETA WORKSPACE",
        "description": "上传报价 Excel 后，系统自动识别业务 Sheet 和物料描述列，并用当前方正规则批量计算价格。",
        "rule_source": "方正价格表 / 基板对照表",
        "output_label": "保持 Excel 结果格式",
        "placeholder": "粘贴一条客户规格，系统会按当前方正规则即时计算",
    },
    {
        "key": "bomin",
        "label": "博敏",
        "feature": "bomin",
        "upload_endpoint": "main.create_bomin_job_view",
        "quote_endpoint": "main.api_bomin_quote",
        "admin_endpoint": "main.admin_bomin_rules",
        "doc_feature": None,
        "panel_tag": "TEST WORKSPACE",
        "description": "按当前博敏 CCL / PP 价格表计算含税价格，并输出结果文件与命中说明。",
        "rule_source": "内置价格表 / 管理员上传",
        "output_label": "博敏计算价格",
        "placeholder": "粘贴一条博敏客户规格，系统会按当前博敏规则即时计算",
    },
    {
        "key": "shennan",
        "label": "深南",
        "feature": "shennan",
        "upload_endpoint": "main.create_shennan_job_view",
        "quote_endpoint": "main.api_shennan_quote",
        "admin_endpoint": "main.admin_shennan_rules",
        "doc_feature": "shennan",
        "panel_tag": "TEST WORKSPACE",
        "description": "复用方正计算公式，并按深南汇总报价单中按胶系拆分的 PP / CCL 价格表取值。",
        "rule_source": "深南报价规则 / 方正基板对照",
        "output_label": "深南计算价格",
        "placeholder": "粘贴一条客户规格，系统会按当前深南规则即时计算",
    },
    {
        "key": "hushi",
        "label": "沪士",
        "feature": "hushi",
        "upload_endpoint": "main.create_hushi_job_view",
        "quote_endpoint": "main.api_hushi_quote",
        "admin_endpoint": "main.admin_hushi_rules",
        "doc_feature": "hushi",
        "panel_tag": "TEST WORKSPACE",
        "description": "按沪士多报价单规则自动匹配胶系、规格和 Rebate / Normal 价格并输出结果。",
        "rule_source": "沪士 ZIP 规则包",
        "output_label": "沪士计算价格",
        "placeholder": "粘贴一条客户规格，系统会按当前沪士规则即时计算",
    },
]


def _price_calculator_options() -> list[dict]:
    options = [dict(item, kind="special") for item in SPECIAL_PRICE_CALCULATORS]
    for customer in PRICE_CALCULATION_CUSTOMERS:
        if not customer.get("enabled"):
            continue
        options.append(
            {
                "key": customer["key"],
                "label": customer["label"],
                "kind": "customer",
                "feature": "price_calculation",
                "customer_key": customer["key"],
                "upload_endpoint": "main.create_price_calculation_job_view",
                "quote_endpoint": "main.api_price_calculation_quote",
                "admin_endpoint": "main.admin_price_calculation_rules",
                "doc_feature": None,
                "panel_tag": "TEST WORKSPACE",
                "description": "选择客户后上传 Excel，系统按该客户当前价格表自动匹配规格并输出刷价结果。",
                "rule_source": "客户独立报价表",
                "output_label": "新未税价" if customer["key"] == "plin" else "注意幅宽 / 每卷米数 / 新单价 / 新总金额",
                "placeholder": "粘贴一条客户规格，系统会按当前客户价格表即时计算",
            }
        )
    return options


def _price_calculator_page_url(calculator_key: str, *, job_id: int | None = None, quote_variant: str | None = None) -> str:
    values = {"calculator_key": calculator_key}
    if job_id is not None:
        values["job_id"] = job_id
    if calculator_key == "jingwang" and quote_variant:
        values["quote_variant"] = quote_variant
    return url_for("main.price_calculation", **values)

ORDER_REPRICE_MODE_META = {
    "block1": {
        "tab": "430匹配",
        "title": "客户明细与430厂内明细匹配",
        "history": "430匹配",
        "customer_hint": "字段：采购订单号、项次、料件编号",
        "factory_label": "430厂内明细",
        "factory_hint": "字段：客户订单、项次、客户产品编号",
        "button": "开始匹配",
    },
    "block2": {
        "tab": "430价格核对",
        "title": "430价格核对",
        "history": "430价格核对",
        "customer_hint": "字段：规格；如有含税单价则自动比对",
        "button": "开始价格核对",
        "quote_hint": "支持多个报价单文件",
    },
    "block3": {
        "tab": "411价格核对",
        "title": "客户改价结果与411厂内价格核对",
        "history": "411价格核对",
        "customer_hint": "字段：采购订单号、项次、料件编号、规格、含税单价",
        "factory_label": "411厂内明细",
        "factory_hint": "字段：客户单号、项次、客户产品编号、单价",
        "button": "开始改价核对",
    },
}

ORDER_REPRICE_CUSTOMERS = [
    {
        "key": "shenghong",
        "label": "胜宏",
        "enabled": True,
        "modes": ["block1", "block2", "block3"],
    },
    {"key": "jingwang", "label": "景旺", "enabled": False, "modes": []},
    {"key": "bomin", "label": "博敏", "enabled": False, "modes": []},
]

TASK_PRIORITY_META = {
    "low": {"label": "低", "rank": 4},
    "normal": {"label": "普通", "rank": 3},
    "high": {"label": "高", "rank": 2},
    "urgent": {"label": "紧急", "rank": 1},
}

TASK_PROGRESS_META = {
    "not_started": {"label": "未开始"},
    "in_progress": {"label": "进行中"},
    "completed": {"label": "已完成"},
}

TASK_GROUP_META = {
    "category": {"label": "任务分类"},
    "priority": {"label": "项目优先级"},
    "progress": {"label": "任务进展"},
}


def current_employee() -> str | None:
    return session.get("employee_id")


def current_user():
    employee_id = current_employee()
    return get_user(employee_id) if employee_id else None


def require_login():
    if not current_employee():
        return redirect(url_for("main.login"))
    user = current_user()
    if user and user["must_change_password"] and request.endpoint not in {"main.change_password", "main.logout"}:
        return redirect(url_for("main.change_password"))
    return None


def require_admin_role():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if not is_admin_user(current_employee()):
        abort(403)
    return None


def _pdf_ai_csrf_token() -> str:
    token = str(session.get("pdf_ai_csrf_token") or "")
    if not token:
        token = secrets.token_urlsafe(32)
        session["pdf_ai_csrf_token"] = token
    return token


def _valid_pdf_ai_csrf_token(value: str) -> bool:
    expected = str(session.get("pdf_ai_csrf_token") or "")
    return bool(expected and value and hmac.compare_digest(expected, value))


def _env_flag_enabled(name: str) -> bool:
    return os.environ.get(name, "").strip().lower() in {"1", "true", "yes", "on"}


def _valid_bomin_quote_token() -> bool:
    expected_token = os.environ.get("BOMIN_QUOTE_API_TOKEN", "").strip()
    if not expected_token:
        return False

    authorization = request.headers.get("Authorization", "").strip()
    bearer_token = ""
    if authorization.lower().startswith("bearer "):
        bearer_token = authorization[7:].strip()

    api_key = request.headers.get("X-API-Key", "").strip()
    return any(
        provided and hmac.compare_digest(provided, expected_token)
        for provided in (bearer_token, api_key)
    )


def _localhost_bomin_quote_allowed() -> bool:
    if not _env_flag_enabled("BOMIN_QUOTE_ALLOW_LOCALHOST"):
        return False
    return request.remote_addr in {"127.0.0.1", "::1", "localhost"}


def _bomin_quote_authorized() -> bool:
    if not require_login():
        return True
    return _valid_bomin_quote_token() or _localhost_bomin_quote_allowed()


def _internal_transcode_agent_error(message: str, status_code: int):
    payload = {
        "status": "失败",
        "formal_code": "",
        "candidate_code": "",
        "pending_code": "",
        "confidence": 0,
        "summary": message,
        "note": message,
        "reason": message,
        "error": message,
        "field_evidence": [],
        "rule_version": "",
        "agent_rule_version": "",
        "requires_manual_completion": False,
        "incomplete_fields": [],
        "aps_query_ready": False,
    }
    return jsonify(payload), status_code


def _authorize_internal_transcode_agent():
    expected_token = os.environ.get("TRANSCODE_AGENT_INTERNAL_TOKEN", "").strip()
    if not expected_token:
        return _internal_transcode_agent_error(
            "内部转码服务未配置 TRANSCODE_AGENT_INTERNAL_TOKEN",
            503,
        )

    provided_token = request.headers.get("X-Internal-Token", "").strip()
    if not provided_token or not hmac.compare_digest(provided_token, expected_token):
        return _internal_transcode_agent_error("内部服务 Token 无效", 401)
    return None


def _normalize_internal_transcode_agent_quote(data: dict) -> dict:
    status = str(data.get("status") or "失败").strip() or "失败"
    candidate_code = str(data.get("candidate_code") or "").strip()
    formal_code = str(data.get("formal_code") or data.get("result") or "").strip()
    if status != "成功":
        formal_code = ""

    pending_code = str(data.get("pending_code") or "").strip()
    if status == "待确认":
        pending_code = pending_code or candidate_code
    else:
        pending_code = ""

    note = str(data.get("summary") or data.get("note") or "").strip()
    error = str(data.get("reason") or data.get("error") or "").strip()
    field_evidence = data.get("field_evidence") or []
    incomplete_fields = []
    for item in field_evidence:
        if not isinstance(item, dict):
            continue
        if str(item.get("field_key") or "").strip() != "structure":
            continue
        structure_code = str(item.get("code") or "").strip()
        hit_type = str(item.get("hit_type") or "").strip()
        if structure_code == "*" or hit_type == "占位符":
            incomplete_fields.append("结构码")
            break

    # Older callers may omit field_evidence; '*' is the established structure placeholder.
    if not incomplete_fields and any(
        "*" in code for code in (formal_code, candidate_code, pending_code) if code
    ):
        incomplete_fields.append("结构码")

    requires_manual_completion = bool(incomplete_fields)
    return {
        "status": status,
        "formal_code": formal_code,
        "candidate_code": candidate_code,
        "pending_code": pending_code,
        "confidence": data.get("confidence", 0),
        "summary": note,
        "note": note,
        "reason": error,
        "error": error,
        "field_evidence": field_evidence,
        "rule_version": str(data.get("rule_version") or ""),
        "agent_rule_version": str(data.get("agent_rule_version") or ""),
        "order_semantic_model": data.get("order_semantic_model") or {},
        "requires_manual_completion": requires_manual_completion,
        "incomplete_fields": incomplete_fields,
        "aps_query_ready": bool(
            status == "成功" and formal_code and not requires_manual_completion
        ),
    }


@bp.app_context_processor
def inject_platform_meta():
    return {
        "platform_name": PLATFORM_NAME,
        "platform_version": PLATFORM_VERSION,
        "stage_meta": STAGE_META,
        "feature_labels": FEATURE_LABELS,
    }


@bp.get("/")
def index():
    if current_employee():
        return redirect(url_for("main.dashboard"))
    return redirect(url_for("main.login"))


@bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        employee_id = request.form.get("employee_id", "").strip()
        password = request.form.get("password", "").strip()
        if not employee_id:
            flash("请输入工号。", "error")
        elif ensure_bootstrap_user(employee_id, password) or verify_user_password(employee_id, password):
            session["employee_id"] = employee_id
            user = get_user(employee_id)
            flash("登录成功。", "success")
            if user and user["must_change_password"]:
                return redirect(url_for("main.change_password"))
            return redirect(url_for("main.dashboard"))
        else:
            flash("登录失败：账号不存在、已停用或密码错误。", "error")
    return render_template("login.html")


@bp.route("/change-password", methods=["GET", "POST"])
def change_password():
    redirect_resp = require_login()
    if redirect_resp and request.endpoint != "main.change_password":
        return redirect_resp
    employee_id = current_employee()
    user = current_user()
    if not employee_id or not user:
        return redirect(url_for("main.login"))

    if request.method == "POST":
        old_password = request.form.get("old_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        if not verify_user_password(employee_id, old_password):
            flash("当前密码错误。", "error")
        elif len(new_password) < 6:
            flash("新密码至少 6 位。", "error")
        elif new_password == employee_id:
            flash("新密码不能继续使用工号。", "error")
        elif new_password != confirm_password:
            flash("两次输入的新密码不一致。", "error")
        else:
            change_user_password(employee_id, new_password)
            flash("密码已更新。", "success")
            return redirect(url_for("main.dashboard"))
    return render_template("change_password.html", user=user)


@bp.post("/logout")
def logout():
    session.clear()
    flash("已退出登录。", "success")
    return redirect(url_for("main.login"))


@bp.get("/dashboard")
def dashboard():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    return render_template("dashboard.html", cards=FUNCTION_CARDS)


def _active_job_for(feature: str, jobs):
    requested_job_id = request.args.get("job_id", type=int)
    if requested_job_id:
        job = get_job(requested_job_id)
        if job and job["employee_id"] == current_employee() and job["feature"] == feature:
            return job
    return jobs[0] if jobs else None


def _job_dict(job) -> dict:
    return {key: job[key] for key in job.keys()} if job else {}


def _order_reprice_manifest(job: dict) -> dict:
    if not job or job.get("feature") != "order_reprice":
        return {}
    manifest_path = Path(job.get("stored_input_path") or "")
    if not manifest_path.exists():
        return {}
    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _price_calculation_manifest(job: dict) -> dict:
    if not job or job.get("feature") != "price_calculation":
        return {}
    manifest_path = Path(job.get("stored_input_path") or "")
    if not manifest_path.exists():
        return {}
    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _pdf_excel_manifest(job: dict) -> dict:
    if not job or job.get("feature") != PDF_EXCEL_FEATURE:
        return {}
    manifest_path = Path(job.get("stored_input_path") or "")
    if not manifest_path.exists():
        return {}
    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _price_customer_from_request() -> str:
    key = (
        request.values.get("customer_key")
        or (request.get_json(silent=True) or {}).get("customer_key")
        or default_price_customer_key()
    )
    return enabled_price_customer(key)["key"]


def _price_quote_variant_from_request(customer_key: str) -> str:
    payload = request.get_json(silent=True) or {}
    return normalize_price_quote_variant(
        customer_key,
        request.values.get("quote_variant") or payload.get("quote_variant") or request.args.get("quote_variant"),
    )


def _order_reprice_mode_from_job(job: dict, manifest: dict | None = None) -> str:
    manifest = manifest if manifest is not None else _order_reprice_manifest(job)
    mode = (manifest.get("mode") or "").strip()
    if mode in ORDER_REPRICE_MODE_META:
        return mode
    text = f"{job.get('source_filename') or ''} {job.get('rule_version') or ''}"
    if "第二块" in text or "430价格核对" in text:
        return "block2"
    if "第三块" in text or "411" in text:
        return "block3"
    return "block1"


def _decorate_job(job) -> dict:
    data = _job_dict(job)
    if not data:
        return data
    data["feature_label"] = FEATURE_LABELS.get(data.get("feature"), "方正价格计算")
    data["customer_label"] = "-"
    data["function_type"] = data["feature_label"]
    data["display_total"] = data.get("total_rows") or 0
    data["display_issue_count"] = (data.get("fail_count") or 0) + (data.get("skip_count") or 0)

    if data.get("feature") == "order_reprice":
        manifest = _order_reprice_manifest(data)
        mode = _order_reprice_mode_from_job(data, manifest)
        data["order_reprice_mode"] = mode
        data["customer_label"] = manifest.get("customer_label") or "胜宏"
        data["function_type"] = ORDER_REPRICE_MODE_META.get(mode, {}).get("history", "订单改价")
        stats = _order_reprice_stats(data, mode)
        data["order_reprice_stats"] = stats
        data["display_total"] = stats.get("total", data["display_total"])
        data["display_issue_count"] = stats.get("issue_count", data["display_issue_count"])
    elif data.get("feature") == "price_calculation":
        manifest = _price_calculation_manifest(data)
        data["price_customer_key"] = manifest.get("customer_key") or default_price_customer_key()
        data["quote_variant"] = normalize_price_quote_variant(data["price_customer_key"], manifest.get("quote_variant"))
        data["customer_label"] = manifest.get("customer_label") or "景旺"
        data["function_type"] = f"{data['customer_label']}价格计算"
    elif data.get("feature") == PDF_EXCEL_FEATURE:
        manifest = _pdf_excel_manifest(data)
        file_count = len(manifest.get("files") or [])
        data["customer_label"] = "批量" if file_count > 1 else "单文件"
        data["function_type"] = "PDF/图片转Excel"
        data["display_total"] = data.get("total_rows") or file_count
    elif data.get("feature") == INVENTORY_DETAIL_FEATURE:
        input_manifest = load_inventory_input_manifest(data)
        result_manifest = load_inventory_result_manifest(data)
        mode = normalize_inventory_mode(
            input_manifest.get("inventory_mode") or result_manifest.get("inventory_mode")
        )
        data["inventory_mode"] = mode
        if mode == PLAN_A_MODE:
            data["customer_label"] = "计划A级"
            data["function_type"] = "计划A级分类"
            expected_results = [{"key": "plan-a", "label": "计划A级结果"}]
        else:
            data["customer_label"] = "仓库分类"
            data["function_type"] = "仓库B级分类"
            expected_results = [
                {"key": "a", "label": "仓库A级结果"},
                {"key": "b", "label": "仓库B级结果"},
                {"key": "b-with-product", "label": "仓库B级含品名品号"},
            ]
        available_keys = set((result_manifest.get("files") or {}).keys())
        data["inventory_results"] = [
            item for item in expected_results if not available_keys or item["key"] in available_keys
        ]
    return data


def _decorate_jobs(jobs) -> list[dict]:
    return [_decorate_job(job) for job in jobs]


def _order_reprice_stats(job: dict, mode: str) -> dict:
    result_path = Path(job.get("stored_result_path") or "")
    if job.get("status") == "completed" and result_path.exists():
        try:
            if mode == "block1":
                return _order_reprice_block1_stats(result_path)
            if mode == "block2":
                return _order_reprice_block2_stats(result_path)
            if mode == "block3":
                return _order_reprice_block3_stats(result_path)
        except Exception:
            pass
    return _order_reprice_fallback_stats(job, mode)


def _order_reprice_fallback_stats(job: dict, mode: str) -> dict:
    total = int(job.get("total_rows") or 0)
    success = int(job.get("success_count") or 0)
    fail = int(job.get("fail_count") or 0)
    if mode == "block1":
        cards = [
            {"label": "总记录数", "value": total},
            {"label": "匹配成功数", "value": success},
            {"label": "项次拆分匹配数", "value": "-"},
            {"label": "未匹配数", "value": fail},
            {"label": "匹配成功率", "value": _format_percent(success, total)},
        ]
        return {"total": total, "issue_count": fail, "cards": cards}
    if mode == "block2":
        cards = [
            {"label": "总记录数", "value": total},
            {"label": "价格正确数", "value": success},
            {"label": "价格异常数", "value": fail},
            {"label": "未匹配数", "value": "-"},
            {"label": "异常率", "value": _format_percent(fail, total)},
        ]
        return {"total": total, "issue_count": fail, "cards": cards}
    cards = [
        {"label": "总记录数", "value": total},
        {"label": "价格正确数", "value": success},
        {"label": "价格错误数", "value": fail},
        {"label": "未匹配数", "value": "-"},
    ]
    return {"total": total, "issue_count": fail, "cards": cards}


def _order_reprice_block1_stats(result_path: Path) -> dict:
    rows = _workbook_records(result_path, "客户明细匹配结果")
    statuses = [str(row.get("匹配状态") or "").strip() for row in rows]
    total = len(rows)
    split = statuses.count("拆分匹配")
    success = statuses.count("已匹配") + split
    unmatched = max(total - success, 0)
    cards = [
        {"label": "总记录数", "value": total},
        {"label": "匹配成功数", "value": success},
        {"label": "项次拆分匹配数", "value": split},
        {"label": "未匹配数", "value": unmatched},
        {"label": "匹配成功率", "value": _format_percent(success, total)},
    ]
    return {"total": total, "issue_count": unmatched, "cards": cards}


def _order_reprice_block2_stats(result_path: Path) -> dict:
    summary = _workbook_summary(result_path, "匹配汇总")
    total = _summary_int(summary, "总记录数")
    correct = _summary_int(summary, "价格一致数量") or _summary_int(summary, "价格正确数量")
    price_error = _summary_int(summary, "价格不一致数量") or _summary_int(summary, "价格错误数量")
    quote_missing = _summary_int(summary, "未命中报价数量")
    factory_missing = _summary_int(summary, "未匹配厂内数量")
    skipped = _summary_int(summary, "不输出数量")
    unmatched = quote_missing + factory_missing
    issue_count = price_error + unmatched
    cards = [
        {"label": "总记录数", "value": total},
        {"label": "价格一致数", "value": correct},
        {"label": "价格不一致数", "value": price_error},
        {"label": "未匹配数", "value": unmatched},
        {"label": "不输出数", "value": skipped},
        {"label": "异常率", "value": _format_percent(issue_count, total)},
    ]
    return {"total": total, "issue_count": issue_count, "cards": cards}


def _order_reprice_block3_stats(result_path: Path) -> dict:
    summary = _workbook_summary(result_path, "核对汇总")
    total = _summary_int(summary, "总记录数")
    correct = _summary_int(summary, "价格正确数量")
    error = _summary_int(summary, "价格错误数量")
    unmatched = _summary_int(summary, "未匹配数量")
    cards = [
        {"label": "总记录数", "value": total},
        {"label": "价格正确数", "value": correct},
        {"label": "价格错误数", "value": error},
        {"label": "未匹配数", "value": unmatched},
    ]
    return {"total": total, "issue_count": error + unmatched, "cards": cards}


def _workbook_records(path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
        records = []
        for row in rows:
            if not any(value is not None and str(value).strip() for value in row):
                continue
            records.append({headers[idx]: value for idx, value in enumerate(row) if idx < len(headers) and headers[idx]})
        return records
    finally:
        wb.close()


def _workbook_summary(path: Path, sheet_name: str) -> dict[str, object]:
    return {str(row.get("项目") or "").strip(): row.get("数量") for row in _workbook_records(path, sheet_name)}


def _summary_int(summary: dict[str, object], key: str) -> int:
    try:
        return int(float(summary.get(key) or 0))
    except (TypeError, ValueError):
        return 0


def _format_percent(numerator: int, denominator: int) -> str:
    if not denominator:
        return "0%"
    return f"{numerator / denominator * 100:.1f}%"


@bp.get("/features/fangzheng")
def fangzheng():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    return redirect(_price_calculator_page_url("fangzheng", job_id=request.args.get("job_id", type=int)))


@bp.get("/features/transcode")
def transcode():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    jobs = list_jobs(current_employee(), limit=20, feature="transcode")
    return render_template(
        "transcode.html",
        jobs=jobs,
        active_rule_version=get_active_transcode_rule_version(),
        active_job=_active_job_for("transcode", jobs),
    )


@bp.get("/features/transcode-agent")
def transcode_agent():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    jobs = list_jobs(current_employee(), limit=20, feature=TRANSCODE_AGENT_FEATURE)
    model_config = load_user_model_config(current_employee())
    return render_template(
        "transcode_agent.html",
        jobs=jobs,
        model_config=model_config,
        active_job=_active_job_for(TRANSCODE_AGENT_FEATURE, jobs),
        auto_confirm=request.args.get("auto_confirm") == "1",
    )


@bp.get("/admin/transcode-agent-pending-rules")
def admin_transcode_agent_pending_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    return redirect(
        url_for(
            "main.admin_transcode_rule_center",
            section="submitted",
        )
    )


@bp.route("/admin/transcode-rule-center", methods=["GET", "POST"])
def admin_transcode_rule_center():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    employee_id = current_employee() or ""
    section = str(request.values.get("section") or "active").strip()
    if section == "overview":
        # 兼容旧收藏地址：仍直接返回页面，但按新的“生效规则”口径渲染。
        section = "active"
    if section == "semantic":
        return redirect(url_for("main.admin_transcode_rule_center", section="customer"))
    if request.method == "POST":
        action = str(request.form.get("action") or "").strip()
        try:
            if action == "save_lookup":
                save_lookup_override(request.form, updated_by=employee_id)
                flash("基础映射已保存并立即生效。", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_rule_center",
                        section="base",
                        business_category=request.form.get("business_category") or "胶系",
                        lookup_group=request.form.get("lookup_group") or "glue_code",
                        lookup_input=request.form.get("lookup_input") or "",
                    )
                )
            if action == "delete_lookup":
                delete_lookup_override(
                    request.form.get("lookup_group") or "",
                    request.form.get("lookup_input") or "",
                    updated_by=employee_id,
                )
                flash("基础映射已停用并保留修改记录。", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_rule_center",
                        section="base",
                        business_category=request.form.get("business_category") or "胶系",
                    )
                )
            if action == "save_asset":
                from .transcode_agent_rules import load_transcode_agent_mapping_tables

                asset_group = request.form.get("asset_group") or "Agent胶系主表"
                existing = find_asset_row(
                    load_transcode_agent_mapping_tables(),
                    asset_group,
                    request.form.get("asset_row_id") or "",
                )
                group, row = build_asset_row_from_form(request.form, existing=existing)
                save_asset_override(group, row, updated_by=employee_id)
                flash("规则资产已保存并立即生效。", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_rule_center",
                        section="base",
                        business_category=request.form.get("business_category") or "胶系",
                        asset_group=group,
                        asset_row_id=row["映射ID"],
                    )
                )
            if action == "delete_asset":
                group = request.form.get("asset_group") or ""
                delete_asset_override(
                    group,
                    request.form.get("asset_row_id") or "",
                    updated_by=employee_id,
                )
                flash("规则资产已停用并保留修改记录。", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_rule_center",
                        section="base",
                        business_category=request.form.get("business_category") or "胶系",
                    )
                )
            if action == "save_score":
                save_score_config(request.form, updated_by=employee_id)
                flash("评分与人工确认标准已保存并立即生效。", "success")
                return redirect(url_for("main.admin_transcode_rule_center", section="scoring"))
            if action == "save_confirmation_policy":
                policy_views = list_confirmation_policy_views()
                selected = next(
                    (
                        item["native_rule"]
                        for item in policy_views
                        if item["rule_id"] == str(request.form.get("confirmation_rule_id") or "")
                    ),
                    None,
                )
                rule = build_confirmation_policy_from_form(request.form, existing=selected)
                save_confirmation_policy(rule, updated_by=employee_id)
                flash("人工确认触发条件已保存并立即生效。", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_rule_center",
                        section="scoring",
                        confirmation_rule_id=rule["rule_id"],
                    )
                )
            if action == "delete_confirmation_policy":
                delete_confirmation_policy(
                    request.form.get("confirmation_rule_id") or "",
                    updated_by=employee_id,
                )
                flash("人工确认触发条件已停用。", "success")
                return redirect(url_for("main.admin_transcode_rule_center", section="scoring"))
            if action == "save_pending_rule":
                pending_rule_id = int(request.form.get("pending_rule_id") or 0)
                update_transcode_agent_pending_rule(
                    pending_rule_id,
                    employee_id,
                    request.form,
                )
                flash("已提交待生效规则已更新。", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_rule_center",
                        section="submitted",
                        pending_rule_id=pending_rule_id,
                    )
                )
            if action == "activate_pending_rule":
                pending_rule_id = int(request.form.get("pending_rule_id") or 0)
                activate_transcode_agent_pending_rule(
                    pending_rule_id,
                    employee_id,
                )
                flash("待生效规则已确认并写入客户特殊规则。", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_rule_center",
                        section="submitted",
                    )
                )
            if action == "delete_pending_rule":
                pending_rule_id = int(request.form.get("pending_rule_id") or 0)
                delete_transcode_agent_pending_rule(
                    pending_rule_id,
                    employee_id,
                )
                flash("待生效规则已删除并保留留痕。", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_rule_center",
                        section="submitted",
                    )
                )
            if action == "backup":
                path = create_rule_center_backup(reason="页面手动备份")
                flash(f"规则已备份：{path.name}", "success")
                return redirect(url_for("main.admin_transcode_rule_center", section="backups"))
            if action == "restore_backup":
                restore_rule_center_backup(
                    request.form.get("backup_name") or "",
                    updated_by=employee_id,
                )
                flash("备份已恢复，规则立即生效。", "success")
                return redirect(url_for("main.admin_transcode_rule_center", section="backups"))
            raise RuleCenterError("未知的规则维护操作。")
        except (
            CustomerRuleMaintenanceError,
            LookupError,
            RuleCenterError,
            ValueError,
            OSError,
            json.JSONDecodeError,
        ) as exc:
            flash(f"规则配置失败：{exc}", "error")

    ensure_rule_center_daily_backup()
    agent_rules = load_transcode_agent_rules()
    semantic_version = get_active_transcode_semantic_rule_version()
    semantic_rules = load_transcode_semantic_rules(semantic_version) if semantic_version else []
    from .transcode_agent_rules import load_transcode_agent_mapping_tables

    mapping_tables = load_transcode_agent_mapping_tables()
    engine = load_transcode_module()
    rule_path = get_transcode_rule_file_path(get_active_transcode_rule_version())
    lookup_tables = engine.build_lookup_tables(engine.load_rule_sheets(str(rule_path)))
    lookup_tables = build_rule_center_lookup_tables(
        lookup_tables,
        mapping_tables,
        official_grade_codes=set(OFFICIAL_GRADE_CODES),
        standard_sizes=STANDARD_MM_SIZE_ALIASES,
        high_speed_mil=HIGH_SPEED_MIL_TO_MM,
        copper_micron=engine.MICRON_COPPER_MAP,
        copper_types=engine.SPECIAL_COPPER_MAP,
        copper_valid=engine.VALID_COPPER_SPECS,
        size_ranges=engine.STANDARD_SIZE_RANGES,
    )
    lookup_groups = lookup_group_meta()
    business_categories = business_rule_category_meta()
    business_category = request.args.get("business_category") or RULE_CENTER_BUSINESS_FIELDS[0]
    if business_category not in business_categories:
        business_category = RULE_CENTER_BUSINESS_FIELDS[0]
    lookup_group = request.args.get("lookup_group") or "glue_code"
    if lookup_group not in lookup_groups:
        lookup_group = "glue_code"
    asset_groups = asset_group_meta()
    asset_group = request.args.get("asset_group") or "Agent胶系主表"
    if asset_group not in asset_groups:
        asset_group = "Agent胶系主表"
    asset_rows = list_asset_rows(mapping_tables, asset_group=asset_group)
    selected_asset = find_asset_row(
        mapping_tables,
        asset_group,
        request.args.get("asset_row_id") or "",
    )
    confirmation_policies = list_confirmation_policy_views()
    selected_confirmation_policy = next(
        (
            item
            for item in confirmation_policies
            if item["rule_id"] == str(request.args.get("confirmation_rule_id") or "")
        ),
        None,
    )
    summary = rule_center_summary(
        mapping_tables=mapping_tables,
        agent_rules=agent_rules,
        semantic_rules=semantic_rules,
    )
    customer_special_rules = project_customer_rule_assets_for_workspace(
        semantic_rules,
        agent_rules,
        mapping_tables,
    )
    customer_rule_summary = customer_rule_workspace(customer_special_rules)
    summary["base_rule_count"] = sum(
        len(list_business_rule_rows(lookup_tables, mapping_tables, category=category))
        for category in business_categories
    )
    summary["customer_special_count"] = len(customer_special_rules)
    summary["customer_count"] = len(
        {
            make_customer_key(rule.get("customer_code"), rule.get("customer_name"))
            for rule in customer_special_rules
            if rule.get("customer_code") or rule.get("customer_name")
        }
    )
    coverage_rows = [
        {"category": "基础规则", "source": "正式业务映射、编码规范、正式胶系表及确定性算法", "count": summary["base_rule_count"], "status": "生效规则"},
        {"category": "客户特殊规则", "source": "按客户统一维护全客户、客户专属和后台运行规则", "count": customer_rule_summary["display_scope_counts"].get("active", 0), "status": "生效规则"},
        {"category": "待完善事项", "source": "待业务确认、待技术支持；不会参与自动出码", "count": customer_rule_summary["display_scope_counts"].get("pending", 0), "status": "待完善"},
        {"category": "历史资料", "source": "历史样本和外部资料；仅在备份与修改记录中追溯", "count": customer_rule_summary["display_scope_counts"].get("reference", 0), "status": "历史追溯"},
        {"category": "出码与人工确认", "source": "统一维护100分出码及人工确认标准", "count": len(confirmation_policies), "status": "统一维护"},
    ]
    is_rule_admin = is_admin_user(employee_id)
    pending_rules = list_transcode_agent_pending_rules(
        employee_id,
        include_all=is_rule_admin,
    )
    selected_pending_rule = None
    selected_pending_id = request.args.get("pending_rule_id", type=int)
    if selected_pending_id:
        selected_pending_rule = get_transcode_agent_pending_rule(
            selected_pending_id,
            employee_id,
            include_all=is_rule_admin,
        )
    lookup_rows = list_lookup_rows(lookup_tables, group_key=lookup_group)
    business_rule_rows = list_business_rule_rows(
        lookup_tables,
        mapping_tables,
        category=business_category,
    )
    return render_template(
        "transcode_rule_center.html",
        section=section,
        summary=summary,
        customer_rule_summary=customer_rule_summary,
        business_fields=RULE_CENTER_BUSINESS_FIELDS,
        business_categories=business_categories,
        business_category=business_category,
        business_rule_rows=business_rule_rows,
        score_config=load_score_config(),
        confirmation_fields=confirmation_field_meta(),
        confirmation_policies=confirmation_policies,
        selected_confirmation_policy=selected_confirmation_policy,
        changes=list_rule_center_changes(),
        customer_changes=list_customer_rule_changes(limit=30),
        backups=list_rule_center_backups(),
        lookup_groups=lookup_groups,
        lookup_group=lookup_group,
        lookup_rows=lookup_rows,
        asset_groups=asset_groups,
        asset_group=asset_group,
        asset_rows=asset_rows,
        selected_asset=selected_asset,
        coverage_rows=coverage_rows,
        pending_rules=pending_rules,
        selected_pending_rule=selected_pending_rule,
        is_rule_admin=is_rule_admin,
        pending_rule_condition_fields=CUSTOMER_RULE_CONDITION_FIELDS,
        pending_rule_condition_operators=CUSTOMER_RULE_CONDITION_OPERATORS,
        pending_rule_condition_operator_labels=CUSTOMER_RULE_CONDITION_OPERATOR_LABELS,
        pending_rule_field_targets=CUSTOMER_RULE_FIELD_TARGETS,
        active_base_version=get_active_transcode_rule_version(),
        active_agent_version=get_active_transcode_agent_rule_version(),
    )


@bp.get("/features/transcode-agent/confirmations/<int:job_id>")
def transcode_agent_confirmation_center(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    try:
        confirmation_data = list_transcode_agent_confirmations(
            job_id,
            current_employee() or "",
            record_scope=request.args.get("scope", "all"),
            record_page=request.args.get("page", 1, type=int) or 1,
            record_page_size=200,
        )
    except LookupError:
        flash("未找到该营销转码Agent确认任务。", "error")
        return redirect(url_for("main.transcode_agent"))
    job = get_job(job_id)
    return render_template(
        "transcode_agent_confirmation.html",
        job=job,
        confirmation_data=confirmation_data,
        condition_fields=CUSTOMER_RULE_CONDITION_FIELDS,
        condition_operators=CUSTOMER_RULE_CONDITION_OPERATORS,
    )


@bp.route("/features/transcode-agent/model-config", methods=["GET", "POST"])
def transcode_agent_model_config():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    employee_id = current_employee() or ""
    if request.method == "POST":
        action = str(request.form.get("action") or "save").strip()
        api_key_text = str(request.form.get("api_key") or "").strip()
        clear_key = bool(request.form.get("clear_api_key"))
        api_key: str | None = "" if clear_key else (api_key_text or None)
        try:
            config = update_user_model_config(
                employee_id,
                enabled=bool(request.form.get("enabled")),
                base_url=request.form.get("base_url") or "",
                api_key=api_key,
                model=request.form.get("model") or "",
            )
            if action == "test":
                result = test_user_model_connection(config)
                flash(f"模型{result['status']}：{result['model']}。", "success")
                return redirect(url_for("main.transcode_agent_model_config"))
            else:
                flash("当前用户的模型配置已保存。", "success")
                return redirect(url_for("main.transcode_agent"))
        except Exception as exc:
            flash(f"模型配置失败：{exc}", "error")
        return redirect(url_for("main.transcode_agent_model_config"))
    return render_template(
        "transcode_agent_model_config.html",
        model_config=load_user_model_config(employee_id),
    )


@bp.route("/admin/transcode-agent-customer-rules", methods=["GET", "POST"])
def admin_transcode_agent_customer_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    employee_id = current_employee() or ""
    active_version = get_active_transcode_semantic_rule_version()
    semantic_rules = load_transcode_semantic_rules(active_version) if active_version else []
    agent_rules = load_transcode_agent_rules()
    from .transcode_agent_rules import load_transcode_agent_mapping_tables

    customer_mapping_tables = load_transcode_agent_mapping_tables()
    rules = project_customer_rule_assets_for_workspace(
        semantic_rules,
        agent_rules,
        customer_mapping_tables,
    )
    if request.method == "POST":
        action = str(request.form.get("action") or "save").strip()
        try:
            if action == "mark_technical_resolved":
                rule = find_rule(rules, request.form.get("rule_id") or "")
                if rule is None or str(rule.get("review_state") or "") != "technical":
                    raise CustomerRuleMaintenanceError("待技术支持规则不存在或已处理")
                mapping_row = dict(rule.get("mapping_row") or {})
                mapping_row["映射ID"] = str(mapping_row.get("映射ID") or "")
                if not mapping_row["映射ID"]:
                    raise CustomerRuleMaintenanceError("待技术支持记录缺少映射ID")
                existing_remark = str(mapping_row.get("备注") or "").strip()
                mapping_row["备注"] = (
                    "已清理：已有对应可执行规则，页面不再展示。"
                    + (f"原备注：{existing_remark}" if existing_remark else "")
                )
                mapping_row["建议处理"] = "已清理：已有对应可执行规则"
                save_asset_override(
                    "待接入规则",
                    mapping_row,
                    updated_by=employee_id,
                    record_change=False,
                )
                flash("待技术支持记录已标记为已清理。", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_agent_customer_rules",
                        customer_key=make_customer_key(
                            mapping_row.get("客户代码"),
                            mapping_row.get("客户简称"),
                        ),
                        business_field=rule.get("business_field") or "基板尺寸",
                        scope="pending",
                        show_history="1",
                    )
                )
            if action == "save":
                existing = find_rule(rules, request.form.get("rule_id") or "")
                asset_type = str(request.form.get("asset_type") or "semantic").strip()
                if asset_type == MAPPING_ASSET_TYPE:
                    mapping_group = str(request.form.get("mapping_group") or "").strip()
                    native_existing = dict((existing or {}).get("mapping_row") or {})
                    group, row = build_asset_row_from_form(request.form, existing=native_existing)
                    if group != mapping_group:
                        raise CustomerRuleMaintenanceError("客户转换规则分类不匹配。")
                    save_asset_override(group, row, updated_by=employee_id)
                    redirect_code = row.get("客户代码") or ""
                    redirect_name = row.get("客户简称") or ""
                    redirect_field = (existing or {}).get("business_field") or "基板尺寸"
                    redirect_rule_id = (existing or {}).get("rule_id") or f"MAP::{group}::{row['映射ID']}"
                elif asset_type == AGENT_ASSET_TYPE:
                    native_existing = dict((existing or {}).get("agent_rule") or {})
                    native_rule = build_agent_rule_from_form(
                        request.form,
                        existing_rule=native_existing,
                    )
                    save_agent_rule_override(
                        native_rule,
                        updated_by=employee_id,
                        previous_rule=native_existing,
                    )
                    redirect_code = native_rule["客户代码"]
                    redirect_name = native_rule["客户简称"]
                    redirect_field = AGENT_OVERRIDE_TO_BUSINESS_FIELD[native_rule["覆盖字段"]]
                    redirect_rule_id = native_rule["规则ID"]
                else:
                    rule = build_rule_from_form(request.form, existing_rule=existing)
                    validate_customer_maintained_rule(rule)
                    save_rule_override(rule, updated_by=employee_id, previous_rule=existing)
                    redirect_code = rule["customer_code"]
                    redirect_name = rule["customer_name"]
                    redirect_field = rule["business_field"]
                    redirect_rule_id = rule["rule_id"]
                flash(f"客户特殊规则已保存并生效：{redirect_rule_id}", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_agent_customer_rules",
                        customer_key=make_customer_key(redirect_code, redirect_name),
                        business_field=redirect_field,
                        rule_id=redirect_rule_id,
                        scope=request.form.get("scope") or "active",
                    )
                )
            if action == "delete":
                rule = find_rule(rules, request.form.get("rule_id") or "")
                if rule is None:
                    raise CustomerRuleMaintenanceError("要删除的规则不存在。")
                if str(rule.get("asset_type") or "") == MAPPING_ASSET_TYPE:
                    delete_asset_override(
                        str(rule.get("mapping_group") or ""),
                        str((rule.get("mapping_row") or {}).get("映射ID") or ""),
                        updated_by=employee_id,
                    )
                elif str(rule.get("asset_type") or "") == AGENT_ASSET_TYPE:
                    delete_agent_rule_override(dict(rule.get("agent_rule") or {}), updated_by=employee_id)
                else:
                    delete_rule_override(rule, updated_by=employee_id)
                flash(f"客户特殊规则已删除并立即停止生效：{rule['rule_id']}", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_agent_customer_rules",
                        customer_key=make_customer_key(rule["customer_code"], rule["customer_name"]),
                        business_field=rule["business_field"],
                        scope=request.form.get("scope") or "active",
                    )
                )
            if action == "restore":
                restored_rule_id = restore_customer_rule_change(
                    int(request.form.get("change_id") or 0),
                    updated_by=employee_id,
                )
                flash(f"修改记录已恢复并生效：{restored_rule_id}", "success")
                return redirect(
                    url_for(
                        "main.admin_transcode_agent_customer_rules",
                        show_history="1",
                    )
                )
            raise CustomerRuleMaintenanceError("未知的客户规则维护操作。")
        except (CustomerRuleMaintenanceError, ValueError) as exc:
            flash(f"客户特殊规则保存失败：{exc}", "error")

    workspace = customer_rule_workspace(
        rules,
        search=request.values.get("search") or "",
        customer_key=request.values.get("customer_key") or "",
        business_field=request.values.get("business_field") or "",
        rule_id=request.values.get("rule_id") or "",
        rule_kind="all",
        rule_scope=request.values.get("scope") or "active",
        status_filter=request.values.get("status") or "all",
    )
    if request.args.get("new") == "1" or request.args.get("new_customer") == "1":
        selected_customer = workspace.get("selected_customer") or {}
        is_new_customer = request.args.get("new_customer") == "1"
        selected_field = workspace["selected_field"]
        new_rule_type = str(request.args.get("rule_type") or "deterministic").strip()
        is_order_semantic = new_rule_type == "order_semantic"
        default_override_field = next(
            (
                key
                for key, field in AGENT_OVERRIDE_TO_BUSINESS_FIELD.items()
                if field == selected_field
            ),
            "grade_code",
        )
        workspace["selected_rule"] = {
            "rule_id": "",
            "asset_type": "semantic" if is_order_semantic else AGENT_ASSET_TYPE,
            "customer_code": "" if is_new_customer else selected_customer.get("code", ""),
            "customer_name": "" if is_new_customer else selected_customer.get("name", ""),
            "business_field": selected_field,
            "source_text": "",
            "input_source": "客户规格",
            "conditions": [
                {
                    "field": "订单备注" if is_order_semantic else "客户规格",
                    "operator": "contains_any" if is_order_semantic else "present",
                    "value": "",
                }
            ],
            "semantic_enabled": is_order_semantic,
            "agent_rule": {
                "覆盖字段": default_override_field,
                "条件胶系": "",
                "条件关键词": "",
                "条件铜厚": "",
                "条件厚度": "",
                "条件尺寸": "",
            },
            "target_fields": [],
            "normalized_values": [],
            "target_field": CUSTOMER_RULE_FIELD_TARGETS[selected_field][0],
            "target_value": "",
            "condition_summary": "",
            "priority": 100,
            "enabled": True,
            "editable": True,
            "status_label": "新增规则",
            "origin": "新增规则",
            "machine_rule": None,
            "approval": {"basis": ""},
            "approval_basis": "",
        }
    elif request.args.get("migrate_rule_id"):
        migration_rule = find_rule(rules, request.args.get("migrate_rule_id") or "")
        if migration_rule is None or str(migration_rule.get("review_state") or "") != "migration":
            flash("要迁移的运行中规则不存在，或其状态已变化。", "error")
        else:
            migration_view = _rule_view(migration_rule, overridden=False)
            workspace["selected_rule"] = {
                **migration_view,
                "rule_id": "",
                "asset_type": "semantic",
                "editable": True,
                "enabled": True,
                "priority": max(100, int(migration_view.get("priority") or 0)),
                "source_text": "迁移确认：" + (migration_view.get("source_text") or ""),
                "approval_basis": "由运行中规则迁移确认：" + (migration_view.get("rule_id") or ""),
                "status_label": "迁移确认",
                "origin": "运行中规则迁移确认",
                "semantic_enabled": False,
                "migration_source_rule_id": migration_view.get("rule_id") or "",
            }
    return render_template(
        "admin_transcode_agent_customer_rules.html",
        workspace=workspace,
        active_semantic_version=active_version or "未发布",
        active_agent_version=get_active_transcode_agent_rule_version() or "未发布",
        agent_override_fields=AGENT_OVERRIDE_TO_BUSINESS_FIELD,
        business_fields=CUSTOMER_RULE_BUSINESS_FIELDS,
        field_targets=CUSTOMER_RULE_FIELD_TARGETS,
        target_field_labels=CUSTOMER_RULE_TARGET_FIELD_LABELS,
        condition_fields=CUSTOMER_RULE_CONDITION_FIELDS,
        condition_operators=CUSTOMER_RULE_CONDITION_OPERATORS,
        condition_operator_labels=CUSTOMER_RULE_CONDITION_OPERATOR_LABELS,
        mapping_asset_type=MAPPING_ASSET_TYPE,
        mapping_asset_groups=asset_group_meta(),
        rule_changes=list_customer_rule_changes(),
        show_history=request.args.get("show_history") == "1",
    )


@bp.get("/features/shennan")
def shennan():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    return redirect(_price_calculator_page_url("shennan", job_id=request.args.get("job_id", type=int)))


@bp.get("/features/hushi")
def hushi():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    return redirect(_price_calculator_page_url("hushi", job_id=request.args.get("job_id", type=int)))


@bp.get("/features/bomin")
def bomin():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    return redirect(_price_calculator_page_url("bomin", job_id=request.args.get("job_id", type=int)))


@bp.get("/features/price-calculation")
def price_calculation():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    calculators = _price_calculator_options()
    calculator_map = {item["key"]: item for item in calculators}
    requested_key = request.args.get("calculator_key") or request.args.get("customer_key") or "fangzheng"
    selected_calculator = requested_key if requested_key in calculator_map else "fangzheng"
    selected_option = calculator_map[selected_calculator]
    selected_customer = selected_option.get("customer_key", "")
    selected_quote_variant = (
        normalize_price_quote_variant(selected_customer, request.args.get("quote_variant"))
        if selected_customer
        else ""
    )

    selected_feature = selected_option["feature"]
    raw_jobs = list_jobs(current_employee(), limit=50 if selected_feature == "price_calculation" else 20, feature=selected_feature)
    if selected_feature == "price_calculation":
        jobs = [
            job
            for job in _decorate_jobs(raw_jobs)
            if job.get("price_customer_key", default_price_customer_key()) == selected_customer
            and (selected_customer != "jingwang" or job.get("quote_variant", "new") == selected_quote_variant)
        ][:20]
        active_job = _decorate_job(_active_job_for(selected_feature, raw_jobs))
        if active_job and (
            active_job.get("price_customer_key") != selected_customer
            or (selected_customer == "jingwang" and active_job.get("quote_variant", "new") != selected_quote_variant)
        ):
            active_job = None
        active_rule_version = get_active_price_rule_version(selected_customer, selected_quote_variant) or "未初始化价格计算规则"
    else:
        jobs = _decorate_jobs(raw_jobs)
        active_job = _decorate_job(_active_job_for(selected_feature, raw_jobs))
        if selected_calculator == "fangzheng":
            active_rule_version = get_active_rule_version()
        elif selected_calculator == "bomin":
            active_rule_version = get_active_bomin_rule_version() or "未初始化博敏价格表"
        elif selected_calculator == "shennan":
            active_rule_version = get_active_shennan_rule_version()
        else:
            active_rule_version = get_active_hushi_rule_version() or "未上传沪士规则"

    admin_values = {"customer_key": selected_customer, "quote_variant": selected_quote_variant} if selected_customer else {}
    return render_template(
        "price_calculation.html",
        calculators=calculators,
        selected_calculator=selected_calculator,
        selected_option=selected_option,
        selected_customer=selected_customer,
        selected_quote_variant=selected_quote_variant,
        quote_variants=JINGWANG_QUOTE_VARIANTS,
        jobs=jobs,
        active_rule_version=active_rule_version,
        active_job=active_job,
        price_rule_doc=get_price_calculation_rule_doc(selected_calculator),
        upload_url=url_for(selected_option["upload_endpoint"]),
        quote_url=url_for(selected_option["quote_endpoint"]),
        admin_url=url_for(selected_option["admin_endpoint"], **admin_values),
        doc_url=url_for("main.rule_doc", feature=selected_option["doc_feature"]) if selected_option.get("doc_feature") else None,
    )


@bp.get("/features/in-transit")
def in_transit():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    reconcile_interrupted_jobs()
    jobs = list_jobs(current_employee(), limit=20, feature="in_transit")
    return render_template(
        "in_transit.html",
        jobs=jobs,
        active_rule_version="内置核对规则 v1",
        active_job=_active_job_for("in_transit", jobs),
    )


@bp.get("/features/inventory-detail")
def inventory_detail():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    reconcile_interrupted_jobs()
    raw_jobs = list_jobs(current_employee(), limit=20, feature=INVENTORY_DETAIL_FEATURE)
    active_job = _active_job_for(INVENTORY_DETAIL_FEATURE, raw_jobs)
    return render_template(
        "inventory_detail.html",
        jobs=_decorate_jobs(raw_jobs),
        active_rule_version="库存明细内置规则 v1",
        active_job=_decorate_job(active_job) if active_job else None,
    )


@bp.get("/features/inventory-bid")
def inventory_bid():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    reconcile_interrupted_jobs()
    jobs = list_jobs(current_employee(), limit=20, feature=INVENTORY_BID_FEATURE)
    return render_template(
        "inventory_bid.html",
        jobs=_decorate_jobs(jobs),
        active_rule_version="库存竞标内置规则 v2",
        active_job=_decorate_job(_active_job_for(INVENTORY_BID_FEATURE, jobs)),
    )


@bp.get("/features/order-reprice")
def order_reprice():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    raw_jobs = list_jobs(current_employee(), limit=20, feature="order_reprice")
    active_raw_job = _active_job_for("order_reprice", raw_jobs)
    jobs = _decorate_jobs(raw_jobs)
    return render_template(
        "order_reprice.html",
        jobs=jobs,
        active_rule_version="订单改价内置规则 v1",
        active_job=_decorate_job(active_raw_job) if active_raw_job else None,
        mode_labels=ORDER_REPRICE_MODE_LABELS,
        mode_meta=ORDER_REPRICE_MODE_META,
        customers=ORDER_REPRICE_CUSTOMERS,
    )


def _render_pdf_excel_page(
    *,
    ai_form_values: dict[str, object] | None = None,
    ai_result: dict[str, str] | None = None,
    open_ai_dialog: bool = False,
):
    jobs = list_jobs(current_employee(), limit=20, feature=PDF_EXCEL_FEATURE)
    ai_admin = None
    if is_admin_user(current_employee()):
        if ai_result is None:
            stored_result = session.pop("pdf_ai_modal_result", None)
            if isinstance(stored_result, dict):
                ai_result = stored_result
        ai_admin = _pdf_ai_admin_context(form_values=ai_form_values, result=ai_result)
    return render_template(
        "pdf_excel.html",
        jobs=_decorate_jobs(jobs),
        active_rule_version="PDF/图片转Excel 内置解析规则 v1",
        active_job=_decorate_job(_active_job_for(PDF_EXCEL_FEATURE, jobs)),
        ai_admin=ai_admin,
        open_ai_dialog=open_ai_dialog,
    )


@bp.get("/features/pdf-excel")
def pdf_excel():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    return _render_pdf_excel_page(open_ai_dialog=request.args.get("ai_config") == "1")


def _pdf_ai_submitted_values() -> dict[str, object]:
    return {
        "enabled": bool(request.form.get("enabled")),
        "api_key": request.form.get("api_key", "").strip(),
        "base_url": request.form.get("base_url", "").strip(),
        "model": request.form.get("model", "").strip(),
        "timeout_seconds": request.form.get("timeout_seconds", "").strip(),
        "max_rows": request.form.get("max_rows", "").strip(),
        "repair_instruction": request.form.get("repair_instruction", "").strip(),
        "rebuild_instruction": request.form.get("rebuild_instruction", "").strip(),
        "header_mapping_instruction": request.form.get("header_mapping_instruction", "").strip(),
    }


def _pdf_ai_expected_version() -> int | None:
    from .ai_repair_config import AiConfigError

    raw_value = request.form.get("expected_active_version_id", "").strip()
    if not raw_value:
        return None
    try:
        value = int(raw_value)
    except ValueError as exc:
        raise AiConfigError("页面配置版本无效，请刷新后重试。") from exc
    if value <= 0:
        raise AiConfigError("页面配置版本无效，请刷新后重试。")
    return value


def _pdf_ai_admin_context(
    *,
    form_values: dict[str, object] | None = None,
    result: dict[str, str] | None = None,
) -> dict[str, object]:
    from .ai_repair_config import (
        ai_config_form_values,
        get_active_ai_config_version_id,
        get_ai_repair_config,
        list_ai_config_versions,
        master_key_configured,
    )

    current_config = get_ai_repair_config()
    versions = list_ai_config_versions()
    active_id = get_active_ai_config_version_id()
    active_history = next((item for item in versions if item["id"] == active_id), None)
    key_configured = (
        bool(active_history and active_history["key_configured"])
        if active_id is not None
        else bool(current_config.api_key)
    )
    values = form_values or ai_config_form_values(current_config)
    values["api_key"] = ""
    return {
        "form_values": values,
        "current_config": current_config.safe_metadata(),
        "current_status": current_config.safe_status(),
        "key_configured": key_configured,
        "master_key_ready": master_key_configured(),
        "active_version_id": active_id,
        "versions": versions,
        "csrf_token": _pdf_ai_csrf_token(),
        "result": result,
    }


@bp.route("/admin/pdf-excel-ai", methods=["GET", "POST"])
def admin_pdf_excel_ai():
    from .ai_repair_config import (
        AiConfigError,
        get_ai_repair_config,
        save_ai_config_version,
        validate_ai_config_input,
    )
    from .deepseek_repair_client import DeepSeekRepairError, test_repair_connection

    access_response = require_admin_role()
    if access_response:
        return access_response
    if request.method == "GET":
        return redirect(url_for("main.pdf_excel", ai_config="1"))
    if not _valid_pdf_ai_csrf_token(request.form.get("csrf_token", "")):
        abort(400)

    action = request.form.get("action", "test").strip()
    submitted = _pdf_ai_submitted_values()
    current_config = get_ai_repair_config()
    try:
        if action == "test":
            candidate = validate_ai_config_input(submitted, current_config)
            if candidate.enabled:
                message = test_repair_connection(candidate)
            else:
                message = "AI 当前为禁用状态，无需连接测试。"
            return _render_pdf_excel_page(
                ai_form_values=submitted,
                ai_result={"category": "success", "message": message},
                open_ai_dialog=True,
            )

        employee_id = current_employee() or ""
        if not verify_user_password(employee_id, request.form.get("current_password", "")):
            raise AiConfigError("当前管理员登录密码错误。")
        expected_version = _pdf_ai_expected_version()

        if action == "save":
            candidate = validate_ai_config_input(submitted, current_config)
            if candidate.enabled:
                test_message = test_repair_connection(candidate)
                test_status = "passed"
            else:
                test_message = "AI 已禁用，未执行连接测试。"
                test_status = "not_required"
            saved = save_ai_config_version(
                candidate,
                employee_id=employee_id,
                expected_active_version_id=expected_version,
                test_status=test_status,
                test_message=test_message,
            )
            session["pdf_ai_modal_result"] = {
                "category": "success",
                "message": f"PDF/Excel AI 配置 v{saved.version_id} 已启用。",
            }
            return redirect(url_for("main.pdf_excel", ai_config="1"))

        if action == "rollback":
            try:
                target_version = int(request.form.get("version_id", ""))
            except ValueError as exc:
                raise AiConfigError("回滚版本无效。") from exc
            target = get_ai_repair_config(target_version, strict=True)
            if target.enabled:
                test_message = test_repair_connection(target)
                test_status = "passed"
            else:
                test_message = "AI 已禁用，未执行连接测试。"
                test_status = "not_required"
            saved = save_ai_config_version(
                target,
                employee_id=employee_id,
                expected_active_version_id=expected_version,
                test_status=test_status,
                test_message=test_message,
                source_version_id=target_version,
            )
            session["pdf_ai_modal_result"] = {
                "category": "success",
                "message": f"已从 v{target_version} 创建并启用回滚版本 v{saved.version_id}。",
            }
            return redirect(url_for("main.pdf_excel", ai_config="1"))

        raise AiConfigError("未知的 AI 配置操作。")
    except (AiConfigError, DeepSeekRepairError) as exc:
        return _render_pdf_excel_page(
            ai_form_values=submitted if action != "rollback" else None,
            ai_result={"category": "error", "message": str(exc)},
            open_ai_dialog=True,
        )


def _today_iso() -> str:
    return date.today().isoformat()


def _task_category_id_from_form(employee_id: str) -> int | None:
    category_id = request.form.get("category_id", type=int)
    if category_id and get_task_category(category_id, employee_id):
        return category_id
    return None


def _task_due_date_from_form() -> str | None:
    due_date = (request.form.get("due_date") or "").strip()
    return due_date or None


def _task_priority_from_form() -> str:
    priority = (request.form.get("priority") or "normal").strip()
    return priority if priority in TASK_PRIORITY_META else "normal"


def _task_progress_from_form() -> str:
    progress = (request.form.get("progress") or "not_started").strip()
    if progress == "waiting":
        return "not_started"
    return progress if progress in TASK_PROGRESS_META else "not_started"


def _progress_label(progress: str) -> str:
    if progress == "waiting":
        progress = "not_started"
    return TASK_PROGRESS_META.get(progress, TASK_PROGRESS_META["not_started"])["label"]


def _group_tasks_by_category(tasks, categories):
    groups = [
        {"group_key": f"category:{category['id']}", "category_id": category["id"], "group_label": category["name"], "tasks": []}
        for category in categories
    ]
    group_map = {group["category_id"]: group for group in groups}
    uncategorized = {"group_key": "category:none", "category_id": None, "group_label": "未分类", "tasks": []}
    for task in tasks:
        group = group_map.get(task["category_id"])
        if group is None:
            group = uncategorized
        group["tasks"].append(task)
    visible_groups = [group for group in groups if group["tasks"]]
    if uncategorized["tasks"]:
        visible_groups.append(uncategorized)
    return visible_groups


def _group_tasks_by_priority(tasks):
    groups = [
        {"group_key": f"priority:{key}", "category_id": None, "priority": key, "group_label": item["label"], "tasks": []}
        for key, item in sorted(TASK_PRIORITY_META.items(), key=lambda pair: pair[1]["rank"])
    ]
    group_map = {group["group_key"].split(":", 1)[1]: group for group in groups}
    for task in tasks:
        group_map.get(task["priority"], group_map["normal"])["tasks"].append(task)
    for group in groups:
        group["tasks"].sort(key=lambda task: (task["priority_sort_order"] or task["id"], task["id"]))
    return [group for group in groups if group["tasks"]]


def _group_tasks_by_progress(tasks):
    groups = [
        {"group_key": "progress:in_progress", "category_id": None, "group_label": "进行中", "tasks": []},
        {"group_key": "progress:not_started", "category_id": None, "group_label": "未开始", "tasks": []},
    ]
    group_map = {"in_progress": groups[0], "not_started": groups[1], "waiting": groups[1]}
    archived_group = {"group_key": "progress:completed", "category_id": None, "group_label": "已完成", "tasks": []}
    for task in tasks:
        if task["archived_at"] or task["progress"] == "completed":
            archived_group["tasks"].append(task)
        else:
            group_map.get(task["progress"], groups[1])["tasks"].append(task)
    visible_groups = [group for group in groups if group["tasks"]]
    if archived_group["tasks"]:
        visible_groups.append(archived_group)
    return visible_groups


def _group_tasks(tasks, categories, group_by: str):
    if group_by == "priority":
        return _group_tasks_by_priority(tasks)
    if group_by == "progress":
        return _group_tasks_by_progress(tasks)
    return _group_tasks_by_category(tasks, categories)


@bp.get("/tasks")
def work_planning():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    employee_id = current_employee()
    today = _today_iso()
    view = request.args.get("view", "all")
    group_by = request.args.get("group_by", "category")
    if group_by not in TASK_GROUP_META:
        group_by = "category"
    selected_category_id = request.args.get("category_id", type=int)
    selected_task_id = request.args.get("task_id", type=int)

    categories = list_task_categories(employee_id)
    archived: bool | None = False
    task_filters: dict[str, object] = {}
    if selected_category_id and get_task_category(selected_category_id, employee_id):
        task_filters["category_id"] = selected_category_id
    if view == "today":
        task_filters["due_on"] = today
    elif view == "overdue":
        task_filters["due_before"] = today
    elif view == "in_progress":
        task_filters["progress"] = "in_progress"
    elif view == "archived":
        archived = True
    else:
        view = "all"

    tasks = list_personal_tasks(employee_id, archived=archived, **task_filters)
    active_tasks = list_personal_tasks(employee_id, archived=False)
    archived_tasks = list_personal_tasks(employee_id, archived=True)
    selected_task = get_personal_task(selected_task_id, employee_id) if selected_task_id else None
    if selected_task and view != "archived" and selected_task["archived_at"]:
        selected_task = None

    stats = {
        "active": len(active_tasks),
        "today": sum(1 for task in active_tasks if task["due_date"] == today),
        "overdue": sum(1 for task in active_tasks if task["due_date"] and task["due_date"] < today),
        "in_progress": sum(1 for task in active_tasks if task["progress"] == "in_progress"),
        "archived": len(archived_tasks),
    }
    view_titles = {
        "all": "进行中项目",
        "today": "今天截止",
        "overdue": "逾期任务",
        "in_progress": "进行中任务",
        "archived": "已归档任务",
    }
    return render_template(
        "work_planning.html",
        categories=categories,
        backup_status=get_task_backup_status(employee_id),
        grouped_tasks=_group_tasks(tasks, categories, group_by),
        group_by=group_by,
        group_meta=TASK_GROUP_META,
        selected_task=selected_task,
        selected_category_id=selected_category_id,
        view=view,
        view_title=view_titles[view],
        today=today,
        stats=stats,
        priority_meta=TASK_PRIORITY_META,
        progress_meta=TASK_PROGRESS_META,
    )


@bp.post("/tasks/backup/save")
def save_task_backup_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    status = save_task_backup(current_employee())
    flash(f"当前任务清单已保存：{status['path']}", "success")
    return redirect(url_for("main.work_planning"))


@bp.post("/tasks/backup/restore")
def restore_task_backup_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    try:
        restore_task_backup(current_employee())
    except FileNotFoundError:
        flash("没有找到可恢复的任务清单。", "error")
    except Exception as exc:
        flash(f"恢复失败：{exc}", "error")
    else:
        flash("已从上次保存的任务清单恢复。", "success")
    return redirect(url_for("main.work_planning"))


@bp.post("/tasks/reorder")
def reorder_personal_tasks_view():
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    raw_ids = payload.get("ordered_ids") or []
    if not isinstance(raw_ids, list):
        return jsonify({"error": "invalid_order"}), 400
    try:
        ordered_ids = [int(item) for item in raw_ids]
    except (TypeError, ValueError):
        return jsonify({"error": "invalid_order"}), 400
    category_id = payload.get("category_id")
    sort_scope = str(payload.get("sort_scope") or "category")
    if sort_scope == "priority":
        priority = str(payload.get("priority") or "")
        if priority not in TASK_PRIORITY_META:
            return jsonify({"error": "invalid_priority"}), 400
        updated = reorder_personal_tasks_by_priority(current_employee(), ordered_ids, priority)
        return jsonify({"ok": True, "updated": updated})
    if category_id in ("", None):
        category_id = None
    else:
        try:
            category_id = int(category_id)
        except (TypeError, ValueError):
            return jsonify({"error": "invalid_category"}), 400
    updated = reorder_personal_tasks(current_employee(), ordered_ids, category_id)
    return jsonify({"ok": True, "updated": updated})


@bp.post("/tasks")
def create_personal_task_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    employee_id = current_employee()
    title = (request.form.get("title") or "").strip()
    if not title:
        flash("请输入任务标题。", "error")
        return redirect(url_for("main.work_planning"))
    task_id = create_personal_task(
        employee_id,
        title=title,
        category_id=_task_category_id_from_form(employee_id),
        description=request.form.get("description") or "",
        task_tag="",
        priority=_task_priority_from_form(),
        progress=_task_progress_from_form(),
        due_date=_task_due_date_from_form(),
    )
    flash("任务已创建。", "success")
    return redirect(url_for("main.work_planning", task_id=task_id))


@bp.post("/tasks/<int:task_id>/update")
def update_personal_task_view(task_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    employee_id = current_employee()
    title = (request.form.get("title") or "").strip()
    if not title:
        flash("任务标题不能为空。", "error")
        return redirect(url_for("main.work_planning", task_id=task_id))
    ok = update_personal_task(
        task_id,
        employee_id,
        title=title,
        category_id=_task_category_id_from_form(employee_id),
        description=request.form.get("description") or "",
        task_tag="",
        priority=_task_priority_from_form(),
        progress=_task_progress_from_form(),
        due_date=_task_due_date_from_form(),
    )
    flash("任务已更新。" if ok else "未找到该任务。", "success" if ok else "error")
    return redirect(url_for("main.work_planning"))


@bp.post("/tasks/<int:task_id>/archive")
def archive_personal_task_view(task_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    ok = archive_personal_task(task_id, current_employee())
    flash("任务已完成并归档。" if ok else "未找到该任务。", "success" if ok else "error")
    return redirect(url_for("main.work_planning"))


@bp.post("/tasks/<int:task_id>/restore")
def restore_personal_task_view(task_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    ok = restore_personal_task(task_id, current_employee())
    flash("任务已恢复到进行中。" if ok else "未找到该任务。", "success" if ok else "error")
    return redirect(url_for("main.work_planning", view="archived"))


@bp.post("/tasks/<int:task_id>/delete")
def delete_personal_task_view(task_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    ok = delete_personal_task(task_id, current_employee())
    flash("任务已删除。" if ok else "未找到该任务。", "success" if ok else "error")
    return redirect(url_for("main.work_planning"))


@bp.post("/tasks/categories")
def create_task_category_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    name = (request.form.get("name") or "").strip()
    short_label = (request.form.get("short_label") or "").strip()
    if not name:
        flash("请输入分类名称。", "error")
        return redirect(url_for("main.work_planning"))
    create_task_category(current_employee(), name, short_label)
    flash("分类已添加。", "success")
    return redirect(url_for("main.work_planning"))


@bp.post("/tasks/categories/<int:category_id>/update")
def update_task_category_view(category_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    name = (request.form.get("name") or "").strip()
    short_label = (request.form.get("short_label") or "").strip()
    if not name:
        flash("分类名称不能为空。", "error")
        return redirect(url_for("main.work_planning"))
    try:
        ok = update_task_category(category_id, current_employee(), name, short_label)
    except sqlite3.IntegrityError:
        ok = False
        flash("分类名称已存在。", "error")
    else:
        flash("分类已更新。" if ok else "未找到该分类。", "success" if ok else "error")
    return redirect(url_for("main.work_planning"))


@bp.post("/tasks/categories/<int:category_id>/delete")
def delete_task_category_view(category_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    ok = delete_task_category(category_id, current_employee())
    flash("分类已删除，原任务已移到未分类。" if ok else "未找到该分类。", "success" if ok else "error")
    return redirect(url_for("main.work_planning"))


@bp.post("/jobs")
def create_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    uploaded_file = request.files.get("excel_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("请先上传 Excel 文件。", "error")
        return redirect(_price_calculator_page_url("fangzheng"))
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        flash("仅支持 Excel 文件。", "error")
        return redirect(_price_calculator_page_url("fangzheng"))
    active_job = get_active_job(current_employee(), "fangzheng")
    if active_job:
        flash("当前已有方正任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(_price_calculator_page_url("fangzheng", job_id=active_job["id"]))
    job_id = queue_job(current_employee(), uploaded_file, original_filename, get_active_rule_version())
    flash("任务已创建，系统正在处理。", "success")
    return redirect(_price_calculator_page_url("fangzheng", job_id=job_id))


@bp.post("/transcode/jobs")
def create_transcode_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    uploaded_file = request.files.get("excel_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("请先上传 Excel 文件。", "error")
        return redirect(url_for("main.transcode"))
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        flash("转码功能仅支持 .xlsx / .xlsm / .xls 文件。", "error")
        return redirect(url_for("main.transcode"))
    active_job = get_active_job(current_employee(), "transcode")
    if active_job:
        flash("当前已有转码任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(url_for("main.transcode", job_id=active_job["id"]))
    job_id = queue_transcode_job(current_employee(), uploaded_file, original_filename)
    flash("转码任务已创建，系统正在处理。", "success")
    return redirect(url_for("main.transcode", job_id=job_id))


@bp.post("/transcode-agent/jobs")
def create_transcode_agent_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    uploaded_file = request.files.get("excel_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("请先上传 Excel 文件。", "error")
        return redirect(url_for("main.transcode_agent"))
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        flash("营销转码Agent仅支持 .xlsx / .xlsm / .xls 文件。", "error")
        return redirect(url_for("main.transcode_agent"))
    active_job = get_active_job(current_employee(), TRANSCODE_AGENT_FEATURE)
    if active_job:
        flash("当前已有营销转码Agent任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(url_for("main.transcode_agent", job_id=active_job["id"]))
    job_id = queue_transcode_agent_job(current_employee(), uploaded_file, original_filename)
    flash("营销转码Agent任务已创建，系统正在处理。", "success")
    return redirect(url_for("main.transcode_agent", job_id=job_id, auto_confirm=1))


@bp.post("/shennan/jobs")
def create_shennan_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    uploaded_file = request.files.get("excel_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("请先上传 Excel 文件。", "error")
        return redirect(_price_calculator_page_url("shennan"))
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        flash("深南价格计算仅支持 .xlsx / .xlsm / .xls 文件。", "error")
        return redirect(_price_calculator_page_url("shennan"))
    active_job = get_active_job(current_employee(), "shennan")
    if active_job:
        flash("当前已有深南任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(_price_calculator_page_url("shennan", job_id=active_job["id"]))
    job_id = queue_shennan_job(current_employee(), uploaded_file, original_filename)
    flash("深南计算任务已创建，系统正在处理。", "success")
    return redirect(_price_calculator_page_url("shennan", job_id=job_id))


@bp.post("/hushi/jobs")
def create_hushi_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    uploaded_file = request.files.get("excel_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("请先上传 Excel 文件。", "error")
        return redirect(_price_calculator_page_url("hushi"))
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        flash("沪士价格计算仅支持 .xlsx / .xlsm / .xls 文件。", "error")
        return redirect(_price_calculator_page_url("hushi"))
    active_job = get_active_job(current_employee(), "hushi")
    if active_job:
        flash("当前已有沪士任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(_price_calculator_page_url("hushi", job_id=active_job["id"]))
    job_id = queue_hushi_job(current_employee(), uploaded_file, original_filename)
    flash("沪士计算任务已创建，系统正在处理。", "success")
    return redirect(_price_calculator_page_url("hushi", job_id=job_id))


@bp.post("/bomin/jobs")
def create_bomin_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    uploaded_file = request.files.get("excel_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("请先上传 Excel 文件。", "error")
        return redirect(_price_calculator_page_url("bomin"))
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        flash("博敏价格计算仅支持 .xlsx / .xlsm / .xls 文件。", "error")
        return redirect(_price_calculator_page_url("bomin"))
    active_job = get_active_job(current_employee(), "bomin")
    if active_job:
        flash("当前已有博敏任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(_price_calculator_page_url("bomin", job_id=active_job["id"]))
    job_id = queue_bomin_job(current_employee(), uploaded_file, original_filename)
    flash("博敏计算任务已创建，系统正在处理。", "success")
    return redirect(_price_calculator_page_url("bomin", job_id=job_id))


@bp.post("/price-calculation/jobs")
def create_price_calculation_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    try:
        customer_key = _price_customer_from_request()
        quote_variant = _price_quote_variant_from_request(customer_key)
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(_price_calculator_page_url("fangzheng"))
    uploaded_file = request.files.get("excel_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("请先上传 Excel 文件。", "error")
        return redirect(_price_calculator_page_url(customer_key, quote_variant=quote_variant))
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        flash("价格计算仅支持 .xlsx / .xlsm / .xls 文件。", "error")
        return redirect(_price_calculator_page_url(customer_key, quote_variant=quote_variant))
    active_job = get_active_job(current_employee(), "price_calculation")
    if active_job:
        flash("当前已有价格计算任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(_price_calculator_page_url(customer_key, quote_variant=quote_variant, job_id=active_job["id"]))
    try:
        job_id = queue_price_calculation_job(current_employee(), customer_key, uploaded_file, original_filename, quote_variant=quote_variant)
    except Exception as exc:
        flash(f"价格计算任务创建失败：{exc}", "error")
        return redirect(_price_calculator_page_url(customer_key, quote_variant=quote_variant))
    flash("价格计算任务已创建，系统正在处理。", "success")
    return redirect(_price_calculator_page_url(customer_key, quote_variant=quote_variant, job_id=job_id))


@bp.post("/in-transit/jobs")
def create_in_transit_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    uploaded_file = request.files.get("in_transit_excel_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("请先上传 Excel 文件。", "error")
        return redirect(url_for("main.in_transit"))
    original_filename = (uploaded_file.filename or "").strip()
    if not original_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        flash("在途核对仅支持 .xlsx / .xlsm / .xls 文件。", "error")
        return redirect(url_for("main.in_transit"))
    active_job = get_active_job(current_employee(), "in_transit")
    if active_job:
        flash("当前已有在途核对任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(url_for("main.in_transit", job_id=active_job["id"]))
    job_id = queue_in_transit_job(current_employee(), uploaded_file, original_filename)
    flash("在途核对任务已创建，系统正在处理。", "success")
    return redirect(url_for("main.in_transit", job_id=job_id))


@bp.post("/inventory-detail/jobs")
def create_inventory_detail_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    requested_mode = (request.form.get("inventory_mode") or "warehouse").strip().lower()
    if requested_mode not in {"warehouse", "plan_a"}:
        flash("库存处理类型无效，请重新选择。", "error")
        return redirect(url_for("main.inventory_detail"))
    inventory_mode = normalize_inventory_mode(requested_mode)
    shanghai_file = request.files.get("shanghai_inventory_file")
    jiangxi_file = request.files.get("jiangxi_inventory_file")
    plan_a_file = request.files.get("plan_a_inventory_file")
    if inventory_mode == PLAN_A_MODE:
        required_files = [plan_a_file]
        if not plan_a_file or not plan_a_file.filename:
            flash("请先上传计划A级库存表。", "error")
            return redirect(url_for("main.inventory_detail"))
    else:
        required_files = [shanghai_file, jiangxi_file]
        if not shanghai_file or not shanghai_file.filename:
            flash("请先上传上海厂库存表。", "error")
            return redirect(url_for("main.inventory_detail"))
        if not jiangxi_file or not jiangxi_file.filename:
            flash("请先上传江西厂库存表。", "error")
            return redirect(url_for("main.inventory_detail"))
    invalid_files = [
        file_obj.filename
        for file_obj in required_files
        if file_obj and Path(file_obj.filename or "").suffix.lower() not in {".xls", ".xlsx"}
    ]
    if invalid_files:
        flash(f"库存明细仅支持 .xls / .xlsx：{', '.join(invalid_files)}", "error")
        return redirect(url_for("main.inventory_detail"))
    active_job = get_active_job(current_employee(), INVENTORY_DETAIL_FEATURE)
    if active_job:
        flash("当前已有库存明细任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(url_for("main.inventory_detail", job_id=active_job["id"]))
    try:
        job_id = queue_inventory_detail_job(
            current_employee(),
            shanghai_file,
            jiangxi_file,
            inventory_mode=inventory_mode,
            plan_a_file=plan_a_file,
        )
    except Exception as exc:
        flash(f"库存明细任务创建失败：{exc}", "error")
        return redirect(url_for("main.inventory_detail"))
    mode_label = "计划A级分类" if inventory_mode == PLAN_A_MODE else "仓库B级分类"
    flash(f"{mode_label}任务已创建，系统正在处理。", "success")
    return redirect(url_for("main.inventory_detail", job_id=job_id))


@bp.post("/inventory-bid/jobs")
def create_inventory_bid_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    shanghai_file = request.files.get("shanghai_inventory_file")
    jiangxi_file = request.files.get("jiangxi_inventory_file")
    if not shanghai_file or not shanghai_file.filename:
        flash("请先上传上海库存表。", "error")
        return redirect(url_for("main.inventory_bid"))
    if not jiangxi_file or not jiangxi_file.filename:
        flash("请先上传江西库存表。", "error")
        return redirect(url_for("main.inventory_bid"))
    invalid_files = [
        file_obj.filename
        for file_obj in [shanghai_file, jiangxi_file]
        if file_obj and Path(file_obj.filename or "").suffix.lower() not in {".xls", ".xlsx"}
    ]
    if invalid_files:
        flash(f"库存竞标仅支持 .xls / .xlsx：{', '.join(invalid_files)}", "error")
        return redirect(url_for("main.inventory_bid"))
    active_job = get_active_job(current_employee(), INVENTORY_BID_FEATURE)
    if active_job:
        flash("当前已有库存竞标任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(url_for("main.inventory_bid", job_id=active_job["id"]))
    try:
        job_id = queue_inventory_bid_job(current_employee(), shanghai_file, jiangxi_file)
    except Exception as exc:
        flash(f"库存竞标任务创建失败：{exc}", "error")
        return redirect(url_for("main.inventory_bid"))
    flash("库存竞标任务已创建，系统正在处理。", "success")
    return redirect(url_for("main.inventory_bid", job_id=job_id))


@bp.post("/inventory-bid/max-jobs")
def create_inventory_bid_max_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    bid_file = request.files.get("bid_result_file")
    if not bid_file or not bid_file.filename:
        flash("请先上传已填写报价的库存竞标汇总表。", "error")
        return redirect(url_for("main.inventory_bid"))
    if Path(bid_file.filename or "").suffix.lower() not in {".xls", ".xlsx"}:
        flash("库存竞标取最大值仅支持 .xls / .xlsx 文件。", "error")
        return redirect(url_for("main.inventory_bid"))
    active_job = get_active_job(current_employee(), INVENTORY_BID_FEATURE)
    if active_job:
        flash("当前已有库存竞标任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(url_for("main.inventory_bid", job_id=active_job["id"]))
    try:
        job_id = queue_inventory_bid_max_job(current_employee(), bid_file)
    except Exception as exc:
        flash(f"库存竞标取最大值任务创建失败：{exc}", "error")
        return redirect(url_for("main.inventory_bid"))
    flash("库存竞标取最大值任务已创建，系统正在处理。", "success")
    return redirect(url_for("main.inventory_bid", job_id=job_id))


@bp.post("/pdf-excel/jobs")
def create_pdf_excel_job_view():
    from .pdf_excel_service import (
        ALLOWED_EXTENSIONS as PDF_EXCEL_ALLOWED_EXTENSIONS,
        FEATURE as PDF_EXCEL_FEATURE,
        queue_pdf_excel_job,
    )

    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    uploaded_files = [file_obj for file_obj in request.files.getlist("pdf_excel_files") if file_obj and file_obj.filename]
    if not uploaded_files:
        flash("请先上传 PDF 或图片文件。", "error")
        return redirect(url_for("main.pdf_excel"))
    invalid_files = [
        file_obj.filename
        for file_obj in uploaded_files
        if Path(file_obj.filename or "").suffix.lower() not in PDF_EXCEL_ALLOWED_EXTENSIONS
    ]
    if invalid_files:
        flash(f"不支持的文件类型：{', '.join(invalid_files)}", "error")
        return redirect(url_for("main.pdf_excel"))
    active_job = get_active_job(current_employee(), PDF_EXCEL_FEATURE)
    if active_job:
        flash("当前已有 PDF/图片转Excel 任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(url_for("main.pdf_excel", job_id=active_job["id"]))
    try:
        job_id = queue_pdf_excel_job(
            current_employee(),
            uploaded_files,
            max_workers=2,
        )
    except Exception as exc:
        flash(f"PDF/图片转Excel 任务创建失败：{exc}", "error")
        return redirect(url_for("main.pdf_excel"))
    flash("PDF/图片转Excel 任务已创建，系统正在处理。", "success")
    return redirect(url_for("main.pdf_excel", job_id=job_id))


@bp.post("/order-reprice/jobs")
def create_order_reprice_job_view():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    mode = request.form.get("mode", "").strip()
    if mode not in ORDER_REPRICE_MODE_LABELS:
        flash("请选择订单改价处理块。", "error")
        return redirect(url_for("main.order_reprice"))

    customer_file = request.files.get("customer_file")
    factory_file = request.files.get("factory_file")
    quote_files = [file_obj for file_obj in request.files.getlist("quote_files") if file_obj and file_obj.filename]

    def _valid_excel(file_obj) -> bool:
        return bool(file_obj and file_obj.filename and file_obj.filename.lower().endswith((".xlsx", ".xlsm", ".xls")))

    if not _valid_excel(customer_file):
        flash("请上传客户明细 Excel 文件。", "error")
        return redirect(url_for("main.order_reprice"))
    if mode != "block2" and not _valid_excel(factory_file):
        flash("请上传客户明细和厂内明细 Excel 文件。", "error")
        return redirect(url_for("main.order_reprice"))
    if mode == "block2" and not quote_files:
        flash("第二块价格核对需要至少上传一份胜宏报价单。", "error")
        return redirect(url_for("main.order_reprice"))
    if any(not _valid_excel(file_obj) for file_obj in quote_files):
        flash("报价单仅支持 .xlsx / .xlsm / .xls 文件。", "error")
        return redirect(url_for("main.order_reprice"))

    active_job = get_active_job(current_employee(), "order_reprice")
    if active_job:
        flash("当前已有订单改价任务正在处理，请先等待完成或停止后再上传。", "error")
        return redirect(url_for("main.order_reprice", job_id=active_job["id"]))
    try:
        job_id = queue_order_reprice_job(current_employee(), mode, customer_file, factory_file, quote_files)
    except Exception as exc:
        flash(f"订单改价任务创建失败：{exc}", "error")
        return redirect(url_for("main.order_reprice"))
    flash("订单改价任务已创建，系统正在处理。", "success")
    return redirect(url_for("main.order_reprice", job_id=job_id))


@bp.get("/jobs/<int:job_id>")
def job_detail(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    job = get_job(job_id)
    if not job or job["employee_id"] != current_employee():
        flash("未找到该任务。", "error")
        return redirect(url_for("main.history"))
    return render_template("job_detail.html", job=_decorate_job(job))


@bp.get("/api/jobs/<int:job_id>")
def api_job_detail(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    job = get_job(job_id)
    if not job or job["employee_id"] != current_employee():
        return jsonify({"error": "not_found"}), 404
    return jsonify(_decorate_job(job))


def _api_quote_response(calculator_func, *, allow_extra_fields: bool = False):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    spec = str(payload.get("spec") or request.form.get("spec") or "").strip()
    if not spec:
        return jsonify({"status": "失败", "price": None, "result": None, "error": "请输入客户规格"}), 400
    try:
        if allow_extra_fields:
            data = calculator_func(
                spec,
                customer=payload.get("customer") or request.form.get("customer") or "",
                customer_code=payload.get("customer_code") or request.form.get("customer_code") or "",
                order_text=payload.get("order_text") or request.form.get("order_text") or "",
            )
        else:
            data = calculator_func(spec)
        status_code = 200 if data.get("status") != "失败" else 422
        return jsonify(data), status_code
    except Exception as exc:
        return jsonify({"status": "失败", "price": None, "result": None, "error": str(exc)}), 500


@bp.post("/api/fangzheng/quote")
def api_fangzheng_quote():
    return _api_quote_response(calculate_fangzheng_quote)


@bp.post("/api/transcode/quote")
def api_transcode_quote():
    return _api_quote_response(calculate_transcode_quote, allow_extra_fields=True)


@bp.post("/api/transcode-agent/quote")
def api_transcode_agent_quote():
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    spec = str(payload.get("spec") or request.form.get("spec") or "").strip()
    if not spec:
        return jsonify({"status": "失败", "result": None, "error": "请输入客户规格"}), 400
    try:
        data = calculate_transcode_agent_quote(
            spec,
            customer=payload.get("customer") or request.form.get("customer") or "",
            customer_code=payload.get("customer_code") or request.form.get("customer_code") or "",
            order_text=payload.get("order_text") or request.form.get("order_text") or "",
            order_remark=payload.get("order_remark") or request.form.get("order_remark") or "",
            employee_id=current_employee() or "",
        )
        return jsonify(data), 200 if data.get("status") != "失败" else 422
    except Exception as exc:
        return jsonify({"status": "失败", "result": None, "error": str(exc)}), 500


@bp.post("/api/internal/transcode-agent/quote")
def api_internal_transcode_agent_quote():
    auth_error = _authorize_internal_transcode_agent()
    if auth_error:
        return auth_error

    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return _internal_transcode_agent_error("请求体必须是 JSON 对象", 400)

    spec = str(payload.get("spec") or "").strip()
    if not spec:
        return _internal_transcode_agent_error("请输入客户规格", 400)

    try:
        data = calculate_transcode_agent_quote(
            spec,
            customer=str(payload.get("customer") or "").strip(),
            customer_code=str(payload.get("customer_code") or "").strip(),
            order_remark=str(payload.get("order_remark") or "").strip(),
            employee_id=os.environ.get("TRANSCODE_AGENT_INTERNAL_EMPLOYEE_ID", "").strip(),
        )
        normalized = _normalize_internal_transcode_agent_quote(data)
        status_code = 200 if normalized["status"] in {"成功", "待确认"} else 422
        return jsonify(normalized), status_code
    except Exception as exc:
        return _internal_transcode_agent_error(f"内部转码执行失败：{exc}", 500)


@bp.post("/api/transcode-agent/single-jobs")
def api_create_transcode_agent_single_job():
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    spec = str(payload.get("spec") or "").strip()
    if not spec:
        return jsonify({"status": "失败", "error": "请输入客户规格"}), 400
    employee_id = current_employee() or ""
    active_job = get_active_job(employee_id, TRANSCODE_AGENT_FEATURE)
    if active_job:
        return (
            jsonify(
                {
                    "status": "失败",
                    "error": "当前已有营销转码Agent任务正在处理，请先等待完成或停止。",
                    "job_id": active_job["id"],
                    "task_url": url_for(
                        "main.transcode_agent",
                        job_id=active_job["id"],
                    ),
                }
            ),
            409,
        )
    try:
        job_id = queue_transcode_agent_single_job(
            employee_id,
            spec=spec,
            customer=payload.get("customer") or "",
            customer_code=payload.get("customer_code") or "",
            order_remark=payload.get("order_remark") or "",
        )
        return (
            jsonify(
                {
                    "status": "已创建",
                    "job_id": job_id,
                    "task_url": url_for(
                        "main.transcode_agent",
                        job_id=job_id,
                        auto_confirm=1,
                    ),
                }
            ),
            202,
        )
    except Exception as exc:
        return jsonify({"status": "失败", "error": str(exc)}), 500


@bp.get("/api/transcode-agent/jobs/<int:job_id>/confirmations")
def api_transcode_agent_confirmations(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    try:
        return jsonify(
            list_transcode_agent_confirmations(
                job_id,
                current_employee() or "",
                record_scope=request.args.get("scope", "all"),
                record_page=request.args.get("page", 1, type=int) or 1,
                record_page_size=request.args.get("page_size", 200, type=int) or 200,
            )
        )
    except LookupError as exc:
        return jsonify({"error": str(exc)}), 404


@bp.post("/api/transcode-agent/confirmations/<int:item_id>/confirm")
def api_confirm_transcode_agent_item(item_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    try:
        data = confirm_transcode_agent_item(
            item_id,
            current_employee() or "",
            confirmed_code=str(payload.get("confirmed_code") or ""),
            basis=str(payload.get("basis") or "").strip(),
            save_long_term=bool(payload.get("save_long_term")),
            long_term_rule=payload.get("long_term_rule")
            if isinstance(payload.get("long_term_rule"), dict)
            else None,
        )
        return jsonify(data)
    except LookupError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 409


@bp.post("/api/transcode-agent/confirmations/<int:item_id>/skip")
def api_skip_transcode_agent_confirmation_row(item_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    try:
        return jsonify(
            skip_transcode_agent_confirmation_row(
                item_id,
                current_employee() or "",
            )
        )
    except LookupError as exc:
        return jsonify({"error": str(exc)}), 404


@bp.post("/api/transcode-agent/rows/<int:job_id>/<int:excel_row>/verify")
def api_verify_transcode_agent_row(job_id: int, excel_row: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    try:
        return jsonify(
            verify_transcode_agent_row(
                job_id,
                excel_row,
                current_employee() or "",
                code=str(payload.get("code") or ""),
                basis=str(payload.get("basis") or "").strip(),
            )
        )
    except LookupError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 409


@bp.post("/api/transcode-agent/jobs/<int:job_id>/verify-all")
def api_verify_all_transcode_agent_rows(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    try:
        return jsonify(
            verify_all_transcode_agent_rows(
                job_id,
                current_employee() or "",
                basis=str(payload.get("basis") or "").strip(),
            )
        )
    except LookupError as exc:
        return jsonify({"error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 409


@bp.post("/api/transcode-agent/jobs/<int:job_id>/finalize-pending")
def api_finalize_transcode_agent_confirmations(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    try:
        return jsonify(
            finalize_transcode_agent_confirmations(
                job_id,
                current_employee() or "",
            )
        )
    except LookupError as exc:
        return jsonify({"error": str(exc)}), 404


@bp.get("/api/transcode-agent/pending-rules")
def api_transcode_agent_pending_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    employee_id = current_employee() or ""
    return jsonify(
        {
            "rules": list_transcode_agent_pending_rules(
                employee_id,
                include_all=is_admin_user(employee_id),
            )
        }
    )


@bp.post("/api/transcode-agent/pending-rules/<int:pending_rule_id>/update")
def api_update_transcode_agent_pending_rule(pending_rule_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    try:
        return jsonify(
            update_transcode_agent_pending_rule(
                pending_rule_id,
                current_employee() or "",
                payload,
            )
        )
    except (LookupError, CustomerRuleMaintenanceError, ValueError) as exc:
        return jsonify({"error": str(exc)}), 409


@bp.post("/api/transcode-agent/pending-rules/<int:pending_rule_id>/activate")
def api_activate_transcode_agent_pending_rule(pending_rule_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    try:
        return jsonify(
            activate_transcode_agent_pending_rule(
                pending_rule_id,
                current_employee() or "",
            )
        )
    except (LookupError, CustomerRuleMaintenanceError) as exc:
        return jsonify({"error": str(exc)}), 409


@bp.post("/api/transcode-agent/pending-rules/<int:pending_rule_id>/delete")
def api_delete_transcode_agent_pending_rule(pending_rule_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    try:
        delete_transcode_agent_pending_rule(
            pending_rule_id,
            current_employee() or "",
        )
        return jsonify({"ok": True})
    except (LookupError, CustomerRuleMaintenanceError) as exc:
        return jsonify({"error": str(exc)}), 409


@bp.post("/api/transcode-agent/jobs/<int:job_id>/reevaluate-pending")
def api_reevaluate_transcode_agent_confirmations(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    try:
        return jsonify(
            reevaluate_transcode_agent_confirmations(
                job_id,
                current_employee() or "",
            )
        )
    except LookupError as exc:
        return jsonify({"error": str(exc)}), 404


@bp.post("/api/shennan/quote")
def api_shennan_quote():
    return _api_quote_response(calculate_shennan_quote)


@bp.post("/api/hushi/quote")
def api_hushi_quote():
    return _api_quote_response(calculate_hushi_quote)


@bp.post("/api/bomin/quote")
def api_bomin_quote():
    if not _bomin_quote_authorized():
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    spec = str(payload.get("spec") or request.form.get("spec") or "").strip()
    if not spec:
        return jsonify({"status": "失败", "price": None, "error": "请输入客户规格"}), 400
    try:
        return jsonify(calculate_bomin_quote(spec))
    except Exception as exc:
        return jsonify({"status": "失败", "price": None, "error": str(exc)}), 500


@bp.post("/api/price-calculation/quote")
def api_price_calculation_quote():
    redirect_resp = require_login()
    if redirect_resp:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(silent=True) or {}
    spec = str(payload.get("spec") or request.form.get("spec") or "").strip()
    customer_key = str(
        payload.get("customer_key") or request.form.get("customer_key") or request.args.get("customer_key") or default_price_customer_key()
    ).strip()
    quote_variant = normalize_price_quote_variant(
        customer_key,
        payload.get("quote_variant") or request.form.get("quote_variant") or request.args.get("quote_variant"),
    )
    quantity = payload.get("quantity") or request.form.get("quantity")
    if not spec:
        return jsonify({"status": "失败", "price": None, "error": "请输入客户规格"}), 400
    try:
        data = calculate_price_quote(customer_key, spec, quantity=quantity, quote_variant=quote_variant)
        status_code = 200 if data.get("status") != "失败" else 422
        return jsonify(data), status_code
    except Exception as exc:
        return jsonify({"status": "失败", "price": None, "error": str(exc)}), 500


def _job_feature_return_url(job, job_id: int) -> str:
    if not job:
        return url_for("main.history")
    feature = job["feature"]
    if feature in {"fangzheng", "bomin", "shennan", "hushi"}:
        return _price_calculator_page_url(feature, job_id=job_id)
    if feature == "price_calculation":
        decorated = _decorate_job(job)
        return _price_calculator_page_url(
            decorated.get("price_customer_key") or default_price_customer_key(),
            job_id=job_id,
            quote_variant=decorated.get("quote_variant"),
        )
    feature_route = {
        "transcode": "main.transcode",
        "transcode_agent": "main.transcode_agent",
        "in_transit": "main.in_transit",
        "inventory_detail": "main.inventory_detail",
        "inventory_bid": "main.inventory_bid",
        "order_reprice": "main.order_reprice",
        "pdf_excel": "main.pdf_excel",
        "transcode_special_import": "main.admin_transcode_special_rules",
    }.get(feature, "main.history")
    return url_for(feature_route, job_id=job_id)


@bp.route("/jobs/<int:job_id>/cancel", methods=["GET", "POST"])
def cancel_job(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if request.method == "GET":
        job = get_job(job_id)
        return redirect(_job_feature_return_url(job, job_id))
    ok, message = cancel_job_process(job_id, current_employee())
    flash(message, "success" if ok else "error")
    job = get_job(job_id)
    return redirect(_job_feature_return_url(job, job_id))


@bp.get("/history")
def history():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    start_date = request.args.get("start_date") or None
    end_date = request.args.get("end_date") or None
    jobs = _decorate_jobs(list_jobs(current_employee(), start_date=start_date, end_date=end_date, limit=100))
    return render_template("history.html", jobs=jobs, start_date=start_date, end_date=end_date)


@bp.post("/history/<int:job_id>/delete")
def delete_history(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    job = get_job(job_id)
    if not job or job["employee_id"] != current_employee():
        flash("未找到该任务。", "error")
        return redirect(url_for("main.history"))
    deleted = delete_job(job_id)
    if deleted:
        if deleted["feature"] == INVENTORY_DETAIL_FEATURE:
            cleanup_inventory_detail_job_files(deleted)
        elif deleted["feature"] == INVENTORY_BID_FEATURE:
            cleanup_inventory_bid_job_files(deleted)
        else:
            for path_key in ["stored_input_path", "stored_result_path"]:
                safe_unlink(deleted[path_key])
        flash("历史记录已删除。", "success")
    return redirect(url_for("main.history"))


@bp.get("/inventory-detail/jobs/<int:job_id>/download/<grade>")
def download_inventory_detail(job_id: int, grade: str):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    job = get_job(job_id)
    if not job or job["employee_id"] != current_employee() or job["feature"] != INVENTORY_DETAIL_FEATURE:
        flash("未找到该库存明细任务。", "error")
        return redirect(url_for("main.history"))
    file_path = get_inventory_result_path(job, grade)
    if not file_path:
        flash("库存明细结果文件不存在。", "error")
        return redirect(url_for("main.job_detail", job_id=job_id))
    return send_file(file_path, as_attachment=True, download_name=file_path.name)


@bp.get("/download/<int:job_id>/<kind>")
def download(job_id: int, kind: str):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    job = get_job(job_id)
    if not job or job["employee_id"] != current_employee():
        flash("未找到该任务。", "error")
        return redirect(url_for("main.history"))
    if job["feature"] == INVENTORY_DETAIL_FEATURE and kind == "result":
        manifest = load_inventory_input_manifest(job)
        result_key = "plan-a" if normalize_inventory_mode(manifest.get("inventory_mode")) == PLAN_A_MODE else "a"
        return redirect(url_for("main.download_inventory_detail", job_id=job_id, grade=result_key))
    file_path = job["stored_result_path"] if kind == "result" else job["stored_input_path"] if kind == "input" else None
    if not file_path or not Path(file_path).exists():
        flash("文件不存在。", "error")
        return redirect(url_for("main.history"))
    if kind == "result" and job["feature"] == TRANSCODE_AGENT_FEATURE:
        try:
            refresh_transcode_agent_audit_sheet(job_id)
        except (OSError, ValueError):
            pass
    return send_file(file_path, as_attachment=True)


@bp.get("/pdf-excel/jobs/<int:job_id>/download/internal-sales")
def download_pdf_excel_internal_sales(job_id: int):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    job = get_job(job_id)
    if not job or job["employee_id"] != current_employee() or job["feature"] != PDF_EXCEL_FEATURE:
        flash("未找到该 PDF 转 Excel 任务。", "error")
        return redirect(url_for("main.history"))
    if job["status"] != "completed":
        flash("任务尚未完成，暂时不能下载内销模板。", "error")
        return redirect(url_for("main.job_detail", job_id=job_id))
    try:
        from .pdf_excel_service import get_or_create_internal_sales_result

        file_path = get_or_create_internal_sales_result(dict(job))
    except (FileNotFoundError, OSError, ValueError) as exc:
        flash(str(exc), "error")
        return redirect(url_for("main.job_detail", job_id=job_id))
    return send_file(file_path, as_attachment=True, download_name=file_path.name)


@bp.route("/admin/rules", methods=["GET", "POST"])
def admin_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if request.method == "POST":
        admin_password = request.form.get("admin_password", "")
        price_file = request.files.get("price_file")
        account_file = request.files.get("account_file")
        remark = request.form.get("remark", "").strip()
        if not verify_admin_password(admin_password):
            flash("管理员密码错误。", "error")
        elif not any(file_obj and file_obj.filename for file_obj in [price_file, account_file]):
            flash("请至少上传一份需要更新的规则文件。", "error")
        else:
            version = save_new_rule_version(price_file, account_file, updated_by=current_employee(), remark=remark)
            flash(f"方正规则已更新，当前生效版本：{version}", "success")
            return redirect(url_for("main.admin_rules"))
    price_path, account_path = get_rule_file_paths()
    return render_template(
        "admin_rules.html",
        active_rule_version=get_active_rule_version(),
        rule_history=get_rule_history(),
        price_path=price_path.name,
        account_path=account_path.name,
    )


@bp.route("/admin/transcode-rules", methods=["GET", "POST"])
def admin_transcode_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if request.method == "POST":
        admin_password = request.form.get("admin_password", "")
        remark = request.form.get("remark", "").strip()
        sheet_files = {sheet: request.files.get(f"sheet_{idx}") for idx, sheet in enumerate(TRANSCODE_RULE_SHEETS)}
        if not verify_admin_password(admin_password):
            flash("管理员密码错误。", "error")
        else:
            try:
                version = save_new_transcode_rule_version_from_sheets(
                    sheet_files,
                    updated_by=current_employee(),
                    remark=remark,
                )
                flash(f"转码规则已更新，当前生效版本：{version}", "success")
                return redirect(url_for("main.admin_transcode_rules"))
            except Exception as exc:
                flash(f"转码规则更新失败：{exc}", "error")
    return render_template(
        "admin_transcode_rules.html",
        active_rule_version=get_active_transcode_rule_version(),
        rule_history=get_transcode_rule_history(),
        rule_path=get_transcode_rule_file_path().name,
        sheets=TRANSCODE_RULE_SHEETS,
    )


@bp.route("/admin/transcode-agent-rules", methods=["GET", "POST"])
def admin_transcode_agent_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if request.method == "POST":
        admin_password = request.form.get("admin_password", "")
        remark = request.form.get("remark", "").strip()
        rule_file = request.files.get("agent_rule_file")
        if not verify_admin_password(admin_password):
            flash("管理员密码错误。", "error")
        elif not rule_file or not rule_file.filename:
            flash("请上传客户特殊清单结构化 Excel 文件。", "error")
        else:
            try:
                version = save_new_transcode_agent_rule_version(
                    rule_file,
                    updated_by=current_employee(),
                    remark=remark,
                )
                flash(f"营销转码Agent规则已更新，当前生效版本：{version}", "success")
                return redirect(url_for("main.admin_transcode_agent_rules"))
            except Exception as exc:
                flash(f"营销转码Agent规则更新失败：{exc}", "error")
    return render_template(
        "admin_transcode_agent_rules.html",
        active_rule_version=get_active_transcode_agent_rule_version() or "未上传",
        base_rule_version=get_active_transcode_rule_version(),
        rule_history=get_transcode_agent_rule_history(),
        rule_path=get_transcode_agent_rule_file_path().name,
        rule_count=get_transcode_agent_rule_count(),
    )


@bp.route("/admin/transcode-agent-rules/export/<export_type>")
def export_transcode_agent_rule_package(export_type: str):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    names = {
        "full": "营销转码Agent完整规则包.xlsx",
        "machine": "营销转码Agent机器规则.xlsx",
        "original": "客户特殊清单原文件.xlsx",
    }
    if export_type not in names:
        flash("未知导出类型。", "error")
        return redirect(url_for("main.admin_transcode_agent_rules"))
    try:
        path = export_transcode_agent_rules(export_type)
    except Exception as exc:
        flash(str(exc), "error")
        return redirect(url_for("main.admin_transcode_agent_rules"))
    return send_file(
        path,
        as_attachment=True,
        download_name=names[export_type],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _collect_special_rule_drafts_from_form() -> list[dict[str, str]]:
    try:
        draft_count = int(request.form.get("draft_count", "0") or 0)
    except ValueError:
        draft_count = 0

    drafts: list[dict[str, str]] = []
    for row_idx in range(draft_count):
        if request.form.get(f"draft_{row_idx}_delete"):
            continue
        row = {
            header: request.form.get(f"draft_{row_idx}_{col_idx}", "").strip()
            for col_idx, header in enumerate(STRUCTURED_HEADERS)
        }
        if any(row.get(header) for header in STRUCTURED_HEADERS if header != "规则ID"):
            drafts.append(row)
    return drafts


@bp.route("/admin/transcode-special-rules", methods=["GET", "POST"])
def admin_transcode_special_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp

    drafts: list[dict[str, str]] = []
    search_results: list[dict[str, str]] = []
    search_view = None
    form_values = {
        "customer_code": request.form.get("customer_code", "").strip(),
        "customer_name": request.form.get("customer_name", "").strip(),
        "material_type": request.form.get("material_type", "全部").strip() or "全部",
        "source_line": request.form.get("source_line", "").strip(),
        "requirement_text": request.form.get("requirement_text", "").strip(),
        "search_query": request.values.get("search_query", "").strip(),
    }

    if request.method == "POST":
        action = request.form.get("action", "parse")
        if action == "save":
            drafts = _collect_special_rule_drafts_from_form()
            admin_password = request.form.get("admin_password", "")
            import_mode = request.form.get("import_mode", "追加")
            if not drafts:
                flash("没有可保存的结构化规则草稿。", "error")
            elif not verify_admin_password(admin_password):
                flash("管理员密码错误。", "error")
            else:
                try:
                    _, saved_ids = save_structured_special_rules(
                        drafts,
                        saved_by=current_employee() or "",
                        import_mode=import_mode,
                    )
                    flash(f"已按“{import_mode}”保存 {len(saved_ids)} 条结构化特殊规则：{', '.join(saved_ids)}", "success")
                    return redirect(url_for("main.admin_transcode_special_rules"))
                except Exception as exc:
                    flash(f"结构化特殊规则保存失败：{exc}", "error")
        elif action == "toggle_structured_rules":
            admin_password = request.form.get("admin_password", "")
            if not verify_admin_password(admin_password):
                flash("管理员密码错误。", "error")
            else:
                enabled = request.form.get("structured_rules_enabled") == "on"
                set_structured_special_rules_enabled(enabled)
                flash(
                    f"结构化客户特殊规则参与转码已{'开启' if enabled else '关闭'}。",
                    "success",
                )
                return redirect(url_for("main.admin_transcode_special_rules"))
        elif action == "bulk_import":
            admin_password = request.form.get("admin_password", "")
            upload = request.files.get("bulk_rule_file")
            if not verify_admin_password(admin_password):
                flash("管理员密码错误。", "error")
            elif not upload or not upload.filename:
                flash("请先选择客户特殊要求 Excel 文件。", "error")
            elif not upload.filename.lower().endswith((".xlsx", ".xlsm")):
                flash("请上传 .xlsx 或 .xlsm 文件。", "error")
            else:
                try:
                    active_job = get_active_job(current_employee(), SPECIAL_IMPORT_FEATURE)
                    if active_job:
                        flash("当前已有客户特殊规则导入任务正在处理，请先等待完成或停止后再上传。", "error")
                        return redirect(url_for("main.admin_transcode_special_rules", job_id=active_job["id"]))
                    job_id = queue_transcode_special_import_job(current_employee(), upload, upload.filename)
                    flash("批量导入任务已创建，系统正在后台解析。", "success")
                    return redirect(url_for("main.admin_transcode_special_rules", job_id=job_id))
                except Exception as exc:
                    flash(f"批量导入失败：{exc}", "error")
        elif action == "search":
            search_results = search_structured_special_rules(form_values["search_query"])
            search_view = build_rule_workspace_view(search_results) if search_results else None
            flash(f"查询到 {len(search_results)} 条规则。", "success" if search_results else "error")
        else:
            drafts = parse_special_requirement(
                requirement_text=form_values["requirement_text"],
                customer_code=form_values["customer_code"],
                customer_name=form_values["customer_name"],
                material_type=form_values["material_type"],
                source_line=form_values["source_line"],
            )
            if drafts:
                flash(f"已生成 {len(drafts)} 条结构化规则草稿，请复核后保存。", "success")
            else:
                flash("请先填写特殊需求文本。", "error")
    elif form_values["search_query"]:
        search_results = search_structured_special_rules(form_values["search_query"])
        search_view = build_rule_workspace_view(search_results) if search_results else None

    draft_view = build_rule_workspace_view(drafts) if drafts else None
    has_existing_customer = customer_has_rules(form_values["customer_code"], form_values["customer_name"]) if (form_values["customer_code"] or form_values["customer_name"]) else False
    import_jobs = list_jobs(current_employee(), limit=8, feature=SPECIAL_IMPORT_FEATURE)
    return render_template(
        "admin_transcode_special_rules.html",
        headers=STRUCTURED_HEADERS,
        drafts=drafts,
        draft_view=draft_view,
        search_results=search_results,
        search_view=search_view,
        form_values=form_values,
        material_types=["全部", "基板", "PP", "基板/PP", "其他"],
        rule_count=get_structured_rule_count(),
        rule_path=get_structured_special_rules_path(),
        has_existing_customer=has_existing_customer,
        structured_rules_enabled=is_structured_special_rules_enabled(),
        special_rule_settings=get_structured_special_rule_settings(),
        latest_original_import_path=get_latest_original_import_path(),
        import_jobs=import_jobs,
        active_import_job=_active_job_for(SPECIAL_IMPORT_FEATURE, import_jobs),
    )


@bp.route("/admin/transcode-special-rules/export/<export_type>")
def export_transcode_special_rules(export_type: str):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    export_names = {
        "full": "客户特殊规则完整规则包.xlsx",
        "structured": "结构化规则总表.xlsx",
        "customer_summary": "客户汇总视图.xlsx",
        "transcode": "参与转码规则.xlsx",
        "manual": "人工确认清单.xlsx",
        "original": "客户特殊要求原文件.xlsx",
    }
    if export_type not in export_names:
        flash("未知导出类型。", "error")
        return redirect(url_for("main.admin_transcode_special_rules"))
    if export_type == "original":
        original_path = get_latest_original_import_path()
        if not original_path.exists():
            flash("还没有批量导入过客户特殊要求原文件。", "error")
            return redirect(url_for("main.admin_transcode_special_rules"))
        settings = get_structured_special_rule_settings()
        return send_file(
            original_path,
            as_attachment=True,
            download_name=settings.get("latest_original_filename") or export_names[export_type],
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    output = build_export_workbook(export_type)
    return send_file(
        output,
        as_attachment=True,
        download_name=export_names[export_type],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/admin/shennan-rules", methods=["GET", "POST"])
def admin_shennan_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if request.method == "POST":
        admin_password = request.form.get("admin_password", "")
        rule_file = request.files.get("rule_file")
        remark = request.form.get("remark", "").strip()
        if not verify_admin_password(admin_password):
            flash("管理员密码错误。", "error")
        elif not rule_file or not rule_file.filename:
            flash("请上传深南报价规则文件。", "error")
        else:
            try:
                version = save_new_shennan_rule_version(rule_file, updated_by=current_employee(), remark=remark)
                flash(f"深南规则已更新，当前生效版本：{version}", "success")
                return redirect(url_for("main.admin_shennan_rules"))
            except Exception as exc:
                flash(f"深南规则更新失败：{exc}", "error")
    return render_template(
        "admin_shennan_rules.html",
        active_rule_version=get_active_shennan_rule_version(),
        rule_history=get_shennan_rule_history(),
        rule_path=get_shennan_rule_file_path().name,
    )


@bp.route("/admin/hushi-rules", methods=["GET", "POST"])
def admin_hushi_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if request.method == "POST":
        admin_password = request.form.get("admin_password", "")
        rule_zip = request.files.get("rule_zip")
        remark = request.form.get("remark", "").strip()
        if not verify_admin_password(admin_password):
            flash("管理员密码错误。", "error")
        elif not rule_zip or not rule_zip.filename:
            flash("请上传沪士报价规则 ZIP 包。", "error")
        else:
            try:
                version = save_new_hushi_rule_version(rule_zip, updated_by=current_employee(), remark=remark)
                flash(f"沪士规则已更新，当前生效版本：{version}", "success")
                return redirect(url_for("main.admin_hushi_rules"))
            except Exception as exc:
                flash(f"沪士规则更新失败：{exc}", "error")
    rule_dir = get_hushi_rule_dir(get_active_hushi_rule_version())
    rule_files = []
    if rule_dir.exists():
        rule_files = sorted(path.name for path in rule_dir.iterdir() if path.is_file())
    return render_template(
        "admin_hushi_rules.html",
        active_rule_version=get_active_hushi_rule_version() or "未上传沪士规则",
        rule_history=get_hushi_rule_history(),
        rule_files=rule_files,
    )


@bp.route("/admin/bomin-rules", methods=["GET", "POST"])
def admin_bomin_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if request.method == "POST":
        admin_password = request.form.get("admin_password", "")
        rule_file = request.files.get("rule_file")
        remark = request.form.get("remark", "").strip()
        if not verify_admin_password(admin_password):
            flash("管理员密码错误。", "error")
        elif not rule_file or not rule_file.filename:
            flash("请上传博敏价格表 Excel 文件。", "error")
        else:
            try:
                version = save_new_bomin_rule_version(rule_file, updated_by=current_employee(), remark=remark)
                flash(f"博敏价格表已更新，当前生效版本：{version}", "success")
                return redirect(url_for("main.admin_bomin_rules"))
            except Exception as exc:
                flash(f"博敏价格表更新失败：{exc}", "error")
    return render_template(
        "admin_bomin_rules.html",
        active_rule_version=get_active_bomin_rule_version() or "未初始化博敏价格表",
        rule_history=get_bomin_rule_history(),
        rule_path=get_bomin_rule_file_path(get_active_bomin_rule_version()).name
        if get_active_bomin_rule_version()
        else "bomin_price_rules.xlsx",
    )


@bp.route("/admin/price-calculation-rules", methods=["GET", "POST"])
def admin_price_calculation_rules():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    selected_customer = request.values.get("customer_key") or default_price_customer_key()
    try:
        selected_customer = enabled_price_customer(selected_customer)["key"]
    except ValueError:
        selected_customer = default_price_customer_key()
    selected_quote_variant = normalize_price_quote_variant(selected_customer, request.values.get("quote_variant"))

    regression = None
    if request.method == "POST":
        action = (request.form.get("action") or "upload").strip().lower()
        admin_password = request.form.get("admin_password", "")
        if action in {"activate", "delete"}:
            version = request.form.get("version", "").strip()
            if not verify_admin_password(admin_password):
                flash("管理员密码错误。", "error")
            else:
                try:
                    if action == "activate":
                        activated_version = activate_price_rule_version(
                            selected_customer,
                            version,
                            selected_quote_variant,
                        )
                        flash(f"规则版本已启用：{activated_version}", "success")
                    else:
                        deleted_version = delete_price_rule_version(
                            selected_customer,
                            version,
                            selected_quote_variant,
                        )
                        flash(f"规则版本已删除：{deleted_version}", "success")
                except Exception as exc:
                    operation = "启用" if action == "activate" else "删除"
                    flash(f"规则版本{operation}失败：{exc}", "error")
            return redirect(
                url_for(
                    "main.admin_price_calculation_rules",
                    customer_key=selected_customer,
                    quote_variant=selected_quote_variant,
                )
            )

        rule_file = request.files.get("rule_file")
        guanghe_huangshi_file = request.files.get("guanghe_huangshi_file")
        guanghe_nanya_file = request.files.get("guanghe_nanya_file")
        suhang_pp_file = request.files.get("suhang_pp_file")
        suhang_ccl_file = request.files.get("suhang_ccl_file")
        remark = request.form.get("remark", "").strip()
        if not verify_admin_password(admin_password):
            flash("管理员密码错误。", "error")
        elif selected_customer == "guanghe" and (not guanghe_huangshi_file or not guanghe_huangshi_file.filename or not guanghe_nanya_file or not guanghe_nanya_file.filename):
            flash("请上传黄石广合单价和南亚新材价格更新两份 Excel。", "error")
        elif selected_customer == "suhang" and (not suhang_pp_file or not suhang_pp_file.filename) and (not suhang_ccl_file or not suhang_ccl_file.filename):
            flash("请至少上传苏杭PP报价单或苏杭CCL报价单其中一份 Excel。", "error")
        elif selected_customer not in {"guanghe", "suhang"} and (not rule_file or not rule_file.filename):
            flash("请上传所选客户的报价表 Excel。", "error")
        else:
            try:
                if selected_customer == "guanghe":
                    version = save_new_guanghe_rule_version(
                        guanghe_huangshi_file,
                        guanghe_nanya_file,
                        updated_by=current_employee(),
                        remark=remark,
                    )
                elif selected_customer == "suhang":
                    version = save_new_suhang_rule_version(
                        suhang_pp_file,
                        suhang_ccl_file,
                        updated_by=current_employee(),
                        remark=remark,
                    )
                else:
                    version = save_new_price_rule_version(
                        selected_customer,
                        rule_file,
                        updated_by=current_employee(),
                        remark=remark,
                        quote_variant=selected_quote_variant,
                    )
                regression = run_jingwang_regression(selected_customer, version, selected_quote_variant)
                if regression["total"]:
                    flash(
                        f"价格计算规则已更新，当前生效版本：{version}；回归通过 {regression['passed']}/{regression['total']}，失败 {regression['failed']}。",
                        "success" if regression["failed"] == 0 else "warning",
                    )
                else:
                    flash(f"价格计算规则已更新，当前生效版本：{version}；未配置测试数据，已跳过回归校验。", "success")
                return redirect(url_for("main.admin_price_calculation_rules", customer_key=selected_customer, quote_variant=selected_quote_variant))
            except Exception as exc:
                flash(f"价格计算规则更新失败：{exc}", "error")

    try:
        active_rule_version = get_active_price_rule_version(selected_customer, selected_quote_variant) or "未初始化价格计算规则"
        rule_history = get_price_rule_history(selected_customer, selected_quote_variant)
    except Exception as exc:
        active_rule_version = f"规则读取失败：{exc}"
        rule_history = []
    return render_template(
        "admin_price_calculation_rules.html",
        customers=PRICE_CALCULATION_CUSTOMERS,
        selected_customer=selected_customer,
        selected_quote_variant=selected_quote_variant,
        quote_variants=JINGWANG_QUOTE_VARIANTS,
        active_rule_version=active_rule_version,
        rule_history=rule_history,
        regression=regression,
    )


@bp.route("/admin/password", methods=["GET", "POST"])
def admin_password():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    mode = request.args.get("mode", "menu")
    admin_verified = bool(session.get("admin_maintenance_verified"))
    if request.method == "POST":
        action = request.form.get("action", "password")
        admin_password_value = request.form.get("admin_password", "")
        if action == "verify_admin":
            if not verify_admin_password(admin_password_value):
                flash("管理员密码错误。", "error")
            else:
                session["admin_maintenance_verified"] = True
                flash("管理员身份验证通过。", "success")
                return redirect(url_for("main.admin_password", mode="admin"))
        elif action == "user_password":
            employee_id = current_employee()
            old_password = request.form.get("old_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")
            if not employee_id or not verify_user_password(employee_id, old_password):
                flash("当前密码错误。", "error")
            elif len(new_password) < 6:
                flash("新密码至少 6 位。", "error")
            elif new_password == employee_id:
                flash("新密码不能继续使用工号。", "error")
            elif new_password != confirm_password:
                flash("两次输入的新密码不一致。", "error")
            else:
                change_user_password(employee_id, new_password)
                flash("个人登录密码已更新。", "success")
                return redirect(url_for("main.admin_password", mode="user"))
        elif action == "password":
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")
            if not verify_admin_password(current_password):
                flash("当前管理员密码错误。", "error")
            elif len(new_password) < 6:
                flash("新管理员密码至少 6 位。", "error")
            elif new_password != confirm_password:
                flash("两次输入的新密码不一致。", "error")
            else:
                update_admin_password(new_password)
                session["admin_maintenance_verified"] = True
                flash("管理员密码已更新。", "success")
                return redirect(url_for("main.admin_password", mode="admin"))
        elif action in {"add_user", "reset_user"} and not admin_verified:
            flash("管理员密码错误。", "error")
            return redirect(url_for("main.admin_password", mode="admin"))
        elif action == "add_user":
            employee_id = request.form.get("employee_id", "").strip()
            if not employee_id:
                flash("请输入工号。", "error")
            else:
                create_user(
                    employee_id,
                    display_name=request.form.get("display_name", "").strip(),
                    department=request.form.get("department", "").strip(),
                    role=request.form.get("role", "user"),
                    enabled=bool(request.form.get("enabled", "1")),
                )
                flash(f"账号 {employee_id} 已保存，初始密码为工号。", "success")
                return redirect(url_for("main.admin_password", mode="admin"))
        elif action == "reset_user":
            employee_id = request.form.get("employee_id", "").strip()
            if not get_user(employee_id):
                flash("未找到该员工账号。", "error")
            else:
                reset_user_password(employee_id)
                flash(f"账号 {employee_id} 密码已重置为工号，下次登录需改密。", "success")
                return redirect(url_for("main.admin_password", mode="admin"))
        else:
            flash("未知的账户操作。", "error")
    admin_verified = bool(session.get("admin_maintenance_verified"))
    return render_template("admin_password.html", users=list_users(), mode=mode, admin_verified=admin_verified)


@bp.get("/rules-docs/<feature>")
def rule_doc(feature: str):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if feature == "transcode_agent":
        return redirect(url_for("main.admin_transcode_rule_center"))
    docs = {
        "fangzheng": {
            "title": "方正价格计算规则说明",
            "items": [
                "系统自动识别业务 Sheet 和物料描述列。",
                "标准 CCL 按胶系、厚度、铜厚、铜箔、叠构和尺寸列匹配价格。",
                "非标准尺寸优先查询基板对照表，再按大板、倍率和小片数量折算。",
                "尾板高度约等于 43 时，按 43/48×1.07 做尾板修正。",
                "厚度近似只允许差异 ≤0.01mm，超过范围输出未命中。",
                "PP 卷料按 48/0.0254/144×RMB/SF 计算。",
            ],
        },
        "transcode": {
            "title": "营销自动化转码规则说明",
            "items": [
                "系统自动识别需求 Sheet 和客户规格列。",
                "转码规则由胶系代码、胶系类别、编码规则、特殊需求、总芯厚转换、客户下单与胶系基板转换六张表组成。",
                "结果写入每行最后一个有数据列之后的新列。",
                "输出文件会新增转码说明 Sheet，记录命中情况和未识别原因。",
            ],
        },
        "shennan": {
            "title": "深南价格计算规则说明",
            "items": [
                "深南功能复用方正核心计算公式。",
                "报价规则从深南报价规则文件中读取，并按深南胶系和 PP/CCL 规则转换。",
                "基板对照表沿用当前方正规则版本。",
                "结果文件同样输出规则说明和未命中原因。",
            ],
        },
        "hushi": {
            "title": "沪士价格计算规则说明",
            "items": [
                "系统自动识别规格列，优先按物料描述中的胶系在沪士报价规则包中匹配报价 Excel。",
                "PP 按产品、玻纤、RC% 匹配报价行；CCL 按产品、厚度、铜厚、叠构、铜箔匹配报价行。",
                "价格优先取 Rebate；如果 Rebate 为空或被删除线划掉，则取 Normal；删除线价格不会被采用。",
                "尺寸按 mm 转 inch，每个边长先除以 25.4 后按 0.5 向下取档，例如 18.27 取 18，18.6 取 18.5。",
                "面积按取档后的 inch 边长相乘再除以 144，面积结果四舍五入保留两位小数后再参与计算。",
                "最终价格 = 匹配单价 × 面积，结果保留两位小数。",
            ],
        },
    }
    if feature not in docs:
        flash("未找到该规则说明。", "error")
        return redirect(url_for("main.dashboard"))
    return render_template("rule_doc.html", doc=docs[feature], feature=feature)


@bp.get("/feedback")
def feedback_center():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    return render_template("feedback.html", feedback_list=list_feedback(current_employee()))


@bp.route("/feedback/<kind>", methods=["GET", "POST"])
def feedback_form(kind: str):
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    valid = {"calculation", "suggestion", "requirement"}
    if kind not in valid:
        flash("未知反馈类型。", "error")
        return redirect(url_for("main.feedback_center"))
    if request.method == "POST":
        create_feedback(
            current_employee(),
            kind,
            feature=request.form.get("feature", ""),
            material_desc=request.form.get("material_desc", "").strip(),
            system_result=request.form.get("system_result", "").strip(),
            expected_result=request.form.get("expected_result", "").strip(),
            content=request.form.get("content", "").strip(),
            daily_workload=request.form.get("daily_workload", "").strip(),
            error_probability=request.form.get("error_probability", "").strip(),
            urgency=request.form.get("urgency", "").strip(),
        )
        flash("反馈已提交。", "success")
        return redirect(url_for("main.feedback_center"))
    return render_template("feedback_form.html", kind=kind)


@bp.route("/admin/feedback", methods=["GET", "POST"])
def admin_feedback():
    redirect_resp = require_login()
    if redirect_resp:
        return redirect_resp
    if request.method == "POST":
        admin_password_value = request.form.get("admin_password", "")
        if not verify_admin_password(admin_password_value):
            flash("管理员密码错误。", "error")
        else:
            session["feedback_admin_verified"] = True
            feedback_id = request.form.get("feedback_id", type=int)
            if feedback_id:
                update_feedback_status(
                    feedback_id,
                    request.form.get("status", "待处理"),
                    request.form.get("admin_note", "").strip(),
                )
                flash("反馈状态已更新。", "success")
            return redirect(url_for("main.admin_feedback"))
    if not session.get("feedback_admin_verified"):
        return render_template("feedback_admin.html", feedback_list=None)
    return render_template("feedback_admin.html", feedback_list=list_feedback())
