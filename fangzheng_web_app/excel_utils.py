from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
import zipfile

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font


OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
ZIP_MAGIC = b"PK"


def excel_format(path: str | Path) -> str:
    """Return xlsx_zip, xls_ole, or unknown based on file signature."""
    file_path = Path(path)
    try:
        with file_path.open("rb") as file:
            header = file.read(8)
    except OSError:
        return "unknown"
    if header.startswith(ZIP_MAGIC):
        return "xlsx_zip"
    if header == OLE_MAGIC:
        return "xls_ole"
    return "unknown"


def is_ole_workbook(path: str | Path) -> bool:
    return excel_format(path) == "xls_ole"


def load_workbook_compat(path: str | Path, *, data_only: bool = False, keep_formatting: bool = False):
    file_path = Path(path)
    detected = excel_format(file_path)
    if detected == "xls_ole":
        return load_xls_as_workbook(file_path, keep_formatting=keep_formatting)
    if detected != "xlsx_zip":
        raise ValueError("无法识别 Excel 文件格式，请使用系统模板另存为 .xlsx 后上传。")
    try:
        # openpyxl rejects xlsx content when the path suffix is .xls. Passing a
        # binary stream lets us honor the file signature instead of the suffix.
        with file_path.open("rb") as file:
            return openpyxl.load_workbook(file, data_only=data_only)
    except Exception as exc:
        repaired = _load_missing_shared_strings_workbook(file_path, data_only=data_only)
        if repaired is not None:
            return repaired
        repaired = _load_truncated_zip_workbook(file_path, data_only=data_only)
        if repaired is not None:
            return repaired
        raise ValueError(f"Excel 文件读取失败：{exc}；该文件可能是损坏的 xlsx，请重新下载或另存为标准 .xlsx 后再上传") from exc


def _load_missing_shared_strings_workbook(path: Path, *, data_only: bool = False):
    """Repair malformed xlsx packages whose shared-string part is misnamed or only falsely declared."""
    try:
        source = BytesIO(path.read_bytes())
        with zipfile.ZipFile(source) as archive:
            names = archive.namelist()
            if "xl/sharedStrings.xml" in names:
                return None

            alias = next(
                (name for name in names if name.lstrip("/").casefold() == "xl/sharedstrings.xml"),
                None,
            )
            if alias:
                shared_strings = archive.read(alias)
            else:
                worksheet_names = [
                    name
                    for name in names
                    if name.lstrip("/").casefold().startswith("xl/worksheets/") and name.casefold().endswith(".xml")
                ]
                shared_string_cell = re.compile(rb"<c\b[^>]*\bt=[\"']s[\"']")
                if any(shared_string_cell.search(archive.read(name)) for name in worksheet_names):
                    return None
                shared_strings = (
                    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0" uniqueCount="0"/>'
                )

            repaired_bytes = BytesIO()
            with zipfile.ZipFile(repaired_bytes, "w", zipfile.ZIP_DEFLATED) as repaired_archive:
                for item in archive.infolist():
                    repaired_archive.writestr(item, archive.read(item.filename))
                repaired_archive.writestr("xl/sharedStrings.xml", shared_strings)
        repaired_bytes.seek(0)
        return openpyxl.load_workbook(repaired_bytes, data_only=data_only)
    except (OSError, KeyError, zipfile.BadZipFile, ValueError, IndexError):
        return None


def _load_truncated_zip_workbook(path: Path, *, data_only: bool = False):
    """Repair lightly truncated xlsx zip files, such as WeCom cache exports."""
    try:
        data = path.read_bytes()
    except OSError:
        return None
    if not data.startswith(ZIP_MAGIC):
        return None
    for padding in range(1, 5):
        candidate = data + (b"\0" * padding)
        if not zipfile.is_zipfile(BytesIO(candidate)):
            continue
        try:
            return openpyxl.load_workbook(BytesIO(candidate), data_only=data_only)
        except Exception:
            continue
    return None


def normalized_xlsx_source(path: str | Path, workbook=None) -> Path:
    """Return an xlsx path that openpyxl can reopen for result writing."""
    file_path = Path(path)
    if not is_ole_workbook(file_path) and zipfile.is_zipfile(file_path):
        return file_path
    normalized = file_path.with_name(f"{file_path.stem}_normalized.xlsx")
    wb = workbook or load_workbook_compat(file_path)
    wb.save(normalized)
    return normalized


def load_xls_as_workbook(path: Path, *, keep_formatting: bool = False):
    import xlrd

    try:
        book = xlrd.open_workbook(path, formatting_info=keep_formatting)
    except Exception as exc:
        raise ValueError(f"旧版 .xls 文件读取失败：{exc}") from exc

    out_wb = Workbook()
    default_ws = out_wb.active
    out_wb.remove(default_ws)

    for sheet_index, sheet in enumerate(book.sheets()):
        title = sheet.name or f"Sheet{sheet_index + 1}"
        ws = out_wb.create_sheet(title[:31])
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                xl_cell = sheet.cell(r, c)
                value = xl_cell.value
                if xl_cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                elif xl_cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                    value = int(value)
                cell = ws.cell(row=r + 1, column=c + 1, value=value)
                if keep_formatting:
                    try:
                        xf = book.xf_list[xl_cell.xf_index]
                        font = book.font_list[xf.font_index]
                        if getattr(font, "struck_out", False):
                            cell.font = Font(strike=True)
                    except Exception:
                        pass
    return out_wb
