from __future__ import annotations

import copy
import io
import tempfile
import unittest
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from PIL import Image
from openpyxl import load_workbook

from fangzheng_web_app.purchase_ai_repair import audit_and_repair_purchase_document
from fangzheng_web_app.purchase_excel_writer import (
    _extract_roll_length,
    _project_detail_standard_fields,
    write_purchase_order_workbook,
)
from fangzheng_web_app.pdf_excel_service import _write_purchase_results
from fangzheng_web_app.cell_ocr import _ocr_thread_count
from fangzheng_web_app.purchase_factory_mapper import FACTORY_DETAIL_HEADERS, project_factory_document
from fangzheng_web_app.purchase_field_rules import header_score
from fangzheng_web_app.purchase_order_pipeline import _native_coordinate_rows, _native_coordinate_rows_reliable
from fangzheng_web_app.purchase_order_segmenter import (
    _repair_sparse_detail_cells,
    _repair_uncertain_detail_headers,
    build_detail_rows_from_table,
)
from fangzheng_web_app.purchase_performance import load_parser_cache, save_parser_cache
from fangzheng_web_app.purchase_result_normalizer import normalize_pp_spec_spacing, normalize_purchase_document


HEADERS = [
    "物料编号",
    "申购单号",
    "物料名称",
    "型号/规格",
    "单位",
    "数量",
    "单价",
    "税率",
    "杂运费",
    "币种",
    "到货日期",
    "采购备注",
]


class PurchaseFieldRuleTests(unittest.TestCase):
    def test_separate_name_and_spec_columns_map_to_name_and_description(self) -> None:
        _score, mapping = header_score(HEADERS)
        self.assertEqual(mapping[2], "物料名称")
        self.assertEqual(mapping[3], "说明")

    def test_combined_spec_column_keeps_legacy_name_mapping(self) -> None:
        _score, mapping = header_score(["物料编号", "型号/规格", "数量", "单价"])
        self.assertEqual(mapping[1], "物料名称")

    def test_pp_spec_spacing_is_conservative(self) -> None:
        self.assertEqual(normalize_pp_spec_spacing("NY3150HCP762848%150米/卷"), "NY3150HCP 7628 48% 150米/卷")
        self.assertEqual(normalize_pp_spec_spacing("NY3170HCP211655%300米/卷"), "NY3170HCP 2116 55% 300米/卷")
        self.assertEqual(normalize_pp_spec_spacing("普通规格ABC123"), "普通规格ABC123")


class SparseOcrTests(unittest.TestCase):
    def test_only_empty_mapped_critical_cell_is_repaired(self) -> None:
        rows = [
            HEADERS,
            ["PP001840", "", "半固化片", "NY3170HCP762848%150米/卷", "卷", "", "6750", "0.13", "0", "RMB", "2026-09-01", ""],
        ]
        cells = [
            {
                "row_index": 1,
                "column_index": 5,
                "bbox": [40, 20, 80, 50],
                "text": "",
                "confidence": 0,
                "method": "page_ocr_row_cluster",
            }
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            image_path = Path(temp_dir) / "table.png"
            Image.new("RGB", (120, 80), "white").save(image_path)
            with patch("fangzheng_web_app.purchase_order_segmenter._is_blank_crop", return_value=False), patch(
                "fangzheng_web_app.purchase_order_segmenter.ocr_cell",
                return_value={"text": "1", "confidence": 0.99, "method": "cell_ocr"},
            ) as mocked_ocr:
                actions = _repair_sparse_detail_cells(image_path, rows, cells, max_repairs=1)
        self.assertEqual(rows[1][5], "1")
        self.assertEqual(actions[0]["field"], "数量")
        self.assertEqual(cells[0]["method"], "cell_ocr_sparse_fallback")
        mocked_ocr.assert_called_once()

    def test_uncertain_header_is_replaced_only_when_mapping_improves(self) -> None:
        rows = [
            ["物料编号", "物料名称", "单价", "数里", "到货日期"],
            ["PP001840", "半固化片", "6750", "1", "2026-09-01"],
        ]
        cells = [
            {
                "row_index": 0,
                "column_index": 3,
                "bbox": [40, 10, 80, 30],
                "text": "数里",
                "confidence": 0.4,
                "method": "page_ocr_region",
            }
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            image_path = Path(temp_dir) / "table.png"
            Image.new("RGB", (120, 80), "white").save(image_path)
            with patch("fangzheng_web_app.purchase_order_segmenter._is_blank_crop", return_value=False), patch(
                "fangzheng_web_app.purchase_order_segmenter.ocr_cell",
                return_value={"text": "数量", "confidence": 0.98, "method": "cell_ocr"},
            ):
                actions = _repair_uncertain_detail_headers(image_path, rows, cells, max_repairs=1)
        self.assertEqual(rows[0][3], "数量")
        self.assertEqual(actions[0]["field"], "数量")
        self.assertEqual(cells[0]["method"], "cell_ocr_header_fallback")

    def test_partial_header_can_be_recovered_to_a_valid_detail_header(self) -> None:
        rows = [["物料编号", "数里", "单价"], ["PP001840", "1", "6750"]]
        cells = [{"row_index": 0, "column_index": 1, "bbox": [40, 10, 80, 30], "text": "数里", "confidence": 0.3}]
        with tempfile.TemporaryDirectory() as temp_dir:
            image_path = Path(temp_dir) / "table.png"
            Image.new("RGB", (120, 80), "white").save(image_path)
            with patch("fangzheng_web_app.purchase_order_segmenter._is_blank_crop", return_value=False), patch(
                "fangzheng_web_app.purchase_order_segmenter.ocr_cell",
                return_value={"text": "数量", "confidence": 0.98, "method": "cell_ocr"},
            ):
                actions = _repair_uncertain_detail_headers(image_path, rows, cells, max_repairs=1)
        header_index, mapping = header_score(rows[0])
        self.assertEqual(header_index, 3)
        self.assertEqual(mapping[1], "数量")
        self.assertEqual(len(actions), 1)


class PurchaseNormalizerTests(unittest.TestCase):
    def test_docling_duplicate_code_spec_is_split_and_relined(self) -> None:
        document = {
            "source_file": "XF77870.pdf",
            "raw_detail_tables": [
                {
                    "page_index": 0,
                    "table_index": 0,
                    "rows": [
                        ["项目", "物料编号", "物料名/规格", "物料要求", "单位", "数量", "税率", "含税单价", "含税金额"],
                        ["1", "20203NY10003-31 70-B-S NY 3170M 0.203mm 1/1(不含铜)1100*1252", "20203NY10003-31 70-B-S NY 3170M 0.203mm 1/1(不含铜)1100*1252", "此订单加急", "张", "800", "13.00", "210.000", "168,000.00"],
                        ["2", "20305NY10003-31 70-B-S NY 3170M 0.305mm 1/1(不含铜)1100*1252", "20305NY10003-31 70-B-S NY 3170M 0.305mm 1/1(不含铜)1100*1252", "张", "", "400", "13.00", "244.650", "97,860.00"],
                        ["3", "NYPP1080-64-317 0 NY PP 3170M 1080 RC64%49.5\"*300米", "NYPP1080-64-317 0 NY PP 3170M 1080 RC64%49.5\"*300米", "卷", "2", "", "13.00", "11190.000", "22,380.00"],
                        ["4", "NYPP2116-55-317 0 NY PP 3170M 2116 RC55%49.5\"*300米", "NYPP2116-55-317 0 NY PP 3170M 2116 RC55%49.5\"*300米", "卷", "1", "", "13.00", "12360.000", "12,360.00"],
                    ],
                    "method": "docling_markdown",
                    "confidence": 0.92,
                }
            ],
            "mapped_detail_rows": [],
            "pages": [],
            "sections": {},
            "issues": [],
            "warnings": [],
        }

        normalized = normalize_purchase_document(document)
        raw_rows = normalized["raw_detail_tables"][0]["rows"]
        standards = [row["standard"] for row in normalized["mapped_detail_rows"]]

        self.assertEqual(
            [row[1] for row in raw_rows[1:]],
            ["20203NY10003-3170-B-S", "20305NY10003-3170-B-S", "NYPP1080-64-3170", "NYPP2116-55-3170"],
        )
        self.assertEqual(raw_rows[1][2], "NY 3170M 0.203mm\n1/1(不含铜)1100*1252")
        self.assertEqual(raw_rows[3][2], 'NY PP 3170M 1080\nRC64%49.5"*300米')
        self.assertEqual([row[3] for row in raw_rows[1:]], ["此订单加急", "", "", ""])
        self.assertEqual([row[4] for row in raw_rows[1:]], ["张", "张", "卷", "卷"])
        self.assertEqual([row[5] for row in raw_rows[1:]], ["800", "400", "2", "1"])
        self.assertEqual([row["物料编码"] for row in standards], [row[1] for row in raw_rows[1:]])
        self.assertEqual([row["物料名称"] for row in standards], [row[2] for row in raw_rows[1:]])

    def test_unique_subtotal_solution_repairs_quantities_and_specs(self) -> None:
        document = {
            "source_file": "P0202607-0229.png",
            "raw_detail_tables": [
                {
                    "page_index": 0,
                    "table_index": 0,
                    "rows": [
                        HEADERS,
                        ["PP001710", "", "半固化片", "NY3150HCP762848%150米/卷", "卷", "4", "6300", "0.13", "0", "RMB", "2026-09-01", ""],
                        ["PP001840", "", "半固化片", "NY3170HCP762848%150米/卷", "卷", "", "6750", "0.13", "0", "RMB", "2026-09-01", ""],
                        ["PP001880", "", "半固化片", "NY3170HCP211655%300米/卷", "卷", "", "11700", "0.13", "0", "RMB", "2026-09-01", ""],
                    ],
                    "method": "test",
                    "confidence": 0.9,
                }
            ],
            "mapped_detail_rows": [],
            "pages": [{"text_lines": ["小计(Sub Total): 43650"]}],
            "sections": {},
            "issues": [],
            "warnings": [],
        }
        normalized = normalize_purchase_document(document, source_text="小计(Sub Total): 43650")
        raw_rows = normalized["raw_detail_tables"][0]["rows"]
        standards = [row["standard"] for row in normalized["mapped_detail_rows"]]
        self.assertEqual([row[5] for row in raw_rows[1:]], ["4", "1", "1"])
        self.assertEqual([row["数量"] for row in standards], ["4", "1", "1"])
        self.assertEqual(standards[0]["说明"], "NY3150HCP 7628 48% 150米/卷")
        self.assertEqual(standards[2]["说明"], "NY3170HCP 2116 55% 300米/卷")
        self.assertEqual(len(normalized.get("recovery_actions") or []), 2)


class RollToMeterExportTests(unittest.TestCase):
    @staticmethod
    def _detail(spec: str, *, quantity: str = "2", price: str = "11190", amount: str = "") -> dict:
        return {
            "original": {
                "项目": "1",
                "物料编号": "NYPP1080-64-3170",
                "物料名/规格": spec,
                "单位": "卷",
                "数量": quantity,
                "含税单价": price,
                "含税金额": amount,
            },
            "standard": {
                "序号": "1",
                "物料编码": "NYPP1080-64-3170",
                "物料名称": spec,
                "说明": "",
                "数量": quantity,
                "单位": "卷",
                "含税单价": price,
                "金额": amount,
                "交货日期": "",
                "备注": "",
            },
            "page_index": 0,
            "table_index": 0,
            "row_index": 1,
        }

    def test_roll_length_formats(self) -> None:
        cases = {
            "NY PP 3170M 1080 RC64% 150米/卷": "150",
            "NY PP 3170M 1080 RC64% 300M/卷": "300",
            'NY PP 3170M 1080 RC64%49.5"*300米': "300",
            "NY PP 3170M 1080 ×150m": "150",
            "规格 50M": "50",
            "规格 50.5米": "50.5",
        }
        for spec, expected in cases.items():
            with self.subTest(spec=spec):
                length, _evidence, reason = _extract_roll_length(self._detail(spec))
                self.assertEqual(str(length), expected)
                self.assertEqual(reason, "")

    def test_model_suffix_is_not_treated_as_roll_length(self) -> None:
        length, _evidence, reason = _extract_roll_length(self._detail("NY PP 3170M 1080"))
        self.assertIsNone(length)
        self.assertIn("未找到", reason)

        length, _evidence, reason = _extract_roll_length(self._detail("NY PP 3170M 1080 300M"))
        self.assertEqual(str(length), "300")
        self.assertEqual(reason, "")

    def test_conflicting_lengths_are_rejected(self) -> None:
        length, _evidence, reason = _extract_roll_length(self._detail("PP 150M/卷 备用规格 300M/卷"))
        self.assertIsNone(length)
        self.assertIn("多个不同卷长", reason)

    def test_standard_fields_are_projected_to_meters(self) -> None:
        projected, issue = _project_detail_standard_fields(
            self._detail('NY PP 3170M 1080 RC64%49.5"*300米', amount="22380")
        )
        self.assertIsNone(issue)
        self.assertEqual(projected["数量"], "600")
        self.assertEqual(projected["单位"], "米")
        self.assertEqual(projected["含税单价"], "37.3")
        self.assertEqual(projected["金额"], "22380")

    def test_failed_projection_keeps_roll_values_and_returns_issue(self) -> None:
        detail = self._detail("NY PP 3170M 1080")
        projected, issue = _project_detail_standard_fields(detail)
        self.assertEqual(projected["数量"], "2")
        self.assertEqual(projected["单位"], "卷")
        self.assertEqual(projected["含税单价"], "11190")
        self.assertIsNotNone(issue)
        self.assertIn("保留卷制口径", issue["message"])

    def test_workbook_keeps_purchase_rolls_and_converts_standard_columns(self) -> None:
        spec = 'NY PP 3170M 1080\nRC64%49.5"*300米'
        detail = self._detail(spec)
        document = {
            "source_file": "XF77870.pdf",
            "page_count": 1,
            "header_info": {"订单号": "XF77870"},
            "raw_detail_tables": [
                {
                    "page_index": 0,
                    "rows": [
                        ["项目", "物料编号", "物料名/规格", "单位", "数量", "含税单价", "含税金额"],
                        ["1", "NYPP1080-64-3170", spec, "卷", "2", "11190", ""],
                    ],
                }
            ],
            "mapped_detail_rows": [detail],
            "sections": {},
            "issues": [],
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "roll_to_meter.xlsx"
            stats = write_purchase_order_workbook([document], output_path)
            workbook = load_workbook(output_path, data_only=False)

        purchase = workbook["采购单"]
        purchase_header_row = next(
            row for row in range(1, purchase.max_row + 1) if purchase.cell(row=row, column=2).value == "物料编号"
        )
        self.assertEqual(purchase.cell(row=purchase_header_row + 1, column=4).value, "卷")
        self.assertEqual(purchase.cell(row=purchase_header_row + 1, column=5).value, "2")
        self.assertEqual(purchase.cell(row=purchase_header_row + 1, column=6).value, "11190")

        details = workbook["明细数据"]
        self.assertEqual([details.cell(row=1, column=column).value or "" for column in range(1, 13)], [
            "单别（必填）", "类型1（必填）", "类型2（必填）", "账款客户编号（必填）",
            "送货客户编号（必填）", "到期付款日（必填）", "送货厂别（选填）", "客户订单号（选填）",
            "", "", "", "",
        ])
        self.assertEqual([details["A2"].value, details["B2"].value, details["C2"].value], ["220", "1", "1"])
        self.assertEqual(details["H2"].value, "XF77870")
        self.assertEqual([details.cell(row=3, column=column).value for column in range(1, 13)], FACTORY_DETAIL_HEADERS)
        self.assertEqual(details["A4"].value, "1")
        self.assertIsNone(details["C4"].value)
        self.assertEqual(details["D4"].value, "NYPP1080-64-3170")
        self.assertEqual(details["F4"].value, 600)
        self.assertIsNone(details["G4"].value)
        self.assertAlmostEqual(details["H4"].value, 37.3)
        self.assertIsNone(details["J4"].value)
        self.assertEqual(details["K4"].value, "XF77870")
        self.assertEqual(details["L4"].value, "2卷&")
        self.assertEqual(stats["issue_count"], 0)

    def test_failed_workbook_conversion_creates_issue_sheet(self) -> None:
        detail = self._detail("NY PP 3170M 1080")
        document = {
            "source_file": "missing_length.pdf",
            "page_count": 1,
            "header_info": {"订单号": "PO-MISSING-LENGTH"},
            "raw_detail_tables": [],
            "mapped_detail_rows": [detail],
            "sections": {},
            "issues": [],
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "missing_length.xlsx"
            stats = write_purchase_order_workbook([document], output_path)
            workbook = load_workbook(output_path, data_only=False)
        self.assertEqual(stats["issue_count"], 1)
        self.assertIn("识别日志", workbook.sheetnames)
        self.assertIn("保留卷制口径", workbook["识别日志"]["H2"].value)


class FactoryImportProjectionTests(unittest.TestCase):
    @staticmethod
    def _document(
        *,
        order: str = "PO-001",
        code: str = "CUST-001",
        item: str = "7",
        quantity: str = "2",
        unit: str = "张",
        price: str = "113",
        original_prices: dict[str, str] | None = None,
        spec: str = "普通板材",
        remark: str = "加急",
    ) -> dict:
        original = {
            "项次": item,
            "物料编号": code,
            "物料名称": "板材",
            "规格": spec,
            "数量": quantity,
            "单位": unit,
            "到货日期": "2026-08-01",
            "备注": remark,
        }
        original.update(original_prices or {"含税单价": price})
        detail = {
            "original": original,
            "standard": {
                "序号": item,
                "物料编码": code,
                "物料名称": "板材",
                "说明": spec,
                "数量": quantity,
                "单位": unit,
                "含税单价": price,
                "金额": "",
                "交货日期": "2026-08-01",
                "备注": remark,
            },
            "raw_text": f"{item} {code} {quantity} {price}",
            "page_index": 0,
            "table_index": 0,
            "row_index": 1,
        }
        return {
            "source_file": f"{order}.pdf",
            "page_count": 1,
            "parser_mode": "test",
            "template_id": "",
            "template_label": "",
            "header_info": {"订单号": order, "客户": "测试客户"},
            "raw_detail_tables": [
                {
                    "page_index": 0,
                    "rows": [
                        list(original.keys()),
                        list(original.values()),
                    ],
                }
            ],
            "mapped_detail_rows": [detail],
            "sections": {},
            "issues": [],
        }

    def test_customer_code_never_populates_factory_product_code(self) -> None:
        document = self._document()
        summary = project_factory_document(document)
        row = document["factory_import"]["rows"][0]
        self.assertEqual(row["项次（必填）"], "7")
        self.assertEqual(row["产品编号"], "")
        self.assertEqual(row["客户产品编号（必填）"], "CUST-001")
        self.assertEqual(row["客户订单序号（选填）"], "")
        self.assertEqual(row["客户订单号"], "PO-001")
        self.assertEqual(row["备注（选填）"], "加急&")
        self.assertEqual(document["factory_import"]["main_values"][:3], ["220", "1", "1"])
        self.assertEqual(summary["ready_rows"], 1)

    def test_customer_identity_uses_document_company_evidence(self) -> None:
        document = self._document()
        document["header_info"].pop("客户")
        document["header_info"]["供应商"] = "南亚新材料科技（江西）有限公司"
        document["pages"] = [{"text_lines": ["深圳万基隆电子科技有限公司", "供应商：南亚新材料科技（江西）有限公司"]}]
        summary = project_factory_document(document)
        self.assertEqual(summary["customer"], "深圳万基隆电子科技有限公司")

    def test_explicit_pre_tax_and_tax_prices_fill_separate_columns(self) -> None:
        document = self._document(original_prices={"未税单价": "100", "含税单价": "113"})
        project_factory_document(document)
        row = document["factory_import"]["rows"][0]
        self.assertEqual(row["税前单价（选填）"], "100")
        self.assertEqual(row["单价（选填）"], "113")

    def test_order_number_is_sanitized_and_filename_is_a_controlled_fallback(self) -> None:
        polluted = self._document(order="XF77870")
        polluted["header_info"]["订单号"] = "码:XF77870 页 码: 1 of 1"
        project_factory_document(polluted)
        self.assertEqual(polluted["factory_import"]["main_values"][7], "XF77870")

        filename_only = self._document(order="P0202607-0229")
        filename_only["header_info"].pop("订单号")
        project_factory_document(filename_only)
        self.assertEqual(filename_only["factory_import"]["main_values"][7], "P0202607-0229")
        self.assertEqual(filename_only["factory_mapping_summary"]["order_number_source"], "来源文件名")

    def test_generic_roll_price_defaults_to_tax_included_and_converts_per_meter(self) -> None:
        document = self._document(
            item="",
            quantity="4",
            unit="卷",
            price="6300",
            original_prices={"单价": "6300"},
            spec="NY3150HCP 7628 48% 150米/卷",
        )
        project_factory_document(document)
        row = document["factory_import"]["rows"][0]
        self.assertEqual(row["项次（必填）"], "1")
        self.assertEqual(row["数量（必填）"], "600")
        self.assertEqual(row["税前单价（选填）"], "")
        self.assertEqual(row["单价（选填）"], "42")
        self.assertEqual(row["备注（选填）"], "加急；4卷&")
        self.assertTrue(any("默认写入厂内“单价”列" in issue["message"] for issue in document["issues"]))
        self.assertTrue(any("按输出顺序生成" in issue["message"] for issue in document["issues"]))

    def test_roll_remark_uses_original_quantity_and_preserves_existing_remark(self) -> None:
        empty_remark = self._document(
            quantity="5",
            unit="卷",
            spec="NY3150HCP 7628 48% 150米/卷",
            remark="",
        )
        project_factory_document(empty_remark)
        self.assertEqual(empty_remark["factory_import"]["rows"][0]["备注（选填）"], "5卷&")

        decimal_quantity = self._document(
            quantity="2.500",
            unit="卷",
            spec="NY3150HCP 7628 48% 150米/卷",
            remark="加急",
        )
        project_factory_document(decimal_quantity)
        row = decimal_quantity["factory_import"]["rows"][0]
        self.assertEqual(row["数量（必填）"], "375")
        self.assertEqual(row["备注（选填）"], "加急；2.5卷&")

    def test_roll_remark_is_not_duplicated_and_failed_conversion_keeps_remark(self) -> None:
        already_noted = self._document(
            quantity="5",
            unit="卷",
            spec="NY3150HCP 7628 48% 150米/卷",
            remark="加急；5卷&",
        )
        project_factory_document(already_noted)
        self.assertEqual(already_noted["factory_import"]["rows"][0]["备注（选填）"], "加急；5卷&")

        failed = self._document(quantity="5", unit="卷", spec="NY PP 3170M 1080", remark="加急")
        project_factory_document(failed)
        failed_row = failed["factory_import"]["rows"][0]
        self.assertEqual(failed_row["数量（必填）"], "5")
        self.assertEqual(failed_row["备注（选填）"], "加急&")

        blank_non_roll = self._document(remark="")
        project_factory_document(blank_non_roll)
        self.assertEqual(blank_non_roll["factory_import"]["rows"][0]["备注（选填）"], "")

    def test_material_name_is_not_used_as_customer_product_code(self) -> None:
        document = self._document(code="")
        document["mapped_detail_rows"][0]["original"].pop("物料编号")
        document["mapped_detail_rows"][0]["original"]["物料名称"] = "ABC123规格名称"
        project_factory_document(document)
        row = document["factory_import"]["rows"][0]
        self.assertEqual(row["客户产品编号（必填）"], "")
        self.assertEqual(document["factory_mapping_summary"]["review_rows"], 1)

    @patch("fangzheng_web_app.purchase_factory_mapper.get_ai_repair_config")
    @patch("fangzheng_web_app.purchase_factory_mapper.request_repair_json")
    def test_ai_header_mapping_only_applies_high_confidence_valid_candidates(self, mocked_request, mocked_config) -> None:
        mocked_config.return_value = SimpleNamespace(available=True)
        mocked_request.return_value = {
            "mappings": [
                {"source_header": "客户料件", "target_field": "customer_product_code", "confidence": 0.95},
                {"source_header": "订购数", "target_field": "quantity", "confidence": 0.95},
            ]
        }
        document = self._document(code="", quantity="")
        detail = document["mapped_detail_rows"][0]
        detail["original"] = {"客户料件": "AI-CODE-1", "订购数": "3", "含税单价": "113"}
        project_factory_document(document)
        row = document["factory_import"]["rows"][0]
        self.assertEqual(row["客户产品编号（必填）"], "AI-CODE-1")
        self.assertEqual(row["数量（必填）"], "3")
        self.assertEqual(document["factory_mapping_summary"]["ai_header_mapping"]["accepted"], 2)

    def test_single_and_multi_order_result_shapes(self) -> None:
        first = self._document(order="PO-A", code="CODE-A")
        second = self._document(order="PO-B", code="CODE-B")
        project_factory_document(first)
        project_factory_document(second)
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            single_path, _single_stats, single_files = _write_purchase_results([first], root / "single", "single")
            self.assertEqual(single_path.suffix, ".xlsx")
            self.assertEqual(len(single_files), 1)

            zip_path, stats, result_files = _write_purchase_results([first, second], root / "multi", "multi")
            self.assertEqual(zip_path.suffix, ".zip")
            self.assertEqual(len(result_files), 2)
            self.assertEqual(stats["ready_count"], 2)
            with zipfile.ZipFile(zip_path) as archive:
                names = archive.namelist()
                self.assertEqual(len(names), 2)
                workbooks = [load_workbook(io.BytesIO(archive.read(name)), data_only=False) for name in names]
        self.assertEqual([workbook["明细数据"]["H2"].value for workbook in workbooks], ["PO-A", "PO-B"])
        self.assertEqual([workbook["明细数据"]["D4"].value for workbook in workbooks], ["CODE-A", "CODE-B"])


class PurchasePerformanceTests(unittest.TestCase):
    def test_parser_cache_round_trip_scrubs_transient_paths(self) -> None:
        document = {
            "source_file": "first.png",
            "pages": [{"image_path": "old/page.png", "clean_image_path": "old/clean.png", "text_lines": ["PO-1"]}],
            "mapped_detail_rows": [{"standard": {"物料编码": "A-1"}}],
            "performance_summary": {"stage_ms": {"total": 99}},
        }
        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "fangzheng_web_app.purchase_performance.PARSER_CACHE_DIR", Path(temp_dir)
        ):
            self.assertTrue(save_parser_cache("abc123", ".png", document))
            cached, _key, reason = load_parser_cache("abc123", ".png")
        self.assertEqual(reason, "hit")
        self.assertEqual(cached["mapped_detail_rows"][0]["standard"]["物料编码"], "A-1")
        self.assertNotIn("image_path", cached["pages"][0])
        self.assertNotIn("clean_image_path", cached["pages"][0])
        self.assertNotIn("performance_summary", cached)

    def test_corrupt_parser_cache_falls_back_to_miss_reason(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "fangzheng_web_app.purchase_performance.PARSER_CACHE_DIR", Path(temp_dir)
        ):
            self.assertTrue(save_parser_cache("broken", ".pdf", {"source_file": "test.pdf"}))
            cache_file = next(Path(temp_dir).glob("*.json"))
            cache_file.write_text("{broken", encoding="utf-8")
            cached, _key, reason = load_parser_cache("broken", ".pdf")
        self.assertIsNone(cached)
        self.assertEqual(reason, "corrupt")

    def test_concurrent_parser_cache_writes_are_atomic(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "fangzheng_web_app.purchase_performance.PARSER_CACHE_DIR", Path(temp_dir)
        ):
            with ThreadPoolExecutor(max_workers=2) as executor:
                results = list(
                    executor.map(
                        lambda index: save_parser_cache(
                            "same-content",
                            ".png",
                            {"source_file": f"source-{index}.png", "mapped_detail_rows": []},
                        ),
                        range(2),
                    )
                )
            cached, _key, reason = load_parser_cache("same-content", ".png")
        self.assertEqual(results, [True, True])
        self.assertEqual(reason, "hit")
        self.assertIn(cached["source_file"], {"source-0.png", "source-1.png"})

    def test_ocr_thread_count_is_adaptive_and_overridable(self) -> None:
        with patch.dict("os.environ", {"PDF_EXCEL_OCR_THREADS": "auto"}), patch("os.cpu_count", return_value=12):
            self.assertEqual(_ocr_thread_count(), 4)
        with patch.dict("os.environ", {"PDF_EXCEL_OCR_THREADS": "2"}):
            self.assertEqual(_ocr_thread_count(), 2)

    def test_native_coordinate_fast_path_rebuilds_xf_fragments(self) -> None:
        rows = [
            ["项目", "物料编号", "", "物料名/规格", "", "物料要求", "单位", "数量", "税率", "", "含税单价", "含税金额"],
            ["", "", "", "NY 3170M 0.203mm", "", "此订单加急", "", "", "", "", "", ""],
            ["1", "20203NY10003-31", "", "1/1(不含铜)1100*1252", "", "", "张", "800", "13.00", "", "210.000", "168,000.00"],
            ["", "70-B-S", "", "", "", "", "", "", "", "", "", ""],
            ["", "", "", "NY PP 3170M 1080", "", "", "", "", "", "", "", ""],
            ["2", "NYPP1080-64-317", "", "RC64%49.5\"*300米", "", "", "卷", "2", "13.00", "", "11190.000", "22,380.00"],
            ["", "0", "", "", "", "", "", "", "", "", "", ""],
            ["含税总金额", "", "", "190,380.00", "", "", "", "", "", "", "", ""],
        ]
        rebuilt, reason = _native_coordinate_rows(rows)
        self.assertEqual(reason, "")
        table = {"page_index": 0, "table_index": 0, "raw_rows": rebuilt, "rows": rebuilt, "method": "pdf_text"}
        mapped_rows, issues = build_detail_rows_from_table(table)
        reliable, validation_reason = _native_coordinate_rows_reliable(mapped_rows, rows)
        self.assertEqual(issues, [])
        self.assertTrue(reliable, validation_reason)
        self.assertEqual(mapped_rows[0]["standard"]["物料编码"], "20203NY10003-3170-B-S")
        self.assertEqual(mapped_rows[0]["standard"]["物料名称"], "NY 3170M 0.203mm\n1/1(不含铜)1100*1252")
        self.assertEqual(mapped_rows[1]["standard"]["物料编码"], "NYPP1080-64-3170")

    @patch("fangzheng_web_app.docling_worker._ensure_worker")
    @patch("fangzheng_web_app.docling_worker.Client")
    def test_docling_worker_client_marks_worker_result(self, mocked_client, mocked_ensure) -> None:
        mocked_ensure.return_value = {"host": "127.0.0.1", "port": 1234, "authkey": "00" * 32}
        connection = mocked_client.return_value
        connection.poll.return_value = True
        connection.recv.return_value = {
            "ok": True,
            "result": {"markdown": "ok", "lines": [], "tables": [], "error": None},
        }
        from fangzheng_web_app.docling_worker import request_docling_parse

        result = request_docling_parse(Path("test.pdf"))
        self.assertTrue(result["worker_used"])
        connection.send.assert_called_once()
        connection.close.assert_called_once()

    @patch("fangzheng_web_app.docling_worker._ensure_worker")
    @patch("fangzheng_web_app.docling_worker.Client")
    def test_docling_worker_timeout_returns_local_fallback_signal(self, mocked_client, mocked_ensure) -> None:
        mocked_ensure.return_value = {"host": "127.0.0.1", "port": 1234, "authkey": "00" * 32}
        connection = mocked_client.return_value
        connection.poll.return_value = False
        from fangzheng_web_app.docling_worker import request_docling_parse

        self.assertIsNone(request_docling_parse(Path("test.pdf")))
        connection.close.assert_called_once()

    @patch("fangzheng_web_app.docling_parser._read_cache", return_value=None)
    @patch("fangzheng_web_app.docling_worker.request_docling_parse", return_value=None)
    @patch("fangzheng_web_app.docling_parser.parse_pdf_with_docling_local")
    def test_docling_public_parser_falls_back_to_local(self, mocked_local, _mocked_worker, _mocked_cache) -> None:
        mocked_local.return_value = {"markdown": "local", "lines": [], "tables": [], "error": None}
        from fangzheng_web_app.docling_parser import parse_pdf_with_docling

        result = parse_pdf_with_docling(Path("test.pdf"))
        self.assertEqual(result["markdown"], "local")
        self.assertFalse(result["worker_used"])


class AiRepairAuditTests(unittest.TestCase):
    @staticmethod
    def _document() -> dict:
        return {
            "source_file": "test.png",
            "mapped_detail_rows": [
                {
                    "standard": {
                        "序号": "",
                        "物料编码": "PP001840",
                        "物料名称": "半固化片",
                        "说明": "NY3170HCP 7628 48% 150米/卷",
                        "数量": "",
                        "单位": "",
                        "含税单价": "6750",
                        "金额": "",
                        "交货日期": "2026-09-01",
                        "备注": "",
                    },
                    "original": {},
                    "raw_text": "PP001840 半固化片 6750",
                    "page_index": 0,
                    "table_index": 0,
                    "row_index": 1,
                    "confidence": 0.5,
                }
            ],
            "raw_detail_tables": [],
            "issues": [],
        }

    @patch("fangzheng_web_app.purchase_ai_repair.get_ai_repair_config")
    @patch("fangzheng_web_app.purchase_ai_repair.request_repair_json")
    def test_ai_summary_records_applied_fields(self, mocked_request, mocked_config) -> None:
        mocked_config.return_value = SimpleNamespace(available=True, model="test-model", max_rows=12)
        mocked_request.return_value = {
            "repairs": [
                {
                    "row_key": 0,
                    "set_fields": {"数量": "1", "单位": "卷"},
                    "confidence": 0.9,
                    "source_evidence": "局部 OCR",
                    "reason": "补足空字段",
                }
            ]
        }
        repaired = audit_and_repair_purchase_document(self._document())
        summary = repaired["ai_repair_summary"]
        self.assertEqual(summary["returned_repairs"], 1)
        self.assertEqual(summary["applied_fields"], 2)
        self.assertEqual(summary["unresolved_rows"], 0)

    @patch("fangzheng_web_app.purchase_ai_repair.get_ai_repair_config")
    @patch("fangzheng_web_app.purchase_ai_repair.request_repair_json", return_value={"repairs": []})
    def test_zero_applied_with_missing_fields_creates_issue(self, _mocked_request, mocked_config) -> None:
        mocked_config.return_value = SimpleNamespace(available=True, model="test-model", max_rows=12)
        repaired = audit_and_repair_purchase_document(self._document())
        self.assertEqual(repaired["ai_repair_summary"]["applied_fields"], 0)
        self.assertEqual(repaired["ai_repair_summary"]["unresolved_rows"], 1)
        self.assertTrue(any("仍缺少关键字段" in issue.get("message", "") for issue in repaired["issues"]))

    @patch("fangzheng_web_app.purchase_ai_repair.get_ai_repair_config")
    @patch("fangzheng_web_app.purchase_ai_repair.request_repair_json", return_value={"repairs": []})
    def test_each_unresolved_row_gets_its_own_issue(self, _mocked_request, mocked_config) -> None:
        mocked_config.return_value = SimpleNamespace(available=True, model="test-model", max_rows=12)
        document = self._document()
        second = copy.deepcopy(document["mapped_detail_rows"][0])
        second["standard"]["物料编码"] = "PP001880"
        second["raw_text"] = "PP001880 半固化片 11700"
        document["mapped_detail_rows"].append(second)
        repaired = audit_and_repair_purchase_document(document)
        self.assertEqual(repaired["ai_repair_summary"]["unresolved_rows"], 2)

if __name__ == "__main__":
    unittest.main()
