from __future__ import annotations

import os
import sys
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


from fangzheng_web_app.hushi_service import (  # noqa: E402
    calculate_hushi_spec,
    detect_description_column,
    normalize_text,
)


DEFAULT_SAMPLE = Path("D:/") / "\u6587\u6863" / "WXWork" / "1688857334773975" / "Cache" / "File" / "2026-07" / "P2.xlsx"
DEFAULT_RULE_DIR = Path("D:/") / "\u684c\u9762" / "20260608\u590f\u8001\u5e08\u6caa\u58eb\u62a5\u4ef7\u5355"


def assert_equal(left, right, message: str) -> None:
    if left != right:
        raise AssertionError(f"{message}: {left!r} != {right!r}")


def assert_close(left: float | None, right: float, message: str) -> None:
    if left is None or abs(left - right) > 0.01:
        raise AssertionError(f"{message}: {left!r} != {right!r}")


def main() -> int:
    sample_path = Path(os.environ.get("HUSHI_P2_SAMPLE_PATH", DEFAULT_SAMPLE))
    rule_dir = Path(os.environ.get("HUSHI_RULE_DIR", DEFAULT_RULE_DIR))
    if not sample_path.exists():
        print(f"SKIP: HUSHI_P2_SAMPLE_PATH not found: {sample_path}")
        return 0
    if not rule_dir.exists():
        print(f"SKIP: HUSHI_RULE_DIR not found: {rule_dir}")
        return 0

    wb = load_workbook(sample_path, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    desc_col = detect_description_column(ws)
    rule_cache = {}
    results = {}

    for row_idx in range(1, ws.max_row + 1):
        spec = normalize_text(ws.cell(row_idx, desc_col).value)
        if spec:
            results[row_idx] = calculate_hushi_spec(spec, rule_dir, rule_cache=rule_cache)

    success_rows = [row_idx for row_idx, result in results.items() if result.final_price is not None]
    fail_rows = [row_idx for row_idx, result in results.items() if result.final_price is None]
    assert_equal(len(success_rows), 13, "P2 success row count")
    assert_equal(fail_rows, [13, 14, 15], "P2 rows without exact FVLP quote")

    assert_close(results[1].final_price, 304.15, "NY-P2(C) mixed PVLP/RTF price")
    assert_close(results[5].final_price, 136.80, "NY-P2 FVLP price should use non-C quote file")
    assert_close(results[8].final_price, 136.63, "NY-P2 mixed FVLP/RTF price")
    assert_close(results[16].final_price, 72.02, "NY-P2 RTF smaller panel price")
    if results[1].rule_file == results[5].rule_file:
        raise AssertionError("NY-P2 and NY-P2(C) should not use the same quote file")

    print("OK: Hushi P2 smoke test passed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
