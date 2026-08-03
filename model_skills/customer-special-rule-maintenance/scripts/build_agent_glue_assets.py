from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from fangzheng_web_app.transcode_agent_rules import MAPPING_TABLE_HEADERS  # noqa: E402


SOURCE_FILENAME = "_ 胶系主表维护view20260724092307.xlsx"
CONFLICT_NAMES = {
    "NY-A2 非汽车板用途": "2G",
    "NY-A1": "2Z",
    "NY3170HF 汽车板": "3C",
    "NY6300S": "6C",
    "NY-P4": "RA",
    "NY-P4(C)": "AB",
    "NY-P5Q 华为可靠性": "BX",
}


def main() -> None:
    parser = argparse.ArgumentParser(description="将最新版胶系主表转换为营销转码Agent胶系资产")
    parser.add_argument("source", type=Path)
    parser.add_argument("legacy_rules", type=Path)
    parser.add_argument("mapping_workbook", type=Path)
    args = parser.parse_args()
    summary = build_assets(args.source, args.legacy_rules, args.mapping_workbook)
    print(summary)


def build_assets(source: Path, legacy_rules: Path, mapping_workbook: Path) -> dict[str, int]:
    source_rows = _read_rows(source, openpyxl.load_workbook(source, data_only=True).active.title)
    legacy_rows = _read_rows(legacy_rules, "胶系代码", require_valid_code=False)
    legacy_ids = {row["id"] for row in legacy_rows}
    legacy_names = {_normalize(row["name"]) for row in legacy_rows}
    grouped_names: dict[str, list[dict]] = defaultdict(list)
    for row in source_rows:
        grouped_names[_normalize(row["name"])].append(row)

    master = []
    for index, row in enumerate(source_rows, 1):
        flags = []
        if row["id"] not in legacy_ids:
            flags.append("新增胶系")
        elif _normalize(row["name"]) not in legacy_names:
            flags.append("新版名称")
        if len({item["code"] for item in grouped_names[_normalize(row["name"])]}) > 1:
            flags.append("同名多码")
        master.append(
            {
                "映射ID": f"TGM-MASTER-{index:04d}",
                "启用": "是",
                "胶系编号": row["id"],
                "胶系名称": row["name"],
                "胶系分类": row["classification"],
                "输出胶系代码": row["code"],
                "来源文件": SOURCE_FILENAME,
                "来源行号": row["row"],
                "备注": "；".join(flags),
            }
        )

    aliases = []
    seen_aliases = set()
    for row in legacy_rows:
        alias_key = _normalize(row["name"])
        if not alias_key or alias_key in seen_aliases or not _valid_code(row["code"]):
            continue
        seen_aliases.add(alias_key)
        canonical = grouped_names.get(alias_key, [])
        standard = canonical[0] if canonical else None
        aliases.append(
            {
                "映射ID": f"TGM-ALIAS-{len(aliases) + 1:04d}",
                "启用": "是",
                "兼容名称": row["name"],
                "标准胶系编号": standard["id"] if standard else "",
                "标准胶系名称": standard["name"] if standard else "",
                "输出胶系代码": row["code"],
                "来源批次": "历史胶系代码表兼容",
                "规则文本": f"历史写法{row['name']}继续输出{row['code']}",
                "备注": "旧正确逻辑兼容；仅在基础解析失败时使用",
            }
        )

    selections = []
    for index, (name, code) in enumerate(CONFLICT_NAMES.items(), 1):
        selections.append(
            {
                "映射ID": f"TGM-SELECT-{index:04d}",
                "启用": "是",
                "胶系名称": name,
                "条件客户代码": "",
                "条件客户简称": "",
                "条件关键词": "",
                "输出胶系代码": code,
                "优先级": 100,
                "来源批次": "20260724新版胶系表冲突兼容",
                "规则文本": f"{name}同名多码未给选择条件，保持历史稳定输出{code}",
                "备注": "出现四位胶系编号时优先按编号；后续可补客户或用途条件",
            }
        )

    workbook = openpyxl.load_workbook(mapping_workbook)
    _replace_sheet(workbook, "Agent胶系主表", master)
    _replace_sheet(workbook, "Agent胶系兼容别名", aliases)
    _replace_sheet(workbook, "Agent胶系选择规则", selections)
    workbook.save(mapping_workbook)
    return {
        "master": len(master),
        "aliases": len(aliases),
        "selections": len(selections),
        "new_ids": sum("新增胶系" in row["备注"] for row in master),
        "renamed": sum("新版名称" in row["备注"] for row in master),
    }


def _read_rows(
    path: Path,
    sheet_name: str,
    *,
    require_valid_code: bool = True,
) -> list[dict[str, str]]:
    worksheet = openpyxl.load_workbook(path, read_only=True, data_only=True)[sheet_name]
    rows = []
    for row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
        cells = [str(value or "").strip() for value in values]
        cells += [""] * (4 - len(cells))
        if not cells[0] or not cells[1] or (require_valid_code and not _valid_code(cells[3])):
            continue
        rows.append(
            {
                "row": str(row_index),
                "id": cells[0],
                "name": cells[1],
                "classification": cells[2],
                "code": cells[3].upper(),
            }
        )
    return rows


def _replace_sheet(workbook, sheet_name: str, rows: list[dict]) -> None:
    existing_index = workbook.sheetnames.index(sheet_name) if sheet_name in workbook.sheetnames else 0
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    worksheet = workbook.create_sheet(sheet_name, existing_index)
    headers = MAPPING_TABLE_HEADERS[sheet_name]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        worksheet.column_dimensions[column[0].column_letter].width = max(width, 12)


def _valid_code(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Z0-9]{2}", str(value or "").strip().upper()))


def _normalize(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).upper()
    return re.sub(r"[\s_\-]+", "", text)


if __name__ == "__main__":
    main()
