from __future__ import annotations

import argparse
import hashlib
import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


HEADERS = [
    "规则ID", "来源候选ID", "启用", "客户代码", "客户简称", "来源行号", "来源列",
    "业务字段", "规则原文", "语义类型", "目标字段", "标准语义值", "原文目标值",
    "条件JSON", "所需订单字段", "执行方式", "优先级", "模型版本", "提示词SHA256",
    "原文证据", "业务确认", "确认依据", "备注",
]


def main() -> None:
    parser = argparse.ArgumentParser(description="合并已确认的多来源客户CCL语义规则")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("rules", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    output = args.output or args.workbook
    payload = json.loads(args.rules.read_text(encoding="utf-8"))
    workbook = load_workbook(args.workbook)
    sheet = workbook["模型语义规则"]
    _ensure_headers(sheet)
    for row in range(sheet.max_row, 1, -1):
        if str(sheet.cell(row=row, column=1).value or "").startswith("TSR-20260723-"):
            sheet.delete_rows(row)
    existing_ids = {
        str(sheet.cell(row=row, column=1).value or "").strip()
        for row in range(2, sheet.max_row + 1)
    }
    prompt_hash = _prompt_hash(args.rules.parents[2] / "marketing-transcode-semantics/references/semantic-normalizer-prompt.txt")
    sequence = 1
    for rule in payload.get("rules") or []:
        for customer in rule.get("customers") or []:
            rule_id = f"TSR-20260723-{sequence:03d}"
            candidate_id = f"MSR-{int(customer['row']):04d}-{_source_code(rule['source_column'])}-{sequence:02d}"
            sequence += 1
            if rule_id in existing_ids:
                continue
            conditions = list(rule.get("conditions") or [])
            priority = 100 + max(0, len(conditions) - 1) * 10
            values = {
                "规则ID": rule_id,
                "来源候选ID": candidate_id,
                "启用": "是",
                "客户代码": customer.get("code", ""),
                "客户简称": customer.get("name", ""),
                "来源行号": customer.get("row"),
                "来源列": rule.get("source_column", ""),
                "业务字段": rule.get("business_field", ""),
                "规则原文": rule.get("source_text", ""),
                "语义类型": rule.get("semantic_type", ""),
                "目标字段": rule.get("target_field", ""),
                "标准语义值": rule.get("normalized_value", ""),
                "原文目标值": rule.get("stated_target_value", ""),
                "条件JSON": json.dumps(conditions, ensure_ascii=False, separators=(",", ":")),
                "所需订单字段": "；".join(dict.fromkeys(str(item.get("field") or "") for item in conditions)),
                "执行方式": "结构化后可确定性执行",
                "优先级": priority,
                "模型版本": "业务确认直接结构化",
                "提示词SHA256": prompt_hash,
                "原文证据": rule.get("evidence_text", ""),
                "业务确认": "确认",
                "确认依据": payload.get("approval_basis", ""),
                "备注": "多来源CCL规则清洗；模型只做运行时订单语义标准化",
            }
            sheet.append([values.get(header, "") for header in HEADERS])
            existing_ids.add(rule_id)
    pending_sheet = workbook["待业务确认"]
    for row in range(pending_sheet.max_row, 1, -1):
        if str(pending_sheet.cell(row=row, column=1).value or "").startswith("PENDING-20260723-"):
            pending_sheet.delete_rows(row)
    existing_pending = {
        str(pending_sheet.cell(row=row, column=1).value or "").strip()
        for row in range(2, pending_sheet.max_row + 1)
    }
    for index, item in enumerate(payload.get("pending") or [], start=1):
        pending_id = f"PENDING-20260723-{index:03d}"
        if pending_id in existing_pending:
            continue
        pending_sheet.append(
            [
                pending_id,
                "",
                item.get("customer", ""),
                "",
                "待字段确认",
                item.get("source_text", ""),
                item.get("reason", ""),
                "待业务确认",
            ]
        )
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(output)


def _ensure_headers(sheet) -> None:
    current = [str(cell.value or "").strip() for cell in sheet[1]]
    if current == HEADERS:
        return
    if "来源列" not in current:
        insert_at = current.index("来源行号") + 2
        sheet.insert_cols(insert_at)
        sheet.cell(row=1, column=insert_at, value="来源列")
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=insert_at, value="CCL特殊规则")
        current = [str(cell.value or "").strip() for cell in sheet[1]]
    if current != HEADERS:
        raise ValueError(f"正式语义规则表头不兼容：{current}")
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(row=1, column=column)
        if source.has_style:
            for row in range(2, sheet.max_row + 1):
                if not sheet.cell(row=row, column=column).has_style:
                    sheet.cell(row=row, column=column)._style = copy(source._style)


def _source_code(value: str) -> str:
    return {"CCL特殊规则": "C", "通用特殊规则": "G", "非影响转码备注": "N"}.get(value, "X")


def _prompt_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest() if path.exists() else ""


if __name__ == "__main__":
    main()
