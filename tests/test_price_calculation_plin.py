from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook

from fangzheng_web_app.price_calculation_service import calculate_plin_spec, load_plin_rules


class PlinPriceCalculationTests(unittest.TestCase):
    def test_copper_adder_uses_the_matched_quote_sheet(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "plin.xlsx"
            workbook = Workbook()
            ny2170 = workbook.active
            ny2170.title = "NY2170基板"
            ny2170.append(["产品类别", "厚度mm", "铜厚", "铜箔类型", "组合叠构", "37*49", "41*49", "43*49"])
            ny2170.append(["NY2170", 0.2, "1/1", "HTE", "7628*1", 170.7, 189.48, 198.01])
            ny2170.append(["2/2比1/1高136.99RMB/SH(36*48)，3/3oz比1/1OZ高285.5RMB/SH(36*48)"])
            ny3170 = workbook.create_sheet("NY3170HC基板")
            ny3170.append(["2/2比1/1高124.54RMB/SH(36*48)，3/3oz比1/1OZ高259.55RMB/SH(36*48)"])
            pp = workbook.create_sheet("NY2170PP")
            pp.append(["Products", "Glass type", "Resin Content", "Length (m)", "Per M"])
            pp.append(["NY2170", "7628", "43%-45%", 100, 35.2])
            workbook.save(path)

            rules = load_plin_rules(path)
            result = calculate_plin_spec(
                "南亚新材料NY2170 FR4 0.2 2/2 37.3经x49.3纬 TG170(7628*1) HTE 有卤 黄 无标 耐CAF 不含铜",
                rules,
            )

        self.assertEqual(("成功", 307.69), (result.status, result.price))
        self.assertIn("加价136.99", result.note)
