from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from fangzheng_web_app import db
from fangzheng_web_app.mail_transcode_agent import mail_store
from fangzheng_web_app.order_entry_service import (
    _apply_customer_spec_matches,
    _line_entry,
    _line_from_pipeline_row,
    _merge_initial_rows,
    _initial_template_data,
    _rows_from_pdf_or_image,
    build_domestic_export,
    get_or_create_template,
    reextract_all_templates,
    save_template,
)
from fangzheng_web_app.purchase_factory_mapper import FACTORY_DETAIL_HEADERS
from fangzheng_web_app.order_intake_service import bootstrap_cases, list_cases


class OrderEntryTemplateTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_patch = patch.object(db, "DATABASE_PATH", Path(self.temp_dir.name) / "entry.sqlite3")
        self.db_patch.start()
        db.init_db()
        self.account_id = mail_store.create_or_update_account(
            "orders@example.com", owner_employee_id="employee-a", auth_code="auth-code"
        )
        mail_store.upsert_message(
            self.account_id, folder="INBOX", uid="entry-1", message_id="<entry-1@example.com>",
            subject="采购订单", sender="buyer@example.com", sent_at="2026-08-19 09:00:00",
            received_at="2026-08-19 09:00:00", body_html="", body_text=(
                "HJ20260818013\nA1A150224149YNNYZ002\n南亚\nNY2150\nFR-4 1.5\n±\n"
                "0.075MM 2/2 41\n英寸\n*49\n英寸含铜无水印\n(TG>150)A\n级\n300"
            ), eml_path="", is_order=1,
        )
        bootstrap_cases("employee-a", self.account_id)
        self.case_id = list_cases("employee-a", "2026-08-19", "new_order", self.account_id)[0]["id"]

    def tearDown(self) -> None:
        self.db_patch.stop()
        self.temp_dir.cleanup()

    def test_empty_configured_spec_field_shows_its_position_as_placeholder(self) -> None:
        template = (
            Path(__file__).resolve().parents[1]
            / "templates"
            / "order_automation_entry_template.html"
        ).read_text(encoding="utf-8")

        self.assertIn("input.placeholder=`第${field.position}位`", template)
        self.assertIn(".oe-match-field input::placeholder{color:#9aa8ba;opacity:1}", template)
        self.assertIn(
            "shipToCode.value=code.value.trim()",
            template,
        )
        self.assertIn("<datalist id=\"orderTypeOptions\">", template)
        self.assertIn("<datalist id=\"customerCodeOptions\">", template)
        self.assertNotIn("customer.customer_name", template)
        self.assertIn("220:['1','1'],221:['3','2'],331:['3','1']", template)
        self.assertIn("normalizeChoice(orderType,3)", template)
        self.assertIn("normalizeCustomerCode();shipToCode.value=code.value.trim()", template)
        self.assertIn('data-field="material_status" aria-label="料号状态"', template)
        self.assertIn('<option value="查询" selected>查询</option>', template)
        self.assertIn("materialCodeOptions", template)
        self.assertIn("syncPair", template)
        self.assertIn("NYEOS订单号：{{ nyeos_order_number }}", template)

    def test_saved_template_reopens_and_exports_same_values(self) -> None:
        _case, template = get_or_create_template(self.case_id, "employee-a")
        self.assertEqual(template["header"]["customer_order_number"], "HJ20260818013")
        self.assertEqual(template["header"]["order_type"], "220")
        self.assertEqual(template["header"]["type_1"], "1")
        self.assertEqual(template["header"]["type_2"], "1")
        self.assertEqual(template["header"]["ledger"], "KL01")
        self.assertTrue(all(line["values"]["material_status"] == "查询" for line in template["lines"]))
        saved = save_template(self.case_id, "employee-a", {
            "header": {
                "order_type": "SO", "bill_to_customer_code": "C001", "ledger": "151",
                "tax_type": "VAT", "customer_invoice_number": "INV-001",
                "commission_rate": "2.5%",
            },
            "lines": [{"values": {
                "line_no": "1", "material_status": "新增", "product_code": "P001", "product_name": "南亚NY2150",
                "customer_product_code": "A1A150224149YNNYZ002", "quantity": "300",
                "customer_spec_match": "NY2150", "product_type": "基板",
            }}],
        })
        self.assertEqual(saved["current_version"], 1)
        _case, reopened = get_or_create_template(self.case_id, "employee-a")
        self.assertEqual(reopened["header"]["ledger"], "151")
        self.assertEqual(reopened["lines"][0]["values"]["material_status"], "新增")
        output, _name = build_domestic_export(self.case_id, "employee-a")
        destination = Path(self.temp_dir.name) / "export.xlsx"
        destination.write_bytes(output.getvalue())
        book = load_workbook(destination, data_only=True)
        sheet = book["内销"]
        self.assertEqual(sheet["A2"].value, "SO")
        self.assertEqual(sheet["D2"].value, "C001")
        self.assertEqual(sheet["H2"].value, "151")
        self.assertEqual(sheet["I2"].value, "VAT")
        self.assertEqual(sheet["J2"].value, "INV-001")
        self.assertEqual(sheet["K2"].value, "2.5%")
        self.assertEqual(sheet["I1"].value, "税种（选填）")
        self.assertEqual(sheet["J1"].value, "客户发票号（选填）")
        self.assertEqual(sheet["K1"].value, "佣金比率（选填）")
        self.assertEqual(sheet["B3"].value, "料号状态（选填）")
        self.assertEqual(sheet["B4"].value, "新增")
        self.assertEqual(sheet["E4"].value, "A1A150224149YNNYZ002")
        self.assertIsNone(sheet["G4"].value)
        self.assertEqual(sheet["H4"].value, "基板")
        self.assertEqual(sheet["J4"].value, 300)
        # The business template is a populated example; exporting a different
        # mail must never carry its sample detail rows into the new file.
        self.assertIsNone(sheet["E5"].value)
        self.assertIsNone(sheet["J6"].value)
        book.close()

    def test_material_status_defaults_to_query_and_rejects_unknown_values(self) -> None:
        get_or_create_template(self.case_id, "employee-a")
        saved = save_template(self.case_id, "employee-a", {
            "header": {"order_type": "220", "bill_to_customer_code": "C001", "ledger": "KL01"},
            "lines": [{"values": {
                "line_no": "1", "customer_product_code": "CUST-1", "quantity": "20",
            }}],
        })
        self.assertEqual(saved["lines"][0]["values"]["material_status"], "查询")

        with self.assertRaisesRegex(ValueError, "料号状态只能选择"):
            save_template(self.case_id, "employee-a", {
                "header": {"order_type": "220", "bill_to_customer_code": "C001", "ledger": "KL01"},
                "lines": [{"values": {
                    "line_no": "1", "material_status": "删除",
                    "customer_product_code": "CUST-1", "quantity": "20",
                }}],
            })

    def test_line_number_tracks_customer_order_sequence_or_auto_increments(self) -> None:
        get_or_create_template(self.case_id, "employee-a")
        saved = save_template(self.case_id, "employee-a", {
            "header": {"order_type": "220", "bill_to_customer_code": "C001", "ledger": "KL01"},
            "lines": [
                {"values": {
                    "line_no": "1", "customer_order_seq": "10",
                    "customer_product_code": "CUST-10", "quantity": "20",
                }},
                {"values": {
                    "line_no": "2", "customer_order_seq": "",
                    "customer_product_code": "CUST-AUTO", "quantity": "30",
                }},
            ],
        })
        values_by_code = {
            line["values"]["customer_product_code"]: line["values"]
            for line in saved["lines"]
        }
        self.assertEqual(values_by_code["CUST-10"]["line_no"], "10")
        self.assertEqual(values_by_code["CUST-10"]["customer_order_seq"], "10")
        self.assertEqual(values_by_code["CUST-AUTO"]["line_no"], "2")
        self.assertEqual(values_by_code["CUST-AUTO"]["customer_order_seq"], "2")

    def test_customer_spec_match_recalculates_from_header_code_and_line_values(self) -> None:
        lines = [{
            "values": {
                "customer_spec": "NY2150_TG150_1.6MM",
                "customer_spec_match": "旧值",
                "product_type": "基板",
            },
            "sources": {"customer_spec_match": {"label": "旧来源", "reference": "旧规则"}},
        }]
        with patch(
            "fangzheng_web_app.order_entry_service.build_customer_spec_match",
            return_value="NY2150_*_1.6MM",
        ) as matcher:
            matched = _apply_customer_spec_matches(
                {"bill_to_customer_code": "C001"}, lines
            )

        self.assertEqual("NY2150_*_1.6MM", matched[0]["values"]["customer_spec_match"])
        self.assertEqual("查询", matched[0]["values"]["material_status"])
        self.assertEqual("客户规格对照表", matched[0]["sources"]["customer_spec_match"]["label"])
        matcher.assert_called_once_with("C001", "基板", "NY2150_TG150_1.6MM")

    def test_matched_customer_code_populates_header_and_spec_match_on_generation(self) -> None:
        with db.db_cursor() as conn:
            conn.execute(
                "UPDATE order_intake_cases SET customer_id=?,customer_match_status=? WHERE id=?",
                (42, "matched", self.case_id),
            )
        with patch(
            "fangzheng_web_app.order_entry_service.get_customer",
            return_value={"id": 42, "customer_code": "123036"},
        ), patch(
            "fangzheng_web_app.order_entry_service.build_customer_spec_match",
            return_value="自动客户规格匹配",
        ) as matcher:
            _case, template = get_or_create_template(self.case_id, "employee-a")

        self.assertEqual("123036", template["header"]["bill_to_customer_code"])
        self.assertEqual("123036", template["header"]["ship_to_customer_code"])
        self.assertTrue(template["lines"])
        self.assertTrue(all(
            line["values"]["customer_spec_match"] == "自动客户规格匹配"
            for line in template["lines"]
        ))
        self.assertTrue(all(
            line["sources"]["customer_spec_match"]["label"] == "客户规格对照表"
            for line in template["lines"]
        ))
        self.assertTrue(all(call.args[0] == "123036" for call in matcher.call_args_list))

    def test_existing_blank_template_is_backfilled_after_customer_match(self) -> None:
        _case, original = get_or_create_template(self.case_id, "employee-a")
        self.assertEqual("", original["header"]["bill_to_customer_code"])
        with db.db_cursor() as conn:
            conn.execute(
                "UPDATE order_intake_cases SET customer_id=?,customer_match_status=? WHERE id=?",
                (43, "matched", self.case_id),
            )
        with patch(
            "fangzheng_web_app.order_entry_service.get_customer",
            return_value={"id": 43, "customer_code": "104253"},
        ), patch(
            "fangzheng_web_app.order_entry_service.build_customer_spec_match",
            return_value="补齐后的规格匹配",
        ):
            _case, template = get_or_create_template(self.case_id, "employee-a")

        self.assertEqual("104253", template["header"]["bill_to_customer_code"])
        self.assertEqual("104253", template["header"]["ship_to_customer_code"])
        self.assertTrue(all(
            line["values"]["customer_spec_match"] == "补齐后的规格匹配"
            for line in template["lines"]
        ))

    def test_existing_bill_to_code_backfills_only_blank_ship_to_code(self) -> None:
        _case, original = get_or_create_template(self.case_id, "employee-a")
        with db.db_cursor() as conn:
            header = dict(original["header"])
            header["bill_to_customer_code"] = "123036"
            header["ship_to_customer_code"] = ""
            conn.execute(
                "UPDATE order_entry_templates SET header_json=? WHERE id=?",
                (json.dumps(header, ensure_ascii=False), original["id"]),
            )
            conn.execute(
                "UPDATE order_intake_cases SET customer_id=?,customer_match_status=? WHERE id=?",
                (44, "matched", self.case_id),
            )
        with patch(
            "fangzheng_web_app.order_entry_service.get_customer",
            return_value={"id": 44, "customer_code": "123036"},
        ):
            _case, template = get_or_create_template(self.case_id, "employee-a")

        self.assertEqual("123036", template["header"]["bill_to_customer_code"])
        self.assertEqual("123036", template["header"]["ship_to_customer_code"])

    def test_empty_customer_code_clears_customer_spec_match(self) -> None:
        lines = [{
            "values": {"customer_spec": "NY2150", "customer_spec_match": "旧值", "product_type": "基板"},
            "sources": {"customer_spec_match": {"label": "旧来源"}},
        }]
        with patch("fangzheng_web_app.order_entry_service.build_customer_spec_match") as matcher:
            matched = _apply_customer_spec_matches({"bill_to_customer_code": ""}, lines)

        self.assertEqual("", matched[0]["values"]["customer_spec_match"])
        self.assertNotIn("customer_spec_match", matched[0]["sources"])
        matcher.assert_not_called()

    def test_manual_customer_spec_match_is_preserved_for_same_context(self) -> None:
        lines = [{
            "values": {
                "customer_spec": "NY2150_TG150_1.6MM",
                "customer_spec_match": "人工调整后的结果",
                "product_type": "基板",
            },
            "sources": {"customer_spec_match": {
                "label": "人工修改",
                "reference": "录单模板手工修改",
                "context": '["C001","基板","NY2150_TG150_1.6MM"]',
            }},
        }]
        with patch("fangzheng_web_app.order_entry_service.build_customer_spec_match") as matcher:
            matched = _apply_customer_spec_matches(
                {"bill_to_customer_code": "C001"}, lines
            )

        self.assertEqual("人工调整后的结果", matched[0]["values"]["customer_spec_match"])
        self.assertEqual("人工修改", matched[0]["sources"]["customer_spec_match"]["label"])
        matcher.assert_not_called()

    def test_manual_customer_spec_match_recalculates_after_context_change(self) -> None:
        lines = [{
            "values": {
                "customer_spec": "NY2150_TG150_1.6MM",
                "customer_spec_match": "人工调整后的结果",
                "product_type": "PP",
            },
            "sources": {"customer_spec_match": {
                "label": "人工修改",
                "context": '["C001","基板","NY2150_TG150_1.6MM"]',
            }},
        }]
        with patch(
            "fangzheng_web_app.order_entry_service.build_customer_spec_match",
            return_value="重新自动匹配",
        ) as matcher:
            matched = _apply_customer_spec_matches(
                {"bill_to_customer_code": "C001"}, lines
            )

        self.assertEqual("重新自动匹配", matched[0]["values"]["customer_spec_match"])
        self.assertEqual("客户规格对照表", matched[0]["sources"]["customer_spec_match"]["label"])
        matcher.assert_called_once_with("C001", "PP", "NY2150_TG150_1.6MM")

    def test_save_template_ignores_manual_match_and_persists_generated_value(self) -> None:
        get_or_create_template(self.case_id, "employee-a")
        with patch(
            "fangzheng_web_app.order_entry_service.build_customer_spec_match",
            return_value="NY2150_*_1.6MM",
        ):
            saved = save_template(self.case_id, "employee-a", {
                "header": {"bill_to_customer_code": "C001"},
                "lines": [{"values": {
                    "customer_spec": "NY2150_TG150_1.6MM",
                    "customer_spec_match": "手工旧值",
                    "product_type": "基板",
                }}],
            })

        self.assertEqual(
            "NY2150_*_1.6MM",
            saved["lines"][0]["values"]["customer_spec_match"],
        )

    def test_excel_attachment_is_merged_into_the_same_email_template(self) -> None:
        attachment_path = Path(self.temp_dir.name) / "客户订单.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "订单"
        sheet.append(["客户订单号", "客户产品编号", "客户规格", "数量", "含税单价", "备注"])
        sheet.append(["PO-20260819", "CUST-001", "FR-4 1.6MM", 500, 12.5, "加急"])
        workbook.save(attachment_path)
        workbook.close()
        with db.db_cursor() as conn:
            mail_id = conn.execute("SELECT id FROM mail_messages WHERE uid='entry-1'").fetchone()["id"]
            conn.execute(
                """INSERT INTO mail_attachments(mail_id,filename,content_type,size_bytes,sha256,stored_path,is_inline,parse_status,created_at)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (mail_id, "客户订单.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", attachment_path.stat().st_size, "", str(attachment_path), 0, "", db.utcnow()),
            )
        _case, template = get_or_create_template(self.case_id, "employee-a")
        matching = [line for line in template["lines"] if line["values"].get("customer_product_code") == "CUST-001"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0]["values"]["quantity"], "500")
        self.assertEqual(matching[0]["values"]["unit_price"], "12.5")
        self.assertEqual(matching[0]["sources"]["quantity"]["label"], "附件：客户订单.xlsx")

    def test_html_mail_table_uses_shared_domestic_template_mapping(self) -> None:
        body_html = """
        <table><tr><th>PO单号</th><th>PO项目号</th><th>物料编码</th><th>物料名称</th>
        <th>物料规格</th><th>数量</th><th>单位</th><th>单价</th><th>交期</th></tr>
        <tr><td>MZPOM12608190027</td><td>2713183</td><td>10101008248</td><td>FR-4</td>
        <td>南亚新材料 NY6180L 0.127 mm</td><td>20</td><td>张</td><td>215.6814</td><td>2026-08-28</td></tr>
        </table>
        """
        with patch("fangzheng_web_app.order_entry_service.get_enabled_extraction_maps", return_value=[]):
            header, rows = _initial_template_data({
                "id": 999,
                "body_html": body_html,
                "body_text": "",
                "attachments": [],
                "detected_fields": {},
                "customer_id": None,
            })
        self.assertEqual(header["customer_order_number"], "MZPOM12608190027")
        self.assertEqual(rows[0]["values"]["customer_product_code"], "10101008248")
        self.assertEqual(rows[0]["values"]["quantity"], "20")
        self.assertEqual(rows[0]["values"]["product_type"], "基板")
        self.assertEqual(rows[0]["sources"]["quantity"]["label"], "邮件正文表格")

    def test_ccl_and_manual_only_fields_follow_conservative_extraction_policy(self) -> None:
        line = _line_entry(
            {
                "customer_product_code": "CUST-001",
                "customer_spec": "FR-4 1.6MM",
                "quantity": "500",
                "product_code": "FACTORY-001",
                "product_name": "不应自动填入的品名",
                "origin": "中国",
                "one_to_many": "一对多关系",
            },
            label="附件：订单.xlsx", reference="订单 第 2 行",
        )
        values = line["values"]
        self.assertEqual(values["quantity"], "500")
        self.assertEqual(values["customer_product_code"], "CUST-001")
        for field in ("product_code", "product_name", "origin", "one_to_many"):
            self.assertEqual(values[field], "")
            self.assertNotIn(field, line["sources"])

    def test_pp_roll_converts_quantity_only_with_explicit_unit_and_metre_value(self) -> None:
        line = _line_entry(
            {
                "customer_product_code": "PP-01",
                "customer_spec": "PP 1080 300M/卷",
                "quantity": "2",
                "remark": "客户加急",
            },
            label="附件：PP订单.xlsx", reference="订单 第 2 行", quantity_unit="卷",
        )
        self.assertEqual(line["values"]["quantity"], "600")
        self.assertEqual(line["values"]["remark"], "客户加急；PP米数：300米")

    def test_pp_sheet_keeps_customer_quantity_and_unknown_pp_is_blank(self) -> None:
        small_piece = _line_entry(
            {"customer_spec": "PP 1080 300m", "quantity": "30"},
            label="附件：PP订单.xlsx", reference="订单 第 2 行", quantity_unit="张",
        )
        unknown = _line_entry(
            {"customer_spec": "PP 1080 300m", "quantity": "2"},
            label="附件：PP订单.xlsx", reference="订单 第 3 行", quantity_unit="",
        )
        self.assertEqual(small_piece["values"]["quantity"], "30")
        self.assertEqual(small_piece["values"]["remark"], "PP米数：300米")
        self.assertEqual(unknown["values"]["quantity"], "")
        self.assertEqual(unknown["values"]["remark"], "PP米数：300米")

    def test_pipeline_mapping_uses_raw_description_and_keeps_pp_detail_rows(self) -> None:
        first = _line_from_pipeline_row(
            {
                "original": {
                    "PO项目号 Project No": "2713186",
                    "物料编码 Material Code": "10601001676",
                    "物料品名 Material Name": "半固化片",
                    "物料描述 Description": "南亚新材料 NY6180LP 106 RC=75% 经300.00 m 纬49.50 inch",
                    "单位 Unit": "卷",
                    "数量 Quantity": "0.1",
                    "不含税单价 Not tax inclusive Unit Price": "10914.1593",
                },
                "standard": {"数量": "0.1", "单位": "卷"},
            },
            "MZPOM12608190027", "附件：采购订单.pdf", "识别明细第 2 行", 1,
        )
        second = _line_from_pipeline_row(
            {
                "original": {
                    "PO项目号 Project No": "2713185",
                    "物料编码 Material Code": "10601001674",
                    "物料品名 Material Name": "半固化片",
                    "物料描述 Description": "南亚新材料 NY6180LP 2116 RC=56% 经200.00 m 纬49.50 inch",
                    "单位 Unit": "卷",
                    "数量 Quantity": "0.1",
                },
                "standard": {"数量": "0.1", "单位": "卷"},
            },
            "MZPOM12608190027", "附件：采购订单.pdf", "识别明细第 3 行", 2,
        )
        self.assertEqual(first["values"]["customer_spec"], "南亚新材料 NY6180LP 106 RC=75% 经300.00 m 纬49.50 inch")
        self.assertEqual(first["values"]["customer_product_code"], "10601001676")
        self.assertEqual(first["values"]["quantity"], "30")
        self.assertEqual(first["values"]["price_before_tax"], "10914.1593")
        self.assertEqual(first["values"]["unit_price"], "")
        self.assertEqual(first["values"]["product_name"], "")
        self.assertEqual(second["values"]["quantity"], "20")
        self.assertEqual(len(_merge_initial_rows([first, second])), 2)

    def test_pipeline_mapping_keeps_a_real_tax_inclusive_unit_price(self) -> None:
        line = _line_from_pipeline_row(
            {
                "original": {
                    "物料描述 Description": "测试规格",
                    "数量 Quantity": "2",
                    "含税单价 Tax inclusive Unit Price": "12.50",
                },
                "standard": {"数量": "2"},
            },
            "PO123456", "附件：采购订单.pdf", "识别明细第 2 行", 1,
        )
        self.assertEqual(line["values"]["price_before_tax"], "")
        self.assertEqual(line["values"]["unit_price"], "12.50")

    def test_pdf_attachment_reuses_pdf_domestic_mapping(self) -> None:
        document = {
            "mapped_detail_rows": [{
                "original": {"物料描述": "PP 1080 300M/卷"},
                "standard": {"物料编码": "CUST-PP", "物料名称": "半固化片"},
            }],
            "factory_import": {
                "main_values": ["", "", "", "", "", "", "", "PO-001"],
                "rows": [{
                    FACTORY_DETAIL_HEADERS[0]: "9",
                    FACTORY_DETAIL_HEADERS[3]: "CUST-PP",
                    FACTORY_DETAIL_HEADERS[4]: "2026-08-30",
                    FACTORY_DETAIL_HEADERS[5]: "600",
                    FACTORY_DETAIL_HEADERS[6]: "10",
                    FACTORY_DETAIL_HEADERS[7]: "11.3",
                    FACTORY_DETAIL_HEADERS[11]: "加急",
                }],
            },
        }
        with patch("fangzheng_web_app.order_entry_service.recognize_purchase_order_document", return_value=document), patch("fangzheng_web_app.order_entry_service.project_factory_document"):
            rows = _rows_from_pdf_or_image(Path("/tmp/PO-001.pdf"), "PO-001.pdf")
        self.assertEqual(rows[0]["values"]["customer_product_code"], "CUST-PP")
        self.assertEqual(rows[0]["values"]["quantity"], "600")
        self.assertEqual(rows[0]["values"]["price_before_tax"], "10")
        self.assertEqual(rows[0]["extracted_header"]["customer_order_number"], "PO-001")
        self.assertEqual(rows[0]["values"]["product_type"], "PP")
        self.assertEqual(rows[0]["sources"]["customer_product_code"]["label"], "附件：PO-001.pdf")

    def test_batch_reextract_keeps_header_and_backs_up_before_replacing_lines(self) -> None:
        get_or_create_template(self.case_id, "employee-a")
        save_template(self.case_id, "employee-a", {
            "header": {"order_type": "SO", "bill_to_customer_code": "C001", "ledger": "151"},
            "lines": [{"values": {
                "line_no": "1", "product_name": "历史错误品名",
                "customer_product_code": "OLD-CODE", "quantity": "999",
            }}],
        })

        summary = reextract_all_templates("employee-a")
        self.assertEqual(summary["template_count"], 1)
        self.assertEqual(summary["previous_line_count"], 1)
        self.assertGreaterEqual(summary["line_count"], 1)
        _case, template = get_or_create_template(self.case_id, "employee-a")
        self.assertEqual(template["header"]["ledger"], "151")
        self.assertEqual(template["current_version"], 1)
        self.assertTrue(all(line["values"]["product_name"] == "" for line in template["lines"]))
        with db.db_cursor() as conn:
            versions = conn.execute(
                "SELECT version_number,lines_json FROM order_entry_template_versions ORDER BY version_number"
            ).fetchall()
        self.assertEqual([row["version_number"] for row in versions], [1])
        self.assertIn("历史错误品名", versions[0]["lines_json"])
