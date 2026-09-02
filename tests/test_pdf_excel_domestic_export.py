from __future__ import annotations

import json
import tempfile
import unittest
import zipfile
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from flask import Flask
from openpyxl import load_workbook

from fangzheng_web_app.pdf_excel_domestic_export import (
    build_domestic_rows,
    build_domestic_workbook,
    build_job_domestic_export,
)
from fangzheng_web_app.purchase_factory_mapper import project_factory_document
from fangzheng_web_app.routes import bp


def _document() -> dict:
    document = {
        "source_file": "PO-001.pdf",
        "template_id": "",
        "header_info": {"订单号": "PO20260821001"},
        "pages": [],
        "issues": [],
        "mapped_detail_rows": [
            {
                "original": {
                    "序号": "9",
                    "物料描述": "PP NY2150 1080 300M/卷",
                    "税前单价": "3000",
                    "备注": "加急",
                },
                "standard": {
                    "序号": "9",
                    "物料编码": "CUST-PP",
                    "物料名称": "半固化片",
                    "说明": "PP NY2150 1080 300M/卷",
                    "数量": "2",
                    "单位": "卷",
                    # The recognizer may expose the same source value through
                    # this generic field; the explicit untaxed heading above
                    # must still keep the tax-inclusive price blank.
                    "含税单价": "3000",
                    "交货日期": "2026-08-30",
                    "备注": "加急",
                },
                "page_index": 0,
                "raw_text": "",
            },
            {
                "original": {"物料描述": "覆铜板 FR-4", "单价": "12.5"},
                "standard": {
                    "序号": "7",
                    "物料编码": "CUST-CCL",
                    "物料名称": "覆铜板",
                    "说明": "FR-4",
                    "数量": "5",
                    "单位": "张",
                    "交货日期": "2026-09-01",
                },
                "page_index": 0,
                "raw_text": "",
            },
            {
                "original": {"物料描述": "PP/CCL 待确认材料", "含税单价": "8.8"},
                "standard": {
                    "物料编码": "CUST-UNKNOWN",
                    "物料名称": "PP/CCL 待确认材料",
                    "数量": "3",
                    "单位": "张",
                },
                "page_index": 0,
                "raw_text": "",
            },
            {
                "original": {"物料描述": "PP NY2150", "单价": "1000"},
                "standard": {
                    "物料编码": "CUST-ROLL-UNKNOWN",
                    "物料名称": "半固化片",
                    "说明": "PP NY2150",
                    "数量": "2",
                    "单位": "卷",
                },
                "page_index": 0,
                "raw_text": "",
            },
        ],
    }
    config = SimpleNamespace(
        available=False,
        version_id=None,
        fingerprint="",
        prompt_digest="",
    )
    project_factory_document(document, config=config)
    return document


class PdfExcelDomesticExportTests(unittest.TestCase):
    def test_explicit_untaxed_price_does_not_duplicate_into_unit_price(self) -> None:
        document = _document()
        factory_row = document["factory_import"]["rows"][0]
        self.assertEqual(factory_row["税前单价（选填）"], "10")
        self.assertEqual(factory_row["单价（选填）"], "")

        _header, domestic_rows = build_domestic_rows(document)
        domestic_row = domestic_rows[0]
        self.assertEqual(domestic_row["税前单价（选填）"], "10")
        self.assertEqual(domestic_row["单价（选填）"], "")

    def test_workbook_uses_new_template_without_changing_existing_result_shape(self) -> None:
        output = build_domestic_workbook(_document())
        book = load_workbook(output, data_only=True)
        self.assertEqual(book.sheetnames, ["内销", "外销"])
        sheet = book["内销"]

        self.assertEqual(
            [sheet.cell(1, column).value for column in range(1, 12)],
            [
                "单别（必填）",
                "类型1（选填）",
                "类型2（选填）",
                "账款客户编号（必填）",
                "送货客户编号（选填）",
                "送货厂别（选填）",
                "客户订单号（选填）",
                "账套（必填）",
                "税种（选填）",
                "客户发票号（选填）",
                "佣金比率（选填）",
            ],
        )
        self.assertEqual(
            [sheet.cell(2, column).value for column in range(1, 12)],
            ["220", "1", "1", None, None, None, "PO20260821001", "KL01", None, None, None],
        )
        self.assertEqual(sheet["A4"].value, "1")
        self.assertIsNone(sheet["B4"].value)
        self.assertIsNone(sheet["C4"].value)
        self.assertEqual(sheet["D4"].value, "CUST-PP")
        self.assertEqual(sheet["E4"].value, "PP NY2150 1080 300M/卷")
        self.assertIsNone(sheet["F4"].value)
        self.assertEqual(sheet["G4"].value, "PP")
        self.assertEqual(sheet["I4"].value, 600)
        self.assertEqual(sheet["J4"].value, 10)
        self.assertIsNone(sheet["K4"].value)
        self.assertEqual(sheet["M4"].value, "9")
        self.assertEqual(sheet["O4"].value, "加急；2卷")

        self.assertEqual(sheet["A5"].value, "2")
        self.assertEqual(sheet["G5"].value, "基板")
        self.assertEqual(sheet["K5"].value, 12.5)
        self.assertEqual(sheet["M5"].value, "7")
        self.assertIsNone(sheet["G6"].value)
        self.assertEqual(sheet["M6"].value, "3")
        self.assertEqual(sheet["G7"].value, "PP")
        self.assertEqual(sheet["I7"].value, 2)
        self.assertEqual(sheet["K7"].value, 1000)
        book.close()

    def test_fr4_spacing_variant_is_classified_as_base_material(self) -> None:
        document = _document()
        document["mapped_detail_rows"] = [
            {
                "original": {"规格": "FR -4 1.6MM", "单价": "12.5"},
                "standard": {"序号": "1", "物料编码": "FR4-001", "数量": "1", "单位": "张"},
            }
        ]
        document["factory_import"]["rows"] = [document["factory_import"]["rows"][1]]
        _header, rows = build_domestic_rows(document)

        self.assertEqual(rows[0]["产品类型（PP、基板）"], "基板")

    def test_template_sample_rows_are_cleared(self) -> None:
        document = _document()
        document["mapped_detail_rows"] = document["mapped_detail_rows"][:1]
        document["factory_import"]["rows"] = document["factory_import"]["rows"][:1]
        output = build_domestic_workbook(document)
        book = load_workbook(output, data_only=True)
        sheet = book["内销"]
        self.assertEqual(sheet["D4"].value, "CUST-PP")
        self.assertIsNone(sheet["D5"].value)
        self.assertIsNone(sheet["E6"].value)
        book.close()

    def test_customer_spec_combines_purchase_spec_and_material_name(self) -> None:
        document = _document()
        document["mapped_detail_rows"][0]["original"] = {
            "规格": "526*626mm纬向",
            "材料名称": "考试板 RTF铜箔 PP(半固化片) 1067 RC73%",
            "单价": "13.30",
        }

        _header, rows = build_domestic_rows(document)

        self.assertEqual(
            rows[0]["客户规格（选填）"],
            "526*626mm纬向 考试板 RTF铜箔 PP(半固化片) 1067 RC73%",
        )

    def test_customer_spec_does_not_repeat_spec_already_in_material_name(self) -> None:
        document = _document()
        document["mapped_detail_rows"][0]["original"] = {
            "规格": "526*626mm纬向",
            "材料名称": "考试板 PP 1067 526*626mm纬向",
            "单价": "13.30",
        }

        _header, rows = build_domestic_rows(document)

        self.assertEqual(rows[0]["客户规格（选填）"], "考试板 PP 1067 526*626mm纬向")

    def test_customer_spec_prefers_material_description_over_item_name(self) -> None:
        document = _document()
        document["mapped_detail_rows"][0]["original"] = {
            "物料品名 Material Name": "FR-4",
            "物料描述 Description": "南亚新材料 NY2150 0.8 mm H/H HTE/HTE 7628x4 TG150",
            "单价": "182.6637",
        }
        document["mapped_detail_rows"][0]["standard"].update(
            {
                "物料名称": "FR-4",
                "说明": "南亚新材料 NY2150 0.8 mm H/H HTE/HTE 7628x4 TG150",
            }
        )

        _header, rows = build_domestic_rows(document)

        self.assertEqual(
            rows[0]["客户规格（选填）"],
            "南亚新材料 NY2150 0.8 mm H/H HTE/HTE 7628x4 TG150",
        )

    def test_job_export_returns_xlsx_or_one_workbook_per_order_in_zip(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            job_dir = Path(temp_dir)
            (job_dir / "json").mkdir()
            manifest = job_dir / "manifest.json"
            manifest.write_text("{}", encoding="utf-8")
            document = _document()
            (job_dir / "json" / "001.json").write_text(
                json.dumps(document, ensure_ascii=False), encoding="utf-8"
            )
            job = {"id": 42, "stored_input_path": str(manifest)}

            output, filename, mimetype = build_job_domestic_export(job)
            self.assertTrue(filename.endswith("_内销模板.xlsx"))
            self.assertIn("spreadsheetml", mimetype)
            self.assertGreater(len(output.getvalue()), 1000)

            second = _document()
            second["source_file"] = "PO-002.pdf"
            (job_dir / "json" / "002.json").write_text(
                json.dumps(second, ensure_ascii=False), encoding="utf-8"
            )
            output, filename, mimetype = build_job_domestic_export(job)
            self.assertEqual(filename, "PDF转Excel任务_42_内销模板.zip")
            self.assertEqual(mimetype, "application/zip")
            with zipfile.ZipFile(output) as archive:
                self.assertEqual(len(archive.namelist()), 2)
                self.assertTrue(all(name.endswith("_内销模板.xlsx") for name in archive.namelist()))

    def test_job_export_rejects_old_task_without_structured_json(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            manifest = Path(temp_dir) / "manifest.json"
            manifest.write_text("{}", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "没有可用的结构化结果"):
                build_job_domestic_export({"id": 42, "stored_input_path": str(manifest)})


class PdfExcelDomesticDownloadRouteTests(unittest.TestCase):
    def setUp(self) -> None:
        self.app = Flask(__name__, template_folder="../templates", static_folder="../static")
        self.app.config.update(SECRET_KEY="test-secret", TESTING=True)
        self.app.register_blueprint(bp)
        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["employee_id"] = "tester"

    def test_completed_owned_pdf_job_downloads_new_template(self) -> None:
        job = {
            "id": 42,
            "employee_id": "tester",
            "feature": "pdf_excel",
            "status": "completed",
        }
        with patch("fangzheng_web_app.routes.get_job", return_value=job), patch(
            "fangzheng_web_app.routes.build_job_domestic_export",
            return_value=(BytesIO(b"xlsx"), "内销模板.xlsx", "application/test"),
        ):
            response = self.client.get("/pdf-excel/jobs/42/domestic-export")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b"xlsx")
        self.assertIn("attachment", response.headers["Content-Disposition"])

    def test_pdf_pages_keep_original_download_and_add_domestic_download(self) -> None:
        job = {
            "id": 42,
            "employee_id": "tester",
            "feature": "pdf_excel",
            "source_filename": "PDF/图片转Excel：PO-001.pdf",
            "stored_input_path": "manifest.json",
            "stored_result_path": "result.xlsx",
            "status": "completed",
            "success_count": 1,
            "fail_count": 0,
            "skip_count": 0,
            "current_row": 1,
            "total_rows": 1,
            "rule_version": "pdf_excel_v1",
            "log_text": "完成",
            "error_message": "",
            "created_at": "2026-08-21 10:00:00",
            "completed_at": "2026-08-21 10:01:00",
        }
        with patch("fangzheng_web_app.routes.get_user", return_value=None), patch(
            "fangzheng_web_app.routes.is_admin_user", return_value=False
        ), patch("fangzheng_web_app.routes.list_jobs", return_value=[job]), patch(
            "fangzheng_web_app.routes.get_job", return_value=job
        ):
            responses = [
                self.client.get("/features/pdf-excel"),
                self.client.get("/jobs/42"),
                self.client.get("/history"),
            ]

        for response in responses:
            self.assertEqual(response.status_code, 200)
            html = response.get_data(as_text=True)
            self.assertIn("/download/42/result", html)
            self.assertIn("/pdf-excel/jobs/42/domestic-export", html)
            self.assertIn("下载内销模板", html)

    def test_cross_user_and_unfinished_jobs_do_not_download(self) -> None:
        cross_user = {
            "id": 42,
            "employee_id": "someone-else",
            "feature": "pdf_excel",
            "status": "completed",
        }
        with patch("fangzheng_web_app.routes.get_job", return_value=cross_user):
            response = self.client.get("/pdf-excel/jobs/42/domestic-export")
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/history"))

        unfinished = {
            "id": 42,
            "employee_id": "tester",
            "feature": "pdf_excel",
            "status": "running",
        }
        with patch("fangzheng_web_app.routes.get_job", return_value=unfinished):
            response = self.client.get("/pdf-excel/jobs/42/domestic-export")
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/jobs/42"))

    def test_unauthenticated_and_non_pdf_jobs_do_not_download(self) -> None:
        client = self.app.test_client()
        response = client.get("/pdf-excel/jobs/42/domestic-export")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers["Location"])

        wrong_feature = {
            "id": 42,
            "employee_id": "tester",
            "feature": "price_calculation",
            "status": "completed",
        }
        with patch("fangzheng_web_app.routes.get_job", return_value=wrong_feature):
            response = self.client.get("/pdf-excel/jobs/42/domestic-export")
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/history"))


if __name__ == "__main__":
    unittest.main()
