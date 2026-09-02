from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from fangzheng_web_app.pp_transcode_service import (
    calculate_pp_transcode_quote,
    queue_pp_transcode_single_job,
    refresh_pp_result_file,
)
from fangzheng_web_app.pp_transcode_rules import resolve_shared_pp_glue


class PPTranscodeServiceTests(unittest.TestCase):
    def test_single_input_creates_one_row_pp_job(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "fangzheng_web_app.pp_transcode_service.JOBS_DIR", Path(temp_dir)
        ), patch(
            "fangzheng_web_app.pp_transcode_service.create_job", return_value=88
        ) as create_job, patch(
            "fangzheng_web_app.pp_transcode_service.launch_job_process"
        ) as launch_job:
            job_id = queue_pp_transcode_single_job(
                "tester",
                spec="PP NY2150 1080 RC70% 300M/卷",
                customer="测试客户",
                customer_code="100001",
                order_remark="测试备注",
            )
            input_path = Path(create_job.call_args.args[2])
            workbook = load_workbook(input_path)
            sheet = workbook["PP转码需求表"]
            self.assertEqual([cell.value for cell in sheet[1]], ["客户代码", "客户简称", "PP客户规格", "订单备注"])
            self.assertEqual([cell.value for cell in sheet[2]], ["100001", "测试客户", "PP NY2150 1080 RC70% 300M/卷", "测试备注"])
            workbook.close()

        self.assertEqual(job_id, 88)
        self.assertEqual(create_job.call_args.kwargs["feature"], "pp_transcode_agent")
        launch_job.assert_called_once_with(88, "pp_transcode_agent", "tester")

    def test_shared_glue_uses_marketing_master_but_excludes_ccl_selection_rules(self) -> None:
        marketing_tables = {
            "Agent胶系主表": [
                {
                    "映射ID": "TGM-MASTER-0001",
                    "启用": "是",
                    "胶系编号": "2HNN",
                    "胶系名称": "NY2150",
                    "输出胶系代码": "2H",
                    "来源行号": 1,
                },
                {
                    "映射ID": "TGM-MASTER-0058",
                    "启用": "是",
                    "胶系编号": "2ZNA",
                    "胶系名称": "NY-A1",
                    "输出胶系代码": "2Z",
                    "来源行号": 2,
                },
            ],
            "Agent胶系兼容别名": [],
            # This is a CCL-specific selection rule. PP must not inherit it.
            "Agent胶系选择规则": [
                {
                    "映射ID": "TGM-SELECT-9999",
                    "启用": "是",
                    "胶系名称": "NY2150",
                    "条件关键词": "汽车板",
                    "输出胶系代码": "3B",
                    "优先级": 999,
                }
            ],
        }
        with patch(
            "fangzheng_web_app.pp_transcode_rules.get_active_transcode_agent_rule_version",
            return_value="agent-test-version",
        ), patch(
            "fangzheng_web_app.pp_transcode_rules.load_transcode_agent_mapping_tables",
            return_value=marketing_tables,
        ):
            resolved = resolve_shared_pp_glue("NY2150 汽车板")
            retired = resolve_shared_pp_glue("NY-A1")

        self.assertIsNotNone(resolved)
        self.assertEqual(resolved["output_value"], "2H")
        self.assertIsNone(retired)

    def test_full_pp_spec_preserves_glue_name_boundaries(self) -> None:
        shared_glue = {
            "id": "shared:TGM-MASTER-0027",
            "input_value": "NY2150",
            "output_value": "2B",
            "business_note": "营销转码Agent胶系主表",
        }
        with patch("fangzheng_web_app.pp_transcode_service.seed_pp_transcode_rules"), patch(
            "fangzheng_web_app.pp_transcode_service.list_base_rules", return_value=[]
        ), patch(
            "fangzheng_web_app.pp_transcode_service.resolve_shared_pp_glue",
            return_value=shared_glue,
        ) as glue_resolver, patch(
            "fangzheng_web_app.pp_transcode_service.list_customer_rules", return_value=[]
        ):
            quote = calculate_pp_transcode_quote("PP NY2150 1080 RC70% 300M/卷")

        glue_resolver.assert_called_once_with("PP NY2150 1080 RC70% 300M/卷")
        self.assertEqual(quote["pending_code"][0:2], "2B")

    def test_quote_only_generates_pending_code(self) -> None:
        def base_rules(*, field_key: str, enabled: str):
            rules = {
                "formula_category": [{"id": 2, "input_value": "NY2150", "output_value": "R"}],
            }
            return rules.get(field_key, [])

        with patch("fangzheng_web_app.pp_transcode_service.seed_pp_transcode_rules"), patch(
            "fangzheng_web_app.pp_transcode_service.list_base_rules", side_effect=base_rules
        ), patch(
            "fangzheng_web_app.pp_transcode_service.resolve_shared_pp_glue",
            return_value={"id": "shared:TGM-MASTER-0001", "input_value": "NY2150", "output_value": "2B", "business_note": "营销转码Agent胶系主表"},
        ), patch("fangzheng_web_app.pp_transcode_service.list_customer_rules", return_value=[]):
            quote = calculate_pp_transcode_quote("NY2150 玻布106 350M RC42%")

        self.assertEqual(quote["status"], "待人工确认")
        self.assertEqual(quote["formal_code"], "")
        self.assertEqual(len(quote["pending_code"]), 27)
        self.assertEqual(quote["pending_code"][0:2], "2B")
        self.assertEqual(quote["pending_code"][2:6], "0106")
        self.assertEqual(quote["pending_code"][6:10], "350M")
        self.assertEqual(quote["pending_code"][10], "R")
        self.assertEqual(quote["pending_code"][11:14], "420")
        self.assertTrue(quote["requires_manual_confirmation"])

    def test_unmaintained_customer_fields_keep_position_width_stars(self) -> None:
        with patch("fangzheng_web_app.pp_transcode_service.seed_pp_transcode_rules"), patch(
            "fangzheng_web_app.pp_transcode_service.list_base_rules", return_value=[]
        ), patch("fangzheng_web_app.pp_transcode_service.resolve_shared_pp_glue", return_value=None), patch(
            "fangzheng_web_app.pp_transcode_service.list_customer_rules", return_value=[]
        ):
            quote = calculate_pp_transcode_quote("未知 PP 规格")

        self.assertEqual(quote["formal_code"], "")
        self.assertEqual(quote["pending_code"][14:16], "**")
        self.assertEqual(quote["pending_code"][16], "*")
        self.assertEqual(quote["pending_code"][18:21], "***")
        self.assertEqual(quote["pending_code"][21:24], "***")
        self.assertEqual(quote["pending_code"][24:26], "**")
        self.assertEqual(quote["pending_code"][26], "*")

    def test_confirmed_value_is_written_as_confirmation_not_formal_code(self) -> None:
        confirmed_code = "2B0106350MR" + "*" * 16
        self.assertEqual(len(confirmed_code), 27)

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "pp_result.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "PP转码结果"
            sheet.append(["客户规格", "PP待人工确认码值", "PP转码状态"])
            sheet.append(["NY2150 玻布106", confirmed_code, "待人工确认"])
            workbook.save(output_path)
            workbook.close()

            job = {"stored_result_path": str(output_path)}
            confirmations = [{
                "excel_row": 2,
                "confirmation_status": "confirmed",
                "confirmed_pending_code": confirmed_code,
            }]
            with patch("fangzheng_web_app.pp_transcode_service.get_job", return_value=job), patch(
                "fangzheng_web_app.pp_transcode_service.list_pp_confirmation_items",
                return_value=confirmations,
            ):
                refresh_pp_result_file(99, "test-user")

            workbook = load_workbook(output_path)
            sheet = workbook["PP转码结果"]
            headers = {sheet.cell(1, column).value: column for column in range(1, sheet.max_column + 1)}
            self.assertEqual(sheet.cell(2, headers["PP本次确认码值"]).value, confirmed_code)
            self.assertEqual(sheet.cell(2, headers["PP人工确认状态"]).value, "本次已确认")
            self.assertEqual(sheet.cell(2, headers["PP转码状态"]).value, "本次已确认")
            self.assertFalse(any("正式码" in str(header or "") for header in headers))
            workbook.close()


if __name__ == "__main__":
    unittest.main()
